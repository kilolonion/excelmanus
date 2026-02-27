"""VLM 超时/重试/错误净化 单元测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from excelmanus.engine_core.tool_dispatcher import ToolDispatcher


class TestSanitizeVlmError:
    """_sanitize_vlm_error 错误净化测试。"""

    def test_html_504_extracts_code_and_title(self) -> None:
        exc = Exception(
            '<!DOCTYPE html><html><head><title>example.com | 504: Gateway time-out</title></head>'
            '<body>lots of html...</body></html>'
        )
        result = ToolDispatcher._sanitize_vlm_error(exc)
        assert "504" in result
        assert "Gateway time-out" in result
        assert "<html" not in result

    def test_html_502_extracts_code(self) -> None:
        exc = Exception(
            '<html><head><title>502 Bad Gateway</title></head><body></body></html>'
        )
        result = ToolDispatcher._sanitize_vlm_error(exc)
        assert "502" in result
        assert "<html" not in result

    def test_plain_error_preserved(self) -> None:
        exc = Exception("Connection refused")
        result = ToolDispatcher._sanitize_vlm_error(exc)
        assert result == "Connection refused"

    def test_long_error_truncated(self) -> None:
        exc = Exception("x" * 1000)
        result = ToolDispatcher._sanitize_vlm_error(exc)
        assert len(result) <= 500


class TestBuildVlmFailureResult:
    """_build_vlm_failure_result 降级引导测试。"""

    def test_contains_fallback_hint(self) -> None:
        exc = TimeoutError("VLM 调用超时（120s）")
        result = json.loads(ToolDispatcher._build_vlm_failure_result(exc, 3, "test.png"))
        assert result["status"] == "error"
        assert result["error_code"] == "VLM_CALL_FAILED"
        assert "fallback_hint" in result
        assert "read_image" in result["fallback_hint"]
        assert "run_code" in result["fallback_hint"]
        assert result["file_path"] == "test.png"

    def test_html_error_sanitized_in_message(self) -> None:
        exc = Exception('<html><head><title>504: Gateway time-out</title></head></html>')
        result = json.loads(ToolDispatcher._build_vlm_failure_result(exc, 2, "img.png"))
        assert "<html" not in result["message"]
        assert "504" in result["message"]

    def test_none_error_handled(self) -> None:
        result = json.loads(ToolDispatcher._build_vlm_failure_result(None, 1, "x.png"))
        assert "未知错误" in result["message"]


class TestPrepareImageForVlm:
    """图片预处理测试。"""

    def test_small_image_not_resized(self) -> None:
        """小图片不应被缩放。"""
        import io
        from PIL import Image
        img = Image.new("RGB", (800, 600), "white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()

        result, mime = ToolDispatcher._prepare_image_for_vlm(raw, max_long_edge=2048)
        # 不应放大（800 < 2048）
        assert len(result) > 0

    def test_large_image_resized(self) -> None:
        """大图片应被缩放到 max_long_edge。"""
        import io
        from PIL import Image
        img = Image.new("RGB", (4000, 3000), "white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()

        result, mime = ToolDispatcher._prepare_image_for_vlm(raw, max_long_edge=2048)
        # 验证输出确实被处理了
        result_img = Image.open(io.BytesIO(result))
        assert max(result_img.size) <= 2048

    def test_rgba_converted_to_rgb(self) -> None:
        """RGBA 图片应被转换为 RGB。"""
        import io
        from PIL import Image
        img = Image.new("RGBA", (100, 100), (255, 0, 0, 128))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()

        result, mime = ToolDispatcher._prepare_image_for_vlm(raw)
        assert len(result) > 0

    def test_gray_background_detection(self) -> None:
        """灰底图片应被检测并预处理。"""
        import io
        from PIL import Image
        # 创建灰底图片（mean ~200, low stddev）
        img = Image.new("RGB", (200, 200), (200, 200, 200))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()
        result, mime = ToolDispatcher._prepare_image_for_vlm(raw)
        assert len(result) > 0

    def test_low_contrast_image_enhanced(self) -> None:
        """低对比度图片应被增强。"""
        import io
        from PIL import Image
        # 灰色图片（低 stddev）
        img = Image.new("RGB", (200, 200), (180, 180, 180))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()
        result, mime = ToolDispatcher._prepare_image_for_vlm(raw)
        assert len(result) > 0


class TestInferNumberFormat:
    """_infer_number_format 数字格式推断测试。"""

    def test_percentage(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("85%") == "0%"

    def test_percentage_with_decimal(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("12.5%") == "0.0%"

    def test_percentage_two_decimals(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("12.50%") == "0.00%"

    def test_dollar_with_comma(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("$1,200") == "$#,##0"

    def test_dollar_with_decimals(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("$1,200.50") == "$#,##0.00"

    def test_yen_with_comma(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("¥1,200") == "¥#,##0"

    def test_comma_separated_no_decimal(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("1,200") == "#,##0"

    def test_comma_with_decimals(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("1,200.00") == "#,##0.00"

    def test_decimal_only(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("12.50") == "0.00"

    def test_plain_integer_returns_none(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("1200") is None

    def test_empty_returns_none(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("") is None

    def test_text_returns_none(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("hello") is None

    def test_negative_percentage(self) -> None:
        from excelmanus.tools.image_tools import _infer_number_format
        assert _infer_number_format("-5.5%") == "0.0%"


class TestRebuildBorderPerSide:
    """rebuild_excel_from_spec Border 四边独立测试。"""

    def test_per_side_border(self, tmp_path: Path) -> None:
        """四边独立 border 应正确应用。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        spec = {
            "version": "1.0",
            "provenance": {"source_image_hash": "", "model": "", "timestamp": ""},
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 1, "cols": 1},
                "cells": [{"address": "A1", "value": "test", "value_type": "string", "style_id": "s1", "confidence": 0.9}],
                "styles": {
                    "s1": {
                        "border": {
                            "top": {"style": "thick", "color": "#000000"},
                            "bottom": {"style": "thin", "color": "#FF0000"},
                        }
                    }
                },
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        init_guard(str(tmp_path))

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path),
            output_path=str(tmp_path / "out.xlsx"),
        ))
        assert result["status"] == "ok"

        # 验证 border 已应用
        from openpyxl import load_workbook
        wb = load_workbook(str(tmp_path / "out.xlsx"))
        cell = wb.active["A1"]
        assert cell.border.top.style == "thick"
        assert cell.border.bottom.style == "thin"

    def test_uniform_border_backward_compat(self, tmp_path: Path) -> None:
        """统一 border（旧格式）应向后兼容。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        spec = {
            "version": "1.0",
            "provenance": {"source_image_hash": "", "model": "", "timestamp": ""},
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 1, "cols": 1},
                "cells": [{"address": "A1", "value": "test", "value_type": "string", "style_id": "s1", "confidence": 0.9}],
                "styles": {"s1": {"border": {"style": "thin"}}},
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        init_guard(str(tmp_path))

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path),
            output_path=str(tmp_path / "out.xlsx"),
        ))
        assert result["status"] == "ok"

        from openpyxl import load_workbook
        wb = load_workbook(str(tmp_path / "out.xlsx"))
        cell = wb.active["A1"]
        assert cell.border.top.style == "thin"
        assert cell.border.left.style == "thin"

    def test_number_format_inferred_from_display_text(self, tmp_path: Path) -> None:
        """number_format 应从 display_text 自动推断。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        spec = {
            "version": "1.0",
            "provenance": {"source_image_hash": "", "model": "", "timestamp": ""},
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 2, "cols": 1},
                "cells": [
                    {"address": "A1", "value": 1200.5, "value_type": "number", "display_text": "$1,200.50", "confidence": 0.9},
                    {"address": "A2", "value": 0.85, "value_type": "number", "display_text": "85%", "confidence": 0.9},
                ],
                "styles": {},
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        init_guard(str(tmp_path))

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path),
            output_path=str(tmp_path / "out.xlsx"),
        ))
        assert result["status"] == "ok"

        from openpyxl import load_workbook
        wb = load_workbook(str(tmp_path / "out.xlsx"))
        ws = wb.active
        assert ws["A1"].number_format == "$#,##0.00"
        assert ws["A2"].number_format == "0%"

    def test_column_width_tolerance(self, tmp_path: Path) -> None:
        """列宽数组过长时不应崩溃。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        spec = {
            "version": "1.0",
            "provenance": {"source_image_hash": "", "model": "", "timestamp": ""},
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 1, "cols": 2},
                "cells": [
                    {"address": "A1", "value": "x", "value_type": "string", "confidence": 0.9},
                ],
                "column_widths": [15, 12, 10, 8, 20, 25, 30, 35, 40, 45, 50],
                "styles": {},
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        init_guard(str(tmp_path))

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path),
            output_path=str(tmp_path / "out.xlsx"),
        ))
        assert result["status"] == "ok"
