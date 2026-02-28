"""image_tools 单元测试。"""

from __future__ import annotations

import base64
import json
import tempfile
from pathlib import Path

import pytest


# 最小有效 PNG（1x1 白色像素）
_MINIMAL_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR4"
    "nGP4z8BQDwAEgAF/pooBPQAAAABJRU5ErkJggg=="
)


class TestReadImage:
    def test_read_png_file(self, tmp_path: Path) -> None:
        """读取 PNG 文件返回正确元数据。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        png_data = base64.b64decode(_MINIMAL_PNG_B64)
        img_path = tmp_path / "test.png"
        img_path.write_bytes(png_data)
        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(img_path)))
        assert result["status"] == "ok"
        assert result["mime_type"] == "image/png"
        assert result["size_bytes"] > 0
        assert "__tool_result_image__" in result

    def test_read_nonexistent_file(self, tmp_path: Path) -> None:
        """读取不存在的文件返回错误。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(tmp_path / "nope.png")))
        assert result["status"] == "error"

    def test_read_unsupported_format(self, tmp_path: Path) -> None:
        """不支持的格式返回错误。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        txt = tmp_path / "test.txt"
        txt.write_text("not an image")
        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(txt)))
        assert result["status"] == "error"

    def test_read_image_too_large(self, tmp_path: Path) -> None:
        """超大文件返回错误。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        big = tmp_path / "big.png"
        big.write_bytes(b"x" * (20_000_001))
        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(big)))
        assert result["status"] == "error"
        assert "超限" in result["message"] or "size" in result["message"].lower()

    def test_image_injection_structure(self, tmp_path: Path) -> None:
        """__tool_result_image__ 结构正确。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        png_data = base64.b64decode(_MINIMAL_PNG_B64)
        img_path = tmp_path / "test.png"
        img_path.write_bytes(png_data)
        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(img_path)))
        injection = result["__tool_result_image__"]
        assert injection["mime_type"] == "image/png"
        assert injection["detail"] == "auto"
        assert len(injection["base64"]) > 0

    def test_get_tools_returns_read_image(self) -> None:
        """get_tools 返回 read_image 工具定义。"""
        from excelmanus.tools.image_tools import get_tools

        tools = get_tools()
        names = [t.name for t in tools]
        assert "read_image" in names

    def test_rejects_outside_workspace_path(self, tmp_path: Path) -> None:
        """workspace 外图片路径应被拒绝。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        outside_dir = Path(tempfile.mkdtemp())
        outside_img = outside_dir / "outside.png"
        outside_img.write_bytes(base64.b64decode(_MINIMAL_PNG_B64))

        init_guard(str(tmp_path))
        result = json.loads(read_image(file_path=str(outside_img)))
        assert result["status"] == "error"
        assert "路径" in result["message"]


class TestTryInjectImage:
    """ToolDispatcher._try_inject_image 单元测试。"""

    def _make_dispatcher(self):
        """创建最小化 mock ToolDispatcher。"""
        from unittest.mock import MagicMock
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher

        engine = MagicMock()
        engine.memory = MagicMock()
        engine._database = None
        engine.config = MagicMock()
        engine.config.code_policy_enabled = False
        engine.state = MagicMock()
        engine.approval = MagicMock()
        engine.approval.is_audit_only_tool = MagicMock(return_value=False)
        engine.approval.is_high_risk_tool = MagicMock(return_value=False)
        engine.full_access_enabled = False
        dispatcher = ToolDispatcher.__new__(ToolDispatcher)
        dispatcher._engine = engine
        dispatcher._pending_vlm_image = None
        dispatcher._deferred_image_injections = []
        dispatcher._injected_image_hashes = set()
        dispatcher._last_vlm_description = None
        dispatcher._last_vlm_description_image_hash = None
        dispatcher._tool_call_store = None
        dispatcher._handlers = []
        return dispatcher, engine

    def test_inject_image_from_result(self) -> None:
        """含 __tool_result_image__ 的结果应延迟注入并移除字段。"""
        dispatcher, engine = self._make_dispatcher()
        # 使用合法 base64，避免解码路径（如 hash 或 VLM）在部分环境下报 Incorrect padding
        b64 = _MINIMAL_PNG_B64.replace("\n", "").strip()
        result = json.dumps({
            "status": "ok",
            "hint": "图片已加载",
            "__tool_result_image__": {"base64": b64, "mime_type": "image/png", "detail": "auto"},
        })
        cleaned = dispatcher._try_inject_image(result)
        parsed = json.loads(cleaned)
        assert "__tool_result_image__" not in parsed
        assert parsed["status"] == "ok"
        # 图片不再立即注入 memory，而是延迟到所有 tool result 写入后
        engine.memory.add_image_message.assert_not_called()
        assert len(dispatcher._deferred_image_injections) == 1
        assert dispatcher._deferred_image_injections[0]["base64"] == b64
        # flush 后才真正注入
        dispatcher.flush_deferred_images()
        engine.memory.add_image_message.assert_called_once_with(
            base64_data=b64, mime_type="image/png", detail="auto",
        )
        assert len(dispatcher._deferred_image_injections) == 0

    def test_no_injection_without_marker(self) -> None:
        """无 __tool_result_image__ 时不触发注入。"""
        dispatcher, engine = self._make_dispatcher()
        result = json.dumps({"status": "ok", "data": "test"})
        cleaned = dispatcher._try_inject_image(result)
        assert cleaned == result
        engine.memory.add_image_message.assert_not_called()

    def test_no_injection_for_non_json(self) -> None:
        """非 JSON 字符串不触发注入。"""
        dispatcher, engine = self._make_dispatcher()
        cleaned = dispatcher._try_inject_image("not json")
        assert cleaned == "not json"
        engine.memory.add_image_message.assert_not_called()

    @pytest.mark.asyncio
    async def test_call_registry_tool_injects_before_truncate(self) -> None:
        """read_image 结果应先注入 __tool_result_image__ 再截断，避免截断破坏 JSON。"""
        from unittest.mock import MagicMock

        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher

        payload = json.dumps({
            "status": "ok",
            "hint": "图片已加载",
            "__tool_result_image__": {
                "base64": "A" * 5000,
                "mime_type": "image/png",
                "detail": "auto",
            },
        }, ensure_ascii=False)

        tool_def = MagicMock()
        tool_def.truncate_result.side_effect = lambda s: s[:100]

        registry = MagicMock()
        registry.call_tool.return_value = payload
        registry.get_tool.return_value = tool_def

        engine = MagicMock()
        engine._registry = registry
        engine.registry = registry
        engine.memory = MagicMock()
        engine._persistent_memory = None
        engine._database = None
        engine.config = MagicMock()
        engine.config.code_policy_enabled = False
        engine.state = MagicMock()
        engine.approval = MagicMock()
        engine.approval.is_audit_only_tool = MagicMock(return_value=False)
        engine.approval.is_high_risk_tool = MagicMock(return_value=False)
        engine.full_access_enabled = False

        dispatcher = ToolDispatcher(engine)
        out = await dispatcher.call_registry_tool(
            tool_name="read_image",
            arguments={"file_path": "x.png"},
            tool_scope=None,
        )
        # 截断后依然应是可解析 JSON 且已移除注入字段
        parsed = json.loads(out)
        assert parsed["status"] == "ok"
        assert "__tool_result_image__" not in parsed
        # 图片延迟注入：call_registry_tool 期间不直接调用 add_image_message
        engine.memory.add_image_message.assert_not_called()
        assert len(dispatcher._deferred_image_injections) == 1
        # flush 后才真正注入
        dispatcher.flush_deferred_images()
        engine.memory.add_image_message.assert_called_once()


_BASIC_SPEC = {
    "version": "1.0",
    "provenance": {"source_image_hash": "sha256:x", "model": "test", "timestamp": "2026-01-01T00:00:00Z"},
    "workbook": {"name": "test"},
    "sheets": [{
        "name": "Sheet1",
        "dimensions": {"rows": 2, "cols": 2},
        "cells": [
            {"address": "A1", "value": "Name", "value_type": "string", "style_id": "header", "confidence": 1.0},
            {"address": "B1", "value": "Age", "value_type": "string", "style_id": "header", "confidence": 1.0},
            {"address": "A2", "value": "Alice", "value_type": "string", "confidence": 1.0},
            {"address": "B2", "value": 30, "value_type": "number", "confidence": 1.0},
        ],
        "styles": {
            "header": {"font": {"bold": True, "size": 12, "color": "#FFFFFF"}, "fill": {"type": "solid", "color": "#4472C4"}},
        },
        "column_widths": [15, 10],
    }],
    "uncertainties": [],
}


class TestRebuildExcelFromSpec:
    def test_basic_rebuild(self, tmp_path: Path) -> None:
        """基础 spec → xlsx 编译。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        init_guard(str(tmp_path))
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_BASIC_SPEC), encoding="utf-8")
        output_path = tmp_path / "output.xlsx"

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path), output_path=str(output_path),
        ))
        assert result["status"] == "ok"
        assert output_path.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Sheet1"]
        assert ws["A1"].value == "Name"
        assert ws["B2"].value == 30
        assert ws["A1"].font.bold is True

    def test_spec_not_found(self, tmp_path: Path) -> None:
        """Spec 文件不存在返回错误。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        init_guard(str(tmp_path))
        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(tmp_path / "nope.json"),
        ))
        assert result["status"] == "error"

    def test_merged_cells(self, tmp_path: Path) -> None:
        """合并单元格正确应用（非锚点无值时正常合并）。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        init_guard(str(tmp_path))
        spec = {
            **_BASIC_SPEC,
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 2, "cols": 2},
                "cells": [
                    {"address": "A1", "value": "Title", "value_type": "string", "confidence": 1.0},
                    {"address": "A2", "value": "Alice", "value_type": "string", "confidence": 1.0},
                    {"address": "B2", "value": 30, "value_type": "number", "confidence": 1.0},
                ],
                "merged_ranges": [{"range": "A1:B1", "confidence": 0.95}],
                "styles": {},
                "column_widths": [15, 10],
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        output_path = tmp_path / "output.xlsx"

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path), output_path=str(output_path),
        ))
        assert result["status"] == "ok"
        assert result["build_summary"]["merges_applied"] == 1

    def test_merge_skipped_when_non_anchor_has_value(self, tmp_path: Path) -> None:
        """非锚点单元格含有值时跳过合并，保留数据完整性。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        init_guard(str(tmp_path))
        spec = {
            **_BASIC_SPEC,
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 2, "cols": 4},
                "cells": [
                    {"address": "A1", "value": "日期：", "value_type": "string", "confidence": 1.0},
                    {"address": "C1", "value": "客户：", "value_type": "string", "confidence": 1.0},
                ],
                "merged_ranges": [{"range": "A1:D1", "confidence": 0.98}],
                "styles": {},
                "column_widths": [10, 10, 10, 10],
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        output_path = tmp_path / "output.xlsx"

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path), output_path=str(output_path),
        ))
        assert result["status"] == "ok"
        # 合并被跳过（因为 C1 是非锚点但有值）
        assert result["build_summary"]["merges_applied"] == 0
        assert any("跳过" in s for s in result["build_summary"]["skipped_items"])

        # 值仍然保留
        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Sheet1"]
        assert ws["A1"].value == "日期："
        assert ws["C1"].value == "客户："

    def test_rejects_outside_workspace_output_path(self, tmp_path: Path) -> None:
        """输出路径在 workspace 外时应被拒绝。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard

        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_BASIC_SPEC), encoding="utf-8")
        init_guard(str(tmp_path))

        outside_dir = Path(tempfile.mkdtemp())
        outside_output = outside_dir / "output.xlsx"
        result = json.loads(
            rebuild_excel_from_spec(spec_path=str(spec_path), output_path=str(outside_output))
        )
        assert result["status"] == "error"
        assert "路径" in result["message"]


class TestVerifyReplica:
    def test_perfect_match(self, tmp_path: Path) -> None:
        """spec 与 rebuild 的 Excel 完全匹配时 match_rate=1.0。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, verify_excel_replica, init_guard

        init_guard(str(tmp_path))
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_BASIC_SPEC), encoding="utf-8")
        excel_path = tmp_path / "output.xlsx"
        rebuild_excel_from_spec(spec_path=str(spec_path), output_path=str(excel_path))

        report_path = tmp_path / "report.md"
        result = json.loads(verify_excel_replica(
            spec_path=str(spec_path), excel_path=str(excel_path), report_path=str(report_path),
        ))
        assert result["status"] == "ok"
        assert result["match_rate"] == 1.0
        assert result["issues"]["total"] == 0
        assert report_path.exists()

    def test_report_file_generated(self, tmp_path: Path) -> None:
        """diff report markdown 文件正确生成。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, verify_excel_replica, init_guard

        init_guard(str(tmp_path))
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_BASIC_SPEC), encoding="utf-8")
        excel_path = tmp_path / "output.xlsx"
        rebuild_excel_from_spec(spec_path=str(spec_path), output_path=str(excel_path))

        report_path = tmp_path / "report.md"
        verify_excel_replica(
            spec_path=str(spec_path), excel_path=str(excel_path), report_path=str(report_path),
        )
        content = report_path.read_text(encoding="utf-8")
        assert "验证报告" in content
        assert "匹配率" in content

    def test_uncertainty_items_in_report(self, tmp_path: Path) -> None:
        """低置信项出现在 diff report 中。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, verify_excel_replica, init_guard

        init_guard(str(tmp_path))
        spec = dict(_BASIC_SPEC)
        spec["uncertainties"] = [
            {"location": "B2", "reason": "数字模糊", "candidate_values": ["30", "38"], "confidence": 0.6},
        ]
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        excel_path = tmp_path / "output.xlsx"
        rebuild_excel_from_spec(spec_path=str(spec_path), output_path=str(excel_path))

        report_path = tmp_path / "report.md"
        result = json.loads(verify_excel_replica(
            spec_path=str(spec_path), excel_path=str(excel_path), report_path=str(report_path),
        ))
        assert result["issues"]["low_confidence"] == 1
        content = report_path.read_text(encoding="utf-8")
        assert "低置信项" in content

    def test_merge_conflict_not_counted_as_mismatch(self, tmp_path: Path) -> None:
        """合并区域非锚点单元格归类为 merge_conflict 而非 mismatch，不降低匹配率。"""
        from excelmanus.tools.image_tools import verify_excel_replica, init_guard
        from openpyxl import Workbook

        init_guard(str(tmp_path))

        # 手动构建一个含合并单元格的 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = "日期："
        # C1 本应有值 "客户：" 但因合并会变成 None
        ws.merge_cells("A1:D1")
        excel_path = tmp_path / "merged.xlsx"
        wb.save(str(excel_path))

        # 构建 spec：A1 和 C1 都有值
        spec = {
            **_BASIC_SPEC,
            "sheets": [{
                "name": "Sheet1",
                "dimensions": {"rows": 1, "cols": 4},
                "cells": [
                    {"address": "A1", "value": "日期：", "value_type": "string", "confidence": 1.0},
                    {"address": "C1", "value": "客户：", "value_type": "string", "confidence": 1.0},
                ],
                "merged_ranges": [{"range": "A1:D1", "confidence": 0.98}],
                "styles": {},
                "column_widths": [],
            }],
        }
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        report_path = tmp_path / "report.md"

        result = json.loads(verify_excel_replica(
            spec_path=str(spec_path), excel_path=str(excel_path), report_path=str(report_path),
        ))
        assert result["status"] == "ok"
        # C1 应被识别为 merge_conflict 而非 mismatch
        assert result["issues"]["merge_conflicts"] == 1
        assert result["issues"]["conflict"] == 0
        # 匹配率应为 100%（merge_conflict 不降低匹配率）
        assert result["match_rate"] == 1.0

        # 报告中应包含合并冲突信息
        content = report_path.read_text(encoding="utf-8")
        assert "合并单元格冲突" in content

    def test_rejects_outside_workspace_report_path(self, tmp_path: Path) -> None:
        """报告路径在 workspace 外时应被拒绝。"""
        from excelmanus.tools.image_tools import (
            rebuild_excel_from_spec,
            verify_excel_replica,
            init_guard,
        )

        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_BASIC_SPEC), encoding="utf-8")
        excel_path = tmp_path / "output.xlsx"
        init_guard(str(tmp_path))
        rebuild_excel_from_spec(spec_path=str(spec_path), output_path=str(excel_path))

        outside_dir = Path(tempfile.mkdtemp())
        outside_report = outside_dir / "report.md"
        result = json.loads(
            verify_excel_replica(
                spec_path=str(spec_path),
                excel_path=str(excel_path),
                report_path=str(outside_report),
            )
        )
        assert result["status"] == "error"
        assert "路径" in result["message"]


class TestValuesMatchDateNormalization:
    """_values_match 日期归一化比较测试。"""

    def test_datetime_vs_date_string(self) -> None:
        from datetime import datetime
        from excelmanus.tools.image_tools import _values_match

        assert _values_match(datetime(2024, 1, 15), "2024-01-15") is True

    def test_date_string_vs_datetime_ignores_time(self) -> None:
        from datetime import datetime
        from excelmanus.tools.image_tools import _values_match

        assert _values_match("2024-01-15", datetime(2024, 1, 15, 10, 30)) is True

    def test_two_datetimes_ignore_time(self) -> None:
        from datetime import datetime
        from excelmanus.tools.image_tools import _values_match

        assert _values_match(datetime(2024, 1, 15), datetime(2024, 1, 15, 10, 30)) is True

    def test_date_object_vs_string(self) -> None:
        from datetime import date
        from excelmanus.tools.image_tools import _values_match

        assert _values_match(date(2024, 1, 15), "2024-01-15") is True

    def test_slash_format(self) -> None:
        from datetime import datetime
        from excelmanus.tools.image_tools import _values_match

        assert _values_match("2024/01/15", datetime(2024, 1, 15)) is True

    def test_different_dates_return_false(self) -> None:
        from excelmanus.tools.image_tools import _values_match

        assert _values_match("2024-01-15", "2024-01-16") is False

    def test_invalid_date_string_returns_false(self) -> None:
        from datetime import datetime
        from excelmanus.tools.image_tools import _values_match

        assert _values_match("not-a-date", datetime(2024, 1, 15)) is False

    def test_numeric_comparison_unchanged(self) -> None:
        from excelmanus.tools.image_tools import _values_match

        assert _values_match(1, 1.0) is True
        assert _values_match(3.14, 3.14) is True
        assert _values_match(1, 2) is False

    def test_none_comparison_unchanged(self) -> None:
        from excelmanus.tools.image_tools import _values_match

        assert _values_match(None, None) is True
        assert _values_match(None, "a") is False
        assert _values_match("a", None) is False
