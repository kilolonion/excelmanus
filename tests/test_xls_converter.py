"""XLS/XLSB → XLSX 转换模块测试。"""
from __future__ import annotations

import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock

from excelmanus.xls_converter import (
    needs_conversion,
    converted_xlsx_path,
    convert_to_xlsx,
    ensure_xlsx,
    ConversionError,
    CONVERTIBLE_EXTENSIONS,
)


# ── 辅助：创建真实 .xls 测试文件 ────────────────────────────

def _create_xls_file(path: Path, data: list[list] | None = None, sheet_name: str = "Sheet1") -> Path:
    """用 xlrd 兼容格式创建 .xls 测试文件（借助 xlwt 或直接二进制）。"""
    # 使用 openpyxl 创建 xlsx 再用 pandas 转 xls 不可行
    # 直接用 xlrd 能读取的最小 BIFF8 二进制写入
    import xlrd
    import struct

    # 最简单的方式：用 pandas to_excel with xlsxwriter 不支持 xls
    # 改用直接构造最小 BIFF8 文件
    # 更简单的方式：先写 xlsx 然后在测试中 mock xlrd
    # 但我们需要真实测试，所以尝试用 xlwt（如果有的话）或构造二进制

    # 最务实的方式：构造 .xlsx 文件，然后在测试中重命名为 .xls
    # 但 xlrd 不能读 xlsx...

    # 用 struct 构造最小 BIFF8 .xls 文件
    return _write_minimal_biff8(path, data or [["Name", "Age"], ["Alice", 30], ["Bob", 25]], sheet_name)


def _write_minimal_biff8(path: Path, data: list[list], sheet_name: str = "Sheet1") -> Path:
    """写入一个最小化的 BIFF8 .xls 文件。

    使用 xlwt 库（如果可用）或跳过。
    """
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not installed, cannot create .xls test files")

    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)

    for r, row in enumerate(data):
        for c, val in enumerate(row):
            if val is not None:
                ws.write(r, c, val)

    wb.save(str(path))
    return path


def _create_xls_with_styles(path: Path) -> Path:
    """创建带样式的 .xls 文件。"""
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not installed")

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Styled")

    # 粗体标题
    bold_font = xlwt.easyxf('font: bold on; alignment: horizontal center')
    ws.write(0, 0, "Name", bold_font)
    ws.write(0, 1, "Score", bold_font)

    # 数据
    ws.write(1, 0, "Alice")
    ws.write(1, 1, 95.5)
    ws.write(2, 0, "Bob")
    ws.write(2, 1, 87.0)

    # 列宽
    ws.col(0).width = 5000
    ws.col(1).width = 3000

    # 合并单元格
    ws.write_merge(3, 3, 0, 1, "Summary")

    wb.save(str(path))
    return path


# ── 纯逻辑测试 ──────────────────────────────────────────────


class TestNeedsConversion:
    def test_xls_needs_conversion(self):
        assert needs_conversion("test.xls") is True
        assert needs_conversion("test.XLS") is True

    def test_xlsb_needs_conversion(self):
        assert needs_conversion("test.xlsb") is True

    def test_xlsx_no_conversion(self):
        assert needs_conversion("test.xlsx") is False
        assert needs_conversion("test.xlsm") is False
        assert needs_conversion("test.csv") is False

    def test_path_object(self):
        assert needs_conversion(Path("/tmp/data.xls")) is True
        assert needs_conversion(Path("/tmp/data.xlsx")) is False


class TestConvertedXlsxPath:
    def test_xls_to_xlsx(self):
        assert converted_xlsx_path("report.xls") == Path("report.xlsx")

    def test_xlsb_to_xlsx(self):
        assert converted_xlsx_path("/tmp/data.xlsb") == Path("/tmp/data.xlsx")

    def test_preserves_directory(self):
        result = converted_xlsx_path("/home/user/uploads/test.xls")
        assert result == Path("/home/user/uploads/test.xlsx")


# ── 真实转换测试 ─────────────────────────────────────────────


class TestConvertXls:
    """需要 xlwt 来创建测试文件。"""

    def test_basic_data_conversion(self, tmp_path):
        """基础数据转换：验证数据完整性。"""
        src = tmp_path / "test.xls"
        data = [["Name", "Age", "City"], ["Alice", 30, "Beijing"], ["Bob", 25, "Shanghai"]]
        try:
            _create_xls_file(src, data)
        except Exception:
            pytest.skip("cannot create .xls test file")

        dst = convert_to_xlsx(src)

        assert dst.exists()
        assert dst.suffix == ".xlsx"

        # 用 openpyxl 验证
        from openpyxl import load_workbook
        wb = load_workbook(dst)
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Age"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 30
        assert ws.cell(3, 3).value == "Shanghai"
        wb.close()

    def test_styled_conversion(self, tmp_path):
        """带样式的转换：验证基础格式保留。"""
        src = tmp_path / "styled.xls"
        try:
            _create_xls_with_styles(src)
        except Exception:
            pytest.skip("cannot create styled .xls test file")

        dst = convert_to_xlsx(src)
        assert dst.exists()

        from openpyxl import load_workbook
        wb = load_workbook(dst)
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        # 合并单元格保留
        merged = list(ws.merged_cells.ranges)
        assert len(merged) >= 1
        wb.close()

    def test_custom_dst(self, tmp_path):
        """自定义目标路径。"""
        src = tmp_path / "test.xls"
        try:
            _create_xls_file(src)
        except Exception:
            pytest.skip("cannot create .xls test file")

        custom_dst = tmp_path / "output" / "converted.xlsx"
        result = convert_to_xlsx(src, custom_dst)
        assert result == custom_dst
        assert custom_dst.exists()

    def test_skip_existing(self, tmp_path):
        """已存在的目标文件不覆盖。"""
        src = tmp_path / "test.xls"
        try:
            _create_xls_file(src)
        except Exception:
            pytest.skip("cannot create .xls test file")

        dst = tmp_path / "test.xlsx"
        dst.write_text("existing")  # 预先创建

        result = convert_to_xlsx(src)
        assert result == dst
        assert dst.read_text() == "existing"  # 未被覆盖

    def test_overwrite_existing(self, tmp_path):
        """overwrite=True 时覆盖已存在的文件。"""
        src = tmp_path / "test.xls"
        try:
            _create_xls_file(src)
        except Exception:
            pytest.skip("cannot create .xls test file")

        dst = tmp_path / "test.xlsx"
        dst.write_text("old")

        result = convert_to_xlsx(src, overwrite=True)
        assert result == dst
        assert dst.stat().st_size > 3  # 不再是 "old" 的 3 字节

    def test_multi_sheet(self, tmp_path):
        """多 sheet 转换。"""
        try:
            import xlwt
        except ImportError:
            pytest.skip("xlwt not installed")

        src = tmp_path / "multi.xls"
        wb = xlwt.Workbook()
        ws1 = wb.add_sheet("Sales")
        ws1.write(0, 0, "Product")
        ws1.write(1, 0, "Widget")
        ws2 = wb.add_sheet("Inventory")
        ws2.write(0, 0, "Item")
        ws2.write(1, 0, "Bolt")
        wb.save(str(src))

        dst = convert_to_xlsx(src)

        from openpyxl import load_workbook
        xlsx_wb = load_workbook(dst)
        assert "Sales" in xlsx_wb.sheetnames
        assert "Inventory" in xlsx_wb.sheetnames
        assert xlsx_wb["Sales"].cell(1, 1).value == "Product"
        assert xlsx_wb["Inventory"].cell(1, 1).value == "Item"
        xlsx_wb.close()


class TestConvertErrors:
    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            convert_to_xlsx(tmp_path / "nonexistent.xls")

    def test_unsupported_format(self, tmp_path):
        f = tmp_path / "test.xlsx"
        f.write_bytes(b"fake")
        with pytest.raises(ValueError, match="不需要转换"):
            convert_to_xlsx(f)


class TestEnsureXlsx:
    def test_xlsx_passthrough(self, tmp_path):
        f = tmp_path / "test.xlsx"
        f.write_bytes(b"fake")
        result, converted = ensure_xlsx(f)
        assert result == f
        assert converted is False

    def test_csv_passthrough(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2")
        result, converted = ensure_xlsx(f)
        assert result == f
        assert converted is False

    def test_xls_converts(self, tmp_path):
        src = tmp_path / "test.xls"
        try:
            _create_xls_file(src)
        except Exception:
            pytest.skip("cannot create .xls test file")

        result, converted = ensure_xlsx(src)
        assert converted is True
        assert result.suffix == ".xlsx"
        assert result.exists()
