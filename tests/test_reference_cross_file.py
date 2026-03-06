"""CrossFileRefScanner 测试。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.reference_graph.cross_file import CrossFileRefScanner


@pytest.fixture()
def two_files(tmp_path: Path) -> tuple[Path, Path]:
    """两个文件，有共同列名（产品ID）和结构相似性。"""
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "订单表"
    ws1["A1"] = "产品ID"
    ws1["B1"] = "数量"
    ws1["C1"] = "金额"
    ws1["A2"] = "P001"
    ws1["B2"] = 5
    ws1["C2"] = 100
    fp1 = tmp_path / "orders.xlsx"
    wb1.save(fp1)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "产品表"
    ws2["A1"] = "产品ID"
    ws2["B1"] = "名称"
    ws2["C1"] = "价格"
    ws2["A2"] = "P001"
    ws2["B2"] = "Widget"
    ws2["C2"] = 20
    fp2 = tmp_path / "products.xlsx"
    wb2.save(fp2)

    return fp1, fp2


@pytest.fixture()
def similar_files(tmp_path: Path) -> tuple[Path, Path]:
    """两个结构相似的文件（月报 Q1 vs Q2）。"""
    for name in ("q1.xlsx", "q2.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "月报"
        ws["A1"] = "月份"
        ws["B1"] = "营收"
        ws["C1"] = "成本"
        ws["D1"] = "利润"
        ws["A2"] = "1月"
        ws["B2"] = 10000
        wb.save(tmp_path / name)
    return tmp_path / "q1.xlsx", tmp_path / "q2.xlsx"


class TestCrossFileRefScanner:
    def test_shared_columns_detected(self, two_files: tuple[Path, Path]) -> None:
        scanner = CrossFileRefScanner()
        rels = scanner.scan_pair(str(two_files[0]), str(two_files[1]))
        shared = rels.get("shared_columns", [])
        assert "产品ID" in shared

    def test_structural_similarity(self, similar_files: tuple[Path, Path]) -> None:
        scanner = CrossFileRefScanner()
        rels = scanner.scan_pair(str(similar_files[0]), str(similar_files[1]))
        assert rels.get("structural_similarity", 0) > 0.7

    def test_no_relationship(self, tmp_path: Path) -> None:
        wb1 = Workbook()
        ws1 = wb1.active
        ws1["A1"] = "x"
        ws1["B1"] = "y"
        wb1.save(tmp_path / "a.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2["A1"] = "alpha"
        ws2["B1"] = "beta"
        wb2.save(tmp_path / "b.xlsx")

        scanner = CrossFileRefScanner()
        rels = scanner.scan_pair(str(tmp_path / "a.xlsx"), str(tmp_path / "b.xlsx"))
        assert rels.get("shared_columns", []) == []
        assert rels.get("structural_similarity", 0) < 0.7
