"""高级格式工具测试：阈值图标、卡片样式、单位缩放、暗色仪表盘。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

from excelmanus.tools.advanced_format_tools import (
    apply_dashboard_dark_theme,
    apply_threshold_icon_format,
    get_tools,
    init_guard,
    scale_range_unit,
    style_card_blocks,
)


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为测试目录。"""
    init_guard(str(tmp_path))


def _rgb_suffix(value: str | None) -> str | None:
    """兼容 openpyxl 的 AARRGGBB 颜色表示，统一取 RGB 后 6 位。"""
    if value is None:
        return None
    if len(value) >= 6:
        return value[-6:]
    return value


def test_apply_threshold_icon_format_supports_dash_symbol(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册"
    ws["I2"] = 1.35
    ws["I3"] = 1.00
    ws["I4"] = 0.72
    fp = tmp_path / "icon.xlsx"
    wb.save(fp)
    wb.close()

    result = json.loads(
        apply_threshold_icon_format(
            file_path=str(fp),
            sheet_name="员工花名册",
            cell_range="I2:I4",
            high_threshold=1.2,
            mid_threshold=0.8,
            high_symbol="↑",
            mid_symbol="—",
            low_symbol="↓",
        )
    )

    assert result["status"] == "success"
    wb2 = load_workbook(fp)
    ws2 = wb2["员工花名册"]
    fmt = ws2["I2"].number_format
    wb2.close()

    assert "[>=1.2]" in fmt
    assert "[>=0.8]" in fmt
    assert "\"—\"" in fmt


def test_style_card_blocks_simulates_round_corner_and_shadow(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI仪表盘"
    ws["A1"] = "KPI"
    ws["B2"] = 123
    fp = tmp_path / "card.xlsx"
    wb.save(fp)
    wb.close()

    result = json.loads(
        style_card_blocks(
            file_path=str(fp),
            sheet_name="KPI仪表盘",
            ranges=["A1:B2"],
            card_fill_color="2E3A46",
            border_color="5A6B7A",
            shadow_color="1A2128",
        )
    )
    assert result["status"] == "success"

    wb2 = load_workbook(fp)
    ws2 = wb2["KPI仪表盘"]
    assert _rgb_suffix(ws2["A1"].fill.fgColor.rgb) == "2E3A46"
    # 左上角使用 corner_style=medium 模拟圆角
    assert ws2["A1"].border.top.style == "medium"
    assert ws2["A1"].border.left.style == "medium"
    # 阴影落在卡片右下外圈
    assert _rgb_suffix(ws2["C3"].fill.fgColor.rgb) == "1A2128"
    wb2.close()


def test_scale_range_unit_scales_values_and_formulas(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws["B2"] = 123500
    ws["B3"] = "=A3*2"
    ws["A3"] = 60000
    fp = tmp_path / "scale.xlsx"
    wb.save(fp)
    wb.close()

    result = json.loads(
        scale_range_unit(
            file_path=str(fp),
            sheet_name="销售明细",
            cell_range="B2:B3",
            divisor=10000,
            suffix="万",
            decimals=2,
        )
    )

    assert result["status"] == "success"
    assert result["numeric_scaled"] == 1
    assert result["formula_scaled"] == 1

    wb2 = load_workbook(fp)
    ws2 = wb2["销售明细"]
    assert ws2["B2"].value == pytest.approx(12.35)
    assert ws2["B3"].value == "=(A3*2)/10000"
    assert ws2["B2"].number_format == '0.00"万"'
    wb2.close()


def test_apply_dashboard_dark_theme_updates_cells_cards_and_chart(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI仪表盘"

    ws.append(["月份", "营收"])
    ws.append(["1月", 100])
    ws.append(["2月", 120])

    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "月度趋势"
    ws.add_chart(chart, "D2")

    fp = tmp_path / "dark.xlsx"
    wb.save(fp)
    wb.close()

    result = json.loads(
        apply_dashboard_dark_theme(
            file_path=str(fp),
            sheet_name="KPI仪表盘",
            card_ranges=["A1:B2"],
            metric_ranges=["B2:B3"],
            chart_style=13,
        )
    )

    assert result["status"] == "success"
    assert result["charts_total"] == 1

    wb2 = load_workbook(fp)
    ws2 = wb2["KPI仪表盘"]
    # 卡片区覆盖基础底色
    assert _rgb_suffix(ws2["A1"].fill.fgColor.rgb) == "2E3A46"
    # 指标数字高亮色
    assert _rgb_suffix(ws2["B2"].font.color.rgb) == "4FC3F7"
    chart_gp = ws2._charts[0].graphical_properties
    assert chart_gp is not None
    assert chart_gp.solidFill.srgbClr == "1B2631"
    wb2.close()


def test_get_tools_contains_advanced_format_tools() -> None:
    """Batch 3 精简：get_tools() 返回空列表。"""
    assert len(get_tools()) == 0
