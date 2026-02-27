"""Excel 内嵌图表工具测试：create_excel_chart。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excelmanus.tools.chart_tools import (
    create_excel_chart,
    init_guard,
)


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为测试目录。"""
    init_guard(str(tmp_path))


def _make_chart_data(tmp_path: Path, name: str = "chart_data.xlsx") -> Path:
    """创建含图表数据的示例 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws["A1"] = "月份"
    ws["B1"] = "营收"
    ws["C1"] = "成本"
    months = ["1月", "2月", "3月", "4月", "5月", "6月"]
    revenues = [100, 150, 120, 180, 200, 170]
    costs = [80, 90, 85, 110, 130, 100]
    for i, (m, r, c) in enumerate(zip(months, revenues, costs), start=2):
        ws[f"A{i}"] = m
        ws[f"B{i}"] = r
        ws[f"C{i}"] = c
    fp = tmp_path / name
    wb.save(fp)
    wb.close()
    return fp


class TestCreateExcelChart:
    """create_excel_chart 工具测试套件。"""

    def test_bar_chart(self, tmp_path: Path) -> None:
        """创建柱状图。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="bar",
                data_range="B1:C7",
                categories_range="A2:A7",
                title="月度营收与成本",
            )
        )
        assert result["status"] == "success"
        assert result["chart_type"] == "bar"

        wb = load_workbook(fp)
        ws = wb["数据"]
        assert len(ws._charts) == 1
        wb.close()

    def test_line_chart(self, tmp_path: Path) -> None:
        """创建折线图。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="line",
                data_range="B1:B7",
                categories_range="A2:A7",
                target_cell="E1",
            )
        )
        assert result["status"] == "success"
        assert result["target_cell"] == "E1"

    def test_pie_chart(self, tmp_path: Path) -> None:
        """创建饼图。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="pie",
                data_range="B1:B7",
                categories_range="A2:A7",
                title="营收占比",
            )
        )
        assert result["status"] == "success"

    def test_area_chart(self, tmp_path: Path) -> None:
        """创建面积图。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="area",
                data_range="B1:C7",
                categories_range="A2:A7",
            )
        )
        assert result["status"] == "success"

    def test_scatter_chart(self, tmp_path: Path) -> None:
        """创建散点图。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="scatter",
                data_range="B1:C7",
                categories_range="B2:B7",
            )
        )
        assert result["status"] == "success"

    def test_chart_on_new_target_sheet(self, tmp_path: Path) -> None:
        """图表放置到新建的目标工作表。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="bar",
                data_range="B1:B7",
                target_sheet="图表汇总",
                target_cell="A1",
            )
        )
        assert result["status"] == "success"
        assert result["target_sheet"] == "图表汇总"

        wb = load_workbook(fp)
        assert "图表汇总" in wb.sheetnames
        assert len(wb["图表汇总"]._charts) == 1
        wb.close()

    def test_chart_with_style_and_size(self, tmp_path: Path) -> None:
        """自定义图表样式和尺寸。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="line",
                data_range="B1:C7",
                style=10,
                width=20.0,
                height=12.0,
                x_title="月份",
                y_title="金额",
            )
        )
        assert result["status"] == "success"

    def test_invalid_chart_type(self, tmp_path: Path) -> None:
        """无效图表类型报错。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(str(fp), chart_type="radar", data_range="B1:B7")
        )
        assert result["status"] == "error"

    def test_from_rows_mode(self, tmp_path: Path) -> None:
        """按行读取数据系列。"""
        fp = _make_chart_data(tmp_path)
        result = json.loads(
            create_excel_chart(
                str(fp),
                chart_type="bar",
                data_range="A1:G3",
                from_rows=True,
            )
        )
        assert result["status"] == "success"
