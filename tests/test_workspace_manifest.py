"""FileRegistry 扫描功能测试（替代旧 WorkspaceManifest 测试）。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.database import Database
from excelmanus.file_registry import FileRegistry, ScanResult


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    """创建含多个 Excel 文件的临时工作区。"""
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "销售数据"
    ws1.append(["姓名", "金额", "日期"])
    ws1.append(["张三", 100, "2025-01-01"])
    wb1.save(tmp_path / "sales.xlsx")

    wb2 = Workbook()
    ws2a = wb2.active
    ws2a.title = "Sheet1"
    ws2a.append(["ID", "产品"])
    ws2b = wb2.create_sheet("Sheet2")
    ws2b.append(["类型", "数量"])
    wb2.save(tmp_path / "products.xlsx")

    sub = tmp_path / "data"
    sub.mkdir()
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "学生花名册"
    ws3.append(["学号", "姓名", "班级", "角色"])
    ws3.append(["001", "张三", "一班", "班长"])
    wb3.save(sub / "迎新活动排班表.xlsx")

    # 隐藏文件（应跳过）
    wb_hidden = Workbook()
    wb_hidden.save(tmp_path / ".hidden.xlsx")
    # 临时文件（应跳过）
    wb_temp = Workbook()
    wb_temp.save(tmp_path / "~$temp.xlsx")
    # 噪音目录中的文件（应跳过）
    noise = tmp_path / ".git"
    noise.mkdir()
    wb_noise = Workbook()
    wb_noise.save(noise / "noise.xlsx")

    return tmp_path


@pytest.fixture()
def registry(workspace: Path) -> FileRegistry:
    db = Database(str(workspace / "test.db"))
    return FileRegistry(db, workspace)


class TestScanWorkspace:
    def test_basic_scan(self, registry: FileRegistry) -> None:
        result = registry.scan_workspace(excel_only=True)
        entries = [e for e in registry.list_all() if e.file_type == "excel"]
        names = [e.original_name for e in entries]
        assert "sales.xlsx" in names
        assert "products.xlsx" in names
        assert "迎新活动排班表.xlsx" in names

    def test_skips_hidden_and_temp(self, registry: FileRegistry) -> None:
        registry.scan_workspace(excel_only=True)
        names = [e.original_name for e in registry.list_all()]
        assert ".hidden.xlsx" not in names
        assert "~$temp.xlsx" not in names

    def test_skips_noise_dirs(self, registry: FileRegistry) -> None:
        registry.scan_workspace(excel_only=True)
        names = [e.original_name for e in registry.list_all()]
        assert "noise.xlsx" not in names

    def test_sheet_metadata(self, registry: FileRegistry) -> None:
        registry.scan_workspace(excel_only=True)
        sales = registry.get_by_path("sales.xlsx")
        assert sales is not None
        assert len(sales.sheet_meta) == 1
        sheet = sales.sheet_meta[0]
        assert sheet["name"] == "销售数据"
        assert sheet["rows"] == 2
        assert sheet["columns"] == 3
        assert "姓名" in sheet["headers"]

    def test_multi_sheet(self, registry: FileRegistry) -> None:
        registry.scan_workspace(excel_only=True)
        products = registry.get_by_path("products.xlsx")
        assert products is not None
        assert len(products.sheet_meta) == 2
        names = [s["name"] for s in products.sheet_meta]
        assert "Sheet1" in names
        assert "Sheet2" in names

    def test_subdirectory_files(self, registry: FileRegistry) -> None:
        registry.scan_workspace(excel_only=True)
        nested = next(
            (e for e in registry.list_all() if e.original_name == "迎新活动排班表.xlsx"),
            None,
        )
        assert nested is not None
        assert "data" in nested.canonical_path
        assert len(nested.sheet_meta) == 1
        assert nested.sheet_meta[0]["name"] == "学生花名册"
        assert "学号" in nested.sheet_meta[0]["headers"]

    def test_max_files_limit(self, registry: FileRegistry) -> None:
        result = registry.scan_workspace(max_files=1, excel_only=True)
        assert result.total_files == 1

    def test_scan_duration_recorded(self, registry: FileRegistry) -> None:
        result = registry.scan_workspace(excel_only=True)
        assert result.scan_duration_ms >= 0

    def test_empty_workspace(self, tmp_path: Path) -> None:
        db = Database(str(tmp_path / "empty.db"))
        reg = FileRegistry(db, tmp_path)
        result = reg.scan_workspace(excel_only=True)
        assert result.total_files == 0

    def test_incremental_scan_no_changes(self, registry: FileRegistry) -> None:
        """无变化时应复用缓存。"""
        r1 = registry.scan_workspace(excel_only=True)
        r2 = registry.scan_workspace(excel_only=True)
        assert r2.cache_hits == r1.total_files
        assert r2.new_files == 0
        assert r2.updated_files == 0

    def test_incremental_detects_new_file(self, registry: FileRegistry, workspace: Path) -> None:
        """新增文件应被检测到。"""
        registry.scan_workspace(excel_only=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "新表"
        ws.append(["A", "B"])
        wb.save(workspace / "new_file.xlsx")

        r2 = registry.scan_workspace(excel_only=True)
        assert r2.new_files >= 1
        names = [e.original_name for e in registry.list_all()]
        assert "new_file.xlsx" in names

    def test_incremental_detects_modified_file(self, registry: FileRegistry, workspace: Path) -> None:
        """修改文件后应重新扫描。"""
        import time
        registry.scan_workspace(excel_only=True)
        time.sleep(0.05)
        wb = Workbook()
        ws = wb.active
        ws.title = "新销售数据"
        ws.append(["新列"])
        wb.save(workspace / "sales.xlsx")

        registry.scan_workspace(excel_only=True)
        sales = registry.get_by_path("sales.xlsx")
        assert sales is not None
        assert sales.sheet_meta[0]["name"] == "新销售数据"


class TestBuildPanorama:
    def test_empty_registry(self, tmp_path: Path) -> None:
        db = Database(str(tmp_path / "empty.db"))
        reg = FileRegistry(db, tmp_path)
        assert reg.build_panorama() == ""

    def test_full_mode_small_workspace(self, registry: FileRegistry) -> None:
        """≤20 文件应使用完整模式。"""
        registry.scan_workspace(excel_only=True)
        panorama = registry.build_panorama()
        assert "## 工作区文件全景" in panorama
        assert "sales.xlsx" in panorama
        assert "迎新活动排班表.xlsx" in panorama

    def test_compact_mode(self, tmp_path: Path) -> None:
        """21-100 文件应使用紧凑模式。"""
        for i in range(25):
            wb = Workbook()
            ws = wb.active
            ws.title = f"Sheet_{i}"
            ws.append([f"col_{i}"])
            wb.save(tmp_path / f"file_{i:03d}.xlsx")

        db = Database(str(tmp_path / "test.db"))
        reg = FileRegistry(db, tmp_path)
        reg.scan_workspace(excel_only=True)
        panorama = reg.build_panorama()
        assert "## 工作区文件全景" in panorama
        assert "Sheet_0" in panorama
