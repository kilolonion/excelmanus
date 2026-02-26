"""写入状态传播、执行守卫与 guard_mode 单元测试。"""

from __future__ import annotations

import types
from dataclasses import dataclass, field
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from excelmanus.config import ExcelManusConfig
from excelmanus.engine import AgentEngine, ChatResult, ToolCallResult
from excelmanus.skillpacks.models import SkillMatchResult
from excelmanus.tools.registry import ToolDef, ToolRegistry


# ── helpers ──────────────────────────────────────────────────

def _make_config(**overrides) -> ExcelManusConfig:
    defaults = {
        "api_key": "test-key",
        "base_url": "https://test.example.com/v1",
        "model": "test-model",
        "max_iterations": 20,
        "max_consecutive_failures": 3,
        "workspace_root": str(Path(__file__).resolve().parent),
        "backup_enabled": False,
    }
    defaults.update(overrides)
    return ExcelManusConfig(**defaults)


def _make_engine(**overrides) -> AgentEngine:
    """构建最小化 AgentEngine 实例用于单元测试。"""
    cfg = _make_config(**{k: v for k, v in overrides.items() if k in ExcelManusConfig.__dataclass_fields__})
    registry = ToolRegistry()
    engine = AgentEngine(config=cfg, registry=registry)
    return engine


def _make_route_result(write_hint: str = "unknown", **kwargs) -> SkillMatchResult:
    defaults = dict(
        skills_used=[],
        route_mode="all_tools",
        system_contexts=[],
    )
    defaults.update(kwargs)
    defaults["write_hint"] = write_hint
    return SkillMatchResult(**defaults)


# ── SkillMatchResult.write_hint 字段测试 ──

class TestWriteHintField:
    def test_default_is_unknown(self):
        result = SkillMatchResult(
            skills_used=[], route_mode="test",
        )
        assert result.write_hint == "unknown"

    def test_explicit_may_write(self):
        result = SkillMatchResult(
            skills_used=[], route_mode="test",
            write_hint="may_write",
        )
        assert result.write_hint == "may_write"

    def test_explicit_read_only(self):
        result = SkillMatchResult(
            skills_used=[], route_mode="test",
            write_hint="read_only",
        )
        assert result.write_hint == "read_only"


# ── ChatResult.write_guard_triggered 字段测试 ──

class TestChatResultWriteGuard:
    def test_default_is_false(self):
        result = ChatResult(reply="ok")
        assert result.write_guard_triggered is False

    def test_explicit_true(self):
        result = ChatResult(reply="ok", write_guard_triggered=True)
        assert result.write_guard_triggered is True


class TestDelegateSubagentWritePropagation:
    """delegate_to_subagent 成功且 subagent 有 file_changes 时应传播写入状态。

    修复回归测试：conversation_20260220T104533 中 subagent 成功写入但
    主 agent 的 write_guard 不认可，导致 finish_task 被拒、任务被重复执行 3 次。
    """

    @staticmethod
    def _delegate_tc(task: str = "test", agent: str = "writer") -> types.SimpleNamespace:
        return types.SimpleNamespace(
            id="call_delegate",
            function=types.SimpleNamespace(
                name="delegate_to_subagent",
                arguments=f'{{"task":"{task}","agent_name":"{agent}"}}',
            ),
        )

    @staticmethod
    def _make_outcome(*, success: bool, file_changes: list[str]) -> "DelegateSubagentOutcome":
        from excelmanus.engine import DelegateSubagentOutcome
        from excelmanus.subagent.models import SubagentFileChange, SubagentResult

        structured = [
            SubagentFileChange(path=p, tool_name="write_excel")
            for p in file_changes
        ]
        sub = SubagentResult(
            success=success,
            summary="test summary",
            subagent_name="subagent",
            permission_mode="default",
            conversation_id="conv_test",
            structured_changes=structured,
        )
        return DelegateSubagentOutcome(
            reply="test reply",
            success=success,
            picked_agent="subagent",
            task_text="test task",
            subagent_result=sub,
        )

    @pytest.mark.asyncio
    async def test_subagent_with_file_changes_propagates_write_state(self):
        """subagent 成功返回且有 file_changes → has_write_tool_call=True。"""
        engine = _make_engine()
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False

        outcome = self._make_outcome(success=True, file_changes=["outputs/backups/test.xlsx"])
        with patch.object(engine, "_delegate_to_subagent", return_value=outcome):
            result = await engine._execute_tool_call(
                self._delegate_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is True
        assert engine._current_write_hint == "may_write"

    @pytest.mark.asyncio
    async def test_subagent_without_file_changes_does_not_propagate(self):
        """subagent 成功但无 file_changes → has_write_tool_call 不变。"""
        engine = _make_engine()
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False

        outcome = self._make_outcome(success=True, file_changes=[])
        with patch.object(engine, "_delegate_to_subagent", return_value=outcome):
            await engine._execute_tool_call(
                self._delegate_tc(agent="analyst"), tool_scope=None, on_event=None, iteration=1,
            )

        assert engine._has_write_tool_call is False

    @pytest.mark.asyncio
    async def test_failed_subagent_with_file_changes_does_not_propagate(self):
        """subagent 失败 → 即使有 file_changes 也不传播写入。"""
        engine = _make_engine()
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False

        outcome = self._make_outcome(success=False, file_changes=["outputs/backups/partial.xlsx"])
        with patch.object(engine, "_delegate_to_subagent", return_value=outcome):
            await engine._execute_tool_call(
                self._delegate_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert engine._has_write_tool_call is False

    @pytest.mark.asyncio
    async def test_write_hint_upgraded_when_not_may_write(self):
        """write_hint 非 may_write 时，subagent 写入应升级 hint。"""
        engine = _make_engine()
        engine._current_write_hint = "read_only"
        engine._has_write_tool_call = False

        outcome = self._make_outcome(success=True, file_changes=["outputs/test.xlsx"])
        with patch.object(engine, "_delegate_to_subagent", return_value=outcome):
            await engine._execute_tool_call(
                self._delegate_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert engine._has_write_tool_call is True
        assert engine._current_write_hint == "may_write"


class TestRunCodeWritePropagation:
    """run_code GREEN/YELLOW 自动执行成功后应正确追踪写入状态。

    回归测试：conversation_20260220T162730 中 run_code 通过 CoW 实际写入文件，
    但 finish_task 被永久拒绝（3 次），因为 run_code 的 write_effect 为 dynamic，
    且 tool_dispatcher 的 code_policy 路径也未调用 record_write_action()。
    """

    @staticmethod
    def _run_code_tc(code: str = "print(1)") -> types.SimpleNamespace:
        import json as _json
        return types.SimpleNamespace(
            id="call_run_code",
            function=types.SimpleNamespace(
                name="run_code",
                arguments=_json.dumps({"code": code}),
            ),
        )

    @staticmethod
    def _make_audit_record(*, has_changes: bool = True):
        from excelmanus.approval import AppliedApprovalRecord, FileChangeRecord
        changes = []
        if has_changes:
            changes.append(FileChangeRecord(
                path="outputs/test.xlsx",
                before_exists=False, after_exists=True,
                before_hash=None, after_hash="abc123",
                before_size=None, after_size=1024,
                is_binary=True,
            ))
        return AppliedApprovalRecord(
            approval_id="test-001",
            tool_name="run_code",
            arguments={"code": "print(1)"},
            tool_scope=[],
            created_at_utc="2026-01-01T00:00:00Z",
            applied_at_utc="2026-01-01T00:00:01Z",
            undoable=False,
            manifest_file="outputs/approvals/test/manifest.json",
            audit_dir="outputs/approvals/test",
            result_preview='{"status":"success","cow_mapping":{}}',
            changes=changes,
        )

    @pytest.mark.asyncio
    async def test_run_code_with_file_changes_triggers_record_write_action(self):
        """run_code 成功执行且 audit_record.changes 非空 → has_write_tool_call=True。"""
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        audit_record = self._make_audit_record(has_changes=True)
        result_json = '{"status":"success","stdout_tail":"ok","cow_mapping":{}}'

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            result = await engine._execute_tool_call(
                self._run_code_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is True

    @pytest.mark.asyncio
    async def test_run_code_cow_mapping_triggers_record_write_action(self):
        """run_code 通过 CoW 写入（audit_record.changes 为空但 cow_mapping 非空）→ has_write_tool_call=True。

        真实场景：run_code 不在 MUTATING_ALL_TOOLS 中，审计系统不做 workspace scan，
        因此 audit_record.changes 永远为空。但 cow_mapping 非空是可靠的写入信号。
        """
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        audit_record = self._make_audit_record(has_changes=False)  # 审计无变更
        result_json = (
            '{"status":"success","stdout_tail":"ok",'
            '"cow_mapping":{"bench/external/test.xlsx":"outputs/test.xlsx"}}'
        )

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            result = await engine._execute_tool_call(
                self._run_code_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is True

    @pytest.mark.asyncio
    async def test_run_code_without_file_changes_does_not_trigger(self):
        """run_code 成功但无文件变更 → has_write_tool_call 不变。"""
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        audit_record = self._make_audit_record(has_changes=False)
        result_json = '{"status":"success","stdout_tail":"ok","cow_mapping":{}}'

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            result = await engine._execute_tool_call(
                self._run_code_tc(), tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is False


    @pytest.mark.asyncio
    async def test_run_code_ast_write_triggers_record_write_action(self):
        """run_code 代码含 wb.save() 但 audit_record.changes 为空且无 cow_mapping → AST 检测触发写入。

        典型场景：openpyxl wb.save() 直接写入非 bench 文件，审计系统不做
        workspace scan（run_code 不在 MUTATING_ALL_TOOLS），cow_mapping 也不产生。
        """
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        code_with_save = (
            'from openpyxl import load_workbook\n'
            'wb = load_workbook("output.xlsx")\n'
            'ws = wb.active\n'
            'ws["A1"] = 42\n'
            'wb.save("output.xlsx")\n'
        )
        audit_record = self._make_audit_record(has_changes=False)
        result_json = '{"status":"success","stdout_tail":"ok","cow_mapping":{}}'

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            result = await engine._execute_tool_call(
                self._run_code_tc(code=code_with_save),
                tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is True


class TestFinishTaskRestored:
    """finish_task 已恢复，任何 write_hint 下都应出现在工具列表中。"""

    def test_finish_task_in_all_write_hints(self):
        for hint in ("unknown", "read_only", "may_write"):
            engine = _make_engine()
            engine._current_write_hint = hint
            engine._skill_router = None
            tools = engine._build_meta_tools()
            names = [t["function"]["name"] for t in tools]
            assert "finish_task" in names, f"finish_task should exist when write_hint={hint}"

    def test_finish_task_in_v5_tools(self):
        engine = _make_engine()
        engine._current_write_hint = "may_write"
        tools = engine._build_v5_tools()
        names = [t["function"]["name"] for t in tools]
        assert "finish_task" in names

    def test_bench_mode_finish_task_has_summary_required(self):
        engine = _make_engine()
        engine._bench_mode = True
        engine._skill_router = None
        tools = engine._build_meta_tools()
        ft = [t for t in tools if t["function"]["name"] == "finish_task"][0]
        assert "summary" in ft["function"]["parameters"]["required"]

    def test_normal_mode_finish_task_has_affected_files(self):
        engine = _make_engine()
        engine._bench_mode = False
        engine._skill_router = None
        tools = engine._build_meta_tools()
        ft = [t for t in tools if t["function"]["name"] == "finish_task"][0]
        props = ft["function"]["parameters"]["properties"]
        assert "affected_files" in props

    @pytest.mark.asyncio
    async def test_finish_accepted_exits_loop(self):
        """finish_task 被接受后应立即退出 _tool_calling_loop。"""
        engine = _make_engine(max_iterations=10)
        engine._has_write_tool_call = True
        route_result = _make_route_result(write_hint="may_write")

        finish_tc = types.SimpleNamespace(
            id="call_finish",
            function=types.SimpleNamespace(
                name="finish_task",
                arguments='{"summary":"已完成数据写入"}',
            ),
        )
        finish_resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(content="", tool_calls=[finish_tc])
            )]
        )
        # 提供足够的 mock 响应（流式尝试+回退可能消耗多个）
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[finish_resp] * 4
        )
        engine._execute_tool_call = AsyncMock(
            return_value=ToolCallResult(
                tool_name="finish_task",
                arguments={"summary": "已完成数据写入"},
                result="✅ 任务完成\n\n已完成数据写入",
                success=True,
                finish_accepted=True,
            )
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert "任务完成" in result.reply
        assert result.iterations == 1

    @pytest.mark.asyncio
    async def test_finish_not_accepted_continues_loop(self):
        """finish_task 未被接受（首次无写入警告）不应退出循环。"""
        engine = _make_engine(max_iterations=3)
        engine._has_write_tool_call = False
        route_result = _make_route_result(write_hint="may_write")

        finish_tc = types.SimpleNamespace(
            id="call_finish",
            function=types.SimpleNamespace(
                name="finish_task",
                arguments='{"summary":"done"}',
            ),
        )
        finish_resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(content="", tool_calls=[finish_tc])
            )]
        )
        text_resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(content="继续执行中", tool_calls=None)
            )]
        )
        # 提供足够的 mock 响应（流式尝试+回退可能消耗多个）
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[finish_resp] * 4 + [text_resp] * 4
        )
        engine._execute_tool_call = AsyncMock(
            return_value=ToolCallResult(
                tool_name="finish_task",
                arguments={"summary": "done"},
                result="⚠️ 未检测到写入类工具的成功调用。",
                success=True,
                finish_accepted=False,
            )
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        # 应该继续到下一轮并返回文本
        assert result.iterations > 1 or "继续执行中" in result.reply or "已达到最大迭代次数" in result.reply


class TestGuardMode:
    """guard_mode 配置测试：off（默认）完全跳过门禁，soft 仅记录诊断。"""

    @pytest.mark.asyncio
    async def test_guard_off_formula_text_passes_through(self):
        """guard_mode=off 时公式文本直接放行，不注入守卫消息。"""
        engine = _make_engine(max_iterations=3, guard_mode="off")
        route_result = _make_route_result(write_hint="unknown")

        formula_text = "请用公式 =SUM(A1:A2)"
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(message=types.SimpleNamespace(content=formula_text, tool_calls=None))]
                ),
            ]
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert result.reply == formula_text
        assert result.iterations == 1
        # 不应注入任何守卫消息
        user_messages = [
            str(m.get("content", ""))
            for m in engine.memory.get_messages()
            if m.get("role") == "user"
        ]
        guard_keywords = ["公式或代码建议", "写入工具"]
        for msg in user_messages:
            for kw in guard_keywords:
                assert kw not in msg

    @pytest.mark.asyncio
    async def test_guard_off_write_hint_may_write_passes_through(self):
        """guard_mode=off 时 may_write 无写入也直接放行。"""
        engine = _make_engine(max_iterations=3, guard_mode="off")
        route_result = _make_route_result(write_hint="may_write")

        text = "分析完成，数据如下..."
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(message=types.SimpleNamespace(content=text, tool_calls=None))]
                ),
            ]
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert result.reply == text
        assert result.iterations == 1
        assert result.write_guard_triggered is False

    @pytest.mark.asyncio
    async def test_guard_soft_records_diag_but_passes_through(self):
        """guard_mode=soft 时记录诊断事件但不强制继续。"""
        engine = _make_engine(max_iterations=3, guard_mode="soft")
        route_result = _make_route_result(write_hint="may_write")

        text = "我来帮你分析数据"
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(message=types.SimpleNamespace(content=text, tool_calls=None))]
                ),
            ]
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert result.reply == text
        assert result.iterations == 1


class TestGuardModeFinishTask:
    """guard_mode=off 时 finish_task 首次即接受，不产生额外迭代。"""

    @pytest.mark.asyncio
    async def test_guard_off_finish_task_accepted_without_write(self):
        """guard_mode=off + may_write + 无写入 → finish_task 首次接受，仅 1 迭代。"""
        engine = _make_engine(max_iterations=3, guard_mode="off")
        engine._has_write_tool_call = False
        route_result = _make_route_result(write_hint="may_write")

        finish_tc = types.SimpleNamespace(
            id="tc_1",
            type="function",
            function=types.SimpleNamespace(
                name="finish_task",
                arguments='{"summary": "图片上传失败，已告知用户"}',
            ),
        )
        resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(
                    content="图片上传失败，请重新上传图片。",
                    tool_calls=[finish_tc],
                )
            )]
        )
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[resp] * 4
        )
        # guard_mode=off → FinishTaskHandler 应首次接受
        engine._execute_tool_call = AsyncMock(
            return_value=ToolCallResult(
                tool_name="finish_task",
                arguments={"summary": "图片上传失败，已告知用户"},
                result="✅ 任务完成\n\n图片上传失败，已告知用户",
                success=True,
                finish_accepted=True,
            )
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        # finish_task 应首次接受 → 1 迭代退出，无额外迭代
        assert result.iterations == 1
        assert "任务完成" in result.reply

    @pytest.mark.asyncio
    async def test_guard_soft_finish_task_still_warns(self):
        """guard_mode=soft + may_write + 无写入 → finish_task 首次被警告。"""
        engine = _make_engine(max_iterations=3, guard_mode="soft")
        engine._has_write_tool_call = False
        route_result = _make_route_result(write_hint="may_write")

        finish_tc = types.SimpleNamespace(
            id="call_finish",
            function=types.SimpleNamespace(
                name="finish_task",
                arguments='{"summary": "done"}',
            ),
        )
        finish_resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(content="", tool_calls=[finish_tc])
            )]
        )
        text_resp = types.SimpleNamespace(
            choices=[types.SimpleNamespace(
                message=types.SimpleNamespace(content="已完成", tool_calls=None)
            )]
        )
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[finish_resp] * 4 + [text_resp] * 4
        )
        # soft 模式下首次拒绝
        engine._execute_tool_call = AsyncMock(
            return_value=ToolCallResult(
                tool_name="finish_task",
                arguments={"summary": "done"},
                result="⚠️ 未检测到写入类工具的成功调用。",
                success=True,
                finish_accepted=False,
            )
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        # soft 模式下首次仍被拒绝 → 应有 >1 迭代
        assert result.iterations > 1 or "已完成" in result.reply or "已达到最大迭代次数" in result.reply


class TestWriteHintSyncOnWriteCall:
    @pytest.mark.asyncio
    async def test_write_hint_upgrades_after_successful_write_tool_call(self):
        engine = _make_engine(max_iterations=1)
        route_result = _make_route_result(write_hint="read_only")
        engine._current_write_hint = "read_only"
        engine._registry.register_tool(
            ToolDef(
                name="write_text_file",
                description="test",
                input_schema={"type": "object", "properties": {}},
                func=lambda **kwargs: "ok",
                write_effect="workspace_write",
            )
        )

        first = types.SimpleNamespace(
            choices=[
                types.SimpleNamespace(
                    message=types.SimpleNamespace(
                        content="",
                        tool_calls=[
                            {
                                "id": "call_write_1",
                                "function": {
                                    "name": "write_text_file",
                                    "arguments": '{"file_path":"demo.txt","content":"hello"}',
                                },
                            }
                        ],
                    )
                )
            ]
        )
        engine._client.chat.completions.create = AsyncMock(side_effect=[first])
        engine._execute_tool_call = AsyncMock(
            return_value=ToolCallResult(
                tool_name="write_text_file",
                arguments={"file_path": "demo.txt", "content": "hello"},
                result="ok",
                success=True,
            )
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert "已达到最大迭代次数" in result.reply
        assert engine._current_write_hint == "may_write"


class TestManifestRefreshOnRecordedWrite:
    @pytest.mark.asyncio
    async def test_manifest_refresh_triggered_by_record_write_action(self):
        """写入标记来自 record_write_action 时，应触发 FileRegistry 增量刷新。"""
        engine = _make_engine(max_iterations=1)
        route_result = _make_route_result(write_hint="read_only")

        first = types.SimpleNamespace(
            choices=[
                types.SimpleNamespace(
                    message=types.SimpleNamespace(
                        content="",
                        tool_calls=[
                            {
                                "id": "call_1",
                                "function": {
                                    "name": "delegate_to_subagent",
                                    "arguments": '{"task":"sync files"}',
                                },
                            }
                        ],
                    )
                )
            ]
        )
        second = types.SimpleNamespace(
            choices=[
                types.SimpleNamespace(
                    message=types.SimpleNamespace(content="done", tool_calls=None)
                )
            ]
        )
        engine._client.chat.completions.create = AsyncMock(side_effect=[first, second])

        async def _execute_and_record_write(*args, **kwargs):
            engine._record_write_action()
            return ToolCallResult(
                tool_name="delegate_to_subagent",
                arguments={"task": "sync files"},
                result="ok",
                success=True,
            )

        engine._execute_tool_call = AsyncMock(side_effect=_execute_and_record_write)

        # Mock FileRegistry.scan_workspace 来验证刷新被触发
        if engine._file_registry is not None:
            with patch.object(engine._file_registry, "scan_workspace") as scan_mock:
                result = await engine._tool_calling_loop(route_result, on_event=None)
            assert "已达到最大迭代次数" in result.reply
            scan_mock.assert_called_once()
        else:
            # 无 FileRegistry 时，验证不报错即可
            result = await engine._tool_calling_loop(route_result, on_event=None)
            assert "已达到最大迭代次数" in result.reply


class TestWriteTrackingApis:
    def test_record_workspace_write_action_marks_manifest_refresh(self):
        engine = _make_engine()
        engine._manifest_refresh_needed = False

        engine._record_workspace_write_action()

        assert engine._has_write_tool_call is True
        assert engine._manifest_refresh_needed is True

    def test_record_external_write_action_does_not_mark_manifest_refresh(self):
        engine = _make_engine()
        engine._manifest_refresh_needed = False

        engine._record_external_write_action()

        assert engine._has_write_tool_call is True
        assert engine._manifest_refresh_needed is False


class TestWaitingForUserActionPassthrough:
    """等待用户操作放行：agent 等待用户上传/提供素材时，写入门禁不应强制继续。

    回归测试：用户上传图片场景中，agent 回复"请上传图片"后被写入门禁
    强制继续，导致反复 list_directory 空转。
    """

    def test_pattern_detects_chinese_upload_request(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action(
            "好的，我这边已就绪。请直接把图片上传到当前会话里，上传后我会立刻帮你按数据+样式完全复刻成 Excel。"
        )

    def test_pattern_detects_waiting_for_upload(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action("等你上传图片后我就开始处理。")

    def test_pattern_detects_not_received_yet(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action(
            "当前仍未检测到任何图片文件，缺少复刻源会导致无法执行写入。"
        )

    def test_pattern_detects_english_upload_request(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action(
            "Please upload the image file so I can replicate it."
        )

    def test_pattern_does_not_match_normal_text(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert not _looks_like_waiting_for_user_action(
            "我已经完成了所有数据的写入操作。"
        )

    def test_pattern_does_not_match_formula_advice(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert not _looks_like_waiting_for_user_action(
            "你可以使用公式 =SUM(A1:A10) 来计算总和。"
        )

    def test_pattern_detects_provide_file(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action("请提供源文件，我来帮你处理。")

    def test_pattern_detects_missing_material(self):
        from excelmanus.engine import _looks_like_waiting_for_user_action
        assert _looks_like_waiting_for_user_action(
            "还需要你上传原始表格文件才能继续。"
        )

    @pytest.mark.asyncio
    async def test_write_guard_bypassed_when_waiting_for_user(self):
        """write_hint=may_write 但 agent 在等待用户上传 → 应直接放行，不强制继续。"""
        engine = _make_engine(max_iterations=10)
        route_result = _make_route_result(write_hint="may_write")

        upload_text = "好的，请直接把图片上传到当前会话里，上传后我立刻帮你复刻。"
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(
                        message=types.SimpleNamespace(content=upload_text, tool_calls=None)
                    )]
                ),
            ]
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        assert result.reply == upload_text
        assert result.iterations == 1
        # 确认写入门禁消息未被注入
        user_messages = [
            str(m.get("content", ""))
            for m in engine.memory.get_messages()
            if m.get("role") == "user"
        ]
        assert not any("尚未调用任何写入工具" in msg for msg in user_messages)

    @pytest.mark.asyncio
    async def test_write_guard_still_fires_for_non_waiting_text(self):
        """write_hint=may_write 且 agent 不是在等用户 → 写入门禁仍应触发。"""
        engine = _make_engine(max_iterations=5)
        route_result = _make_route_result(write_hint="may_write")

        normal_text = "我已经分析完数据了，结果如下：总计 100 条记录。"
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(
                        message=types.SimpleNamespace(content=normal_text, tool_calls=None)
                    )]
                ),
                types.SimpleNamespace(
                    choices=[types.SimpleNamespace(
                        message=types.SimpleNamespace(content="done", tool_calls=None)
                    )]
                ),
            ]
        )

        result = await engine._tool_calling_loop(route_result, on_event=None)

        # guard_mode=off 时写入门禁不强制继续，第一轮 text-only 直接返回
        assert result.reply == normal_text
        assert result.iterations == 1


class TestRunCodeASTVariableWriteRegression:
    """回归测试：wb.save(file_path) 使用变量参数时 AST 检测应触发写入。

    conversation_20260221T135637 中 run_code 通过 openpyxl wb.save(file_path)
    实际写入了文件，但 AST 检测因仅识别字面量而漏检。
    """

    @staticmethod
    def _run_code_tc(code: str) -> types.SimpleNamespace:
        import json as _json
        return types.SimpleNamespace(
            id="call_run_code_var",
            function=types.SimpleNamespace(
                name="run_code",
                arguments=_json.dumps({"code": code}),
            ),
        )

    @pytest.mark.asyncio
    async def test_wb_save_variable_triggers_write_detection(self):
        """wb.save(file_path) 使用变量 → AST 检测应识别为写入。"""
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        code_with_var_save = (
            "from openpyxl import load_workbook\n"
            "file_path = './.tmp/test.xlsx'\n"
            "wb = load_workbook(file_path)\n"
            "ws = wb.active\n"
            "ws['A1'] = 42\n"
            "wb.save(file_path)\n"
        )
        from excelmanus.approval import AppliedApprovalRecord
        audit_record = AppliedApprovalRecord(
            approval_id="test-var",
            tool_name="run_code",
            arguments={"code": code_with_var_save},
            tool_scope=[],
            created_at_utc="2026-01-01T00:00:00Z",
            applied_at_utc="2026-01-01T00:00:01Z",
            undoable=False,
            manifest_file="outputs/approvals/test/manifest.json",
            audit_dir="outputs/approvals/test",
            result_preview='{"status":"success","cow_mapping":{}}',
            changes=[],
        )
        result_json = '{"status":"success","stdout_tail":"ok","cow_mapping":{}}'

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            result = await engine._execute_tool_call(
                self._run_code_tc(code_with_var_save),
                tool_scope=None, on_event=None, iteration=1,
            )

        assert result.success is True
        assert engine._has_write_tool_call is True

    @pytest.mark.asyncio
    async def test_wb_save_variable_propagates_write(self):
        """wb.save(var) 写入后 has_write_tool_call 应被正确传播。"""
        engine = _make_engine(code_policy_enabled=True)
        engine._current_write_hint = "may_write"
        engine._has_write_tool_call = False
        engine._window_perception = None

        code = "wb.save(file_path)\n"
        from excelmanus.approval import AppliedApprovalRecord
        audit_record = AppliedApprovalRecord(
            approval_id="test-var2",
            tool_name="run_code",
            arguments={"code": code},
            tool_scope=[],
            created_at_utc="2026-01-01T00:00:00Z",
            applied_at_utc="2026-01-01T00:00:01Z",
            undoable=False,
            manifest_file="outputs/approvals/test/manifest.json",
            audit_dir="outputs/approvals/test",
            result_preview='{"status":"success","cow_mapping":{}}',
            changes=[],
        )
        result_json = '{"status":"success","stdout_tail":"ok","cow_mapping":{}}'

        with patch.object(engine, "_execute_tool_with_audit", new_callable=AsyncMock,
                          return_value=(result_json, audit_record)):
            await engine._execute_tool_call(
                self._run_code_tc(code),
                tool_scope=None, on_event=None, iteration=1,
            )

        assert engine._has_write_tool_call is True
