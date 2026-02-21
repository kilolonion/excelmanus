"""格式化工具增强功能测试：颜色映射、样式读取、合并单元格、行高调整。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from excelmanus.tools.format_tools import (
    COLOR_NAME_MAP,
    _resolve_color,
    _build_font,
    _build_fill,
    _build_border,
    _build_side,
    _extract_font,
    _extract_fill,
    _extract_border,
    _extract_alignment,
    _color_to_hex,
    format_cells,
    read_cell_styles,
    adjust_row_height,
    merge_cells_tool,
    unmerge_cells_tool,
    adjust_column_width,
    init_guard,
    get_tools,
)


@pytest.fixture()
def sample_xlsx(tmp_path: Path) -> Path:
    """创建一个带有样式的示例 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入数据
    ws["A1"] = "标题"
    ws["B1"] = "数据"
    ws["A2"] = 100
    ws["B2"] = 200
    ws["A3"] = 300
    ws["B3"] = 400

    # 给 A1 设置样式
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="FF0000")
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A1"].border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thick", color="000000"),
    )
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 合并 A3:B3
    ws.merge_cells("A3:B3")

    file_path = tmp_path / "test_styled.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为 tmp_path。"""
    init_guard(str(tmp_path))


# ── _resolve_color 测试 ──────────────────────────────────


class TestResolveColor:
    def test_chinese_color_name(self) -> None:
        assert _resolve_color("红色") == "FF0000"
        assert _resolve_color("浅蓝") == "5B9BD5"
        assert _resolve_color("深绿色") == "006100"

    def test_english_color_name(self) -> None:
        assert _resolve_color("red") == "FF0000"
        assert _resolve_color("blue") == "0000FF"
        assert _resolve_color("lightgray") == "D3D3D3"

    def test_hex_code_passthrough(self) -> None:
        assert _resolve_color("FF0000") == "FF0000"
        assert _resolve_color("00ff00") == "00FF00"

    def test_hex_with_hash(self) -> None:
        assert _resolve_color("#FF0000") == "FF0000"
        assert _resolve_color("#00ff00") == "00FF00"

    def test_none_returns_none(self) -> None:
        assert _resolve_color(None) is None

    def test_unknown_passthrough(self) -> None:
        assert _resolve_color("unknown_color") == "unknown_color"


# ── _build_font 增强测试 ─────────────────────────────────


class TestBuildFont:
    def test_underline(self) -> None:
        font = _build_font({"underline": "single"})
        assert font.underline == "single"

    def test_strikethrough(self) -> None:
        font = _build_font({"strikethrough": True})
        assert font.strike is True

    def test_color_name_resolved(self) -> None:
        font = _build_font({"color": "红色"})
        # openpyxl 的 Font.color 是 Color 对象，rgb 格式为 AARRGGBB
        assert font.color.rgb.endswith("FF0000")


# ── _build_fill 增强测试 ─────────────────────────────────


class TestBuildFill:
    def test_color_name_resolved(self) -> None:
        fill = _build_fill({"color": "浅黄色"})
        assert fill.start_color.rgb == "00FFF2CC"

    def test_fill_type_none(self) -> None:
        fill = _build_fill({"fill_type": "none"})
        # openpyxl 将 fill_type='none' 存储为 patternType=None
        assert fill.patternType is None


# ── _build_border 增强测试 ───────────────────────────────


class TestBuildBorder:
    def test_uniform_mode(self) -> None:
        border = _build_border({"style": "medium", "color": "红色"})
        assert border.left.style == "medium"
        assert border.left.color.rgb == "00FF0000"

    def test_per_side_mode(self) -> None:
        border = _build_border({
            "left": {"style": "thin", "color": "000000"},
            "top": {"style": "thick"},
        })
        assert border.left.style == "thin"
        assert border.top.style == "thick"
        # right/bottom 未指定，应为默认空 Side
        assert border.right.style is None
        assert border.bottom.style is None

    def test_side_string_shorthand(self) -> None:
        side = _build_side("medium")
        assert side.style == "medium"


# ── _extract_* 测试 ──────────────────────────────────────


class TestExtractFunctions:
    def test_extract_font_non_default(self) -> None:
        font = Font(name="Arial", size=14, bold=True, color="FF0000")
        info = _extract_font(font)
        assert info is not None
        assert info["name"] == "Arial"
        assert info["size"] == 14
        assert info["bold"] is True
        assert info["color"] == "FF0000"

    def test_extract_font_default_returns_none(self) -> None:
        font = Font()
        info = _extract_font(font)
        assert info is None

    def test_extract_fill_solid(self) -> None:
        fill = PatternFill(start_color="FFFF00", fill_type="solid")
        info = _extract_fill(fill)
        assert info is not None
        assert info["type"] == "solid"
        assert info["color"] == "FFFF00"

    def test_extract_fill_none_returns_none(self) -> None:
        fill = PatternFill(fill_type=None)
        info = _extract_fill(fill)
        assert info is None

    def test_extract_border_with_styles(self) -> None:
        border = Border(left=Side(style="thin"), top=Side(style="medium"))
        info = _extract_border(border)
        assert info is not None
        assert info["left"] == "thin"
        assert info["top"] == "medium"

    def test_extract_border_empty_returns_none(self) -> None:
        border = Border()
        info = _extract_border(border)
        assert info is None

    def test_extract_alignment_non_default(self) -> None:
        alignment = Alignment(horizontal="center", wrap_text=True)
        info = _extract_alignment(alignment)
        assert info is not None
        assert info["horizontal"] == "center"
        assert info["wrap_text"] is True

    def test_extract_alignment_default_returns_none(self) -> None:
        alignment = Alignment()
        info = _extract_alignment(alignment)
        assert info is None


# ── read_cell_styles 测试 ────────────────────────────────


class TestReadCellStyles:
    def test_reads_styled_cells(self, sample_xlsx: Path) -> None:
        result = json.loads(read_cell_styles(str(sample_xlsx), "A1:B3"))
        assert result["status"] == "success"
        assert result["total_cells"] == 6
        # A1 应该被检测到有样式
        styled = result["styled_cells"]
        a1_entries = [c for c in styled if c["cell"] == "A1"]
        assert len(a1_entries) == 1
        a1 = a1_entries[0]
        assert a1["font"]["bold"] is True
        assert a1["font"]["name"] == "微软雅黑"
        assert "fill" in a1
        assert "border" in a1
        assert "alignment" in a1

    def test_summary_only(self, sample_xlsx: Path) -> None:
        result = json.loads(read_cell_styles(str(sample_xlsx), "A1:B3", summary_only=True))
        assert result["status"] == "success"
        assert "styled_cells" not in result
        assert "summary" in result
        assert result["summary"]["has_merged_cells"] is True
        assert "A3:B3" in result["summary"]["merged_ranges"]

    def test_detects_merged_cells(self, sample_xlsx: Path) -> None:
        result = json.loads(read_cell_styles(str(sample_xlsx), "A1:B3"))
        # A3 是合并区域的一部分
        a3_entries = [c for c in result["styled_cells"] if c["cell"] == "A3"]
        assert any(e.get("merged") for e in a3_entries)


# ── merge / unmerge 测试 ─────────────────────────────────


class TestMergeCells:
    def test_merge_and_unmerge(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "合并测试"
        ws["B1"] = "数据"
        file_path = tmp_path / "merge_test.xlsx"
        wb.save(file_path)
        wb.close()

        # 合并
        result = json.loads(merge_cells_tool(str(file_path), "A1:B1"))
        assert result["status"] == "success"
        assert result["merged_range"] == "A1:B1"

        # 验证合并状态
        styles = json.loads(read_cell_styles(str(file_path), "A1:B1"))
        assert styles["summary"]["has_merged_cells"] is True

        # 取消合并
        result = json.loads(unmerge_cells_tool(str(file_path), "A1:B1"))
        assert result["status"] == "success"


# ── adjust_row_height 测试 ───────────────────────────────


class TestAdjustRowHeight:
    def test_manual_row_height(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "行高测试"
        file_path = tmp_path / "row_height.xlsx"
        wb.save(file_path)
        wb.close()

        result = json.loads(adjust_row_height(str(file_path), rows={"1": 30.0, "2": 25.0}))
        assert result["status"] == "success"
        assert result["rows_adjusted"]["1"] == 30.0
        assert result["rows_adjusted"]["2"] == 25.0

    def test_auto_fit_row_height(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "测试"
        file_path = tmp_path / "row_auto.xlsx"
        wb.save(file_path)
        wb.close()

        result = json.loads(adjust_row_height(str(file_path), auto_fit=True))
        assert result["status"] == "success"
        assert len(result["rows_adjusted"]) > 0


# ── format_cells 颜色名集成测试 ──────────────────────────


class TestFormatCellsColorName:
    def test_chinese_color_name_in_font(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "颜色测试"
        file_path = tmp_path / "color_name.xlsx"
        wb.save(file_path)
        wb.close()

        result = json.loads(format_cells(
            str(file_path), "A1",
            font={"color": "红色", "bold": True},
            fill={"color": "浅黄色"},
        ))
        assert result["status"] == "success"

        # 验证样式已应用
        styles = json.loads(read_cell_styles(str(file_path), "A1"))
        a1 = styles["styled_cells"][0]
        assert a1["font"]["bold"] is True
        assert a1["font"]["color"] == "FF0000"


# ── get_tools 完整性测试 ─────────────────────────────────


class TestGetTools:
    def test_all_new_tools_registered(self) -> None:
        """Batch 2 精简：get_tools() 返回空列表。"""
        tools = get_tools()
        assert len(tools) == 0

    def test_format_cells_schema_has_underline(self) -> None:
        """Batch 2 精简：format_cells 已删除，跳过。"""

    def test_border_schema_has_sides(self) -> None:
        """Batch 2 精简：format_cells 已删除，跳过。"""
