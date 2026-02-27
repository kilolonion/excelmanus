"""read_excel include 参数测试：验证按需维度查询功能。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from excelmanus.tools import data_tools


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为 tmp_path，允许临时目录访问。"""
    data_tools.init_guard(str(tmp_path))


@pytest.fixture()
def styled_xlsx(tmp_path: Path) -> Path:
    """创建一个包含多种样式、图表、冻结窗格等的测试 Excel 文件。"""
    fp = tmp_path / "styled.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头行 — 加粗、居中、蓝底
    headers = ["月份", "销售额", "成本", "利润"]
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    data_rows = [
        ["1月", 10000, 6000, 4000],
        ["2月", 12000, 7000, 5000],
        ["3月", 15000, 8000, 7000],
        ["4月", 11000, 6500, 4500],
        ["5月", 13000, 7500, 5500],
    ]
    data_font = Font(size=11)
    num_fmt = "#,##0"
    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            if c_idx >= 2:
                cell.number_format = num_fmt

    # 利润列公式（第6行开始）
    ws.cell(row=7, column=1, value="6月")
    ws.cell(row=7, column=2, value=14000)
    ws.cell(row=7, column=3, value=8500)
    ws.cell(row=7, column=4, value="=B7-C7")

    # 冻结窗格
    ws.freeze_panes = "A2"

    # 列宽
    ws.column_dimensions["A"].width = 12.0
    ws.column_dimensions["B"].width = 15.0

    # 合并单元格（标题区域）
    ws["F1"] = "年度汇总"
    ws.merge_cells("F1:H1")

    # 边框
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # 条件格式 — 简单的 cellIs 规则
    from openpyxl.formatting.rule import CellIsRule

    ws.conditional_formatting.add(
        "D2:D7",
        CellIsRule(operator="greaterThan", formula=["5000"], fill=PatternFill(bgColor="C6EFCE")),
    )

    # 数据验证
    dv = DataValidation(type="list", formula1='"1月,2月,3月,4月,5月,6月"', allow_blank=True)
    dv.add("A8:A20")
    ws.add_data_validation(dv)

    # 打印设置
    ws.print_area = "A1:D7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.print_title_rows = "1:1"

    # 嵌入图表
    chart = BarChart()
    chart.title = "月度销售"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=7), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=7))
    ws.add_chart(chart, "F3")

    wb.save(fp)
    wb.close()
    return fp


@pytest.fixture()
def simple_xlsx(tmp_path: Path) -> Path:
    """最简单的 Excel 文件，用于回归测试。"""
    fp = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名称"
    ws["B1"] = "数量"
    ws["A2"] = "苹果"
    ws["B2"] = 10
    ws["A3"] = "香蕉"
    ws["B3"] = 20
    wb.save(fp)
    wb.close()
    return fp


# ── 回归测试：无 include 时行为不变 ──────────────────────

class TestReadExcelRegression:
    """确保无 include 时返回与旧版一致。"""

    def test_basic_read_no_include(self, simple_xlsx: Path) -> None:
        result = json.loads(data_tools.read_excel(str(simple_xlsx)))
        assert result["shape"] == {"rows": 2, "columns": 2}
        assert result["columns"] == ["名称", "数量"]
        assert "styles" not in result
        assert "charts" not in result
        assert "freeze_panes" not in result

    def test_include_style_summary_backward_compat(self, styled_xlsx: Path) -> None:
        """include_style_summary=True 应触发 styles 维度。"""
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include_style_summary=True)
        )
        # 向后兼容：应返回 styles 键（压缩样式类格式）
        assert "styles" in result
        assert "style_classes" in result["styles"]
        assert "cell_style_map" in result["styles"]

    def test_tool_def_uses_higher_result_cap(self) -> None:
        """read_excel ToolDef 应保留更高截断上限，减少关键预览信息丢失。"""
        tools = {tool.name: tool for tool in data_tools.get_tools()}
        assert tools["read_excel"].max_result_chars == 6000


# ── include=["styles"] ──────────────────────────────────

class TestIncludeStyles:
    def test_styles_returns_compressed_classes(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["styles"])
        )
        styles = result["styles"]
        assert "style_classes" in styles
        assert "cell_style_map" in styles
        assert "merged_ranges" in styles

        # 至少有表头样式类
        assert len(styles["style_classes"]) >= 1
        # cell_style_map 不为空
        assert len(styles["cell_style_map"]) >= 1
        # 合并范围应包含 F1:H1
        assert any("F1" in mr for mr in styles["merged_ranges"])

    def test_style_classes_contain_font_info(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["styles"])
        )
        classes = result["styles"]["style_classes"]
        # 应至少有一个类包含 font.bold
        has_bold = any(
            cls.get("font", {}).get("bold") is True
            for cls in classes.values()
        )
        assert has_bold, f"未找到 bold 样式类: {classes}"


# ── include=["charts"] ──────────────────────────────────

class TestIncludeCharts:
    def test_charts_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["charts"])
        )
        assert "charts" in result
        assert len(result["charts"]) == 1
        chart = result["charts"][0]
        assert chart["type"] == "bar"
        assert "title" in chart or "series_count" in chart

    def test_no_charts_returns_empty_list(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["charts"])
        )
        assert result["charts"] == []


# ── include=["freeze_panes"] ────────────────────────────

class TestIncludeFreezePanes:
    def test_freeze_panes_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["freeze_panes"])
        )
        assert result["freeze_panes"] == "A2"

    def test_no_freeze_panes(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["freeze_panes"])
        )
        assert result["freeze_panes"] is None


# ── include=["conditional_formatting"] ──────────────────

class TestIncludeConditionalFormatting:
    def test_conditional_formatting_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["conditional_formatting"])
        )
        assert "conditional_formatting" in result
        cf = result["conditional_formatting"]
        assert len(cf) >= 1
        assert "type" in cf[0]

    def test_no_conditional_formatting(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["conditional_formatting"])
        )
        assert result["conditional_formatting"] == []


# ── include=["data_validation"] ─────────────────────────

class TestIncludeDataValidation:
    def test_data_validation_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["data_validation"])
        )
        assert "data_validation" in result
        dv = result["data_validation"]
        assert len(dv) >= 1
        assert dv[0]["type"] == "list"

    def test_no_data_validation(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["data_validation"])
        )
        assert result["data_validation"] == []


# ── include=["print_settings"] ──────────────────────────

class TestIncludePrintSettings:
    def test_print_settings_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["print_settings"])
        )
        ps = result["print_settings"]
        assert "print_area" in ps
        assert ps["orientation"] == "landscape"

    def test_no_print_settings(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["print_settings"])
        )
        # 简单文件应返回空或很少的设置
        assert isinstance(result["print_settings"], dict)


# ── include=["column_widths"] ───────────────────────────

class TestIncludeColumnWidths:
    def test_column_widths_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["column_widths"])
        )
        widths = result["column_widths"]
        assert "A" in widths
        assert widths["A"] == 12.0
        assert "B" in widths
        assert widths["B"] == 15.0


# ── include=["formulas"] ────────────────────────────────

class TestIncludeFormulas:
    def test_formulas_detected(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(styled_xlsx), include=["formulas"])
        )
        formulas = result["formulas"]
        assert len(formulas) >= 1
        # 应检测到 D7 的公式
        d7 = [f for f in formulas if f["cell"] == "D7"]
        assert len(d7) == 1
        assert d7[0]["formula"] == "=B7-C7"


# ── 多维度组合 ──────────────────────────────────────────

class TestMultiDimensions:
    def test_multiple_include_dimensions(self, styled_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(
                str(styled_xlsx),
                include=["styles", "charts", "freeze_panes", "column_widths"],
            )
        )
        assert "styles" in result
        assert "charts" in result
        assert "freeze_panes" in result
        assert "column_widths" in result
        # 基础信息始终存在
        assert "columns" in result
        assert "shape" in result
        assert "preview" in result


# ── 无效维度警告 ─────────────────────────────────────────

class TestInvalidDimension:
    def test_unknown_dimension_warning(self, simple_xlsx: Path) -> None:
        result = json.loads(
            data_tools.read_excel(str(simple_xlsx), include=["nonexistent_dim"])
        )
        assert "include_warning" in result
        assert "nonexistent_dim" in result["include_warning"]
