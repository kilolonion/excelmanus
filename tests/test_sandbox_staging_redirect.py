"""sandbox_hook staging 映射重定向测试。

验证 EXCELMANUS_STAGING_MAP 环境变量注入后，sandbox_hook 的
_guarded_open 和 _patch_openpyxl_save 能正确将文件读写重定向到
transaction staging 副本。
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

import pytest

from excelmanus.security.sandbox_hook import generate_wrapper_script


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    return tmp_path


def _run_in_sandbox(
    workspace: Path,
    script_content: str,
    tier: str,
    *,
    staging_map: dict[str, str] | None = None,
) -> subprocess.CompletedProcess:
    """在沙盒 wrapper 中执行脚本，可选注入 staging 映射。"""
    script = workspace / "test_script.py"
    script.write_text(script_content, encoding="utf-8")
    wrapper = generate_wrapper_script(tier, str(workspace))
    wrapper_path = workspace / "_wrapper.py"
    wrapper_path.write_text(wrapper, encoding="utf-8")
    env = os.environ.copy()
    if staging_map:
        env["EXCELMANUS_STAGING_MAP"] = json.dumps(staging_map)
    return subprocess.run(
        [sys.executable, str(wrapper_path), str(script)],
        capture_output=True,
        text=True,
        timeout=15,
        env=env,
    )


class TestStagingRedirectOpen:
    """open() 写入时自动重定向到 staging 副本。"""

    def test_write_redirected_to_staged_copy(self, workspace: Path) -> None:
        """写模式下，原始路径被重定向到 staging 副本。"""
        original = workspace / "data.txt"
        original.write_text("original_content", encoding="utf-8")
        staged = workspace / "outputs" / "backups" / "data_staged.txt"
        staged.parent.mkdir(parents=True)
        staged.write_text("staged_content", encoding="utf-8")

        staging_map = {str(original): str(staged)}
        code = (
            f"with open(r'{original}', 'w') as f:\n"
            f"    f.write('modified')\n"
            f"print('done')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN", staging_map=staging_map)
        assert result.returncode == 0, f"stderr: {result.stderr}"
        # 原始文件未被修改
        assert original.read_text() == "original_content"
        # staging 副本被修改
        assert staged.read_text() == "modified"

    def test_read_redirected_to_staged_copy(self, workspace: Path) -> None:
        """读模式下，有 staging 映射时也重定向，确保读到最新 staged 版本。"""
        original = workspace / "data.txt"
        original.write_text("original", encoding="utf-8")
        staged = workspace / "outputs" / "backups" / "data_staged.txt"
        staged.parent.mkdir(parents=True)
        staged.write_text("staged_version", encoding="utf-8")

        staging_map = {str(original): str(staged)}
        code = (
            f"with open(r'{original}', 'r') as f:\n"
            f"    content = f.read()\n"
            f"print(content)\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN", staging_map=staging_map)
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert "staged_version" in result.stdout

    def test_no_staging_map_writes_normally(self, workspace: Path) -> None:
        """无 staging 映射时正常写入。"""
        target = workspace / "output.txt"
        code = (
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('hello')\n"
            f"print('done')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert target.read_text() == "hello"

    def test_empty_staging_map_no_effect(self, workspace: Path) -> None:
        """空 staging 映射不影响正常行为。"""
        target = workspace / "output.txt"
        code = (
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('hello')\n"
            f"print('done')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN", staging_map={})
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert target.read_text() == "hello"


class TestStagingRedirectOpenpyxl:
    """openpyxl wb.save() 写入时重定向到 staging 副本。"""

    def test_openpyxl_save_redirected(self, workspace: Path) -> None:
        """wb.save() 调用被重定向到 staging 副本。"""
        from openpyxl import Workbook, load_workbook

        original = workspace / "report.xlsx"
        wb = Workbook()
        wb.active["A1"] = "original_data"
        wb.save(str(original))
        wb.close()

        staged = workspace / "outputs" / "backups" / "report_staged.xlsx"
        staged.parent.mkdir(parents=True)
        shutil.copy2(str(original), str(staged))

        staging_map = {str(original): str(staged)}
        code = (
            "from openpyxl import load_workbook\n"
            f"wb = load_workbook(r'{original}')\n"
            "wb.active['A1'] = 'modified_data'\n"
            f"wb.save(r'{original}')\n"
            "print('saved')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN", staging_map=staging_map)
        assert result.returncode == 0, f"stderr: {result.stderr}"

        # 原始文件 A1 仍为 original_data
        wb_orig = load_workbook(original)
        assert wb_orig.active["A1"].value == "original_data"
        wb_orig.close()

        # staging 副本 A1 已更新
        wb_staged = load_workbook(staged)
        assert wb_staged.active["A1"].value == "modified_data"
        wb_staged.close()

    def test_openpyxl_save_no_staging_map(self, workspace: Path) -> None:
        """无 staging 映射时 openpyxl 正常写入。"""
        from openpyxl import Workbook, load_workbook

        output_dir = workspace / "outputs"
        output_dir.mkdir()
        target = output_dir / "test.xlsx"

        code = (
            "from openpyxl import Workbook\n"
            "wb = Workbook()\n"
            "wb.active['A1'] = 'hello'\n"
            f"wb.save(r'{target}')\n"
            "print('saved')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert target.exists()

        wb = load_workbook(target)
        assert wb.active["A1"].value == "hello"
        wb.close()


class TestStagingMapParsing:
    """staging 映射 JSON 解析的边界情况。"""

    def test_invalid_json_ignored(self, workspace: Path) -> None:
        """无效 JSON 不影响正常执行。"""
        target = workspace / "output.txt"
        code = (
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('ok')\n"
            f"print('done')\n"
        )
        script = workspace / "test_script.py"
        script.write_text(code, encoding="utf-8")
        wrapper = generate_wrapper_script("GREEN", str(workspace))
        wrapper_path = workspace / "_wrapper.py"
        wrapper_path.write_text(wrapper, encoding="utf-8")
        env = os.environ.copy()
        env["EXCELMANUS_STAGING_MAP"] = "not-valid-json"
        result = subprocess.run(
            [sys.executable, str(wrapper_path), str(script)],
            capture_output=True, text=True, timeout=15, env=env,
        )
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert target.read_text() == "ok"

    def test_non_dict_json_ignored(self, workspace: Path) -> None:
        """非 dict 类型 JSON 不影响正常执行。"""
        target = workspace / "output.txt"
        code = (
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('ok')\n"
            f"print('done')\n"
        )
        script = workspace / "test_script.py"
        script.write_text(code, encoding="utf-8")
        wrapper = generate_wrapper_script("GREEN", str(workspace))
        wrapper_path = workspace / "_wrapper.py"
        wrapper_path.write_text(wrapper, encoding="utf-8")
        env = os.environ.copy()
        env["EXCELMANUS_STAGING_MAP"] = "[1, 2, 3]"
        result = subprocess.run(
            [sys.executable, str(wrapper_path), str(script)],
            capture_output=True, text=True, timeout=15, env=env,
        )
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert target.read_text() == "ok"
