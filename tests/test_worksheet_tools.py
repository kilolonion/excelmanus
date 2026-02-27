"""工作表级设置工具测试：set_print_layout、set_page_header_footer。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excelmanus.tools.worksheet_tools import (
    get_tools,
    init_guard,
    set_page_header_footer,
    set_print_layout,
)


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为测试目录。"""
    init_guard(str(tmp_path))


def _make_sample(tmp_path: Path, name: str = "sample.xlsx") -> Path:
    """创建含数据的示例 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws["A1"] = "产品"
    ws["B1"] = "金额"
    for i in range(2, 12):
        ws[f"A{i}"] = f"产品{i - 1}"
        ws[f"B{i}"] = i * 100
    fp = tmp_path / name
    wb.save(fp)
    wb.close()
    return fp


# ── set_print_layout 测试 ──────────────────────────────────


class TestSetPrintLayout:
    """set_print_layout 工具测试套件。"""

    def test_set_print_area(self, tmp_path: Path) -> None:
        """设置打印区域。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", print_area="A1:B11"))
        assert result["status"] == "success"
        assert "print_area=A1:B11" in result["settings_applied"]

        wb = load_workbook(fp)
        assert wb["销售明细"].print_area is not None
        wb.close()

    def test_set_orientation_landscape(self, tmp_path: Path) -> None:
        """设置横向打印。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", orientation="landscape"))
        assert result["status"] == "success"

        wb = load_workbook(fp)
        assert wb["销售明细"].page_setup.orientation == "landscape"
        wb.close()

    def test_set_paper_size(self, tmp_path: Path) -> None:
        """设置 A3 纸张。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", paper_size="a3"))
        assert result["status"] == "success"

        wb = load_workbook(fp)
        assert wb["销售明细"].page_setup.paperSize == 8  # A3 = 8
        wb.close()

    def test_fit_to_width(self, tmp_path: Path) -> None:
        """缩放到页面宽度。"""
        fp = _make_sample(tmp_path)
        result = json.loads(
            set_print_layout(str(fp), "销售明细", fit_to_width=1, fit_to_height=0)
        )
        assert result["status"] == "success"

        wb = load_workbook(fp)
        ws = wb["销售明细"]
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        wb.close()

    def test_repeat_rows(self, tmp_path: Path) -> None:
        """每页重复表头行。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", repeat_rows_top="1:1"))
        assert result["status"] == "success"
        assert "repeat_rows=1:1" in result["settings_applied"]

    def test_combined_settings(self, tmp_path: Path) -> None:
        """一次调用设置多个打印参数。"""
        fp = _make_sample(tmp_path)
        result = json.loads(
            set_print_layout(
                str(fp),
                "销售明细",
                print_area="A1:B11",
                orientation="landscape",
                fit_to_width=1,
                repeat_rows_top="1:1",
                center_horizontally=True,
            )
        )
        assert result["status"] == "success"
        assert len(result["settings_applied"]) == 5

    def test_invalid_sheet_name(self, tmp_path: Path) -> None:
        """不存在的工作表报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "不存在的表"))
        assert "error" in result

    def test_invalid_orientation(self, tmp_path: Path) -> None:
        """无效的纸张方向报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", orientation="diagonal"))
        assert "error" in result

    def test_invalid_paper_size(self, tmp_path: Path) -> None:
        """无效的纸张大小报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_print_layout(str(fp), "销售明细", paper_size="a0"))
        assert "error" in result


# ── set_page_header_footer 测试 ──────────────────────────────


class TestSetPageHeaderFooter:
    """set_page_header_footer 工具测试套件。"""

    def test_set_header_center(self, tmp_path: Path) -> None:
        """设置中间页眉。"""
        fp = _make_sample(tmp_path)
        result = json.loads(
            set_page_header_footer(str(fp), "销售明细", header_center="XX公司2025年1月考勤表")
        )
        assert result["status"] == "success"

        wb = load_workbook(fp)
        assert "XX公司" in wb["销售明细"].oddHeader.center.text
        wb.close()

    def test_set_footer_with_page_number(self, tmp_path: Path) -> None:
        """设置页脚含页码占位符。"""
        fp = _make_sample(tmp_path)
        result = json.loads(
            set_page_header_footer(
                str(fp),
                "销售明细",
                footer_center="第 &[Page] 页 / 共 &[Pages] 页",
            )
        )
        assert result["status"] == "success"

        wb = load_workbook(fp)
        footer_text = wb["销售明细"].oddFooter.center.text
        # 占位符应被转换为 openpyxl 格式
        assert "&P" in footer_text
        assert "&N" in footer_text
        wb.close()

    def test_set_all_positions(self, tmp_path: Path) -> None:
        """设置全部六个位置。"""
        fp = _make_sample(tmp_path)
        result = json.loads(
            set_page_header_footer(
                str(fp),
                "销售明细",
                header_left="左页眉",
                header_center="中页眉",
                header_right="右页眉",
                footer_left="左页脚",
                footer_center="中页脚",
                footer_right="右页脚",
            )
        )
        assert result["status"] == "success"
        assert len(result["settings_applied"]) == 6

    def test_invalid_sheet_name(self, tmp_path: Path) -> None:
        """不存在的工作表报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(set_page_header_footer(str(fp), "不存在", header_center="test"))
        assert "error" in result


# ── get_tools 测试 ──────────────────────────────────────


def test_get_tools_returns_all() -> None:
    """验证 get_tools 返回正确数量和名称。"""
    tools = get_tools()
    assert len(tools) == 0  # Batch 3 精简：set_print_layout/set_page_header_footer 已删除
