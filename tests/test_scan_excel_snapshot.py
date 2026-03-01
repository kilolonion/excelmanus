"""scan_excel_snapshot 工具单元测试。"""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.tools.data_tools import init_guard, scan_excel_snapshot


@pytest.fixture(autouse=True)
def _set_guard(tmp_path: Path):
    """每个测试前将 FileAccessGuard 指向 tmp_path。"""
    init_guard(str(tmp_path))


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """创建一个包含多 Sheet 的测试 Excel 文件。"""
    wb = Workbook()
    # Sheet1: 混合数据
    ws1 = wb.active
    ws1.title = "销售数据"
    ws1.append(["ID", "姓名", "金额", "日期", "状态"])
    ws1.append([1, "张三", 100.5, "2024-01-01", "已完成"])
    ws1.append([2, "李四", 200.0, "2024-01-02", "已完成"])
    ws1.append([3, "王五", None, "2024-01-03", "已完成"])
    ws1.append([1, "张三", 100.5, "2024-01-01", "已完成"])  # 重复行
    ws1.append([5, "赵六", 99999.0, "2024-01-04", "待处理"])
    ws1.append([6, "钱七", -50.0, "2024-01-05", "已完成"])

    # Sheet2: 关联表（ID 列共享）
    ws2 = wb.create_sheet("员工信息")
    ws2.append(["ID", "部门", "电话"])
    ws2.append([1, "技术部", "13800138000"])
    ws2.append([2, "市场部", 13900139000])  # int 类型电话
    ws2.append([3, "技术部", "N/A"])

    fp = tmp_path / "test_data.xlsx"
    wb.save(fp)
    return fp


class TestScanExcelSnapshotBasic:
    """基础功能测试。"""

    def test_returns_valid_json(self, sample_xlsx: Path) -> None:
        result = scan_excel_snapshot(file_path=str(sample_xlsx))
        data = json.loads(result)
        assert "sheets" in data
        assert data["sheet_count"] == 2
        assert len(data["sheets"]) == 2

    def test_sheet_metadata(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(file_path=str(sample_xlsx)))
        sheet1 = data["sheets"][0]
        assert sheet1["name"] == "销售数据"
        assert sheet1["rows"] >= 6
        assert sheet1["cols"] == 5

    def test_column_stats_numeric(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(file_path=str(sample_xlsx)))
        sheet1 = data["sheets"][0]
        cols = {c["name"]: c for c in sheet1["columns"]}
        assert "金额" in cols
        amount = cols["金额"]
        assert amount["inferred_type"] == "numeric"
        assert amount["null_count"] >= 1
        assert "min" in amount
        assert "max" in amount
        assert "mean" in amount

    def test_column_stats_string(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(file_path=str(sample_xlsx)))
        sheet1 = data["sheets"][0]
        cols = {c["name"]: c for c in sheet1["columns"]}
        assert "姓名" in cols
        name_col = cols["姓名"]
        assert name_col["inferred_type"] == "string"
        assert name_col["unique_count"] >= 1
        assert "sample_values" in name_col

    def test_quality_signals_missing_data(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(file_path=str(sample_xlsx)))
        signals = data["quality_signals"]
        signal_types = [s["type"] for s in signals]
        assert "missing_data" in signal_types

    def test_duplicate_rows_detected(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(file_path=str(sample_xlsx)))
        sheet1 = data["sheets"][0]
        assert sheet1["duplicate_row_count"] >= 1

    def test_relationships_shared_column(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(
            file_path=str(sample_xlsx), include_relationships=True,
        ))
        rels = data.get("relationships", [])
        shared = [r for r in rels if r["type"] == "shared_column_name"]
        assert any("ID" in r["columns"] for r in shared)

    def test_no_relationships_when_disabled(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(
            file_path=str(sample_xlsx), include_relationships=False,
        ))
        assert data.get("relationships", []) == []


class TestScanExcelSnapshotSampling:
    """采样策略测试。"""

    def test_small_file_no_sampling(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(
            file_path=str(sample_xlsx), max_sample_rows=500,
        ))
        sheet1 = data["sheets"][0]
        assert sheet1.get("sampled") is not True

    def test_sampling_flag_when_forced(self, sample_xlsx: Path) -> None:
        data = json.loads(scan_excel_snapshot(
            file_path=str(sample_xlsx), max_sample_rows=2,
        ))
        sheet1 = data["sheets"][0]
        if sheet1["rows"] > 3:  # header + 2 data rows
            assert sheet1.get("sampled") is True
            assert "sample_size" in sheet1


class TestScanExcelSnapshotEdgeCases:
    """边界情况测试。"""

    def test_file_not_found(self, tmp_path: Path) -> None:
        result = scan_excel_snapshot(file_path=str(tmp_path / "nonexistent.xlsx"))
        data = json.loads(result)
        assert "error" in data

    def test_empty_sheet(self, tmp_path: Path) -> None:
        wb = Workbook()
        fp = tmp_path / "empty.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        assert data["sheet_count"] == 1
        assert data["sheets"][0]["rows"] <= 1

    def test_single_column(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["名称"])
        ws.append(["A"])
        ws.append(["B"])
        fp = tmp_path / "single_col.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        assert len(data["sheets"][0]["columns"]) == 1

    def test_all_null_column(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["空列", "有值列"])
        ws.append([None, "A"])
        ws.append([None, "B"])
        ws.append([None, "C"])
        fp = tmp_path / "null_col.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        cols = {c["name"]: c for c in data["sheets"][0]["columns"]}
        assert cols["空列"]["null_rate"] == 1.0
        signals = data["quality_signals"]
        empty_signals = [s for s in signals if s["type"] == "empty_column"]
        assert len(empty_signals) >= 1

    def test_constant_column(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["状态"])
        for _ in range(5):
            ws.append(["已完成"])
        fp = tmp_path / "constant.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        cols = {c["name"]: c for c in data["sheets"][0]["columns"]}
        assert cols["状态"]["unique_count"] == 1
        signals = data["quality_signals"]
        const_signals = [s for s in signals if s["type"] == "constant_column"]
        assert len(const_signals) >= 1


class TestScanExcelSnapshotMixedTypes:
    """类型混杂检测测试。

    注意：pandas read_excel 会将 str+int 混合列统一转为 float64，
    因此 openpyxl 层面的混合需要用不同方式检测。
    这里测试 pandas 层面可检测到的混合（object dtype 列中真正混合了 str 和 numeric）。
    """

    def test_mixed_type_detected_via_object_dtype(self, tmp_path: Path) -> None:
        """object dtype 列中混合了 str 和 int 时应检测到 mixed。"""
        # 使用 openpyxl 直接写入让 pandas 保持 object dtype
        wb = Workbook()
        ws = wb.active
        ws.append(["标签"])
        ws.append(["正常"])
        ws.append([100])       # int 值
        ws.append(["异常"])
        ws.append([True])      # bool 值
        fp = tmp_path / "mixed_object.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        cols = {c["name"]: c for c in data["sheets"][0]["columns"]}
        tag_col = cols["标签"]
        # pandas 读取后这个列应该是 object dtype，包含 str + int + bool 混合
        assert tag_col["inferred_type"] == "mixed"
        assert "mixed_type_counts" in tag_col

    def test_pure_string_not_mixed(self, tmp_path: Path) -> None:
        """纯字符串列不应被标记为 mixed。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["名称"])
        ws.append(["张三"])
        ws.append(["李四"])
        fp = tmp_path / "pure_str.xlsx"
        wb.save(fp)
        data = json.loads(scan_excel_snapshot(file_path=str(fp)))
        cols = {c["name"]: c for c in data["sheets"][0]["columns"]}
        assert cols["名称"]["inferred_type"] == "string"
        assert "mixed_type_counts" not in cols["名称"]
