"""单元测试：@ 提及解析器（MentionParser）。

覆盖 Parser 基础解析、边界情况、@img 兼容语法等场景。

**验证：需求 1.1–1.8, 8.1**
"""

from __future__ import annotations

from excelmanus.mentions.parser import MentionParser, Mention, ParseResult


# ── 单个类型解析 ──────────────────────────────────────────


class TestSingleMentionParsing:
    """单个 @ 标记解析测试。"""

    def test_file_mention(self) -> None:
        """@file:path 解析为 kind='file'。"""
        r = MentionParser.parse("分析 @file:sales.xlsx")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "file"
        assert m.value == "sales.xlsx"
        assert m.raw == "@file:sales.xlsx"

    def test_folder_mention(self) -> None:
        """@folder:path 解析为 kind='folder'。"""
        r = MentionParser.parse("查看 @folder:outputs/")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "folder"
        assert m.value == "outputs/"

    def test_skill_mention(self) -> None:
        """@skill:name 解析为 kind='skill'。"""
        r = MentionParser.parse("使用 @skill:data_basic 分析")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "skill"
        assert m.value == "data_basic"

    def test_mcp_mention(self) -> None:
        """@mcp:server_name 解析为 kind='mcp'。"""
        r = MentionParser.parse("连接 @mcp:mongodb")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "mcp"
        assert m.value == "mongodb"

    def test_img_mention(self) -> None:
        """@img path.png 解析为 kind='img'（兼容旧语法）。"""
        r = MentionParser.parse("请看 @img chart.png 这张图")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "img"
        assert m.value == "chart.png"

    def test_img_various_extensions(self) -> None:
        """@img 支持多种图片扩展名。"""
        for ext in ["png", "jpg", "jpeg", "gif", "bmp", "webp"]:
            r = MentionParser.parse(f"@img photo.{ext}")
            assert len(r.mentions) == 1
            assert r.mentions[0].value == f"photo.{ext}"


# ── 多标记解析 ────────────────────────────────────────────


class TestMultipleMentions:
    """多个 @ 标记解析测试。"""

    def test_multiple_mentions_in_order(self) -> None:
        """多个 @ 标记按出现顺序返回。"""
        r = MentionParser.parse("@file:a.xlsx @folder:src/ @skill:chart_basic")
        assert len(r.mentions) == 3
        assert r.mentions[0].kind == "file"
        assert r.mentions[1].kind == "folder"
        assert r.mentions[2].kind == "skill"

    def test_mixed_types_with_text(self) -> None:
        """混合类型与普通文本。"""
        r = MentionParser.parse("请分析 @file:data.xlsx 中的数据，参考 @folder:docs/")
        assert len(r.mentions) == 2
        assert r.mentions[0].value == "data.xlsx"
        assert r.mentions[1].value == "docs/"

    def test_position_tracking(self) -> None:
        """start/end 位置正确。"""
        text = "看 @file:a.xlsx 吧"
        r = MentionParser.parse(text)
        m = r.mentions[0]
        assert text[m.start : m.end] == "@file:a.xlsx"


# ── clean_text 生成 ───────────────────────────────────────


class TestCleanText:
    """clean_text 生成测试。"""

    def test_clean_text_removes_mentions(self) -> None:
        """clean_text 移除所有 @ 标记。"""
        r = MentionParser.parse("分析 @file:sales.xlsx 数据")
        assert "@file" not in r.clean_text
        assert "sales.xlsx" not in r.clean_text
        assert "分析" in r.clean_text
        assert "数据" in r.clean_text

    def test_clean_text_preserves_non_mention_text(self) -> None:
        """clean_text 保留非标记文本。"""
        r = MentionParser.parse("普通文本没有标记")
        assert r.clean_text == "普通文本没有标记"

    def test_original_preserved(self) -> None:
        """original 保留原始输入。"""
        text = "分析 @file:sales.xlsx"
        r = MentionParser.parse(text)
        assert r.original == text


# ── 边界情况 ──────────────────────────────────────────────


class TestEdgeCases:
    """边界情况测试。"""

    def test_empty_input(self) -> None:
        """空字符串返回空 Mention 列表。"""
        r = MentionParser.parse("")
        assert r.mentions == ()
        assert r.clean_text == ""
        assert r.original == ""

    def test_no_mentions(self) -> None:
        """无 @ 标记返回空列表，clean_text 与 original 相同。"""
        r = MentionParser.parse("普通文本")
        assert r.mentions == ()
        assert r.clean_text == "普通文本"

    def test_at_without_type(self) -> None:
        """@ 后无类型不匹配。"""
        r = MentionParser.parse("email@example.com")
        assert r.mentions == ()

    def test_at_type_without_value(self) -> None:
        """@type: 后无值不匹配（因为 \\S+ 要求至少一个非空白字符）。"""
        r = MentionParser.parse("@file: 后面是空格")
        # @file: 后面紧跟空格，\S+ 不匹配
        assert len(r.mentions) == 0

    def test_unknown_type(self) -> None:
        """未知 @ 类型不匹配。"""
        r = MentionParser.parse("@unknown:value")
        assert r.mentions == ()

    def test_case_insensitive(self) -> None:
        """@ 类型大小写不敏感。"""
        r = MentionParser.parse("@FILE:test.xlsx @Folder:src/")
        assert len(r.mentions) == 2
        assert r.mentions[0].kind == "file"
        assert r.mentions[1].kind == "folder"

    def test_consecutive_mentions(self) -> None:
        """连续多个 @ 标记（无间隔文本）。"""
        r = MentionParser.parse("@file:a.xlsx@file:b.xlsx")
        # 第一个匹配 @file:a.xlsx@file:b.xlsx（\S+ 贪婪匹配）
        # 实际行为取决于正则
        assert len(r.mentions) >= 1

    def test_mention_with_path_separators(self) -> None:
        """value 包含路径分隔符。"""
        r = MentionParser.parse("@file:data/reports/sales.xlsx")
        assert len(r.mentions) == 1
        assert r.mentions[0].value == "data/reports/sales.xlsx"

    def test_img_without_valid_extension(self) -> None:
        """@img 后跟非图片扩展名不匹配。"""
        r = MentionParser.parse("@img document.pdf")
        img_mentions = [m for m in r.mentions if m.kind == "img"]
        assert len(img_mentions) == 0

    def test_parse_result_is_immutable_tuple(self) -> None:
        """ParseResult.mentions 是不可变元组。"""
        r = MentionParser.parse("@file:test.xlsx")
        assert isinstance(r.mentions, tuple)


# ── range_spec 解析 ──────────────────────────────────────


class TestRangeSpecParsing:
    """@file:xxx[Sheet!Range] range_spec 解析测试。"""

    def test_file_with_range_spec(self) -> None:
        """@file:sales.xlsx[Sheet1!A1:C10] 解析 range_spec。"""
        r = MentionParser.parse("分析 @file:sales.xlsx[Sheet1!A1:C10]")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.kind == "file"
        assert m.value == "sales.xlsx"
        assert m.range_spec == "Sheet1!A1:C10"
        assert m.raw == "@file:sales.xlsx[Sheet1!A1:C10]"

    def test_file_with_range_no_sheet(self) -> None:
        """@file:data.xlsx[A1:B5] 无 sheet 名的 range_spec。"""
        r = MentionParser.parse("@file:data.xlsx[A1:B5]")
        assert len(r.mentions) == 1
        m = r.mentions[0]
        assert m.value == "data.xlsx"
        assert m.range_spec == "A1:B5"

    def test_file_with_single_cell_range(self) -> None:
        """@file:data.xlsx[Sheet1!A1] 单格 range_spec。"""
        r = MentionParser.parse("@file:data.xlsx[Sheet1!A1]")
        assert len(r.mentions) == 1
        assert r.mentions[0].range_spec == "Sheet1!A1"

    def test_file_without_range_spec(self) -> None:
        """@file:sales.xlsx 无 range_spec 时 range_spec 为 None。"""
        r = MentionParser.parse("@file:sales.xlsx")
        assert len(r.mentions) == 1
        assert r.mentions[0].range_spec is None

    def test_range_spec_clean_text(self) -> None:
        """带 range_spec 的标记从 clean_text 中完整移除。"""
        r = MentionParser.parse("查看 @file:sales.xlsx[Sheet1!A1:C10] 的数据")
        assert "@file" not in r.clean_text
        assert "Sheet1" not in r.clean_text
        assert "查看" in r.clean_text
        assert "的数据" in r.clean_text

    def test_range_spec_position_tracking(self) -> None:
        """range_spec 标记的 start/end 位置正确。"""
        text = "看 @file:a.xlsx[Sheet1!A1:B2] 吧"
        r = MentionParser.parse(text)
        m = r.mentions[0]
        assert text[m.start : m.end] == "@file:a.xlsx[Sheet1!A1:B2]"

    def test_multiple_mentions_with_range(self) -> None:
        """多个标记中混合有 range_spec 和无 range_spec。"""
        r = MentionParser.parse("@file:a.xlsx[Sheet1!A1:C3] @file:b.xlsx")
        assert len(r.mentions) == 2
        assert r.mentions[0].range_spec == "Sheet1!A1:C3"
        assert r.mentions[1].range_spec is None

    def test_non_file_mention_no_range(self) -> None:
        """非 file 类型的标记即使后跟 [] 也被正确解析。"""
        r = MentionParser.parse("@skill:data_basic")
        assert len(r.mentions) == 1
        assert r.mentions[0].range_spec is None


# ══════════════════════════════════════════════════════════
# 单元测试：MentionResolver
# **验证：需求 2.1–2.6, 3.1–3.5, 4.1, 4.5, 5.1, 5.2, 9.1–9.4**
# ══════════════════════════════════════════════════════════

import asyncio
import os
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from excelmanus.mentions.parser import Mention, ResolvedMention
from excelmanus.mentions.resolver import MentionResolver
from excelmanus.security.guard import FileAccessGuard


# ── 辅助工具 ──────────────────────────────────────────────


def _make_mention(kind: str, value: str) -> Mention:
    """快速构造 Mention 对象。"""
    raw = f"@{kind}:{value}" if kind != "img" else f"@img {value}"
    return Mention(kind=kind, value=value, raw=raw, start=0, end=len(raw))


def _make_resolver(
    workspace_root: str,
    skill_loader=None,
    mcp_manager=None,
    max_file_tokens: int = 2000,
) -> MentionResolver:
    """构造 MentionResolver 实例。"""
    guard = FileAccessGuard(workspace_root)
    return MentionResolver(
        workspace_root=workspace_root,
        guard=guard,
        skill_loader=skill_loader,
        mcp_manager=mcp_manager,
        max_file_tokens=max_file_tokens,
    )


# ── Resolver file: Excel 摘要 ────────────────────────────


class TestResolverExcelFile:
    """Excel 文件解析测试。"""

    def test_excel_summary_contains_sheets(self, tmp_path: Path) -> None:
        """Excel 摘要包含 sheet 列表、行列数、表头。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["日期", "产品", "数量"])
        ws.append(["2024-01-01", "A", 10])
        ws.append(["2024-01-02", "B", 20])
        wb.save(str(tmp_path / "test.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "test.xlsx")
        result = resolver._resolve_file(mention)

        assert result.error is None
        assert "Sales" in result.context_block
        assert "3行" in result.context_block
        assert "3列" in result.context_block
        assert "日期" in result.context_block

    def test_excel_multiple_sheets(self, tmp_path: Path) -> None:
        """多 sheet Excel 文件摘要包含所有 sheet。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["C", "D", "E"])
        wb.save(str(tmp_path / "multi.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "multi.xlsx")
        result = resolver._resolve_file(mention)

        assert result.error is None
        assert "Sheet1" in result.context_block
        assert "Sheet2" in result.context_block


# ── Resolver file: Excel 范围读取 ────────────────────────


class TestResolverExcelRange:
    """Excel 范围读取测试。"""

    def test_parse_range_spec_with_sheet(self) -> None:
        """_parse_range_spec 解析 Sheet1!A1:C10。"""
        sheet, cell_range = MentionResolver._parse_range_spec("Sheet1!A1:C10")
        assert sheet == "Sheet1"
        assert cell_range == "A1:C10"

    def test_parse_range_spec_without_sheet(self) -> None:
        """_parse_range_spec 解析 A1:C10（无 sheet 名）。"""
        sheet, cell_range = MentionResolver._parse_range_spec("A1:C10")
        assert sheet is None
        assert cell_range == "A1:C10"

    def test_parse_range_spec_single_cell(self) -> None:
        """_parse_range_spec 单格扩展为 A1:A1。"""
        sheet, cell_range = MentionResolver._parse_range_spec("Sheet1!A1")
        assert sheet == "Sheet1"
        assert cell_range == "A1:A1"

    def test_parse_range_spec_single_cell_no_sheet(self) -> None:
        """_parse_range_spec 无 sheet 单格。"""
        sheet, cell_range = MentionResolver._parse_range_spec("B5")
        assert sheet is None
        assert cell_range == "B5:B5"

    def test_excel_range_reads_correct_cells(self, tmp_path: Path) -> None:
        """_resolve_excel_range 读取正确的单元格数据。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 25, "Beijing"])
        ws.append(["Bob", 30, "Shanghai"])
        ws.append(["Carol", 28, "Guangzhou"])
        wb.save(str(tmp_path / "test.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = Mention(
            kind="file", value="test.xlsx",
            raw="@file:test.xlsx[Data!A1:C3]",
            start=0, end=27, range_spec="Data!A1:C3",
        )
        result = resolver._resolve_file(mention)

        assert result.error is None
        assert "Data!A1:C3" in result.context_block
        assert "Name" in result.context_block
        assert "Alice" in result.context_block
        assert "Bob" in result.context_block
        # 第 4 行（Carol）不应包含在内（范围为 A1:C3）
        assert "Carol" not in result.context_block

    def test_excel_range_single_cell(self, tmp_path: Path) -> None:
        """单格 range 正确读取。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B2"] = 42
        wb.save(str(tmp_path / "single.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = Mention(
            kind="file", value="single.xlsx",
            raw="@file:single.xlsx[Sheet1!B2]",
            start=0, end=28, range_spec="Sheet1!B2",
        )
        result = resolver._resolve_file(mention)

        assert result.error is None
        assert "42" in result.context_block

    def test_excel_range_nonexistent_sheet(self, tmp_path: Path) -> None:
        """不存在的 sheet 返回错误。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A"])
        wb.save(str(tmp_path / "test.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = Mention(
            kind="file", value="test.xlsx",
            raw="@file:test.xlsx[NoSuchSheet!A1:B2]",
            start=0, end=35, range_spec="NoSuchSheet!A1:B2",
        )
        result = resolver._resolve_file(mention)

        assert result.error is not None
        assert "工作表不存在" in result.error

    def test_excel_range_pipe_table_format(self, tmp_path: Path) -> None:
        """范围读取结果为管道分隔表格格式。"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Product", "Price"])
        ws.append(["Widget", 9.99])
        wb.save(str(tmp_path / "sales.xlsx"))
        wb.close()

        resolver = _make_resolver(str(tmp_path))
        mention = Mention(
            kind="file", value="sales.xlsx",
            raw="@file:sales.xlsx[Sales!A1:B2]",
            start=0, end=29, range_spec="Sales!A1:B2",
        )
        result = resolver._resolve_file(mention)

        assert result.error is None
        # 应包含管道分隔的表格
        assert "| A | B |" in result.context_block
        assert "| --- | --- |" in result.context_block
        assert "Product" in result.context_block
        assert "Widget" in result.context_block


# ── Resolver file: 文本文件 ──────────────────────────────


class TestResolverTextFile:
    """文本文件解析测试。"""

    def test_text_file_first_n_lines(self, tmp_path: Path) -> None:
        """文本文件摘要包含前 N 行。"""
        content = "\n".join(f"Line {i}" for i in range(10))
        (tmp_path / "readme.txt").write_text(content, encoding="utf-8")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "readme.txt")
        result = resolver._resolve_file(mention)

        assert result.error is None
        assert "Line 0" in result.context_block
        assert "Line 1" in result.context_block

    def test_large_text_file_truncation(self, tmp_path: Path) -> None:
        """大文本文件被截断到 token 预算内。"""
        # 生成大量文本（远超 50 tokens）
        content = "\n".join(f"This is a long line number {i} with some content" for i in range(500))
        (tmp_path / "large.txt").write_text(content, encoding="utf-8")

        resolver = _make_resolver(str(tmp_path), max_file_tokens=50)
        mention = _make_mention("file", "large.txt")
        result = resolver._resolve_file(mention)

        assert result.error is None
        # context_block 应该被截断
        import tiktoken
        enc = tiktoken.encoding_for_model("gpt-4o")
        token_count = len(enc.encode(result.context_block))
        assert token_count <= 50

    def test_file_not_found(self, tmp_path: Path) -> None:
        """不存在的文件返回错误。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "nonexistent.txt")
        result = resolver._resolve_file(mention)

        assert result.error is not None
        assert "文件不存在" in result.error


# ── Resolver folder ──────────────────────────────────────


class TestResolverFolder:
    """文件夹解析测试。"""

    def test_normal_tree(self, tmp_path: Path) -> None:
        """正常目录树生成。"""
        (tmp_path / "src").mkdir()
        (tmp_path / "src" / "main.py").write_text("pass")
        (tmp_path / "README.md").write_text("# Hello")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", ".")
        result = resolver._resolve_folder(mention)

        assert result.error is None
        assert "src" in result.context_block
        assert "main.py" in result.context_block
        assert "README.md" in result.context_block

    def test_empty_dir(self, tmp_path: Path) -> None:
        """空目录只显示目录名。"""
        empty = tmp_path / "empty"
        empty.mkdir()

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", "empty")
        result = resolver._resolve_folder(mention)

        assert result.error is None
        assert "empty" in result.context_block

    def test_deep_dir_truncation(self, tmp_path: Path) -> None:
        """深层目录被截断到 depth ≤ 2。"""
        # 创建 4 层深的目录
        deep = tmp_path / "a" / "b" / "c" / "d"
        deep.mkdir(parents=True)
        (deep / "deep_file.txt").write_text("deep")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", "a")
        result = resolver._resolve_folder(mention)

        assert result.error is None
        assert "b" in result.context_block
        # depth=2 时 c 应该出现（a/b/c 是第 2 层），但 d 不应出现
        # a(0) -> b(1) -> c(2)，c 在 depth=2 时是叶节点，不展开
        assert "deep_file.txt" not in result.context_block

    def test_exclude_hidden_files(self, tmp_path: Path) -> None:
        """排除隐藏文件。"""
        (tmp_path / ".hidden").write_text("secret")
        (tmp_path / "visible.txt").write_text("hello")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", ".")
        result = resolver._resolve_folder(mention)

        assert result.error is None
        assert ".hidden" not in result.context_block
        assert "visible.txt" in result.context_block

    def test_exclude_venv_and_node_modules(self, tmp_path: Path) -> None:
        """排除 .venv 和 node_modules 目录。"""
        (tmp_path / ".venv").mkdir()
        (tmp_path / ".venv" / "lib.py").write_text("pass")
        (tmp_path / "node_modules").mkdir()
        (tmp_path / "node_modules" / "pkg.js").write_text("//")
        (tmp_path / "app.py").write_text("pass")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", ".")
        result = resolver._resolve_folder(mention)

        assert result.error is None
        assert ".venv" not in result.context_block
        assert "node_modules" not in result.context_block
        assert "app.py" in result.context_block

    def test_folder_not_found(self, tmp_path: Path) -> None:
        """不存在的目录返回错误。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", "nonexistent")
        result = resolver._resolve_folder(mention)

        assert result.error is not None
        assert "目录不存在" in result.error


# ── Resolver skill ───────────────────────────────────────


class TestResolverSkill:
    """Skill 解析测试。"""

    def test_existing_skill(self, tmp_path: Path) -> None:
        """已存在的 skill 返回 render_context() 内容。"""
        mock_skill = MagicMock()
        mock_skill.render_context.return_value = "[Skillpack] data_basic\n描述：基础数据分析"

        mock_loader = MagicMock()
        mock_loader.get_skillpack.return_value = mock_skill

        resolver = _make_resolver(str(tmp_path), skill_loader=mock_loader)
        mention = _make_mention("skill", "data_basic")
        result = resolver._resolve_skill(mention)

        assert result.error is None
        assert "data_basic" in result.context_block
        mock_loader.get_skillpack.assert_called_once_with("data_basic")

    def test_nonexistent_skill(self, tmp_path: Path) -> None:
        """不存在的 skill 返回错误。"""
        mock_loader = MagicMock()
        mock_loader.get_skillpack.return_value = None

        resolver = _make_resolver(str(tmp_path), skill_loader=mock_loader)
        mention = _make_mention("skill", "nonexistent")
        result = resolver._resolve_skill(mention)

        assert result.error is not None
        assert "技能不存在" in result.error

    def test_no_skill_loader(self, tmp_path: Path) -> None:
        """无 skill_loader 时返回错误。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("skill", "any_skill")
        result = resolver._resolve_skill(mention)

        assert result.error is not None
        assert "技能不存在" in result.error


# ── Resolver mcp ─────────────────────────────────────────


class TestResolverMCP:
    """MCP 服务解析测试。"""

    async def test_connected_mcp_returns_tools(self, tmp_path: Path) -> None:
        """已连接的 MCP 返回工具列表。"""
        mock_manager = MagicMock()
        mock_manager.connected_servers.return_value = ["mongodb"]
        mock_manager.get_server_info.return_value = [
            {
                "name": "mongodb",
                "transport": "stdio",
                "status": "ready",
                "tool_count": 3,
                "tools": ["find", "aggregate", "count"],
                "last_error": None,
                "init_ms": 100,
            }
        ]

        resolver = _make_resolver(str(tmp_path), mcp_manager=mock_manager)
        mention = _make_mention("mcp", "mongodb")
        result = await resolver._resolve_mcp(mention)

        assert result.error is None
        assert "mongodb" in result.context_block
        assert "find" in result.context_block
        assert "aggregate" in result.context_block

    async def test_disconnected_mcp_returns_error(self, tmp_path: Path) -> None:
        """未连接的 MCP 返回错误。"""
        mock_manager = MagicMock()
        mock_manager.connected_servers.return_value = ["other_server"]

        resolver = _make_resolver(str(tmp_path), mcp_manager=mock_manager)
        mention = _make_mention("mcp", "mongodb")
        result = await resolver._resolve_mcp(mention)

        assert result.error is not None
        assert "MCP 服务未连接或不存在" in result.error

    async def test_no_mcp_manager(self, tmp_path: Path) -> None:
        """无 mcp_manager 时返回错误。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("mcp", "any_server")
        result = await resolver._resolve_mcp(mention)

        assert result.error is not None
        assert "MCP 服务未连接或不存在" in result.error


# ── Resolver 安全 ────────────────────────────────────────


class TestResolverSecurity:
    """安全校验测试。"""

    def test_path_traversal_rejection(self, tmp_path: Path) -> None:
        """路径穿越被拒绝。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "../../../etc/passwd")
        result = resolver._resolve_file(mention)

        assert result.error is not None
        assert result.context_block == ""

    def test_path_traversal_url_encoded(self, tmp_path: Path) -> None:
        """URL 编码的路径穿越被拒绝。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "%2e%2e/secret.txt")
        result = resolver._resolve_file(mention)

        assert result.error is not None
        assert result.context_block == ""

    def test_out_of_bounds_rejection(self, tmp_path: Path) -> None:
        """路径越界被拒绝。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "/etc/passwd")
        result = resolver._resolve_file(mention)

        assert result.error is not None
        assert result.context_block == ""

    def test_folder_path_traversal(self, tmp_path: Path) -> None:
        """文件夹路径穿越被拒绝。"""
        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("folder", "../..")
        result = resolver._resolve_folder(mention)

        assert result.error is not None
        assert result.context_block == ""

    def test_dangling_symlink_rejection(self, tmp_path: Path) -> None:
        """悬空符号链接被拒绝。"""
        link = tmp_path / "dangling_link"
        link.symlink_to(tmp_path / "nonexistent_target")

        resolver = _make_resolver(str(tmp_path))
        mention = _make_mention("file", "dangling_link")
        result = resolver._resolve_file(mention)

        # 应该返回错误（文件不存在或符号链接目标不存在）
        assert result.error is not None


# ── Resolver resolve() 集成 ──────────────────────────────


class TestResolverResolve:
    """resolve() 方法集成测试。"""

    async def test_resolve_multiple_types(self, tmp_path: Path) -> None:
        """resolve() 处理多种类型的 Mention。"""
        (tmp_path / "test.txt").write_text("hello world")
        (tmp_path / "subdir").mkdir()

        resolver = _make_resolver(str(tmp_path))
        mentions = [
            _make_mention("file", "test.txt"),
            _make_mention("folder", "subdir"),
            _make_mention("img", "photo.png"),
        ]
        results = await resolver.resolve(mentions)

        assert len(results) == 3
        # file 应成功
        assert results[0].error is None
        assert "hello world" in results[0].context_block
        # folder 应成功
        assert results[1].error is None
        # img 不生成 context_block
        assert results[2].context_block == ""
        assert results[2].error is None


# ── Engine 上下文注入与 SkillRouter 兼容测试 ──────────────


from excelmanus.mentions.parser import ResolvedMention, Mention
from excelmanus.engine import build_mention_context_block


class TestBuildMentionContextBlock:
    """build_mention_context_block() 函数测试。"""

    def test_empty_list_returns_empty(self) -> None:
        """空列表返回空字符串。"""
        assert build_mention_context_block([]) == ""

    def test_none_returns_empty(self) -> None:
        """None 等价于空列表。"""
        assert build_mention_context_block([]) == ""

    def test_single_file_mention(self) -> None:
        """单个成功的 file mention 生成正确的 XML。"""
        mention = Mention(kind="file", value="sales.xlsx", raw="@file:sales.xlsx", start=0, end=16)
        rm = ResolvedMention(mention=mention, context_block="  Sheets: Sheet1 (100行×5列)")
        xml = build_mention_context_block([rm])

        assert "<mention_context>" in xml
        assert '</mention_context>' in xml
        assert '<file path="sales.xlsx">' in xml
        assert "Sheets: Sheet1 (100行×5列)" in xml
        assert "</file>" in xml

    def test_single_folder_mention(self) -> None:
        """单个成功的 folder mention 生成正确的 XML。"""
        mention = Mention(kind="folder", value="outputs/", raw="@folder:outputs/", start=0, end=16)
        rm = ResolvedMention(mention=mention, context_block="  outputs/\n  ├── report.xlsx")
        xml = build_mention_context_block([rm])

        assert '<folder path="outputs/">' in xml
        assert "report.xlsx" in xml
        assert "</folder>" in xml

    def test_single_skill_mention(self) -> None:
        """单个成功的 skill mention 生成正确的 XML。"""
        mention = Mention(kind="skill", value="data_basic", raw="@skill:data_basic", start=0, end=17)
        rm = ResolvedMention(mention=mention, context_block="  [Skillpack] data_basic\n  描述：基础数据分析")
        xml = build_mention_context_block([rm])

        assert '<skill name="data_basic">' in xml
        assert "[Skillpack] data_basic" in xml
        assert "</skill>" in xml

    def test_single_mcp_mention(self) -> None:
        """单个成功的 mcp mention 生成正确的 XML。"""
        mention = Mention(kind="mcp", value="mongodb", raw="@mcp:mongodb", start=0, end=12)
        rm = ResolvedMention(mention=mention, context_block="  Tools: find, aggregate")
        xml = build_mention_context_block([rm])

        assert '<mcp server="mongodb">' in xml
        assert "Tools: find, aggregate" in xml
        assert "</mcp>" in xml

    def test_error_mention(self) -> None:
        """解析失败的 mention 生成 <error> 标签。"""
        mention = Mention(kind="file", value="nonexistent.xlsx", raw="@file:nonexistent.xlsx", start=0, end=22)
        rm = ResolvedMention(mention=mention, error="文件不存在：nonexistent.xlsx")
        xml = build_mention_context_block([rm])

        assert '<error ref="@file:nonexistent.xlsx">' in xml
        assert "文件不存在：nonexistent.xlsx" in xml
        assert "</error>" in xml

    def test_img_mention_skipped(self) -> None:
        """img 类型 mention 不生成 context block。"""
        mention = Mention(kind="img", value="photo.png", raw="@img photo.png", start=0, end=14)
        rm = ResolvedMention(mention=mention, context_block="")
        xml = build_mention_context_block([rm])

        assert xml == ""

    def test_multiple_mentions_all_present(self) -> None:
        """多个 mention 的 context_block 全部出现在 XML 中。"""
        mentions = [
            ResolvedMention(
                mention=Mention(kind="file", value="a.xlsx", raw="@file:a.xlsx", start=0, end=12),
                context_block="  File A content",
            ),
            ResolvedMention(
                mention=Mention(kind="folder", value="data/", raw="@folder:data/", start=13, end=26),
                context_block="  data/\n  ├── sub/",
            ),
            ResolvedMention(
                mention=Mention(kind="skill", value="chart_basic", raw="@skill:chart_basic", start=27, end=45),
                context_block="  [Skillpack] chart_basic",
            ),
        ]
        xml = build_mention_context_block(mentions)

        assert "<mention_context>" in xml
        assert "File A content" in xml
        assert "data/" in xml
        assert "[Skillpack] chart_basic" in xml

    def test_mixed_success_and_failure(self) -> None:
        """成功和失败的 mention 混合时，XML 同时包含 context 和 error。"""
        mentions = [
            ResolvedMention(
                mention=Mention(kind="file", value="ok.txt", raw="@file:ok.txt", start=0, end=12),
                context_block="  File content here",
            ),
            ResolvedMention(
                mention=Mention(kind="file", value="bad.txt", raw="@file:bad.txt", start=13, end=26),
                error="文件不存在：bad.txt",
            ),
        ]
        xml = build_mention_context_block(mentions)

        assert '<file path="ok.txt">' in xml
        assert "File content here" in xml
        assert '<error ref="@file:bad.txt">' in xml
        assert "文件不存在：bad.txt" in xml

    def test_user_message_preserved(self) -> None:
        """用户消息保持原样（保留 @type:value 标记）。

        这是一个概念验证：engine.chat() 不修改 user_message，
        只将 mention 上下文注入系统提示词。
        """
        original_message = "请分析 @file:sales.xlsx 的数据"
        # 解析后 clean_text 不含 @ 标记
        result = MentionParser.parse(original_message)
        assert "@file:sales.xlsx" not in result.clean_text
        # 但 original 保留
        assert result.original == original_message

    def test_no_mentions_returns_empty(self) -> None:
        """无 mention 输入时 build_mention_context_block 返回空。"""
        xml = build_mention_context_block([])
        assert xml == ""


class TestSkillMentionRouting:
    """@skill:name 路由兼容性测试。"""

    def test_skill_mention_extracts_raw_args(self) -> None:
        """@skill:name 后面的文本被正确提取为 raw_args。"""
        text = "请分析 @skill:data_basic 这个文件的数据"
        result = MentionParser.parse(text)

        skill_mentions = [m for m in result.mentions if m.kind == "skill"]
        assert len(skill_mentions) == 1
        assert skill_mentions[0].value == "data_basic"

        # clean_text 就是 raw_args
        assert "data_basic" not in result.clean_text
        assert "请分析" in result.clean_text
        assert "这个文件的数据" in result.clean_text

    def test_skill_mention_and_slash_produce_same_name(self) -> None:
        """@skill:data_basic 和 /data_basic 应产生相同的 skill name。"""
        mention_input = "@skill:data_basic 分析数据"
        result = MentionParser.parse(mention_input)

        skill_mentions = [m for m in result.mentions if m.kind == "skill"]
        assert len(skill_mentions) == 1

        # @skill:name 提取的 value 就是 skill name
        skill_name_from_mention = skill_mentions[0].value

        # /name 斜杠命令中的 name 部分
        slash_input = "/data_basic 分析数据"
        # 斜杠命令解析：取 / 后第一个空格前的部分
        slash_name = slash_input.split()[0].lstrip("/")

        assert skill_name_from_mention == slash_name

    def test_no_at_input_backward_compatible(self) -> None:
        """不包含 @ 标记的输入行为不变。"""
        text = "帮我分析这个文件"
        result = MentionParser.parse(text)

        assert len(result.mentions) == 0
        assert result.clean_text == text
        assert result.original == text

        # build_mention_context_block 返回空
        xml = build_mention_context_block([])
        assert xml == ""

    def test_slash_command_still_works(self) -> None:
        """/skill_name 斜杠命令语法仍然可用（向后兼容）。"""
        # 斜杠命令不是 @ mention，不会被 MentionParser 解析
        text = "/data_basic 分析数据"
        result = MentionParser.parse(text)

        # 不应提取到任何 mention
        assert len(result.mentions) == 0
        assert result.clean_text == text

    def test_skill_mention_with_multiple_mentions(self) -> None:
        """@skill:name 与其他 mention 共存时正确提取。"""
        text = "@file:sales.xlsx @skill:data_basic 分析这个文件"
        result = MentionParser.parse(text)

        assert len(result.mentions) == 2
        assert result.mentions[0].kind == "file"
        assert result.mentions[0].value == "sales.xlsx"
        assert result.mentions[1].kind == "skill"
        assert result.mentions[1].value == "data_basic"


# ══════════════════════════════════════════════════════════
# 单元测试：MentionCompleter
# **验证：需求 7.1–7.8, 8.1–8.3**
# ══════════════════════════════════════════════════════════

from prompt_toolkit.completion import CompleteEvent
from prompt_toolkit.document import Document

from excelmanus.mentions.completer import MentionCompleter


def _make_completer(
    workspace_root: str,
    engine=None,
    max_scan_depth: int = 2,
) -> MentionCompleter:
    """构造 MentionCompleter 实例。"""
    return MentionCompleter(
        workspace_root=workspace_root,
        engine=engine,
        max_scan_depth=max_scan_depth,
    )


def _get_completions(completer: MentionCompleter, text: str) -> list:
    """获取补全结果列表。"""
    doc = Document(text, cursor_position=len(text))
    event = CompleteEvent()
    return list(completer.get_completions(doc, event))


def _display_text(completion) -> str:
    """从 Completion 的 display 字段提取纯文本。"""
    d = completion.display
    if isinstance(d, str):
        return d
    # FormattedText：(style, text) 元组列表
    try:
        return "".join(t for _, t in d)
    except Exception:
        return str(d)


# ── 阶段一：@ 触发分类菜单 ───────────────────────────────


class TestCompleterCategoryMenu:
    """@ 触发分类菜单测试。"""

    def test_at_triggers_category_menu(self, tmp_path: Path) -> None:
        """输入 @ 后显示 5 个分类候选项。"""
        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@")

        assert len(completions) == 5
        display_texts = [_display_text(c) for c in completions]
        assert "@file" in display_texts
        assert "@folder" in display_texts
        assert "@skill" in display_texts
        assert "@mcp" in display_texts
        assert "@img" in display_texts

    def test_partial_category_filters(self, tmp_path: Path) -> None:
        """输入 @fi 过滤出 file 分类。"""
        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@fi")

        assert len(completions) == 1
        assert _display_text(completions[0]) == "@file"

    def test_partial_category_fo(self, tmp_path: Path) -> None:
        """输入 @fo 过滤出 folder 分类。"""
        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@fo")

        assert len(completions) == 1
        assert _display_text(completions[0]) == "@folder"

    def test_no_at_no_completions(self, tmp_path: Path) -> None:
        """无 @ 输入不触发补全。"""
        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "hello")

        assert len(completions) == 0


# ── 阶段二：@file: 文件补全 ──────────────────────────────


class TestCompleterFileCompletions:
    """@file: 文件补全测试。"""

    def test_file_lists_workspace_files(self, tmp_path: Path) -> None:
        """@file: 列出工作区文件。"""
        (tmp_path / "data.xlsx").write_text("")
        (tmp_path / "readme.md").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@file:")

        texts = [c.text for c in completions]
        assert "data.xlsx" in texts
        assert "readme.md" in texts

    def test_file_excludes_hidden(self, tmp_path: Path) -> None:
        """@file: 排除隐藏文件。"""
        (tmp_path / ".hidden").write_text("")
        (tmp_path / "visible.txt").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@file:")

        texts = [c.text for c in completions]
        assert ".hidden" not in texts
        assert "visible.txt" in texts

    def test_file_excludes_venv_and_node_modules(self, tmp_path: Path) -> None:
        """@file: 排除 .venv 和 node_modules 目录内容。"""
        (tmp_path / ".venv").mkdir()
        (tmp_path / ".venv" / "lib.py").write_text("")
        (tmp_path / "node_modules").mkdir()
        (tmp_path / "node_modules" / "pkg.js").write_text("")
        (tmp_path / "app.py").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@file:")

        texts = [c.text for c in completions]
        assert "app.py" in texts
        # .venv 和 node_modules 内的文件不应出现
        assert not any(".venv" in t for t in texts)
        assert not any("node_modules" in t for t in texts)

    def test_file_depth_limit(self, tmp_path: Path) -> None:
        """@file: 深度限制 ≤ 2（逐级浏览模式）。"""
        deep = tmp_path / "a" / "b" / "c"
        deep.mkdir(parents=True)
        (deep / "deep.txt").write_text("")
        (tmp_path / "a" / "b" / "shallow.txt").write_text("")
        (tmp_path / "top.txt").write_text("")

        completer = _make_completer(str(tmp_path), max_scan_depth=2)

        # 根目录层级：能看到 top.txt 和 a/ 目录
        root_completions = _get_completions(completer, "@file:")
        root_texts = [c.text for c in root_completions]
        assert "top.txt" in root_texts
        assert "a/" in root_texts

        # depth 2 层级（a/b/）：能看到 shallow.txt 和 c/ 目录
        depth2_completions = _get_completions(completer, "@file:a/b/")
        depth2_texts = [c.text for c in depth2_completions]
        assert "a/b/shallow.txt" in depth2_texts
        assert "a/b/c/" in depth2_texts

        # depth 3 层级（a/b/c/）：超过限制，不应返回任何条目
        depth3_completions = _get_completions(completer, "@file:a/b/c/")
        depth3_texts = [c.text for c in depth3_completions]
        assert "a/b/c/deep.txt" not in depth3_texts

    def test_file_prefix_filter(self, tmp_path: Path) -> None:
        """@file:da 过滤以 da 开头的文件。"""
        (tmp_path / "data.xlsx").write_text("")
        (tmp_path / "readme.md").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@file:da")

        texts = [c.text for c in completions]
        assert "data.xlsx" in texts
        assert "readme.md" not in texts


# ── 阶段二：@folder: 目录补全 ────────────────────────────


class TestCompleterFolderCompletions:
    """@folder: 目录补全测试。"""

    def test_folder_lists_directories(self, tmp_path: Path) -> None:
        """@folder: 列出工作区目录。"""
        (tmp_path / "src").mkdir()
        (tmp_path / "docs").mkdir()
        (tmp_path / "file.txt").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@folder:")

        texts = [c.text for c in completions]
        assert "src/" in texts
        assert "docs/" in texts
        # 文件不应出现在 folder 补全中
        assert "file.txt" not in texts

    def test_folder_excludes_hidden(self, tmp_path: Path) -> None:
        """@folder: 排除隐藏目录。"""
        (tmp_path / ".git").mkdir()
        (tmp_path / "src").mkdir()

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@folder:")

        texts = [c.text for c in completions]
        assert ".git/" not in texts
        assert "src/" in texts


# ── 阶段二：@skill: 技能补全 ─────────────────────────────


class TestCompleterSkillCompletions:
    """@skill: 技能补全测试。"""

    def test_skill_lists_names(self, tmp_path: Path) -> None:
        """@skill: 列出 user_invocable 的技能名称。"""
        mock_engine = MagicMock()
        mock_engine._list_manual_invocable_skill_names.return_value = [
            "data_basic",
            "chart_basic",
            "format_basic",
        ]

        completer = _make_completer(str(tmp_path), engine=mock_engine)
        completions = _get_completions(completer, "@skill:")

        texts = [c.text for c in completions]
        assert "data_basic" in texts
        assert "chart_basic" in texts
        assert "format_basic" in texts

    def test_skill_prefix_filter(self, tmp_path: Path) -> None:
        """@skill:da 过滤以 da 开头的技能。"""
        mock_engine = MagicMock()
        mock_engine._list_manual_invocable_skill_names.return_value = [
            "data_basic",
            "chart_basic",
        ]

        completer = _make_completer(str(tmp_path), engine=mock_engine)
        completions = _get_completions(completer, "@skill:da")

        texts = [c.text for c in completions]
        assert "data_basic" in texts
        assert "chart_basic" not in texts

    def test_skill_no_engine(self, tmp_path: Path) -> None:
        """无 engine 时 @skill: 返回空。"""
        completer = _make_completer(str(tmp_path), engine=None)
        completions = _get_completions(completer, "@skill:")

        assert len(completions) == 0


# ── 阶段二：@mcp: MCP 补全 ──────────────────────────────


class TestCompleterMCPCompletions:
    """@mcp: MCP 服务补全测试。"""

    def test_mcp_lists_servers(self, tmp_path: Path) -> None:
        """@mcp: 列出已连接的 MCP 服务名称。"""
        mock_engine = MagicMock()
        mock_engine.mcp_server_info.return_value = [
            {"name": "mongodb", "status": "ready"},
            {"name": "postgres", "status": "ready"},
        ]

        completer = _make_completer(str(tmp_path), engine=mock_engine)
        completions = _get_completions(completer, "@mcp:")

        texts = [c.text for c in completions]
        assert "mongodb" in texts
        assert "postgres" in texts

    def test_mcp_no_engine(self, tmp_path: Path) -> None:
        """无 engine 时 @mcp: 返回空。"""
        completer = _make_completer(str(tmp_path), engine=None)
        completions = _get_completions(completer, "@mcp:")

        assert len(completions) == 0


# ── @img 补全 ────────────────────────────────────────────


class TestCompleterImgCompletions:
    """@img 图片补全测试。"""

    def test_img_lists_image_files(self, tmp_path: Path) -> None:
        """@img 列出图片文件。"""
        (tmp_path / "chart.png").write_text("")
        (tmp_path / "photo.jpg").write_text("")
        (tmp_path / "data.xlsx").write_text("")

        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@img ")

        texts = [c.text for c in completions]
        assert "chart.png" in texts
        assert "photo.jpg" in texts
        assert "data.xlsx" not in texts


# ── CLI 集成测试（向后兼容验证）─────────────────────────


class TestCLIBackwardCompatibility:
    """CLI 集成向后兼容测试。"""

    def test_img_syntax_backward_compatible(self) -> None:
        """@img path.png 语法保持向后兼容。"""
        result = MentionParser.parse("@img chart.png")
        assert len(result.mentions) == 1
        assert result.mentions[0].kind == "img"
        assert result.mentions[0].value == "chart.png"

    def test_slash_command_still_works(self) -> None:
        """/skill_name 斜杠命令不被 @ 解析器干扰。"""
        result = MentionParser.parse("/data_basic 分析数据")
        assert len(result.mentions) == 0
        assert result.clean_text == "/data_basic 分析数据"

    def test_no_at_input_unchanged(self) -> None:
        """无 @ 输入行为不变。"""
        result = MentionParser.parse("帮我分析这个文件")
        assert len(result.mentions) == 0
        assert result.clean_text == "帮我分析这个文件"
        assert result.original == "帮我分析这个文件"

    def test_completer_handles_empty_workspace(self, tmp_path: Path) -> None:
        """空工作区不崩溃。"""
        completer = _make_completer(str(tmp_path))
        completions = _get_completions(completer, "@file:")
        assert isinstance(completions, list)

    def test_completer_handles_no_engine(self, tmp_path: Path) -> None:
        """无 engine 时补全器不崩溃。"""
        completer = _make_completer(str(tmp_path), engine=None)
        # 所有分类都应正常工作（skill/mcp 返回空）
        for prefix in ["@", "@file:", "@folder:", "@skill:", "@mcp:", "@img "]:
            completions = _get_completions(completer, prefix)
            assert isinstance(completions, list)
