"""bench_validator 断言校验引擎的单元测试。"""

from __future__ import annotations

import pytest

from excelmanus.bench_validator import (
    AssertionResult,
    SuiteValidationSummary,
    ValidationSummary,
    aggregate_suite_validation,
    merge_assertions,
    validate_case,
)


# ── 测试数据工厂 ─────────────────────────────────────────


def _make_result_dict(
    *,
    status: str = "ok",
    iterations: int = 2,
    duration_seconds: float = 3.5,
    llm_call_count: int = 2,
    tool_call_count: int = 3,
    tool_successes: int = 3,
    tool_failures: int = 0,
    total_tokens: int = 12000,
    skills_used: list[str] | None = None,
    route_mode: str = "fallback",
    tool_calls: list[dict] | None = None,
    llm_calls: list[dict] | None = None,
    reply: str = "这是测试回复，包含列名信息。",
) -> dict:
    """构造一个最小化的 BenchResult.to_dict() 输出。"""
    if skills_used is None:
        skills_used = ["data_basic"]
    if tool_calls is None:
        tool_calls = [
            {"tool_name": "read_excel", "success": True},
            {"tool_name": "list_sheets", "success": True},
            {"tool_name": "format_cells", "success": True},
        ][:tool_call_count]
    if llm_calls is None:
        llm_calls = [
            {
                "response": {
                    "content": "",
                    "tool_calls": [{"function": {"name": "read_excel"}}],
                },
            },
            {
                "response": {
                    "content": reply,
                    "tool_calls": None,
                },
            },
        ][:llm_call_count]
    return {
        "execution": {
            "status": status,
            "iterations": iterations,
            "duration_seconds": duration_seconds,
            "skills_used": skills_used,
            "route_mode": route_mode,
        },
        "stats": {
            "llm_call_count": llm_call_count,
            "tool_call_count": tool_call_count,
            "tool_successes": tool_successes,
            "tool_failures": tool_failures,
            "total_tokens": total_tokens,
        },
        "artifacts": {
            "tool_calls": tool_calls,
            "llm_calls": llm_calls,
        },
        "result": {
            "reply": reply,
        },
    }


# ── merge_assertions ─────────────────────────────────────


class TestMergeAssertions:
    def test_both_none(self):
        assert merge_assertions(None, None) == {}

    def test_suite_only(self):
        result = merge_assertions({"max_iterations": 5}, None)
        assert result == {"max_iterations": 5}

    def test_case_only(self):
        result = merge_assertions(None, {"max_iterations": 3})
        assert result == {"max_iterations": 3}

    def test_case_overrides_suite(self):
        result = merge_assertions(
            {"max_iterations": 5, "status": "ok"},
            {"max_iterations": 3},
        )
        assert result == {"max_iterations": 3, "status": "ok"}

    def test_disjoint_fields_merged(self):
        result = merge_assertions(
            {"max_iterations": 5},
            {"expected_skill": "data_basic"},
        )
        assert result == {"max_iterations": 5, "expected_skill": "data_basic"}


# ── validate_case: 个别规则 ──────────────────────────────


class TestValidateStatus:
    def test_status_ok_passes(self):
        r = _make_result_dict(status="ok")
        v = validate_case(r, {"status": "ok"})
        assert v.total == 1
        assert v.passed == 1
        assert v.failed == 0

    def test_status_mismatch_fails(self):
        r = _make_result_dict(status="error")
        v = validate_case(r, {"status": "ok"})
        assert v.failed == 1
        assert "error" in v.results[0].message


class TestValidateMaxIterations:
    def test_within_limit(self):
        r = _make_result_dict(iterations=2)
        v = validate_case(r, {"max_iterations": 5})
        assert v.passed == 1

    def test_at_limit(self):
        r = _make_result_dict(iterations=5)
        v = validate_case(r, {"max_iterations": 5})
        assert v.passed == 1

    def test_exceeds_limit(self):
        r = _make_result_dict(iterations=6)
        v = validate_case(r, {"max_iterations": 5})
        assert v.failed == 1
        assert "超过上限" in v.results[0].message


class TestValidateMaxTokens:
    def test_within_limit(self):
        r = _make_result_dict(total_tokens=10000)
        v = validate_case(r, {"max_tokens": 50000})
        assert v.passed == 1

    def test_exceeds_limit(self):
        r = _make_result_dict(total_tokens=60000)
        v = validate_case(r, {"max_tokens": 50000})
        assert v.failed == 1


class TestValidateExpectedSkill:
    def test_skill_in_skills_used(self):
        r = _make_result_dict(skills_used=["data_basic", "data_basic"])
        v = validate_case(r, {"expected_skill": "data_basic"})
        assert v.passed == 1

    def test_skill_matches_route_mode(self):
        r = _make_result_dict(skills_used=[], route_mode="data_basic")
        v = validate_case(r, {"expected_skill": "data_basic"})
        assert v.passed == 1

    def test_skill_mismatch(self):
        r = _make_result_dict(skills_used=["other_skill"], route_mode="fallback")
        v = validate_case(r, {"expected_skill": "data_basic"})
        assert v.failed == 1
        assert "路由未命中" in v.results[0].message


class TestValidateRequiredTools:
    def test_all_present(self):
        r = _make_result_dict(tool_calls=[
            {"tool_name": "read_excel", "success": True},
            {"tool_name": "list_sheets", "success": True},
        ])
        v = validate_case(r, {"required_tools": ["read_excel"]})
        assert v.passed == 1

    def test_missing_tool(self):
        r = _make_result_dict(tool_calls=[
            {"tool_name": "list_sheets", "success": True},
        ])
        v = validate_case(r, {"required_tools": ["read_excel", "list_sheets"]})
        assert v.failed == 1
        assert "read_excel" in v.results[0].message


class TestValidateForbiddenTools:
    def test_no_violations(self):
        r = _make_result_dict(tool_calls=[
            {"tool_name": "read_excel", "success": True},
        ])
        v = validate_case(r, {"forbidden_tools": ["activate_skill"]})
        assert v.passed == 1

    def test_violation_detected(self):
        r = _make_result_dict(tool_calls=[
            {"tool_name": "read_excel", "success": True},
            {"tool_name": "activate_skill", "success": True},
        ])
        v = validate_case(r, {"forbidden_tools": ["activate_skill"]})
        assert v.failed == 1
        assert "activate_skill" in v.results[0].message


class TestValidateNoEmptyPromise:
    def test_first_call_has_tool_calls(self):
        """首轮有 tool_calls → 通过"""
        r = _make_result_dict(llm_calls=[
            {"response": {"content": "", "tool_calls": [{"function": {"name": "read_excel"}}]}},
        ])
        v = validate_case(r, {"no_empty_promise": True})
        assert v.passed == 1

    def test_first_call_empty_promise(self):
        """首轮有文字但无 tool_calls → 失败"""
        r = _make_result_dict(llm_calls=[
            {"response": {"content": "好的，请稍等，我来帮你查看...", "tool_calls": None}},
        ])
        v = validate_case(r, {"no_empty_promise": True})
        assert v.failed == 1
        assert "空承诺" in v.results[0].message

    def test_pure_text_reply_no_tools_needed(self):
        """首轮纯文字回复（如问候）+ 无 tool_calls → 也判为空承诺
        注：对于真正不需要工具的 case，不应启用此断言。"""
        r = _make_result_dict(llm_calls=[
            {"response": {"content": "你好！我可以帮你处理 Excel 文件。", "tool_calls": []}},
        ])
        v = validate_case(r, {"no_empty_promise": True})
        # tool_calls 是空列表，不是 None → 仍然判为空承诺
        assert v.failed == 1

    def test_no_llm_calls(self):
        """无 LLM 调用 → 跳过检查（warning）"""
        r = _make_result_dict(llm_calls=[])
        v = validate_case(r, {"no_empty_promise": True})
        assert v.passed == 1
        assert v.results[0].severity == "warning"


class TestValidateReplyContains:
    def test_all_keywords_present(self):
        r = _make_result_dict(reply="数据包含以下列名：城市、金额、日期")
        v = validate_case(r, {"reply_contains": ["列名", "城市"]})
        assert v.passed == 1

    def test_missing_keyword(self):
        r = _make_result_dict(reply="数据读取完成")
        v = validate_case(r, {"reply_contains": ["列名"]})
        assert v.failed == 1
        assert "列名" in v.results[0].message


class TestValidateReplyNotContains:
    def test_no_violations(self):
        r = _make_result_dict(reply="已完成数据读取")
        v = validate_case(r, {"reply_not_contains": ["抱歉", "无法"]})
        assert v.passed == 1

    def test_violation_detected(self):
        r = _make_result_dict(reply="抱歉，我无法完成这个任务")
        v = validate_case(r, {"reply_not_contains": ["抱歉", "无法"]})
        assert v.failed == 1


# ── validate_case: 多规则组合 ────────────────────────────


class TestValidateMultipleRules:
    def test_all_pass(self):
        r = _make_result_dict(
            status="ok",
            iterations=2,
            total_tokens=10000,
            skills_used=["data_basic"],
        )
        v = validate_case(r, {
            "status": "ok",
            "max_iterations": 5,
            "max_tokens": 50000,
            "expected_skill": "data_basic",
        })
        assert v.total == 4
        assert v.passed == 4
        assert v.failed == 0

    def test_mixed_pass_fail(self):
        r = _make_result_dict(
            status="ok",
            iterations=8,
            total_tokens=10000,
            skills_used=["other_skill"],
        )
        v = validate_case(r, {
            "status": "ok",
            "max_iterations": 5,
            "expected_skill": "data_basic",
        })
        assert v.total == 3
        assert v.passed == 1  # status ok
        assert v.failed == 2  # iterations + skill

    def test_empty_assertions(self):
        r = _make_result_dict()
        v = validate_case(r, {})
        assert v.total == 0
        assert v.passed == 0


# ── aggregate_suite_validation ───────────────────────────


class TestAggregateSuiteValidation:
    def test_all_passed(self):
        validations = [
            ("case_1", ValidationSummary(total=3, passed=3, failed=0, results=[])),
            ("case_2", ValidationSummary(total=2, passed=2, failed=0, results=[])),
        ]
        sv = aggregate_suite_validation(validations)
        assert sv.total_assertions == 5
        assert sv.passed == 5
        assert sv.failed == 0
        assert sv.pass_rate == 100.0
        assert sv.failed_cases == []

    def test_some_failed(self):
        validations = [
            ("case_1", ValidationSummary(total=3, passed=3, failed=0, results=[])),
            ("case_2", ValidationSummary(total=3, passed=1, failed=2, results=[])),
        ]
        sv = aggregate_suite_validation(validations)
        assert sv.total_assertions == 6
        assert sv.passed == 4
        assert sv.failed == 2
        assert sv.failed_cases == ["case_2"]
        assert 60.0 < sv.pass_rate < 70.0  # 66.7%

    def test_empty(self):
        sv = aggregate_suite_validation([])
        assert sv.total_assertions == 0
        assert sv.pass_rate == 100.0


# ── to_dict 序列化 ───────────────────────────────────────


class TestSerialization:
    def test_assertion_result_to_dict(self):
        ar = AssertionResult(
            rule="max_iterations",
            passed=False,
            expected="<= 3",
            actual=5,
            message="超过上限",
        )
        d = ar.to_dict()
        assert d["rule"] == "max_iterations"
        assert d["passed"] is False
        assert d["expected"] == "<= 3"
        assert d["actual"] == 5
        assert "severity" not in d  # default "error" is omitted

    def test_validation_summary_to_dict(self):
        vs = ValidationSummary(total=2, passed=1, failed=1, results=[
            AssertionResult(rule="status", passed=True),
            AssertionResult(rule="max_iterations", passed=False, message="fail"),
        ])
        d = vs.to_dict()
        assert d["total"] == 2
        assert d["passed"] == 1
        assert d["failed"] == 1
        assert len(d["results"]) == 2

    def test_suite_validation_to_dict(self):
        sv = SuiteValidationSummary(
            total_assertions=10,
            passed=8,
            failed=2,
            pass_rate=80.0,
            failed_cases=["case_x"],
        )
        d = sv.to_dict()
        assert d["pass_rate"] == 80.0
        assert d["failed_cases"] == ["case_x"]


# ── golden_cells 断言 ────────────────────────────────────


def _make_golden_pair(tmp_path, *, sheet_name="Sheet1", golden_data=None, output_data=None):
    """创建一对 golden + output xlsx 文件用于测试。"""
    from openpyxl import Workbook

    # golden 文件
    golden_path = tmp_path / "golden.xlsx"
    wb_g = Workbook()
    ws_g = wb_g.active
    ws_g.title = sheet_name
    for row_idx, row_data in enumerate(golden_data or [], 1):
        for col_idx, val in enumerate(row_data, 1):
            ws_g.cell(row=row_idx, column=col_idx, value=val)
    wb_g.save(golden_path)
    wb_g.close()

    # output 文件
    workdir = tmp_path / "workfiles"
    workdir.mkdir()
    output_path = workdir / "output.xlsx"
    wb_o = Workbook()
    ws_o = wb_o.active
    ws_o.title = sheet_name
    for row_idx, row_data in enumerate(output_data or [], 1):
        for col_idx, val in enumerate(row_data, 1):
            ws_o.cell(row=row_idx, column=col_idx, value=val)
    wb_o.save(output_path)
    wb_o.close()

    return workdir, golden_path


class TestGoldenCells:
    """golden_cells 断言的单元测试。"""

    def test_all_cells_match(self, tmp_path):
        data = [["A", 1], ["B", 2], ["C", 3]]
        workdir, golden = _make_golden_pair(
            tmp_path, golden_data=data, output_data=data,
        )
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden),
                "answer_position": "'Sheet1'!A1:B3",
            },
            workfile_dir=workdir,
        )
        assert v.total == 1
        assert v.passed == 1
        golden_result = v.results[0]
        assert golden_result.rule == "golden_cells"
        assert golden_result.passed is True
        assert golden_result.actual["accuracy_pct"] == 100.0

    def test_cells_mismatch(self, tmp_path):
        golden_data = [["A", 1], ["B", 2]]
        output_data = [["A", 99], ["X", 2]]
        workdir, golden = _make_golden_pair(
            tmp_path, golden_data=golden_data, output_data=output_data,
        )
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden),
                "answer_position": "'Sheet1'!A1:B2",
            },
            workfile_dir=workdir,
        )
        assert v.failed == 1
        golden_result = v.results[0]
        assert golden_result.rule == "golden_cells"
        assert golden_result.passed is False
        assert golden_result.actual["matched"] == 2  # A1 and B2 match
        assert golden_result.actual["total_cells"] == 4
        assert len(golden_result.actual["mismatches_sample"]) == 2

    def test_none_vs_empty_string_treated_equal(self, tmp_path):
        golden_data = [[None, ""]]
        output_data = [["", None]]
        workdir, golden = _make_golden_pair(
            tmp_path, golden_data=golden_data, output_data=output_data,
        )
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden),
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=workdir,
        )
        assert v.passed == 1

    def test_float_int_tolerance(self, tmp_path):
        golden_data = [[1, 2.0]]
        output_data = [[1.0, 2]]
        workdir, golden = _make_golden_pair(
            tmp_path, golden_data=golden_data, output_data=output_data,
        )
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden),
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=workdir,
        )
        assert v.passed == 1

    def test_missing_output_file(self, tmp_path):
        golden_path = tmp_path / "golden.xlsx"
        from openpyxl import Workbook
        wb = Workbook(); wb.save(golden_path); wb.close()

        empty_dir = tmp_path / "empty_workdir"
        empty_dir.mkdir()
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden_path),
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=empty_dir,
        )
        assert v.failed == 1
        assert "未找到" in v.results[0].message

    def test_missing_golden_file(self, tmp_path):
        workdir = tmp_path / "workfiles"
        workdir.mkdir()
        from openpyxl import Workbook
        wb = Workbook(); wb.save(workdir / "out.xlsx"); wb.close()

        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": "/nonexistent/golden.xlsx",
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=workdir,
        )
        assert v.failed == 1
        assert "不存在" in v.results[0].message

    def test_no_workfile_dir_skips_golden(self):
        """当 workfile_dir 为 None 时，跳过 golden 断言。"""
        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": "some/golden.xlsx",
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=None,
        )
        # 没有 assertions 也没有可执行的 golden → 空结果
        assert v.total == 0

    def test_golden_combined_with_regular_assertions(self, tmp_path):
        """golden_cells 与常规断言组合使用。"""
        data = [["A", 1]]
        workdir, golden = _make_golden_pair(
            tmp_path, golden_data=data, output_data=data,
        )
        r = _make_result_dict(status="ok", iterations=3)
        v = validate_case(
            r, {"status": "ok", "max_iterations": 5},
            expected={
                "golden_file": str(golden),
                "answer_position": "'Sheet1'!A1:B1",
            },
            workfile_dir=workdir,
        )
        # status + max_iterations + golden_cells = 3 assertions
        assert v.total == 3
        assert v.passed == 3

    def test_formula_output_fails_golden(self, tmp_path):
        """写入公式的输出文件（data_only=True 读取为 None）应当断言失败。"""
        from openpyxl import Workbook

        golden_data = [["(P)", 1], ["(O)", 2]]
        workdir = tmp_path / "workfiles"
        workdir.mkdir()

        # golden 文件：有具体值
        golden_path = tmp_path / "golden.xlsx"
        wb_g = Workbook()
        ws_g = wb_g.active
        ws_g.title = "Sheet1"
        for r, row in enumerate(golden_data, 1):
            for c, val in enumerate(row, 1):
                ws_g.cell(row=r, column=c, value=val)
        wb_g.save(golden_path)
        wb_g.close()

        # output 文件：写入公式（无缓存值）
        output_path = workdir / "output.xlsx"
        wb_o = Workbook()
        ws_o = wb_o.active
        ws_o.title = "Sheet1"
        ws_o["A1"] = '=INDEX(Sheet2!A:A,1)'
        ws_o["B1"] = '=SUM(Sheet2!B:B)'
        ws_o["A2"] = '=INDEX(Sheet2!A:A,2)'
        ws_o["B2"] = '=SUM(Sheet2!C:C)'
        wb_o.save(output_path)
        wb_o.close()

        r = _make_result_dict()
        v = validate_case(
            r, {},
            expected={
                "golden_file": str(golden_path),
                "answer_position": "'Sheet1'!A1:B2",
            },
            workfile_dir=workdir,
        )
        assert v.failed == 1
        golden_result = v.results[0]
        assert golden_result.actual["matched"] == 0
        assert golden_result.actual["total_cells"] == 4
