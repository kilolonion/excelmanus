"""单元格级操作工具测试：write_cells、insert_rows、insert_columns。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excelmanus.tools.cell_tools import (
    get_tools,
    init_guard,
    insert_columns,
    insert_rows,
    write_cells,
)


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为测试目录。"""
    init_guard(str(tmp_path))


def _make_sample(tmp_path: Path, name: str = "sample.xlsx") -> Path:
    """创建一个 3x3 示例 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "姓名"
    ws["B1"] = "年龄"
    ws["C1"] = "城市"
    ws["A2"] = "张三"
    ws["B2"] = 28
    ws["C2"] = "北京"
    ws["A3"] = "李四"
    ws["B3"] = 35
    ws["C3"] = "上海"
    fp = tmp_path / name
    wb.save(fp)
    wb.close()
    return fp


# ── write_cells 测试 ──────────────────────────────────────


class TestWriteCells:
    """write_cells 工具测试套件。"""

    def test_single_cell_value(self, tmp_path: Path) -> None:
        """写入单个数值到指定单元格。"""
        fp = _make_sample(tmp_path)
        result = json.loads(write_cells(str(fp), cell="D1", value="得分"))
        assert result["status"] == "success"
        assert result["cells_written"] == 1

        wb = load_workbook(fp)
        assert wb.active["D1"].value == "得分"
        # 确认其他数据未被影响
        assert wb.active["A1"].value == "姓名"
        assert wb.active["B2"].value == 28
        wb.close()

    def test_single_cell_formula(self, tmp_path: Path) -> None:
        """写入公式到单元格。"""
        fp = _make_sample(tmp_path)
        result = json.loads(write_cells(str(fp), cell="D2", value="=B2*2"))
        assert result["status"] == "success"

        wb = load_workbook(fp)
        assert wb.active["D2"].value == "=B2*2"
        wb.close()

    def test_range_mode_batch_write(self, tmp_path: Path) -> None:
        """范围模式批量写入二维数据。"""
        fp = _make_sample(tmp_path)
        values = [
            ["王五", 42, "广州"],
            ["赵六", 29, "深圳"],
        ]
        result = json.loads(write_cells(str(fp), cell_range="A4", values=values))
        assert result["status"] == "success"
        assert result["rows_written"] == 2
        assert result["cells_written"] == 6

        wb = load_workbook(fp)
        ws = wb.active
        assert ws["A4"].value == "王五"
        assert ws["B4"].value == 42
        assert ws["C5"].value == "深圳"
        # 原数据不受影响
        assert ws["A2"].value == "张三"
        wb.close()

    def test_range_mode_with_full_range(self, tmp_path: Path) -> None:
        """范围模式使用完整范围引用（如 A1:B2）。"""
        fp = _make_sample(tmp_path)
        values = [[100, 200], [300, 400]]
        result = json.loads(write_cells(str(fp), cell_range="E1:F2", values=values))
        assert result["status"] == "success"
        assert result["range"] == "E1:F2"

        wb = load_workbook(fp)
        assert wb.active["E1"].value == 100
        assert wb.active["F2"].value == 400
        wb.close()

    def test_specific_sheet(self, tmp_path: Path) -> None:
        """指定工作表名写入。"""
        fp = _make_sample(tmp_path)
        # 新建第二个 sheet
        wb = load_workbook(fp)
        wb.create_sheet("目标表")
        wb.save(fp)
        wb.close()

        result = json.loads(write_cells(str(fp), sheet_name="目标表", cell="A1", value="测试"))
        assert result["status"] == "success"

        wb = load_workbook(fp)
        assert wb["目标表"]["A1"].value == "测试"
        # 原 Sheet1 不受影响
        assert wb["Sheet1"]["A1"].value == "姓名"
        wb.close()

    def test_mutual_exclusive_modes_error(self, tmp_path: Path) -> None:
        """同时传 cell 和 cell_range 报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(write_cells(str(fp), cell="A1", value=1, cell_range="A1", values=[[1]]))
        assert "error" in result

    def test_no_params_error(self, tmp_path: Path) -> None:
        """不传 cell 也不传 cell_range 报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(write_cells(str(fp)))
        assert "error" in result

    def test_numeric_string_coercion(self, tmp_path: Path) -> None:
        """数字字符串自动转换为数值类型。"""
        fp = _make_sample(tmp_path)
        write_cells(str(fp), cell="D2", value="42.5")
        wb = load_workbook(fp)
        assert wb.active["D2"].value == 42.5
        wb.close()


# ── insert_rows 测试 ──────────────────────────────────────


class TestInsertRows:
    """insert_rows 工具测试套件。"""

    def test_insert_single_row(self, tmp_path: Path) -> None:
        """在第 2 行前插入 1 行。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_rows(str(fp), row=2))
        assert result["status"] == "success"
        assert result["count"] == 1
        assert result["rows_after"] == result["rows_before"] + 1

        wb = load_workbook(fp)
        ws = wb.active
        # 原 A2（张三）应移到 A3
        assert ws["A1"].value == "姓名"
        assert ws["A2"].value is None  # 新插入的空行
        assert ws["A3"].value == "张三"
        wb.close()

    def test_insert_multiple_rows(self, tmp_path: Path) -> None:
        """插入多行。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_rows(str(fp), row=1, count=3))
        assert result["status"] == "success"
        assert result["count"] == 3

        wb = load_workbook(fp)
        ws = wb.active
        # 原 A1（姓名）应移到 A4
        assert ws["A4"].value == "姓名"
        wb.close()

    def test_insert_row_invalid_params(self, tmp_path: Path) -> None:
        """无效参数报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_rows(str(fp), row=0))
        assert "error" in result
        result = json.loads(insert_rows(str(fp), row=1, count=0))
        assert "error" in result


# ── insert_columns 测试 ──────────────────────────────────────


class TestInsertColumns:
    """insert_columns 工具测试套件。"""

    def test_insert_single_column_by_letter(self, tmp_path: Path) -> None:
        """用字母指定位置插入 1 列。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_columns(str(fp), column="B"))
        assert result["status"] == "success"
        assert result["inserted_at_column"] == "B"
        assert result["count"] == 1

        wb = load_workbook(fp)
        ws = wb.active
        # 原 B1（年龄）应移到 C1
        assert ws["A1"].value == "姓名"
        assert ws["B1"].value is None  # 新插入的空列
        assert ws["C1"].value == "年龄"
        wb.close()

    def test_insert_column_by_number(self, tmp_path: Path) -> None:
        """用数字指定位置插入列。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_columns(str(fp), column=1, count=2))
        assert result["status"] == "success"
        assert result["count"] == 2

        wb = load_workbook(fp)
        ws = wb.active
        # 原 A1（姓名）应移到 C1
        assert ws["C1"].value == "姓名"
        wb.close()

    def test_insert_column_invalid_letter(self, tmp_path: Path) -> None:
        """无效列字母报错。"""
        fp = _make_sample(tmp_path)
        result = json.loads(insert_columns(str(fp), column="1A"))
        assert "error" in result


# ── get_tools 测试 ──────────────────────────────────────


def test_get_tools_returns_all() -> None:
    """验证 get_tools 返回正确数量和名称。"""
    tools = get_tools()
    assert len(tools) == 3
    names = {t.name for t in tools}
    assert names == {"write_cells", "insert_rows", "insert_columns"}
