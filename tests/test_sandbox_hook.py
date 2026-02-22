"""运行时沙盒钩子集成测试。"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest

from excelmanus.security.sandbox_hook import generate_wrapper_script


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    return tmp_path


def _run_in_sandbox(workspace: Path, script_content: str, tier: str) -> subprocess.CompletedProcess:
    """辅助函数：在沙盒 wrapper 中执行脚本。"""
    script = workspace / "test_script.py"
    script.write_text(script_content, encoding="utf-8")
    wrapper = generate_wrapper_script(tier, str(workspace))
    wrapper_path = workspace / "_wrapper.py"
    wrapper_path.write_text(wrapper, encoding="utf-8")
    return subprocess.run(
        [sys.executable, str(wrapper_path), str(script)],
        capture_output=True, text=True, timeout=10,
    )


class TestGreenSandbox:
    """GREEN 模式：禁止网络模块导入 + subprocess 函数级拦截。"""

    def test_import_requests_blocked(self, workspace: Path) -> None:
        result = _run_in_sandbox(workspace, "import requests\nprint('should not reach')", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_import_subprocess_allowed(self, workspace: Path) -> None:
        """subprocess 模块允许导入（pandas/matplotlib 内部依赖）。"""
        result = _run_in_sandbox(workspace, "import subprocess\nprint('import_ok')", "GREEN")
        assert result.returncode == 0
        assert "import_ok" in result.stdout

    def test_subprocess_run_blocked(self, workspace: Path) -> None:
        """subprocess.run() 等进程创建函数被拦截。"""
        result = _run_in_sandbox(workspace, "import subprocess\nsubprocess.run(['echo','hi'])", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_subprocess_popen_blocked(self, workspace: Path) -> None:
        result = _run_in_sandbox(workspace, "import subprocess\nsubprocess.Popen(['echo','hi'])", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_subprocess_check_output_blocked(self, workspace: Path) -> None:
        result = _run_in_sandbox(workspace, "import subprocess\nsubprocess.check_output(['echo','hi'])", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_from_os_import_execv_blocked(self, workspace: Path) -> None:
        """回归测试：from os import execv 不应绕过进程创建拦截。"""
        code = "from os import execv\nexecv('/bin/echo', ('echo', 'hi'))"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_import_socket_allowed(self, workspace: Path) -> None:
        """socket 模块允许导入（matplotlib.pyplot 内部依赖）。"""
        result = _run_in_sandbox(workspace, "import socket\nprint('socket_ok')", "GREEN")
        assert result.returncode == 0
        assert "socket_ok" in result.stdout

    def test_socket_create_blocked(self, workspace: Path) -> None:
        """创建 socket 实例被拦截（禁止实际网络通信）。"""
        result = _run_in_sandbox(workspace, "import socket\ns = socket.socket()", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_raw_socket_create_blocked(self, workspace: Path) -> None:
        """回归测试：import _socket 不应绕过网络拦截。"""
        result = _run_in_sandbox(workspace, "import _socket\ns = _socket.socket()", "GREEN")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_socket_gethostname_allowed(self, workspace: Path) -> None:
        """只读信息函数仍可用。"""
        result = _run_in_sandbox(workspace, "import socket\nprint('host:', socket.gethostname())", "GREEN")
        assert result.returncode == 0
        assert "host:" in result.stdout

    def test_matplotlib_pyplot_allowed(self, workspace: Path) -> None:
        """回归测试：matplotlib.pyplot 依赖 socket，确保 GREEN 沙盒不拦截。"""
        code = "import matplotlib; matplotlib.use('Agg'); import matplotlib.pyplot as plt; print('pyplot_ok')"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert "pyplot_ok" in result.stdout

    def test_safe_import_allowed(self, workspace: Path) -> None:
        result = _run_in_sandbox(workspace, "import json\nprint(json.dumps({'ok': True}))", "GREEN")
        assert result.returncode == 0
        assert "ok" in result.stdout

    def test_pandas_import_allowed(self, workspace: Path) -> None:
        """回归测试：pandas 3.x 内部依赖 subprocess，确保 GREEN 沙盒不拦截。"""
        code = "import pandas\nprint('pandas_version:', pandas.__version__)"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert "pandas_version:" in result.stdout

    def test_file_write_inside_workspace_allowed(self, workspace: Path) -> None:
        out = workspace / "output.txt"
        code = f"with open(r'{out}', 'w') as f:\n    f.write('hello')\nprint('done')"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert out.read_text() == "hello"

    def test_file_write_outside_workspace_blocked(self, workspace: Path) -> None:
        # 写入一个不存在的路径（不在工作区内，也不在系统临时目录下）
        target = Path("/tmp/_sandbox_test_should_not_exist/escape.txt")
        code = (
            "import os\n"
            f"os.makedirs(os.path.dirname(r'{target}'), exist_ok=True)\n"
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('evil')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode != 0
        assert not target.exists()

    def test_file_read_outside_workspace_allowed(self, workspace: Path) -> None:
        import tempfile
        outside = Path(tempfile.mkdtemp())
        outside_file = outside / "readable.txt"
        outside_file.write_text("safe_data", encoding="utf-8")
        code = f"with open(r'{outside_file}', 'r') as f:\n    print(f.read())"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert "safe_data" in result.stdout


class TestYellowSandbox:
    """YELLOW 模式：允许网络模块，subprocess 允许导入但函数被拦截。"""

    def test_import_subprocess_allowed(self, workspace: Path) -> None:
        """subprocess 模块允许导入。"""
        result = _run_in_sandbox(workspace, "import subprocess\nprint('import_ok')", "YELLOW")
        assert result.returncode == 0
        assert "import_ok" in result.stdout

    def test_subprocess_run_blocked(self, workspace: Path) -> None:
        """subprocess.run() 仍被拦截。"""
        result = _run_in_sandbox(workspace, "import subprocess\nsubprocess.run(['echo','hi'])", "YELLOW")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_import_ctypes_allowed(self, workspace: Path) -> None:
        """ctypes 已从 YELLOW 禁止列表移除（pandas 等数据处理库间接依赖）。"""
        result = _run_in_sandbox(workspace, "import ctypes\nprint('ctypes_ok')", "YELLOW")
        assert result.returncode == 0
        assert "ctypes_ok" in result.stdout

    def test_import_socket_allowed(self, workspace: Path) -> None:
        # socket should NOT be blocked in YELLOW
        result = _run_in_sandbox(workspace, "import socket\nprint('socket_ok')", "YELLOW")
        assert result.returncode == 0
        assert "socket_ok" in result.stdout

    def test_raw_socket_create_blocked(self, workspace: Path) -> None:
        result = _run_in_sandbox(workspace, "import _socket\ns = _socket.socket()", "YELLOW")
        assert result.returncode != 0
        assert "安全策略禁止" in result.stderr

    def test_network_module_not_blocked(self, workspace: Path) -> None:
        # requests may not be installed, just verify the hook doesn't block it
        code = "try:\n    import requests\n    print('import_ok')\nexcept ImportError:\n    print('not_installed_ok')"
        result = _run_in_sandbox(workspace, code, "YELLOW")
        assert result.returncode == 0
        assert "ok" in result.stdout


class TestRedSandbox:
    """RED 模式：无钩子注入，所有操作允许。"""

    def test_no_restrictions(self, workspace: Path) -> None:
        code = "import json, os\nprint(json.dumps({'pid': os.getpid()}))"
        result = _run_in_sandbox(workspace, code, "RED")
        assert result.returncode == 0
        data = json.loads(result.stdout.strip())
        assert "pid" in data


class TestAutoCoW:
    """自动 Copy-on-Write 行为测试。"""

    def test_auto_cow_on_protected_dir(self, workspace: Path) -> None:
        import os
        
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True, exist_ok=True)
        target = bench_dir / "protected.txt"
        target.write_text("original_data", encoding="utf-8")
        
        outputs_dir = workspace / "outputs"
        
        # EXCELMANUS_BENCH_PROTECTED_DIRS="bench/external" is default
        code = (
            "import os\n"
            f"with open(r'{target}', 'w') as f:\n"
            f"    f.write('new_data')\n"
            f"with open(r'{target}', 'r') as f:\n"
            f"    print('READ:', f.read())\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert "READ: new_data" in result.stdout
        
        # 原文件未被修改
        assert target.read_text(encoding="utf-8") == "original_data"
        
        # 副本已生成并被修改
        cow_file = outputs_dir / "protected.txt"
        assert cow_file.exists()
        assert cow_file.read_text(encoding="utf-8") == "new_data"
        
    def test_auto_cow_openpyxl_save(self, workspace: Path) -> None:
        import os
        
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True, exist_ok=True)
        target = bench_dir / "protected.xlsx"
        
        # 创建一个合法的空 excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "original"
        wb.save(target)
        
        outputs_dir = workspace / "outputs"
        
        code = (
            "import openpyxl\n"
            f"wb = openpyxl.load_workbook(r'{target}')\n"
            "ws = wb.active\n"
            "ws['A1'] = 'new_data'\n"
            f"wb.save(r'{target}')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        
        # 原文件未被修改
        wb_orig = openpyxl.load_workbook(target)
        assert wb_orig.active["A1"].value == "original"
        
        # 副本已生成并被修改
        cow_file = outputs_dir / "protected.xlsx"
        assert cow_file.exists()
        wb_cow = openpyxl.load_workbook(cow_file)
        assert wb_cow.active["A1"].value == "new_data"


class TestWrapperPreservesSemantics:
    """Wrapper 不应破坏 __file__ / __name__ 语义。"""

    def test_file_and_name(self, workspace: Path) -> None:
        script = workspace / "check_env.py"
        script.write_text(
            "import json\nprint(json.dumps({'file': __file__, 'name': __name__}))",
            encoding="utf-8",
        )
        wrapper = generate_wrapper_script("GREEN", str(workspace))
        wrapper_path = workspace / "_wrapper.py"
        wrapper_path.write_text(wrapper, encoding="utf-8")
        result = subprocess.run(
            [sys.executable, str(wrapper_path), str(script)],
            capture_output=True, text=True, timeout=10,
        )
        assert result.returncode == 0
        data = json.loads(result.stdout.strip())
        assert data["name"] == "__main__"
        assert str(script) in data["file"]
