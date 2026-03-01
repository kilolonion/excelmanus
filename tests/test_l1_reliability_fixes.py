"""L1 系统级可靠性修复回归测试。

覆盖三个修复：
- L1-1: Stuck Detection 动态阈值（task_tags 放宽只读循环阈值）
- L1-2: search_excel_values fuzzy 模糊匹配
- L1-3: FinishTaskHandler 输出空值率检测
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from excelmanus.engine_core.session_state import (
    SessionState,
    _ACTION_REPEAT_THRESHOLD,
    _READ_ONLY_LOOP_THRESHOLD,
)


# ═══════════════════════════════════════════════════════════════════
# L1-1: Stuck Detection 动态阈值
# ═══════════════════════════════════════════════════════════════════


class TestStuckDetectionDynamicThreshold:
    """验证 detect_stuck_pattern 根据 task_tags 动态调整阈值。"""

    def _fill_read_only_calls(self, state: SessionState, count: int) -> None:
        """填充 count 个不同参数的只读工具调用。"""
        for i in range(count):
            state.record_tool_call_for_stuck_detection("read_excel", {"file": f"f_{i}.xlsx"})

    def test_default_threshold_triggers(self) -> None:
        """无 task_tags 时，默认阈值 (8) 触发。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, _READ_ONLY_LOOP_THRESHOLD)
        warning = state.detect_stuck_pattern()
        assert warning is not None
        assert "只读循环" in warning

    def test_relaxed_tags_delay_trigger(self) -> None:
        """含 cross_sheet tag 时，默认阈值 (8) 不触发。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, _READ_ONLY_LOOP_THRESHOLD)
        warning = state.detect_stuck_pattern(task_tags=("cross_sheet",))
        assert warning is None, "complex task should NOT trigger at default threshold"

    def test_relaxed_tags_trigger_at_higher_threshold(self) -> None:
        """含 cross_sheet tag 时，放宽阈值 (12) 触发。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, SessionState._RELAXED_READ_ONLY_THRESHOLD)
        warning = state.detect_stuck_pattern(task_tags=("cross_sheet",))
        assert warning is not None
        assert "只读循环" in warning
        assert str(SessionState._RELAXED_READ_ONLY_THRESHOLD) in warning

    def test_non_relaxed_tag_uses_default(self) -> None:
        """非放宽标签不影响阈值。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, _READ_ONLY_LOOP_THRESHOLD)
        warning = state.detect_stuck_pattern(task_tags=("simple",))
        assert warning is not None

    def test_multiple_relaxed_tags(self) -> None:
        """多个放宽标签同时存在时也只放宽一次。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, _READ_ONLY_LOOP_THRESHOLD)
        warning = state.detect_stuck_pattern(task_tags=("cross_sheet", "formatting", "large_data"))
        assert warning is None

    def test_empty_tags_uses_default(self) -> None:
        """空 task_tags 使用默认阈值。"""
        state = SessionState()
        state.current_write_hint = "may_write"
        self._fill_read_only_calls(state, _READ_ONLY_LOOP_THRESHOLD)
        warning = state.detect_stuck_pattern(task_tags=())
        assert warning is not None

    def test_action_repeat_unaffected_by_tags(self) -> None:
        """动作重复检测不受 task_tags 影响。"""
        state = SessionState()
        for _ in range(_ACTION_REPEAT_THRESHOLD):
            state.record_tool_call_for_stuck_detection("read_excel", {"file": "a.xlsx"})
        warning = state.detect_stuck_pattern(task_tags=("cross_sheet",))
        assert warning is not None
        assert "重复操作" in warning


# ═══════════════════════════════════════════════════════════════════
# L1-2: search_excel_values fuzzy 模糊匹配
# ═══════════════════════════════════════════════════════════════════


@pytest.fixture()
def fuzzy_test_workbook(tmp_path: Path) -> Path:
    """创建一个用于 fuzzy 搜索测试的 Excel 文件。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["姓名", "班级", "成绩"])
    ws.append(["张三", "24级电子信息科学与技术1班", 90])
    ws.append(["李四", "24级计算机科学与技术2班", 85])
    ws.append(["王五", "23级软件工程3班", 78])
    ws.append(["赵六", "Advanced Mathematics Class", 92])
    fp = tmp_path / "fuzzy_test.xlsx"
    wb.save(fp)
    return fp


class TestSearchFuzzyMatch:
    """验证 search_excel_values fuzzy 匹配模式。"""

    def _search(self, file_path: Path, query: str, **kwargs) -> dict:
        from excelmanus.tools.data_tools import search_excel_values
        # Monkey-patch guard to allow test paths
        import excelmanus.tools.data_tools as dt
        original_get_guard = dt._get_guard
        mock_guard = MagicMock()
        mock_guard.resolve_and_validate = lambda p: file_path  # return Path, not str
        mock_guard.workspace_root = file_path.parent
        dt._get_guard = lambda: mock_guard
        try:
            result = search_excel_values(str(file_path), query, match_mode="fuzzy", **kwargs)
            return json.loads(result)
        finally:
            dt._get_guard = original_get_guard

    def test_fuzzy_chinese_digit_conversion(self, fuzzy_test_workbook: Path) -> None:
        """搜索'电子一班'应匹配'电子信息科学与技术1班'（中文数字→阿拉伯数字）。"""
        result = self._search(fuzzy_test_workbook, "电子一班")
        assert result["total_matches"] >= 1
        match_val = result["matches"][0]["value"]
        assert "电子" in match_val and "1班" in match_val

    def test_fuzzy_token_split(self, fuzzy_test_workbook: Path) -> None:
        """搜索'计算机2班'应匹配'计算机科学与技术2班'。"""
        result = self._search(fuzzy_test_workbook, "计算机2班")
        assert result["total_matches"] >= 1

    def test_fuzzy_no_match(self, fuzzy_test_workbook: Path) -> None:
        """搜索'物理4班'不应有匹配。"""
        result = self._search(fuzzy_test_workbook, "物理4班")
        assert result["total_matches"] == 0

    def test_fuzzy_english_tokens(self, fuzzy_test_workbook: Path) -> None:
        """搜索'Advanced Class'应匹配'Advanced Mathematics Class'。"""
        result = self._search(fuzzy_test_workbook, "Advanced Class")
        assert result["total_matches"] >= 1

    def test_fuzzy_case_insensitive(self, fuzzy_test_workbook: Path) -> None:
        """模糊匹配默认大小写不敏感。"""
        result = self._search(fuzzy_test_workbook, "advanced class")
        assert result["total_matches"] >= 1

    def test_fuzzy_single_token_fallback(self, fuzzy_test_workbook: Path) -> None:
        """单个 token 时行为类似 contains。"""
        result = self._search(fuzzy_test_workbook, "张三")
        assert result["total_matches"] >= 1


# ═══════════════════════════════════════════════════════════════════
# L1-3: FinishTaskHandler 输出空值率检测
# ═══════════════════════════════════════════════════════════════════


class TestOutputNullRateCheck:
    """验证 FinishTaskHandler._check_output_null_rate 空值率前置检查。"""

    @staticmethod
    def _make_engine_mock(file_path: str, sheet: str = "Sheet1") -> MagicMock:
        """构造一个 mock engine，带 _state.write_operations_log 和 _guard。"""
        engine = MagicMock()
        state = MagicMock()
        state.write_operations_log = [{"file_path": file_path, "sheet": sheet}]
        engine._state = state
        guard = MagicMock()
        guard.resolve_and_validate = lambda p: p
        engine._guard = guard
        return engine

    def test_detects_high_null_rate(self, tmp_path: Path) -> None:
        """空值率 >50% 应触发告警。"""
        fp = tmp_path / "mostly_empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", "B", "C", "D", "E"])
        # 10 行数据，每行只有第一列有值，其余 4 列为空 → 80% 空值率
        for i in range(10):
            ws.append([f"val_{i}", None, None, None, None])
        wb.save(fp)

        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = self._make_engine_mock(str(fp))
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is not None
        assert "空值率异常" in result

    def test_passes_normal_file(self, tmp_path: Path) -> None:
        """正常文件（空值率 <50%）不触发告警。"""
        fp = tmp_path / "normal.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", "B", "C"])
        for i in range(10):
            ws.append([f"a_{i}", f"b_{i}", f"c_{i}"])
        wb.save(fp)

        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = self._make_engine_mock(str(fp))
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

    def test_no_write_log_returns_none(self) -> None:
        """无写入日志时返回 None。"""
        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = MagicMock()
        state = MagicMock()
        state.write_operations_log = []
        engine._state = state
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

    def test_single_row_not_flagged(self, tmp_path: Path) -> None:
        """只有 header + 1 行数据的文件不应触发（row_count <= 1）。"""
        fp = tmp_path / "single_row.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", "B", "C"])
        ws.append([None, None, None])  # 1 row data, all null, but row_count == 1
        wb.save(fp)

        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = self._make_engine_mock(str(fp))
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

    def test_nonexistent_file_skipped(self) -> None:
        """不存在的文件静默跳过。"""
        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = self._make_engine_mock("/nonexistent/path.xlsx")
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

    def test_non_excel_file_skipped(self, tmp_path: Path) -> None:
        """非 Excel 文件静默跳过。"""
        fp = tmp_path / "data.csv"
        fp.write_text("a,b,c\n1,2,3\n")

        from excelmanus.engine_core.tool_handlers import FinishTaskHandler
        engine = self._make_engine_mock(str(fp))
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

    def test_specific_sheet_checked(self, tmp_path: Path) -> None:
        """指定 sheet 只检查该 sheet。"""
        fp = tmp_path / "multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        # Sheet1: 正常
        ws1 = wb.active
        ws1.title = "Good"
        ws1.append(["A", "B"])
        for i in range(5):
            ws1.append([f"a_{i}", f"b_{i}"])
        # Sheet2: 大量空值
        ws2 = wb.create_sheet("Bad")
        ws2.append(["A", "B", "C", "D"])
        for i in range(5):
            ws2.append([f"v_{i}", None, None, None])
        wb.save(fp)

        from excelmanus.engine_core.tool_handlers import FinishTaskHandler

        # 只检查 Good sheet → 不告警
        engine = self._make_engine_mock(str(fp), sheet="Good")
        result = FinishTaskHandler._check_output_null_rate(engine)
        assert result is None

        # 只检查 Bad sheet → 告警
        engine2 = self._make_engine_mock(str(fp), sheet="Bad")
        result2 = FinishTaskHandler._check_output_null_rate(engine2)
        assert result2 is not None
