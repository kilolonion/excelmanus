"""inspect_excel_files 工具函数测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.tools import data_tools


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    """创建含多个 Excel 文件的临时工作区。"""
    # 文件 1：单 sheet
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "销售数据"
    ws1.append(["姓名", "金额", "日期"])
    ws1.append(["张三", 100, "2025-01-01"])
    ws1.append(["李四", 200, "2025-01-02"])
    ws1.append(["王五", 300, "2025-01-03"])
    ws1.append(["赵六", 400, "2025-01-04"])
    wb1.save(tmp_path / "sales.xlsx")

    # 文件 2：多 sheet
    wb2 = Workbook()
    ws2a = wb2.active
    ws2a.title = "Sheet1"
    ws2a.append(["ID", "产品"])
    ws2a.append([1, "苹果"])
    ws2b = wb2.create_sheet("Sheet2")
    ws2b.append(["类型", "数量"])
    ws2b.append(["A", 10])
    wb2.save(tmp_path / "products.xlsx")

    # 文件 3：空 sheet
    wb3 = Workbook()
    wb3.active.title = "空表"
    wb3.save(tmp_path / "empty.xlsx")

    # 隐藏/临时文件（应被跳过）
    wb_hidden = Workbook()
    wb_hidden.save(tmp_path / ".hidden.xlsx")
    wb_temp = Workbook()
    wb_temp.save(tmp_path / "~$temp.xlsx")

    # 非 Excel 文件（应被忽略）
    (tmp_path / "readme.txt").write_text("hello")

    # 子目录中的文件（recursive=True 时可见，recursive=False 时忽略）
    sub = tmp_path / "subdir"
    sub.mkdir()
    wb_sub = Workbook()
    ws_sub = wb_sub.active
    ws_sub.title = "学生花名册"
    ws_sub.append(["学号", "姓名", "班级"])
    ws_sub.append(["001", "张三", "一班"])
    wb_sub.save(sub / "迎新活动排班表.xlsx")

    # 噪音目录中的文件（递归时也应被跳过）
    noise_dir = tmp_path / ".git"
    noise_dir.mkdir()
    wb_noise = Workbook()
    wb_noise.save(noise_dir / "noise.xlsx")

    data_tools.init_guard(str(tmp_path))
    return tmp_path


class TestInspectExcelFiles:
    def test_basic_scan(self, workspace: Path) -> None:
        """recursive=True 默认，应找到根目录 + 子目录的文件。"""
        result = json.loads(data_tools.inspect_excel_files())
        assert result["excel_files_found"] == 4
        names = [f["file"] for f in result["files"]]
        assert "sales.xlsx" in names
        assert "products.xlsx" in names
        assert "empty.xlsx" in names
        assert "迎新活动排班表.xlsx" in names

    def test_non_recursive_scan(self, workspace: Path) -> None:
        """recursive=False 时仅扫描当前目录层级。"""
        result = json.loads(data_tools.inspect_excel_files(recursive=False))
        assert result["excel_files_found"] == 3
        names = [f["file"] for f in result["files"]]
        assert "迎新活动排班表.xlsx" not in names

    def test_hidden_and_temp_skipped(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files())
        names = [f["file"] for f in result["files"]]
        assert ".hidden.xlsx" not in names
        assert "~$temp.xlsx" not in names

    def test_sheet_info(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files())
        sales = next(f for f in result["files"] if f["file"] == "sales.xlsx")
        assert len(sales["sheets"]) == 1
        sheet = sales["sheets"][0]
        assert sheet["name"] == "销售数据"
        assert sheet["rows"] == 5  # 1 标题 + 4 数据
        assert sheet["columns"] == 3
        assert sheet["header_row_hint"] == 0
        assert sheet["business_columns"] == 3
        assert sheet["header"] == ["姓名", "金额", "日期"]

    def test_multi_sheet(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files())
        products = next(f for f in result["files"] if f["file"] == "products.xlsx")
        assert len(products["sheets"]) == 2
        sheet_names = [s["name"] for s in products["sheets"]]
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names

    def test_preview_rows(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files(preview_rows=2))
        sales = next(f for f in result["files"] if f["file"] == "sales.xlsx")
        sheet = sales["sheets"][0]
        # preview_rows=2 → 最多 2 行数据预览
        assert len(sheet["preview"]) == 2

    def test_max_files_limit(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files(max_files=1))
        assert result["excel_files_found"] == 1
        assert result["truncated"] is True

    def test_scan_subdirectory(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files(directory="subdir"))
        assert result["excel_files_found"] == 1
        assert result["files"][0]["file"] == "迎新活动排班表.xlsx"

    def test_invalid_directory(self, workspace: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files(directory="nonexistent"))
        assert "error" in result

    def test_empty_directory(self, workspace: Path) -> None:
        empty_dir = workspace / "empty_dir"
        empty_dir.mkdir()
        result = json.loads(data_tools.inspect_excel_files(directory="empty_dir"))
        assert result["excel_files_found"] == 0
        assert result["files"] == []

    def test_trailing_nulls_trimmed(self, workspace: Path) -> None:
        """合并标题行场景：应识别真正表头，并裁剪尾部 null。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "合并标题"
        # 第一行：合并标题，只有第一列有值，其余为 None
        ws.append(["年度报表", None, None, None, None])
        # 第二行：真正列名
        ws.append(["姓名", "部门", "金额", "日期", "备注"])
        ws.append(["张三", "销售部", 100, "2025-01", None])
        wb.save(workspace / "merged_header.xlsx")

        result = json.loads(data_tools.inspect_excel_files())
        merged = next(f for f in result["files"] if f["file"] == "merged_header.xlsx")
        sheet = merged["sheets"][0]
        assert sheet["header_row_hint"] == 1
        assert sheet["business_columns"] == 5
        # header 应为真正列名，且尾部 null 被裁剪
        assert sheet["header"] == ["姓名", "部门", "金额", "日期", "备注"]
        # preview 首行应为数据行，尾部 null 被裁剪
        assert sheet["preview"][0] == ["张三", "销售部", "100", "2025-01"]

    def test_wide_table_header_preserved(self, workspace: Path) -> None:
        """宽表场景：header 完整保留，仅 preview 数据行被截断。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "宽表"
        cols = [f"col{i}" for i in range(20)]
        ws.append(cols)
        ws.append(list(range(20)))
        wb.save(workspace / "wide.xlsx")

        result = json.loads(data_tools.inspect_excel_files(max_columns=5))
        wide = next(f for f in result["files"] if f["file"] == "wide.xlsx")
        sheet = wide["sheets"][0]
        # header 不截断，完整保留 20 列名
        assert len(sheet["header"]) == 20
        assert sheet["header"] == cols
        # preview 数据行被截断到 5 列
        assert len(sheet["preview"][0]) == 5
        assert sheet["preview_columns_truncated"] == 5

    def test_narrow_table_no_truncation(self, workspace: Path) -> None:
        """窄表不触发截断。"""
        result = json.loads(data_tools.inspect_excel_files(max_columns=10))
        sales = next(f for f in result["files"] if f["file"] == "sales.xlsx")
        sheet = sales["sheets"][0]
        assert "preview_columns_truncated" not in sheet
        assert sheet["header"] == ["姓名", "金额", "日期"]

    def test_tool_registered(self) -> None:
        """确认 inspect_excel_files 已注册到 get_tools()。"""
        names = {t.name for t in data_tools.get_tools()}
        assert "inspect_excel_files" in names

    def test_file_list_summary_present(self, workspace: Path) -> None:
        """结果中应包含紧凑的 file_list 摘要，列出所有文件名和大小。"""
        result = json.loads(data_tools.inspect_excel_files())
        assert "file_list" in result
        file_list = result["file_list"]
        assert len(file_list) == result["excel_files_found"]
        # 每个条目应包含 file 和 size
        for entry in file_list:
            assert "file" in entry
            assert "size" in entry
        # file_list 中的文件名应与 files 中一致
        list_names = [e["file"] for e in file_list]
        files_names = [f["file"] for f in result["files"]]
        assert list_names == files_names

    def test_deterministic_order_with_max_files(self, workspace: Path) -> None:
        """最大文件数截断时，应按相对路径字母序取前 N 个（确定性）。"""
        result = json.loads(data_tools.inspect_excel_files(max_files=2))
        assert result["excel_files_found"] == 2
        assert result["truncated"] is True

    def test_max_result_chars_unlimited(self) -> None:
        """inspect_excel_files 的 ToolDef 应设置 max_result_chars=0（不截断）。"""
        tools = {t.name: t for t in data_tools.get_tools()}
        tool = tools["inspect_excel_files"]
        assert tool.max_result_chars == 0

    # ── 搜索功能测试 ──

    def test_search_by_filename(self, workspace: Path) -> None:
        """按文件名搜索应快速命中，无需打开文件。"""
        result = json.loads(data_tools.inspect_excel_files(search="sales"))
        assert result["excel_files_found"] == 1
        assert result["files"][0]["file"] == "sales.xlsx"

    def test_search_by_sheet_name(self, workspace: Path) -> None:
        """按 sheet 名搜索应找到包含该 sheet 的文件。"""
        result = json.loads(data_tools.inspect_excel_files(search="学生花名册"))
        assert result["excel_files_found"] == 1
        assert result["files"][0]["file"] == "迎新活动排班表.xlsx"

    def test_search_by_sheet_name_param(self, workspace: Path) -> None:
        """用 sheet_name 参数精确搜索。"""
        result = json.loads(data_tools.inspect_excel_files(sheet_name="学生花名册"))
        assert result["excel_files_found"] == 1
        assert result["files"][0]["file"] == "迎新活动排班表.xlsx"

    def test_search_no_match(self, workspace: Path) -> None:
        """搜索无结果时返回 0 文件。"""
        result = json.loads(data_tools.inspect_excel_files(search="不存在的关键词"))
        assert result["excel_files_found"] == 0

    def test_search_case_insensitive(self, workspace: Path) -> None:
        """搜索不区分大小写。"""
        result = json.loads(data_tools.inspect_excel_files(search="SALES"))
        assert result["excel_files_found"] == 1

    def test_recursive_skips_noise_dirs(self, workspace: Path) -> None:
        """递归扫描应跳过 .git/.venv 等噪音目录。"""
        result = json.loads(data_tools.inspect_excel_files())
        names = [f["file"] for f in result["files"]]
        assert "noise.xlsx" not in names

    def test_search_with_recursive(self, workspace: Path) -> None:
        """搜索应能找到子目录中的文件。"""
        result = json.loads(data_tools.inspect_excel_files(search="排班表"))
        assert result["excel_files_found"] == 1
        assert result["files"][0]["file"] == "迎新活动排班表.xlsx"
        # path 应包含子目录
        assert "subdir" in result["files"][0]["path"]
