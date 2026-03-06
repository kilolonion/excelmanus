"""reference_tools 工具测试。"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.tools import reference_tools


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    """创建测试工作簿并初始化 guard。"""
    wb = Workbook()
    ws_product = wb.active
    ws_product.title = "产品表"
    ws_product["A1"] = "ID"
    ws_product["B1"] = "价格"
    ws_product["A2"] = "P001"
    ws_product["B2"] = 100

    ws_order = wb.create_sheet("订单表")
    ws_order["A1"] = "ID"
    ws_order["B1"] = "数量"
    ws_order["C1"] = "单价"
    ws_order["D1"] = "金额"
    ws_order["A2"] = "P001"
    ws_order["B2"] = 5
    ws_order["C2"] = "=VLOOKUP(A2,产品表!A:B,2,0)"
    ws_order["D2"] = "=B2*C2"

    ws_summary = wb.create_sheet("汇总表")
    ws_summary["A1"] = "总金额"
    ws_summary["A2"] = "=SUM(订单表!D:D)"

    fp = tmp_path / "test.xlsx"
    wb.save(fp)

    reference_tools.init_guard(str(tmp_path))
    return tmp_path


class TestGetReferenceMap:
    def test_returns_valid_json(self, workspace: Path) -> None:
        result = reference_tools.get_reference_map("test.xlsx")
        data = json.loads(result)
        assert "sheets" in data
        assert "cross_sheet_edges" in data

    def test_detects_sheets(self, workspace: Path) -> None:
        data = json.loads(reference_tools.get_reference_map("test.xlsx"))
        assert "订单表" in data["sheets"]
        assert "产品表" in data["sheets"]
        assert "汇总表" in data["sheets"]

    def test_detects_cross_sheet_edges(self, workspace: Path) -> None:
        data = json.loads(reference_tools.get_reference_map("test.xlsx"))
        edges = data["cross_sheet_edges"]
        assert len(edges) >= 1
        targets = {e["target_sheet"] for e in edges}
        assert "产品表" in targets


class TestTraceReferences:
    def test_precedents(self, workspace: Path) -> None:
        result = reference_tools.trace_references("test.xlsx", "订单表!C2", direction="precedents")
        data = json.loads(result)
        assert data["target"] == "订单表!C2"
        assert len(data["precedents"]) >= 1

    def test_dependents(self, workspace: Path) -> None:
        result = reference_tools.trace_references("test.xlsx", "订单表!D2", direction="dependents")
        data = json.loads(result)
        assert "dependents" in data


class TestGetImpactAnalysis:
    def test_basic(self, workspace: Path) -> None:
        result = reference_tools.get_impact_analysis("test.xlsx", "产品表!B2")
        data = json.loads(result)
        assert "direct_impact" in data
        assert data["total_affected_cells"] >= 0


class TestGetTools:
    def test_returns_tools(self) -> None:
        tools = reference_tools.get_tools()
        assert len(tools) >= 3
        names = {t.name for t in tools}
        assert "get_reference_map" in names
        assert "trace_references" in names
        assert "get_impact_analysis" in names
