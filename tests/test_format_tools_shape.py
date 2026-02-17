"""测试 read_cell_styles 返回 shape 信息。"""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture()
def sample_xlsx(tmp_path: Path) -> Path:
    """创建一个 3 行 5 列的测试 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 4):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")
    ws["A1"].font = ws["A1"].font.copy(bold=True)
    path = tmp_path / "test_shape.xlsx"
    wb.save(path)
    wb.close()
    return path


class TestReadCellStylesShape:
    """read_cell_styles 应在返回中包含 rows 和 columns 字段。"""

    def test_returns_rows_and_columns(self, sample_xlsx: Path, monkeypatch) -> None:
        from excelmanus.tools.format_tools import read_cell_styles
        from excelmanus.security.guard import FileAccessGuard

        guard = FileAccessGuard(str(sample_xlsx.parent))
        monkeypatch.setattr(
            "excelmanus.tools.format_tools._get_guard", lambda: guard
        )

        result_str = read_cell_styles(
            file_path=str(sample_xlsx),
            cell_range="A1:E1",
        )
        result = json.loads(result_str)

        assert result["rows"] == 3, f"expected rows=3, got {result.get('rows')}"
        assert result["columns"] == 5, f"expected columns=5, got {result.get('columns')}"

    def test_shape_fields_compatible_with_extract_shape(self, sample_xlsx: Path, monkeypatch) -> None:
        from excelmanus.tools.format_tools import read_cell_styles
        from excelmanus.security.guard import FileAccessGuard
        from excelmanus.window_perception.extractor import extract_shape

        guard = FileAccessGuard(str(sample_xlsx.parent))
        monkeypatch.setattr(
            "excelmanus.tools.format_tools._get_guard", lambda: guard
        )

        result_str = read_cell_styles(
            file_path=str(sample_xlsx),
            cell_range="1:1",
        )
        result = json.loads(result_str)
        total_rows, total_cols = extract_shape(result)

        assert total_rows == 3
        assert total_cols == 5
