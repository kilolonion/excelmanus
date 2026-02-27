"""Post-Write Inline Checkpoint 测试。

覆盖：
- _post_write_checkpoint 各写入工具的回读验证
- _checkpoint_write_cells 单元格/范围模式
- _checkpoint_create_sheet / _checkpoint_delete_sheet
- _checkpoint_insert (rows/columns)
- 异常静默处理（文件不存在、路径为空）
- checkpoint 接入 ToolDispatcher 主流程
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl
import pytest


# ── 辅助工厂 ──────────────────────────────────────────────

def _create_test_xlsx(tmp_path: Path, *, sheets: dict[str, list[list]] | None = None) -> Path:
    """在 tmp_path 中创建测试 xlsx 文件。"""
    fp = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if sheets:
        for name, data in sheets.items():
            if name == "Sheet1":
                target = ws
            else:
                target = wb.create_sheet(name)
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, val in enumerate(row_data, 1):
                    target.cell(row=row_idx, column=col_idx, value=val)
    wb.save(str(fp))
    wb.close()
    return fp


# ── _post_write_checkpoint 单元测试 ──────────────────────


class TestPostWriteCheckpointWriteCells:
    """write_cells checkpoint 验证。"""

    def test_single_cell_mode(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={"Sheet1": [["hello"]]})
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells",
            {"file_path": str(fp), "cell": "A1", "value": "hello"},
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "A1" in result
        assert "hello" in result

    def test_range_mode(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={
            "Sheet1": [[1, 2, 3], [4, 5, 6], [7, 8, 9]],
        })
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells",
            {
                "file_path": str(fp),
                "cell_range": "A1",
                "values": [[1, 2, 3], [4, 5, 6], [7, 8, 9]],
            },
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "3行" in result
        assert "3列" in result

    def test_with_sheet_name(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={
            "Sheet1": [],
            "数据": [["a", "b"]],
        })
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells",
            {"file_path": str(fp), "sheet_name": "数据", "cell": "A1", "value": "a"},
            str(tmp_path),
        )
        assert "数据" in result


class TestPostWriteCheckpointCreateSheet:
    def test_sheet_exists(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={"Sheet1": [], "新表": []})
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "create_sheet",
            {"file_path": str(fp), "sheet_name": "新表"},
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "已创建" in result
        assert "新表" in result

    def test_sheet_not_found(self, tmp_path):
        fp = _create_test_xlsx(tmp_path)
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "create_sheet",
            {"file_path": str(fp), "sheet_name": "不存在"},
            str(tmp_path),
        )
        assert "回读异常" in result


class TestPostWriteCheckpointDeleteSheet:
    def test_sheet_deleted(self, tmp_path):
        fp = _create_test_xlsx(tmp_path)  # 只有 Sheet1
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "delete_sheet",
            {"file_path": str(fp), "sheet_name": "不存在的sheet"},
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "已删除" in result

    def test_sheet_still_exists(self, tmp_path):
        fp = _create_test_xlsx(tmp_path)
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "delete_sheet",
            {"file_path": str(fp), "sheet_name": "Sheet1"},
            str(tmp_path),
        )
        assert "回读异常" in result
        assert "仍存在" in result


class TestPostWriteCheckpointInsert:
    def test_insert_rows(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={"Sheet1": [[1], [2], [3]]})
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "insert_rows",
            {"file_path": str(fp), "row": 2, "count": 3},
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "总行数" in result

    def test_insert_columns(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={"Sheet1": [[1, 2, 3]]})
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "insert_columns",
            {"file_path": str(fp), "column": 2, "count": 2},
            str(tmp_path),
        )
        assert "回读确认" in result
        assert "总列数" in result


class TestPostWriteCheckpointEdgeCases:
    def test_empty_file_path(self, tmp_path):
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells", {"file_path": ""}, str(tmp_path),
        )
        assert result == ""

    def test_nonexistent_file(self, tmp_path):
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells",
            {"file_path": str(tmp_path / "nonexistent.xlsx")},
            str(tmp_path),
        )
        assert result == ""

    def test_unknown_tool(self, tmp_path):
        fp = _create_test_xlsx(tmp_path)
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "some_other_tool", {"file_path": str(fp)}, str(tmp_path),
        )
        assert result == ""

    def test_relative_path_resolved(self, tmp_path):
        fp = _create_test_xlsx(tmp_path, sheets={"Sheet1": [["val"]]})
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "write_cells",
            {"file_path": "test.xlsx", "cell": "A1", "value": "val"},
            str(tmp_path),
        )
        assert "回读确认" in result

    def test_create_sheet_no_name(self, tmp_path):
        fp = _create_test_xlsx(tmp_path)
        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        result = ToolDispatcher._post_write_checkpoint(
            "create_sheet", {"file_path": str(fp)}, str(tmp_path),
        )
        assert result == ""
