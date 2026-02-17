"""datetime 序列化 Bug 回归测试。

验证当 header_row 猜错导致 datetime 值成为列名时，
read_excel / filter_data / analyze_data 不会崩溃。
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.tools import data_tools


@pytest.fixture()
def excel_with_datetime_header(tmp_path: Path) -> Path:
    """创建一个含 datetime 值在第一行（模拟标题行）的 Excel 文件。

    结构：
      第 0 行：合并标题 "员工花名册（机密）"
      第 1 行：真正的列头（工号、姓名、性别、出生日期、入职日期、部门、职级）
      第 2 行起：数据（含 datetime 值）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册"
    ws.append(["员工花名册（机密）", None, None, None, None, None, None])
    ws.append(["工号", "姓名", "性别", "出生日期", "入职日期", "部门", "职级"])
    ws.append(["EMP0001", "张三", "男", datetime(1990, 1, 4), datetime(2022, 9, 20), "技术部", "P7"])
    ws.append(["EMP0002", "李四", "女", datetime(1992, 5, 15), datetime(2021, 3, 10), "市场部", "P5"])
    fp = tmp_path / "employees.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture()
def excel_with_deep_header(tmp_path: Path) -> Path:
    """创建一个前 7 行为摘要，第 8 行才是表头的工作簿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"
    ws.append(["2025年度 KPI 仪表盘"])
    ws.append([None])
    ws.append(["总营收", None, "净利润", None, "客户数"])
    ws.append(["98,000,000元", None, "12,250,000元", None, "15,832"])
    ws.append(["↑ 19.5%", None, "↑ 22.3%", None, "↑ 8.7%"])
    ws.append([None])
    ws.append(["月度关键指标趋势"])
    ws.append(["月份", "营收", "成本", "利润", "订单量"])
    ws.append(["2025年01月", 1000, 800, "=B9-C9", 100])
    ws.append(["2025年02月", 1200, 900, "=B10-C10", 120])
    fp = tmp_path / "deep_header.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    data_tools.init_guard(str(tmp_path))


class TestReadExcelDatetimeSafe:
    """read_excel 在 datetime 列名场景下不崩溃。"""

    def test_wrong_header_row_no_crash(self, excel_with_datetime_header: Path) -> None:
        """header_row=2 导致 datetime 值成为列名，应返回有效 JSON 而非崩溃。"""
        result = data_tools.read_excel(
            str(excel_with_datetime_header),
            sheet_name="员工花名册",
            header_row=2,
        )
        parsed = json.loads(result)
        assert "columns" in parsed
        # 所有列名应为字符串
        for col in parsed["columns"]:
            assert isinstance(col, str)

    def test_auto_detect_header_row(self, excel_with_datetime_header: Path) -> None:
        """不指定 header_row 时，应自动检测到 header_row=1。"""
        result = data_tools.read_excel(
            str(excel_with_datetime_header),
            sheet_name="员工花名册",
        )
        parsed = json.loads(result)
        # 自动检测 header_row=1 后，列名应为中文字段名
        assert "工号" in parsed["columns"]
        assert "姓名" in parsed["columns"]
        assert "部门" in parsed["columns"]

    def test_explicit_header_row_0_still_works(self, excel_with_datetime_header: Path) -> None:
        """显式指定 header_row=0 时不应触发自动检测。"""
        result = data_tools.read_excel(
            str(excel_with_datetime_header),
            sheet_name="员工花名册",
            header_row=0,
        )
        parsed = json.loads(result)
        assert "columns" in parsed
        # header_row=0 时第一列是"员工花名册（机密）"
        assert any("员工花名册" in col for col in parsed["columns"])


class TestFilterDataDatetimeSafe:
    """filter_data 在列名不存在时不崩溃。"""

    def test_wrong_header_returns_error_json(self, excel_with_datetime_header: Path) -> None:
        """header_row=2 导致列名匹配失败，应返回错误 JSON 而非异常。"""
        result = data_tools.filter_data(
            str(excel_with_datetime_header),
            column="部门",
            operator="eq",
            value="技术部",
            sheet_name="员工花名册",
            header_row=2,
        )
        parsed = json.loads(result)
        assert "error" in parsed
        # 错误消息中的列名应为字符串
        assert isinstance(parsed["error"], str)

    def test_auto_detect_filter_success(self, excel_with_datetime_header: Path) -> None:
        """不指定 header_row 时自动检测后，filter_data 应成功筛选。"""
        result = data_tools.filter_data(
            str(excel_with_datetime_header),
            column="部门",
            operator="eq",
            value="技术部",
            sheet_name="员工花名册",
        )
        parsed = json.loads(result)
        assert "error" not in parsed
        assert parsed["filtered_rows"] == 1
        assert parsed["data"][0]["姓名"] == "张三"


class TestAnalyzeDataDatetimeSafe:
    """analyze_data 在 datetime 列名场景下不崩溃。"""

    def test_wrong_header_no_crash(self, excel_with_datetime_header: Path) -> None:
        result = data_tools.analyze_data(
            str(excel_with_datetime_header),
            sheet_name="员工花名册",
            header_row=2,
        )
        parsed = json.loads(result)
        assert "columns" in parsed
        for col in parsed["columns"]:
            assert isinstance(col, str)


class TestDetectHeaderRow:
    """_detect_header_row 启发式检测。"""

    def test_standard_header_at_row_0(self, tmp_path: Path) -> None:
        """标准格式（第 0 行即 header）应返回 0。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C", "D"])
        ws.append([1, 2, 3, 4])
        fp = tmp_path / "standard.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) == 0

    def test_title_row_then_header(self, tmp_path: Path) -> None:
        """第 0 行是标题行（1个非空值），第 1 行是 header → 应返回 1。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["报表标题", None, None, None])
        ws.append(["列A", "列B", "列C", "列D"])
        ws.append([1, 2, 3, 4])
        fp = tmp_path / "title_header.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) == 1

    def test_two_title_rows_then_header(self, tmp_path: Path) -> None:
        """前两行都是标题行，第 2 行是 header → 应返回 2。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["大标题", None, None])
        ws.append(["副标题", None, None])
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        fp = tmp_path / "two_titles.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) == 2

    def test_empty_sheet_returns_none(self, tmp_path: Path) -> None:
        """空表应返回 None。"""
        wb = Workbook()
        fp = tmp_path / "empty.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) is None

    def test_specific_sheet_name(self, excel_with_datetime_header: Path) -> None:
        """指定 sheet_name 时应正确检测。"""
        result = data_tools._detect_header_row(excel_with_datetime_header, "员工花名册")
        assert result == 1

    def test_deep_header_detected(self, excel_with_deep_header: Path) -> None:
        """深层表头（第 8 行）应被正确识别。"""
        result = data_tools._detect_header_row(excel_with_deep_header, "KPI")
        assert result == 7

    def test_merged_title_row_skipped(self, tmp_path: Path) -> None:
        """合并标题行（跨多列）应被跳过，检测到真正的 header。"""
        wb = Workbook()
        ws = wb.active
        # 第 0 行：合并标题 "2024年销售数据" 跨 A1:F1
        ws.append(["2024年销售数据", None, None, None, None, None])
        ws.merge_cells("A1:F1")
        # 第 1 行：真正的表头
        ws.append(["月份", "产品", "地区", "销售额", "成本", "利润"])
        # 第 2 行：数据
        ws.append(["1月", "产品A", "华东", 10000, 6000, 4000])
        ws.append(["2月", "产品B", "华北", 12000, 7000, 5000])
        fp = tmp_path / "merged_title.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) == 1

    def test_merged_title_two_rows_skipped(self, tmp_path: Path) -> None:
        """两行合并标题都应被跳过。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["年度销售报表", None, None, None])
        ws.merge_cells("A1:D1")
        ws.append(["生成时间：2024-01", None, None, None])
        ws.merge_cells("A2:D2")
        ws.append(["月份", "产品", "销售额", "利润"])
        ws.append(["1月", "A", 10000, 4000])
        fp = tmp_path / "two_merged_titles.xlsx"
        wb.save(fp)
        assert data_tools._detect_header_row(fp, None) == 2



class TestDeepHeaderRead:
    """深层表头自动读取回归。"""

    def test_read_excel_auto_detect_deep_header(self, excel_with_deep_header: Path) -> None:
        result = data_tools.read_excel(
            str(excel_with_deep_header),
            sheet_name="KPI",
            max_rows=5,
        )
        parsed = json.loads(result)
        assert parsed.get("detected_header_row") == 7
        assert parsed["columns"][:3] == ["月份", "营收", "成本"]


class TestUnnamedFallback:
    """当自动检测的 header_row 产生 Unnamed 列名时，应自动回退到下一行。"""

    def test_fallback_on_unnamed_columns(self, tmp_path: Path) -> None:
        """非合并但内容为空的标题行导致 Unnamed 时，_read_df 应自动重试。"""
        wb = Workbook()
        ws = wb.active
        # 第 0 行：3 个非空值但不是好的列名（会被选为 header 但产生 Unnamed）
        ws["A1"] = "标题A"
        ws["B1"] = "标题B"
        ws["C1"] = "标题C"
        ws["D1"] = None
        ws["E1"] = None
        ws["F1"] = None
        # 第 1 行：真正的表头
        ws["A2"] = "月份"
        ws["B2"] = "产品"
        ws["C2"] = "销售额"
        ws["D2"] = "成本"
        ws["E2"] = "利润"
        ws["F2"] = "地区"
        # 数据行
        ws["A3"] = "1月"
        ws["B3"] = "产品A"
        ws["C3"] = 10000
        ws["D3"] = 6000
        ws["E3"] = 4000
        ws["F3"] = "华东"
        fp = tmp_path / "unnamed_fallback.xlsx"
        wb.save(fp)

        from excelmanus.tools.data_tools import _read_df
        df, effective_header = _read_df(fp, None)
        # 列名不应包含 Unnamed
        unnamed_count = sum(1 for c in df.columns if str(c).startswith("Unnamed"))
        assert unnamed_count == 0, f"列名中仍有 Unnamed: {list(df.columns)}"

class TestUnnamedWarning:
    """read_excel 返回结果中应包含 Unnamed 列名警告。"""

    def test_unnamed_warning_present(self, tmp_path: Path) -> None:
        """当列名中仍有 Unnamed 时，summary 应包含警告字段。"""
        wb = Workbook()
        ws = wb.active
        # 构造一个即使回退也无法消除 Unnamed 的场景：
        # 表头行有 3 个非空 + 3 个空（Unnamed 占比 50%，不触发回退）
        ws.append(["A", "B", "C", None, None, None])
        ws.append([1, 2, 3, 4, 5, 6])
        ws.append([7, 8, 9, 10, 11, 12])
        fp = tmp_path / "all_unnamed.xlsx"
        wb.save(fp)

        data_tools.init_guard(str(tmp_path))
        result_json = json.loads(data_tools.read_excel(str(fp)))
        # 应该有 unnamed_columns_warning 字段
        assert "unnamed_columns_warning" in result_json

    def test_no_warning_when_clean(self, tmp_path: Path) -> None:
        """列名正常时不应有警告。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["月份", "产品", "销售额"])
        ws.append(["1月", "A", 10000])
        fp = tmp_path / "clean.xlsx"
        wb.save(fp)

        data_tools.init_guard(str(tmp_path))
        result_json = json.loads(data_tools.read_excel(str(fp)))
        assert "unnamed_columns_warning" not in result_json

