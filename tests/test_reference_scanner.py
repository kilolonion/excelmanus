"""Tier1Scanner 测试。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.reference_graph.scanner import Tier1Scanner


@pytest.fixture()
def simple_workbook(tmp_path: Path) -> Path:
    """两张表，订单表 VLOOKUP 引用产品表。"""
    wb = Workbook()
    ws_product = wb.active
    ws_product.title = "产品表"
    ws_product["A1"] = "产品ID"
    ws_product["B1"] = "价格"
    ws_product["A2"] = "P001"
    ws_product["B2"] = 100

    ws_order = wb.create_sheet("订单表")
    ws_order["A1"] = "产品ID"
    ws_order["B1"] = "数量"
    ws_order["C1"] = "单价"
    ws_order["D1"] = "金额"
    ws_order["A2"] = "P001"
    ws_order["B2"] = 5
    ws_order["C2"] = "=VLOOKUP(A2,产品表!A:B,2,0)"
    ws_order["D2"] = "=B2*C2"

    fp = tmp_path / "orders.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture()
def named_range_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName("PriceRange", attr_text="Data!$A$1:$A$10")
    wb.defined_names.add(dn)
    fp = tmp_path / "named.xlsx"
    wb.save(fp)
    return fp


class TestTier1Scanner:
    def test_detects_cross_sheet_formula(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        assert index.file_path == str(simple_workbook)
        assert "订单表" in index.sheets
        assert "产品表" in index.sheets

        order_summary = index.sheets["订单表"]
        assert order_summary.formula_count >= 2
        assert len(order_summary.outgoing_refs) >= 1
        target_sheets = {e.target_sheet for e in order_summary.outgoing_refs}
        assert "产品表" in target_sheets

    def test_cross_sheet_edges(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        assert len(index.cross_sheet_edges) >= 1
        edge = index.cross_sheet_edges[0]
        assert edge.source_sheet == "订单表"
        assert edge.target_sheet == "产品表"
        assert edge.ref_count >= 1

    def test_incoming_refs(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        product = index.sheets["产品表"]
        assert len(product.incoming_refs) >= 1

    def test_self_refs(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        order = index.sheets["订单表"]
        assert order.self_refs >= 1  # =B2*C2

    def test_formula_patterns(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        order = index.sheets["订单表"]
        assert "VLOOKUP" in order.formula_patterns

    def test_named_ranges(self, named_range_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(named_range_workbook))
        assert "PriceRange" in index.named_ranges

    def test_render_summary(self, simple_workbook: Path) -> None:
        scanner = Tier1Scanner()
        index = scanner.scan(str(simple_workbook))
        summary = index.render_summary()
        assert "订单表" in summary
        assert "产品表" in summary

    def test_empty_workbook(self, tmp_path: Path) -> None:
        wb = Workbook()
        fp = tmp_path / "empty.xlsx"
        wb.save(fp)
        scanner = Tier1Scanner()
        index = scanner.scan(str(fp))
        assert index.sheets
        assert all(s.formula_count == 0 for s in index.sheets.values())
