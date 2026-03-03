"""discover_file_relationships 工具函数单元测试。

覆盖场景：
- _normalize_column_name 归一化逻辑
- _detect_cross_file_relationships 核心算法
- discover_file_relationships 完整工具（含文件 I/O）
- 边界：单文件、空目录、CSV 文件
- context_builder _try_auto_prescan 跨文件关系注入
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest


# ── _normalize_column_name 测试 ───────────────────────────


class TestNormalizeColumnName:
    """列名归一化函数测试。"""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excelmanus.tools.data_tools import _normalize_column_name
        self.normalize = _normalize_column_name

    def test_exact_synonym_match(self):
        assert self.normalize("customer_id") == "客户id"
        assert self.normalize("CustomerID") == "客户id"
        assert self.normalize("客户编号") == "客户id"

    def test_compact_synonym_match(self):
        """去除空格/下划线/连字符后匹配。"""
        assert self.normalize("Customer ID") == "客户id"
        assert self.normalize("customer-id") == "客户id"

    def test_suffix_strip_match(self):
        """去后缀后匹配同义词表。"""
        assert self.normalize("客户id") == "客户id"
        assert self.normalize("订单编号") == "订单id"

    def test_passthrough(self):
        """不在同义词表中的列名返回 compact 形式。"""
        result = self.normalize("Total Amount")
        assert result == "totalamount"

    def test_empty_and_whitespace(self):
        assert self.normalize("") == ""
        assert self.normalize("   ") == ""

    def test_generic_names_not_mapped(self):
        """name/姓名/date/日期 等泛化列名不应被归一化映射，避免误匹配。"""
        # 这些应该返回 compact 形式，不是同义词映射
        assert self.normalize("name") == "name"
        assert self.normalize("姓名") == "姓名"
        assert self.normalize("名称") == "名称"
        assert self.normalize("date") == "date"
        assert self.normalize("日期") == "日期"
        assert self.normalize("时间") == "时间"
        assert self.normalize("department") == "department"
        assert self.normalize("dept") == "dept"


# ── _detect_cross_file_relationships 测试 ─────────────────


class TestDetectCrossFileRelationships:
    """跨文件关系检测核心算法测试。"""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excelmanus.tools.data_tools import _detect_cross_file_relationships
        self.detect = _detect_cross_file_relationships

    def test_exact_column_match_with_overlap(self):
        """精确列名匹配 + 高值重叠 + 新增字段验证。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["客户ID", "金额"]},
            "b.xlsx": {"Sheet1": ["客户ID", "产品"]},
        }
        df_a = pd.DataFrame({"客户ID": ["C001", "C002", "C003"], "金额": [100, 200, 300]})
        df_b = pd.DataFrame({"客户ID": ["C001", "C002", "C004"], "产品": ["X", "Y", "Z"]})
        file_dfs = {
            "a.xlsx": {"Sheet1": df_a},
            "b.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert len(pairs) == 1
        pair = pairs[0]
        assert pair["file_a"] == "a.xlsx"
        assert pair["file_b"] == "b.xlsx"
        assert len(pair["shared_columns"]) == 1
        col = pair["shared_columns"][0]
        assert col["col_a"] == "客户ID"
        assert col["col_b"] == "客户ID"
        assert col["match_type"] == "exact"
        assert col["overlap_ratio"] > 0.5  # C001, C002 out of 3
        # 新增字段验证
        assert "relationship" in col
        assert col["relationship"] in ("one_to_one", "one_to_many", "many_to_one", "many_to_many")
        assert "suggested_join" in col
        assert col["suggested_join"] in ("left", "right", "inner", "outer")
        assert "type_a" in col
        assert "type_b" in col
        assert "type_compatible" in col
        assert isinstance(col["type_compatible"], bool)

    def test_normalized_match(self):
        """归一化匹配：不同列名但归一化后相同。"""
        file_columns = {
            "sales.xlsx": {"Sheet1": ["customer_id", "amount"]},
            "clients.xlsx": {"Sheet1": ["客户编号", "name"]},
        }
        df_a = pd.DataFrame({"customer_id": ["C001", "C002"], "amount": [100, 200]})
        df_b = pd.DataFrame({"客户编号": ["C001", "C002"], "name": ["Alice", "Bob"]})
        file_dfs = {
            "sales.xlsx": {"Sheet1": df_a},
            "clients.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert len(pairs) == 1
        col = pairs[0]["shared_columns"][0]
        assert col["match_type"] == "normalized"
        assert col["overlap_ratio"] == 1.0
        assert col["relationship"] == "one_to_one"  # 两边唯一值占比都是 100%

    def test_no_relationship(self):
        """无关联列。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["col_x"]},
            "b.xlsx": {"Sheet1": ["col_y"]},
        }
        file_dfs = {
            "a.xlsx": {"Sheet1": pd.DataFrame({"col_x": [1, 2]})},
            "b.xlsx": {"Sheet1": pd.DataFrame({"col_y": [3, 4]})},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert pairs == []

    def test_exact_match_low_overlap_still_reported(self):
        """精确列名匹配但值不重叠 — 仍然报告（exact 匹配不受阈值限制）。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["ID"]},
            "b.xlsx": {"Sheet1": ["ID"]},
        }
        df_a = pd.DataFrame({"ID": ["A1", "A2", "A3"]})
        df_b = pd.DataFrame({"ID": ["B1", "B2", "B3"]})
        file_dfs = {
            "a.xlsx": {"Sheet1": df_a},
            "b.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert len(pairs) == 1
        col = pairs[0]["shared_columns"][0]
        assert col["match_type"] == "exact"
        assert col["overlap_ratio"] == 0.0

    def test_normalized_match_below_threshold_filtered(self):
        """归一化匹配但值重叠低于阈值 — 不报告。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["customer_id"]},
            "b.xlsx": {"Sheet1": ["客户编号"]},
        }
        df_a = pd.DataFrame({"customer_id": ["X1", "X2", "X3", "X4", "X5"]})
        df_b = pd.DataFrame({"客户编号": ["Y1", "Y2", "Y3", "Y4", "Y5"]})
        file_dfs = {
            "a.xlsx": {"Sheet1": df_a},
            "b.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert pairs == []

    def test_multiple_files(self):
        """3 个文件之间的关系检测。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["ID", "val"]},
            "b.xlsx": {"Sheet1": ["ID", "info"]},
            "c.xlsx": {"Sheet1": ["ID", "extra"]},
        }
        shared_ids = ["001", "002", "003"]
        file_dfs = {
            "a.xlsx": {"Sheet1": pd.DataFrame({"ID": shared_ids, "val": [1, 2, 3]})},
            "b.xlsx": {"Sheet1": pd.DataFrame({"ID": shared_ids, "info": ["x", "y", "z"]})},
            "c.xlsx": {"Sheet1": pd.DataFrame({"ID": shared_ids, "extra": [10, 20, 30]})},
        }

        pairs = self.detect(file_columns, file_dfs)
        # 3 个文件两两配对 → C(3,2) = 3 对
        assert len(pairs) == 3

    def test_unnamed_columns_skipped(self):
        """Unnamed 列被过滤。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["Unnamed: 0", "ID"]},
            "b.xlsx": {"Sheet1": ["Unnamed: 0", "ID"]},
        }
        df_a = pd.DataFrame({"Unnamed: 0": [0, 1], "ID": ["A", "B"]})
        df_b = pd.DataFrame({"Unnamed: 0": [0, 1], "ID": ["A", "C"]})
        file_dfs = {
            "a.xlsx": {"Sheet1": df_a},
            "b.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert len(pairs) == 1
        # 只有 ID 列匹配，Unnamed 被跳过
        assert len(pairs[0]["shared_columns"]) == 1
        assert pairs[0]["shared_columns"][0]["col_a"] == "ID"

    def test_generic_column_no_false_positive(self):
        """泛化列名（名称 vs 姓名）不应产生归一化匹配。"""
        file_columns = {
            "a.xlsx": {"Sheet1": ["产品名称", "金额"]},
            "b.xlsx": {"Sheet1": ["客户姓名", "地址"]},
        }
        df_a = pd.DataFrame({"产品名称": ["Widget"], "金额": [100]})
        df_b = pd.DataFrame({"客户姓名": ["Alice"], "地址": ["北京"]})
        file_dfs = {
            "a.xlsx": {"Sheet1": df_a},
            "b.xlsx": {"Sheet1": df_b},
        }

        pairs = self.detect(file_columns, file_dfs)
        # "产品名称" 和 "客户姓名" 不应匹配
        assert pairs == []

    def test_multi_sheet_cross_file(self):
        """跨文件 + 跨 sheet 场景。"""
        file_columns = {
            "a.xlsx": {
                "订单": ["订单号", "客户ID"],
                "产品": ["产品ID", "名称"],
            },
            "b.xlsx": {
                "客户表": ["客户ID", "姓名"],
            },
        }
        df_orders = pd.DataFrame({"订单号": ["O1"], "客户ID": ["C1"]})
        df_products = pd.DataFrame({"产品ID": ["P1"], "名称": ["Widget"]})
        df_clients = pd.DataFrame({"客户ID": ["C1"], "姓名": ["Alice"]})
        file_dfs = {
            "a.xlsx": {"订单": df_orders, "产品": df_products},
            "b.xlsx": {"客户表": df_clients},
        }

        pairs = self.detect(file_columns, file_dfs)
        assert len(pairs) == 1
        shared = pairs[0]["shared_columns"]
        # 应该发现 a.xlsx:订单.客户ID ↔ b.xlsx:客户表.客户ID
        client_match = [c for c in shared if c["col_a"] == "客户ID"]
        assert len(client_match) >= 1
        assert client_match[0]["sheet_a"] == "订单"
        assert client_match[0]["sheet_b"] == "客户表"


# ── discover_file_relationships 完整工具测试 ───────────────


class TestDiscoverFileRelationships:
    """discover_file_relationships 工具函数集成测试。"""

    @pytest.fixture
    def workspace(self, tmp_path: Path) -> Path:
        """创建含多个 Excel 文件的临时工作区。"""
        from openpyxl import Workbook

        # 文件 A：销售数据
        wb_a = Workbook()
        ws_a = wb_a.active
        ws_a.title = "销售"
        ws_a.append(["客户ID", "金额", "日期"])
        ws_a.append(["C001", 100, "2024-01-01"])
        ws_a.append(["C002", 200, "2024-01-02"])
        ws_a.append(["C003", 300, "2024-01-03"])
        wb_a.save(tmp_path / "sales.xlsx")

        # 文件 B：客户数据
        wb_b = Workbook()
        ws_b = wb_b.active
        ws_b.title = "客户表"
        ws_b.append(["客户ID", "姓名", "城市"])
        ws_b.append(["C001", "Alice", "北京"])
        ws_b.append(["C002", "Bob", "上海"])
        ws_b.append(["C004", "Charlie", "广州"])
        wb_b.save(tmp_path / "clients.xlsx")

        # 文件 C：无关数据
        wb_c = Workbook()
        ws_c = wb_c.active
        ws_c.title = "日志"
        ws_c.append(["时间戳", "事件"])
        ws_c.append(["2024-01-01", "login"])
        wb_c.save(tmp_path / "logs.xlsx")

        return tmp_path

    @pytest.fixture
    def _init_guard(self, workspace: Path):
        """初始化 FileAccessGuard。"""
        from excelmanus.tools.data_tools import init_guard
        init_guard(str(workspace))
        yield
        # conftest 的 _reset_tool_guards 会清理

    def test_with_file_paths(self, workspace: Path, _init_guard):
        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[
                str(workspace / "sales.xlsx"),
                str(workspace / "clients.xlsx"),
            ],
        )
        result = json.loads(result_str)
        assert result["files_analyzed"] == 2
        assert len(result["file_pairs"]) >= 1

        # 应该发现 客户ID 列关联
        pair = result["file_pairs"][0]
        cols = pair["shared_columns"]
        client_cols = [c for c in cols if "客户" in c["col_a"] or "客户" in c["col_b"]]
        assert len(client_cols) >= 1

    def test_with_directory(self, workspace: Path, _init_guard):
        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(directory=str(workspace))
        result = json.loads(result_str)
        assert result["files_analyzed"] >= 2

    def test_single_file_returns_empty(self, workspace: Path, _init_guard):
        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[str(workspace / "sales.xlsx")],
        )
        result = json.loads(result_str)
        assert result["files_analyzed"] <= 1
        assert result["file_pairs"] == []

    def test_summary_generated(self, workspace: Path, _init_guard):
        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[
                str(workspace / "sales.xlsx"),
                str(workspace / "clients.xlsx"),
            ],
        )
        result = json.loads(result_str)
        assert "summary" in result
        assert len(result["summary"]) > 0

    def test_csv_support(self, workspace: Path, _init_guard):
        """CSV 文件也能参与跨文件关系发现。"""
        import csv

        csv_path = workspace / "orders.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["客户ID", "订单号"])
            writer.writerow(["C001", "ORD001"])
            writer.writerow(["C002", "ORD002"])

        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[
                str(workspace / "sales.xlsx"),
                str(csv_path),
            ],
        )
        result = json.loads(result_str)
        assert result["files_analyzed"] == 2
        # 应发现 客户ID 关联
        assert len(result["file_pairs"]) >= 1

    def test_merge_hints_and_suggested_groups(self, workspace: Path, _init_guard):
        """验证 merge_hints 和 suggested_groups 字段存在。"""
        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[
                str(workspace / "sales.xlsx"),
                str(workspace / "clients.xlsx"),
            ],
        )
        result = json.loads(result_str)
        assert result["files_analyzed"] == 2

        # merge_hints 应包含可操作的合并建议
        assert "merge_hints" in result
        assert len(result["merge_hints"]) >= 1
        hint = result["merge_hints"][0]
        assert "key_column_a" in hint
        assert "key_column_b" in hint
        assert "suggested_join" in hint
        assert "pandas_hint" in hint
        assert "pd.merge" in hint["pandas_hint"]
        assert "relationship" in hint

        # suggested_groups 应建议将关联文件分组（方面4联动）
        assert "suggested_groups" in result
        groups = result["suggested_groups"]
        assert len(groups) >= 1
        assert "name" in groups[0]
        assert "files" in groups[0]
        assert len(groups[0]["files"]) >= 2

    def test_relationship_directionality(self, workspace: Path, _init_guard):
        """验证关系方向性检测：主表（唯一键）vs 多端。"""
        from openpyxl import Workbook

        # 文件 A：客户主表（ID 唯一）
        wb_master = Workbook()
        ws = wb_master.active
        ws.append(["客户ID", "姓名"])
        for i in range(1, 11):
            ws.append([f"C{i:03d}", f"Name_{i}"])
        wb_master.save(workspace / "master.xlsx")

        # 文件 B：订单表（客户ID 有重复，多端）
        wb_orders = Workbook()
        ws2 = wb_orders.active
        ws2.append(["客户ID", "订单号"])
        for i in range(1, 21):
            ws2.append([f"C{(i % 5) + 1:03d}", f"ORD{i:03d}"])
        wb_orders.save(workspace / "orders.xlsx")

        from excelmanus.tools.data_tools import discover_file_relationships

        result_str = discover_file_relationships(
            file_paths=[
                str(workspace / "master.xlsx"),
                str(workspace / "orders.xlsx"),
            ],
        )
        result = json.loads(result_str)
        assert len(result["file_pairs"]) >= 1
        col = result["file_pairs"][0]["shared_columns"][0]
        # master 侧 unique_ratio 应接近 1.0（主键），orders 侧较低（多端）
        assert col["unique_ratio_a"] >= 0.9  # master 表 ID 唯一
        assert col["unique_ratio_b"] < 0.9   # orders 表 ID 有重复
        assert col["relationship"] == "one_to_many"


# ── context_builder 自动注入测试 ────────────────────────────


class TestAutoPrescanCrossFileInjection:
    """验证 _try_auto_prescan 在 ≥2 文件时注入跨文件关系。"""

    @pytest.fixture
    def workspace(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook

        wb_a = Workbook()
        ws_a = wb_a.active
        ws_a.append(["ID", "Value"])
        ws_a.append(["001", 10])
        ws_a.append(["002", 20])
        wb_a.save(tmp_path / "file_a.xlsx")

        wb_b = Workbook()
        ws_b = wb_b.active
        ws_b.append(["ID", "Name"])
        ws_b.append(["001", "Alice"])
        ws_b.append(["002", "Bob"])
        wb_b.save(tmp_path / "file_b.xlsx")

        return tmp_path

    @pytest.fixture
    def _init_guard(self, workspace: Path):
        from excelmanus.tools.data_tools import init_guard
        init_guard(str(workspace))
        yield

    def test_prescan_injects_cross_file_relationships(self, workspace: Path, _init_guard):
        from excelmanus.engine_core.context_builder import ContextBuilder

        state = MagicMock()
        state.explorer_reports = []

        excel_paths = [
            str(workspace / "file_a.xlsx"),
            str(workspace / "file_b.xlsx"),
        ]

        result = ContextBuilder._try_auto_prescan(excel_paths, state)
        assert result is True

        # 应该有 ≥3 个 report：2 个文件扫描 + 1 个跨文件关系
        reports = state.explorer_reports
        assert len(reports) >= 2

        # 检查是否有跨文件关系报告
        rel_reports = [
            r for r in reports
            if any(
                f.get("type") == "cross_file_relationship"
                for f in r.get("findings", [])
            )
        ]
        assert len(rel_reports) >= 1
        rel_report = rel_reports[0]
        assert "recommendation" in rel_report
        assert any("ID" in f["detail"] for f in rel_report["findings"])

    def test_prescan_single_file_no_cross_file_report(self, workspace: Path, _init_guard):
        """单文件时不应有跨文件关系报告。"""
        from excelmanus.engine_core.context_builder import ContextBuilder

        state = MagicMock()
        state.explorer_reports = []

        excel_paths = [str(workspace / "file_a.xlsx")]

        ContextBuilder._try_auto_prescan(excel_paths, state)

        # 只有 1 个文件扫描报告，无跨文件关系
        rel_reports = [
            r for r in state.explorer_reports
            if any(
                f.get("type") == "cross_file_relationship"
                for f in r.get("findings", [])
            )
        ]
        assert len(rel_reports) == 0
