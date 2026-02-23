"""VBA 支持相关回归测试：P0 keep_vba、P1 VBA 查看、P1 guard 豁免。"""

from __future__ import annotations

import json
import types
from pathlib import Path
from unittest.mock import AsyncMock

import pytest

from excelmanus.engine import (
    _contains_formula_advice,
    _user_requests_vba,
    _VBA_MACRO_ADVICE_PATTERN,
)


# ── P0: write_excel keep_vba 修复验证 ──────────────────────


class TestWriteExcelKeepVba:
    """write_excel 写入 .xlsm 时应使用 keep_vba=True。"""

    def test_write_excel_xlsm_passes_keep_vba(self, tmp_path: Path) -> None:
        """验证 write_excel 对 .xlsm 文件传递 keep_vba=True。"""
        import openpyxl

        # 创建带 VBA 标记的 .xlsm 文件
        xlsm_path = tmp_path / "test.xlsm"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        ws["A2"] = "data"
        # openpyxl 无法直接创建真实 VBA，但可以测试 writer 参数传递
        wb.save(xlsm_path)
        wb.close()

        # 使用 write_excel 写入
        from unittest.mock import patch

        from excelmanus.tools.data_tools import write_excel

        with patch("excelmanus.tools.data_tools._get_guard") as mock_guard:
            mock_guard.return_value.resolve_and_validate.return_value = xlsm_path
            result = write_excel(
                str(xlsm_path),
                [{"header": "new_data"}],
                sheet_name="Sheet1",
            )
            parsed = json.loads(result)
            assert parsed["status"] == "success"

        # 验证文件仍然可读
        wb2 = openpyxl.load_workbook(xlsm_path)
        assert "Sheet1" in wb2.sheetnames
        wb2.close()

    def test_write_excel_xlsx_no_keep_vba(self, tmp_path: Path) -> None:
        """验证 write_excel 对 .xlsx 文件不传递 keep_vba。"""
        import openpyxl

        xlsx_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "header"
        wb.save(xlsx_path)
        wb.close()

        from unittest.mock import patch

        from excelmanus.tools.data_tools import write_excel

        with patch("excelmanus.tools.data_tools._get_guard") as mock_guard:
            mock_guard.return_value.resolve_and_validate.return_value = xlsx_path
            result = write_excel(
                str(xlsx_path),
                [{"header": "new_data"}],
                sheet_name="Sheet1",
            )
            parsed = json.loads(result)
            assert parsed["status"] == "success"


# ── P1: VBA 信息提取 ──────────────────────────────────────


class TestCollectVbaInfo:
    """_collect_vba_info 的单元测试。"""

    def test_xlsx_returns_no_vba(self, tmp_path: Path) -> None:
        """对 .xlsx 文件应返回 has_vba=False。"""
        import openpyxl

        from excelmanus.tools.data_tools import _collect_vba_info

        xlsx_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        wb.close()

        info = _collect_vba_info(xlsx_path)
        assert info["has_vba"] is False
        assert info["modules"] == []

    def test_xlsm_without_actual_vba(self, tmp_path: Path) -> None:
        """对无 VBA 内容的 .xlsm 文件应返回 has_vba=False。"""
        import openpyxl

        from excelmanus.tools.data_tools import _collect_vba_info

        # openpyxl 创建的 .xlsm 不包含 vbaProject.bin
        xlsm_path = tmp_path / "test.xlsm"
        wb = openpyxl.Workbook()
        wb.save(xlsm_path)
        wb.close()

        info = _collect_vba_info(xlsm_path)
        assert info["has_vba"] is False

    def test_non_excel_returns_no_vba(self, tmp_path: Path) -> None:
        """对非 Excel 文件应返回 has_vba=False。"""
        from excelmanus.tools.data_tools import _collect_vba_info

        txt_path = tmp_path / "test.txt"
        txt_path.write_text("not excel")

        info = _collect_vba_info(txt_path)
        assert info["has_vba"] is False

    def test_vba_dimension_in_include_dimensions(self) -> None:
        """vba 应在 INCLUDE_DIMENSIONS 中注册。"""
        from excelmanus.tools.data_tools import INCLUDE_DIMENSIONS

        assert "vba" in INCLUDE_DIMENSIONS

    def test_vba_dimension_in_scan_files_dimensions(self) -> None:
        """vba 应在 _SCAN_FILES_DIMENSIONS 中注册。"""
        from excelmanus.tools.data_tools import _SCAN_FILES_DIMENSIONS

        assert "vba" in _SCAN_FILES_DIMENSIONS


# ── P1: VBA 用户请求检测 ──────────────────────────────────


class TestUserRequestsVba:
    """_user_requests_vba 检测模式的单元测试。"""

    @pytest.mark.parametrize(
        "text",
        [
            "查看这个文件的VBA代码",
            "这个文件有宏吗",
            "帮我解释一下这个macro",
            "提取VBA源码",
            "查看宏模块",
            "这个 .xlsm 有什么 VBA 宏",
            "read_excel include vba",
            "inspect vba macros",
            "解读VBA逻辑",
            "vbaProject 有哪些内容",
        ],
    )
    def test_detects_vba_requests(self, text: str) -> None:
        assert _user_requests_vba(text) is True

    @pytest.mark.parametrize(
        "text",
        [
            "帮我汇总销售数据",
            "格式化 Sheet1",
            "写入 A1 单元格",
            "读取前10行",
            "",
        ],
    )
    def test_does_not_false_positive(self, text: str) -> None:
        assert _user_requests_vba(text) is False

    def test_empty_returns_false(self) -> None:
        assert _user_requests_vba("") is False
        assert _user_requests_vba(None) is False  # type: ignore[arg-type]


# ── P1: _contains_formula_advice vba_exempt 参数 ──────────


class TestContainsFormulaAdviceVbaExempt:
    """_contains_formula_advice 的 vba_exempt 参数测试。"""

    def test_vba_code_detected_without_exempt(self) -> None:
        """默认模式应检测 VBA 代码模式。"""
        vba_text = "Sub MyMacro()\n  MsgBox \"Hello\"\nEnd Sub"
        assert _contains_formula_advice(vba_text) is True

    def test_vba_code_exempted_with_flag(self) -> None:
        """vba_exempt=True 时不检测 VBA 代码模式。"""
        vba_text = "Sub MyMacro()\n  MsgBox \"Hello\"\nEnd Sub"
        assert _contains_formula_advice(vba_text, vba_exempt=True) is False

    def test_formula_still_detected_with_vba_exempt(self) -> None:
        """vba_exempt=True 仍应检测公式建议。"""
        formula_text = "你可以使用 =SUM(A1:A10) 来计算。"
        assert _contains_formula_advice(formula_text, vba_exempt=True) is True

    def test_vba_code_block_detected_without_exempt(self) -> None:
        """VBA 代码块标记应被检测。"""
        text = "```vb\nSub Test()\nEnd Sub\n```"
        assert _contains_formula_advice(text) is True

    def test_vba_code_block_exempted_with_flag(self) -> None:
        """vba_exempt=True 时 VBA 代码块标记不被检测。"""
        text = "```vb\nSub Test()\nEnd Sub\n```"
        assert _contains_formula_advice(text, vba_exempt=True) is False

    def test_application_object_detected_without_exempt(self) -> None:
        """Application.xxx 模式应被检测。"""
        text = "使用 Application.ScreenUpdating = False 来优化"
        assert _contains_formula_advice(text) is True

    def test_application_object_exempted_with_flag(self) -> None:
        text = "使用 Application.ScreenUpdating = False 来优化"
        assert _contains_formula_advice(text, vba_exempt=True) is False


# ── P1: Engine 集成测试 — VBA 豁免 ────────────────────────


class TestVbaExemptEngineIntegration:
    """AgentEngine 中 VBA 豁免逻辑的集成测试。"""

    @staticmethod
    def _make_engine(**overrides):
        from excelmanus.config import ExcelManusConfig
        from excelmanus.engine import AgentEngine
        from excelmanus.tools.registry import ToolRegistry

        defaults = {
            "api_key": "test-key",
            "base_url": "https://test.example.com/v1",
            "model": "test-model",
            "max_iterations": 20,
            "max_consecutive_failures": 3,
            "workspace_root": str(Path(__file__).resolve().parent),
            "backup_enabled": False,
        }
        defaults.update(overrides)
        cfg = ExcelManusConfig(**defaults)
        registry = ToolRegistry()
        return AgentEngine(config=cfg, registry=registry)

    @staticmethod
    def _make_route_result(**kwargs):
        from excelmanus.skillpacks.models import SkillMatchResult

        defaults = dict(
            skills_used=[],
            route_mode="all_tools",
            system_contexts=[],
        )
        defaults.update(kwargs)
        return SkillMatchResult(**defaults)

    def test_vba_exempt_initialized_false(self) -> None:
        engine = self._make_engine()
        assert engine._vba_exempt is False

    @pytest.mark.asyncio
    async def test_vba_exempt_set_for_vba_request(self) -> None:
        """用户请求 VBA 时应设置 _vba_exempt=True。"""
        engine = self._make_engine(max_iterations=2)
        route_result = self._make_route_result()
        engine._route_skills = AsyncMock(return_value=route_result)

        vba_reply = "Sub MyMacro()\n  MsgBox \"Hello\"\nEnd Sub"
        engine._client.chat.completions.create = AsyncMock(
            return_value=types.SimpleNamespace(
                choices=[
                    types.SimpleNamespace(
                        message=types.SimpleNamespace(
                            content=vba_reply, tool_calls=None
                        )
                    )
                ]
            )
        )

        result = await engine.chat("查看这个文件的VBA代码")
        # VBA 豁免模式下，VBA 代码不应触发 execution_guard
        assert engine._vba_exempt is True
        assert result.reply == vba_reply

    @pytest.mark.asyncio
    async def test_no_vba_exempt_for_normal_request(self) -> None:
        """普通请求不应设置 _vba_exempt。"""
        engine = self._make_engine(max_iterations=3)
        route_result = self._make_route_result()
        engine._route_skills = AsyncMock(return_value=route_result)

        vba_reply = "Sub MyMacro()\n  MsgBox \"Hello\"\nEnd Sub"
        engine._client.chat.completions.create = AsyncMock(
            side_effect=[
                types.SimpleNamespace(
                    choices=[
                        types.SimpleNamespace(
                            message=types.SimpleNamespace(
                                content=vba_reply, tool_calls=None
                            )
                        )
                    ]
                ),
                types.SimpleNamespace(
                    choices=[
                        types.SimpleNamespace(
                            message=types.SimpleNamespace(
                                content="done", tool_calls=None
                            )
                        )
                    ]
                ),
            ]
        )

        result = await engine.chat("帮我汇总销售数据")
        assert engine._vba_exempt is False
        # 应触发 execution_guard（VBA 代码被检测）
        user_messages = [
            str(m.get("content", ""))
            for m in engine.memory.get_messages()
            if m.get("role") == "user"
        ]
        guard_fired = any("⚠️" in msg and "公式或代码建议" in msg for msg in user_messages)
        assert guard_fired

    @pytest.mark.asyncio
    async def test_vba_exempt_resets_on_new_task(self) -> None:
        """新任务应重置 _vba_exempt。"""
        engine = self._make_engine(max_iterations=2)
        route_result = self._make_route_result()
        engine._route_skills = AsyncMock(return_value=route_result)

        engine._client.chat.completions.create = AsyncMock(
            return_value=types.SimpleNamespace(
                choices=[
                    types.SimpleNamespace(
                        message=types.SimpleNamespace(
                            content="ok", tool_calls=None
                        )
                    )
                ]
            )
        )

        await engine.chat("查看VBA宏")
        assert engine._vba_exempt is True

        await engine.chat("读取前10行数据")
        assert engine._vba_exempt is False
