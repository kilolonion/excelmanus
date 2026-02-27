"""统一测试：5 个工具的增强返回功能。

覆盖：
1. list_sheets + include
2. format_cells + return_styles
3. write_cells + return_preview
4. inspect_excel_files + include
5. create_excel_chart 返回图表元信息
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from excelmanus.tools import cell_tools, chart_tools, data_tools, format_tools, sheet_tools


@pytest.fixture(autouse=True)
def _init_guards(tmp_path: Path) -> None:
    """初始化所有模块的 FileAccessGuard。"""
    for mod in (data_tools, format_tools, sheet_tools, cell_tools, chart_tools):
        mod.init_guard(str(tmp_path))


@pytest.fixture()
def rich_xlsx(tmp_path: Path) -> Path:
    """包含多种格式信息的测试 Excel 文件。"""
    fp = tmp_path / "report.xlsx"
    wb = Workbook()

    # ── Sheet1: 销售明细 ──
    ws1 = wb.active
    ws1.title = "销售明细"
    headers = ["订单号", "城市", "产品", "数量", "金额"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=12)
    for r in range(2, 12):
        ws1.cell(row=r, column=1, value=f"ORD-{r:03d}")
        ws1.cell(row=r, column=2, value="北京")
        ws1.cell(row=r, column=3, value="笔记本")
        ws1.cell(row=r, column=4, value=r * 2)
        ws1.cell(row=r, column=5, value=r * 1000)
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 15.0
    ws1.column_dimensions["E"].width = 12.0
    ws1.conditional_formatting.add(
        "E2:E11",
        CellIsRule(operator="greaterThan", formula=["5000"], fill=PatternFill(bgColor="C6EFCE")),
    )
    chart = BarChart()
    chart.title = "销售额"
    chart.add_data(Reference(ws1, min_col=5, min_row=1, max_row=11), titles_from_data=True)
    ws1.add_chart(chart, "G2")

    # ── Sheet2: 员工表 ──
    ws2 = wb.create_sheet("员工表")
    for c, h in enumerate(["姓名", "部门", "工资"], 1):
        ws2.cell(row=1, column=c, value=h)
    ws2.cell(row=2, column=1, value="张三")
    ws2.cell(row=2, column=2, value="技术部")
    ws2.cell(row=2, column=3, value=15000)
    ws2.freeze_panes = "A2"

    wb.save(fp)
    wb.close()
    return fp


@pytest.fixture()
def simple_xlsx(tmp_path: Path) -> Path:
    """最简单的 Excel 文件。"""
    fp = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名称"
    ws["B1"] = "数量"
    ws["A2"] = "苹果"
    ws["B2"] = 10
    ws["A3"] = "香蕉"
    ws["B3"] = 20
    wb.save(fp)
    wb.close()
    return fp


# ════════════════════════════════════════════════════════════
# 1. list_sheets + include
# ════════════════════════════════════════════════════════════


class TestListSheetsInclude:
    """list_sheets include 参数测试。"""

    def test_no_include_regression(self, rich_xlsx: Path) -> None:
        result = json.loads(sheet_tools.list_sheets(str(rich_xlsx)))
        assert result["sheet_count"] == 2
        assert "column_names" not in result["sheets"][0]
        assert "freeze_panes" not in result["sheets"][0]

    def test_include_columns(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["columns"])
        )
        s1 = result["sheets"][0]
        assert "column_names" in s1
        assert "订单号" in s1["column_names"]
        assert "金额" in s1["column_names"]

    def test_include_freeze_panes(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["freeze_panes"])
        )
        assert result["sheets"][0]["freeze_panes"] == "A2"
        assert result["sheets"][1]["freeze_panes"] == "A2"

    def test_include_preview(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["preview"], max_preview_rows=3)
        )
        s1 = result["sheets"][0]
        assert "preview" in s1
        assert len(s1["preview"]) == 3

    def test_include_charts(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["charts"])
        )
        assert len(result["sheets"][0]["charts"]) == 1
        assert result["sheets"][1]["charts"] == []

    def test_include_conditional_formatting(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["conditional_formatting"])
        )
        assert len(result["sheets"][0]["conditional_formatting"]) >= 1
        assert result["sheets"][1]["conditional_formatting"] == []

    def test_include_column_widths(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(rich_xlsx), include=["column_widths"])
        )
        widths = result["sheets"][0]["column_widths"]
        assert widths["A"] == 15.0
        assert widths["E"] == 12.0

    def test_multi_include(self, rich_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(
                str(rich_xlsx),
                include=["columns", "freeze_panes", "charts", "column_widths"],
            )
        )
        s1 = result["sheets"][0]
        assert "column_names" in s1
        assert "freeze_panes" in s1
        assert "charts" in s1
        assert "column_widths" in s1

    def test_invalid_dimension_warning(self, simple_xlsx: Path) -> None:
        result = json.loads(
            sheet_tools.list_sheets(str(simple_xlsx), include=["nonexistent"])
        )
        assert "include_warning" in result


# ════════════════════════════════════════════════════════════
# 2. format_cells + return_styles
# ════════════════════════════════════════════════════════════


class TestFormatCellsReturnStyles:
    """format_cells return_styles 参数测试。"""

    def test_no_return_styles_regression(self, simple_xlsx: Path) -> None:
        result = json.loads(
            format_tools.format_cells(
                str(simple_xlsx), "A1:B1", font={"bold": True},
            )
        )
        assert result["status"] == "success"
        assert "after_styles" not in result

    def test_return_styles_true(self, simple_xlsx: Path) -> None:
        result = json.loads(
            format_tools.format_cells(
                str(simple_xlsx), "A1:B1",
                font={"bold": True, "color": "FF0000"},
                return_styles=True,
            )
        )
        assert result["status"] == "success"
        assert "after_styles" in result
        styles = result["after_styles"]
        assert "style_classes" in styles
        assert "cell_style_map" in styles
        # 应有包含 bold 的样式类
        has_bold = any(
            cls.get("font", {}).get("bold") is True
            for cls in styles["style_classes"].values()
        )
        assert has_bold


# ════════════════════════════════════════════════════════════
# 3. write_cells + return_preview
# ════════════════════════════════════════════════════════════


class TestWriteCellsReturnPreview:
    """write_cells return_preview 参数测试。"""

    def test_no_return_preview_regression(self, simple_xlsx: Path) -> None:
        result = json.loads(
            cell_tools.write_cells(str(simple_xlsx), cell="C1", value="价格")
        )
        assert result["status"] == "success"
        assert "preview_after" not in result

    def test_single_cell_preview(self, simple_xlsx: Path) -> None:
        result = json.loads(
            cell_tools.write_cells(
                str(simple_xlsx), cell="C1", value="价格", return_preview=True,
            )
        )
        assert result["status"] == "success"
        assert result["preview_after"] == [["价格"]]

    def test_range_preview(self, simple_xlsx: Path) -> None:
        result = json.loads(
            cell_tools.write_cells(
                str(simple_xlsx),
                cell_range="C1",
                values=[["价格", "备注"], [100, "VIP"], [200, "普通"]],
                return_preview=True,
            )
        )
        assert result["status"] == "success"
        preview = result["preview_after"]
        assert len(preview) == 3
        assert preview[0][0] == "价格"
        assert preview[1][0] == "100"

    def test_formula_cell_preview(self, simple_xlsx: Path) -> None:
        result = json.loads(
            cell_tools.write_cells(
                str(simple_xlsx), cell="C2", value="=B2*2", return_preview=True,
            )
        )
        assert result["status"] == "success"
        # data_only=True 模式下公式可能返回 None（未求值），但不应报错
        assert "preview_after" in result


# ════════════════════════════════════════════════════════════
# 4. inspect_excel_files + include
# ════════════════════════════════════════════════════════════


class TestInspectExcelFilesInclude:
    """inspect_excel_files include 参数测试。"""

    def test_no_include_regression(self, rich_xlsx: Path, tmp_path: Path) -> None:
        result = json.loads(data_tools.inspect_excel_files(str(tmp_path)))
        assert result["excel_files_found"] >= 1
        sheet0 = result["files"][0]["sheets"][0]
        assert "freeze_panes" not in sheet0

    def test_include_freeze_panes(self, rich_xlsx: Path, tmp_path: Path) -> None:
        result = json.loads(
            data_tools.inspect_excel_files(str(tmp_path), include=["freeze_panes"])
        )
        sheet0 = result["files"][0]["sheets"][0]
        assert sheet0["freeze_panes"] == "A2"

    def test_include_charts(self, rich_xlsx: Path, tmp_path: Path) -> None:
        result = json.loads(
            data_tools.inspect_excel_files(str(tmp_path), include=["charts"])
        )
        # 销售明细有 1 个图表
        sheets = result["files"][0]["sheets"]
        sales_sheet = [s for s in sheets if s["name"] == "销售明细"][0]
        assert len(sales_sheet["charts"]) == 1

    def test_include_column_widths(self, rich_xlsx: Path, tmp_path: Path) -> None:
        result = json.loads(
            data_tools.inspect_excel_files(str(tmp_path), include=["column_widths"])
        )
        sheet0 = result["files"][0]["sheets"][0]
        assert "column_widths" in sheet0
        assert sheet0["column_widths"]["A"] == 15.0

    def test_invalid_dimension_warning(self, simple_xlsx: Path, tmp_path: Path) -> None:
        result = json.loads(
            data_tools.inspect_excel_files(str(tmp_path), include=["nonexistent"])
        )
        assert "include_warning" in result


# ════════════════════════════════════════════════════════════
# 5. create_excel_chart 返回图表元信息
# ════════════════════════════════════════════════════════════


class TestCreateExcelChartInfo:
    """create_excel_chart 返回增强元信息测试。"""

    def test_chart_info_returned(self, simple_xlsx: Path) -> None:
        result = json.loads(
            chart_tools.create_excel_chart(
                file_path=str(simple_xlsx),
                chart_type="bar",
                data_range="B1:B3",
                categories_range="A2:A3",
                target_cell="D1",
            )
        )
        assert result["status"] == "success"
        assert "chart_info" in result
        assert "total_charts_on_sheet" in result
        assert result["total_charts_on_sheet"] >= 1
        info = result["chart_info"]
        assert info.get("type") == "bar"

    def test_multiple_charts_count(self, simple_xlsx: Path) -> None:
        # 创建第一个图表
        chart_tools.create_excel_chart(
            file_path=str(simple_xlsx),
            chart_type="bar",
            data_range="B1:B3",
            target_cell="D1",
        )
        # 创建第二个图表
        result = json.loads(
            chart_tools.create_excel_chart(
                file_path=str(simple_xlsx),
                chart_type="line",
                data_range="B1:B3",
                target_cell="D15",
            )
        )
        assert result["total_charts_on_sheet"] == 2
