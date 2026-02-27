"""group_aggregate 工具函数的单元测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest

from excelmanus.tools.data_tools import (
    _coerce_numeric,
    _resolve_formula_columns,
    group_aggregate,
    init_guard,
)


# ── fixtures ─────────────────────────────────────────────


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    """初始化 workspace guard 并返回临时目录。"""
    init_guard(str(tmp_path))
    return tmp_path


@pytest.fixture()
def simple_excel(workspace: Path) -> Path:
    """创建简单销售数据 Excel 文件。"""
    df = pd.DataFrame({
        "城市": ["北京", "上海", "北京", "上海", "广州", "广州", "北京"],
        "产品": ["手机", "手机", "电脑", "电脑", "手机", "电脑", "手机"],
        "数量": [10, 20, 5, 15, 8, 12, 7],
        "单价": [5000, 6000, 8000, 9000, 4500, 7500, 5500],
        "金额": [50000, 120000, 40000, 135000, 36000, 90000, 38500],
    })
    fp = workspace / "sales.xlsx"
    df.to_excel(fp, index=False)
    return fp


@pytest.fixture()
def formula_excel(workspace: Path) -> Path:
    """创建含公式列的 Excel 文件（模拟 stress_test 场景）。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"

    # 合并标题行
    ws.merge_cells("A1:E1")
    ws["A1"] = "测试数据"
    ws.merge_cells("A2:E2")
    ws["A2"] = "生成时间：2026-01-01"

    # header 行 (row 3)
    headers = ["城市", "数量", "单价", "总金额", "折扣后"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)

    # 数据行 (rows 4-8)
    data = [
        ("北京", 10, 100),
        ("上海", 20, 200),
        ("北京", 5, 300),
        ("上海", 15, 150),
        ("广州", 8, 250),
    ]
    for row_idx, (city, qty, price) in enumerate(data, 4):
        ws.cell(row=row_idx, column=1, value=city)
        ws.cell(row=row_idx, column=2, value=qty)
        ws.cell(row=row_idx, column=3, value=price)
        # 公式列：总金额 = 数量 * 单价
        ws.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}")
        # 公式列：折扣后 = 总金额 * 0.9
        ws.cell(row=row_idx, column=5, value=f"=D{row_idx}*0.9")

    fp = workspace / "formula_test.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture()
def text_number_excel(workspace: Path) -> Path:
    """创建含文本格式数值的 Excel 文件。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "城市"
    ws["B1"] = "金额"
    rows = [
        ("北京", "1,234.56元"),
        ("上海", "2,345.67元"),
        ("北京", "3,456.78元"),
    ]
    for i, (city, amount) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=city)
        ws.cell(row=i, column=2, value=amount)

    fp = workspace / "text_number.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture()
def unsupported_formula_excel(workspace: Path) -> Path:
    """创建包含函数公式（SUM）的 Excel，用于未解析公式防误聚合测试。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "测试"
    ws.append(["城市", "A", "B", "函数公式"])
    ws.append(["北京", 1, 2, "=SUM(B2:C2)"])
    ws.append(["上海", 3, 4, "=SUM(B3:C3)"])
    ws.append(["北京", 5, 6, "=SUM(B4:C4)"])
    fp = workspace / "unsupported_formula.xlsx"
    wb.save(fp)
    return fp


# ── _coerce_numeric 测试 ─────────────────────────────────


class TestCoerceNumeric:
    """测试数值强制转换辅助函数。"""

    def test_already_numeric(self) -> None:
        s = pd.Series([1.0, 2.5, 3.0])
        result = _coerce_numeric(s)
        assert result.dtype == float
        assert list(result) == [1.0, 2.5, 3.0]

    def test_comma_separated(self) -> None:
        s = pd.Series(["1,234.56", "2,345.67"])
        result = _coerce_numeric(s)
        assert abs(result.iloc[0] - 1234.56) < 0.01
        assert abs(result.iloc[1] - 2345.67) < 0.01

    def test_yuan_suffix(self) -> None:
        s = pd.Series(["1,234.56元", "2,345.67元"])
        result = _coerce_numeric(s)
        assert abs(result.iloc[0] - 1234.56) < 0.01

    def test_percent(self) -> None:
        s = pd.Series(["10%", "20%", "30%"])
        result = _coerce_numeric(s)
        assert abs(result.iloc[0] - 0.10) < 0.001
        assert abs(result.iloc[2] - 0.30) < 0.001

    def test_unconvertible_stays_nan(self) -> None:
        s = pd.Series(["abc", "def"])
        result = _coerce_numeric(s)
        assert result.isna().all()


# ── group_aggregate 基本功能测试 ─────────────────────────


class TestGroupAggregate:
    """测试 group_aggregate 工具函数。"""

    def test_single_group_count(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"*": "count"},
        ))
        assert "error" not in result
        assert result["total_groups"] == 3
        data = {r["城市"]: r["count"] for r in result["data"]}
        assert data["北京"] == 3
        assert data["上海"] == 2
        assert data["广州"] == 2

    def test_sum_aggregation(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"金额": "sum"},
        ))
        assert "error" not in result
        data = {r["城市"]: r["金额_sum"] for r in result["data"]}
        assert data["北京"] == 128500
        assert data["上海"] == 255000
        assert data["广州"] == 126000

    def test_multiple_aggregations(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"金额": ["sum", "mean"], "*": "count"},
        ))
        assert "error" not in result
        cols = result["columns"]
        assert "金额_sum" in cols
        assert "金额_mean" in cols
        assert "count" in cols

    def test_multi_group_by(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by=["城市", "产品"],
            aggregations={"*": "count"},
        ))
        assert "error" not in result
        assert result["total_groups"] >= 5

    def test_sort_descending(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"金额": "sum"},
            sort_by="金额_sum",
            ascending=False,
        ))
        data = result["data"]
        assert data[0]["城市"] == "上海"  # 最高销售额
        values = [r["金额_sum"] for r in data]
        assert values == sorted(values, reverse=True)

    def test_limit(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"金额": "sum"},
            sort_by="金额_sum",
            ascending=False,
            limit=2,
        ))
        assert result["rows_returned"] == 2
        assert len(result["data"]) == 2

    def test_missing_group_col(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="不存在的列",
            aggregations={"*": "count"},
        ))
        assert "error" in result

    def test_missing_agg_col(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"不存在的列": "sum"},
        ))
        assert "error" in result

    def test_invalid_agg_func(self, simple_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(simple_excel),
            group_by="城市",
            aggregations={"金额": "invalid_func"},
        ))
        assert "error" in result


# ── 公式列求值测试 ───────────────────────────────────────


class TestFormulaResolution:
    """测试公式列自动求值。"""

    def test_formula_columns_resolved(self, formula_excel: Path) -> None:
        """公式列（总金额 = 数量*单价）应被自动求值。"""
        result = json.loads(group_aggregate(
            file_path=str(formula_excel),
            group_by="城市",
            aggregations={"总金额": "sum"},
            sheet_name="销售明细",
        ))
        assert "error" not in result
        data = {r["城市"]: r["总金额_sum"] for r in result["data"]}
        # 北京：10*100 + 5*300 = 2500
        assert abs(data["北京"] - 2500) < 0.01
        # 上海：20*200 + 15*150 = 6250
        assert abs(data["上海"] - 6250) < 0.01
        # 广州：8*250 = 2000
        assert abs(data["广州"] - 2000) < 0.01

    def test_chained_formula_resolved(self, formula_excel: Path) -> None:
        """链式公式列（折扣后 = 总金额*0.9）应被自动求值。"""
        result = json.loads(group_aggregate(
            file_path=str(formula_excel),
            group_by="城市",
            aggregations={"折扣后": "sum"},
            sheet_name="销售明细",
        ))
        assert "error" not in result
        data = {r["城市"]: r["折扣后_sum"] for r in result["data"]}
        # 北京：2500 * 0.9 = 2250
        assert abs(data["北京"] - 2250) < 0.01


# ── 文本数值列自动转换测试 ───────────────────────────────


class TestTextNumberCoercion:
    """测试含文本格式数值列的聚合。"""

    def test_text_yuan_sum(self, text_number_excel: Path) -> None:
        """含 '元' 后缀和千分位的文本型数值列应被自动转换后聚合。"""
        result = json.loads(group_aggregate(
            file_path=str(text_number_excel),
            group_by="城市",
            aggregations={"金额": "sum"},
        ))
        assert "error" not in result
        data = {r["城市"]: r["金额_sum"] for r in result["data"]}
        # 北京：1234.56 + 3456.78 = 4691.34
        assert abs(data["北京"] - 4691.34) < 0.01
        # 上海：2345.67
        assert abs(data["上海"] - 2345.67) < 0.01


class TestFormulaSafetyGuard:
    """测试未解析公式列的安全防护。"""

    def test_unresolved_formula_blocked_for_numeric_agg(self, unsupported_formula_excel: Path) -> None:
        result = json.loads(group_aggregate(
            file_path=str(unsupported_formula_excel),
            group_by="城市",
            aggregations={"函数公式": "sum"},
            sheet_name="测试",
        ))
        assert "error" in result
        assert "未解析公式列" in result["error"]
        assert "函数公式" in result.get("blocked_columns", [])
