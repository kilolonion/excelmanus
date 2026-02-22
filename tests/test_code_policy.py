"""CodePolicyEngine 静态分析单元测试。"""
from __future__ import annotations

import pytest

from excelmanus.security.code_policy import CodePolicyEngine, CodeRiskTier, extract_excel_targets, strip_exit_calls


class TestGreenTier:
    """纯计算 + 安全 IO 代码应归类为 GREEN。"""

    @pytest.mark.parametrize("code", [
        "import pandas as pd\ndf = pd.read_excel('data.xlsx')\nprint(df.head())",
        "import numpy as np\nx = np.array([1,2,3])\nprint(x.mean())",
        "from pathlib import Path\ndata = Path('out.csv').read_text()\nprint(len(data))",
        "import json, csv, re, math, datetime\nprint('ok')",
        "x = 1 + 2\nprint(x)",
        "# just a comment",
        "",
    ])
    def test_safe_code_is_green(self, code: str) -> None:
        result = CodePolicyEngine().analyze(code)
        assert result.tier == CodeRiskTier.GREEN


class TestYellowTier:
    """含网络但无危险操作的代码应归类为 YELLOW。"""

    @pytest.mark.parametrize("code", [
        "import requests\nr = requests.get('https://example.com')\nprint(r.status_code)",
        "from urllib.request import urlopen\ndata = urlopen('https://example.com').read()",
        "import httpx\nclient = httpx.Client()",
        "import socket\ns = socket.socket()",
        "import _socket\ns = _socket.socket()",
    ])
    def test_network_code_is_yellow(self, code: str) -> None:
        result = CodePolicyEngine().analyze(code)
        assert result.tier == CodeRiskTier.YELLOW


class TestRedTier:
    """含危险操作的代码应归类为 RED。"""

    @pytest.mark.parametrize("code", [
        "import subprocess\nsubprocess.run(['ls'])",
        "import os\nos.system('rm -rf /')",
        "exec('print(1)')",
        "eval('__import__(\"os\").system(\"ls\")')",
        "import ctypes",
        "import signal",
        "compile('print(1)', '<string>', 'exec')",
        "__import__('subprocess')",
        "import importlib\nimportlib.import_module('subprocess')",
    ])
    def test_dangerous_code_is_red(self, code: str) -> None:
        result = CodePolicyEngine().analyze(code)
        assert result.tier == CodeRiskTier.RED

    def test_syntax_error_is_red(self) -> None:
        result = CodePolicyEngine().analyze("def f(\n")
        assert result.tier == CodeRiskTier.RED

    def test_obfuscation_base64_exec(self) -> None:
        code = "import base64\nexec(base64.b64decode('cHJpbnQoMSk='))"
        result = CodePolicyEngine().analyze(code)
        assert result.tier == CodeRiskTier.RED


class TestCapabilities:
    """检测到的能力标签应正确反映。"""

    def test_pandas_has_safe_compute(self) -> None:
        result = CodePolicyEngine().analyze("import pandas")
        assert "SAFE_COMPUTE" in result.capabilities

    def test_requests_has_network(self) -> None:
        result = CodePolicyEngine().analyze("import requests")
        assert "NETWORK" in result.capabilities

    def test_subprocess_has_subprocess(self) -> None:
        result = CodePolicyEngine().analyze("import subprocess")
        assert "SUBPROCESS" in result.capabilities


class TestExtraModules:
    """用户自定义白名单/黑名单。"""

    def test_extra_safe_module(self) -> None:
        engine = CodePolicyEngine(extra_safe_modules=("my_custom_lib",))
        result = engine.analyze("import my_custom_lib")
        assert result.tier == CodeRiskTier.GREEN

    def test_extra_blocked_module(self) -> None:
        engine = CodePolicyEngine(extra_blocked_modules=("pandas",))
        result = engine.analyze("import pandas")
        assert result.tier == CodeRiskTier.RED


class TestExtractExcelTargetsVariableArgs:
    """extract_excel_targets 应识别变量参数的写入调用。

    回归测试：conversation_20260221T135637 中 wb.save(file_path) 使用变量
    导致 AST 检测漏检，finish_task 被拒。
    """

    def test_wb_save_with_variable_detected_as_write(self) -> None:
        code = (
            "from openpyxl import load_workbook\n"
            "file_path = './.tmp/test.xlsx'\n"
            "wb = load_workbook(file_path)\n"
            "ws = wb.active\n"
            "ws['A1'] = 42\n"
            "wb.save(file_path)\n"
        )
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) >= 1
        assert write_targets[0].source == "wb.save"
        assert write_targets[0].file_path == "<variable>"

    def test_wb_save_with_literal_still_works(self) -> None:
        code = 'wb.save("output.xlsx")\n'
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) == 1
        assert write_targets[0].file_path == "output.xlsx"

    def test_wb_save_non_excel_literal_skipped(self) -> None:
        code = 'wb.save("output.json")\n'
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) == 0

    def test_df_to_excel_with_variable_detected_as_write(self) -> None:
        code = (
            "import pandas as pd\n"
            "output_path = 'result.xlsx'\n"
            "df = pd.DataFrame({'a': [1]})\n"
            "df.to_excel(output_path, sheet_name='Sheet1')\n"
        )
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) >= 1
        assert write_targets[0].source == "df.to_excel"
        assert write_targets[0].file_path == "<variable>"

    def test_df_to_csv_with_variable_detected_as_write(self) -> None:
        code = "df.to_csv(out_file)\n"
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) == 1
        assert write_targets[0].source == "df.to_csv"

    def test_df_to_excel_with_literal_still_works(self) -> None:
        code = 'df.to_excel("result.xlsx", sheet_name="Summary")\n'
        targets = extract_excel_targets(code)
        write_targets = [t for t in targets if t.operation == "write"]
        assert len(write_targets) == 1
        assert write_targets[0].file_path == "result.xlsx"
        assert write_targets[0].sheet_name == "Summary"


class TestSysExitIsRed:
    """回归测试：sys.exit() 必须被检测为 RED。

    conversation_20260221T152608: LLM 生成的脚本含 sys.exit(1)，
    触发 RED 拦截后无降级重试，导致任务完全失败。
    """

    @pytest.mark.parametrize("code", [
        "import sys\ntry:\n    pass\nexcept Exception:\n    sys.exit(1)",
        "import sys\nsys.exit(0)",
        "import os\nos._exit(1)",
    ])
    def test_exit_calls_are_red(self, code: str) -> None:
        result = CodePolicyEngine().analyze(code)
        assert result.tier == CodeRiskTier.RED
        assert "SUBPROCESS" in result.capabilities


class TestStripExitCalls:
    """strip_exit_calls 自动清洗退出调用。

    回归测试：conversation_20260221T152608 根因修复的防御层。
    清洗后的代码应可降级为 GREEN/YELLOW。
    """

    def test_strips_sys_exit(self) -> None:
        code = (
            "import sys\n"
            "import pandas as pd\n"
            "try:\n"
            "    df = pd.read_excel('data.xlsx')\n"
            "    print(df.head())\n"
            "except FileNotFoundError as e:\n"
            "    print(f'错误: {e}', file=sys.stderr)\n"
            "    sys.exit(1)\n"
            "except Exception as e:\n"
            "    print(f'错误: {e}', file=sys.stderr)\n"
            "    sys.exit(1)\n"
        )
        # 原始代码是 RED
        original = CodePolicyEngine().analyze(code)
        assert original.tier == CodeRiskTier.RED

        # 清洗后应降级
        sanitized = strip_exit_calls(code)
        assert sanitized is not None
        result = CodePolicyEngine().analyze(sanitized)
        assert result.tier == CodeRiskTier.GREEN

    def test_strips_os_exit(self) -> None:
        code = "import os\nprint('hello')\nos._exit(1)\n"
        sanitized = strip_exit_calls(code)
        assert sanitized is not None
        result = CodePolicyEngine().analyze(sanitized)
        assert result.tier != CodeRiskTier.RED

    def test_strips_builtin_exit(self) -> None:
        code = "print('hello')\nexit(1)\n"
        sanitized = strip_exit_calls(code)
        assert sanitized is not None
        assert "exit" not in sanitized or "# sanitized" in sanitized or "pass" in sanitized

    def test_no_exit_returns_none(self) -> None:
        code = "import pandas as pd\ndf = pd.read_excel('data.xlsx')\n"
        assert strip_exit_calls(code) is None

    def test_empty_code_returns_none(self) -> None:
        assert strip_exit_calls("") is None
        assert strip_exit_calls("   ") is None

    def test_syntax_error_returns_none(self) -> None:
        assert strip_exit_calls("def f(\n") is None

    def test_non_exit_red_not_downgraded(self) -> None:
        """含 subprocess 等真正危险操作的代码，清洗退出调用不应降级。"""
        code = "import subprocess\nsubprocess.run(['ls'])\nimport sys\nsys.exit(1)\n"
        sanitized = strip_exit_calls(code)
        # 清洗了 exit，但 subprocess 仍在
        if sanitized is not None:
            result = CodePolicyEngine().analyze(sanitized)
            assert result.tier == CodeRiskTier.RED
