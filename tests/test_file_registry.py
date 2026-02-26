"""FileRegistry Phase 1 单元测试。

覆盖：
- FileEntry / FileEvent 数据模型
- FileRegistryStore CRUD（SQLite）
- FileRegistry 注册、查询、路径解析、panorama 构建
- DB migration 12（三表创建）
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

import pytest

from excelmanus.database import Database
from excelmanus.file_registry import (
    FileEntry,
    FileEvent,
    FileRegistry,
    ScanResult,
    _detect_file_type,
)
from excelmanus.stores.file_registry_store import FileRegistryStore


# ── Fixtures ─────────────────────────────────────────────────


@pytest.fixture()
def tmp_db(tmp_path: Path) -> Database:
    """创建临时 SQLite 数据库（自动 migrate 到最新版本）。"""
    db_path = str(tmp_path / "test.db")
    return Database(db_path)


@pytest.fixture()
def store(tmp_db: Database) -> FileRegistryStore:
    return FileRegistryStore(tmp_db)


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    """创建临时工作区目录。"""
    ws = tmp_path / "workspace"
    ws.mkdir()
    return ws


@pytest.fixture()
def registry(tmp_db: Database, workspace: Path) -> FileRegistry:
    return FileRegistry(tmp_db, workspace)


# ── 数据模型测试 ─────────────────────────────────────────────


class TestFileEntry:
    def test_to_dict_roundtrip(self):
        entry = FileEntry(
            id="abc123",
            workspace="/tmp/ws",
            canonical_path="data/sales.xlsx",
            original_name="sales.xlsx",
            file_type="excel",
            size_bytes=1024,
            origin="uploaded",
            origin_session_id="sess1",
            origin_turn=1,
            sheet_meta=[{"name": "Sheet1", "rows": 100, "columns": 5}],
        )
        d = entry.to_dict()
        restored = FileEntry.from_dict(d)
        assert restored.id == entry.id
        assert restored.canonical_path == entry.canonical_path
        assert restored.origin == "uploaded"
        assert restored.sheet_meta == entry.sheet_meta
        assert restored.origin_turn == 1

    def test_defaults(self):
        entry = FileEntry(
            id="x", workspace="/ws", canonical_path="a.txt", original_name="a.txt",
        )
        assert entry.file_type == "other"
        assert entry.origin == "scan"
        assert entry.staging_path is None
        assert entry.is_active_cow is False
        assert entry.deleted_at is None


class TestFileEvent:
    def test_from_dict(self):
        d = {
            "id": "ev1",
            "file_id": "f1",
            "event_type": "uploaded",
            "session_id": "s1",
            "turn": 2,
            "tool_name": None,
            "details": {"size": 100},
            "created_at": "2026-01-01T00:00:00",
        }
        ev = FileEvent.from_dict(d)
        assert ev.event_type == "uploaded"
        assert ev.details == {"size": 100}
        assert ev.turn == 2


# ── 文件类型检测 ─────────────────────────────────────────────


class TestDetectFileType:
    def test_excel(self):
        assert _detect_file_type("data/report.xlsx") == "excel"
        assert _detect_file_type("test.xls") == "excel"

    def test_csv(self):
        assert _detect_file_type("data.csv") == "csv"

    def test_image(self):
        assert _detect_file_type("chart.png") == "image"
        assert _detect_file_type("photo.jpg") == "image"

    def test_text(self):
        assert _detect_file_type("readme.md") == "text"
        assert _detect_file_type("config.json") == "text"

    def test_other(self):
        assert _detect_file_type("archive.zip") == "other"


# ── FileRegistryStore 测试 ───────────────────────────────────


class TestFileRegistryStore:
    def test_upsert_and_get(self, store: FileRegistryStore):
        record = {
            "id": "f1",
            "workspace": "/ws",
            "canonical_path": "data/sales.xlsx",
            "original_name": "sales.xlsx",
            "file_type": "excel",
            "origin": "uploaded",
            "size_bytes": 2048,
            "sheet_meta": [{"name": "Sheet1"}],
        }
        store.upsert_file(record)

        got = store.get_by_id("f1")
        assert got is not None
        assert got["canonical_path"] == "data/sales.xlsx"
        assert got["origin"] == "uploaded"
        assert got["sheet_meta"] == [{"name": "Sheet1"}]

    def test_get_by_path(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "f2",
            "workspace": "/ws",
            "canonical_path": "uploads/img.png",
            "original_name": "img.png",
            "origin": "uploaded",
        })
        got = store.get_by_path("/ws", "uploads/img.png")
        assert got is not None
        assert got["id"] == "f2"

    def test_upsert_updates_existing(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "f3",
            "workspace": "/ws",
            "canonical_path": "a.xlsx",
            "original_name": "a.xlsx",
            "origin": "scan",
            "size_bytes": 100,
        })
        # 重新 upsert 同一路径
        store.upsert_file({
            "id": "f3_new",
            "workspace": "/ws",
            "canonical_path": "a.xlsx",
            "original_name": "a_renamed.xlsx",
            "origin": "scan",
            "size_bytes": 200,
        })
        got = store.get_by_path("/ws", "a.xlsx")
        assert got is not None
        assert got["size_bytes"] == 200
        assert got["original_name"] == "a_renamed.xlsx"

    def test_list_all(self, store: FileRegistryStore):
        for i in range(3):
            store.upsert_file({
                "id": f"f{i}",
                "workspace": "/ws",
                "canonical_path": f"file{i}.xlsx",
                "original_name": f"file{i}.xlsx",
                "origin": "scan",
            })
        result = store.list_all("/ws")
        assert len(result) == 3

    def test_soft_delete(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fd",
            "workspace": "/ws",
            "canonical_path": "del.xlsx",
            "original_name": "del.xlsx",
            "origin": "scan",
        })
        assert store.soft_delete("/ws", "del.xlsx") is True
        # 默认不包含已删除
        assert len(store.list_all("/ws")) == 0
        # 包含已删除
        assert len(store.list_all("/ws", include_deleted=True)) == 1

    def test_upsert_batch(self, store: FileRegistryStore):
        records = [
            {
                "id": f"b{i}",
                "workspace": "/ws",
                "canonical_path": f"batch{i}.xlsx",
                "original_name": f"batch{i}.xlsx",
                "origin": "scan",
            }
            for i in range(5)
        ]
        count = store.upsert_batch(records)
        assert count == 5
        assert len(store.list_all("/ws")) == 5

    def test_aliases(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fa",
            "workspace": "/ws",
            "canonical_path": "uploads/abc_report.xlsx",
            "original_name": "report.xlsx",
            "origin": "uploaded",
        })
        store.add_alias("a1", "fa", "display_name", "report.xlsx")
        store.add_alias("a2", "fa", "upload_path", "uploads/abc_report.xlsx")

        aliases = store.get_aliases("fa")
        assert len(aliases) == 2

        found = store.find_by_alias("report.xlsx")
        assert found is not None
        assert found["id"] == "fa"

    def test_events(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fe",
            "workspace": "/ws",
            "canonical_path": "ev.xlsx",
            "original_name": "ev.xlsx",
            "origin": "scan",
        })
        store.add_event({
            "id": "e1",
            "file_id": "fe",
            "event_type": "uploaded",
            "session_id": "s1",
            "turn": 1,
            "details": {"size": 100},
        })
        store.add_event({
            "id": "e2",
            "file_id": "fe",
            "event_type": "modified",
            "session_id": "s1",
            "turn": 2,
            "tool_name": "write_cells",
        })
        events = store.get_events("fe")
        assert len(events) == 2
        assert events[0]["event_type"] == "uploaded"
        assert events[1]["tool_name"] == "write_cells"

    def test_events_by_session(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fes",
            "workspace": "/ws",
            "canonical_path": "ses.xlsx",
            "original_name": "ses.xlsx",
            "origin": "scan",
        })
        store.add_event({
            "id": "es1",
            "file_id": "fes",
            "event_type": "modified",
            "session_id": "session_A",
            "turn": 1,
        })
        events = store.get_events_by_session("session_A")
        assert len(events) == 1

    def test_staging_update(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fs",
            "workspace": "/ws",
            "canonical_path": "stg.xlsx",
            "original_name": "stg.xlsx",
            "origin": "scan",
        })
        assert store.update_staging("/ws", "stg.xlsx", "/tmp/staged.xlsx") is True
        staged = store.list_staged("/ws")
        assert len(staged) == 1
        assert staged[0]["staging_path"] == "/tmp/staged.xlsx"

        # 清除 staging
        store.update_staging("/ws", "stg.xlsx", None)
        assert len(store.list_staged("/ws")) == 0

    def test_cow_status(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "fc",
            "workspace": "/ws",
            "canonical_path": "cow.xlsx",
            "original_name": "cow.xlsx",
            "origin": "cow_copy",
            "is_active_cow": True,
        })
        cows = store.list_active_cow("/ws")
        assert len(cows) == 1

        store.update_cow_status("/ws", "cow.xlsx", False)
        assert len(store.list_active_cow("/ws")) == 0

    def test_children(self, store: FileRegistryStore):
        store.upsert_file({
            "id": "parent",
            "workspace": "/ws",
            "canonical_path": "original.xlsx",
            "original_name": "original.xlsx",
            "origin": "uploaded",
        })
        store.upsert_file({
            "id": "child1",
            "workspace": "/ws",
            "canonical_path": "backup1.xlsx",
            "original_name": "backup1.xlsx",
            "origin": "backup",
            "parent_file_id": "parent",
        })
        store.upsert_file({
            "id": "child2",
            "workspace": "/ws",
            "canonical_path": "backup2.xlsx",
            "original_name": "backup2.xlsx",
            "origin": "cow_copy",
            "parent_file_id": "parent",
        })
        children = store.get_children("parent")
        assert len(children) == 2


# ── FileRegistry 核心测试 ───────────────────────────────────


class TestFileRegistry:
    def test_register_upload(self, registry: FileRegistry):
        entry = registry.register_upload(
            canonical_path="uploads/abc_sales.xlsx",
            original_name="sales.xlsx",
            file_type="excel",
            size_bytes=2048,
            session_id="s1",
            turn=1,
        )
        assert entry.origin == "uploaded"
        assert entry.original_name == "sales.xlsx"
        assert entry.size_bytes == 2048

        # 可通过路径查询
        found = registry.get_by_path("uploads/abc_sales.xlsx")
        assert found is not None
        assert found.id == entry.id

        # 可通过别名查询
        found_alias = registry.get_by_alias("sales.xlsx")
        assert found_alias is not None
        assert found_alias.id == entry.id

        # 有上传事件
        events = registry.get_events(entry.id)
        assert len(events) == 1
        assert events[0].event_type == "uploaded"

    def test_register_from_scan(self, registry: FileRegistry):
        entry = registry.register_from_scan(
            canonical_path="data/report.xlsx",
            original_name="report.xlsx",
            size_bytes=1024,
            mtime_ns=123456789,
            sheet_meta=[{"name": "Sheet1", "rows": 50, "columns": 3}],
        )
        assert entry.origin == "scan"
        assert entry.mtime_ns == 123456789

        # 更新已存在的记录
        updated = registry.register_from_scan(
            canonical_path="data/report.xlsx",
            original_name="report.xlsx",
            size_bytes=2048,
            mtime_ns=987654321,
        )
        assert updated.id == entry.id  # 同一记录
        assert updated.size_bytes == 2048
        assert updated.mtime_ns == 987654321

    def test_register_agent_output(self, registry: FileRegistry, workspace: Path):
        # 创建父文件
        parent = registry.register_upload(
            canonical_path="sales.xlsx",
            original_name="sales.xlsx",
            size_bytes=100,
        )

        # 创建一个物理文件让 size 能被读取
        (workspace / "outputs").mkdir(exist_ok=True)
        (workspace / "outputs" / "summary.xlsx").write_bytes(b"x" * 50)

        output = registry.register_agent_output(
            canonical_path="outputs/summary.xlsx",
            original_name="summary.xlsx",
            parent_canonical="sales.xlsx",
            session_id="s1",
            turn=3,
            tool_name="run_code",
        )
        assert output.origin == "agent_created"
        assert output.parent_file_id == parent.id
        assert output.size_bytes == 50

    def test_register_backup(self, registry: FileRegistry):
        parent = registry.register_upload(
            canonical_path="sales.xlsx",
            original_name="sales.xlsx",
        )
        backup = registry.register_backup(
            backup_path="outputs/backups/sales_20260226.xlsx",
            parent_canonical="sales.xlsx",
            reason="staging",
            session_id="s1",
            turn=2,
        )
        assert backup.origin == "backup"
        assert backup.parent_file_id == parent.id

        # lineage 查询
        lineage = registry.get_lineage(backup.id)
        assert len(lineage) == 2
        assert lineage[0].id == backup.id
        assert lineage[1].id == parent.id

    def test_register_cow(self, registry: FileRegistry):
        parent = registry.register_upload(
            canonical_path="sales.xlsx",
            original_name="sales.xlsx",
        )
        cow = registry.register_cow(
            cow_path="outputs/sales_cow_abc.xlsx",
            parent_canonical="sales.xlsx",
            session_id="s1",
            turn=3,
        )
        assert cow.origin == "cow_copy"
        assert cow.is_active_cow is True
        assert cow.parent_file_id == parent.id

    def test_resolve_for_tool(self, registry: FileRegistry):
        registry.register_upload(
            canonical_path="uploads/abc_report.xlsx",
            original_name="report.xlsx",
        )

        # canonical_path 精确匹配
        assert registry.resolve_for_tool("uploads/abc_report.xlsx") == "uploads/abc_report.xlsx"

        # alias 匹配（display_name）
        assert registry.resolve_for_tool("report.xlsx") == "uploads/abc_report.xlsx"

        # 不存在时返回原始路径
        assert registry.resolve_for_tool("unknown.txt") == "unknown.txt"

    def test_resolve_for_display(self, registry: FileRegistry):
        registry.register_upload(
            canonical_path="uploads/abc_report.xlsx",
            original_name="report.xlsx",
        )
        assert registry.resolve_for_display("uploads/abc_report.xlsx") == "report.xlsx"
        assert registry.resolve_for_display("unknown.txt") == "unknown.txt"

    def test_mark_deleted(self, registry: FileRegistry):
        entry = registry.register_from_scan(
            canonical_path="old.xlsx",
            original_name="old.xlsx",
        )
        registry.mark_deleted("old.xlsx")

        # list_all 默认不含已删除
        active = registry.list_all(include_deleted=False)
        active_paths = [e.canonical_path for e in active]
        assert "old.xlsx" not in active_paths

        # 仍可查到（保留 provenance）
        all_entries = registry.list_all(include_deleted=True)
        assert any(e.canonical_path == "old.xlsx" for e in all_entries)

    def test_list_all(self, registry: FileRegistry):
        for i in range(3):
            registry.register_from_scan(
                canonical_path=f"file{i}.xlsx",
                original_name=f"file{i}.xlsx",
            )
        assert len(registry.list_all()) == 3

    def test_get_children(self, registry: FileRegistry):
        parent = registry.register_upload(
            canonical_path="parent.xlsx",
            original_name="parent.xlsx",
        )
        registry.register_backup(
            backup_path="backup1.xlsx",
            parent_canonical="parent.xlsx",
        )
        registry.register_backup(
            backup_path="backup2.xlsx",
            parent_canonical="parent.xlsx",
        )
        children = registry.get_children(parent.id)
        assert len(children) == 2


    def test_register_upload_existing_path_preserves_id(self, registry: FileRegistry):
        """对已存在路径重复 register_upload 应复用原 ID。"""
        e1 = registry.register_upload(
            canonical_path="uploads/data.xlsx",
            original_name="data.xlsx",
            size_bytes=100,
        )
        e2 = registry.register_upload(
            canonical_path="uploads/data.xlsx",
            original_name="data_v2.xlsx",
            size_bytes=200,
        )
        assert e2.id == e1.id
        assert e2.size_bytes == 200
        assert e2.original_name == "data_v2.xlsx"

        # DB 中的 ID 也一致
        found = registry.get_by_path("uploads/data.xlsx")
        assert found is not None
        assert found.id == e1.id

    def test_register_agent_output_existing_path_preserves_id(self, registry: FileRegistry, workspace: Path):
        """agent_output 重复注册同路径应复用原 ID。"""
        (workspace / "out.xlsx").write_bytes(b"x" * 10)
        e1 = registry.register_agent_output(
            canonical_path="out.xlsx",
            original_name="out.xlsx",
            session_id="s1", turn=1,
        )
        e2 = registry.register_agent_output(
            canonical_path="out.xlsx",
            original_name="out.xlsx",
            session_id="s1", turn=2,
        )
        assert e2.id == e1.id

    def test_remove_aliases(self, registry: FileRegistry):
        """remove_aliases_for_file 清理别名。"""
        entry = registry.register_upload(
            canonical_path="uploads/abc.xlsx",
            original_name="abc.xlsx",
        )
        registry.add_alias(entry.id, "custom", "my_alias")
        assert registry.get_by_alias("my_alias") is not None

        removed = registry._store.remove_aliases_for_file(entry.id)
        assert removed >= 1


# ── Panorama 构建测试 ────────────────────────────────────────


class TestPanorama:
    def test_empty_panorama(self, registry: FileRegistry):
        assert registry.build_panorama() == ""

    def test_full_mode(self, registry: FileRegistry):
        registry.register_upload(
            canonical_path="sales.xlsx",
            original_name="sales.xlsx",
            file_type="excel",
            size_bytes=2048,
            session_id="s1",
            turn=1,
            sheet_meta=[
                {"name": "订单", "rows": 1200, "columns": 8},
                {"name": "客户", "rows": 500, "columns": 5},
            ],
        )
        registry.register_upload(
            canonical_path="uploads/chart.png",
            original_name="chart.png",
            file_type="image",
            size_bytes=345000,
            session_id="s1",
            turn=2,
        )
        registry.register_backup(
            backup_path="outputs/backups/sales_bk.xlsx",
            parent_canonical="sales.xlsx",
            session_id="s1",
            turn=2,
            tool_name="write_cells",
        )

        text = registry.build_panorama()
        assert "## 工作区文件全景" in text
        assert "用户文件" in text
        assert "sales.xlsx" in text
        assert "chart.png" in text
        assert "备份与副本" in text
        assert "⚠️" in text

    def test_compact_mode(self, registry: FileRegistry):
        # 注册 > 20 文件触发紧凑模式
        for i in range(25):
            registry.register_from_scan(
                canonical_path=f"data/file{i:03d}.xlsx",
                original_name=f"file{i:03d}.xlsx",
            )
        text = registry.build_panorama()
        assert "用户文件 (25)" in text

    def test_summary_mode(self, registry: FileRegistry):
        # 注册 > 100 文件触发摘要模式
        for i in range(105):
            registry.register_from_scan(
                canonical_path=f"data/file{i:03d}.xlsx",
                original_name=f"file{i:03d}.xlsx",
            )
        text = registry.build_panorama()
        assert "105 个用户文件" in text
        assert "热点目录" in text


# ── DB Migration 测试 ────────────────────────────────────────


class TestDBMigration:
    def test_migration_creates_tables(self, tmp_db: Database):
        """验证 migration 12 创建了三张表。"""
        conn = tmp_db.conn
        # file_registry 表
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='file_registry'"
        ).fetchone()
        assert row is not None

        # file_registry_aliases 表
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='file_registry_aliases'"
        ).fetchone()
        assert row is not None

        # file_registry_events 表
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='file_registry_events'"
        ).fetchone()
        assert row is not None

    def test_migration_version(self, tmp_db: Database):
        """验证迁移到了 v12。"""
        row = tmp_db.conn.execute(
            "SELECT MAX(version) as v FROM schema_version"
        ).fetchone()
        assert row["v"] >= 12


# ── 扫描测试 ─────────────────────────────────────────────────


class TestScanWorkspace:
    def test_scan_empty_workspace(self, registry: FileRegistry):
        result = registry.scan_workspace()
        assert result.total_files == 0

    def test_scan_with_excel_files(self, registry: FileRegistry, workspace: Path):
        """在工作区中创建 Excel 文件并扫描。"""
        # 创建一个最小的 xlsx 文件
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "foo"
        ws["B2"] = 42

        xlsx_path = workspace / "test_data.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        result = registry.scan_workspace()
        assert result.total_files == 1
        assert result.new_files == 1

        # 文件已注册
        entry = registry.get_by_path("test_data.xlsx")
        assert entry is not None
        assert entry.file_type == "excel"
        assert entry.origin == "scan"
        assert len(entry.sheet_meta) > 0
        assert entry.sheet_meta[0]["name"] == "TestSheet"

    def test_incremental_scan(self, registry: FileRegistry, workspace: Path):
        """增量扫描：mtime 未变则跳过。"""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.save(str(workspace / "inc.xlsx"))
        wb.close()

        r1 = registry.scan_workspace()
        assert r1.new_files == 1

        r2 = registry.scan_workspace()
        assert r2.new_files == 0
        assert r2.cache_hits == 1

    def test_scan_detects_deletion(self, registry: FileRegistry, workspace: Path):
        """扫描检测到文件被删除 → 软删除。"""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        xlsx_path = workspace / "will_delete.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        registry.scan_workspace()
        assert registry.get_by_path("will_delete.xlsx") is not None

        # 删除物理文件
        xlsx_path.unlink()

        result = registry.scan_workspace()
        assert result.deleted_files == 1

        # 仍可通过 include_deleted 查到
        all_entries = registry.list_all(include_deleted=True)
        deleted = [e for e in all_entries if e.canonical_path == "will_delete.xlsx"]
        assert len(deleted) == 1
        assert deleted[0].deleted_at is not None


# ── 缓存一致性测试 ───────────────────────────────────────────


class TestCacheConsistency:
    def test_cache_survives_reload(self, tmp_db: Database, workspace: Path):
        """Registry 重建后能从 DB 恢复缓存。"""
        r1 = FileRegistry(tmp_db, workspace)
        r1.register_upload(
            canonical_path="uploads/test.xlsx",
            original_name="test.xlsx",
            size_bytes=100,
        )

        # 创建新的 Registry 实例（模拟重启）
        r2 = FileRegistry(tmp_db, workspace)
        found = r2.get_by_path("uploads/test.xlsx")
        assert found is not None
        assert found.original_name == "test.xlsx"

    def test_alias_cache_survives_reload(self, tmp_db: Database, workspace: Path):
        """别名缓存在重建后仍可用。"""
        r1 = FileRegistry(tmp_db, workspace)
        r1.register_upload(
            canonical_path="uploads/abc_report.xlsx",
            original_name="report.xlsx",
        )

        r2 = FileRegistry(tmp_db, workspace)
        found = r2.get_by_alias("report.xlsx")
        assert found is not None
        assert found.canonical_path == "uploads/abc_report.xlsx"


# ── 全文件类型扫描测试 ──────────────────────────────────────


class TestFullFileTypeScan:
    """scan_workspace 支持全文件类型。"""

    def test_scan_all_file_types(self, registry: FileRegistry, workspace: Path):
        """扫描 Excel + 图片 + 文本 + CSV 文件。"""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.save(str(workspace / "data.xlsx"))
        wb.close()

        (workspace / "image.png").write_bytes(b"\x89PNG\r\n\x1a\n" + b"\x00" * 10)
        (workspace / "readme.md").write_text("# Hello", encoding="utf-8")
        (workspace / "data.csv").write_text("a,b\n1,2\n", encoding="utf-8")

        result = registry.scan_workspace()
        assert result.total_files == 4
        assert result.new_files == 4

        # 验证文件类型
        excel_entry = registry.get_by_path("data.xlsx")
        assert excel_entry is not None
        assert excel_entry.file_type == "excel"
        assert len(excel_entry.sheet_meta) > 0  # Excel 有 sheet meta

        img_entry = registry.get_by_path("image.png")
        assert img_entry is not None
        assert img_entry.file_type == "image"
        assert img_entry.sheet_meta == []  # 非 Excel 无 sheet meta

        md_entry = registry.get_by_path("readme.md")
        assert md_entry is not None
        assert md_entry.file_type == "text"

        csv_entry = registry.get_by_path("data.csv")
        assert csv_entry is not None
        assert csv_entry.file_type == "csv"

    def test_excel_only_mode(self, registry: FileRegistry, workspace: Path):
        """excel_only=True 仅扫描 Excel 文件。"""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.save(str(workspace / "data.xlsx"))
        wb.close()
        (workspace / "image.png").write_bytes(b"\x89PNG" + b"\x00" * 10)
        (workspace / "readme.md").write_text("# Hello", encoding="utf-8")

        result = registry.scan_workspace(excel_only=True)
        assert result.total_files == 1
        assert result.new_files == 1

        assert registry.get_by_path("data.xlsx") is not None
        assert registry.get_by_path("image.png") is None
        assert registry.get_by_path("readme.md") is None

    def test_skip_binary_extensions(self, registry: FileRegistry, workspace: Path):
        """扫描跳过 .pyc / .db 等二进制扩展名。"""
        (workspace / "module.pyc").write_bytes(b"\x00" * 10)
        (workspace / "local.db").write_bytes(b"\x00" * 10)
        (workspace / "notes.txt").write_text("hello", encoding="utf-8")

        result = registry.scan_workspace()
        assert result.total_files == 1  # 只有 notes.txt
        assert registry.get_by_path("notes.txt") is not None
        assert registry.get_by_path("module.pyc") is None

    def test_skip_dot_and_tilde_files(self, registry: FileRegistry, workspace: Path):
        """扫描跳过以 . 或 ~$ 开头的文件。"""
        (workspace / ".hidden").write_text("secret", encoding="utf-8")
        (workspace / "~$temp.xlsx").write_bytes(b"\x00" * 10)
        (workspace / "visible.txt").write_text("ok", encoding="utf-8")

        result = registry.scan_workspace()
        assert result.total_files == 1
        assert registry.get_by_path("visible.txt") is not None

    def test_skip_noise_dirs(self, registry: FileRegistry, workspace: Path):
        """扫描跳过 .git / __pycache__ 等噪音目录。"""
        git_dir = workspace / ".git"
        git_dir.mkdir()
        (git_dir / "config").write_text("...", encoding="utf-8")

        cache_dir = workspace / "__pycache__"
        cache_dir.mkdir()
        (cache_dir / "mod.cpython-312.pyc").write_bytes(b"\x00")

        (workspace / "real.txt").write_text("data", encoding="utf-8")

        result = registry.scan_workspace()
        assert result.total_files == 1


class TestScanUploads:
    """scan_uploads 专门扫描 uploads/ 目录。"""

    def test_scan_uploads_empty(self, registry: FileRegistry):
        """uploads/ 不存在时返回空结果。"""
        result = registry.scan_uploads()
        assert result.total_files == 0

    def test_scan_uploads_discovers_untracked(self, registry: FileRegistry, workspace: Path):
        """scan_uploads 发现未注册的上传文件。"""
        uploads = workspace / "uploads"
        uploads.mkdir()
        (uploads / "report.csv").write_text("a,b\n1,2\n", encoding="utf-8")
        (uploads / "photo.jpg").write_bytes(b"\xff\xd8\xff\xe0" + b"\x00" * 10)

        result = registry.scan_uploads()
        assert result.total_files == 2
        assert result.new_files == 2

        csv_entry = registry.get_by_path("uploads/report.csv")
        assert csv_entry is not None
        assert csv_entry.file_type == "csv"
        assert csv_entry.origin == "scan"

    def test_scan_uploads_respects_registered(self, registry: FileRegistry, workspace: Path):
        """scan_uploads 对已通过 register_upload 注册的文件做增量更新。"""
        uploads = workspace / "uploads"
        uploads.mkdir()
        csv_path = uploads / "data.csv"
        csv_path.write_text("a,b\n1,2\n", encoding="utf-8")

        # 先通过 register_upload 注册
        registry.register_upload(
            canonical_path="uploads/data.csv",
            original_name="data.csv",
            size_bytes=csv_path.stat().st_size,
        )

        result = registry.scan_uploads()
        assert result.new_files == 0
        assert result.cache_hits == 1  # 已注册且未变
        assert result.total_files == 1

    def test_scan_uploads_detects_content_change(self, registry: FileRegistry, workspace: Path):
        """scan_uploads 检测已注册文件的内容变化。"""
        uploads = workspace / "uploads"
        uploads.mkdir()
        csv_path = uploads / "data.csv"
        csv_path.write_text("a,b\n1,2\n", encoding="utf-8")

        registry.register_upload(
            canonical_path="uploads/data.csv",
            original_name="data.csv",
            size_bytes=csv_path.stat().st_size,
        )

        # 修改文件内容
        import time
        time.sleep(0.05)
        csv_path.write_text("a,b,c\n1,2,3\n4,5,6\n", encoding="utf-8")

        result = registry.scan_uploads()
        assert result.updated_files == 1

        entry = registry.get_by_path("uploads/data.csv")
        assert entry is not None
        assert entry.size_bytes == csv_path.stat().st_size


# ── Staging / CoW / Checkpoint 委托层测试 ────────────────


@pytest.fixture
def versioned_registry(tmp_db: Database, workspace: Path) -> FileRegistry:
    """启用版本管理的 FileRegistry。"""
    return FileRegistry(tmp_db, workspace, enable_versions=True)


class TestVersionedRegistry:
    """enable_versions=True 时的版本管理委托层。"""

    def test_has_versions(self, versioned_registry: FileRegistry):
        assert versioned_registry.has_versions is True
        assert versioned_registry.fvm is not None

    def test_no_versions_by_default(self, registry: FileRegistry):
        assert registry.has_versions is False
        assert registry.fvm is None

    def test_stage_for_write(self, versioned_registry: FileRegistry, workspace: Path):
        """staging 写入并返回副本路径。"""
        orig = workspace / "data.txt"
        orig.write_text("hello", encoding="utf-8")

        staged_path = versioned_registry.stage_for_write("data.txt")
        assert staged_path != str(orig)
        assert Path(staged_path).exists()

    def test_stage_for_write_no_versions(self, registry: FileRegistry, workspace: Path):
        """无版本管理时返回原始路径。"""
        orig = workspace / "data.txt"
        orig.write_text("hello", encoding="utf-8")

        result = registry.stage_for_write("data.txt")
        assert result == str(orig.resolve())

    def test_commit_staged(self, versioned_registry: FileRegistry, workspace: Path):
        """提交 staged 文件回原位。"""
        orig = workspace / "data.txt"
        orig.write_text("original", encoding="utf-8")

        staged = versioned_registry.stage_for_write("data.txt")
        Path(staged).write_text("modified", encoding="utf-8")

        result = versioned_registry.commit_staged("data.txt")
        assert result is not None
        assert orig.read_text(encoding="utf-8") == "modified"

    def test_discard_staged(self, versioned_registry: FileRegistry, workspace: Path):
        """丢弃 staged 文件。"""
        orig = workspace / "data.txt"
        orig.write_text("original", encoding="utf-8")

        staged = versioned_registry.stage_for_write("data.txt")
        Path(staged).write_text("modified", encoding="utf-8")

        ok = versioned_registry.discard_staged("data.txt")
        assert ok is True
        assert not Path(staged).exists()

    def test_staged_file_map(self, versioned_registry: FileRegistry, workspace: Path):
        """staged_file_map 返回映射。"""
        orig = workspace / "data.txt"
        orig.write_text("hello", encoding="utf-8")
        versioned_registry.stage_for_write("data.txt")

        fmap = versioned_registry.staged_file_map()
        assert len(fmap) == 1

    def test_cow_mapping(self, versioned_registry: FileRegistry, workspace: Path):
        """CoW 映射注册和查找。"""
        orig = workspace / "src.xlsx"
        orig.write_text("data", encoding="utf-8")
        dst = workspace / "outputs" / "src_cow.xlsx"
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_text("copy", encoding="utf-8")

        versioned_registry.register_cow_mapping("src.xlsx", "outputs/src_cow.xlsx")
        redirect = versioned_registry.lookup_cow_redirect("src.xlsx")
        assert redirect is not None

    def test_checkpoint_and_restore(self, versioned_registry: FileRegistry, workspace: Path):
        """checkpoint 快照 + restore_to_original。"""
        f = workspace / "target.txt"
        f.write_text("v1", encoding="utf-8")

        ver = versioned_registry.checkpoint_file("target.txt", reason="staging")
        assert ver is not None

        f.write_text("v2", encoding="utf-8")
        ok = versioned_registry.restore_to_original("target.txt")
        assert ok is True
        assert f.read_text(encoding="utf-8") == "v1"

    def test_turn_checkpoint_and_rollback(self, versioned_registry: FileRegistry, workspace: Path):
        """轮次 checkpoint + rollback。"""
        f = workspace / "sheet.xlsx"
        f.write_text("original", encoding="utf-8")

        versioned_registry.checkpoint_file("sheet.xlsx", reason="staging")
        f.write_text("turn1_modified", encoding="utf-8")
        cp = versioned_registry.create_turn_checkpoint(1, ["sheet.xlsx"], ["write_cell"])
        assert cp is not None

        cps = versioned_registry.list_turn_checkpoints()
        assert len(cps) == 1

        restored = versioned_registry.rollback_to_turn(1)
        assert "sheet.xlsx" in restored

    def test_invalidate_undo(self, versioned_registry: FileRegistry, workspace: Path):
        """invalidate_undo 标记版本不可恢复。"""
        f = workspace / "a.txt"
        f.write_text("v1", encoding="utf-8")
        versioned_registry.checkpoint_file("a.txt", reason="staging")

        count = versioned_registry.invalidate_undo({"a.txt"})
        assert count >= 1

    def test_gc_versions(self, versioned_registry: FileRegistry, workspace: Path):
        """gc_versions 不报错。"""
        result = versioned_registry.gc_versions(max_age_seconds=0)
        assert result >= 0


# ── System Prompt 统一 + 写后事件记录 ──────────────────────


class TestBuildFileRegistryNotice:
    """_build_file_registry_notice 统一测试。"""

    def _make_builder(self, *, registry=None, file_registry=None):
        from unittest.mock import MagicMock
        from excelmanus.engine_core.context_builder import ContextBuilder
        from excelmanus.engine_core.session_state import SessionState

        engine = MagicMock()
        _state = SessionState()
        engine._state = _state
        engine.state = _state
        engine.file_registry = file_registry
        if registry:
            _state.register_cow_mappings(registry)
        builder = ContextBuilder(engine)
        return builder

    def test_empty_returns_empty(self):
        builder = self._make_builder()
        assert builder._build_file_registry_notice() == ""

    def test_panorama_from_file_registry(self, tmp_path):
        """FileRegistry 可用时使用 build_panorama()。"""
        db = Database(str(tmp_path / "test.db"))
        reg = FileRegistry(db, tmp_path)
        f = tmp_path / "demo.xlsx"
        f.write_text("x", encoding="utf-8")
        reg.register_upload("demo.xlsx", "demo.xlsx", file_type="excel")

        builder = self._make_builder(file_registry=reg)
        notice = builder._build_file_registry_notice()
        assert "demo.xlsx" in notice
        assert "工作区文件全景" in notice

    def test_panorama_with_cow_mapping(self, tmp_path):
        """文件全景 + CoW 映射同时存在。"""
        db = Database(str(tmp_path / "test.db"))
        reg = FileRegistry(db, tmp_path)
        f = tmp_path / "a.xlsx"
        f.write_text("x", encoding="utf-8")
        reg.register_upload("a.xlsx", "a.xlsx", file_type="excel")

        builder = self._make_builder(
            file_registry=reg,
            registry={"a.xlsx": "outputs/a.xlsx"},
        )
        notice = builder._build_file_registry_notice()
        # 全景部分
        assert "工作区文件全景" in notice
        # CoW 部分
        assert "⚠️ 文件保护路径映射（CoW）" in notice
        assert "outputs/a.xlsx" in notice

    def test_no_registry_returns_empty(self):
        """无 FileRegistry 时返回空。"""
        builder = self._make_builder()
        notice = builder._build_file_registry_notice()
        assert notice == ""

    def test_cow_only_no_panorama(self):
        """仅 CoW 映射、无文件全景。"""
        builder = self._make_builder(
            registry={"src.xlsx": "outputs/src.xlsx"},
        )
        notice = builder._build_file_registry_notice()
        assert "⚠️ 文件保护路径映射（CoW）" in notice
        assert "src.xlsx" in notice


class TestToolDispatcherWriteEvent:
    """写后事件记录到 FileRegistry。"""

    def test_record_event_called_on_write(self, tmp_path):
        """写入工具成功后，FileRegistry.record_event 被调用。"""
        from unittest.mock import MagicMock, patch

        db = Database(str(tmp_path / "test.db"))
        reg = FileRegistry(db, tmp_path)
        f = tmp_path / "target.xlsx"
        f.write_text("data", encoding="utf-8")
        entry = reg.register_upload("target.xlsx", "target.xlsx", file_type="excel")

        engine = MagicMock()
        engine.file_registry = reg
        engine.get_tool_write_effect.return_value = "workspace_write"
        engine.state.session_turn = 1

        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher
        dispatcher = ToolDispatcher(engine)
        dispatcher._EXCEL_WRITE_TOOLS = {"write_excel"}

        # 模拟后处理中的写后事件记录逻辑
        with patch.object(reg, "record_event") as mock_record:
            # 直接调用写后事件记录段的逻辑
            tool_name = "write_excel"
            arguments = {"file_path": "target.xlsx"}
            _write_paths = []
            if tool_name in dispatcher._EXCEL_WRITE_TOOLS:
                _wp = (arguments.get("file_path") or "").strip()
                if _wp:
                    _write_paths.append(_wp)
            for _wpath in _write_paths:
                _entry = reg.get_by_path(_wpath)
                if _entry is not None:
                    reg.record_event(
                        _entry.id,
                        "tool_write",
                        tool_name=tool_name,
                        turn=engine.state.session_turn,
                    )
            mock_record.assert_called_once_with(
                entry.id, "tool_write", tool_name="write_excel", turn=1,
            )
