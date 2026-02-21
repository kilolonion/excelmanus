"""Image-to-Excel 端到端集成测试。"""

from __future__ import annotations

import base64
import json
from pathlib import Path

import pytest


# 最小有效 PNG（1x1 白色像素）
_MINIMAL_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR4"
    "nGP4z8BQDwAEgAF/pooBPQAAAABJRU5ErkJggg=="
)

_FULL_SPEC = {
    "version": "1.0",
    "provenance": {
        "source_image_hash": "sha256:abc123",
        "model": "test-model",
        "timestamp": "2026-02-21T11:30:00Z",
    },
    "workbook": {
        "name": "replica",
        "locale": "zh-CN",
        "default_font": {"name": "等线", "size": 11},
    },
    "sheets": [{
        "name": "Sheet1",
        "dimensions": {"rows": 3, "cols": 3},
        "freeze_panes": "A2",
        "cells": [
            {"address": "A1", "value": "产品", "value_type": "string", "style_id": "header", "confidence": 0.98},
            {"address": "B1", "value": "数量", "value_type": "string", "style_id": "header", "confidence": 0.98},
            {"address": "C1", "value": "金额", "value_type": "string", "style_id": "header", "confidence": 0.98},
            {"address": "A2", "value": "苹果", "value_type": "string", "confidence": 1.0},
            {"address": "B2", "value": 100, "value_type": "number", "confidence": 1.0},
            {"address": "C2", "value": 500.5, "value_type": "number", "confidence": 0.95},
            {"address": "A3", "value": "香蕉", "value_type": "string", "confidence": 1.0},
            {"address": "B3", "value": 200, "value_type": "number", "confidence": 1.0},
            {"address": "C3", "value": 300, "value_type": "number", "confidence": 0.9},
        ],
        "merged_ranges": [],
        "styles": {
            "header": {
                "font": {"bold": True, "size": 12, "color": "#FFFFFF", "name": "微软雅黑"},
                "fill": {"type": "solid", "color": "#4472C4"},
                "alignment": {"horizontal": "center"},
            },
        },
        "column_widths": [15, 10, 12],
        "row_heights": {"1": 28},
    }],
    "uncertainties": [
        {
            "location": "C2",
            "reason": "小数部分模糊",
            "candidate_values": ["500.5", "500.8"],
            "confidence": 0.65,
        },
    ],
}


class TestImageToExcelPipeline:
    """端到端集成测试：spec → rebuild → verify 完整闭环。"""

    def test_spec_roundtrip(self, tmp_path: Path) -> None:
        """ReplicaSpec → rebuild → verify 完整闭环，match_rate 应为 1.0。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, verify_excel_replica, init_guard

        init_guard(str(tmp_path))
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(_FULL_SPEC), encoding="utf-8")
        excel_path = tmp_path / "draft.xlsx"
        report_path = tmp_path / "diff_report.md"

        # rebuild
        rebuild_result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path), output_path=str(excel_path),
        ))
        assert rebuild_result["status"] == "ok"
        assert excel_path.exists()
        assert rebuild_result["build_summary"]["cells_written"] == 9
        assert rebuild_result["build_summary"]["styles_applied"] == 3

        # verify
        verify_result = json.loads(verify_excel_replica(
            spec_path=str(spec_path), excel_path=str(excel_path), report_path=str(report_path),
        ))
        assert verify_result["status"] == "ok"
        assert verify_result["match_rate"] == 1.0
        assert verify_result["issues"]["low_confidence"] == 1  # 来自 uncertainties
        assert report_path.exists()

        # 报告内容检查
        report_content = report_path.read_text(encoding="utf-8")
        assert "100.0%" in report_content
        assert "低置信项" in report_content

    def test_read_image_returns_injection(self, tmp_path: Path) -> None:
        """read_image → _image_injection 结构正确。"""
        from excelmanus.tools.image_tools import read_image, init_guard

        png_data = base64.b64decode(_MINIMAL_PNG_B64)
        img_path = tmp_path / "test.png"
        img_path.write_bytes(png_data)
        init_guard(str(tmp_path))

        result = json.loads(read_image(file_path=str(img_path)))
        assert result["status"] == "ok"
        injection = result["_image_injection"]
        assert injection["mime_type"] == "image/png"
        # base64 应能解码回原始数据
        decoded = base64.b64decode(injection["base64"])
        assert decoded == png_data

    def test_rebuild_with_merged_cells_and_styles(self, tmp_path: Path) -> None:
        """复杂 spec（合并+样式+freeze）编译正确。"""
        from excelmanus.tools.image_tools import rebuild_excel_from_spec, init_guard
        from openpyxl import load_workbook

        init_guard(str(tmp_path))
        spec = dict(_FULL_SPEC)
        spec["sheets"] = [{
            **_FULL_SPEC["sheets"][0],
            "merged_ranges": [{"range": "A1:C1", "confidence": 0.95}],
        }]
        spec_path = tmp_path / "spec.json"
        spec_path.write_text(json.dumps(spec), encoding="utf-8")
        excel_path = tmp_path / "complex.xlsx"

        result = json.loads(rebuild_excel_from_spec(
            spec_path=str(spec_path), output_path=str(excel_path),
        ))
        assert result["status"] == "ok"
        assert result["build_summary"]["merges_applied"] == 1

        # 验证 openpyxl 结构
        wb = load_workbook(str(excel_path))
        ws = wb["Sheet1"]
        assert ws.freeze_panes == "A2"
        assert len(ws.merged_cells.ranges) == 1
        assert ws["A1"].font.bold is True

    def test_replica_spec_pydantic_roundtrip(self) -> None:
        """ReplicaSpec JSON 序列化/反序列化一致性。"""
        from excelmanus.replica_spec import ReplicaSpec

        spec = ReplicaSpec.model_validate(_FULL_SPEC)
        json_str = spec.model_dump_json()
        spec2 = ReplicaSpec.model_validate_json(json_str)
        assert spec.version == spec2.version
        assert len(spec.sheets) == len(spec2.sheets)
        assert len(spec.sheets[0].cells) == len(spec2.sheets[0].cells)
        assert spec.uncertainties[0].location == spec2.uncertainties[0].location

    def test_multimodal_memory_integration(self) -> None:
        """Memory 层多模态消息与 TokenCounter 集成。"""
        from excelmanus.config import ExcelManusConfig
        from excelmanus.memory import ConversationMemory, TokenCounter, IMAGE_TOKEN_ESTIMATE

        config = ExcelManusConfig(
            api_key="test", base_url="https://test.example.com/v1", model="test",
        )
        mem = ConversationMemory(config)

        # 添加图片消息
        mem.add_image_message(base64_data="abc123", mime_type="image/png")
        msgs = mem.get_messages()
        last = msgs[-1]
        assert last["role"] == "user"
        assert isinstance(last["content"], list)

        # token 计数包含图片估算
        count = TokenCounter.count_message(last)
        assert count >= IMAGE_TOKEN_ESTIMATE

    def test_provider_image_conversion(self) -> None:
        """Gemini 和 Claude provider 正确转换图片 content part。"""
        from excelmanus.providers.gemini import _openai_messages_to_gemini
        from excelmanus.providers.claude import _openai_messages_to_claude

        messages = [
            {"role": "user", "content": [
                {"type": "text", "text": "分析图片"},
                {"type": "image_url", "image_url": {
                    "url": "data:image/png;base64,abc123",
                }},
            ]},
        ]

        # Gemini
        _, contents = _openai_messages_to_gemini(messages)
        parts = contents[0]["parts"]
        assert any("inlineData" in p for p in parts)

        # Claude
        _, claude_msgs = _openai_messages_to_claude(messages)
        content = claude_msgs[0]["content"]
        assert any(b.get("type") == "image" for b in content)
