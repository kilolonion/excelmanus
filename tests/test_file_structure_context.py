"""路由层文件结构预注入测试。

验证 _build_file_structure_context 和 _guess_header_row
能正确提取 Excel 文件结构并给出 header_row 建议。
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from openpyxl import Workbook

from excelmanus.config import ExcelManusConfig
from excelmanus.skillpacks.router import SkillRouter


@pytest.fixture()
def router() -> SkillRouter:
    """创建一个最小化的 SkillRouter 实例。"""
    config = ExcelManusConfig(api_key="test", base_url="http://test", model="test")
    loader = MagicMock()
    loader.get_skillpacks.return_value = {}
    loader.load_all.return_value = {}
    return SkillRouter(config, loader)


@pytest.fixture()
def excel_with_title_row(tmp_path: Path) -> Path:
    """第 0 行是标题行，第 1 行是真正 header。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册"
    ws.append(["员工花名册（机密）", None, None, None, None])
    ws.append(["工号", "姓名", "性别", "部门", "职级"])
    ws.append(["EMP001", "张三", "男", "技术部", "P7"])
    fp = tmp_path / "employees.xlsx"
    wb.save(fp)
    return fp


@pytest.fixture()
def excel_standard_header(tmp_path: Path) -> Path:
    """标准格式，第 0 行即 header。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售数据"
    ws.append(["产品", "数量", "单价", "总计"])
    ws.append(["苹果", 10, 5.0, 50.0])
    fp = tmp_path / "sales.xlsx"
    wb.save(fp)
    return fp


class TestBuildFileStructureContext:
    """_build_file_structure_context 测试。"""

    def test_no_paths_returns_empty(self, router: SkillRouter) -> None:
        text, sc, mr = router._build_file_structure_context_sync(candidate_file_paths=None)
        assert text == ""
        text, sc, mr = router._build_file_structure_context_sync(candidate_file_paths=[])
        assert text == ""

    def test_title_row_file_suggests_header_1(
        self, router: SkillRouter, excel_with_title_row: Path
    ) -> None:
        result, sc, mr = router._build_file_structure_context_sync(
            candidate_file_paths=[str(excel_with_title_row)]
        )
        assert "[文件结构预览]" in result
        assert "员工花名册" in result
        assert "header_row=1" in result

    def test_standard_header_suggests_header_0(
        self, router: SkillRouter, excel_standard_header: Path
    ) -> None:
        result, sc, mr = router._build_file_structure_context_sync(
            candidate_file_paths=[str(excel_standard_header)]
        )
        assert "[文件结构预览]" in result
        assert "销售数据" in result
        # header_row=0 是默认值，不需要显式建议
        assert "header_row=0" in result

    def test_nonexistent_file_skipped(self, router: SkillRouter, tmp_path: Path) -> None:
        result, sc, mr = router._build_file_structure_context_sync(
            candidate_file_paths=[str(tmp_path / "nonexistent.xlsx")]
        )
        assert result == ""

    def test_non_excel_file_skipped(self, router: SkillRouter, tmp_path: Path) -> None:
        txt_file = tmp_path / "data.txt"
        txt_file.write_text("hello")
        result, sc, mr = router._build_file_structure_context_sync(
            candidate_file_paths=[str(txt_file)]
        )
        assert result == ""

    def test_max_files_limit(
        self, router: SkillRouter, tmp_path: Path
    ) -> None:
        """超过 max_files 限制时只处理前 N 个文件。"""
        paths = []
        for i in range(5):
            wb = Workbook()
            ws = wb.active
            ws.append(["A", "B", "C"])
            ws.append([1, 2, 3])
            fp = tmp_path / f"file_{i}.xlsx"
            wb.save(fp)
            paths.append(str(fp))

        result, sc, mr = router._build_file_structure_context_sync(
            candidate_file_paths=paths, max_files=2
        )
        # 应该只包含 2 个文件的信息
        assert result.count("文件:") == 2

    def test_reuse_file_structure_cache_avoids_reopen(
        self, router: SkillRouter, excel_standard_header: Path
    ) -> None:
        """同一文件未变化时应命中缓存，避免重复打开工作簿。"""
        with patch("openpyxl.load_workbook", wraps=openpyxl.load_workbook) as mocked:
            router._build_file_structure_context_sync(candidate_file_paths=[str(excel_standard_header)])
            router._build_file_structure_context_sync(candidate_file_paths=[str(excel_standard_header)])

        assert mocked.call_count == 1

    def test_cache_invalidated_when_file_changes(
        self, router: SkillRouter, excel_standard_header: Path
    ) -> None:
        """文件 mtime/size 变化后应重新读取，不复用旧缓存。"""
        with patch("openpyxl.load_workbook", wraps=openpyxl.load_workbook) as mocked:
            router._build_file_structure_context_sync(candidate_file_paths=[str(excel_standard_header)])

            wb = Workbook()
            ws = wb.active
            ws.title = "销售数据"
            ws.append(["产品", "数量", "单价", "总计"])
            ws.append(["苹果", 10, 5.0, 50.0])
            ws.append(["超长文本" * 20, 99, 1.0, 99.0])
            wb.save(excel_standard_header)

            router._build_file_structure_context_sync(candidate_file_paths=[str(excel_standard_header)])

        assert mocked.call_count == 2


class TestGuessHeaderRow:
    """_guess_header_row 静态方法测试。"""

    def test_standard_header(self) -> None:
        rows = [["A", "B", "C"], [1, 2, 3]]
        assert SkillRouter._guess_header_row(rows) == 0

    def test_title_then_header(self) -> None:
        rows = [["报表标题", None, None], ["A", "B", "C"], [1, 2, 3]]
        assert SkillRouter._guess_header_row(rows) == 1

    def test_two_titles_then_header(self) -> None:
        rows = [["大标题", None], ["副标题", None], ["A", "B", "C"]]
        assert SkillRouter._guess_header_row(rows) == 2

    def test_empty_rows(self) -> None:
        assert SkillRouter._guess_header_row([]) is None

    def test_all_sparse_rows(self) -> None:
        rows = [["A", None], ["B", None], ["C", None]]
        assert SkillRouter._guess_header_row(rows) is None

    def test_deep_header_row(self) -> None:
        rows = [
            ["2025年度 KPI 仪表盘", None, None],
            [None, None, None],
            ["总营收", None, "净利润", None, "客户数"],
            ["98,000,000元", None, "12,250,000元", None, "15,832"],
            ["↑ 19.5%", None, "↑ 22.3%", None, "↑ 8.7%"],
            [None, None, None],
            ["月度关键指标趋势", None, None],
            ["月份", "营收", "成本", "利润", "订单量"],
            ["2025年01月", 1000, 800, 200, 100],
        ]
        assert SkillRouter._guess_header_row(rows) == 7
