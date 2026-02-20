"""sandbox_hook bench 保护目录写拦截 + openpyxl save 原子写入 测试。"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import pytest

from excelmanus.security.sandbox_hook import generate_wrapper_script


@pytest.fixture()
def workspace(tmp_path: Path) -> Path:
    return tmp_path


def _run_in_sandbox(
    workspace: Path,
    script_content: str,
    tier: str,
    *,
    env_override: dict[str, str] | None = None,
) -> subprocess.CompletedProcess:
    """在沙盒 wrapper 中执行脚本。"""
    script = workspace / "test_script.py"
    script.write_text(script_content, encoding="utf-8")
    wrapper = generate_wrapper_script(tier, str(workspace))
    wrapper_path = workspace / "_wrapper.py"
    wrapper_path.write_text(wrapper, encoding="utf-8")
    env = os.environ.copy()
    if env_override:
        env.update(env_override)
    return subprocess.run(
        [sys.executable, str(wrapper_path), str(script)],
        capture_output=True, text=True, timeout=15, env=env,
    )


class TestBenchProtectedDirWrite:
    """bench 保护目录写拦截。"""

    def test_write_to_bench_external_blocked(self, workspace: Path) -> None:
        """写入 bench/external/ 下的文件应被拦截。"""
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True)
        target = bench_dir / "source.xlsx"
        target.write_text("original", encoding="utf-8")
        code = f"with open(r'{target}', 'w') as f:\n    f.write('overwritten')"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode != 0
        assert "bench 保护目录" in result.stderr
        assert target.read_text(encoding="utf-8") == "original"

    def test_read_from_bench_external_allowed(self, workspace: Path) -> None:
        """读取 bench/external/ 下的文件不受影响。"""
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True)
        target = bench_dir / "source.xlsx"
        target.write_text("data123", encoding="utf-8")
        code = f"with open(r'{target}', 'r') as f:\n    print(f.read())"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert "data123" in result.stdout

    def test_write_to_non_protected_dir_allowed(self, workspace: Path) -> None:
        """写入非保护目录（工作区内）正常。"""
        output_dir = workspace / "outputs"
        output_dir.mkdir()
        target = output_dir / "result.txt"
        code = f"with open(r'{target}', 'w') as f:\n    f.write('ok')\nprint('done')"
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0
        assert target.read_text(encoding="utf-8") == "ok"

    def test_env_var_override_protected_dirs(self, workspace: Path) -> None:
        """环境变量 EXCELMANUS_BENCH_PROTECTED_DIRS 覆盖保护目录。"""
        custom_dir = workspace / "my_data"
        custom_dir.mkdir()
        target = custom_dir / "file.txt"
        target.write_text("original", encoding="utf-8")
        code = f"with open(r'{target}', 'w') as f:\n    f.write('overwritten')"
        result = _run_in_sandbox(
            workspace, code, "GREEN",
            env_override={"EXCELMANUS_BENCH_PROTECTED_DIRS": "my_data"},
        )
        assert result.returncode != 0
        assert "bench 保护目录" in result.stderr
        assert target.read_text(encoding="utf-8") == "original"

    def test_bench_protection_yellow_mode(self, workspace: Path) -> None:
        """YELLOW 模式同样拦截 bench 保护目录写入。"""
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True)
        target = bench_dir / "source.xlsx"
        target.write_text("original", encoding="utf-8")
        code = f"with open(r'{target}', 'w') as f:\n    f.write('overwritten')"
        result = _run_in_sandbox(workspace, code, "YELLOW")
        assert result.returncode != 0
        assert "bench 保护目录" in result.stderr

    def test_bench_protection_red_mode_no_restriction(self, workspace: Path) -> None:
        """RED 模式不受 bench 保护限制。"""
        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True)
        target = bench_dir / "source.txt"
        target.write_text("original", encoding="utf-8")
        code = f"with open(r'{target}', 'w') as f:\n    f.write('overwritten')\nprint('done')"
        result = _run_in_sandbox(workspace, code, "RED")
        assert result.returncode == 0
        assert target.read_text(encoding="utf-8") == "overwritten"


class TestOpenpyxlAtomicSave:
    """openpyxl save 原子写入保护 — 直接测试核心逻辑。

    由于 openpyxl 内部通过 et_xmlfile 在系统临时目录创建文件，
    无法在沙盒子进程中完整运行（_guarded_open 会拦截工作区外写入），
    因此提取 _patch_openpyxl_save 的核心逻辑在进程内直接测试。
    """

    def test_atomic_save_success(self, workspace: Path) -> None:
        """正常 save 成功：先写临时文件再原子替换。"""
        from openpyxl import Workbook
        from openpyxl.workbook import Workbook as _Wb

        target = workspace / "output.xlsx"
        # 创建初始文件
        wb = Workbook()
        wb.active["A1"] = "initial"
        wb.save(str(target))
        wb.close()
        original_size = target.stat().st_size

        # 手动应用 atomic save patch
        _original_save = _Wb.save

        def _atomic_save(self, filename):
            import tempfile as _tf
            resolved = os.path.realpath(str(filename))
            if not os.path.exists(resolved):
                return _original_save(self, filename)
            dir_name = os.path.dirname(resolved)
            fd, tmp_path = _tf.mkstemp(suffix=".xlsx", dir=dir_name)
            os.close(fd)
            try:
                _original_save(self, tmp_path)
                os.replace(tmp_path, resolved)
            except BaseException:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
                raise

        _Wb.save = _atomic_save
        try:
            wb2 = Workbook()
            wb2.active["A1"] = "updated"
            wb2.save(str(target))
            wb2.close()
        finally:
            _Wb.save = _original_save

        # 验证文件已更新
        from openpyxl import load_workbook
        wb3 = load_workbook(target)
        assert wb3.active["A1"].value == "updated"
        wb3.close()

    def test_atomic_save_failure_preserves_original(self, workspace: Path) -> None:
        """原子 save 失败时原文件不损坏。"""
        from openpyxl import Workbook, load_workbook
        from openpyxl.workbook import Workbook as _Wb

        target = workspace / "output.xlsx"
        # 创建初始文件
        wb = Workbook()
        wb.active["A1"] = "precious_data"
        wb.save(str(target))
        wb.close()
        original_content = target.read_bytes()

        # 构造一个会失败的 _original_save
        _real_save = _Wb.save

        def _failing_save(self, filename):
            raise IOError("模拟写入失败")

        def _atomic_save(self, filename):
            import tempfile as _tf
            resolved = os.path.realpath(str(filename))
            if not os.path.exists(resolved):
                return _failing_save(self, filename)
            dir_name = os.path.dirname(resolved)
            fd, tmp_path = _tf.mkstemp(suffix=".xlsx", dir=dir_name)
            os.close(fd)
            try:
                _failing_save(self, tmp_path)
                os.replace(tmp_path, resolved)
            except BaseException:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
                raise

        _Wb.save = _atomic_save
        try:
            wb2 = Workbook()
            wb2.active["A1"] = "bad_data"
            with pytest.raises(IOError, match="模拟写入失败"):
                wb2.save(str(target))
        finally:
            _Wb.save = _real_save

        # 原文件内容不变
        assert target.read_bytes() == original_content
        # 无残留临时文件
        xlsx_files = list(workspace.glob("*.xlsx"))
        assert len(xlsx_files) == 1
        assert xlsx_files[0].name == "output.xlsx"

    def test_atomic_save_new_file_direct_write(self, workspace: Path) -> None:
        """新建文件时直接写入（无需原子替换）。"""
        from openpyxl import Workbook, load_workbook
        from openpyxl.workbook import Workbook as _Wb

        target = workspace / "brand_new.xlsx"
        assert not target.exists()

        _original_save = _Wb.save

        def _atomic_save(self, filename):
            import tempfile as _tf
            resolved = os.path.realpath(str(filename))
            if not os.path.exists(resolved):
                return _original_save(self, filename)
            dir_name = os.path.dirname(resolved)
            fd, tmp_path = _tf.mkstemp(suffix=".xlsx", dir=dir_name)
            os.close(fd)
            try:
                _original_save(self, tmp_path)
                os.replace(tmp_path, resolved)
            except BaseException:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
                raise

        _Wb.save = _atomic_save
        try:
            wb = Workbook()
            wb.active["A1"] = "new_file"
            wb.save(str(target))
            wb.close()
        finally:
            _Wb.save = _original_save

        assert target.exists()
        wb2 = load_workbook(target)
        assert wb2.active["A1"].value == "new_file"
        wb2.close()



class TestOpenpyxlSandboxSave:
    """openpyxl save 在沙盒子进程中的完整端到端测试。"""

    def test_openpyxl_save_new_file_in_sandbox(self, workspace: Path) -> None:
        """openpyxl wb.save() 新建文件在 GREEN 沙盒子进程中应成功。"""
        output = workspace / "outputs"
        output.mkdir()
        target = output / "test.xlsx"
        code = (
            "from openpyxl import Workbook\n"
            f"wb = Workbook()\n"
            f"wb.active['A1'] = 'hello'\n"
            f"wb.save(r'{target}')\n"
            f"print('saved')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert "saved" in result.stdout
        assert target.exists()

    def test_openpyxl_save_existing_file_in_sandbox(self, workspace: Path) -> None:
        """覆盖已有 xlsx 文件（走 _atomic_save 路径）应成功。"""
        from openpyxl import Workbook as _Wb

        output = workspace / "outputs"
        output.mkdir()
        target = output / "existing.xlsx"
        wb = _Wb()
        wb.active["A1"] = "initial"
        wb.save(str(target))
        wb.close()

        code = (
            "from openpyxl import Workbook, load_workbook\n"
            f"wb = load_workbook(r'{target}')\n"
            f"wb.active['A1'] = 'updated'\n"
            f"wb.save(r'{target}')\n"
            f"wb2 = load_workbook(r'{target}')\n"
            f"print(wb2.active['A1'].value)\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert "updated" in result.stdout

    def test_openpyxl_save_bench_protected_blocked(self, workspace: Path) -> None:
        """openpyxl wb.save() 写入 bench/external 应被拦截。"""
        from openpyxl import Workbook as _Wb

        bench_dir = workspace / "bench" / "external"
        bench_dir.mkdir(parents=True)
        target = bench_dir / "source.xlsx"
        wb = _Wb()
        wb.active["A1"] = "original"
        wb.save(str(target))
        wb.close()

        code = (
            "from openpyxl import Workbook, load_workbook\n"
            f"wb = load_workbook(r'{target}')\n"
            f"wb.active['A1'] = 'hacked'\n"
            f"wb.save(r'{target}')\n"
        )
        result = _run_in_sandbox(workspace, code, "GREEN")
        assert result.returncode != 0
        assert "bench 保护目录" in (result.stderr + result.stdout)


class TestWrapperTemplateContent:
    """验证生成的 wrapper 脚本包含预期的保护代码。"""

    def test_wrapper_contains_bench_protected_dirs(self) -> None:
        """GREEN wrapper 包含 bench 保护目录相关代码。"""
        wrapper = generate_wrapper_script("GREEN", "/tmp/ws")
        assert "EXCELMANUS_BENCH_PROTECTED_DIRS" in wrapper
        assert "_BENCH_PROTECTED_DIRS" in wrapper
        assert "bench 保护目录" in wrapper

    def test_wrapper_contains_openpyxl_patch(self) -> None:
        """GREEN wrapper 包含 openpyxl atomic save patch。"""
        wrapper = generate_wrapper_script("GREEN", "/tmp/ws")
        assert "_patch_openpyxl_save" in wrapper
        assert "_atomic_save" in wrapper
        assert "os.replace" in wrapper

    def test_red_wrapper_no_protection(self) -> None:
        """RED wrapper 不包含保护代码。"""
        wrapper = generate_wrapper_script("RED", "/tmp/ws")
        assert "_BENCH_PROTECTED_DIRS" not in wrapper
        assert "_patch_openpyxl_save" not in wrapper
