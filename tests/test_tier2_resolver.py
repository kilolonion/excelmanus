"""Tier2Resolver 测试。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excelmanus.reference_graph.scanner import Tier2Resolver


@pytest.fixture()
def chain_workbook(tmp_path: Path) -> Path:
    """A2=10, B2=A2*2, C2=B2+A2 → 链式依赖。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Base"
    ws["B1"] = "Double"
    ws["C1"] = "Sum"
    ws["A2"] = 10
    ws["B2"] = "=A2*2"
    ws["C2"] = "=B2+A2"
    fp = tmp_path / "chain.xlsx"
    wb.save(fp)
    return fp


class TestTier2Resolver:
    def test_precedents(self, chain_workbook: Path) -> None:
        resolver = Tier2Resolver()
        node = resolver.resolve(str(chain_workbook), "Sheet1", "C2", direction="precedents", depth=1)
        assert node.address == "C2"
        assert node.formula == "=B2+A2"
        prec_addrs = {r.cell_or_range for r in node.precedents}
        assert "B2" in prec_addrs
        assert "A2" in prec_addrs

    def test_dependents(self, chain_workbook: Path) -> None:
        resolver = Tier2Resolver()
        node = resolver.resolve(str(chain_workbook), "Sheet1", "A2", direction="dependents", depth=1)
        dep_addrs = {r.cell_or_range for r in node.dependents}
        assert "B2" in dep_addrs
        assert "C2" in dep_addrs

    def test_both(self, chain_workbook: Path) -> None:
        resolver = Tier2Resolver()
        node = resolver.resolve(str(chain_workbook), "Sheet1", "B2", direction="both", depth=1)
        assert len(node.precedents) >= 1
        assert len(node.dependents) >= 1

    def test_non_formula_cell(self, chain_workbook: Path) -> None:
        resolver = Tier2Resolver()
        node = resolver.resolve(str(chain_workbook), "Sheet1", "A2", direction="precedents", depth=1)
        assert node.formula is None
        assert node.precedents == []
