"""auto_fit 智能列宽/行高估算函数的单元测试。"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from excelmanus.tools.format_tools import (
    _display_char_width,
    _estimate_display_width,
    _estimate_row_height,
    _format_number_display,
    _is_wide_char,
)


# ── _is_wide_char ──────────────────────────────────────────


class TestIsWideChar:
    def test_cjk_chars(self):
        assert _is_wide_char("你") is True
        assert _is_wide_char("好") is True
        assert _is_wide_char("表") is True

    def test_fullwidth_chars(self):
        assert _is_wide_char("Ａ") is True  # fullwidth A
        assert _is_wide_char("１") is True  # fullwidth 1

    def test_latin_chars(self):
        assert _is_wide_char("A") is False
        assert _is_wide_char("z") is False
        assert _is_wide_char("1") is False

    def test_symbols(self):
        assert _is_wide_char(".") is False
        assert _is_wide_char("%") is False


# ── _display_char_width ───────────────────────────────────


class TestDisplayCharWidth:
    def test_pure_latin(self):
        assert _display_char_width("Hello") == 5.0

    def test_pure_cjk(self):
        assert _display_char_width("你好") == 4.0

    def test_mixed(self):
        # "A你B" = 1.0 + 2.0 + 1.0 = 4.0
        assert _display_char_width("A你B") == 4.0

    def test_empty(self):
        assert _display_char_width("") == 0.0


# ── _format_number_display ────────────────────────────────


class TestFormatNumberDisplay:
    def test_percentage(self):
        result = _format_number_display(0.85, "0%")
        assert result == "85%"

    def test_percentage_with_decimals(self):
        result = _format_number_display(0.856, "0.0%")
        assert result == "85.6%"

    def test_comma_format(self):
        result = _format_number_display(1234567, "#,##0")
        assert result == "1,234,567"

    def test_comma_with_decimals(self):
        result = _format_number_display(1234.5, "#,##0.00")
        assert result == "1,234.50"

    def test_currency_prefix(self):
        result = _format_number_display(1234, "$#,##0")
        assert result == "$1,234"

    def test_yen_currency(self):
        result = _format_number_display(1234.5, "¥#,##0.00")
        assert result == "¥1,234.50"

    def test_general_returns_none(self):
        assert _format_number_display(123, "General") is None

    def test_none_value(self):
        assert _format_number_display(None, "#,##0") is None

    def test_non_numeric_string(self):
        assert _format_number_display("hello", "#,##0") is None

    def test_plain_decimal(self):
        result = _format_number_display(3.14159, "0.00")
        assert result == "3.14"


# ── _estimate_display_width ───────────────────────────────


class TestEstimateDisplayWidth:
    def test_none_value(self):
        assert _estimate_display_width(None) == 0.0

    def test_basic_string(self):
        w = _estimate_display_width("Hello")
        assert w == pytest.approx(5.0, abs=0.1)

    def test_cjk_string(self):
        w = _estimate_display_width("你好世界")
        assert w == pytest.approx(8.0, abs=0.1)

    def test_bold_is_wider(self):
        normal = _estimate_display_width("Test", is_bold=False)
        bold = _estimate_display_width("Test", is_bold=True)
        assert bold > normal

    def test_larger_font_is_wider(self):
        small = _estimate_display_width("Test", font_size=11.0)
        large = _estimate_display_width("Test", font_size=16.0)
        assert large > small

    def test_number_format_applied(self):
        # 1234567 with "#,##0" → "1,234,567" (9 chars) vs "1234567" (7 chars)
        with_fmt = _estimate_display_width(1234567, number_format="#,##0")
        without_fmt = _estimate_display_width(1234567)
        assert with_fmt > without_fmt

    def test_multiline_takes_longest(self):
        w = _estimate_display_width("short\nthis is a longer line")
        # Should be width of "this is a longer line"
        expected = _estimate_display_width("this is a longer line")
        assert w == pytest.approx(expected, abs=0.1)


# ── _estimate_row_height ──────────────────────────────────


def _make_mock_cell(
    value=None,
    font_size=11.0,
    bold=False,
    wrap_text=False,
    column=1,
    number_format="General",
):
    """Create a mock cell for row height estimation."""
    cell = MagicMock()
    cell.value = value
    cell.column = column

    font = MagicMock()
    font.size = font_size
    font.bold = bold
    cell.font = font

    alignment = MagicMock()
    alignment.wrap_text = wrap_text
    cell.alignment = alignment

    cell.number_format = number_format
    return cell


class TestEstimateRowHeight:
    def test_empty_row(self):
        cells = [_make_mock_cell(value=None)]
        h = _estimate_row_height(cells)
        assert h >= 15.0  # minimum

    def test_default_font_height(self):
        cells = [_make_mock_cell(value="Hello")]
        h = _estimate_row_height(cells)
        # 11pt * 1.35 = 14.85, but min is 15.0
        assert h >= 14.0
        assert h <= 20.0

    def test_large_font_taller(self):
        small = _estimate_row_height([_make_mock_cell(value="A", font_size=11)])
        large = _estimate_row_height([_make_mock_cell(value="A", font_size=18)])
        assert large > small

    def test_cjk_taller_than_latin(self):
        latin = _estimate_row_height([_make_mock_cell(value="Hello")])
        cjk = _estimate_row_height([_make_mock_cell(value="你好")])
        assert cjk >= latin  # CJK uses higher line_height_factor

    def test_wrap_text_increases_height(self):
        # Narrow column + long text + wrap_text → multi-line → taller
        col_widths = {"A": 10.0}
        no_wrap = _estimate_row_height(
            [_make_mock_cell(value="This is a very long text that should wrap", wrap_text=False)],
            col_widths,
        )
        with_wrap = _estimate_row_height(
            [_make_mock_cell(value="This is a very long text that should wrap", wrap_text=True)],
            col_widths,
        )
        assert with_wrap > no_wrap

    def test_explicit_newlines(self):
        single = _estimate_row_height([_make_mock_cell(value="Line1")])
        multi = _estimate_row_height([_make_mock_cell(value="Line1\nLine2\nLine3")])
        assert multi > single


# ── 集成测试：adjust_column_width / adjust_row_height ─────


class TestAutoFitIntegration:
    """Integration tests using actual openpyxl workbooks."""

    @pytest.fixture()
    def workbook_path(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        from excelmanus.tools.format_tools import init_guard

        init_guard(str(tmp_path))

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        # Row 1: header (bold, center)
        for col, header in enumerate(["ID", "名称", "金额", "百分比"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2-4: data
        data = [
            (1, "苹果公司", 1234567.89, 0.856),
            (2, "谷歌母公司Alphabet", 9876543.21, 0.123),
            (3, "微软", 5555555.55, 0.5),
        ]
        for r, (id_, name, amount, pct) in enumerate(data, 2):
            ws.cell(row=r, column=1, value=id_)
            ws.cell(row=r, column=2, value=name)
            c = ws.cell(row=r, column=3, value=amount)
            c.number_format = "#,##0.00"
            p = ws.cell(row=r, column=4, value=pct)
            p.number_format = "0.0%"

        path = tmp_path / "test_autofit.xlsx"
        wb.save(path)
        return path

    def test_auto_fit_column_width(self, workbook_path: Path):
        from excelmanus.tools.format_tools import adjust_column_width

        result = json.loads(
            adjust_column_width(
                file_path=str(workbook_path),
                auto_fit=True,
                sheet_name="Test",
            )
        )
        assert result["status"] == "success"
        cols = result["columns_adjusted"]

        # Column B (名称) should be wider than column A (ID) due to CJK content
        assert cols["B"] > cols["A"]
        # Column C (金额) should be wide due to number format
        assert cols["C"] > 10

    def test_auto_fit_row_height(self, workbook_path: Path):
        from excelmanus.tools.format_tools import adjust_row_height

        result = json.loads(
            adjust_row_height(
                file_path=str(workbook_path),
                auto_fit=True,
                sheet_name="Test",
            )
        )
        assert result["status"] == "success"
        rows = result["rows_adjusted"]

        # Header row (12pt bold) should be taller than data rows (11pt)
        assert float(rows["1"]) >= float(rows["2"])
