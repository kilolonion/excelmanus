"""transform_data 回归测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excelmanus.tools import data_tools


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    data_tools.init_guard(str(tmp_path))


def _build_multi_sheet_book(path: Path, first_sheet_name: str = "Data") -> None:
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = first_sheet_name
    ws_data.append(["id", "val"])
    ws_data.append([1, 10])
    ws_data.append([2, 20])

    ws_lookup = wb.create_sheet("Lookup")
    ws_lookup.append(["code", "name"])
    ws_lookup.append(["A", "Alpha"])
    ws_lookup.append(["B", "Beta"])

    wb.save(path)
    wb.close()


def _header_values(path: Path, sheet_name: str) -> list[str]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        return [str(cell.value) for cell in ws[1] if cell.value is not None]
    finally:
        wb.close()


class TestTransformDataSheetSafety:
    def test_inplace_transform_preserves_other_sheets(self, tmp_path: Path) -> None:
        source = tmp_path / "input.xlsx"
        _build_multi_sheet_book(source)

        result = json.loads(
            data_tools.transform_data(
                file_path=str(source),
                sheet_name="Data",
                operations=[{"type": "add_column", "name": "tag", "value": "ok"}],
            )
        )
        assert result["status"] == "success"
        assert result["sheet"] == "Data"

        wb = load_workbook(source, data_only=True)
        try:
            assert "Data" in wb.sheetnames
            assert "Lookup" in wb.sheetnames
            assert wb["Lookup"]["A2"].value == "A"
            assert wb["Lookup"]["B2"].value == "Alpha"
            headers = [cell.value for cell in wb["Data"][1]]
            assert "tag" in headers
        finally:
            wb.close()

    def test_output_new_file_preserves_other_sheets(self, tmp_path: Path) -> None:
        source = tmp_path / "source.xlsx"
        output = tmp_path / "output.xlsx"
        _build_multi_sheet_book(source)

        result = json.loads(
            data_tools.transform_data(
                file_path=str(source),
                sheet_name="Data",
                output_path=str(output),
                operations=[{"type": "rename", "columns": {"val": "value"}}],
            )
        )
        assert result["status"] == "success"
        assert result["file"] == "output.xlsx"

        assert "val" in _header_values(source, "Data")
        assert "value" not in _header_values(source, "Data")

        wb_out = load_workbook(output, data_only=True)
        try:
            assert "Data" in wb_out.sheetnames
            assert "Lookup" in wb_out.sheetnames
            out_headers = [cell.value for cell in wb_out["Data"][1]]
            assert "value" in out_headers
            assert wb_out["Lookup"]["A3"].value == "B"
            assert wb_out["Lookup"]["B3"].value == "Beta"
        finally:
            wb_out.close()

    def test_default_sheet_name_replaces_first_sheet(self, tmp_path: Path) -> None:
        source = tmp_path / "first_sheet.xlsx"
        _build_multi_sheet_book(source, first_sheet_name="MainSheet")

        result = json.loads(
            data_tools.transform_data(
                file_path=str(source),
                operations=[{"type": "add_column", "name": "flag", "value": 1}],
            )
        )
        assert result["status"] == "success"
        assert result["sheet"] == "MainSheet"

        wb = load_workbook(source, data_only=True)
        try:
            headers_main = [cell.value for cell in wb["MainSheet"][1]]
            assert "flag" in headers_main
            assert wb["Lookup"]["B2"].value == "Alpha"
        finally:
            wb.close()

