"""条件格式工具测试：add_color_scale、add_data_bar、add_conditional_rule。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excelmanus.tools.advanced_format_tools import (
    add_color_scale,
    add_conditional_rule,
    add_data_bar,
    init_guard,
)


@pytest.fixture(autouse=True)
def _init_guard(tmp_path: Path) -> None:
    """初始化 FileAccessGuard 为测试目录。"""
    init_guard(str(tmp_path))


def _make_numeric_sheet(tmp_path: Path, name: str = "data.xlsx") -> Path:
    """创建含数值数据的示例 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws["A1"] = "产品"
    ws["B1"] = "单价"
    ws["C1"] = "级别"
    for i in range(2, 22):
        ws[f"A{i}"] = f"产品{i - 1}"
        ws[f"B{i}"] = i * 10
        ws[f"C{i}"] = "FATAL" if i % 5 == 0 else "INFO"
    fp = tmp_path / name
    wb.save(fp)
    wb.close()
    return fp


# ── add_color_scale 测试 ──────────────────────────────────


class TestAddColorScale:
    """add_color_scale 工具测试套件。"""

    def test_three_color_scale(self, tmp_path: Path) -> None:
        """三色色阶（默认参数）。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(add_color_scale(str(fp), "B2:B21", sheet_name="销售明细"))
        assert result["status"] == "success"
        assert result["scale_type"] == "three_color"

        wb = load_workbook(fp)
        ws = wb["销售明细"]
        cf_rules = ws.conditional_formatting
        assert len(list(cf_rules)) >= 1
        wb.close()

    def test_two_color_scale(self, tmp_path: Path) -> None:
        """二色色阶（mid_color=None）。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_color_scale(
                str(fp), "B2:B21",
                sheet_name="销售明细",
                mid_color=None,
                min_color="FFFFFF",
                max_color="FF0000",
            )
        )
        assert result["status"] == "success"
        assert result["scale_type"] == "two_color"

    def test_custom_thresholds(self, tmp_path: Path) -> None:
        """自定义阈值和类型。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_color_scale(
                str(fp), "B2:B21",
                min_type="num", min_value=0,
                mid_type="num", mid_value=100,
                max_type="num", max_value=200,
            )
        )
        assert result["status"] == "success"


# ── add_data_bar 测试 ──────────────────────────────────────


class TestAddDataBar:
    """add_data_bar 工具测试套件。"""

    def test_default_data_bar(self, tmp_path: Path) -> None:
        """默认参数数据条。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(add_data_bar(str(fp), "B2:B21", sheet_name="销售明细"))
        assert result["status"] == "success"
        assert result["show_value"] is True

        wb = load_workbook(fp)
        ws = wb["销售明细"]
        assert len(list(ws.conditional_formatting)) >= 1
        wb.close()

    def test_custom_color_and_hide_value(self, tmp_path: Path) -> None:
        """自定义颜色并隐藏数值。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_data_bar(str(fp), "B2:B21", color="FF6600", show_value=False)
        )
        assert result["status"] == "success"
        assert result["show_value"] is False


# ── add_conditional_rule 测试 ──────────────────────────────


class TestAddConditionalRule:
    """add_conditional_rule 工具测试套件。"""

    def test_cell_is_greater_than(self, tmp_path: Path) -> None:
        """cell_is 规则：大于某值高亮。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="cell_is",
                operator="greaterThan",
                values=[100],
                fill_color="FF0000",
                font_bold=True,
            )
        )
        assert result["status"] == "success"
        assert "cell_is(greaterThan)" in result["rule_detail"]

    def test_cell_is_between(self, tmp_path: Path) -> None:
        """cell_is 规则：between 范围。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="cell_is",
                operator="between",
                values=[50, 150],
                fill_color="FFFF00",
            )
        )
        assert result["status"] == "success"

    def test_cell_is_between_requires_two_values(self, tmp_path: Path) -> None:
        """between 需要恰好 2 个值。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="cell_is",
                operator="between",
                values=[50],
            )
        )
        assert "error" in result

    def test_formula_rule(self, tmp_path: Path) -> None:
        """formula 规则：FATAL 行整行变色。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "A2:C21",
                rule_type="formula",
                formula='=$C2="FATAL"',
                fill_color="8B0000",
                font_color="FFFFFF",
            )
        )
        assert result["status"] == "success"
        assert "formula" in result["rule_detail"]

    def test_icon_set_rule(self, tmp_path: Path) -> None:
        """icon_set 规则：三箭头。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="icon_set",
                icon_style="3Arrows",
            )
        )
        assert result["status"] == "success"
        assert "icon_set(3Arrows)" in result["rule_detail"]

    def test_invalid_rule_type(self, tmp_path: Path) -> None:
        """无效 rule_type 报错。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(str(fp), "B2:B21", rule_type="invalid")
        )
        assert "error" in result

    def test_cell_is_missing_operator(self, tmp_path: Path) -> None:
        """cell_is 缺少 operator 报错。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(str(fp), "B2:B21", rule_type="cell_is")
        )
        assert "error" in result

    def test_formula_missing_formula(self, tmp_path: Path) -> None:
        """formula 规则缺少 formula 报错。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(str(fp), "B2:B21", rule_type="formula")
        )
        assert "error" in result

    def test_icon_set_missing_style(self, tmp_path: Path) -> None:
        """icon_set 缺少 icon_style 报错。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(str(fp), "B2:B21", rule_type="icon_set")
        )
        assert "error" in result

    def test_numeric_formula_auto_coercion(self, tmp_path: Path) -> None:
        """formula 传入数值时自动转为 values，兼容 LLM 误传场景。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="cell_is",
                operator="lessThan",
                formula=1000,
                fill_color="FF0000",
            )
        )
        assert result["status"] == "success"
        assert "cell_is(lessThan)" in result["rule_detail"]

    def test_numeric_formula_float_auto_coercion(self, tmp_path: Path) -> None:
        """formula 传入浮点数时自动转为 values。"""
        fp = _make_numeric_sheet(tmp_path)
        result = json.loads(
            add_conditional_rule(
                str(fp), "B2:B21",
                rule_type="cell_is",
                operator="greaterThan",
                formula=99.5,
                fill_color="00FF00",
            )
        )
        assert result["status"] == "success"
