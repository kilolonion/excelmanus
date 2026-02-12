"""格式化 Skill 单元测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from excelmanus.security import SecurityViolationError
from excelmanus.skills import ToolDef
from excelmanus.skills.format_skill import (
    SKILL_DESCRIPTION,
    SKILL_NAME,
    adjust_column_width,
    format_cells,
    get_tools,
    init_guard,
)


# ── Fixtures ──────────────────────────────────────────────


@pytest.fixture
def workspace(tmp_path: Path) -> Path:
    """创建临时工作目录并初始化 FileAccessGuard。"""
    init_guard(str(tmp_path))
    return tmp_path


@pytest.fixture
def sample_excel(workspace: Path) -> Path:
    """创建包含示例数据的 Excel 文件。"""
    df = pd.DataFrame(
        {
            "姓名": ["张三", "李四", "王五"],
            "销售额": [12000, 15000, 9000],
            "部门": ["技术部", "市场部", "财务部"],
        }
    )
    path = workspace / "data.xlsx"
    df.to_excel(path, index=False)
    return path


# ── Skill 元数据测试 ──────────────────────────────────────


class TestSkillMetadata:
    """Skill 模块约定测试。"""

    def test_skill_name_defined(self) -> None:
        assert isinstance(SKILL_NAME, str)
        assert len(SKILL_NAME) > 0

    def test_skill_description_defined(self) -> None:
        assert isinstance(SKILL_DESCRIPTION, str)
        assert len(SKILL_DESCRIPTION) > 0

    def test_get_tools_returns_two_tools(self) -> None:
        """get_tools() 应返回 2 个 ToolDef（format_cells、adjust_column_width）。"""
        tools = get_tools()
        assert len(tools) == 2
        assert all(isinstance(t, ToolDef) for t in tools)

    def test_tool_names(self) -> None:
        names = {t.name for t in get_tools()}
        assert names == {"format_cells", "adjust_column_width"}

    def test_tools_have_input_schema(self) -> None:
        for tool in get_tools():
            assert tool.input_schema.get("type") == "object"
            assert "properties" in tool.input_schema
            assert "file_path" in tool.input_schema["required"]

    def test_tool_funcs_callable(self) -> None:
        for tool in get_tools():
            assert callable(tool.func)


# ── format_cells 测试 ─────────────────────────────────────


class TestFormatCells:
    """format_cells 工具测试。"""

    def test_apply_font(self, sample_excel: Path, workspace: Path) -> None:
        """应用字体样式应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "A1:C1", font={"bold": True, "size": 14})
        )
        assert result["status"] == "success"
        assert result["cells_formatted"] == 3

        # 验证样式已写入
        wb = load_workbook(workspace / "data.xlsx")
        cell = wb.active["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
        wb.close()

    def test_apply_fill(self, sample_excel: Path, workspace: Path) -> None:
        """应用填充颜色应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "A1:C1", fill={"color": "FFFF00"})
        )
        assert result["status"] == "success"

        wb = load_workbook(workspace / "data.xlsx")
        cell = wb.active["A1"]
        assert cell.fill.start_color.rgb == "00FFFF00"
        wb.close()

    def test_apply_border(self, sample_excel: Path, workspace: Path) -> None:
        """应用边框样式应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "A1:C3", border={"style": "thin", "color": "000000"})
        )
        assert result["status"] == "success"
        assert result["cells_formatted"] == 9

        wb = load_workbook(workspace / "data.xlsx")
        cell = wb.active["B2"]
        assert cell.border.left.style == "thin"
        wb.close()

    def test_apply_alignment(self, sample_excel: Path, workspace: Path) -> None:
        """应用对齐设置应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "A1:C1", alignment={"horizontal": "center", "wrap_text": True})
        )
        assert result["status"] == "success"

        wb = load_workbook(workspace / "data.xlsx")
        cell = wb.active["A1"]
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.wrap_text is True
        wb.close()

    def test_apply_number_format(self, sample_excel: Path, workspace: Path) -> None:
        """应用数字格式应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "B2:B4", number_format="#,##0.00")
        )
        assert result["status"] == "success"

        wb = load_workbook(workspace / "data.xlsx")
        cell = wb.active["B2"]
        assert cell.number_format == "#,##0.00"
        wb.close()

    def test_single_cell(self, sample_excel: Path) -> None:
        """单个单元格格式化应成功。"""
        result = json.loads(
            format_cells("data.xlsx", "A1", font={"bold": True})
        )
        assert result["status"] == "success"
        assert result["cells_formatted"] == 1

    def test_with_sheet_name(self, workspace: Path) -> None:
        """指定 sheet_name 应在对应工作表上操作。"""
        df = pd.DataFrame({"值": [1, 2, 3]})
        path = workspace / "multi.xlsx"
        with pd.ExcelWriter(path) as writer:
            df.to_excel(writer, sheet_name="数据表", index=False)

        result = json.loads(
            format_cells("multi.xlsx", "A1", sheet_name="数据表", font={"bold": True})
        )
        assert result["status"] == "success"

    def test_path_traversal_rejected(self, workspace: Path) -> None:
        """路径穿越应被拒绝。"""
        with pytest.raises(SecurityViolationError, match="路径"):
            format_cells("../../etc/data.xlsx", "A1", font={"bold": True})


# ── adjust_column_width 测试 ──────────────────────────────


class TestAdjustColumnWidth:
    """adjust_column_width 工具测试。"""

    def test_manual_width(self, sample_excel: Path, workspace: Path) -> None:
        """手动指定列宽应成功。"""
        result = json.loads(
            adjust_column_width("data.xlsx", columns={"A": 25, "B": 15})
        )
        assert result["status"] == "success"
        assert result["columns_adjusted"]["A"] == 25
        assert result["columns_adjusted"]["B"] == 15

        wb = load_workbook(workspace / "data.xlsx")
        assert wb.active.column_dimensions["A"].width == 25
        assert wb.active.column_dimensions["B"].width == 15
        wb.close()

    def test_auto_fit(self, sample_excel: Path, workspace: Path) -> None:
        """自动适配列宽应成功，且宽度大于 0。"""
        result = json.loads(
            adjust_column_width("data.xlsx", auto_fit=True)
        )
        assert result["status"] == "success"
        adjusted = result["columns_adjusted"]
        assert len(adjusted) > 0
        for width in adjusted.values():
            assert width > 0

    def test_no_columns_no_autofit(self, sample_excel: Path) -> None:
        """既不指定 columns 也不开启 auto_fit 时，不调整任何列。"""
        result = json.loads(
            adjust_column_width("data.xlsx")
        )
        assert result["status"] == "success"
        assert result["columns_adjusted"] == {}

    def test_with_sheet_name(self, workspace: Path) -> None:
        """指定 sheet_name 应在对应工作表上操作。"""
        df = pd.DataFrame({"数据": [100, 200]})
        path = workspace / "sheets.xlsx"
        with pd.ExcelWriter(path) as writer:
            df.to_excel(writer, sheet_name="报表", index=False)

        result = json.loads(
            adjust_column_width("sheets.xlsx", columns={"A": 20}, sheet_name="报表")
        )
        assert result["status"] == "success"

    def test_path_traversal_rejected(self, workspace: Path) -> None:
        """路径穿越应被拒绝。"""
        with pytest.raises(SecurityViolationError, match="路径"):
            adjust_column_width("../../../etc/data.xlsx", auto_fit=True)
