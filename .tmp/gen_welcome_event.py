"""生成迎新活动排班 Excel — 含 VLOOKUP 公式"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import random, string

random.seed(42)

wb = openpyxl.Workbook()

# ── 样式 ──
header_font = Font(bold=True, size=12)
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center")

def style_header(ws, row, cols, fill=header_fill, font=header_font_white):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

# ── 姓名池 ──
surnames = list("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜")
given_chars = list("伟芳娜敏静丽强磊洋勇艳杰娟涛明超秀霞平刚桂英华慧建文晓玲博宇哲翔鑫")

def rand_name():
    return random.choice(surnames) + "".join(random.choices(given_chars, k=random.choice([1, 2])))

# ── 班级与角色 ──
classes = [f"物理{i}班" for i in range(1, 9)] + ["电子1班", "电子2班"]

# 每班角色分配
roles_mandatory = ["班长", "副班长", "团支书", "学习委员", "生活委员", "文体委员"]
union_roles = ["团委干事", "学生会干事"]  # 每班各抽 1-2 人

students_all = []  # (学号, 姓名, 班级, 角色, 是否必须参加)
sid_counter = 202501001

for cls in classes:
    # 班干部 6 人
    for role in roles_mandatory:
        students_all.append((str(sid_counter), rand_name(), cls, role, "是"))
        sid_counter += 1
    # 团委/学生会干事 2 人
    for role in random.sample(union_roles * 2, 2):
        students_all.append((str(sid_counter), rand_name(), cls, role, "是"))
        sid_counter += 1
    # 自愿报名同学 12-18 人
    vol_count = random.randint(12, 18)
    for _ in range(vol_count):
        students_all.append((str(sid_counter), rand_name(), cls, "普通同学", "否（自愿）"))
        sid_counter += 1

random.shuffle(students_all)

# ===== Sheet 1: 学生花名册 =====
ws1 = wb.active
ws1.title = "学生花名册"
ws1.append(["学号", "姓名", "班级", "角色", "是否必须参加", "联系电话", "报名状态"])
style_header(ws1, 1, 7)

for i, (sid, name, cls, role, must) in enumerate(students_all, start=2):
    phone = f"138{random.randint(10000000, 99999999)}"
    status = "已确认" if must == "是" else random.choice(["已报名", "未报名", "已报名"])
    ws1.append([sid, name, cls, role, must, phone, status])
    for c in range(1, 8):
        ws1.cell(row=i, column=c).border = thin_border
        ws1.cell(row=i, column=c).alignment = center

for c in range(1, 8):
    ws1.column_dimensions[get_column_letter(c)].width = 16

total_students = len(students_all)

# ===== Sheet 2: 角色参与要求表（VLOOKUP 源表）=====
ws2 = wb.create_sheet("角色参与要求")
ws2.append(["角色", "参与要求", "职责说明", "优先级"])
style_header(ws2, 1, 4)

role_data = [
    ("班长", "必须参加", "负责本班同学的组织协调", "A"),
    ("副班长", "必须参加", "协助班长，负责后勤保障", "A"),
    ("团支书", "必须参加", "负责思想动员与宣传", "A"),
    ("学习委员", "必须参加", "负责活动记录与总结", "A"),
    ("生活委员", "必须参加", "负责物资采购与分发", "A"),
    ("文体委员", "必须参加", "负责文艺节目与体育项目策划", "A"),
    ("团委干事", "必须参加", "协助团委工作，负责签到考勤", "A"),
    ("学生会干事", "必须参加", "协助学生会工作，负责场地布置", "A"),
    ("普通同学", "自愿报名", "按分组参与活动", "B"),
]
for r, row in enumerate(role_data, start=2):
    ws2.append(list(row))
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).alignment = center

for c in range(1, 5):
    ws2.column_dimensions[get_column_letter(c)].width = 22

# ===== Sheet 3: 活动时段与任务安排 =====
ws3 = wb.create_sheet("活动时段安排")
ws3.append(["时段编号", "日期", "时段", "大组长", "大组长班级", "大组长学号",
            "组别", "任务名称", "任务地点", "组长", "组长班级", "组长学号", "组员人数"])
style_header(ws3, 1, 13)

time_slots = [
    ("T1", "周六", "上午 8:30-11:30"),
    ("T2", "周六", "下午 14:00-17:00"),
    ("T3", "周日", "上午 8:30-11:30"),
    ("T4", "周日", "下午 14:00-17:00"),
]

tasks_pool = [
    ("校园导览讲解", "教学楼A区"),
    ("迎新横幅布置", "体育馆前广场"),
    ("新生注册引导", "行政楼一楼大厅"),
    ("行李搬运志愿", "学生宿舍区"),
    ("校史馆参观带队", "校史馆"),
    ("社团招新协助", "学生活动中心"),
    ("迎新晚会彩排", "大礼堂"),
    ("新生家长接待", "校门口接待站"),
    ("宿舍文化布置", "宿舍楼公共区"),
    ("校园安全巡逻", "校园主干道"),
    ("餐厅引导服务", "第一食堂"),
    ("体育器材搬运", "体育馆"),
    ("迎新物资分发", "新生报到处"),
    ("摄影摄像记录", "全校区"),
    ("医疗应急保障", "校医院旁"),
    ("交通疏导指挥", "校门口"),
]

# 筛选已报名/已确认的学生
confirmed = [s for s in students_all if s[4] == "是" or True]  # 简化：全部参与分配
random.shuffle(confirmed)

# 年级长
grade_leader = None
for s in students_all:
    if s[3] == "班长" and "物理1" in s[2]:
        grade_leader = s
        break

group_names = ["A组", "B组", "C组", "D组"]
student_idx = 0
row_num = 2

# 大组长候选（各班班长/副班长）
leader_candidates = [s for s in students_all if s[3] in ("班长", "副班长")]
random.shuffle(leader_candidates)
leader_idx = 0

# 组长候选（各班其他班干部）
group_leader_candidates = [s for s in students_all if s[3] in ("团支书", "学习委员", "生活委员", "文体委员", "团委干事", "学生会干事")]
random.shuffle(group_leader_candidates)
gl_idx = 0

task_idx = 0

for slot_id, day, period in time_slots:
    # 大组长
    dl = leader_candidates[leader_idx % len(leader_candidates)]
    leader_idx += 1
    
    slot_tasks = []
    for g in range(4):
        task_name, task_loc = tasks_pool[task_idx % len(tasks_pool)]
        task_idx += 1
        # 组长
        gl = group_leader_candidates[gl_idx % len(group_leader_candidates)]
        gl_idx += 1
        member_count = random.randint(8, 15)
        
        ws3.append([
            slot_id, day, period,
            dl[1], dl[2], dl[0],
            group_names[g], task_name, task_loc,
            gl[1], gl[2], gl[0],
            member_count
        ])
        for c in range(1, 14):
            ws3.cell(row=row_num, column=c).border = thin_border
            ws3.cell(row=row_num, column=c).alignment = center
        row_num += 1

for c in range(1, 14):
    ws3.column_dimensions[get_column_letter(c)].width = 18

# ===== Sheet 4: 分组明细（含 VLOOKUP 公式）=====
ws4 = wb.create_sheet("分组明细")
ws4.append(["时段编号", "组别", "学号", "姓名", "班级", "角色（VLOOKUP）",
            "参与要求（VLOOKUP）", "职责说明（VLOOKUP）", "优先级（VLOOKUP）"])
style_header(ws4, 1, 9)

# 分配学生到各组
assign_idx = 0
row4 = 2
for slot_i, (slot_id, day, period) in enumerate(time_slots):
    for g_i, gname in enumerate(group_names):
        count = random.randint(8, 12)
        for _ in range(count):
            s = confirmed[assign_idx % len(confirmed)]
            assign_idx += 1
            sid_val = s[0]
            name_val = s[1]
            cls_val = s[2]
            
            # VLOOKUP: 根据学号从花名册查角色
            # 花名册在 Sheet1 A:D 列，学号在 A 列，角色在 D 列
            role_formula = f'=VLOOKUP(C{row4},学生花名册!$A:$D,4,FALSE)'
            
            # VLOOKUP: 根据角色从角色参与要求表查参与要求
            req_formula = f'=VLOOKUP(F{row4},角色参与要求!$A:$B,2,FALSE)'
            
            # VLOOKUP: 根据角色查职责说明
            duty_formula = f'=VLOOKUP(F{row4},角色参与要求!$A:$C,3,FALSE)'
            
            # VLOOKUP: 根据角色查优先级
            priority_formula = f'=VLOOKUP(F{row4},角色参与要求!$A:$D,4,FALSE)'
            
            ws4.append([slot_id, gname, sid_val, name_val, cls_val,
                        role_formula, req_formula, duty_formula, priority_formula])
            for c in range(1, 10):
                ws4.cell(row=row4, column=c).border = thin_border
                ws4.cell(row=row4, column=c).alignment = center
            row4 += 1

for c in range(1, 10):
    ws4.column_dimensions[get_column_letter(c)].width = 20

# ===== Sheet 5: 管理架构 =====
ws5 = wb.create_sheet("管理架构")
ws5.append(["层级", "职务", "姓名", "学号", "班级", "管辖范围", "联系电话（VLOOKUP）"])
style_header(ws5, 1, 7)

# 年级长
r5 = 2
ws5.append([
    "第一层", "年级长（全程）",
    grade_leader[1], grade_leader[0], grade_leader[2],
    "管辖全部4个时段的大组长",
    f'=VLOOKUP(D{r5},学生花名册!$A:$F,6,FALSE)'
])
for c in range(1, 8):
    ws5.cell(row=r5, column=c).border = thin_border
    ws5.cell(row=r5, column=c).alignment = center
r5 += 1

# 大组长
for slot_i, (slot_id, day, period) in enumerate(time_slots):
    dl = leader_candidates[slot_i % len(leader_candidates)]
    ws5.append([
        "第二层", f"大组长（{day}{period}）",
        dl[1], dl[0], dl[2],
        f"管辖{day}{period}的A/B/C/D四个小组",
        f'=VLOOKUP(D{r5},学生花名册!$A:$F,6,FALSE)'
    ])
    for c in range(1, 8):
        ws5.cell(row=r5, column=c).border = thin_border
        ws5.cell(row=r5, column=c).alignment = center
    r5 += 1

# 组长（16个）
gl_used = 0
for slot_i, (slot_id, day, period) in enumerate(time_slots):
    for gname in group_names:
        gl = group_leader_candidates[gl_used % len(group_leader_candidates)]
        gl_used += 1
        ws5.append([
            "第三层", f"组长（{day}{period}-{gname}）",
            gl[1], gl[0], gl[2],
            f"负责{gname}全部组员",
            f'=VLOOKUP(D{r5},学生花名册!$A:$F,6,FALSE)'
        ])
        for c in range(1, 8):
            ws5.cell(row=r5, column=c).border = thin_border
            ws5.cell(row=r5, column=c).alignment = center
        r5 += 1

for c in range(1, 8):
    ws5.column_dimensions[get_column_letter(c)].width = 28

# ===== Sheet 6: 统计汇总（含 VLOOKUP + COUNTIF 等）=====
ws6 = wb.create_sheet("统计汇总")
ws6.append(["班级", "总人数", "必须参加人数", "自愿报名人数", "已报名人数", "未报名人数", "参与率"])
style_header(ws6, 1, 7)

for r6_i, cls in enumerate(classes, start=2):
    ws6.append([
        cls,
        f'=COUNTIF(学生花名册!$C:$C,A{r6_i})',
        f'=COUNTIFS(学生花名册!$C:$C,A{r6_i},学生花名册!$E:$E,"是")',
        f'=COUNTIFS(学生花名册!$C:$C,A{r6_i},学生花名册!$E:$E,"否（自愿）")',
        f'=COUNTIFS(学生花名册!$C:$C,A{r6_i},学生花名册!$G:$G,"已报名")+COUNTIFS(学生花名册!$C:$C,A{r6_i},学生花名册!$G:$G,"已确认")',
        f'=COUNTIFS(学生花名册!$C:$C,A{r6_i},学生花名册!$G:$G,"未报名")',
        f'=IF(B{r6_i}=0,"",E{r6_i}/B{r6_i})',
    ])
    for c in range(1, 8):
        ws6.cell(row=r6_i, column=c).border = thin_border
        ws6.cell(row=r6_i, column=c).alignment = center

# 参与率列设为百分比格式
for r in range(2, 12):
    ws6.cell(row=r, column=7).number_format = '0.0%'

# 汇总行
r_total = 12
ws6.cell(row=r_total, column=1, value="合计")
ws6.cell(row=r_total, column=1).font = Font(bold=True)
for c in range(2, 7):
    ws6.cell(row=r_total, column=c, value=f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}11)')
    ws6.cell(row=r_total, column=c).border = thin_border
    ws6.cell(row=r_total, column=c).alignment = center
ws6.cell(row=r_total, column=7, value=f'=IF(B{r_total}=0,"",E{r_total}/B{r_total})')
ws6.cell(row=r_total, column=7).number_format = '0.0%'
ws6.cell(row=r_total, column=7).border = thin_border

for c in range(1, 8):
    ws6.column_dimensions[get_column_letter(c)].width = 18

# ===== Sheet 7: VLOOKUP 查询工具表 =====
ws7 = wb.create_sheet("快速查询")
ws7.append(["【输入学号查询学生信息】"])
ws7.cell(row=1, column=1).font = Font(bold=True, size=14, color="C00000")
ws7.merge_cells("A1:F1")

ws7.append(["查询学号", "", "姓名", "班级", "角色", "联系电话"])
style_header(ws7, 2, 6, fill=PatternFill("solid", fgColor="FFC000"), font=Font(bold=True, size=12))

# 示例查询行
sample_sids = random.sample([s[0] for s in students_all], 5)
for ri, sid in enumerate(sample_sids, start=3):
    ws7.cell(row=ri, column=1, value=int(sid))
    ws7.cell(row=ri, column=3, value=f'=VLOOKUP(A{ri},学生花名册!$A:$F,2,FALSE)')
    ws7.cell(row=ri, column=4, value=f'=VLOOKUP(A{ri},学生花名册!$A:$F,3,FALSE)')
    ws7.cell(row=ri, column=5, value=f'=VLOOKUP(A{ri},学生花名册!$A:$F,4,FALSE)')
    ws7.cell(row=ri, column=6, value=f'=VLOOKUP(A{ri},学生花名册!$A:$F,6,FALSE)')
    for c in range(1, 7):
        ws7.cell(row=ri, column=c).border = thin_border
        ws7.cell(row=ri, column=c).alignment = center

ws7.append([])
ws7.append(["【输入角色查询参与要求】"])
r_role_title = ws7.max_row
ws7.cell(row=r_role_title, column=1).font = Font(bold=True, size=14, color="C00000")
ws7.merge_cells(f"A{r_role_title}:D{r_role_title}")

ws7.append(["查询角色", "", "参与要求", "职责说明"])
rh = ws7.max_row
style_header(ws7, rh, 4, fill=PatternFill("solid", fgColor="70AD47"), font=Font(bold=True, size=12))

sample_roles = ["班长", "团支书", "普通同学", "学生会干事", "文体委员"]
for role in sample_roles:
    rr = ws7.max_row + 1
    ws7.cell(row=rr, column=1, value=role)
    ws7.cell(row=rr, column=3, value=f'=VLOOKUP(A{rr},角色参与要求!$A:$B,2,FALSE)')
    ws7.cell(row=rr, column=4, value=f'=VLOOKUP(A{rr},角色参与要求!$A:$C,3,FALSE)')
    for c in range(1, 5):
        ws7.cell(row=rr, column=c).border = thin_border
        ws7.cell(row=rr, column=c).alignment = center

for c in range(1, 7):
    ws7.column_dimensions[get_column_letter(c)].width = 22

# ── 保存 ──
output_path = ".tmp/迎新活动排班表.xlsx"
wb.save(output_path)
print(f"✅ 已生成: {output_path}")
print(f"   学生总数: {total_students}")
print(f"   Sheet 数: {len(wb.sheetnames)}")
print(f"   Sheets: {wb.sheetnames}")
