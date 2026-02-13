#!/usr/bin/env python3
"""
生成一个大型、多工作表、样式丰富的 Excel 压力测试文件。
用于全面挑战 ExcelManus agent 的处理能力。

包含：
- 10+ 工作表
- 数千行数据
- 丰富的样式（字体、填充、边框、对齐、数字格式）
- 合并单元格
- 条件格式
- 数据验证
- 公式（SUM/AVERAGE/VLOOKUP/IF/COUNTIF 等）
- 图表（柱状图、折线图、饼图）
- 冻结窗格
- 筛选器
- 超链接
- 批注
- 多种数据类型（日期、百分比、货币、布尔值、长文本）
"""

import random
import string
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers,
    NamedStyle, Protection
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, DataBarRule, IconSetRule
)


# ── 辅助函数 ──────────────────────────────────────────────

def rand_name():
    """随机中文姓名"""
    surnames = "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜"
    given = "伟芳娜秀英敏静丽强磊洋勇艳杰娟涛明超秀华达飞刚平志明建国文辉力诚"
    return random.choice(surnames) + "".join(random.choices(given, k=random.randint(1, 2)))


def rand_company():
    prefixes = ["华", "中", "新", "大", "金", "万", "天", "国", "盛", "恒"]
    suffixes = ["科技", "电子", "贸易", "实业", "集团", "投资", "材料", "能源", "医药", "食品"]
    return random.choice(prefixes) + random.choice(prefixes) + random.choice(suffixes)


def rand_product():
    cats = ["笔记本电脑", "智能手机", "平板电脑", "无线耳机", "智能手表",
            "机械键盘", "显示器", "路由器", "移动电源", "摄像头",
            "打印机", "扫描仪", "投影仪", "服务器", "交换机"]
    return random.choice(cats)


def rand_city():
    cities = ["北京", "上海", "广州", "深圳", "杭州", "成都", "武汉", "南京",
              "重庆", "西安", "苏州", "天津", "长沙", "郑州", "青岛", "大连",
              "厦门", "宁波", "合肥", "福州", "昆明", "贵阳", "南宁", "海口"]
    return random.choice(cities)


def rand_department():
    return random.choice(["销售部", "市场部", "技术部", "财务部", "人事部",
                          "运营部", "产品部", "客服部", "法务部", "采购部"])


def rand_date(start_year=2022, end_year=2025):
    start = datetime.date(start_year, 1, 1)
    end = datetime.date(end_year, 12, 31)
    delta = (end - start).days
    return start + datetime.timedelta(days=random.randint(0, delta))


# ── 样式预设 ──────────────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="微软雅黑", size=11, italic=True, color="808080")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
THICK_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium")
)

MONEY_FMT = '#,##0.00"元"'
PCT_FMT = '0.00%'
DATE_FMT = 'YYYY-MM-DD'
INT_FMT = '#,##0'

FILL_LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_LIGHT_RED = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
FILL_LIGHT_PURPLE = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
FILL_LIGHT_ORANGE = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")


def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def apply_data_border(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet 1: 销售明细（2000行） ──────────────────────────

def create_sales_detail(wb: Workbook):
    ws = wb.active
    ws.title = "销售明细"
    ws.sheet_properties.tabColor = "2F5496"

    # 标题区
    ws.merge_cells("A1:L1")
    ws["A1"] = "2022-2025 年度销售明细数据"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:L2")
    ws["A2"] = f"生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 2000 条记录"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["订单编号", "日期", "客户名称", "公司", "城市", "产品",
               "数量", "单价(元)", "总金额(元)", "折扣率", "实付金额(元)", "状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, len(headers))

    statuses = ["已完成", "进行中", "已取消", "待审核", "已退款"]
    status_fills = {
        "已完成": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "进行中": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "已取消": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "待审核": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "已退款": PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
    }

    for i in range(2000):
        row = i + 4
        order_id = f"ORD-{2022 + i // 500}-{i + 1:05d}"
        date = rand_date()
        name = rand_name()
        company = rand_company()
        city = rand_city()
        product = rand_product()
        qty = random.randint(1, 200)
        price = round(random.uniform(50, 15000), 2)
        discount = round(random.uniform(0, 0.3), 4)
        status = random.choice(statuses)

        ws.cell(row=row, column=1, value=order_id)
        ws.cell(row=row, column=2, value=date).number_format = DATE_FMT
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=company)
        ws.cell(row=row, column=5, value=city)
        ws.cell(row=row, column=6, value=product)
        ws.cell(row=row, column=7, value=qty).number_format = INT_FMT
        ws.cell(row=row, column=8, value=price).number_format = MONEY_FMT
        # 公式：数量 * 单价
        ws.cell(row=row, column=9).value = f"=G{row}*H{row}"
        ws.cell(row=row, column=9).number_format = MONEY_FMT
        ws.cell(row=row, column=10, value=discount).number_format = PCT_FMT
        # 公式：总金额 * (1 - 折扣率)
        ws.cell(row=row, column=11).value = f"=I{row}*(1-J{row})"
        ws.cell(row=row, column=11).number_format = MONEY_FMT
        status_cell = ws.cell(row=row, column=12, value=status)
        if status in status_fills:
            status_cell.fill = status_fills[status]

        # 交替行颜色
        if i % 2 == 0:
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_BLUE

    apply_data_border(ws, 3, 2003, 12)

    # 汇总行
    summary_row = 2004
    ws.cell(row=summary_row, column=6, value="合计").font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=7).value = f"=SUM(G4:G2003)"
    ws.cell(row=summary_row, column=7).number_format = INT_FMT
    ws.cell(row=summary_row, column=9).value = f"=SUM(I4:I2003)"
    ws.cell(row=summary_row, column=9).number_format = MONEY_FMT
    ws.cell(row=summary_row, column=11).value = f"=SUM(K4:K2003)"
    ws.cell(row=summary_row, column=11).number_format = MONEY_FMT
    for c in range(1, 13):
        ws.cell(row=summary_row, column=c).font = Font(bold=True)
        ws.cell(row=summary_row, column=c).border = THICK_BORDER
        ws.cell(row=summary_row, column=c).fill = FILL_LIGHT_YELLOW

    # 条件格式：总金额 > 100000 高亮
    ws.conditional_formatting.add(
        "I4:I2003",
        CellIsRule(operator="greaterThan", formula=["100000"],
                   fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
                   font=Font(bold=True, color="FFFFFF"))
    )
    # 数据条
    ws.conditional_formatting.add(
        "K4:K2003",
        DataBarRule(start_type="min", end_type="max",
                    color="5B9BD5")
    )

    # 冻结窗格
    ws.freeze_panes = "A4"
    # 自动筛选
    ws.auto_filter.ref = f"A3:L2003"

    set_col_widths(ws, {"A": 18, "B": 14, "C": 12, "D": 16, "E": 10,
                        "F": 16, "G": 10, "H": 14, "I": 16, "J": 10,
                        "K": 16, "L": 12})

    return ws


# ── Sheet 2: 员工花名册（500行） ─────────────────────────

def create_employee_roster(wb: Workbook):
    ws = wb.create_sheet("员工花名册")
    ws.sheet_properties.tabColor = "00B050"

    headers = ["工号", "姓名", "性别", "出生日期", "入职日期", "部门",
               "职级", "基本工资", "绩效系数", "实发工资", "手机号", "邮箱",
               "学历", "婚姻状况", "紧急联系人", "备注"]

    ws.merge_cells("A1:P1")
    ws["A1"] = "员工花名册（机密）"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="006100")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = FILL_LIGHT_GREEN
    ws.row_dimensions[1].height = 35

    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    levels = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "M1", "M2", "M3"]
    educations = ["大专", "本科", "硕士", "博士", "MBA"]
    marriage = ["已婚", "未婚", "离异"]

    for i in range(500):
        row = i + 3
        emp_id = f"EMP{i + 1:04d}"
        name = rand_name()
        gender = random.choice(["男", "女"])
        birth = rand_date(1975, 2000)
        hire = rand_date(2015, 2025)
        dept = rand_department()
        level = random.choice(levels)
        base_salary = random.randint(5000, 80000)
        perf = round(random.uniform(0.6, 1.5), 2)
        phone = f"1{random.choice(['3','5','7','8','9'])}{random.randint(100000000, 999999999)}"
        email = f"{''.join(random.choices(string.ascii_lowercase, k=6))}@example.com"
        edu = random.choice(educations)
        mar = random.choice(marriage)
        emergency = rand_name()
        note = random.choice(["", "", "", "试用期", "即将转正", "优秀员工", "外派中", ""])

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        gender_cell = ws.cell(row=row, column=3, value=gender)
        if gender == "女":
            gender_cell.font = Font(color="FF69B4")
        else:
            gender_cell.font = Font(color="4169E1")
        ws.cell(row=row, column=4, value=birth).number_format = DATE_FMT
        ws.cell(row=row, column=5, value=hire).number_format = DATE_FMT
        ws.cell(row=row, column=6, value=dept)
        level_cell = ws.cell(row=row, column=7, value=level)
        if level.startswith("M"):
            level_cell.fill = FILL_LIGHT_PURPLE
            level_cell.font = Font(bold=True)
        ws.cell(row=row, column=8, value=base_salary).number_format = MONEY_FMT
        ws.cell(row=row, column=9, value=perf).number_format = "0.00"
        ws.cell(row=row, column=10).value = f"=H{row}*I{row}"
        ws.cell(row=row, column=10).number_format = MONEY_FMT
        ws.cell(row=row, column=11, value=phone)
        ws.cell(row=row, column=12, value=email)
        ws.cell(row=row, column=13, value=edu)
        ws.cell(row=row, column=14, value=mar)
        ws.cell(row=row, column=15, value=emergency)
        ws.cell(row=row, column=16, value=note)

        if i % 2 == 1:
            for c in range(1, 17):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_GREEN

    apply_data_border(ws, 2, 502, 16)

    # 数据验证：性别列
    dv_gender = DataValidation(type="list", formula1='"男,女"', allow_blank=False)
    dv_gender.error = "请选择男或女"
    dv_gender.errorTitle = "输入错误"
    ws.add_data_validation(dv_gender)
    dv_gender.add(f"C3:C502")

    # 数据验证：学历列
    dv_edu = DataValidation(type="list", formula1='"大专,本科,硕士,博士,MBA"')
    ws.add_data_validation(dv_edu)
    dv_edu.add(f"M3:M502")

    # 条件格式：工资色阶
    ws.conditional_formatting.add(
        "J3:J502",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )

    # 批注
    ws["A1"].comment = Comment("此表包含员工敏感信息，请注意保密", "系统管理员")
    ws["H2"].comment = Comment("基本工资不含绩效奖金和补贴", "HR部门")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:P502"
    ws.protection = Protection(locked=True)

    set_col_widths(ws, {get_column_letter(i): w for i, w in enumerate([
        0, 10, 10, 8, 14, 14, 12, 8, 14, 10, 14, 14, 22, 8, 10, 12, 20
    ], 0) if i > 0})

    return ws


# ── Sheet 3: 月度汇总透视表 ──────────────────────────────

def create_monthly_pivot(wb: Workbook):
    ws = wb.create_sheet("月度汇总")
    ws.sheet_properties.tabColor = "FFC000"

    ws.merge_cells("A1:N1")
    ws["A1"] = "月度销售汇总分析表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    months = [f"{y}年{m:02d}月" for y in range(2022, 2026) for m in range(1, 13)]
    products = ["笔记本电脑", "智能手机", "平板电脑", "无线耳机", "智能手表",
                "机械键盘", "显示器", "路由器", "移动电源", "摄像头"]

    # 表头
    headers = ["月份"] + products + ["月度合计", "环比增长", "同比增长"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    for i, month in enumerate(months):
        row = i + 3
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = FILL_LIGHT_YELLOW

        for j in range(len(products)):
            col = j + 2
            val = round(random.uniform(10000, 500000), 2)
            ws.cell(row=row, column=col, value=val).number_format = MONEY_FMT

        # 月度合计公式
        last_prod_col = get_column_letter(len(products) + 1)
        ws.cell(row=row, column=len(products) + 2).value = \
            f"=SUM(B{row}:{last_prod_col}{row})"
        ws.cell(row=row, column=len(products) + 2).number_format = MONEY_FMT
        ws.cell(row=row, column=len(products) + 2).font = Font(bold=True)

        # 环比增长
        total_col = get_column_letter(len(products) + 2)
        if i > 0:
            ws.cell(row=row, column=len(products) + 3).value = \
                f"=({total_col}{row}-{total_col}{row - 1})/{total_col}{row - 1}"
        else:
            ws.cell(row=row, column=len(products) + 3, value="N/A")
        ws.cell(row=row, column=len(products) + 3).number_format = PCT_FMT

        # 同比增长
        if i >= 12:
            ws.cell(row=row, column=len(products) + 4).value = \
                f"=({total_col}{row}-{total_col}{row - 12})/{total_col}{row - 12}"
        else:
            ws.cell(row=row, column=len(products) + 4, value="N/A")
        ws.cell(row=row, column=len(products) + 4).number_format = PCT_FMT

        # 年份分组颜色
        year_fills = [FILL_LIGHT_BLUE, FILL_LIGHT_GREEN, FILL_LIGHT_YELLOW, FILL_LIGHT_ORANGE]
        year_idx = i // 12
        if year_idx < len(year_fills):
            for c in range(2, len(products) + 2):
                ws.cell(row=row, column=c).fill = year_fills[year_idx]

    total_rows = len(months)
    apply_data_border(ws, 2, total_rows + 2, len(headers))

    # 条件格式：环比增长图标集
    growth_col = get_column_letter(len(products) + 3)
    ws.conditional_formatting.add(
        f"{growth_col}3:{growth_col}{total_rows + 2}",
        IconSetRule("3Arrows", "num", [0, -0.05, 0.05])
    )

    ws.freeze_panes = "B3"
    set_col_widths(ws, {"A": 14})
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    return ws


# ── Sheet 4: 城市分布（带图表） ──────────────────────────

def create_city_analysis(wb: Workbook):
    ws = wb.create_sheet("城市分析")
    ws.sheet_properties.tabColor = "FF0000"

    cities = ["北京", "上海", "广州", "深圳", "杭州", "成都", "武汉", "南京",
              "重庆", "西安", "苏州", "天津", "长沙", "郑州", "青岛", "大连"]

    headers = ["城市", "订单数", "总销售额", "平均客单价", "退货率", "客户满意度", "市场份额"]
    ws.merge_cells("A1:G1")
    ws["A1"] = "城市销售分析"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    for i, city in enumerate(cities):
        row = i + 3
        orders = random.randint(50, 500)
        total = round(random.uniform(100000, 5000000), 2)
        avg = round(total / orders, 2)
        return_rate = round(random.uniform(0.01, 0.15), 4)
        satisfaction = round(random.uniform(3.5, 5.0), 2)
        share = round(random.uniform(0.02, 0.15), 4)

        ws.cell(row=row, column=1, value=city).font = Font(bold=True)
        ws.cell(row=row, column=2, value=orders).number_format = INT_FMT
        ws.cell(row=row, column=3, value=total).number_format = MONEY_FMT
        ws.cell(row=row, column=4, value=avg).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value=return_rate).number_format = PCT_FMT
        ws.cell(row=row, column=6, value=satisfaction).number_format = "0.00"
        ws.cell(row=row, column=7, value=share).number_format = PCT_FMT

        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_RED

    apply_data_border(ws, 2, len(cities) + 2, len(headers))

    # 柱状图：总销售额
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "各城市总销售额"
    chart1.y_axis.title = "销售额（元）"
    chart1.x_axis.title = "城市"
    chart1.style = 10
    data = Reference(ws, min_col=3, min_row=2, max_row=len(cities) + 2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(cities) + 2)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "I2")

    # 饼图：市场份额
    chart2 = PieChart()
    chart2.title = "市场份额分布"
    chart2.style = 26
    data2 = Reference(ws, min_col=7, min_row=2, max_row=len(cities) + 2)
    cats2 = Reference(ws, min_col=1, min_row=3, max_row=len(cities) + 2)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "I20")

    set_col_widths(ws, {"A": 10, "B": 10, "C": 16, "D": 14, "E": 10, "F": 12, "G": 12})
    return ws


# ── Sheet 5: 产品目录（含超链接、图片占位） ──────────────

def create_product_catalog(wb: Workbook):
    ws = wb.create_sheet("产品目录")
    ws.sheet_properties.tabColor = "7030A0"

    ws.merge_cells("A1:J1")
    ws["A1"] = "产品目录与规格参数"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="7030A0")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = FILL_LIGHT_PURPLE

    headers = ["产品编号", "产品名称", "类别", "品牌", "规格", "成本价",
               "零售价", "利润率", "库存量", "供应商链接"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    categories = {
        "笔记本电脑": ["联想", "戴尔", "华为", "苹果", "华硕"],
        "智能手机": ["华为", "小米", "OPPO", "vivo", "苹果"],
        "平板电脑": ["苹果", "华为", "三星", "联想", "小米"],
        "无线耳机": ["苹果", "索尼", "华为", "JBL", "漫步者"],
        "智能手表": ["苹果", "华为", "小米", "三星", "Garmin"],
        "机械键盘": ["Cherry", "罗技", "雷蛇", "达尔优", "HHKB"],
        "显示器": ["戴尔", "LG", "三星", "华硕", "明基"],
        "路由器": ["华为", "TP-Link", "小米", "华硕", "网件"],
    }

    specs_templates = {
        "笔记本电脑": "i7/16GB/512GB SSD/14英寸",
        "智能手机": "8GB+256GB/6.7英寸/5000mAh",
        "平板电脑": "8GB+128GB/11英寸/WiFi6",
        "无线耳机": "蓝牙5.3/主动降噪/30h续航",
        "智能手表": "1.5英寸AMOLED/GPS/心率监测",
        "机械键盘": "87键/红轴/RGB背光/Type-C",
        "显示器": "27英寸/4K/IPS/HDR400",
        "路由器": "WiFi6/AX3000/双频/Mesh",
    }

    row_idx = 3
    for cat, brands in categories.items():
        for j, brand in enumerate(brands):
            for variant in range(3):  # 每品牌3个型号
                pid = f"PRD-{row_idx - 2:04d}"
                name = f"{brand} {cat} {'旗舰' if variant == 0 else '标准' if variant == 1 else '入门'}版"
                spec = specs_templates.get(cat, "标准配置")
                cost = round(random.uniform(200, 8000), 2)
                retail = round(cost * random.uniform(1.2, 2.5), 2)

                ws.cell(row=row_idx, column=1, value=pid)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=cat)
                ws.cell(row=row_idx, column=4, value=brand)
                ws.cell(row=row_idx, column=5, value=spec)
                ws.cell(row=row_idx, column=6, value=cost).number_format = MONEY_FMT
                ws.cell(row=row_idx, column=7, value=retail).number_format = MONEY_FMT
                # 利润率公式
                ws.cell(row=row_idx, column=8).value = f"=(G{row_idx}-F{row_idx})/G{row_idx}"
                ws.cell(row=row_idx, column=8).number_format = PCT_FMT
                ws.cell(row=row_idx, column=9, value=random.randint(0, 5000)).number_format = INT_FMT
                # 超链接
                link_cell = ws.cell(row=row_idx, column=10, value="查看详情")
                link_cell.hyperlink = f"https://example.com/product/{pid}"
                link_cell.font = Font(color="0563C1", underline="single")

                # 库存预警：库存<100 红色
                if ws.cell(row=row_idx, column=9).value < 100:
                    ws.cell(row=row_idx, column=9).fill = FILL_LIGHT_RED
                    ws.cell(row=row_idx, column=9).font = Font(bold=True, color="FF0000")

                # 类别分组颜色
                cat_fills = {
                    "笔记本电脑": FILL_LIGHT_BLUE, "智能手机": FILL_LIGHT_GREEN,
                    "平板电脑": FILL_LIGHT_YELLOW, "无线耳机": FILL_LIGHT_PURPLE,
                    "智能手表": FILL_LIGHT_ORANGE, "机械键盘": FILL_LIGHT_RED,
                }
                if cat in cat_fills:
                    ws.cell(row=row_idx, column=3).fill = cat_fills[cat]

                row_idx += 1

    apply_data_border(ws, 2, row_idx - 1, len(headers))

    # 条件格式：利润率色阶
    ws.conditional_formatting.add(
        f"H3:H{row_idx - 1}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{row_idx - 1}"
    set_col_widths(ws, {"A": 12, "B": 28, "C": 14, "D": 10, "E": 28,
                        "F": 14, "G": 14, "H": 10, "I": 10, "J": 14})
    return ws


# ── Sheet 6: 财务报表（复杂公式） ────────────────────────

def create_financial_report(wb: Workbook):
    ws = wb.create_sheet("财务报表")
    ws.sheet_properties.tabColor = "002060"

    ws.merge_cells("A1:F1")
    ws["A1"] = "2022-2025 年度财务报表"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="002060")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 40

    years = ["2022", "2023", "2024", "2025"]
    headers = ["科目"] + years + ["复合增长率"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    # 收入类
    income_items = [
        ("一、营业收入", True, [50000000, 65000000, 82000000, 98000000]),
        ("  1. 产品销售收入", False, [35000000, 45000000, 58000000, 70000000]),
        ("  2. 服务收入", False, [10000000, 13000000, 16000000, 19000000]),
        ("  3. 其他收入", False, [5000000, 7000000, 8000000, 9000000]),
        ("二、营业成本", True, [30000000, 37000000, 45000000, 52000000]),
        ("  1. 原材料成本", False, [18000000, 22000000, 27000000, 31000000]),
        ("  2. 人工成本", False, [8000000, 10000000, 12000000, 14000000]),
        ("  3. 制造费用", False, [4000000, 5000000, 6000000, 7000000]),
        ("三、毛利润", True, None),  # 公式
        ("四、期间费用", True, [12000000, 15000000, 18000000, 21000000]),
        ("  1. 销售费用", False, [5000000, 6000000, 7500000, 9000000]),
        ("  2. 管理费用", False, [4000000, 5000000, 6000000, 7000000]),
        ("  3. 研发费用", False, [3000000, 4000000, 4500000, 5000000]),
        ("五、营业利润", True, None),  # 公式
        ("六、所得税费用", True, None),  # 公式
        ("七、净利润", True, None),  # 公式
    ]

    for i, (item, is_bold, values) in enumerate(income_items):
        row = i + 3
        ws.cell(row=row, column=1, value=item)
        if is_bold:
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_BLUE
        else:
            ws.cell(row=row, column=1).font = Font(size=10)

        if values:
            for j, val in enumerate(values):
                ws.cell(row=row, column=j + 2, value=val).number_format = MONEY_FMT
        elif item == "三、毛利润":
            for j in range(4):
                col_l = get_column_letter(j + 2)
                ws.cell(row=row, column=j + 2).value = f"={col_l}3-{col_l}7"
                ws.cell(row=row, column=j + 2).number_format = MONEY_FMT
        elif item == "五、营业利润":
            for j in range(4):
                col_l = get_column_letter(j + 2)
                ws.cell(row=row, column=j + 2).value = f"={col_l}11-{col_l}12"
                ws.cell(row=row, column=j + 2).number_format = MONEY_FMT
        elif item == "六、所得税费用":
            for j in range(4):
                col_l = get_column_letter(j + 2)
                ws.cell(row=row, column=j + 2).value = f"={col_l}16*0.25"
                ws.cell(row=row, column=j + 2).number_format = MONEY_FMT
        elif item == "七、净利润":
            for j in range(4):
                col_l = get_column_letter(j + 2)
                ws.cell(row=row, column=j + 2).value = f"={col_l}16-{col_l}17"
                ws.cell(row=row, column=j + 2).number_format = MONEY_FMT

        # 复合增长率 CAGR = (终值/初值)^(1/年数) - 1
        if values or item in ["三、毛利润", "五、营业利润", "六、所得税费用", "七、净利润"]:
            ws.cell(row=row, column=6).value = f"=(E{row}/B{row})^(1/3)-1"
            ws.cell(row=row, column=6).number_format = PCT_FMT

    apply_data_border(ws, 2, len(income_items) + 2, len(headers))

    # 折线图
    chart = LineChart()
    chart.title = "年度收入与利润趋势"
    chart.y_axis.title = "金额（元）"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # 营业收入
    data1 = Reference(ws, min_col=2, max_col=5, min_row=3, max_row=3)
    chart.add_data(data1, from_rows=True, titles_from_data=False)
    chart.series[0].tx = SeriesLabel(v="营业收入")

    # 净利润
    data2 = Reference(ws, min_col=2, max_col=5, min_row=18, max_row=18)
    chart.add_data(data2, from_rows=True, titles_from_data=False)
    chart.series[1].tx = SeriesLabel(v="净利润")

    cats = Reference(ws, min_col=2, max_col=5, min_row=2)
    chart.set_categories(cats)
    ws.add_chart(chart, "A21")

    set_col_widths(ws, {"A": 20, "B": 18, "C": 18, "D": 18, "E": 18, "F": 14})
    return ws


# ── Sheet 7: 跨表引用与VLOOKUP ──────────────────────────

def create_cross_reference(wb: Workbook):
    ws = wb.create_sheet("跨表引用")
    ws.sheet_properties.tabColor = "ED7D31"

    ws.merge_cells("A1:H1")
    ws["A1"] = "跨工作表引用与高级公式演示"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = FILL_LIGHT_ORANGE

    headers = ["序号", "引用说明", "公式", "结果", "公式类型", "复杂度", "说明", "验证状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    formulas = [
        ("销售明细总订单数", "=COUNTA(销售明细!A4:A2003)", "COUNTA", "★★", "统计非空单元格数"),
        ("销售明细总金额", "=SUM(销售明细!I4:I2003)", "SUM跨表", "★★", "跨表求和"),
        ("平均客单价", "=AVERAGE(销售明细!I4:I2003)", "AVERAGE跨表", "★★", "跨表平均值"),
        ("最大单笔订单", "=MAX(销售明细!I4:I2003)", "MAX跨表", "★★", "跨表最大值"),
        ("最小单笔订单", "=MIN(销售明细!I4:I2003)", "MIN跨表", "★★", "跨表最小值"),
        ("已完成订单数", '=COUNTIF(销售明细!L4:L2003,"已完成")', "COUNTIF", "★★★", "条件计数"),
        ("已取消订单数", '=COUNTIF(销售明细!L4:L2003,"已取消")', "COUNTIF", "★★★", "条件计数"),
        ("员工总数", "=COUNTA(员工花名册!A3:A502)", "COUNTA跨表", "★★", "跨表统计"),
        ("平均工资", "=AVERAGE(员工花名册!J3:J502)", "AVERAGE跨表", "★★", "跨表平均"),
        ("最高工资", "=MAX(员工花名册!J3:J502)", "MAX跨表", "★★", "跨表最大值"),
        ("产品种类数", "=COUNTA(产品目录!A3:A200)", "COUNTA跨表", "★★", "跨表统计"),
        ("IF嵌套示例", '=IF(D2>1000,"大额",IF(D2>500,"中额","小额"))', "IF嵌套", "★★★★", "多层条件判断"),
        ("TEXT格式化", '=TEXT(NOW(),"YYYY年MM月DD日 HH:MM")', "TEXT", "★★★", "日期格式化"),
        ("CONCATENATE", '=CONCATENATE("报表生成于: ",TEXT(NOW(),"YYYY-MM-DD"))', "CONCATENATE", "★★★", "文本拼接"),
    ]

    for i, (desc, formula, ftype, complexity, note) in enumerate(formulas):
        row = i + 3
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=formula).font = Font(name="Consolas", size=9, color="0000FF")
        ws.cell(row=row, column=4).value = formula  # 实际公式
        ws.cell(row=row, column=4).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value=ftype)
        ws.cell(row=row, column=6, value=complexity)
        ws.cell(row=row, column=7, value=note)
        ws.cell(row=row, column=8, value="待验证")
        ws.cell(row=row, column=8).fill = FILL_LIGHT_YELLOW

        if i % 2 == 0:
            for c in range(1, 9):
                if c != 8:
                    ws.cell(row=row, column=c).fill = FILL_LIGHT_ORANGE

    apply_data_border(ws, 2, len(formulas) + 2, len(headers))
    set_col_widths(ws, {"A": 8, "B": 20, "C": 45, "D": 18, "E": 14, "F": 10, "G": 20, "H": 12})
    return ws


# ── Sheet 8: 多维数据矩阵（大量合并单元格） ─────────────

def create_matrix_sheet(wb: Workbook):
    ws = wb.create_sheet("多维矩阵")
    ws.sheet_properties.tabColor = "00B0F0"

    ws.merge_cells("A1:R1")
    ws["A1"] = "产品-地区-季度 三维销售矩阵"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    regions = ["华北", "华东", "华南", "华中", "西南", "西北"]
    quarters = ["Q1", "Q2", "Q3", "Q4"]
    products = ["笔记本", "手机", "平板"]

    # 构建复杂表头（三层合并）
    # 第2行：地区（每个地区占4列）
    col = 2
    for region in regions:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        cell = ws.cell(row=2, column=col, value=region)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = HEADER_ALIGN
        cell.border = THICK_BORDER
        col += 4

    # 第3行：季度
    col = 2
    quarter_fills = [
        PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    ]
    for _ in regions:
        for qi, q in enumerate(quarters):
            cell = ws.cell(row=3, column=col, value=q)
            cell.font = Font(bold=True, size=10)
            cell.fill = quarter_fills[qi]
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            col += 1

    # 第1列：产品（合并）
    ws.cell(row=2, column=1, value="产品\\地区")
    ws.cell(row=2, column=1).font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).alignment = HEADER_ALIGN
    ws.merge_cells("A2:A3")

    # 数据行
    row = 4
    for year in range(2022, 2026):
        # 年份标题行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(regions) * 4 + 1)
        ws.cell(row=row, column=1, value=f"── {year}年 ──")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

        for product in products:
            ws.cell(row=row, column=1, value=product)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = FILL_LIGHT_BLUE

            col = 2
            for _ in regions:
                for qi in range(4):
                    val = round(random.uniform(10000, 200000), 2)
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = '#,##0'
                    cell.fill = quarter_fills[qi]
                    cell.border = THIN_BORDER
                    col += 1
            row += 1

    apply_data_border(ws, 2, row - 1, len(regions) * 4 + 1)
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(regions) * 4 + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12

    return ws


# ── Sheet 9: 日志数据（长文本、时间戳） ──────────────────

def create_log_sheet(wb: Workbook):
    ws = wb.create_sheet("系统日志")
    ws.sheet_properties.tabColor = "808080"

    headers = ["时间戳", "级别", "模块", "用户", "操作", "详细信息", "IP地址", "耗时(ms)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, len(headers))

    levels = ["INFO", "WARN", "ERROR", "DEBUG", "FATAL"]
    level_fills = {
        "INFO": FILL_LIGHT_BLUE, "WARN": FILL_LIGHT_YELLOW,
        "ERROR": FILL_LIGHT_RED, "DEBUG": FILL_LIGHT_GREEN,
        "FATAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }
    level_fonts = {
        "FATAL": Font(bold=True, color="FFFFFF"),
        "ERROR": Font(bold=True, color="CC0000"),
    }

    modules = ["auth", "payment", "order", "inventory", "report", "api", "scheduler", "cache"]
    actions = [
        "用户登录成功", "用户登录失败：密码错误", "创建订单", "取消订单",
        "支付成功", "支付超时", "库存不足告警", "缓存命中",
        "缓存未命中", "API调用超时", "数据库连接池耗尽", "定时任务执行完成",
        "文件上传成功", "权限校验失败", "数据导出完成", "系统健康检查通过",
    ]

    for i in range(1000):
        row = i + 2
        ts = datetime.datetime(2025, 1, 1) + datetime.timedelta(
            seconds=random.randint(0, 86400 * 30))
        level = random.choices(levels, weights=[50, 20, 15, 10, 5])[0]
        module = random.choice(modules)
        user = rand_name()
        action = random.choice(actions)

        detail = f"[{module.upper()}] {action} | session={random.randint(10000,99999)} | " \
                 f"trace_id={''.join(random.choices(string.hexdigits[:16], k=32))}"
        ip = f"192.168.{random.randint(1,254)}.{random.randint(1,254)}"
        elapsed = round(random.uniform(0.5, 5000), 2)

        ws.cell(row=row, column=1, value=ts).number_format = "YYYY-MM-DD HH:MM:SS"
        lvl_cell = ws.cell(row=row, column=2, value=level)
        if level in level_fills:
            lvl_cell.fill = level_fills[level]
        if level in level_fonts:
            lvl_cell.font = level_fonts[level]
        ws.cell(row=row, column=3, value=module)
        ws.cell(row=row, column=4, value=user)
        ws.cell(row=row, column=5, value=action)

        ws.cell(row=row, column=6, value=detail).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=7, value=ip)
        elapsed_cell = ws.cell(row=row, column=8, value=elapsed)
        elapsed_cell.number_format = "#,##0.00"
        if elapsed > 3000:
            elapsed_cell.fill = FILL_LIGHT_RED
            elapsed_cell.font = Font(bold=True, color="FF0000")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:H1001"
    set_col_widths(ws, {"A": 22, "B": 8, "C": 12, "D": 10,
                        "E": 20, "F": 60, "G": 16, "H": 12})
    return ws


# ── Sheet 10: KPI仪表盘（复杂布局） ─────────────────────

def create_kpi_dashboard(wb: Workbook):
    ws = wb.create_sheet("KPI仪表盘")
    ws.sheet_properties.tabColor = "C00000"

    ws.merge_cells("A1:L1")
    ws["A1"] = "2025年度 KPI 仪表盘"
    ws["A1"].font = Font(name="微软雅黑", size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    # KPI 卡片区域
    kpis = [
        ("总营收", "98,000,000元", "↑ 19.5%", "C00000"),
        ("净利润", "12,250,000元", "↑ 22.3%", "00B050"),
        ("客户数", "15,832", "↑ 8.7%", "0070C0"),
        ("订单量", "42,156", "↑ 15.2%", "7030A0"),
        ("退货率", "3.2%", "↓ 1.1%", "ED7D31"),
        ("满意度", "4.6/5.0", "↑ 0.3", "FFC000"),
    ]

    for i, (label, value, change, color) in enumerate(kpis):
        col_start = i * 2 + 1
        col_end = col_start + 1
        # 标签行
        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
        cell = ws.cell(row=3, column=col_start, value=label)
        cell.font = Font(size=10, color="808080")
        cell.alignment = Alignment(horizontal="center")

        # 数值行
        ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)
        val_cell = ws.cell(row=4, column=col_start, value=value)
        val_cell.font = Font(size=18, bold=True, color=color)
        val_cell.alignment = Alignment(horizontal="center")
        # 变化行
        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_end)
        chg_cell = ws.cell(row=5, column=col_start, value=change)
        chg_color = "00B050" if "↑" in change else "FF0000"
        chg_cell.font = Font(size=11, color=chg_color)
        chg_cell.alignment = Alignment(horizontal="center")

        # 卡片边框
        for r in range(3, 6):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = Border(
                    left=Side(style="medium" if c == col_start else "thin"),
                    right=Side(style="medium" if c == col_end else "thin"),
                    top=Side(style="medium" if r == 3 else "thin"),
                    bottom=Side(style="medium" if r == 5 else "thin"),
                )

    # 月度趋势数据表
    ws.merge_cells("A7:L7")
    ws["A7"] = "月度关键指标趋势"
    ws["A7"].font = Font(size=14, bold=True, color="333333")
    ws["A7"].alignment = Alignment(horizontal="center")

    trend_headers = ["月份", "营收", "成本", "利润", "订单量",
                     "客单价", "新客户", "流失客户", "NPS评分",
                     "转化率", "复购率", "库存周转"]
    for col, h in enumerate(trend_headers, 1):
        ws.cell(row=8, column=col, value=h)
    apply_header_style(ws, 8, len(trend_headers))

    for i in range(12):
        row = i + 9
        ws.cell(row=row, column=1, value=f"2025年{i+1:02d}月")
        ws.cell(row=row, column=2, value=round(random.uniform(6e6, 12e6), 2)).number_format = MONEY_FMT
        ws.cell(row=row, column=3, value=round(random.uniform(3e6, 7e6), 2)).number_format = MONEY_FMT
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value=random.randint(2000, 5000)).number_format = INT_FMT
        ws.cell(row=row, column=6).value = f"=B{row}/E{row}"
        ws.cell(row=row, column=6).number_format = MONEY_FMT
        ws.cell(row=row, column=7, value=random.randint(100, 800)).number_format = INT_FMT
        ws.cell(row=row, column=8, value=random.randint(20, 150)).number_format = INT_FMT
        ws.cell(row=row, column=9, value=round(random.uniform(30, 80), 1))
        ws.cell(row=row, column=10, value=round(random.uniform(0.02, 0.12), 4)).number_format = PCT_FMT
        ws.cell(row=row, column=11, value=round(random.uniform(0.15, 0.45), 4)).number_format = PCT_FMT
        ws.cell(row=row, column=12, value=round(random.uniform(3, 12), 2))
        if i % 2 == 0:
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_RED

    apply_data_border(ws, 8, 20, 12)

    # 营收利润折线图
    chart = LineChart()
    chart.title = "月度营收与利润趋势"
    chart.style = 10
    chart.width = 24
    chart.height = 14
    rev_data = Reference(ws, min_col=2, min_row=8, max_row=20)
    profit_data = Reference(ws, min_col=4, min_row=8, max_row=20)
    cats = Reference(ws, min_col=1, min_row=9, max_row=20)
    chart.add_data(rev_data, titles_from_data=True)
    chart.add_data(profit_data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A22")

    set_col_widths(ws, {get_column_letter(i): 14 for i in range(1, 13)})
    return ws


# ── Sheet 11: 数据类型大全 ──────────────────────────────

def create_data_types_sheet(wb: Workbook):
    """各种数据类型和格式的综合展示"""
    ws = wb.create_sheet("数据类型大全")
    ws.sheet_properties.tabColor = "00B0F0"

    ws.merge_cells("A1:F1")
    ws["A1"] = "数据类型与格式化综合测试"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    sections = [
        ("数字格式", [
            ("整数", 12345, "#,##0"),
            ("负数", -9876, "#,##0;[Red]-#,##0"),
            ("小数", 3.14159265, "0.0000"),
            ("科学计数", 0.00000123, "0.00E+00"),
            ("百分比", 0.8567, "0.00%"),
            ("分数", 0.333333, "# ?/?"),
            ("人民币", 99999.99, '¥#,##0.00'),
            ("美元", 1234.56, '$#,##0.00'),
            ("千分位", 1234567890, "#,##0"),
            ("自定义", 42, '000000'),
        ]),

        ("日期时间格式", [
            ("标准日期", datetime.date(2025, 6, 15), "YYYY-MM-DD"),
            ("中文日期", datetime.date(2025, 6, 15), 'YYYY"年"MM"月"DD"日"'),
            ("短日期", datetime.date(2025, 6, 15), "MM/DD"),
            ("时间", datetime.time(14, 30, 45), "HH:MM:SS"),
            ("日期时间", datetime.datetime(2025, 6, 15, 14, 30), "YYYY-MM-DD HH:MM"),
            ("星期", datetime.date(2025, 6, 15), "DDDD"),
            ("月份名", datetime.date(2025, 6, 15), "MMMM YYYY"),
        ]),
        ("文本格式", [
            ("短文本", "Hello World", "@"),
            ("中文长文本", "这是一段较长的中文文本，用于测试单元格的自动换行和文本溢出处理能力。" * 3, "@"),
            ("特殊字符", "!@#$%^&*()_+-=[]{}|;':\",./<>?", "@"),
            ("Unicode", "🎉🚀💡📊🔥✅❌⚠️🎯📈", "@"),
            ("换行文本", "第一行\n第二行\n第三行", "@"),
            ("空字符串", "", "@"),
            ("纯空格", "   ", "@"),
        ]),

        ("布尔与特殊值", [
            ("True", True, ""),
            ("False", False, ""),
            ("零", 0, "#,##0"),
            ("None/空", None, ""),
            ("极大数", 99999999999999, "#,##0"),
            ("极小数", 0.000000001, "0.000000000"),
            ("负百分比", -0.1234, "0.00%"),
        ]),
    ]

    row = 3
    for section_name, items in sections:
        # 分区标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"▶ {section_name}")
        ws.cell(row=row, column=1).font = Font(size=13, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        sub_headers = ["描述", "原始值", "格式化显示", "数据类型", "格式代码", "字节长度"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=col, value=h)
        apply_header_style(ws, row, len(sub_headers))
        row += 1

        for desc, value, fmt in items:
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=repr(value) if value is not None else "None")
            cell = ws.cell(row=row, column=3, value=value)
            if fmt:
                cell.number_format = fmt
            if isinstance(value, str) and "\n" in value:
                cell.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 45
            ws.cell(row=row, column=4, value=type(value).__name__ if value is not None else "NoneType")
            ws.cell(row=row, column=5, value=fmt if fmt else "(默认)")
            ws.cell(row=row, column=5).font = Font(name="Consolas", size=9)
            byte_len = len(str(value).encode("utf-8")) if value is not None else 0
            ws.cell(row=row, column=6, value=byte_len)
            if row % 2 == 0:
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = FILL_LIGHT_BLUE
            row += 1
        row += 1  # 分区间空行

    apply_data_border(ws, 3, row - 2, 6)
    set_col_widths(ws, {"A": 16, "B": 30, "C": 30, "D": 14, "E": 24, "F": 12})
    return ws


# ── Sheet 12: 考勤表（复杂合并+条件格式） ────────────────

def create_attendance_sheet(wb: Workbook):
    ws = wb.create_sheet("考勤表")
    ws.sheet_properties.tabColor = "FF6600"

    ws.merge_cells("A1:AG1")
    ws["A1"] = "2025年1月 员工考勤表"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 表头：工号、姓名、部门 + 31天 + 出勤天数、迟到次数、缺勤天数
    headers = ["工号", "姓名", "部门"]
    for d in range(1, 32):
        headers.append(f"{d}日")
    headers.extend(["出勤", "迟到", "缺勤"])
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    apply_header_style(ws, 2, len(headers))

    # 日期列宽度较窄
    for i in range(4, 35):
        ws.column_dimensions[get_column_letter(i)].width = 5

    marks = ["✓", "✓", "✓", "✓", "✓", "迟", "×", "假", "✓", "✓"]
    mark_fills = {
        "✓": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "迟": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "×": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "假": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    }
    mark_fonts = {
        "✓": Font(color="006100"), "迟": Font(color="9C6500"),
        "×": Font(color="9C0006", bold=True), "假": Font(color="003399"),
    }

    for i in range(80):
        row = i + 3
        ws.cell(row=row, column=1, value=f"EMP{i+1:04d}")
        ws.cell(row=row, column=2, value=rand_name())
        ws.cell(row=row, column=3, value=rand_department())

        for d in range(31):
            col = d + 4
            # 周末自动标灰
            day_date = datetime.date(2025, 1, d + 1)
            if day_date.weekday() >= 5:
                ws.cell(row=row, column=col, value="休")
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(color="808080")
            else:
                mark = random.choice(marks)
                cell = ws.cell(row=row, column=col, value=mark)
                cell.alignment = Alignment(horizontal="center")
                if mark in mark_fills:
                    cell.fill = mark_fills[mark]
                if mark in mark_fonts:
                    cell.font = mark_fonts[mark]

        # 统计公式
        day_range_start = get_column_letter(4)
        day_range_end = get_column_letter(34)
        # 出勤天数 = COUNTIF(D:AH, "✓")
        ws.cell(row=row, column=35).value = \
            f'=COUNTIF({day_range_start}{row}:{day_range_end}{row},"✓")'
        ws.cell(row=row, column=35).number_format = INT_FMT
        ws.cell(row=row, column=35).font = Font(bold=True)
        # 迟到次数
        ws.cell(row=row, column=36).value = \
            f'=COUNTIF({day_range_start}{row}:{day_range_end}{row},"迟")'
        ws.cell(row=row, column=36).number_format = INT_FMT
        # 缺勤天数
        ws.cell(row=row, column=37).value = \
            f'=COUNTIF({day_range_start}{row}:{day_range_end}{row},"×")'
        ws.cell(row=row, column=37).number_format = INT_FMT

    apply_data_border(ws, 2, 82, 37)
    ws.freeze_panes = "D3"
    set_col_widths(ws, {"A": 10, "B": 10, "C": 10,
                        "AI": 8, "AJ": 8, "AK": 8})
    return ws


# ── Sheet 13: 隐藏工作表（测试隐藏属性） ────────────────

def create_hidden_sheet(wb: Workbook):
    ws = wb.create_sheet("_隐藏配置")
    ws.sheet_state = "hidden"

    ws["A1"] = "此工作表为隐藏配置表"
    ws["A1"].font = Font(bold=True, color="FF0000")

    # 配置数据
    configs = [
        ("系统版本", "3.0.0"),
        ("数据库连接", "postgresql://localhost:5432/excelmanus"),
        ("缓存TTL", "3600"),
        ("最大并发", "100"),
        ("日志级别", "INFO"),
        ("密钥哈希", "sha256:a1b2c3d4e5f6..."),
        ("生成时间", str(datetime.datetime.now())),
        ("生成工具", "generate_stress_test_excel.py"),
    ]
    for i, (key, val) in enumerate(configs):
        ws.cell(row=i + 3, column=1, value=key).font = Font(bold=True)
        ws.cell(row=i + 3, column=2, value=val)

    # 下拉列表数据源
    ws["D1"] = "状态列表"
    ws["D1"].font = Font(bold=True)
    statuses = ["已完成", "进行中", "已取消", "待审核", "已退款"]
    for i, s in enumerate(statuses):
        ws.cell(row=i + 2, column=4, value=s)

    ws["E1"] = "部门列表"
    ws["E1"].font = Font(bold=True)
    depts = ["销售部", "市场部", "技术部", "财务部", "人事部",
             "运营部", "产品部", "客服部", "法务部", "采购部"]
    for i, d in enumerate(depts):
        ws.cell(row=i + 2, column=5, value=d)

    return ws


# ── 主函数 ────────────────────────────────────────────────

def main():
    """生成压力测试 Excel 文件"""
    print("🚀 开始生成压力测试 Excel 文件...")
    random.seed(42)  # 固定种子，保证可复现

    wb = Workbook()

    print("  📊 [1/12] 销售明细（2000行）...")
    create_sales_detail(wb)

    print("  👥 [2/12] 员工花名册（500行）...")
    create_employee_roster(wb)

    print("  📈 [3/12] 月度汇总透视表...")
    create_monthly_pivot(wb)

    print("  🏙️ [4/12] 城市分析（含图表）...")
    create_city_analysis(wb)

    print("  📦 [5/12] 产品目录（120+产品）...")
    create_product_catalog(wb)

    print("  💰 [6/12] 财务报表（复杂公式）...")
    create_financial_report(wb)

    print("  🔗 [7/12] 跨表引用...")
    create_cross_reference(wb)

    print("  🧮 [8/12] 多维矩阵（大量合并）...")
    create_matrix_sheet(wb)

    print("  📝 [9/12] 系统日志（1000行）...")
    create_log_sheet(wb)

    print("  📊 [10/12] KPI仪表盘...")
    create_kpi_dashboard(wb)

    print("  🔢 [11/12] 数据类型大全...")
    create_data_types_sheet(wb)

    print("  📅 [12/12] 考勤表（80人×31天）...")
    create_attendance_sheet(wb)

    print("  🔒 [bonus] 隐藏配置表...")
    create_hidden_sheet(wb)

    # 保存文件
    output_path = Path("stress_test_comprehensive.xlsx")
    print(f"\n  💾 保存到 {output_path}...")
    wb.save(str(output_path))

    # 输出统计信息
    file_size = output_path.stat().st_size
    size_mb = file_size / (1024 * 1024)
    print(f"\n✅ 生成完成！")
    print(f"   文件：{output_path.absolute()}")
    print(f"   大小：{size_mb:.2f} MB ({file_size:,} 字节)")
    print(f"   工作表数：{len(wb.sheetnames)}")
    print(f"   工作表列表：")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"     {i:2d}. {name}")
    print(f"\n📋 覆盖特性：")
    features = [
        "2000+ 行销售数据", "500 行员工数据", "48 个月汇总",
        "16 城市分析", "120+ 产品目录", "4 年财务报表",
        "跨表公式引用", "三维矩阵（合并单元格）", "1000 行日志",
        "KPI 仪表盘", "数据类型全覆盖", "80人考勤表",
        "隐藏工作表", "条件格式（色阶/数据条/图标集）",
        "数据验证（下拉列表）", "冻结窗格", "自动筛选",
        "超链接", "批注", "多种图表（柱状/折线/饼图）",
        "丰富样式（字体/填充/边框/对齐/数字格式）",
        "SUM/AVERAGE/MAX/MIN/COUNTIF/IF/TEXT 等公式",
    ]
    for f in features:
        print(f"     ✓ {f}")


if __name__ == "__main__":
    main()
