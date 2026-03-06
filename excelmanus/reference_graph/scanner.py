"""引用图谱扫描器 — Tier 1 工作表级 + Tier 2 单元格级。"""
from __future__ import annotations

import time
from collections import Counter, defaultdict
from typing import Any

from .formula_parser import FormulaRefExtractor, address_in_ref
from .models import (
    CellNode,
    CellRef,
    ExternalRef,
    RefType,
    SheetRefEdge,
    SheetRefSummary,
    WorkbookRefIndex,
)

_MAX_SAMPLE_FORMULAS = 3


class Tier1Scanner:
    """轻量级工作表级引用扫描（文件上传时执行）。"""

    def __init__(self) -> None:
        self._extractor = FormulaRefExtractor()

    def scan(self, file_path: str) -> WorkbookRefIndex:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=False, read_only=True)
        try:
            return self._scan_workbook(file_path, wb)
        finally:
            wb.close()

    def _scan_workbook(self, file_path: str, wb: Any) -> WorkbookRefIndex:
        sheets: dict[str, SheetRefSummary] = {}
        edge_map: dict[tuple[str, str], dict[str, Any]] = {}
        external_refs: list[ExternalRef] = []
        named_ranges: dict[str, str] = {}

        for dn in wb.defined_names.values():
            if dn.attr_text and not dn.attr_text.startswith("#"):
                named_ranges[dn.name] = dn.attr_text

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            formula_count = 0
            self_ref_count = 0
            func_counter: Counter[str] = Counter()
            outgoing: dict[str, int] = defaultdict(int)
            outgoing_samples: dict[str, list[str]] = defaultdict(list)

            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str) or not val.startswith("="):
                        continue
                    formula_count += 1
                    refs = self._extractor.extract(val)
                    funcs = self._extractor.extract_functions(val)
                    for f in funcs:
                        func_counter[f] += 1

                    has_self = False
                    for ref in refs:
                        if ref.file_path:
                            external_refs.append(ExternalRef(
                                book_name=ref.file_path,
                                sheet_name=ref.sheet_name or "",
                                cell_or_range=ref.cell_or_range,
                                source_sheet=ws_name,
                                source_cell=cell.coordinate if hasattr(cell, "coordinate") else "",
                            ))
                        elif ref.sheet_name and ref.sheet_name != ws_name:
                            target = ref.sheet_name
                            outgoing[target] += 1
                            if len(outgoing_samples[target]) < _MAX_SAMPLE_FORMULAS:
                                outgoing_samples[target].append(val)
                        else:
                            has_self = True
                    if has_self:
                        self_ref_count += 1

            top_funcs = [f for f, _ in func_counter.most_common(10)]
            sheets[ws_name] = SheetRefSummary(
                sheet_name=ws_name,
                formula_count=formula_count,
                outgoing_refs=[],
                incoming_refs=[],
                self_refs=self_ref_count,
                formula_patterns=top_funcs,
            )

            for target, count in outgoing.items():
                key = (ws_name, target)
                edge_map[key] = {
                    "source": ws_name,
                    "target": target,
                    "count": count,
                    "samples": outgoing_samples[target],
                }

        cross_sheet_edges: list[SheetRefEdge] = []
        for (src, tgt), info in edge_map.items():
            edge = SheetRefEdge(
                source_sheet=src,
                target_sheet=tgt,
                ref_type=RefType.FORMULA,
                ref_count=info["count"],
                sample_formulas=info["samples"],
                column_pairs=[],
            )
            cross_sheet_edges.append(edge)
            if src in sheets:
                sheets[src].outgoing_refs.append(edge)
            if tgt in sheets:
                sheets[tgt].incoming_refs.append(edge)

        return WorkbookRefIndex(
            file_path=file_path,
            sheets=sheets,
            cross_sheet_edges=cross_sheet_edges,
            external_refs=external_refs,
            named_ranges=named_ranges,
            built_at=time.time(),
        )


class Tier2Resolver:
    """单元格级深度引用解析（按需调用）。"""

    _MAX_DEPTH = 5

    def __init__(self) -> None:
        self._extractor = FormulaRefExtractor()

    def resolve(
        self,
        file_path: str,
        sheet_name: str,
        address: str,
        *,
        direction: str = "both",
        depth: int = 1,
    ) -> CellNode:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=False, read_only=True)
        try:
            return self._resolve(wb, sheet_name, address, direction,
                                 min(depth, self._MAX_DEPTH))
        finally:
            wb.close()

    # ------------------------------------------------------------------

    def _find_formula(self, wb: Any, sheet_name: str, address: str) -> str | None:
        """读取指定单元格的公式（无公式返回 None）。"""
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                if coord == address:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        return val
                    return None
        return None

    def _resolve(
        self,
        wb: Any,
        sheet_name: str,
        address: str,
        direction: str,
        depth: int,
    ) -> CellNode:
        formula = self._find_formula(wb, sheet_name, address)

        direct_prec: list[CellRef] = []
        if formula and direction in ("both", "precedents"):
            direct_prec = self._extractor.extract(formula)

        direct_deps: list[CellRef] = []
        if direction in ("both", "dependents"):
            direct_deps = self._find_dependents(wb, sheet_name, address)

        all_prec = list(direct_prec)
        all_deps = list(direct_deps)

        if depth > 1 and direction in ("both", "precedents"):
            all_prec = self._expand_precedents(
                wb, sheet_name, direct_prec, depth - 1,
            )

        if depth > 1 and direction in ("both", "dependents"):
            all_deps = self._expand_dependents(
                wb, sheet_name, address, direct_deps, depth - 1,
            )

        return CellNode(
            sheet=sheet_name,
            address=address,
            formula=formula,
            precedents=all_prec,
            dependents=all_deps,
        )

    # ------------------------------------------------------------------

    def _expand_precedents(
        self,
        wb: Any,
        origin_sheet: str,
        direct: list[CellRef],
        remaining_depth: int,
    ) -> list[CellRef]:
        """BFS 展开 precedents 至 remaining_depth 层。"""
        result = list(direct)
        seen: set[str] = {r.display() for r in direct}
        frontier = list(direct)

        for _ in range(remaining_depth):
            if not frontier:
                break
            next_frontier: list[CellRef] = []
            for ref in frontier:
                if ":" in ref.cell_or_range:
                    continue
                ref_sheet = ref.sheet_name or origin_sheet
                sub_formula = self._find_formula(wb, ref_sheet, ref.cell_or_range)
                if not sub_formula:
                    continue
                sub_refs = self._extractor.extract(sub_formula)
                for sr in sub_refs:
                    key = sr.display()
                    if key not in seen:
                        seen.add(key)
                        result.append(sr)
                        next_frontier.append(sr)
            frontier = next_frontier
        return result

    def _expand_dependents(
        self,
        wb: Any,
        origin_sheet: str,
        origin_address: str,
        direct: list[CellRef],
        remaining_depth: int,
    ) -> list[CellRef]:
        """BFS 展开 dependents 至 remaining_depth 层。"""
        result = list(direct)
        seen: set[str] = {
            f"{(r.sheet_name or origin_sheet)}!{r.cell_or_range}"
            for r in direct
        }
        seen.add(f"{origin_sheet}!{origin_address}")
        frontier = list(direct)

        for _ in range(remaining_depth):
            if not frontier:
                break
            next_frontier: list[CellRef] = []
            for ref in frontier:
                ref_sheet = ref.sheet_name or origin_sheet
                sub_deps = self._find_dependents(wb, ref_sheet, ref.cell_or_range)
                for sd in sub_deps:
                    key = f"{(sd.sheet_name or ref_sheet)}!{sd.cell_or_range}"
                    if key not in seen:
                        seen.add(key)
                        result.append(sd)
                        next_frontier.append(sd)
            frontier = next_frontier
        return result

    # ------------------------------------------------------------------

    def _find_dependents(
        self, wb: Any, sheet_name: str, address: str,
    ) -> list[CellRef]:
        """在工作簿中搜索引用了指定单元格的所有公式。"""
        results: list[CellRef] = []
        seen: set[str] = set()

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str) or not val.startswith("="):
                        continue
                    refs = self._extractor.extract(val)
                    for ref in refs:
                        ref_sheet = ref.sheet_name or ws_name
                        if ref_sheet == sheet_name and address_in_ref(address, ref.cell_or_range):
                            coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                            key = f"{ws_name}!{coord}"
                            if coord and key not in seen:
                                seen.add(key)
                                results.append(CellRef(
                                    sheet_name=ws_name if ws_name != sheet_name else None,
                                    cell_or_range=coord,
                                ))
        return results
