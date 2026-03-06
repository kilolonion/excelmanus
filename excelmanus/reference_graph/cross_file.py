"""跨文件关联扫描器。"""
from __future__ import annotations

from typing import Any


class CrossFileRefScanner:
    """检测两个 Excel 文件之间的关联关系。"""

    def scan_pair(self, file_path_a: str, file_path_b: str) -> dict[str, Any]:
        from openpyxl import load_workbook

        headers_a = self._extract_headers(file_path_a)
        headers_b = self._extract_headers(file_path_b)

        shared_columns = sorted(set(headers_a) & set(headers_b))

        if headers_a and headers_b:
            union = set(headers_a) | set(headers_b)
            intersection = set(headers_a) & set(headers_b)
            similarity = len(intersection) / len(union) if union else 0.0
        else:
            similarity = 0.0

        return {
            "file_a": file_path_a,
            "file_b": file_path_b,
            "shared_columns": shared_columns,
            "structural_similarity": round(similarity, 3),
            "headers_a": headers_a,
            "headers_b": headers_b,
        }

    @staticmethod
    def _extract_headers(file_path: str) -> list[str]:
        """提取工作簿第一个工作表的表头列名。"""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            headers: list[str] = []
            for row in ws.iter_rows(max_row=1):
                for cell in row:
                    val = cell.value
                    if val is not None:
                        headers.append(str(val).strip())
            return headers
        finally:
            wb.close()
