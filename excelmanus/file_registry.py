"""FileRegistry：统一文件注册表 — 元数据 + provenance + 物理操作 + 路径解析。

当前实现：核心数据模型、注册/查询/事件记录、路径解析、panorama 构建。
目标为统一收敛为文件管理的唯一外部接口，后续可扩展全文件类型扫描、
uploads 接入、Staging/CoW/checkpoint 等。
"""
from __future__ import annotations

import logging
import os
import secrets
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any

from excelmanus.security.path_utils import resolve_in_workspace, to_workspace_relative

if TYPE_CHECKING:
    from excelmanus.database import Database

logger = logging.getLogger(__name__)

# Excel 扩展名集合
_EXCEL_EXTENSIONS = frozenset({".xlsx", ".xls", ".xlsm", ".xlsb", ".csv"})

# 图片扩展名
_IMAGE_EXTENSIONS = frozenset({".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"})

# 递归扫描时跳过的噪音目录
_SKIP_DIRS: frozenset[str] = frozenset({
    ".git", ".venv", "node_modules", "__pycache__",
    ".worktrees", "dist", "build",
})

# 全文件扫描时跳过的二进制/编译文件扩展名
_SKIP_EXTENSIONS: frozenset[str] = frozenset({
    ".pyc", ".pyo", ".so", ".dylib", ".dll", ".exe",
    ".o", ".a", ".class", ".jar", ".war",
    ".whl", ".egg", ".tar", ".gz", ".bz2", ".xz", ".zst",
    ".db", ".sqlite", ".sqlite3",
    ".woff", ".woff2", ".ttf", ".otf", ".eot",
    ".DS_Store",
})

# Panorama 自适应阈值
_PANORAMA_FULL_THRESHOLD = 20
_PANORAMA_COMPACT_THRESHOLD = 100

# 目录语义标签
_DIR_LABELS: dict[str, str] = {
    "uploads": "用户上传",
    "outputs": "产出物",
    "outputs/backups": "备份副本",
}


def _dir_label(parent: str) -> str:
    normalized = parent.replace("\\", "/").strip("/")
    for prefix, label_ in _DIR_LABELS.items():
        if normalized == prefix or normalized.startswith(prefix + "/"):
            return label_
    return ""


def _detect_file_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in _EXCEL_EXTENSIONS:
        return "excel" if ext != ".csv" else "csv"
    if ext in _IMAGE_EXTENSIONS:
        return "image"
    if ext in (".txt", ".md", ".json", ".xml", ".yaml", ".yml", ".log"):
        return "text"
    return "other"


def _new_id() -> str:
    return secrets.token_hex(8)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ── 数据模型 ──────────────────────────────────────────────────


@dataclass
class FileEntry:
    """Registry 中的文件记录。"""

    id: str
    workspace: str
    canonical_path: str
    original_name: str
    file_type: str = "other"
    size_bytes: int = 0
    origin: str = "scan"
    origin_session_id: str | None = None
    origin_turn: int | None = None
    origin_tool: str | None = None
    parent_file_id: str | None = None
    sheet_meta: list[dict] = field(default_factory=list)
    content_hash: str = ""
    mtime_ns: int = 0
    staging_path: str | None = None
    is_active_cow: bool = False
    created_at: str = ""
    updated_at: str = ""
    deleted_at: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "workspace": self.workspace,
            "canonical_path": self.canonical_path,
            "original_name": self.original_name,
            "file_type": self.file_type,
            "size_bytes": self.size_bytes,
            "origin": self.origin,
            "origin_session_id": self.origin_session_id,
            "origin_turn": self.origin_turn,
            "origin_tool": self.origin_tool,
            "parent_file_id": self.parent_file_id,
            "sheet_meta": self.sheet_meta,
            "content_hash": self.content_hash,
            "mtime_ns": self.mtime_ns,
            "staging_path": self.staging_path,
            "is_active_cow": self.is_active_cow,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
            "deleted_at": self.deleted_at,
        }

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> FileEntry:
        return cls(
            id=d["id"],
            workspace=d["workspace"],
            canonical_path=d["canonical_path"],
            original_name=d["original_name"],
            file_type=d.get("file_type", "other"),
            size_bytes=d.get("size_bytes", 0),
            origin=d.get("origin", "scan"),
            origin_session_id=d.get("origin_session_id"),
            origin_turn=d.get("origin_turn"),
            origin_tool=d.get("origin_tool"),
            parent_file_id=d.get("parent_file_id"),
            sheet_meta=d.get("sheet_meta", []),
            content_hash=d.get("content_hash", ""),
            mtime_ns=d.get("mtime_ns", 0),
            staging_path=d.get("staging_path"),
            is_active_cow=d.get("is_active_cow", False),
            created_at=d.get("created_at", ""),
            updated_at=d.get("updated_at", ""),
            deleted_at=d.get("deleted_at"),
        )


@dataclass
class FileEvent:
    """文件生命周期事件。"""

    id: str
    file_id: str
    event_type: str
    session_id: str | None = None
    turn: int | None = None
    tool_name: str | None = None
    details: dict = field(default_factory=dict)
    created_at: str = ""

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> FileEvent:
        return cls(
            id=d["id"],
            file_id=d["file_id"],
            event_type=d["event_type"],
            session_id=d.get("session_id"),
            turn=d.get("turn"),
            tool_name=d.get("tool_name"),
            details=d.get("details", {}),
            created_at=d.get("created_at", ""),
        )


@dataclass
class ScanResult:
    """scan_workspace() 的返回值。"""

    total_files: int = 0
    new_files: int = 0
    updated_files: int = 0
    deleted_files: int = 0
    cache_hits: int = 0
    scan_duration_ms: int = 0


# ── FileRegistry 核心类 ──────────────────────────────────────


class FileRegistry:
    """全局文件注册表 — 统一元数据 + provenance + 路径解析 + 物理操作。

    enable_versions=True 时内部组合 FileVersionManager，
    提供 staging / CoW / checkpoint / rollback 统一接口。
    """

    def __init__(
        self,
        database: "Database",
        workspace_root: str | Path,
        *,
        enable_versions: bool = False,
    ) -> None:
        from excelmanus.stores.file_registry_store import FileRegistryStore

        self._store = FileRegistryStore(database)
        self._workspace_root = Path(workspace_root).resolve()
        self._workspace_key = str(self._workspace_root)

        # 内存热缓存（启动时从 DB 加载）
        self._path_cache: dict[str, FileEntry] = {}  # canonical_path → entry
        self._id_to_path: dict[str, str] = {}  # file_id → canonical_path
        self._alias_cache: dict[str, str] = {}  # alias_value → file_id

        # 可选版本管理层（session 级）
        self._fvm: Any = None
        if enable_versions:
            from excelmanus.file_versions import FileVersionManager
            self._fvm = FileVersionManager(self._workspace_root)

        self._load_cache()

    @property
    def workspace_root(self) -> Path:
        return self._workspace_root

    # ── 缓存管理 ─────────────────────────────────────────────

    def _load_cache(self) -> None:
        """从 DB 加载到内存缓存。"""
        try:
            rows = self._store.list_all(self._workspace_key, include_deleted=False)
            file_ids: list[str] = []
            for row in rows:
                entry = FileEntry.from_dict(row)
                self._path_cache[entry.canonical_path] = entry
                self._id_to_path[entry.id] = entry.canonical_path
                file_ids.append(entry.id)
            # 单次批量加载所有别名（避免 N+1 查询）
            if file_ids:
                all_aliases = self._store.get_all_aliases_for_files(file_ids)
                for fid, aliases in all_aliases.items():
                    for a in aliases:
                        self._alias_cache[a["alias_value"]] = fid
        except Exception:
            logger.debug("FileRegistry 缓存加载失败", exc_info=True)

    def _cache_entry(self, entry: FileEntry) -> None:
        """更新缓存中的 entry。"""
        self._path_cache[entry.canonical_path] = entry
        self._id_to_path[entry.id] = entry.canonical_path

    def _invalidate_entry(self, canonical_path: str) -> None:
        """从缓存中移除 entry。"""
        entry = self._path_cache.pop(canonical_path, None)
        if entry:
            self._id_to_path.pop(entry.id, None)

    # ── 路径工具 ─────────────────────────────────────────────

    def _resolve(self, file_path: str) -> Path:
        return resolve_in_workspace(file_path, self._workspace_root)

    def _to_rel(self, abs_path: Path) -> str:
        return to_workspace_relative(abs_path, self._workspace_root)

    # ── 注册入口 ─────────────────────────────────────────────

    def register_upload(
        self,
        canonical_path: str,
        original_name: str,
        file_type: str = "",
        size_bytes: int = 0,
        session_id: str | None = None,
        turn: int | None = None,
        sheet_meta: list[dict] | None = None,
    ) -> FileEntry:
        """注册上传文件。"""
        if not file_type:
            file_type = _detect_file_type(canonical_path)
        now = _now_iso()

        # 复用已有记录的 ID，避免缓存/DB 不一致
        existing = self._path_cache.get(canonical_path)
        entry = FileEntry(
            id=existing.id if existing else _new_id(),
            workspace=self._workspace_key,
            canonical_path=canonical_path,
            original_name=original_name,
            file_type=file_type,
            size_bytes=size_bytes,
            origin="uploaded",
            origin_session_id=session_id,
            origin_turn=turn,
            sheet_meta=sheet_meta or [],
            created_at=existing.created_at if existing else now,
            updated_at=now,
        )
        self._store.upsert_file(entry.to_dict())
        self._cache_entry(entry)

        # 添加 display_name 别名
        if original_name != canonical_path:
            alias_id = _new_id()
            self._store.add_alias(alias_id, entry.id, "display_name", original_name)
            self._alias_cache[original_name] = entry.id

        # 记录上传事件
        self.record_event(
            entry.id, "uploaded",
            session_id=session_id, turn=turn,
            details={"original_name": original_name, "size_bytes": size_bytes},
        )
        return entry

    def register_from_scan(
        self,
        canonical_path: str,
        original_name: str,
        size_bytes: int = 0,
        mtime_ns: int = 0,
        file_type: str = "",
        sheet_meta: list[dict] | None = None,
        content_hash: str = "",
    ) -> FileEntry:
        """从工作区扫描注册文件（新增或更新）。"""
        if not file_type:
            file_type = _detect_file_type(canonical_path)
        now = _now_iso()

        # 检查是否已存在
        existing = self._path_cache.get(canonical_path)
        if existing:
            # 更新已有记录
            existing.size_bytes = size_bytes
            existing.mtime_ns = mtime_ns
            existing.file_type = file_type
            existing.content_hash = content_hash
            if sheet_meta is not None:
                existing.sheet_meta = sheet_meta
            existing.updated_at = now
            existing.deleted_at = None  # 复活
            self._store.upsert_file(existing.to_dict())
            self._cache_entry(existing)
            return existing

        entry = FileEntry(
            id=_new_id(),
            workspace=self._workspace_key,
            canonical_path=canonical_path,
            original_name=original_name,
            file_type=file_type,
            size_bytes=size_bytes,
            origin="scan",
            sheet_meta=sheet_meta or [],
            content_hash=content_hash,
            mtime_ns=mtime_ns,
            created_at=now,
            updated_at=now,
        )
        self._store.upsert_file(entry.to_dict())
        self._cache_entry(entry)
        return entry

    def register_agent_output(
        self,
        canonical_path: str,
        original_name: str,
        parent_canonical: str | None = None,
        session_id: str | None = None,
        turn: int | None = None,
        tool_name: str | None = None,
        sheet_meta: list[dict] | None = None,
    ) -> FileEntry:
        """注册 agent 产出文件。"""
        file_type = _detect_file_type(canonical_path)
        now = _now_iso()
        parent_id = None
        if parent_canonical:
            parent = self._path_cache.get(parent_canonical)
            if parent:
                parent_id = parent.id

        existing = self._path_cache.get(canonical_path)
        entry = FileEntry(
            id=existing.id if existing else _new_id(),
            workspace=self._workspace_key,
            canonical_path=canonical_path,
            original_name=original_name,
            file_type=file_type,
            origin="agent_created",
            origin_session_id=session_id,
            origin_turn=turn,
            origin_tool=tool_name,
            parent_file_id=parent_id,
            sheet_meta=sheet_meta or [],
            created_at=existing.created_at if existing else now,
            updated_at=now,
        )

        # 尝试获取文件大小
        try:
            resolved = self._resolve(canonical_path)
            if resolved.exists():
                entry.size_bytes = resolved.stat().st_size
        except (ValueError, OSError):
            pass

        self._store.upsert_file(entry.to_dict())
        self._cache_entry(entry)

        self.record_event(
            entry.id, "created",
            session_id=session_id, turn=turn, tool_name=tool_name,
            details={"parent": parent_canonical},
        )
        return entry

    def register_backup(
        self,
        backup_path: str,
        parent_canonical: str,
        reason: str = "staging",
        session_id: str | None = None,
        turn: int | None = None,
        tool_name: str | None = None,
    ) -> FileEntry:
        """注册备份/副本文件。"""
        file_type = _detect_file_type(backup_path)
        now = _now_iso()
        parent = self._path_cache.get(parent_canonical)
        parent_id = parent.id if parent else None
        original_name = Path(backup_path).name

        existing = self._path_cache.get(backup_path)
        entry = FileEntry(
            id=existing.id if existing else _new_id(),
            workspace=self._workspace_key,
            canonical_path=backup_path,
            original_name=original_name,
            file_type=file_type,
            origin="backup",
            origin_session_id=session_id,
            origin_turn=turn,
            origin_tool=tool_name,
            parent_file_id=parent_id,
            created_at=existing.created_at if existing else now,
            updated_at=now,
        )
        self._store.upsert_file(entry.to_dict())
        self._cache_entry(entry)

        self.record_event(
            entry.id, "backed_up",
            session_id=session_id, turn=turn, tool_name=tool_name,
            details={"reason": reason, "parent": parent_canonical},
        )
        return entry

    def register_cow(
        self,
        cow_path: str,
        parent_canonical: str,
        session_id: str | None = None,
        turn: int | None = None,
    ) -> FileEntry:
        """注册 CoW 副本。"""
        file_type = _detect_file_type(cow_path)
        now = _now_iso()
        parent = self._path_cache.get(parent_canonical)
        parent_id = parent.id if parent else None

        existing = self._path_cache.get(cow_path)
        entry = FileEntry(
            id=existing.id if existing else _new_id(),
            workspace=self._workspace_key,
            canonical_path=cow_path,
            original_name=Path(cow_path).name,
            file_type=file_type,
            origin="cow_copy",
            origin_session_id=session_id,
            origin_turn=turn,
            parent_file_id=parent_id,
            is_active_cow=True,
            created_at=existing.created_at if existing else now,
            updated_at=now,
        )
        self._store.upsert_file(entry.to_dict())
        self._cache_entry(entry)

        # 添加 CoW 路径别名
        alias_id = _new_id()
        self._store.add_alias(alias_id, entry.id, "cow_path", cow_path)
        self._alias_cache[cow_path] = entry.id

        self.record_event(
            entry.id, "cow_created",
            session_id=session_id, turn=turn,
            details={"parent": parent_canonical},
        )
        return entry

    # ── 事件记录 ─────────────────────────────────────────────

    def record_event(
        self,
        file_id: str,
        event_type: str,
        *,
        session_id: str | None = None,
        turn: int | None = None,
        tool_name: str | None = None,
        details: dict | None = None,
    ) -> None:
        """记录文件生命周期事件。"""
        self._store.add_event({
            "id": _new_id(),
            "file_id": file_id,
            "event_type": event_type,
            "session_id": session_id,
            "turn": turn,
            "tool_name": tool_name,
            "details": details or {},
            "created_at": _now_iso(),
        })

    # ── 查询 ─────────────────────────────────────────────────

    def get_by_path(self, canonical_path: str) -> FileEntry | None:
        """按规范路径查询（优先缓存）。"""
        cached = self._path_cache.get(canonical_path)
        if cached:
            return cached
        row = self._store.get_by_path(self._workspace_key, canonical_path)
        if row:
            entry = FileEntry.from_dict(row)
            self._cache_entry(entry)
            return entry
        return None

    def get_by_alias(self, alias_value: str) -> FileEntry | None:
        """通过别名查找文件。"""
        file_id = self._alias_cache.get(alias_value)
        if file_id:
            path = self._id_to_path.get(file_id)
            if path and path in self._path_cache:
                return self._path_cache[path]
        row = self._store.find_by_alias(alias_value)
        if row:
            entry = FileEntry.from_dict(row)
            self._cache_entry(entry)
            return entry
        return None

    def get_by_id(self, file_id: str) -> FileEntry | None:
        """按 ID 查询文件。"""
        # 先查内存缓存
        path = self._id_to_path.get(file_id)
        if path and path in self._path_cache:
            return self._path_cache[path]
        row = self._store.get_by_id(file_id)
        if row:
            entry = FileEntry.from_dict(row)
            self._cache_entry(entry)
            return entry
        return None

    def list_all(self, include_deleted: bool = False) -> list[FileEntry]:
        """列出所有文件。"""
        if not include_deleted:
            return [
                e for e in self._path_cache.values()
                if e.deleted_at is None
            ]
        rows = self._store.list_all(self._workspace_key, include_deleted=True)
        return [FileEntry.from_dict(r) for r in rows]

    def get_children(self, file_id: str) -> list[FileEntry]:
        """获取文件的子文件（备份/副本）。"""
        rows = self._store.get_children(file_id)
        return [FileEntry.from_dict(r) for r in rows]

    def get_lineage(self, file_id: str) -> list[FileEntry]:
        """获取文件的祖先链（从当前到根）。"""
        result: list[FileEntry] = []
        seen: set[str] = set()
        current_id: str | None = file_id
        while current_id and current_id not in seen:
            seen.add(current_id)
            row = self._store.get_by_id(current_id)
            if not row:
                break
            entry = FileEntry.from_dict(row)
            result.append(entry)
            current_id = entry.parent_file_id
        return result

    def get_events(self, file_id: str) -> list[FileEvent]:
        """获取文件的事件历史。"""
        rows = self._store.get_events(file_id)
        return [FileEvent.from_dict(r) for r in rows]

    # ── 路径解析 ─────────────────────────────────────────────

    def resolve_for_tool(self, path_or_alias: str) -> str:
        """任何路径/别名 → 实际可用的规范路径。

        查找顺序：canonical_path 精确 → alias → original_name 模糊匹配。
        """
        # 1. canonical_path 精确匹配
        if path_or_alias in self._path_cache:
            return path_or_alias

        # 2. alias 匹配
        entry = self.get_by_alias(path_or_alias)
        if entry:
            return entry.canonical_path

        # 3. original_name 模糊匹配
        for e in self._path_cache.values():
            if e.deleted_at is None and e.original_name == path_or_alias:
                return e.canonical_path

        # 4. 直接返回原始路径
        return path_or_alias

    def resolve_for_display(self, canonical_path: str) -> str:
        """规范路径 → 用户友好的原始名。"""
        entry = self._path_cache.get(canonical_path)
        if entry:
            return entry.original_name
        return Path(canonical_path).name

    # ── 软删除 ───────────────────────────────────────────────

    def mark_deleted(self, canonical_path: str) -> None:
        """软删除（文件从磁盘消失时调用）。"""
        entry = self._path_cache.get(canonical_path)
        if entry:
            entry.deleted_at = _now_iso()
            entry.updated_at = entry.deleted_at
        self._store.soft_delete(self._workspace_key, canonical_path)
        # 不从缓存移除，保留 provenance

    # ── 添加别名 ─────────────────────────────────────────────

    def add_alias(
        self,
        file_id: str,
        alias_type: str,
        alias_value: str,
    ) -> None:
        """为文件添加别名。"""
        alias_id = _new_id()
        self._store.add_alias(alias_id, file_id, alias_type, alias_value)
        self._alias_cache[alias_value] = file_id

    # ── System Prompt 构建 ───────────────────────────────────

    def build_panorama(self, max_tokens: int = 1500) -> str:
        """构建文件全景图文本，用于 system prompt 注入。

        统一覆盖文件全景、上传文件提示与 CoW 路径提示。

        TODO: 使用 max_tokens 参数截断超长输出。
        """
        active = [
            e for e in self._path_cache.values()
            if e.deleted_at is None
        ]
        if not active:
            return ""

        # 分类
        user_files: list[FileEntry] = []
        backups: list[FileEntry] = []
        agent_outputs: list[FileEntry] = []

        for e in active:
            if e.origin in ("backup", "cow_copy", "staged"):
                backups.append(e)
            elif e.origin == "agent_created":
                agent_outputs.append(e)
            else:
                user_files.append(e)

        total = len(active)
        lines: list[str] = ["## 工作区文件全景"]

        if total <= _PANORAMA_FULL_THRESHOLD:
            self._panorama_full(lines, user_files, backups, agent_outputs)
        elif total <= _PANORAMA_COMPACT_THRESHOLD:
            self._panorama_compact(lines, user_files, backups, agent_outputs)
        else:
            self._panorama_summary(lines, user_files, backups, agent_outputs)

        lines.append("")
        lines.append("⚠️ 路径规则：读写操作使用「位置」列路径。向用户展示使用「文件」列名称。")
        lines.append("备份副本不可直接修改，操作原始文件即可。")

        return "\n".join(lines)

    def _panorama_full(
        self,
        lines: list[str],
        user_files: list[FileEntry],
        backups: list[FileEntry],
        agent_outputs: list[FileEntry],
    ) -> None:
        """完整模式：表格 + sheet 详情。"""
        if user_files:
            lines.append(f"\n### 用户文件 ({len(user_files)})")
            lines.append("| 文件 | 位置 | 来源 | 结构 |")
            lines.append("|---|---|---|---|")
            for e in sorted(user_files, key=lambda x: x.canonical_path):
                parent = str(Path(e.canonical_path).parent)
                loc = parent + "/" if parent != "." else "./"
                origin_str = self._format_origin(e)
                struct = self._format_structure(e)
                lines.append(f"| {e.original_name} | {loc} | {origin_str} | {struct} |")

        if backups:
            lines.append(f"\n### 备份与副本 ({len(backups)})")
            lines.append("| 副本 | 原始文件 | 类型 | 产生于 |")
            lines.append("|---|---|---|---|")
            for e in sorted(backups, key=lambda x: x.created_at):
                parent_name = self._get_parent_name(e)
                btype = "CoW保护" if e.origin == "cow_copy" else "事务备份"
                origin_str = self._format_origin(e)
                lines.append(f"| {e.canonical_path} | {parent_name} | {btype} | {origin_str} |")

        if agent_outputs:
            lines.append(f"\n### Agent 产出 ({len(agent_outputs)})")
            lines.append("| 文件 | 位置 | 派生自 | 产生于 |")
            lines.append("|---|---|---|---|")
            for e in sorted(agent_outputs, key=lambda x: x.created_at):
                parent_name = self._get_parent_name(e)
                parent_dir = str(Path(e.canonical_path).parent)
                loc = parent_dir + "/" if parent_dir != "." else "./"
                origin_str = self._format_origin(e)
                lines.append(f"| {e.original_name} | {loc} | {parent_name} | {origin_str} |")

    def _panorama_compact(
        self,
        lines: list[str],
        user_files: list[FileEntry],
        backups: list[FileEntry],
        agent_outputs: list[FileEntry],
    ) -> None:
        """紧凑模式：文件列表。"""
        if user_files:
            lines.append(f"\n### 用户文件 ({len(user_files)})")
            for e in sorted(user_files, key=lambda x: x.canonical_path):
                sheets = ""
                if e.sheet_meta:
                    sheet_names = [s.get("name", "") for s in e.sheet_meta]
                    sheets = f" [{', '.join(sheet_names)}]"
                lines.append(f"- `{e.canonical_path}`{sheets}")

        if backups:
            lines.append(f"\n### 备份与副本 ({len(backups)})")
            for e in sorted(backups, key=lambda x: x.created_at):
                parent_name = self._get_parent_name(e)
                lines.append(f"- `{e.canonical_path}` ← {parent_name}")

        if agent_outputs:
            lines.append(f"\n### Agent 产出 ({len(agent_outputs)})")
            for e in sorted(agent_outputs, key=lambda x: x.created_at):
                lines.append(f"- `{e.canonical_path}`")

    def _panorama_summary(
        self,
        lines: list[str],
        user_files: list[FileEntry],
        backups: list[FileEntry],
        agent_outputs: list[FileEntry],
    ) -> None:
        """统计摘要模式。"""
        lines.append(f"\n共 {len(user_files)} 个用户文件, "
                      f"{len(backups)} 个备份/副本, "
                      f"{len(agent_outputs)} 个 agent 产出")

        # 热点目录
        dir_counts: dict[str, int] = {}
        for e in user_files:
            parent = str(Path(e.canonical_path).parent)
            dir_counts[parent] = dir_counts.get(parent, 0) + 1
        if dir_counts:
            top_dirs = sorted(dir_counts.items(), key=lambda x: -x[1])[:10]
            lines.append("热点目录：")
            for d, count in top_dirs:
                label_ = _dir_label(d) if d != "." else ""
                suffix = f"（{label_}）" if label_ else ""
                dn = d if d != "." else "(根目录)"
                lines.append(f"  - `{dn}/` ({count} 个文件){suffix}")

    def _format_origin(self, e: FileEntry) -> str:
        """格式化来源信息。"""
        parts: list[str] = []
        if e.origin == "uploaded":
            parts.append("上传")
        elif e.origin == "scan":
            parts.append("扫描")
        elif e.origin == "agent_created":
            parts.append("agent")
        elif e.origin == "backup":
            parts.append("备份")
        elif e.origin == "cow_copy":
            parts.append("CoW")
        if e.origin_turn is not None:
            parts.append(f"T{e.origin_turn}")
        if e.origin_tool:
            parts.append(e.origin_tool)
        return "(" + " ".join(parts) + ")" if parts else ""

    def _format_structure(self, e: FileEntry) -> str:
        """格式化文件结构（sheet 详情或文件类型）。"""
        if e.file_type in ("excel", "csv") and e.sheet_meta:
            parts: list[str] = []
            for s in e.sheet_meta:
                name = s.get("name", "")
                rows = s.get("rows", 0)
                cols = s.get("columns", 0)
                parts.append(f"{name}({rows}×{cols})")
            return f"{len(e.sheet_meta)}表: " + ", ".join(parts)
        if e.file_type == "image":
            return f"图片 {self._format_size(e.size_bytes)}"
        if e.size_bytes:
            return self._format_size(e.size_bytes)
        return e.file_type

    def _get_parent_name(self, e: FileEntry) -> str:
        """获取父文件的用户友好名。"""
        if not e.parent_file_id:
            return "-"
        path = self._id_to_path.get(e.parent_file_id)
        if path:
            parent = self._path_cache.get(path)
            if parent:
                return parent.original_name
        return "-"

    @staticmethod
    def _format_size(size_bytes: int) -> str:
        if size_bytes < 1024:
            return f"{size_bytes}B"
        if size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.0f}KB"
        return f"{size_bytes / (1024 * 1024):.1f}MB"

    # ── 扫描 ─────────────────────────────────────────────────

    def scan_workspace(
        self,
        *,
        max_files: int = 500,
        header_scan_rows: int = 5,
        excel_only: bool = False,
    ) -> ScanResult:
        """递归扫描工作区，注册/更新文件到 registry。

        扫描范围：工作区根目录 + uploads/ + outputs/ 下的所有文件。
        Excel 文件额外提取 sheet 元数据（表名、行列数、表头）。

        Args:
            max_files: 最大文件数量限制。
            header_scan_rows: Excel 文件表头探测的行数。
            excel_only: 仅扫描 Excel 文件（兼容旧 manifest 行为）。
        """
        start_ts = time.monotonic()
        result = ScanResult()

        collected = self._collect_file_paths(max_files, excel_only=excel_only)
        current_rel_paths: set[str] = set()

        for fp in collected:
            try:
                stat = fp.stat()
            except OSError:
                continue

            rel_path = self._to_rel(fp)
            current_rel_paths.add(rel_path)
            mtime_ns = stat.st_mtime_ns

            existing = self._path_cache.get(rel_path)
            if existing and existing.mtime_ns == mtime_ns and existing.size_bytes == stat.st_size:
                result.cache_hits += 1
                continue

            # Excel 文件：扫描 sheet 元数据
            file_type = _detect_file_type(rel_path)
            sheet_meta: list[dict] = []
            if file_type in ("excel",):
                try:
                    sheet_meta = self._scan_file_sheets(fp, header_scan_rows)
                except Exception:
                    logger.debug("扫描文件 %s 失败", fp, exc_info=True)

            if existing:
                result.updated_files += 1
            else:
                result.new_files += 1

            self.register_from_scan(
                canonical_path=rel_path,
                original_name=fp.name,
                size_bytes=stat.st_size,
                mtime_ns=mtime_ns,
                file_type=file_type,
                sheet_meta=sheet_meta,
            )

        # 软删除磁盘已不存在的文件（仅 scan origin 的文件）
        for path, entry in list(self._path_cache.items()):
            if (
                entry.origin == "scan"
                and entry.deleted_at is None
                and path not in current_rel_paths
            ):
                self.mark_deleted(path)
                result.deleted_files += 1

        result.total_files = len(current_rel_paths)
        result.scan_duration_ms = int((time.monotonic() - start_ts) * 1000)

        logger.info(
            "FileRegistry scan: %d 文件 (新增 %d, 更新 %d, 删除 %d, 缓存命中 %d), 耗时 %dms",
            result.total_files, result.new_files, result.updated_files,
            result.deleted_files, result.cache_hits, result.scan_duration_ms,
        )
        return result

    def scan_uploads(self, *, header_scan_rows: int = 5) -> ScanResult:
        """专门扫描 uploads/ 目录并注册未跟踪的上传文件。

        区别于 scan_workspace：仅扫描 uploads/ 子目录，
        且对已通过 register_upload 注册的文件只做增量更新。
        """
        uploads_dir = self._workspace_root / "uploads"
        if not uploads_dir.exists():
            return ScanResult()

        start_ts = time.monotonic()
        result = ScanResult()

        for walk_root, dirs, files in os.walk(uploads_dir):
            dirs[:] = [d for d in dirs if d not in _SKIP_DIRS]
            for name in files:
                if name.startswith((".", "~$")):
                    continue
                fp = Path(walk_root, name)
                try:
                    stat = fp.stat()
                except OSError:
                    continue

                rel_path = self._to_rel(fp)
                mtime_ns = stat.st_mtime_ns

                existing = self._path_cache.get(rel_path)
                if existing:
                    if existing.mtime_ns == mtime_ns and existing.size_bytes == stat.st_size:
                        result.cache_hits += 1
                    elif existing.mtime_ns == 0 and existing.size_bytes == stat.st_size:
                        # register_upload 未设 mtime → 补填 mtime，视为缓存命中
                        existing.mtime_ns = mtime_ns
                        self._store.upsert_file(existing.to_dict())
                        self._cache_entry(existing)
                        result.cache_hits += 1
                    else:
                        # 已注册但内容变化 → 更新 mtime/size
                        existing.mtime_ns = mtime_ns
                        existing.size_bytes = stat.st_size
                        existing.updated_at = _now_iso()
                        self._store.upsert_file(existing.to_dict())
                        self._cache_entry(existing)
                        result.updated_files += 1
                    result.total_files += 1
                    continue

                # 未注册的上传文件 → 按 scan origin 注册
                file_type = _detect_file_type(rel_path)
                sheet_meta: list[dict] = []
                if file_type in ("excel",):
                    try:
                        sheet_meta = self._scan_file_sheets(fp, header_scan_rows)
                    except Exception:
                        logger.debug("扫描上传文件 %s 失败", fp, exc_info=True)

                self.register_from_scan(
                    canonical_path=rel_path,
                    original_name=name,
                    size_bytes=stat.st_size,
                    mtime_ns=mtime_ns,
                    file_type=file_type,
                    sheet_meta=sheet_meta,
                )
                result.new_files += 1
                result.total_files += 1

        result.scan_duration_ms = int((time.monotonic() - start_ts) * 1000)
        return result

    def _collect_file_paths(self, max_files: int, *, excel_only: bool = False) -> list[Path]:
        """递归收集工作区中的文件路径。"""
        root = self._workspace_root
        paths: list[Path] = []

        for walk_root, dirs, files in os.walk(root):
            dirs[:] = [d for d in dirs if d not in _SKIP_DIRS]
            for name in files:
                if name.startswith((".", "~$")):
                    continue
                _, ext = os.path.splitext(name)
                ext_lower = ext.lower()

                if excel_only:
                    if ext_lower not in _EXCEL_EXTENSIONS:
                        continue
                else:
                    # 跳过完全无用的二进制文件
                    if ext_lower in _SKIP_EXTENSIONS:
                        continue

                paths.append(Path(walk_root, name))
                if len(paths) >= max_files:
                    paths.sort(key=lambda p: str(p.relative_to(root)).lower())
                    return paths

        paths.sort(key=lambda p: str(p.relative_to(root)).lower())
        return paths

    @staticmethod
    def _scan_file_sheets(fp: Path, header_scan_rows: int) -> list[dict]:
        """用 openpyxl 扫描单个 Excel 文件的 sheet 元数据。"""
        # CSV 文件不走 openpyxl
        if fp.suffix.lower() == ".csv":
            import csv as _csv

            _enc = "utf-8"
            for _try_enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
                try:
                    with open(fp, "r", encoding=_try_enc) as _f:
                        _f.read(4096)
                    _enc = _try_enc
                    break
                except (UnicodeDecodeError, LookupError):
                    continue

            with open(fp, "r", encoding=_enc, newline="") as _f:
                reader = _csv.reader(_f)
                rows_raw: list[list[str]] = []
                for row in reader:
                    rows_raw.append(row)
                    if len(rows_raw) >= header_scan_rows + 1:
                        break

            total_rows = len(rows_raw)
            total_cols = max((len(r) for r in rows_raw), default=0)
            headers: list[str] = []
            if rows_raw:
                best_idx = 0
                best_score = -1
                for idx, r in enumerate(rows_raw):
                    non_empty = [v for v in r if v and v.strip()]
                    score = len(non_empty) * 2
                    if score > best_score:
                        best_score = score
                        best_idx = idx
                headers = [v.strip() for v in rows_raw[best_idx] if v and v.strip()]

            return [{
                "name": "Sheet1",
                "rows": total_rows,
                "columns": total_cols,
                "headers": headers,
            }]

        from openpyxl import load_workbook

        sheets: list[dict] = []
        wb = load_workbook(fp, read_only=True, data_only=True)
        try:
            for sn in wb.sheetnames:
                ws = wb[sn]
                total_rows = ws.max_row or 0
                total_cols = ws.max_column or 0

                headers: list[str] = []
                if total_rows > 0:
                    scan_limit = min(header_scan_rows, total_rows)
                    rows_raw: list[list[Any]] = []
                    for row in ws.iter_rows(
                        min_row=1,
                        max_row=scan_limit,
                        min_col=1,
                        max_col=min(total_cols, 30),
                        values_only=True,
                    ):
                        rows_raw.append(list(row))

                    if rows_raw:
                        best_idx = 0
                        best_score = -1
                        for idx, r in enumerate(rows_raw):
                            non_empty = [v for v in r if v is not None and str(v).strip()]
                            str_count = sum(1 for v in non_empty if isinstance(v, str))
                            score = str_count * 2 + len(non_empty)
                            if score > best_score:
                                best_score = score
                                best_idx = idx
                        header_row = rows_raw[best_idx]
                        headers = [
                            str(v).strip()
                            for v in header_row
                            if v is not None and str(v).strip()
                        ]

                sheets.append({
                    "name": sn,
                    "rows": total_rows,
                    "columns": total_cols,
                    "headers": headers,
                })
        finally:
            wb.close()
        return sheets

    # ── Staging / CoW / Checkpoint 委托层 ────────────────────

    @property
    def has_versions(self) -> bool:
        """是否启用了版本管理层。"""
        return self._fvm is not None

    @property
    def fvm(self) -> Any:
        """底层 FileVersionManager（仅 enable_versions=True 时可用）。"""
        return self._fvm

    # -- Staging -------------------------------------------------------

    def stage_for_write(
        self,
        file_path: str,
        *,
        ref_id: str = "",
        scope: str = "all",
    ) -> str:
        """确保文件有原始快照，返回 staged 副本路径。

        委托 FVM.stage_for_write() + 注册备份到 registry 元数据。
        """
        if self._fvm is None:
            return str(self._resolve(file_path))

        staged_path = self._fvm.stage_for_write(
            file_path, ref_id=ref_id, scope=scope,
        )

        # 在 registry 元数据中注册备份副本
        try:
            resolved = self._resolve(file_path)
            rel = self._to_rel(resolved)
            staged_rel = self._to_rel(Path(staged_path))
            if staged_rel != rel:
                self.register_backup(
                    canonical_path=staged_rel,
                    parent_path=rel,
                )
        except Exception:
            logger.debug("stage_for_write registry 注册失败", exc_info=True)

        return staged_path

    def get_staged_path(self, file_path: str) -> str | None:
        """查询文件的 staged 副本路径。"""
        if self._fvm is None:
            return None
        return self._fvm.get_staged_path(file_path)

    def commit_staged(self, file_path: str) -> dict[str, str] | None:
        """将单个 staged 文件提交回原位置。"""
        if self._fvm is None:
            return None
        return self._fvm.commit_staged(file_path)

    def commit_all_staged(self) -> list[dict[str, str]]:
        """将所有 staged 文件提交回原位置。"""
        if self._fvm is None:
            return []
        return self._fvm.commit_all_staged()

    def discard_staged(self, file_path: str) -> bool:
        """丢弃单个 staged 文件。"""
        if self._fvm is None:
            return False
        return self._fvm.discard_staged(file_path)

    def discard_all_staged(self) -> int:
        """丢弃所有 staged 文件。"""
        if self._fvm is None:
            return 0
        return self._fvm.discard_all_staged()

    def list_staged(self) -> list[dict[str, str]]:
        """列出所有活跃的 staged 文件。"""
        if self._fvm is None:
            return []
        return self._fvm.list_staged()

    def staged_file_map(self) -> dict[str, str]:
        """返回 original_abs → staged_abs 映射。"""
        if self._fvm is None:
            return {}
        return self._fvm.staged_file_map()

    def has_staging(self, file_path: str) -> bool:
        """检查文件是否有活跃的 staging 条目。"""
        if self._fvm is None:
            return False
        return self._fvm.has_staging(file_path)

    def remove_staging_for_path(self, file_path: str) -> bool:
        """移除指定文件的 staging 条目。"""
        if self._fvm is None:
            return False
        return self._fvm.remove_staging_for_path(file_path)

    def rename_staging_path(self, old_path: str, new_path: str) -> bool:
        """重命名 staging 条目的原始路径。"""
        if self._fvm is None:
            return False
        return self._fvm.rename_staging_path(old_path, new_path)

    def prune_stale_staging(self) -> int:
        """移除 staged 物理文件已不存在的条目。"""
        if self._fvm is None:
            return 0
        return self._fvm.prune_stale_staging()

    # -- CoW -----------------------------------------------------------

    def register_cow_mapping(self, src_rel: str, dst_rel: str) -> None:
        """注册 CoW 路径映射。

        委托 FVM + 注册到 registry 元数据。
        """
        if self._fvm is not None:
            self._fvm.register_cow_mapping(src_rel, dst_rel)

        # 在 registry 中注册 CoW 条目
        try:
            self.register_cow(
                canonical_path=dst_rel,
                source_path=src_rel,
            )
        except Exception:
            logger.debug("register_cow_mapping registry 注册失败", exc_info=True)

    def lookup_cow_redirect(self, rel_path: str) -> str | None:
        """查找相对路径是否有 CoW/staging 副本。"""
        if self._fvm is not None:
            return self._fvm.lookup_cow_redirect(rel_path)
        return None

    def get_cow_mappings(self) -> dict[str, str]:
        """返回当前所有活跃的 CoW 映射（src_rel → dst_rel）。"""
        if self._fvm is not None:
            return dict(getattr(self._fvm, "_cow_registry", {}))
        return {}

    # -- Checkpoint / Version ------------------------------------------

    def checkpoint_file(
        self,
        file_path: str,
        *,
        reason: str = "staging",
        ref_id: str = "",
    ) -> Any:
        """为文件创建版本快照。返回 FileVersion 或 None（去重）。"""
        if self._fvm is None:
            return None
        return self._fvm.checkpoint(file_path, reason=reason, ref_id=ref_id)

    def create_turn_checkpoint(
        self,
        turn_number: int,
        dirty_files: list[str],
        tool_names: list[str] | None = None,
    ) -> Any:
        """对 dirty_files 做快照，记录为一个轮次 checkpoint。"""
        if self._fvm is None:
            return None
        return self._fvm.create_turn_checkpoint(
            turn_number, dirty_files, tool_names=tool_names,
        )

    def rollback_to_turn(self, turn_number: int) -> list[str]:
        """回退到指定轮次之前的状态。返回被恢复的文件路径列表。"""
        if self._fvm is None:
            return []
        return self._fvm.rollback_to_turn(turn_number)

    def list_turn_checkpoints(self) -> list[Any]:
        """返回所有轮次 checkpoint（时间正序）。"""
        if self._fvm is None:
            return []
        return self._fvm.list_turn_checkpoints()

    def restore_to_original(self, file_path: str) -> bool:
        """将文件恢复到最早的原始版本。"""
        if self._fvm is None:
            return False
        return self._fvm.restore_to_original(file_path)

    def invalidate_undo(self, rel_paths: set[str]) -> int:
        """标记指定文件的版本链为不可恢复。"""
        if self._fvm is None:
            return 0
        return self._fvm.invalidate_undo(rel_paths)

    def get_version_original(self, file_path: str) -> Any:
        """获取文件的最早版本。"""
        if self._fvm is None:
            return None
        return self._fvm.get_original(file_path)

    def get_version_latest(self, file_path: str) -> Any:
        """获取文件的最新版本。"""
        if self._fvm is None:
            return None
        return self._fvm.get_latest(file_path)

    def list_versions(self, file_path: str) -> list[Any]:
        """获取文件的完整版本链。"""
        if self._fvm is None:
            return []
        return self._fvm.list_versions(file_path)

    def list_all_tracked(self) -> list[str]:
        """返回所有有版本记录的文件相对路径。"""
        if self._fvm is None:
            return []
        return self._fvm.list_all_tracked()

    def gc_versions(self, max_age_seconds: float = 3600) -> int:
        """清理过期版本快照。"""
        if self._fvm is None:
            return 0
        return self._fvm.gc(max_age_seconds)
