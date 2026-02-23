"""WorkspaceManifest — 工作区 Excel 文件轻量级元数据清单。

会话启动时递归扫描工作区所有 Excel 文件的 sheet 名称、行列数、列头，
构建紧凑的 JSON 清单用于：
1. 注入 system prompt，让 agent 首轮即知工作区全貌；
2. 供 inspect_excel_files 的 search 功能使用缓存加速。

设计原则：
- 仅读取元信息（sheet names + 前 2 行用于表头识别），不加载数据
- openpyxl read_only 模式，内存占用极低
- 基于 mtime 的增量更新，避免重复 IO
"""

from __future__ import annotations

import json
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any

from excelmanus.excel_extensions import EXCEL_EXTENSIONS
from excelmanus.logger import get_logger

if TYPE_CHECKING:
    from excelmanus.database import Database

logger = get_logger("workspace_manifest")

# 递归扫描时跳过的噪音目录
_SKIP_DIRS: frozenset[str] = frozenset({
    ".git", ".venv", "node_modules", "__pycache__",
    ".worktrees", "dist", "build", "outputs",
})

# Manifest 摘要注入 system prompt 的阈值
_INJECT_FULL_THRESHOLD = 20      # ≤20 文件：完整注入
_INJECT_COMPACT_THRESHOLD = 100  # 20-100 文件：紧凑注入
# >100 文件：仅注入统计摘要


def _excel_glob_patterns() -> tuple[str, ...]:
    """将共享扩展名集合转换为稳定的 glob 列表。"""
    return tuple(f"*{suffix}" for suffix in sorted(EXCEL_EXTENSIONS))


@dataclass
class SheetMeta:
    """单个 sheet 的元数据。"""
    name: str
    rows: int
    columns: int
    headers: list[str] = field(default_factory=list)


@dataclass
class ExcelFileMeta:
    """单个 Excel 文件的元数据。"""
    path: str          # 相对于 workspace_root 的路径
    name: str          # 文件名
    size_bytes: int
    modified_ts: float  # mtime 时间戳
    sheets: list[SheetMeta] = field(default_factory=list)


@dataclass
class WorkspaceManifest:
    """工作区 Excel 文件清单。"""
    workspace_root: str
    scan_time: str          # ISO 格式扫描时间
    scan_duration_ms: int   # 扫描耗时（毫秒）
    total_files: int
    files: list[ExcelFileMeta] = field(default_factory=list)

    # ── 内部状态 ──
    _mtime_cache: dict[str, float] = field(default_factory=dict, repr=False)

    def get_system_prompt_summary(self) -> str:
        """生成用于注入 system prompt 的工作区概览文本。

        根据文件数量自动选择详细度：
        - ≤20 文件：完整列出每个文件的 sheet 名 + 行列数 + 列头
        - 20-100 文件：仅列出文件路径 + sheet 名列表
        - >100 文件：仅统计摘要
        """
        if not self.files:
            return ""

        lines: list[str] = [
            "## 工作区 Excel 文件概览",
            f"共 {self.total_files} 个 Excel 文件"
            f"（扫描于 {self.scan_time}，耗时 {self.scan_duration_ms}ms）：",
        ]

        if self.total_files <= _INJECT_FULL_THRESHOLD:
            # 完整模式
            for fm in self.files:
                sheet_parts: list[str] = []
                for sm in fm.sheets:
                    header_hint = ""
                    if sm.headers:
                        cols_str = ", ".join(sm.headers[:6])
                        if len(sm.headers) > 6:
                            cols_str += f" +{len(sm.headers) - 6}列"
                        header_hint = f" [{cols_str}]"
                    sheet_parts.append(f"{sm.name}({sm.rows}×{sm.columns}){header_hint}")
                sheets_str = " | ".join(sheet_parts) if sheet_parts else "(空)"
                lines.append(f"- `{fm.path}` → {sheets_str}")

        elif self.total_files <= _INJECT_COMPACT_THRESHOLD:
            # 紧凑模式：文件路径 + sheet 名列表
            for fm in self.files:
                sheet_names = [sm.name for sm in fm.sheets]
                sheets_str = ", ".join(sheet_names) if sheet_names else "(空)"
                lines.append(f"- `{fm.path}` → [{sheets_str}]")

        else:
            # 统计摘要模式
            total_sheets = sum(len(fm.sheets) for fm in self.files)
            # 按目录分组统计
            dir_counts: dict[str, int] = {}
            for fm in self.files:
                parent = str(Path(fm.path).parent)
                if parent == ".":
                    parent = "(根目录)"
                dir_counts[parent] = dir_counts.get(parent, 0) + 1
            top_dirs = sorted(dir_counts.items(), key=lambda x: -x[1])[:10]
            lines.append(f"共 {total_sheets} 个工作表")
            lines.append("热点目录：")
            for d, count in top_dirs:
                lines.append(f"  - `{d}/` ({count} 个文件)")
            lines.append(
                "使用 `inspect_excel_files(search=\"关键词\")` 可按文件名或 sheet 名快速定位。"
            )

        return "\n".join(lines)


def _sheets_to_json(sheets: list[SheetMeta]) -> str:
    """将 SheetMeta 列表序列化为 JSON 字符串。"""
    return json.dumps(
        [{"name": s.name, "rows": s.rows, "columns": s.columns, "headers": s.headers} for s in sheets],
        ensure_ascii=False,
    )


def _sheets_from_json(raw: str) -> list[SheetMeta]:
    """从 JSON 字符串反序列化为 SheetMeta 列表。"""
    try:
        items = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []
    return [
        SheetMeta(
            name=item.get("name", ""),
            rows=item.get("rows", 0),
            columns=item.get("columns", 0),
            headers=item.get("headers", []),
        )
        for item in items
        if isinstance(item, dict)
    ]


def _collect_excel_paths(root: Path, max_files: int) -> list[Path]:
    """递归收集工作区中的 Excel 文件路径。"""
    excel_paths: list[Path] = []
    for ext in _excel_glob_patterns():
        for p in root.rglob(ext):
            if p.name.startswith((".", "~$")):
                continue
            rel_parts = p.relative_to(root).parts[:-1]
            if any(part in _SKIP_DIRS for part in rel_parts):
                continue
            excel_paths.append(p)
            if len(excel_paths) >= max_files:
                break
        if len(excel_paths) >= max_files:
            break
    excel_paths.sort(key=lambda p: str(p.relative_to(root)).lower())
    return excel_paths


def _scan_file_sheets(fp: Path, header_scan_rows: int) -> list[SheetMeta]:
    """用 openpyxl 扫描单个 Excel 文件的 sheet 元数据。"""
    from openpyxl import load_workbook

    sheets: list[SheetMeta] = []
    wb = load_workbook(fp, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0

            headers: list[str] = []
            if total_rows > 0:
                scan_limit = min(header_scan_rows, total_rows)
                rows_raw: list[list[Any]] = []
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=scan_limit,
                    min_col=1,
                    max_col=min(total_cols, 30),
                    values_only=True,
                ):
                    rows_raw.append(list(row))

                if rows_raw:
                    best_idx = 0
                    best_score = -1
                    for idx, r in enumerate(rows_raw):
                        non_empty = [v for v in r if v is not None and str(v).strip()]
                        str_count = sum(1 for v in non_empty if isinstance(v, str))
                        score = str_count * 2 + len(non_empty)
                        if score > best_score:
                            best_score = score
                            best_idx = idx
                    header_row = rows_raw[best_idx]
                    headers = [
                        str(v).strip()
                        for v in header_row
                        if v is not None and str(v).strip()
                    ]

            sheets.append(SheetMeta(
                name=sn, rows=total_rows, columns=total_cols, headers=headers,
            ))
    finally:
        wb.close()
    return sheets


def build_manifest(
    workspace_root: str,
    *,
    max_files: int = 500,
    header_scan_rows: int = 5,
    silent: bool = False,
    database: "Database | None" = None,
) -> WorkspaceManifest:
    """递归扫描工作区，构建 Excel 文件元数据清单。

    有 DB 时优先从缓存加载，仅扫描 mtime/size 变更的文件（跨会话增量）。
    无 DB 时行为与旧版完全一致。

    Args:
        workspace_root: 工作区根目录路径。
        max_files: 最多扫描文件数，默认 500。
        header_scan_rows: 每个 sheet 扫描前 N 行用于表头识别。
        silent: 静默模式（仅 DEBUG 日志）。
        database: 可选统一数据库实例。

    Returns:
        WorkspaceManifest 实例。
    """
    root = Path(workspace_root).resolve()
    workspace_key = str(root)
    start_ts = time.monotonic()

    # 加载 DB 缓存（如果可用）
    db_cache: dict[str, dict[str, Any]] = {}
    manifest_store = None
    if database is not None:
        try:
            from excelmanus.stores.manifest_store import ManifestStore
            manifest_store = ManifestStore(database)
            db_cache = manifest_store.load_cached(workspace_key)
        except Exception:
            logger.debug("加载 manifest DB 缓存失败，回退全量扫描", exc_info=True)

    excel_paths = _collect_excel_paths(root, max_files)

    files: list[ExcelFileMeta] = []
    mtime_cache: dict[str, float] = {}
    db_writes: list[dict[str, Any]] = []  # 需写回 DB 的新/变更记录
    cache_hits = 0

    for fp in excel_paths:
        try:
            stat = fp.stat()
        except OSError:
            continue

        rel_path = str(fp.relative_to(root))
        mtime_cache[rel_path] = stat.st_mtime
        mtime_ns = stat.st_mtime_ns

        # DB 缓存命中检查：mtime_ns + size_bytes 均匹配则直接复用
        cached = db_cache.get(rel_path)
        if (
            cached is not None
            and cached["mtime_ns"] == mtime_ns
            and cached["size_bytes"] == stat.st_size
        ):
            sheets = _sheets_from_json(cached["sheets_json"])
            files.append(ExcelFileMeta(
                path=rel_path,
                name=fp.name,
                size_bytes=stat.st_size,
                modified_ts=stat.st_mtime,
                sheets=sheets,
            ))
            cache_hits += 1
            continue

        # 缓存未命中 → openpyxl 扫描
        sheets = []
        try:
            sheets = _scan_file_sheets(fp, header_scan_rows)
        except Exception as exc:  # noqa: BLE001
            logger.debug("扫描文件 %s 失败: %s", fp, exc)

        files.append(ExcelFileMeta(
            path=rel_path,
            name=fp.name,
            size_bytes=stat.st_size,
            modified_ts=stat.st_mtime,
            sheets=sheets,
        ))

        # 记录需要写回 DB 的条目
        if manifest_store is not None:
            db_writes.append({
                "path": rel_path,
                "name": fp.name,
                "size_bytes": stat.st_size,
                "mtime_ns": mtime_ns,
                "sheets_json": _sheets_to_json(sheets),
            })

    # 写回 DB 缓存 & 清理陈旧记录
    if manifest_store is not None:
        try:
            if db_writes:
                manifest_store.upsert_batch(workspace_key, db_writes)
            current_rel_paths = {str(fp.relative_to(root)) for fp in excel_paths}
            manifest_store.remove_stale(workspace_key, current_rel_paths)
        except Exception:
            logger.debug("manifest DB 缓存写回失败", exc_info=True)

    elapsed_ms = int((time.monotonic() - start_ts) * 1000)
    scan_time = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    manifest = WorkspaceManifest(
        workspace_root=workspace_key,
        scan_time=scan_time,
        scan_duration_ms=elapsed_ms,
        total_files=len(files),
        files=files,
        _mtime_cache=mtime_cache,
    )

    scanned = len(files) - cache_hits
    _log = logger.debug if silent else logger.info
    _log(
        "Workspace manifest 构建完成: %d 文件 (缓存命中 %d, 扫描 %d), 耗时 %dms",
        len(files), cache_hits, scanned, elapsed_ms,
    )
    return manifest


def refresh_manifest(
    manifest: WorkspaceManifest,
    *,
    max_files: int = 500,
    header_scan_rows: int = 5,
    database: "Database | None" = None,
) -> WorkspaceManifest:
    """增量更新 Manifest：仅重新扫描 mtime 变化或新增的文件。

    有 DB 时同步更新缓存。

    Args:
        manifest: 现有 Manifest。
        max_files: 最多扫描文件数。
        header_scan_rows: 表头识别行数。
        database: 可选统一数据库实例。

    Returns:
        更新后的 WorkspaceManifest（新实例）。
    """
    root = Path(manifest.workspace_root).resolve()
    workspace_key = str(root)
    start_ts = time.monotonic()

    manifest_store = None
    if database is not None:
        try:
            from excelmanus.stores.manifest_store import ManifestStore
            manifest_store = ManifestStore(database)
        except Exception:
            logger.debug("ManifestStore 初始化失败", exc_info=True)

    current_paths = _collect_excel_paths(root, max_files)

    old_by_path: dict[str, ExcelFileMeta] = {
        fm.path: fm for fm in manifest.files
    }
    old_mtime = manifest._mtime_cache

    files: list[ExcelFileMeta] = []
    new_mtime_cache: dict[str, float] = {}
    changed_count = 0
    db_writes: list[dict[str, Any]] = []

    for fp in current_paths:
        try:
            stat = fp.stat()
        except OSError:
            continue

        rel_path = str(fp.relative_to(root))
        new_mtime_cache[rel_path] = stat.st_mtime

        # mtime 未变且已有内存缓存 → 复用
        if rel_path in old_by_path and old_mtime.get(rel_path) == stat.st_mtime:
            files.append(old_by_path[rel_path])
            continue

        # 需要重新扫描
        changed_count += 1
        sheets: list[SheetMeta] = []
        try:
            sheets = _scan_file_sheets(fp, header_scan_rows)
        except Exception as exc:  # noqa: BLE001
            logger.debug("增量扫描文件 %s 失败: %s", fp, exc)

        files.append(ExcelFileMeta(
            path=rel_path,
            name=fp.name,
            size_bytes=stat.st_size,
            modified_ts=stat.st_mtime,
            sheets=sheets,
        ))

        if manifest_store is not None:
            db_writes.append({
                "path": rel_path,
                "name": fp.name,
                "size_bytes": stat.st_size,
                "mtime_ns": stat.st_mtime_ns,
                "sheets_json": _sheets_to_json(sheets),
            })

    # 写回 DB 缓存 & 清理
    if manifest_store is not None:
        try:
            if db_writes:
                manifest_store.upsert_batch(workspace_key, db_writes)
            current_rel_paths = {str(fp.relative_to(root)) for fp in current_paths}
            manifest_store.remove_stale(workspace_key, current_rel_paths)
        except Exception:
            logger.debug("manifest DB 缓存写回失败", exc_info=True)

    elapsed_ms = int((time.monotonic() - start_ts) * 1000)
    scan_time = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    logger.info(
        "Workspace manifest 增量更新: %d 文件 (变更 %d), 耗时 %dms",
        len(files), changed_count, elapsed_ms,
    )

    return WorkspaceManifest(
        workspace_root=manifest.workspace_root,
        scan_time=scan_time,
        scan_duration_ms=elapsed_ms,
        total_files=len(files),
        files=files,
        _mtime_cache=new_mtime_cache,
    )
