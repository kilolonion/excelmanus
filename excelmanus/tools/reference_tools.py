"""引用关系图工具：get_reference_map, trace_references, get_impact_analysis。"""
from __future__ import annotations

import json
from dataclasses import asdict
from pathlib import Path
from typing import Any

from excelmanus.reference_graph.cache import RefCache
from excelmanus.reference_graph.formula_parser import FormulaRefExtractor
from excelmanus.reference_graph.models import CellNode, CellRef, WorkbookRefIndex
from excelmanus.reference_graph.scanner import Tier1Scanner
from excelmanus.tools.registry import ToolDef

_workspace_root: str | None = None
_cache = RefCache()
_scanner = Tier1Scanner()
_extractor = FormulaRefExtractor()


def init_guard(workspace_root: str) -> None:
    """设置工具的工作空间根目录。"""
    global _workspace_root
    _workspace_root = workspace_root


def _resolve_path(file_path: str) -> str:
    """将相对路径解析为绝对路径。"""
    if _workspace_root and not Path(file_path).is_absolute():
        return str(Path(_workspace_root) / file_path)
    return file_path


def _ensure_index(file_path: str) -> WorkbookRefIndex:
    """确保 Tier 1 索引已缓存。"""
    abs_path = _resolve_path(file_path)
    cached = _cache.get_tier1(abs_path)
    if cached is not None:
        return cached
    index = _scanner.scan(abs_path)
    _cache.put_tier1(abs_path, index)
    return index


def _error_json(message: str) -> str:
    return json.dumps({"status": "error", "message": message}, ensure_ascii=False)


def _parse_target(target: str) -> tuple[str | None, str]:
    """解析 'Sheet!Cell' 格式的目标。"""
    if "!" in target:
        sheet, addr = target.split("!", 1)
        return sheet.strip("'"), addr
    return None, target


def get_reference_map(file_path: str, detail: str = "summary") -> str:
    """获取工作簿引用全景图。"""
    try:
        index = _ensure_index(file_path)
    except Exception as e:
        return _error_json(f"无法扫描文件: {e}")

    result: dict[str, Any] = {
        "file_path": file_path,
        "sheets": {},
        "cross_sheet_edges": [],
        "named_ranges": index.named_ranges,
    }

    for name, summary in index.sheets.items():
        result["sheets"][name] = {
            "formula_count": summary.formula_count,
            "self_refs": summary.self_refs,
            "formula_patterns": summary.formula_patterns,
            "outgoing": [e.target_sheet for e in summary.outgoing_refs],
            "incoming": [e.source_sheet for e in summary.incoming_refs],
        }

    for edge in index.cross_sheet_edges:
        result["cross_sheet_edges"].append({
            "source_sheet": edge.source_sheet,
            "target_sheet": edge.target_sheet,
            "ref_type": edge.ref_type.value,
            "ref_count": edge.ref_count,
            "sample_formulas": edge.sample_formulas,
        })

    if detail == "summary":
        summary_text = index.render_summary()
        if summary_text:
            result["summary_text"] = summary_text

    return json.dumps(result, ensure_ascii=False, indent=2)


def trace_references(
    file_path: str,
    target: str,
    direction: str = "both",
    depth: int = 2,
) -> str:
    """追踪单元格引用链。"""
    try:
        abs_path = _resolve_path(file_path)
        sheet_name, address = _parse_target(target)

        from openpyxl import load_workbook
        wb = load_workbook(abs_path, data_only=False, read_only=True)
        try:
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            formula = None
            for row in ws.iter_rows():
                for cell in row:
                    coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                    if coord == address:
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            formula = val
                        break

            precedents: list[dict[str, Any]] = []
            dependents: list[dict[str, Any]] = []

            if formula and direction in ("both", "precedents"):
                refs = _extractor.extract(formula)
                for ref in refs:
                    precedents.append({
                        "cell_or_range": ref.cell_or_range,
                        "sheet_name": ref.sheet_name or sheet_name,
                        "display": ref.display(),
                    })

            if direction in ("both", "dependents"):
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if not isinstance(val, str) or not val.startswith("="):
                            continue
                        refs = _extractor.extract(val)
                        for ref in refs:
                            ref_sheet = ref.sheet_name or sheet_name
                            if ref_sheet == sheet_name and address in ref.cell_or_range:
                                coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                                if coord and coord != address:
                                    dependents.append({
                                        "cell": coord,
                                        "sheet": sheet_name,
                                        "formula": val,
                                    })

            _ensure_index(file_path)
            index = _cache.get_tier1(abs_path)
            if index and direction in ("both", "dependents"):
                for other_sheet in wb.sheetnames:
                    if other_sheet == sheet_name:
                        continue
                    ws_other = wb[other_sheet]
                    for row in ws_other.iter_rows():
                        for cell in row:
                            val = cell.value
                            if not isinstance(val, str) or not val.startswith("="):
                                continue
                            refs = _extractor.extract(val)
                            for ref in refs:
                                if ref.sheet_name == sheet_name and address in ref.cell_or_range:
                                    coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                                    dependents.append({
                                        "cell": coord,
                                        "sheet": other_sheet,
                                        "formula": val,
                                    })
        finally:
            wb.close()

        return json.dumps({
            "target": target,
            "formula": formula,
            "precedents": precedents,
            "dependents": dependents,
        }, ensure_ascii=False, indent=2)
    except Exception as e:
        return _error_json(f"追踪引用失败: {e}")


def get_impact_analysis(
    file_path: str,
    target: str,
    scope: str = "all",
) -> str:
    """分析修改影响范围。"""
    try:
        abs_path = _resolve_path(file_path)
        sheet_name, address = _parse_target(target)

        from openpyxl import load_workbook
        wb = load_workbook(abs_path, data_only=False, read_only=True)
        try:
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]

            direct: list[dict[str, Any]] = []
            affected_sheets: set[str] = set()

            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if not isinstance(val, str) or not val.startswith("="):
                            continue
                        refs = _extractor.extract(val)
                        for ref in refs:
                            ref_sheet = ref.sheet_name or ws_name
                            if ref_sheet == sheet_name and address in ref.cell_or_range:
                                coord = cell.coordinate if hasattr(cell, "coordinate") else ""
                                direct.append({
                                    "cell": coord,
                                    "sheet": ws_name,
                                    "formula": val,
                                })
                                affected_sheets.add(ws_name)
        finally:
            wb.close()

        return json.dumps({
            "target": target,
            "direct_impact": direct,
            "total_affected_cells": len(direct),
            "affected_sheets": sorted(affected_sheets),
        }, ensure_ascii=False, indent=2)
    except Exception as e:
        return _error_json(f"影响分析失败: {e}")


def get_cache() -> RefCache:
    """返回模块级缓存实例（供集成使用）。"""
    return _cache


def get_tools() -> list[ToolDef]:
    """返回引用关系图的所有工具定义。"""
    return [
        ToolDef(
            name="get_reference_map",
            description=(
                "获取工作簿的引用关系全景图：工作表间公式引用、外部引用、命名范围等。"
                "适用场景：理解工作簿结构、了解表间数据流动、发现隐藏的依赖关系。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "detail": {
                        "type": "string",
                        "enum": ["summary", "full"],
                        "description": "详细程度：summary（默认）或 full",
                        "default": "summary",
                    },
                },
                "required": ["file_path"],
            },
            func=get_reference_map,
            write_effect="none",
        ),
        ToolDef(
            name="trace_references",
            description=(
                "追踪单元格的引用链：查看某个单元格引用了哪些源（precedents）"
                "以及被哪些单元格依赖（dependents）。"
                "适用场景：公式追踪、数据血缘分析、理解计算逻辑。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "target": {
                        "type": "string",
                        "description": "目标单元格，格式：Sheet!Cell（如 订单表!C2）",
                    },
                    "direction": {
                        "type": "string",
                        "enum": ["precedents", "dependents", "both"],
                        "description": "追踪方向",
                        "default": "both",
                    },
                    "depth": {
                        "type": "integer",
                        "description": "递归深度（默认 2）",
                        "default": 2,
                    },
                },
                "required": ["file_path", "target"],
            },
            func=trace_references,
            write_effect="none",
        ),
        ToolDef(
            name="get_impact_analysis",
            description=(
                "分析修改某个单元格/区域后的影响范围：哪些单元格会受到影响。"
                "适用场景：修改数据前的风险评估、理解数据变更的传播路径。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "target": {
                        "type": "string",
                        "description": "目标单元格/区域，格式：Sheet!Cell（如 产品表!B2）",
                    },
                    "scope": {
                        "type": "string",
                        "enum": ["all", "direct"],
                        "description": "分析范围：all（全部影响）或 direct（仅直接引用）",
                        "default": "all",
                    },
                },
                "required": ["file_path", "target"],
            },
            func=get_impact_analysis,
            write_effect="none",
        ),
    ]
