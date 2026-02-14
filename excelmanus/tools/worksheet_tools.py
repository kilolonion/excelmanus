"""工作表级设置工具：打印布局、页眉页脚等工作表属性操作。

与 sheet_tools（工作表 CRUD）和 format_tools（单元格样式）不同，
本模块处理工作表级别的全局属性设置。
"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.header_footer import HeaderFooter, HeaderFooterItem

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.worksheet")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "worksheet"
SKILL_DESCRIPTION = "工作表级设置工具集：打印布局、页眉页脚"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard 单例。"""
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


# ── 纸张大小映射 ──────────────────────────────────────────

# openpyxl 纸张大小常量（对应 Excel PageSetup.paperSize）
_PAPER_SIZE_MAP: dict[str, int] = {
    "letter": 1,
    "a3": 8,
    "a4": 9,
    "a5": 11,
    "b4": 12,
    "b5": 13,
    "legal": 5,
    "tabloid": 3,
    "executive": 7,
}


# ── 工具函数 ──────────────────────────────────────────────


def set_print_layout(
    file_path: str,
    sheet_name: str,
    print_area: str | None = None,
    orientation: str | None = None,
    paper_size: str | None = None,
    fit_to_width: int | None = None,
    fit_to_height: int | None = None,
    scale: int | None = None,
    repeat_rows_top: str | None = None,
    repeat_columns_left: str | None = None,
    center_horizontally: bool = False,
    center_vertically: bool = False,
) -> str:
    """设置工作表的打印布局参数。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称。
        print_area: 打印区域范围（如 "A1:L2004"），None 表示不修改。
        orientation: 纸张方向，"landscape"（横向）或 "portrait"（纵向）。
        paper_size: 纸张大小名称（letter/a3/a4/a5/b4/b5/legal/tabloid/executive）。
        fit_to_width: 缩放到指定页宽（页数），设为 1 表示"适合页面宽度"。
        fit_to_height: 缩放到指定页高（页数），设为 0 或 None 表示不限制高度。
        scale: 缩放百分比（10-400），与 fit_to_width/fit_to_height 互斥。
        repeat_rows_top: 每页重复的标题行范围（如 "1:1" 或 "1:3"）。
        repeat_columns_left: 每页重复的标题列范围（如 "A:A" 或 "A:B"）。
        center_horizontally: 是否水平居中打印。
        center_vertically: 是否垂直居中打印。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return json.dumps(
            {"error": f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}"},
            ensure_ascii=False,
        )
    ws = wb[sheet_name]

    applied: list[str] = []

    # 打印区域
    if print_area is not None:
        ws.print_area = print_area
        applied.append(f"print_area={print_area}")

    # 纸张方向
    if orientation is not None:
        orientation_lower = orientation.lower()
        if orientation_lower not in ("landscape", "portrait"):
            wb.close()
            return json.dumps(
                {"error": f"orientation 必须为 'landscape' 或 'portrait'，收到: '{orientation}'"},
                ensure_ascii=False,
            )
        ws.page_setup.orientation = orientation_lower
        applied.append(f"orientation={orientation_lower}")

    # 纸张大小
    if paper_size is not None:
        size_key = paper_size.lower().strip()
        if size_key not in _PAPER_SIZE_MAP:
            wb.close()
            return json.dumps(
                {"error": f"不支持的纸张大小: '{paper_size}'，支持: {list(_PAPER_SIZE_MAP.keys())}"},
                ensure_ascii=False,
            )
        ws.page_setup.paperSize = _PAPER_SIZE_MAP[size_key]
        applied.append(f"paper_size={size_key}")

    # 缩放设置（scale 与 fit_to 互斥）
    if scale is not None:
        if not (10 <= scale <= 400):
            wb.close()
            return json.dumps(
                {"error": "scale 必须在 10-400 之间"},
                ensure_ascii=False,
            )
        ws.page_setup.scale = scale
        # 清除 fit_to 设置
        ws.page_setup.fitToWidth = None
        ws.page_setup.fitToHeight = None
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        applied.append(f"scale={scale}%")
    else:
        if fit_to_width is not None or fit_to_height is not None:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            if fit_to_width is not None:
                ws.page_setup.fitToWidth = fit_to_width
                applied.append(f"fit_to_width={fit_to_width}")
            if fit_to_height is not None:
                ws.page_setup.fitToHeight = fit_to_height
                applied.append(f"fit_to_height={fit_to_height}")

    # 重复标题行
    if repeat_rows_top is not None:
        ws.print_title_rows = repeat_rows_top
        applied.append(f"repeat_rows={repeat_rows_top}")

    # 重复标题列
    if repeat_columns_left is not None:
        ws.print_title_cols = repeat_columns_left
        applied.append(f"repeat_columns={repeat_columns_left}")

    # 居中打印
    if center_horizontally:
        ws.print_options.horizontalCentered = True
        applied.append("center_horizontally=true")
    if center_vertically:
        ws.print_options.verticalCentered = True
        applied.append("center_vertically=true")

    wb.save(safe_path)
    wb.close()

    logger.info("set_print_layout: %s[%s] => %s", safe_path.name, sheet_name, applied)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "sheet": sheet_name,
            "settings_applied": applied,
        },
        ensure_ascii=False,
        indent=2,
    )


def set_page_header_footer(
    file_path: str,
    sheet_name: str,
    header_left: str | None = None,
    header_center: str | None = None,
    header_right: str | None = None,
    footer_left: str | None = None,
    footer_center: str | None = None,
    footer_right: str | None = None,
) -> str:
    """设置工作表的页眉和页脚内容。

    支持 Excel 占位符：
    - &[Page] 或 &P — 当前页码
    - &[Pages] 或 &N — 总页数
    - &[Date] 或 &D — 当前日期
    - &[Time] 或 &T — 当前时间
    - &[Tab] 或 &A — 工作表名
    - &[File] 或 &F — 文件名

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称。
        header_left: 左侧页眉内容。
        header_center: 中间页眉内容。
        header_right: 右侧页眉内容。
        footer_left: 左侧页脚内容。
        footer_center: 中间页脚内容。
        footer_right: 右侧页脚内容。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return json.dumps(
            {"error": f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}"},
            ensure_ascii=False,
        )
    ws = wb[sheet_name]

    applied: list[str] = []

    # 转换用户友好的占位符为 openpyxl 格式
    def _normalize_placeholder(text: str) -> str:
        replacements = {
            "&[Page]": "&P",
            "&[Pages]": "&N",
            "&[Date]": "&D",
            "&[Time]": "&T",
            "&[Tab]": "&A",
            "&[File]": "&F",
        }
        result = text
        for friendly, code in replacements.items():
            result = result.replace(friendly, code)
        return result

    # 页眉
    if header_left is not None:
        ws.oddHeader.left.text = _normalize_placeholder(header_left)
        applied.append(f"header_left={header_left}")
    if header_center is not None:
        ws.oddHeader.center.text = _normalize_placeholder(header_center)
        applied.append(f"header_center={header_center}")
    if header_right is not None:
        ws.oddHeader.right.text = _normalize_placeholder(header_right)
        applied.append(f"header_right={header_right}")

    # 页脚
    if footer_left is not None:
        ws.oddFooter.left.text = _normalize_placeholder(footer_left)
        applied.append(f"footer_left={footer_left}")
    if footer_center is not None:
        ws.oddFooter.center.text = _normalize_placeholder(footer_center)
        applied.append(f"footer_center={footer_center}")
    if footer_right is not None:
        ws.oddFooter.right.text = _normalize_placeholder(footer_right)
        applied.append(f"footer_right={footer_right}")

    wb.save(safe_path)
    wb.close()

    logger.info("set_page_header_footer: %s[%s] => %s", safe_path.name, sheet_name, applied)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "sheet": sheet_name,
            "settings_applied": applied,
        },
        ensure_ascii=False,
        indent=2,
    )


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回工作表级设置工具的所有工具定义。"""
    return [
        ToolDef(
            name="set_print_layout",
            description=(
                "设置 Excel 工作表的打印布局：打印区域、纸张方向/大小、"
                "缩放到页面宽度、每页重复表头行等"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称",
                    },
                    "print_area": {
                        "type": "string",
                        "description": "打印区域范围（如 'A1:L2004'）",
                    },
                    "orientation": {
                        "type": "string",
                        "enum": ["landscape", "portrait"],
                        "description": "纸张方向：landscape（横向）或 portrait（纵向）",
                    },
                    "paper_size": {
                        "type": "string",
                        "enum": ["letter", "a3", "a4", "a5", "b4", "b5", "legal", "tabloid", "executive"],
                        "description": "纸张大小",
                    },
                    "fit_to_width": {
                        "type": "integer",
                        "description": "缩放到指定页宽（1 = 适合页面宽度）。与 scale 互斥",
                    },
                    "fit_to_height": {
                        "type": "integer",
                        "description": "缩放到指定页高（0 = 不限制高度）。与 scale 互斥",
                    },
                    "scale": {
                        "type": "integer",
                        "description": "缩放百分比（10-400）。与 fit_to_width/fit_to_height 互斥",
                    },
                    "repeat_rows_top": {
                        "type": "string",
                        "description": "每页重复的标题行范围（如 '1:1' 或 '1:3'）",
                    },
                    "repeat_columns_left": {
                        "type": "string",
                        "description": "每页重复的标题列范围（如 'A:A'）",
                    },
                    "center_horizontally": {
                        "type": "boolean",
                        "description": "是否水平居中打印",
                        "default": False,
                    },
                    "center_vertically": {
                        "type": "boolean",
                        "description": "是否垂直居中打印",
                        "default": False,
                    },
                },
                "required": ["file_path", "sheet_name"],
                "additionalProperties": False,
            },
            func=set_print_layout,
        ),
        ToolDef(
            name="set_page_header_footer",
            description=(
                "设置 Excel 工作表的页眉和页脚。"
                "支持占位符：&[Page]（页码）、&[Pages]（总页数）、"
                "&[Date]（日期）、&[Time]（时间）、&[Tab]（工作表名）、&[File]（文件名）"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称",
                    },
                    "header_left": {
                        "type": "string",
                        "description": "左侧页眉内容",
                    },
                    "header_center": {
                        "type": "string",
                        "description": "中间页眉内容",
                    },
                    "header_right": {
                        "type": "string",
                        "description": "右侧页眉内容",
                    },
                    "footer_left": {
                        "type": "string",
                        "description": "左侧页脚内容",
                    },
                    "footer_center": {
                        "type": "string",
                        "description": "中间页脚内容（常用 '第 &[Page] 页 / 共 &[Pages] 页'）",
                    },
                    "footer_right": {
                        "type": "string",
                        "description": "右侧页脚内容",
                    },
                },
                "required": ["file_path", "sheet_name"],
                "additionalProperties": False,
            },
            func=set_page_header_footer,
        ),
    ]
