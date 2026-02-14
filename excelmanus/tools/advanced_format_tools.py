"""高级格式工具：阈值图标、卡片样式、单位缩放、暗色仪表盘。"""

from __future__ import annotations

import json
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools.format_tools import COLOR_NAME_MAP
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.advanced_format")
SKILL_NAME = "advanced_format"
SKILL_DESCRIPTION = "高级格式工具：阈值图标、卡片样式、单位缩放、暗色仪表盘"
_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    global _guard
    _guard = FileAccessGuard(workspace_root)


def _iter_cells(ws: Any, cell_range: str) -> list[Cell | MergedCell]:
    cell_data = ws[cell_range]
    if isinstance(cell_data, (Cell, MergedCell)):
        return [cell_data]
    if isinstance(cell_data, tuple) and cell_data and not isinstance(cell_data[0], tuple):
        return list(cell_data)
    return [cell for row in cell_data for cell in row]


def _normalize_hex_color(value: str | None, fallback: str) -> str:
    if value is None:
        return fallback
    raw = str(value).strip()
    if not raw:
        return fallback
    if raw in COLOR_NAME_MAP:
        return COLOR_NAME_MAP[raw]
    lower = raw.lower()
    if lower in COLOR_NAME_MAP:
        return COLOR_NAME_MAP[lower]
    if raw.startswith("#"):
        raw = raw[1:]
    if re.fullmatch(r"[0-9A-Fa-f]{6}", raw):
        return raw.upper()
    return fallback


def _to_display_number(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else f"{value:.12g}"


def _to_divisor_text(divisor: float) -> str:
    return str(int(divisor)) if float(divisor).is_integer() else f"{divisor:.12g}"


def _is_formula_scaled(formula: str, divisor_text: str) -> bool:
    compact = re.sub(r"\s+", "", formula)
    return bool(re.search(rf"/{re.escape(divisor_text)}$", compact))


def _build_unit_number_format(*, decimals: int, suffix: str, thousand_separator: bool) -> str:
    int_part = "#,##0" if thousand_separator else "0"
    decimal_part = f".{('0' * decimals)}" if decimals > 0 else ""
    safe_suffix = suffix.replace('"', '""')
    return f'{int_part}{decimal_part}"{safe_suffix}"'


def _set_chart_title_color(chart: Any, color_hex: str) -> None:
    title = getattr(chart, "title", None)
    if title is None or getattr(title, "tx", None) is None:
        return
    rich = getattr(title.tx, "rich", None)
    if rich is None:
        return
    for paragraph in rich.p or []:
        for run in paragraph.r or []:
            run.rPr = run.rPr or CharacterProperties()
            run.rPr.solidFill = color_hex


def apply_threshold_icon_format(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    high_threshold: float = 1.2,
    mid_threshold: float = 0.8,
    high_symbol: str = "↑",
    mid_symbol: str = "—",
    low_symbol: str = "↓",
    high_color: str = "Green",
    mid_color: str = "Yellow",
    low_color: str = "Red",
    show_value: bool = False,
    decimals: int = 2,
) -> str:
    """三段阈值图标化显示。使用 number_format 精确支持中间“横线”符号。"""
    if high_threshold <= mid_threshold:
        raise ValueError("high_threshold 必须大于 mid_threshold")
    if decimals < 0:
        raise ValueError("decimals 不能小于 0")

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    high_s, mid_s, low_s = (s.replace('"', '""') for s in (high_symbol, mid_symbol, low_symbol))
    high_t, mid_t = _to_display_number(high_threshold), _to_display_number(mid_threshold)
    value_fmt = f"0.{('0' * decimals)}" if decimals > 0 else "0"
    if show_value:
        number_format = (
            f'[{high_color}][>={high_t}]{value_fmt}" {high_s}";'
            f'[{mid_color}][>={mid_t}]{value_fmt}" {mid_s}";'
            f'[{low_color}]{value_fmt}" {low_s}"'
        )
    else:
        number_format = (
            f'[{high_color}][>={high_t}]"{high_s}";'
            f'[{mid_color}][>={mid_t}]"{mid_s}";'
            f'[{low_color}]"{low_s}"'
        )

    cells = _iter_cells(ws, cell_range)
    for cell in cells:
        cell.number_format = number_format

    wb.save(safe_path)
    wb.close()
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "cells_formatted": len(cells),
            "number_format": number_format,
        },
        ensure_ascii=False,
    )


def style_card_blocks(
    file_path: str,
    sheet_name: str,
    ranges: list[str],
    card_fill_color: str = "2E3A46",
    border_color: str = "5A6B7A",
    shadow_color: str = "1A2128",
    text_color: str = "FFFFFF",
    border_style: str = "thick",
    corner_style: str = "medium",
    add_shadow: bool = True,
) -> str:
    """批量应用卡片样式（粗边框 + 圆角模拟 + 阴影模拟）。"""
    if not ranges:
        raise ValueError("ranges 不能为空")

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name]

    card_fill = PatternFill(fill_type="solid", fgColor=_normalize_hex_color(card_fill_color, "2E3A46"))
    shadow_fill = PatternFill(fill_type="solid", fgColor=_normalize_hex_color(shadow_color, "1A2128"))
    text_font = Font(color=_normalize_hex_color(text_color, "FFFFFF"))

    border_hex = _normalize_hex_color(border_color, "5A6B7A")
    edge_side = Side(style=border_style, color=border_hex)
    corner_side = Side(style=corner_style, color=border_hex)

    details: list[dict[str, Any]] = []
    for block in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(block)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = card_fill
                cell.font = text_font
                top = edge_side if r == min_row else Side()
                bottom = edge_side if r == max_row else Side()
                left = edge_side if c == min_col else Side()
                right = edge_side if c == max_col else Side()
                if r == min_row and c == min_col:
                    top, left = corner_side, corner_side
                elif r == min_row and c == max_col:
                    top, right = corner_side, corner_side
                elif r == max_row and c == min_col:
                    bottom, left = corner_side, corner_side
                elif r == max_row and c == max_col:
                    bottom, right = corner_side, corner_side
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        if add_shadow:
            shadow_row, shadow_col = max_row + 1, max_col + 1
            for c in range(min_col + 1, max_col + 2):
                ws.cell(shadow_row, c).fill = shadow_fill
            for r in range(min_row + 1, max_row + 2):
                ws.cell(r, shadow_col).fill = shadow_fill

        details.append(
            {
                "range": block,
                "shadow_applied": add_shadow,
                "anchor": f"{get_column_letter(min_col)}{min_row}",
            }
        )

    wb.save(safe_path)
    wb.close()
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "sheet": sheet_name,
            "blocks_styled": len(details),
            "details": details,
        },
        ensure_ascii=False,
    )


def scale_range_unit(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    divisor: float = 10000.0,
    suffix: str = "万",
    decimals: int = 2,
    thousand_separator: bool = False,
    scale_formulas: bool = True,
) -> str:
    """按除数缩放范围内数值/公式，并设置统一单位格式。"""
    if divisor <= 0:
        raise ValueError("divisor 必须大于 0")
    if decimals < 0:
        raise ValueError("decimals 不能小于 0")

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    number_format = _build_unit_number_format(decimals=decimals, suffix=suffix, thousand_separator=thousand_separator)
    divisor_text = _to_divisor_text(divisor)
    numeric_scaled = formula_scaled = skipped = 0

    for cell in _iter_cells(ws, cell_range):
        value = cell.value
        converted = False
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.value = float(value) / divisor
            numeric_scaled += 1
            converted = True
        elif isinstance(value, str) and value.startswith("=") and scale_formulas:
            formula = value[1:].strip()
            if not _is_formula_scaled(formula, divisor_text):
                cell.value = f"=({formula})/{divisor_text}"
                formula_scaled += 1
                converted = True
        elif isinstance(value, str):
            raw = value.replace(",", "").strip()
            if raw:
                try:
                    parsed = float(raw)
                except ValueError:
                    parsed = None
                if parsed is not None:
                    cell.value = parsed / divisor
                    numeric_scaled += 1
                    converted = True

        if converted or value is not None:
            cell.number_format = number_format
        if not converted and value is not None:
            skipped += 1

    wb.save(safe_path)
    wb.close()
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "number_format": number_format,
            "divisor": divisor,
            "numeric_scaled": numeric_scaled,
            "formula_scaled": formula_scaled,
            "skipped": skipped,
        },
        ensure_ascii=False,
    )


def apply_dashboard_dark_theme(
    file_path: str,
    sheet_name: str,
    cell_range: str | None = None,
    background_color: str = "1B2631",
    text_color: str = "DDE3EA",
    card_ranges: list[str] | None = None,
    card_fill_color: str = "2E3A46",
    card_border_color: str = "455A64",
    metric_ranges: list[str] | None = None,
    metric_color: str = "4FC3F7",
    chart_style: int = 13,
    chart_area_color: str = "1B2631",
    plot_area_color: str = "243447",
    chart_title_color: str = "E5E7EB",
) -> str:
    """暗色仪表盘主题：基础底色、卡片区、指标高亮、图表区域样式。"""
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name]

    if cell_range is None:
        cell_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    bg_fill = PatternFill(fill_type="solid", fgColor=_normalize_hex_color(background_color, "1B2631"))
    base_font = Font(color=_normalize_hex_color(text_color, "DDE3EA"))
    base_cells = _iter_cells(ws, cell_range)
    for cell in base_cells:
        cell.fill = bg_fill
        cell.font = base_font

    if card_ranges:
        card_fill = PatternFill(fill_type="solid", fgColor=_normalize_hex_color(card_fill_color, "2E3A46"))
        side = Side(style="medium", color=_normalize_hex_color(card_border_color, "455A64"))
        for block in card_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(block)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(r, c)
                    cell.fill = card_fill
                    cell.border = Border(
                        left=side if c == min_col else Side(),
                        right=side if c == max_col else Side(),
                        top=side if r == min_row else Side(),
                        bottom=side if r == max_row else Side(),
                    )

    if metric_ranges:
        metric_font = Font(color=_normalize_hex_color(metric_color, "4FC3F7"), bold=True)
        for rg in metric_ranges:
            for cell in _iter_cells(ws, rg):
                cell.font = metric_font

    chart_count = 0
    chart_errors: list[str] = []
    chart_area_hex = _normalize_hex_color(chart_area_color, "1B2631")
    plot_area_hex = _normalize_hex_color(plot_area_color, "243447")
    title_hex = _normalize_hex_color(chart_title_color, "E5E7EB")

    for idx, chart in enumerate(getattr(ws, "_charts", []), start=1):
        chart_count += 1
        try:
            if chart_style > 0:
                chart.style = chart_style
            chart.graphical_properties = GraphicalProperties(solidFill=chart_area_hex)
            plot_area = getattr(chart, "plot_area", None)
            if plot_area is not None:
                plot_area.graphicalProperties = GraphicalProperties(solidFill=plot_area_hex)
            _set_chart_title_color(chart, title_hex)
        except Exception as exc:  # noqa: BLE001
            chart_errors.append(f"chart#{idx}: {exc}")

    wb.save(safe_path)
    wb.close()
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "sheet": sheet_name,
            "base_cells": len(base_cells),
            "card_blocks": len(card_ranges or []),
            "metric_ranges": len(metric_ranges or []),
            "charts_total": chart_count,
            "chart_errors": chart_errors,
        },
        ensure_ascii=False,
    )


def add_color_scale(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    min_color: str = "63BE7B",
    mid_color: str | None = "FFEB84",
    max_color: str = "F8696B",
    min_type: str = "min",
    mid_type: str | None = "percentile",
    max_type: str = "max",
    min_value: float | None = None,
    mid_value: float | None = 50,
    max_value: float | None = None,
) -> str:
    """为指定范围添加二色或三色色阶条件格式。

    Args:
        file_path: Excel 文件路径。
        cell_range: 应用范围（如 "D2:D2004"）。
        sheet_name: 工作表名称，默认活动工作表。
        min_color: 最小值颜色（十六进制或颜色名），默认绿色。
        mid_color: 中间值颜色，传 None 则使用二色色阶。默认黄色。
        max_color: 最大值颜色，默认红色。
        min_type: 最小值类型（min/num/percent/percentile/formula）。
        mid_type: 中间值类型（num/percent/percentile/formula），仅三色色阶。
        max_type: 最大值类型（max/num/percent/percentile/formula）。
        min_value: 最小值阈值，min_type 为 min 时可不填。
        mid_value: 中间值阈值，默认 50。
        max_value: 最大值阈值，max_type 为 max 时可不填。

    Returns:
        JSON 格式的操作结果。
    """
    from openpyxl.formatting.rule import ColorScaleRule

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    min_hex = _normalize_hex_color(min_color, "63BE7B")
    max_hex = _normalize_hex_color(max_color, "F8696B")

    kwargs: dict[str, Any] = {
        "start_type": min_type,
        "start_value": min_value,
        "start_color": min_hex,
        "end_type": max_type,
        "end_value": max_value,
        "end_color": max_hex,
    }

    scale_type = "two_color"
    if mid_color is not None:
        mid_hex = _normalize_hex_color(mid_color, "FFEB84")
        kwargs["mid_type"] = mid_type or "percentile"
        kwargs["mid_value"] = mid_value
        kwargs["mid_color"] = mid_hex
        scale_type = "three_color"

    rule = ColorScaleRule(**kwargs)
    ws.conditional_formatting.add(cell_range, rule)

    wb.save(safe_path)
    wb.close()

    logger.info("add_color_scale: %s[%s] %s (%s)", safe_path.name, sheet_name, cell_range, scale_type)
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "scale_type": scale_type,
            "colors": {"min": min_hex, "mid": mid_color and _normalize_hex_color(mid_color, "FFEB84"), "max": max_hex},
        },
        ensure_ascii=False,
    )


def add_data_bar(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    color: str = "638EC6",
    min_type: str = "min",
    max_type: str = "max",
    min_value: float | None = None,
    max_value: float | None = None,
    show_value: bool = True,
) -> str:
    """为指定范围添加数据条条件格式。

    Args:
        file_path: Excel 文件路径。
        cell_range: 应用范围（如 "E2:E100"）。
        sheet_name: 工作表名称，默认活动工作表。
        color: 数据条颜色（十六进制或颜色名），默认蓝色。
        min_type: 最小值类型（min/num/percent/percentile）。
        max_type: 最大值类型（max/num/percent/percentile）。
        min_value: 最小值阈值。
        max_value: 最大值阈值。
        show_value: 是否同时显示单元格数值，默认 True。

    Returns:
        JSON 格式的操作结果。
    """
    from openpyxl.formatting.rule import DataBarRule

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    color_hex = _normalize_hex_color(color, "638EC6")

    rule = DataBarRule(
        start_type=min_type,
        start_value=min_value,
        end_type=max_type,
        end_value=max_value,
        color=color_hex,
        showValue=show_value,
    )
    ws.conditional_formatting.add(cell_range, rule)

    wb.save(safe_path)
    wb.close()

    logger.info("add_data_bar: %s[%s] %s", safe_path.name, sheet_name, cell_range)
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "color": color_hex,
            "show_value": show_value,
        },
        ensure_ascii=False,
    )


def add_conditional_rule(
    file_path: str,
    cell_range: str,
    rule_type: str,
    sheet_name: str | None = None,
    operator: str | None = None,
    formula: str | list[str] | None = None,
    values: list[float] | None = None,
    font_color: str | None = None,
    fill_color: str | None = None,
    font_bold: bool | None = None,
    icon_style: str | None = None,
    reverse_icons: bool = False,
) -> str:
    """添加通用条件格式规则（CellIs、Formula、IconSet）。

    rule_type 说明：
    - "cell_is": 基于单元格值的条件（需 operator + formula/values）。
      operator 支持：greaterThan, lessThan, greaterThanOrEqual, lessThanOrEqual,
      equal, notEqual, between, notBetween, containsText。
    - "formula": 基于公式的条件（需 formula 参数）。
      公式为真时应用样式，如 '=$C2="FATAL"' 表示当 C 列为 FATAL 时整行变色。
    - "icon_set": 图标集（需 icon_style 参数）。
      icon_style 支持：3Arrows, 3ArrowsGray, 3Flags, 3TrafficLights1,
      3TrafficLights2, 3Signs, 3Symbols, 3Symbols2, 4Arrows, 4ArrowsGray,
      4RedToBlack, 4Rating, 4TrafficLights, 5Arrows, 5ArrowsGray,
      5Rating, 5Quarters。

    Args:
        file_path: Excel 文件路径。
        cell_range: 应用范围（如 "A2:L2004"）。
        rule_type: 规则类型（cell_is / formula / icon_set）。
        sheet_name: 工作表名称，默认活动工作表。
        operator: cell_is 模式的比较运算符。
        formula: 公式字符串或列表。cell_is 模式下是比较值/公式，formula 模式下是条件公式。
        values: cell_is between/notBetween 模式下的两个边界值。
        font_color: 条件满足时的字体颜色。
        fill_color: 条件满足时的填充颜色。
        font_bold: 条件满足时是否加粗。
        icon_style: icon_set 模式的图标样式名称。
        reverse_icons: icon_set 模式是否反转图标顺序。

    Returns:
        JSON 格式的操作结果。
    """
    from openpyxl.formatting.rule import CellIsRule, FormulaRule, IconSetRule

    valid_rule_types = ("cell_is", "formula", "icon_set")
    if rule_type not in valid_rule_types:
        return json.dumps(
            {"error": f"rule_type 必须为 {valid_rule_types} 之一，收到: '{rule_type}'"},
            ensure_ascii=False,
        )

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)
    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # 构建样式参数
    style_font = None
    style_fill = None
    if font_color is not None or font_bold is not None:
        font_kwargs: dict[str, Any] = {}
        if font_color is not None:
            font_kwargs["color"] = _normalize_hex_color(font_color, "000000")
        if font_bold is not None:
            font_kwargs["bold"] = font_bold
        style_font = Font(**font_kwargs)
    if fill_color is not None:
        hex_color = _normalize_hex_color(fill_color, "FFFFFF")
        style_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    rule_detail = rule_type

    if rule_type == "cell_is":
        if operator is None:
            wb.close()
            return json.dumps(
                {"error": "cell_is 规则需要 operator 参数"},
                ensure_ascii=False,
            )
        rule_kwargs: dict[str, Any] = {"operator": operator}
        if operator in ("between", "notBetween"):
            if values is None or len(values) != 2:
                wb.close()
                return json.dumps(
                    {"error": f"operator='{operator}' 需要 values 参数包含恰好 2 个值"},
                    ensure_ascii=False,
                )
            rule_kwargs["formula"] = [str(values[0]), str(values[1])]
        else:
            if formula is not None:
                rule_kwargs["formula"] = [formula] if isinstance(formula, str) else formula
            elif values is not None and len(values) >= 1:
                rule_kwargs["formula"] = [str(values[0])]
            else:
                wb.close()
                return json.dumps(
                    {"error": "cell_is 规则需要 formula 或 values 参数指定比较值"},
                    ensure_ascii=False,
                )
        if style_font:
            rule_kwargs["font"] = style_font
        if style_fill:
            rule_kwargs["fill"] = style_fill
        rule = CellIsRule(**rule_kwargs)
        rule_detail = f"cell_is({operator})"

    elif rule_type == "formula":
        if formula is None:
            wb.close()
            return json.dumps(
                {"error": "formula 规则需要 formula 参数"},
                ensure_ascii=False,
            )
        formula_list = [formula] if isinstance(formula, str) else formula
        rule_kwargs = {"formula": formula_list}
        if style_font:
            rule_kwargs["font"] = style_font
        if style_fill:
            rule_kwargs["fill"] = style_fill
        rule = FormulaRule(**rule_kwargs)
        rule_detail = f"formula({formula_list[0][:60]})"

    elif rule_type == "icon_set":
        if icon_style is None:
            wb.close()
            return json.dumps(
                {"error": "icon_set 规则需要 icon_style 参数"},
                ensure_ascii=False,
            )
        rule = IconSetRule(
            icon_style=icon_style,
            type="percent",
            values=[0, 33, 67],
            showValue=True,
            reverse=reverse_icons,
        )
        rule_detail = f"icon_set({icon_style})"

    ws.conditional_formatting.add(cell_range, rule)

    wb.save(safe_path)
    wb.close()

    logger.info("add_conditional_rule: %s[%s] %s %s", safe_path.name, sheet_name, cell_range, rule_detail)
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "rule_type": rule_type,
            "rule_detail": rule_detail,
        },
        ensure_ascii=False,
    )


def get_tools() -> list[ToolDef]:
    """返回高级格式工具定义。"""
    border_style_enum = ["thin", "medium", "thick", "double", "dotted", "dashed"]
    return [
        ToolDef(
            name="apply_threshold_icon_format",
            description="三段阈值图标化显示（支持精确上/横/下符号）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string"},
                    "cell_range": {"type": "string"},
                    "high_threshold": {"type": "number", "default": 1.2},
                    "mid_threshold": {"type": "number", "default": 0.8},
                    "high_symbol": {"type": "string", "default": "↑"},
                    "mid_symbol": {"type": "string", "default": "—"},
                    "low_symbol": {"type": "string", "default": "↓"},
                    "high_color": {"type": "string", "default": "Green"},
                    "mid_color": {"type": "string", "default": "Yellow"},
                    "low_color": {"type": "string", "default": "Red"},
                    "show_value": {"type": "boolean", "default": False},
                    "decimals": {"type": "integer", "default": 2},
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=apply_threshold_icon_format,
        ),
        ToolDef(
            name="style_card_blocks",
            description="批量卡片化区域（粗边框+圆角与阴影模拟）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string"},
                    "ranges": {"type": "array", "items": {"type": "string"}},
                    "card_fill_color": {"type": "string", "default": "2E3A46"},
                    "border_color": {"type": "string", "default": "5A6B7A"},
                    "shadow_color": {"type": "string", "default": "1A2128"},
                    "text_color": {"type": "string", "default": "FFFFFF"},
                    "border_style": {"type": "string", "enum": border_style_enum, "default": "thick"},
                    "corner_style": {"type": "string", "enum": border_style_enum, "default": "medium"},
                    "add_shadow": {"type": "boolean", "default": True},
                },
                "required": ["file_path", "sheet_name", "ranges"],
                "additionalProperties": False,
            },
            func=style_card_blocks,
        ),
        ToolDef(
            name="scale_range_unit",
            description="按除数缩放数值/公式并统一单位格式（如元转万）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string"},
                    "cell_range": {"type": "string"},
                    "divisor": {"type": "number", "default": 10000.0},
                    "suffix": {"type": "string", "default": "万"},
                    "decimals": {"type": "integer", "default": 2},
                    "thousand_separator": {"type": "boolean", "default": False},
                    "scale_formulas": {"type": "boolean", "default": True},
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=scale_range_unit,
        ),
        ToolDef(
            name="apply_dashboard_dark_theme",
            description="暗色仪表盘主题（单元格、卡片和图表样式）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string"},
                    "cell_range": {"type": "string"},
                    "background_color": {"type": "string", "default": "1B2631"},
                    "text_color": {"type": "string", "default": "DDE3EA"},
                    "card_ranges": {"type": "array", "items": {"type": "string"}},
                    "card_fill_color": {"type": "string", "default": "2E3A46"},
                    "card_border_color": {"type": "string", "default": "455A64"},
                    "metric_ranges": {"type": "array", "items": {"type": "string"}},
                    "metric_color": {"type": "string", "default": "4FC3F7"},
                    "chart_style": {"type": "integer", "default": 13},
                    "chart_area_color": {"type": "string", "default": "1B2631"},
                    "plot_area_color": {"type": "string", "default": "243447"},
                    "chart_title_color": {"type": "string", "default": "E5E7EB"},
                },
                "required": ["file_path", "sheet_name"],
                "additionalProperties": False,
            },
            func=apply_dashboard_dark_theme,
        ),
        ToolDef(
            name="add_color_scale",
            description="为 Excel 范围添加二色或三色色阶条件格式（低值→高值渐变色）。颜色支持十六进制或中文名",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Excel 文件路径"},
                    "cell_range": {"type": "string", "description": "应用范围（如 'D2:D2004'）"},
                    "sheet_name": {"type": "string", "description": "工作表名称，默认活动工作表"},
                    "min_color": {"type": "string", "default": "63BE7B", "description": "最小值颜色，默认绿色"},
                    "mid_color": {"type": "string", "default": "FFEB84", "description": "中间值颜色，不传则使用二色色阶"},
                    "max_color": {"type": "string", "default": "F8696B", "description": "最大值颜色，默认红色"},
                    "min_type": {"type": "string", "enum": ["min", "num", "percent", "percentile", "formula"], "default": "min"},
                    "mid_type": {"type": "string", "enum": ["num", "percent", "percentile", "formula"], "default": "percentile"},
                    "max_type": {"type": "string", "enum": ["max", "num", "percent", "percentile", "formula"], "default": "max"},
                    "min_value": {"type": "number", "description": "最小值阈值"},
                    "mid_value": {"type": "number", "default": 50, "description": "中间值阈值"},
                    "max_value": {"type": "number", "description": "最大值阈值"},
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=add_color_scale,
        ),
        ToolDef(
            name="add_data_bar",
            description="为 Excel 范围添加数据条条件格式（单元格内按比例显示彩色条）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Excel 文件路径"},
                    "cell_range": {"type": "string", "description": "应用范围（如 'E2:E100'）"},
                    "sheet_name": {"type": "string", "description": "工作表名称，默认活动工作表"},
                    "color": {"type": "string", "default": "638EC6", "description": "数据条颜色，默认蓝色"},
                    "min_type": {"type": "string", "enum": ["min", "num", "percent", "percentile"], "default": "min"},
                    "max_type": {"type": "string", "enum": ["max", "num", "percent", "percentile"], "default": "max"},
                    "min_value": {"type": "number", "description": "最小值阈值"},
                    "max_value": {"type": "number", "description": "最大值阈值"},
                    "show_value": {"type": "boolean", "default": True, "description": "是否同时显示数值"},
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=add_data_bar,
        ),
        ToolDef(
            name="add_conditional_rule",
            description=(
                "添加通用条件格式规则。rule_type 支持三种："
                "(1) cell_is — 基于值比较（如 >1000 高亮），需 operator + formula/values；"
                "(2) formula — 基于公式条件（如 FATAL 行整行变色），需 formula；"
                "(3) icon_set — 图标集（如 3Arrows 三箭头），需 icon_style"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Excel 文件路径"},
                    "cell_range": {"type": "string", "description": "应用范围（如 'A2:L2004'）"},
                    "rule_type": {
                        "type": "string",
                        "enum": ["cell_is", "formula", "icon_set"],
                        "description": "规则类型",
                    },
                    "sheet_name": {"type": "string", "description": "工作表名称，默认活动工作表"},
                    "operator": {
                        "type": "string",
                        "enum": [
                            "greaterThan", "lessThan", "greaterThanOrEqual", "lessThanOrEqual",
                            "equal", "notEqual", "between", "notBetween", "containsText",
                        ],
                        "description": "cell_is 模式的比较运算符",
                    },
                    "formula": {
                        "description": "公式字符串。cell_is 模式下是比较值，formula 模式下是条件公式（如 '=$C2=\"FATAL\"'）",
                    },
                    "values": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "between/notBetween 的两个边界值，或单值比较的数值",
                    },
                    "font_color": {"type": "string", "description": "条件满足时的字体颜色"},
                    "fill_color": {"type": "string", "description": "条件满足时的填充颜色"},
                    "font_bold": {"type": "boolean", "description": "条件满足时是否加粗"},
                    "icon_style": {
                        "type": "string",
                        "description": "icon_set 图标样式（如 '3Arrows', '3TrafficLights1', '4Rating', '5Quarters'）",
                    },
                    "reverse_icons": {"type": "boolean", "default": False, "description": "是否反转图标顺序"},
                },
                "required": ["file_path", "cell_range", "rule_type"],
                "additionalProperties": False,
            },
            func=add_conditional_rule,
        ),
    ]
