"""格式化工具：提供单元格格式化、样式读取、合并单元格和行列尺寸调整能力。"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools._helpers import get_worksheet
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.format")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "format"
SKILL_DESCRIPTION = "格式化工具集：样式读取与设置、合并单元格、行列尺寸调整"

# ── 中文颜色名 → 十六进制映射 ────────────────────────────

COLOR_NAME_MAP: dict[str, str] = {
    # 基础色
    "红": "FF0000", "红色": "FF0000", "red": "FF0000",
    "绿": "00B050", "绿色": "00B050", "green": "00B050",
    "蓝": "0000FF", "蓝色": "0000FF", "blue": "0000FF",
    "黄": "FFFF00", "黄色": "FFFF00", "yellow": "FFFF00",
    "白": "FFFFFF", "白色": "FFFFFF", "white": "FFFFFF",
    "黑": "000000", "黑色": "000000", "black": "000000",
    # 常用色
    "橙": "FFC000", "橙色": "FFC000", "orange": "FFC000",
    "紫": "7030A0", "紫色": "7030A0", "purple": "7030A0",
    "粉": "FF69B4", "粉色": "FF69B4", "pink": "FF69B4",
    "棕": "8B4513", "棕色": "8B4513", "brown": "8B4513",
    "灰": "808080", "灰色": "808080", "gray": "808080", "grey": "808080",
    "青": "00CED1", "青色": "00CED1", "cyan": "00FFFF",
    # 浅色系
    "浅蓝": "5B9BD5", "浅蓝色": "5B9BD5", "lightblue": "ADD8E6",
    "浅绿": "92D050", "浅绿色": "92D050", "lightgreen": "90EE90",
    "浅黄": "FFF2CC", "浅黄色": "FFF2CC", "lightyellow": "FFFFE0",
    "浅灰": "D9D9D9", "浅灰色": "D9D9D9", "lightgray": "D3D3D3",
    "浅红": "FF7F7F", "浅红色": "FF7F7F",
    "浅紫": "B4A7D6", "浅紫色": "B4A7D6",
    # 深色系
    "深蓝": "002060", "深蓝色": "002060", "darkblue": "00008B",
    "深绿": "006100", "深绿色": "006100", "darkgreen": "006400",
    "深红": "C00000", "深红色": "C00000", "darkred": "8B0000",
    "深灰": "404040", "深灰色": "404040", "darkgray": "A9A9A9",
    # Excel 主题常用色
    "金": "FFD700", "金色": "FFD700", "gold": "FFD700",
    "银": "C0C0C0", "银色": "C0C0C0", "silver": "C0C0C0",
    "天蓝": "4472C4", "天蓝色": "4472C4",
    "草绿": "70AD47", "草绿色": "70AD47",
    "珊瑚": "FF7F50", "珊瑚色": "FF7F50", "coral": "FF7F50",
}

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard 单例。"""
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


# ── 工具函数 ──────────────────────────────────────────────


def format_cells(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    font: dict[str, Any] | None = None,
    fill: dict[str, Any] | None = None,
    border: dict[str, Any] | None = None,
    alignment: dict[str, Any] | None = None,
    number_format: str | None = None,
    return_styles: bool = False,
) -> str:
    """对指定单元格范围应用格式化样式。

    Args:
        file_path: Excel 文件路径。
        cell_range: 单元格范围，如 "A1:C3" 或 "A1"。
        sheet_name: 工作表名称，默认活动工作表。
        font: 字体设置，支持 name/size/bold/italic/color。
        fill: 填充设置，支持 color（十六进制颜色码）。
        border: 边框设置，支持 style（thin/medium/thick）和 color。
        alignment: 对齐设置，支持 horizontal/vertical/wrap_text。
        number_format: 数字格式字符串，如 "#,##0.00"。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    # 构建 openpyxl 样式对象
    style_font = _build_font(font) if font else None
    style_fill = _build_fill(fill) if fill else None
    style_border = _build_border(border) if border else None
    style_alignment = _build_alignment(alignment) if alignment else None

    # 应用样式到范围内每个单元格
    applied_count = 0
    cell_data = ws[cell_range]

    # ws[cell_range] 对单个单元格返回 Cell，对范围返回 tuple of tuples
    if isinstance(cell_data, (Cell, MergedCell)):
        rows = ((cell_data,),)
    elif isinstance(cell_data, tuple) and cell_data and not isinstance(cell_data[0], tuple):
        # 单行范围如 "A1:C1" 返回 tuple of Cell
        rows = (cell_data,)
    else:
        rows = cell_data

    for row in rows:
        cells = row if isinstance(row, tuple) else (row,)
        for cell in cells:
            if style_font:
                cell.font = style_font
            if style_fill:
                cell.fill = style_fill
            if style_border:
                cell.border = style_border
            if style_alignment:
                cell.alignment = style_alignment
            if number_format:
                cell.number_format = number_format
            applied_count += 1

    wb.save(safe_path)

    result_data: dict[str, Any] = {
        "status": "success",
        "file": safe_path.name,
        "range": cell_range,
        "cells_formatted": applied_count,
    }

    # 格式化后返回样式快照
    if return_styles:
        from excelmanus.tools.data_tools import _collect_styles_compressed

        # 需要重新打开文件读取写入后的样式
        wb2 = load_workbook(safe_path)
        try:
            ws2 = get_worksheet(wb2, sheet_name)
            result_data["after_styles"] = _collect_styles_compressed(ws2, max_rows=200)
        finally:
            wb2.close()
    
    wb.close()

    logger.info("已格式化 %s 范围 %s（%d 个单元格）", safe_path.name, cell_range, applied_count)

    return json.dumps(result_data, ensure_ascii=False, indent=2)


def adjust_column_width(
    file_path: str,
    columns: dict[str, float] | None = None,
    auto_fit: bool = False,
    sheet_name: str | None = None,
) -> str:
    """调整列宽：支持指定宽度或自动适配。

    Args:
        file_path: Excel 文件路径。
        columns: 列宽映射，如 {"A": 20, "B": 15}。与 auto_fit 互斥时优先使用。
        auto_fit: 是否自动适配所有列宽（基于内容最大长度）。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    adjusted: dict[str, float] = {}

    if columns:
        # 手动指定列宽
        for col_letter, width in columns.items():
            ws.column_dimensions[col_letter.upper()].width = width
            adjusted[col_letter.upper()] = width

    elif auto_fit:
        # 自动适配：遍历所有列，取每列最大内容长度
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            # 加 2 作为边距
            width = max_len + 2
            ws.column_dimensions[col_letter].width = width
            adjusted[col_letter] = width

    wb.save(safe_path)
    wb.close()

    logger.info("已调整 %s 列宽（%d 列）", safe_path.name, len(adjusted))

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "columns_adjusted": adjusted,
        },
        ensure_ascii=False,
    )


def read_cell_styles(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    summary_only: bool = False,
) -> str:
    """读取指定单元格范围的样式信息（字体、填充、边框、对齐等）。

    Args:
        file_path: Excel 文件路径。
        cell_range: 单元格范围，如 "A1:C3" 或 "A1"。
        sheet_name: 工作表名称，默认活动工作表。
        summary_only: 仅返回样式统计汇总而非逐单元格明细。

    Returns:
        JSON 格式的样式信息。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    # 获取合并单元格范围集合
    merged_ranges = ws.merged_cells.ranges
    merged_set: set[str] = set()
    for mr in merged_ranges:
        for row in mr.rows:
            for cell_coord in row:
                merged_set.add(f"{get_column_letter(cell_coord[1])}{cell_coord[0]}")

    cell_data = ws[cell_range]
    # 归一化为 tuple of tuples
    if isinstance(cell_data, (Cell, MergedCell)):
        rows: tuple = ((cell_data,),)
    elif isinstance(cell_data, tuple) and cell_data and not isinstance(cell_data[0], tuple):
        rows = (cell_data,)
    else:
        rows = cell_data

    max_cells = 200
    cell_styles: list[dict[str, Any]] = []
    fill_colors: set[str] = set()
    font_colors: set[str] = set()
    font_names: set[str] = set()
    border_styles_used: set[str] = set()
    total_cells = 0

    for row in rows:
        row_cells = row if isinstance(row, tuple) else (row,)
        for cell in row_cells:
            total_cells += 1
            coord = f"{get_column_letter(cell.column)}{cell.row}"
            is_merged = coord in merged_set

            # 提取样式信息
            font_info = _extract_font(cell.font)
            fill_info = _extract_fill(cell.fill)
            border_info = _extract_border(cell.border)
            align_info = _extract_alignment(cell.alignment)
            num_fmt = cell.number_format if cell.number_format != "General" else None

            # 收集统计信息
            if fill_info and fill_info.get("color"):
                fill_colors.add(fill_info["color"])
            if font_info:
                if font_info.get("color"):
                    font_colors.add(font_info["color"])
                if font_info.get("name"):
                    font_names.add(font_info["name"])
            if border_info:
                for side_name in ("left", "right", "top", "bottom"):
                    s = border_info.get(side_name)
                    if s and s != "none":
                        border_styles_used.add(s)

            if not summary_only and len(cell_styles) < max_cells:
                # 只输出有非默认样式的单元格
                has_style = any([font_info, fill_info, border_info, align_info, num_fmt, is_merged])
                if has_style:
                    entry: dict[str, Any] = {"cell": coord}
                    val = cell.value
                    if val is not None:
                        entry["value"] = str(val) if not isinstance(val, (int, float, bool)) else val
                    if font_info:
                        entry["font"] = font_info
                    if fill_info:
                        entry["fill"] = fill_info
                    if border_info:
                        entry["border"] = border_info
                    if align_info:
                        entry["alignment"] = align_info
                    if num_fmt:
                        entry["number_format"] = num_fmt
                    if is_merged:
                        entry["merged"] = True
                    cell_styles.append(entry)

    # 在 wb.close() 之前保存 shape 信息
    sheet_max_row = ws.max_row or 0
    sheet_max_col = ws.max_column or 0

    wb.close()

    # 构建合并范围列表
    range_merged: list[str] = [str(mr) for mr in merged_ranges]

    result: dict[str, Any] = {
        "status": "success",
        "file": safe_path.name,
        "range": cell_range,
        "total_cells": total_cells,
        "rows": sheet_max_row,
        "columns": sheet_max_col,
        "summary": {
            "fill_colors_used": sorted(fill_colors),
            "font_colors_used": sorted(font_colors),
            "font_names_used": sorted(font_names),
            "border_styles_used": sorted(border_styles_used),
            "merged_ranges": range_merged,
            "has_merged_cells": len(range_merged) > 0,
        },
    }
    if not summary_only:
        result["styled_cells"] = cell_styles
        if len(cell_styles) >= max_cells:
            result["truncated"] = True
            result["truncated_message"] = f"仅展示前 {max_cells} 个有样式的单元格"

    logger.info("已读取 %s 范围 %s 的样式（%d 个单元格）", safe_path.name, cell_range, total_cells)
    return json.dumps(result, ensure_ascii=False, indent=2)


def adjust_row_height(
    file_path: str,
    rows: dict[str, float] | None = None,
    auto_fit: bool = False,
    sheet_name: str | None = None,
) -> str:
    """调整行高：支持指定高度或自动适配。

    Args:
        file_path: Excel 文件路径。
        rows: 行高映射，如 {"1": 30, "2": 25}（键为行号字符串）。
        auto_fit: 是否自动适配所有行高（基于默认行高 * 1.2）。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    adjusted: dict[str, float] = {}

    if rows:
        for row_num_str, height in rows.items():
            row_num = int(row_num_str)
            ws.row_dimensions[row_num].height = height
            adjusted[str(row_num)] = height
    elif auto_fit:
        for row_idx in range(1, ws.max_row + 1):
            # 基于行内最大字体大小估算行高
            max_font_size = 11.0
            for cell in ws[row_idx]:
                if cell.font and cell.font.size:
                    max_font_size = max(max_font_size, float(cell.font.size))
            height = max_font_size * 1.5
            ws.row_dimensions[row_idx].height = height
            adjusted[str(row_idx)] = height

    wb.save(safe_path)
    wb.close()

    logger.info("已调整 %s 行高（%d 行）", safe_path.name, len(adjusted))

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "rows_adjusted": adjusted,
        },
        ensure_ascii=False,
    )


def merge_cells_tool(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
) -> str:
    """合并指定范围的单元格。

    Args:
        file_path: Excel 文件路径。
        cell_range: 要合并的单元格范围，如 "A1:C1"。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    ws.merge_cells(cell_range)
    wb.save(safe_path)
    wb.close()

    logger.info("已合并 %s 范围 %s", safe_path.name, cell_range)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "merged_range": cell_range,
        },
        ensure_ascii=False,
    )


def unmerge_cells_tool(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
) -> str:
    """取消合并指定范围的单元格。

    Args:
        file_path: Excel 文件路径。
        cell_range: 要取消合并的单元格范围，如 "A1:C1"。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    ws.unmerge_cells(cell_range)
    wb.save(safe_path)
    wb.close()

    logger.info("已取消合并 %s 范围 %s", safe_path.name, cell_range)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "unmerged_range": cell_range,
        },
        ensure_ascii=False,
    )


# ── 内部辅助函数 ──────────────────────────────────────────


def _resolve_color(value: str | None) -> str | None:
    """将颜色名称或十六进制码统一解析为十六进制码。"""
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in COLOR_NAME_MAP:
        return COLOR_NAME_MAP[normalized]
    # 去除可能的 # 前缀
    hex_value = value.strip().lstrip("#")
    if len(hex_value) in (6, 8) and all(c in "0123456789abcdefABCDEF" for c in hex_value):
        return hex_value.upper()
    return value


def _build_font(config: dict[str, Any]) -> Font:
    """从配置字典构建 openpyxl Font 对象。"""
    return Font(
        name=config.get("name"),
        size=config.get("size"),
        bold=config.get("bold"),
        italic=config.get("italic"),
        color=_resolve_color(config.get("color")),
        underline=config.get("underline"),
        strike=config.get("strikethrough"),
    )


def _build_fill(config: dict[str, Any]) -> PatternFill:
    """从配置字典构建 openpyxl PatternFill 对象。"""
    color = _resolve_color(config.get("color")) or "FFFFFF"
    fill_type = config.get("fill_type", "solid")
    return PatternFill(
        start_color=color,
        end_color=color,
        fill_type=fill_type,
    )


def _build_side(side_config: dict[str, Any] | str) -> Side:
    """从配置构建单个 Side 对象。"""
    if isinstance(side_config, str):
        return Side(style=side_config, color="000000")
    return Side(
        style=side_config.get("style", "thin"),
        color=_resolve_color(side_config.get("color")) or "000000",
    )


def _build_border(config: dict[str, Any]) -> Border:
    """从配置字典构建 openpyxl Border 对象。支持统一设置或单边差异化。"""
    # 如果指定了 left/right/top/bottom 中任意一个，使用单边模式
    has_sides = any(k in config for k in ("left", "right", "top", "bottom"))
    if has_sides:
        return Border(
            left=_build_side(config["left"]) if "left" in config else Side(),
            right=_build_side(config["right"]) if "right" in config else Side(),
            top=_build_side(config["top"]) if "top" in config else Side(),
            bottom=_build_side(config["bottom"]) if "bottom" in config else Side(),
        )
    # 统一模式：四边相同
    style = config.get("style", "thin")
    color = _resolve_color(config.get("color")) or "000000"
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _build_alignment(config: dict[str, Any]) -> Alignment:
    """从配置字典构建 openpyxl Alignment 对象。"""
    return Alignment(
        horizontal=config.get("horizontal"),
        vertical=config.get("vertical"),
        wrap_text=config.get("wrap_text"),
    )


# ── 样式提取辅助函数（用于 read_cell_styles）──────────────


def _color_to_hex(color: Any) -> str | None:
    """将 openpyxl Color 对象转换为十六进制字符串。"""
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        # openpyxl 的 rgb 可能是 AARRGGBB 格式
        if len(rgb) == 8:
            return rgb[2:]  # 去掉 alpha 通道
        return rgb
    if hasattr(color, "theme") and color.theme is not None:
        return f"theme:{color.theme}"
    if hasattr(color, "indexed") and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return None


def _extract_font(font: Font | None) -> dict[str, Any] | None:
    """从 openpyxl Font 提取非默认属性字典。"""
    if font is None:
        return None
    info: dict[str, Any] = {}
    if font.name and font.name != "Calibri":
        info["name"] = font.name
    if font.size and font.size != 11:
        info["size"] = font.size
    if font.bold:
        info["bold"] = True
    if font.italic:
        info["italic"] = True
    if font.underline and font.underline != "none":
        info["underline"] = font.underline
    if font.strike:
        info["strikethrough"] = True
    color_hex = _color_to_hex(font.color)
    if color_hex and color_hex != "000000":
        info["color"] = color_hex
    return info or None


def _extract_fill(fill: PatternFill | None) -> dict[str, Any] | None:
    """从 openpyxl PatternFill 提取非默认属性字典。"""
    if fill is None:
        return None
    fill_type = fill.fill_type or fill.patternType
    if not fill_type or fill_type == "none":
        return None
    info: dict[str, Any] = {"type": fill_type}
    color_hex = _color_to_hex(fill.fgColor)
    if color_hex:
        info["color"] = color_hex
    return info


def _extract_border(border: Border | None) -> dict[str, Any] | None:
    """从 openpyxl Border 提取非默认属性字典。"""
    if border is None:
        return None
    info: dict[str, Any] = {}
    for side_name in ("left", "right", "top", "bottom"):
        side: Side = getattr(border, side_name, None)
        if side and side.style and side.style != "none":
            info[side_name] = side.style
    return info or None


def _extract_alignment(alignment: Alignment | None) -> dict[str, Any] | None:
    """从 openpyxl Alignment 提取非默认属性字典。"""
    if alignment is None:
        return None
    info: dict[str, Any] = {}
    if alignment.horizontal and alignment.horizontal != "general":
        info["horizontal"] = alignment.horizontal
    if alignment.vertical and alignment.vertical != "bottom":
        info["vertical"] = alignment.vertical
    if alignment.wrap_text:
        info["wrap_text"] = True
    return info or None


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回格式化 Skill 的所有工具定义。"""
    _side_schema = {
        "type": "object",
        "description": "单边边框设置",
        "properties": {
            "style": {"type": "string", "enum": ["thin", "medium", "thick", "double", "dotted", "dashed"]},
            "color": {"type": "string", "description": "颜色码或颜色名（如 '红色'、'FF0000'）"},
        },
    }
    return [
        ToolDef(
            name="format_cells",
            description="对 Excel 单元格范围应用格式化样式（字体、填充、边框、对齐、数字格式）。颜色参数支持中文名（如 '红色'）或十六进制码（如 'FF0000'）。设置 return_styles=true 可在格式化后直接返回样式快照，省去额外 read_cell_styles 验证",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "单元格范围，如 'A1:C3' 或 'A1'",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                    "font": {
                        "type": "object",
                        "description": "字体设置：name/size/bold/italic/color/underline/strikethrough",
                        "properties": {
                            "name": {"type": "string"},
                            "size": {"type": "number"},
                            "bold": {"type": "boolean"},
                            "italic": {"type": "boolean"},
                            "color": {"type": "string", "description": "颜色码或颜色名（如 '红色'、'FF0000'）"},
                            "underline": {"type": "string", "enum": ["single", "double", "none"], "description": "下划线样式"},
                            "strikethrough": {"type": "boolean", "description": "是否添加删除线"},
                        },
                    },
                    "fill": {
                        "type": "object",
                        "description": "填充设置：color + 可选 fill_type",
                        "properties": {
                            "color": {"type": "string", "description": "颜色码或颜色名（如 '浅黄色'、'FFFF00'）"},
                            "fill_type": {"type": "string", "enum": ["solid", "none"], "default": "solid"},
                        },
                    },
                    "border": {
                        "type": "object",
                        "description": "边框设置：统一模式用 style+color，差异化模式用 left/right/top/bottom",
                        "properties": {
                            "style": {"type": "string", "enum": ["thin", "medium", "thick", "double", "dotted", "dashed"]},
                            "color": {"type": "string", "description": "颜色码或颜色名"},
                            "left": _side_schema,
                            "right": _side_schema,
                            "top": _side_schema,
                            "bottom": _side_schema,
                        },
                    },
                    "alignment": {
                        "type": "object",
                        "description": "对齐设置：horizontal/vertical/wrap_text",
                        "properties": {
                            "horizontal": {"type": "string", "enum": ["left", "center", "right"]},
                            "vertical": {"type": "string", "enum": ["top", "center", "bottom"]},
                            "wrap_text": {"type": "boolean"},
                        },
                    },
                    "number_format": {
                        "type": "string",
                        "description": "数字格式字符串，如 '#,##0.00'",
                    },
                    "return_styles": {
                        "type": "boolean",
                        "description": "格式化后返回压缩样式快照（省去额外 read_cell_styles 验证）",
                        "default": False,
                    },
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=format_cells,
        ),
        ToolDef(
            name="adjust_column_width",
            description="调整 Excel 列宽：支持手动指定宽度或自动适配内容",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "columns": {
                        "type": "object",
                        "description": "列宽映射，如 {'A': 20, 'B': 15}",
                        "additionalProperties": {"type": "number"},
                    },
                    "auto_fit": {
                        "type": "boolean",
                        "description": "是否自动适配所有列宽",
                        "default": False,
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=adjust_column_width,
        ),
        ToolDef(
            name="read_cell_styles",
            description="读取 Excel 单元格范围的样式信息（字体、颜色、填充、边框、对齐、合并状态），用于感知现有格式",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "单元格范围，如 'A1:C10' 或 'A1'",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                    "summary_only": {
                        "type": "boolean",
                        "description": "仅返回样式统计汇总（使用了哪些颜色/字体/边框），不返回逐单元格明细",
                        "default": False,
                    },
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=read_cell_styles,
        ),
        ToolDef(
            name="adjust_row_height",
            description="调整 Excel 行高：支持手动指定高度或自动适配",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "rows": {
                        "type": "object",
                        "description": "行高映射，如 {'1': 30, '2': 25}（键为行号字符串）",
                        "additionalProperties": {"type": "number"},
                    },
                    "auto_fit": {
                        "type": "boolean",
                        "description": "是否自动适配所有行高",
                        "default": False,
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=adjust_row_height,
        ),
        ToolDef(
            name="merge_cells",
            description="合并 Excel 指定范围的单元格",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "要合并的单元格范围，如 'A1:C1'",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=merge_cells_tool,
        ),
        ToolDef(
            name="unmerge_cells",
            description="取消合并 Excel 指定范围的单元格",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "要取消合并的单元格范围，如 'A1:C1'",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=unmerge_cells_tool,
        ),
    ]
