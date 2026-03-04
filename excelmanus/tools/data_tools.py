"""数据工具：提供 Excel 读写、分析、过滤和转换能力。"""

from __future__ import annotations

import functools
import json
import shutil
from datetime import date, datetime
from typing import Any

import pandas as pd

_builtin_range = range  # 保存内置 range，避免被同名函数参数遮蔽

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools._guard_ctx import get_guard as _get_ctx_guard
from excelmanus.tools._helpers import check_file_exists, get_worksheet, resolve_sheet_name
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.data")

# ── 表头识别配置 ──────────────────────────────────────────

_HEADER_MIN_NON_EMPTY = 3
_HEADER_SCAN_ROWS = 30
_HEADER_SCAN_COLS = 200
_HEADER_KEYWORDS = (
    "月份",
    "日期",
    "时间",
    "城市",
    "地区",
    "产品",
    "部门",
    "姓名",
    "工号",
    "编号",
    "金额",
    "数量",
    "状态",
    "营收",
    "利润",
    "成本",
)
_TITLE_HINT_PREFIXES = (
    "生成时间",
    "汇总",
    "分析",
    "报表",
    "仪表盘",
    "机密",
)

# ── 表单类文档识别配置 ────────────────────────────────────
# 表单类文档特征：大量标签-值对，标签行占比高，大量合并单元格
_FORM_LABEL_KEYWORDS = frozenset({
    "交款人", "交款单位", "联系人", "联系方式", "付款账户", "付款人",
    "付款事由", "付款方式", "付款金额", "其他金额", "费用合计", "大写金额",
    "备注", "说明", "日期", "编号", "单位", "姓名", "电话", "地址",
    "客户", "供应商", "发票", "账号", "开户行",
})
_FORM_LABEL_RATIO_THRESHOLD = 0.3  # 标签行占比超过此阈值认为是表单类文档
_FORM_MERGED_CELL_RATIO_THRESHOLD = 0.15  # 合并单元格占比超过此阈值认为是表单类文档

# ── 合并单元格摘要配置 ─────────────────────────────────────
_MERGED_SUMMARY_MAX_SPANS = 20  # 摘要中最多列出的合并区域数


def _collect_merged_cell_summary(ws: Any) -> dict[str, Any] | None:
    """收集工作表的合并单元格摘要信息。

    将合并区域按语义角色分类（标题跨列、列组标头、数据区合并），
    并计算合并单元格占比，附带处理建议。

    Args:
        ws: openpyxl Worksheet 对象（非 read_only 模式）。

    Returns:
        合并摘要字典，无合并时返回 None。
    """
    from openpyxl.utils import get_column_letter

    merged_ranges = list(ws.merged_cells.ranges)
    if not merged_ranges:
        return None

    total_rows = ws.max_row or 1
    total_cols = ws.max_column or 1
    total_cells = total_rows * total_cols

    # 统计合并单元格总数
    merged_cell_count = 0
    for mr in merged_ranges:
        merged_cell_count += (mr.max_row - mr.min_row + 1) * (mr.max_col - mr.min_col + 1)
    merged_ratio = merged_cell_count / max(total_cells, 1)

    # 分类合并区域
    header_spans: list[str] = []       # 宽跨列（标题/分组标头）
    column_group_spans: list[str] = []  # 列组标头（如 "星期一" 跨若干列）
    data_merged_count = 0               # 数据区合并（跨行，暗示 NaN）

    for mr in merged_ranges:
        col_span = mr.max_col - mr.min_col + 1
        row_span = mr.max_row - mr.min_row + 1
        start_col_letter = get_column_letter(mr.min_col)
        end_col_letter = get_column_letter(mr.max_col)

        # 读取合并区域左上角值
        top_left_value = ws.cell(row=mr.min_row, column=mr.min_col).value
        label = f"'{top_left_value}'" if top_left_value else "(空)"

        range_str = str(mr)

        if col_span > total_cols * 0.5:
            # 宽跨列：跨度超过总列数 50%，通常是标题行
            header_spans.append(f"{range_str} → {label}")
        elif col_span >= 2 and row_span <= 2 and mr.min_row <= 5:
            # 列组标头：前 5 行内、跨 2+ 列但不太宽，通常是分组标头
            column_group_spans.append(
                f"{range_str} → {label} (cols {start_col_letter}:{end_col_letter})"
            )
        elif row_span >= 2:
            # 数据区跨行合并：pandas 读取时仅首格有值，其余为 NaN
            data_merged_count += 1

    summary: dict[str, Any] = {
        "merged_range_count": len(merged_ranges),
        "merged_cell_ratio": f"{merged_ratio:.1%}",
    }

    if header_spans:
        summary["header_spans"] = header_spans[:_MERGED_SUMMARY_MAX_SPANS]
    if column_group_spans:
        summary["column_group_spans"] = column_group_spans[:_MERGED_SUMMARY_MAX_SPANS]
    if data_merged_count > 0:
        summary["data_merged_ranges"] = data_merged_count
        summary["hint"] = (
            f"数据区存在 {data_merged_count} 处跨行合并，"
            "pandas 读取时仅合并区域左上角单元格有值，其余为 NaN。"
            "建议用 openpyxl ws.merged_cells.ranges 获取合并信息后做值传播（forward-fill）。"
        )

    return summary


# ── CSV/TSV 支持 ──────────────────────────────────────────

_CSV_EXTENSIONS: frozenset[str] = frozenset({".csv", ".tsv", ".txt"})

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "data"
SKILL_DESCRIPTION = "Excel 数据操作工具集：读取、写入、分析、过滤和转换"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard（优先 per-session contextvar）。"""
    ctx_guard = _get_ctx_guard()
    if ctx_guard is not None:
        return ctx_guard
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def build_completeness_meta(
    total_available: int,
    returned: int,
    *,
    entity_name: str = "行",
) -> dict[str, Any]:
    """构建数据完整性元数据，供工具统一使用。

    当 returned < total_available 时，附加截断标记和自然语言提示，
    帮助 LLM 正确理解数据范围，避免将预览数据误判为全量数据。
    """
    meta: dict[str, Any] = {
        "total_available": total_available,
        "returned": returned,
    }
    if returned < total_available:
        meta["is_truncated"] = True
        meta["truncation_note"] = (
            f"⚠️ 仅返回 {returned} {entity_name}（共 {total_available} {entity_name}）。"
            f"如需操作全量数据，请注意实际数据范围。"
        )
    return meta


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


def _is_date_like(value: Any) -> bool:
    """判断值是否为日期/时间类型。"""
    return isinstance(value, (date, datetime, pd.Timestamp))


def _normalize_cell(value: Any) -> Any:
    """将单元格值规范化为便于表头检测的格式。"""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def _is_csv_file(path: Any) -> bool:
    """判断文件是否为 CSV/TSV 格式。"""
    from pathlib import Path

    p = Path(path) if not isinstance(path, Path) else path
    return p.suffix.lower() in _CSV_EXTENSIONS


def _serialize_cell_value(value: Any) -> Any:
    """将单元格值序列化为 JSON 兼容类型。"""
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _trim_trailing_nulls_generic(row: list[Any]) -> list[Any]:
    """裁剪尾部空值，减少噪音列影响。"""
    end = len(row)
    while end > 0 and row[end - 1] is None:
        end -= 1
    return row[:end]


def _looks_like_title_row(first_cell: Any) -> bool:
    """判断首单元格是否更像标题而非字段名。"""
    if not isinstance(first_cell, str):
        return False
    text = first_cell.strip()
    if not text:
        return False
    if any(token in text for token in ("──", "——", "年度", "季度")):
        return True
    return any(hint in text for hint in _TITLE_HINT_PREFIXES)


def _is_form_type_document(
    safe_path: Any,
    sheet_name: str | None,
    max_scan: int = _HEADER_SCAN_ROWS,
) -> tuple[bool, str]:
    """检测是否为表单类文档（如收据、模板等非标准数据表格）。

    表单类文档特征：
    1. 大量标签-值对（标签行占比高）
    2. 大量合并单元格
    3. 缺少标准列头

    Returns:
        (is_form_document, reason) 元组
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(safe_path, read_only=False, data_only=True)
    except Exception:
        return False, ""

    try:
        if sheet_name:
            resolved = resolve_sheet_name(sheet_name, wb.sheetnames)
            ws = wb[resolved] if resolved else wb.active
        else:
            ws = wb.active
        if ws is None:
            return False, ""

        max_row = min(max_scan, ws.max_row or max_scan)
        max_col = ws.max_column or 10

        # 统计合并单元格占比
        total_cells = max_row * max_col
        merged_cells = 0
        for merged_range in ws.merged_cells.ranges:
            merged_cells += (merged_range.max_row - merged_range.min_row + 1) * \
                           (merged_range.max_col - merged_range.min_col + 1)

        merged_ratio = merged_cells / max(total_cells, 1)

        # 统计标签行（包含表单标签关键词的非空行）占比
        label_rows = 0
        total_scannable_rows = 0

        for row in ws.iter_rows(min_row=1, max_row=max_scan, min_col=1, max_col=max_col, values_only=True):
            row_values = [_normalize_cell(c) for c in row]
            non_empty = [v for v in row_values if v is not None]

            if len(non_empty) >= 2:  # 至少2个非空单元格才计入
                total_scannable_rows += 1
                # 检查是否包含表单标签关键词（精确匹配单元格值，避免数据行误判）
                cell_texts = {str(v).strip() for v in non_empty if isinstance(v, str)}
                if any(ct in _FORM_LABEL_KEYWORDS for ct in cell_texts):
                    label_rows += 1

        label_ratio = label_rows / max(total_scannable_rows, 1)

        # 判断逻辑
        if merged_ratio > _FORM_MERGED_CELL_RATIO_THRESHOLD:
            return True, f"合并单元格占比 {merged_ratio:.1%} 超过阈值 {_FORM_MERGED_CELL_RATIO_THRESHOLD:.1%}"

        if label_ratio > _FORM_LABEL_RATIO_THRESHOLD:
            return True, f"表单标签行占比 {label_ratio:.1%} 超过阈值 {_FORM_LABEL_RATIO_THRESHOLD:.1%}"

        return False, ""
    finally:
        wb.close()


def _header_row_score(
    row_values: list[Any],
    row_idx0: int,
    next_row_values: list[Any] | None = None,
) -> float:
    """计算候选表头行分数。分数越高越可能是表头。"""
    non_empty = [v for v in row_values if v is not None]
    if len(non_empty) < _HEADER_MIN_NON_EMPTY:
        return float("-inf")

    text_values = [str(v).strip() for v in non_empty if isinstance(v, str) and str(v).strip()]
    numeric_count = sum(1 for v in non_empty if isinstance(v, (int, float)))
    date_count = sum(1 for v in non_empty if _is_date_like(v))
    string_count = len(text_values)
    unique_ratio = len(set(map(str, non_empty))) / max(len(non_empty), 1)
    keyword_hits = sum(
        1
        for text in text_values
        if any(k in text for k in _HEADER_KEYWORDS)
    )

    score = 0.0
    score += len(non_empty) * 2.0
    score += string_count * 1.6
    score -= numeric_count * 1.4
    score -= date_count * 1.0
    score += unique_ratio * 2.5
    score += keyword_hits * 2.8
    score -= row_idx0 * 0.03  # 轻微偏好靠前行

    first_cell = row_values[0] if row_values else None
    if _looks_like_title_row(first_cell):
        score -= 6.0

    if next_row_values is not None:
        next_non_empty = [v for v in next_row_values if v is not None]
        if next_non_empty:
            next_numeric_ratio = sum(1 for v in next_non_empty if isinstance(v, (int, float))) / len(next_non_empty)
            # 表头下一行常见“数据占比更高”
            score += next_numeric_ratio * 1.8
    return score


def _guess_header_row_from_rows(rows: list[list[Any]], *, max_scan: int | None = None, skip_rows: set[int] | None = None) -> int | None:
    """基于抽样行猜测 header 行号（0-indexed）。"""
    if not rows:
        return None

    upper = len(rows) if max_scan is None else min(len(rows), max_scan)
    best_row: int | None = None
    best_score = float("-inf")

    for idx in range(upper):
        if skip_rows and idx in skip_rows:
            continue
        row = _trim_trailing_nulls_generic(rows[idx])
        next_row = _trim_trailing_nulls_generic(rows[idx + 1]) if idx + 1 < upper else None
        score = _header_row_score(row, idx, next_row)
        if score > best_score:
            best_score = score
            best_row = idx

    if best_row is None or best_score == float("-inf"):
        return None
    return best_row

def _detect_header_row(
    safe_path: Any,
    sheet_name: str | None,
    max_scan: int = _HEADER_SCAN_ROWS,
    max_scan_columns: int = _HEADER_SCAN_COLS,
) -> int | None:
    """启发式检测 header 行号（0-indexed）。

    策略：
    1. 首先检测是否为表单类文档（收据、模板等），如果是则返回 -1 表示无需 header
    2. 扫描前 N 行（默认 30）和前 M 列（默认 200）；
    3. 对每一行按"文本占比、关键字、唯一性、数据行特征"打分；
    4. 选择分数最高者作为表头。

    Returns:
        检测到的 header 行号（从0开始），无法确定时返回 None。
        返回 -1 表示检测为表单类文档，不应使用 header。
    """
    # 首先检测是否为表单类文档
    is_form, reason = _is_form_type_document(safe_path, sheet_name, max_scan)
    if is_form:
        logger.info("检测为表单类文档：%s", reason)
        return -1  # 特殊标记：表单类文档，不使用 header

    try:
        from openpyxl import load_workbook
        wb = load_workbook(safe_path, read_only=False, data_only=True)
    except Exception:
        return None

    try:
        if sheet_name:
            resolved = resolve_sheet_name(sheet_name, wb.sheetnames)
            if resolved:
                ws = wb[resolved]
            else:
                ws = wb.active
        else:
            ws = wb.active
        if ws is None:
            return None

        # 收集宽合并行（列跨度 > 50% 总列数）
        scan_cols = max(1, min(max_scan_columns, ws.max_column or max_scan_columns))
        wide_merged_rows: set[int] = set()
        for merged_range in ws.merged_cells.ranges:
            col_span = merged_range.max_col - merged_range.min_col + 1
            if col_span > scan_cols * 0.5:
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    if r <= max_scan:
                        wide_merged_rows.add(r - 1)  # 转为 0-indexed

        rows: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_scan,
            min_col=1,
            max_col=scan_cols,
            values_only=True,
        ):
            rows.append([_normalize_cell(c) for c in row])

        if not rows:
            return None

        return _guess_header_row_from_rows(rows, max_scan=max_scan, skip_rows=wide_merged_rows)
    finally:
        wb.close()


def _build_read_kwargs(
    safe_path: Any,
    sheet_name: str | None,
    max_rows: int | None = None,
    header_row: int | None = None,
) -> dict[str, Any]:
    """构建 pd.read_excel 的公共参数，统一处理 header_row。

    Args:
        safe_path: 已校验的文件路径。
        sheet_name: 工作表名称。
        max_rows: 最大读取行数。
        header_row: 列头所在行号（从0开始），默认自动检测。
            不传此参数时工具会启发式检测真正的表头行；
            仅在自动检测不准确时才显式指定。
            特殊值 -1 表示表单类文档，不使用 header（header=None）。

    Returns:
        可直接传给 pd.read_excel 的关键字参数字典。
        包含特殊键 "_form_type_document" 表示是否被检测为表单类文档。
    """
    kwargs: dict[str, Any] = {"io": safe_path}
    if sheet_name is not None:
        kwargs["sheet_name"] = sheet_name
    if max_rows is not None:
        kwargs["nrows"] = max_rows
    if header_row is not None:
        # header_row=-1 表示表单类文档，不使用 header
        if header_row == -1:
            kwargs["header"] = None
            kwargs["_form_type_document"] = True
            logger.info("检测为表单类文档，使用 header=None 读取 (sheet=%s)", sheet_name)
        else:
            kwargs["header"] = header_row
    else:
        # 启发式自动检测 header 行（仅当用户未显式指定时）
        detected = _detect_header_row(safe_path, sheet_name)
        if detected is not None:
            if detected == -1:
                # 表单类文档，不使用 header
                kwargs["header"] = None
                kwargs["_form_type_document"] = True
                logger.info("自动检测为表单类文档，使用 header=None 读取 (sheet=%s)", sheet_name)
            elif detected > 0:
                kwargs["header"] = detected
                logger.info("自动检测 header_row=%d (sheet=%s)", detected, sheet_name)
    return kwargs


def _resolve_formula_columns(
    df: pd.DataFrame,
    safe_path: Any,
    sheet_name: str | None,
    header_row: int,
) -> pd.DataFrame:
    """对公式列进行求值：当 data_only 模式读到全 NaN 时，
    尝试解析公式并用已有列数据计算。

    仅处理同行引用的简单算术公式（如 =G4*H4, =I4*(1-J4)）。
    复杂公式（跨行引用、函数调用等）会被静默跳过。
    """
    import re

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    formula_meta: dict[str, Any] = {
        "resolved_columns": [],
        "unresolved_columns": [],
        "unresolved_details": {},
    }
    if df.empty:
        df.attrs["formula_resolution"] = formula_meta
        return df

    # 找出全 NaN 的列
    nan_cols_idx = [
        i for i, col in enumerate(df.columns)
        if df[col].isna().all()
    ]
    if not nan_cols_idx:
        df.attrs["formula_resolution"] = formula_meta
        return df

    wb = load_workbook(safe_path, data_only=False, read_only=True)
    try:
        try:
            ws = get_worksheet(wb, sheet_name)
        except ValueError:
            df.attrs["formula_resolution"] = formula_meta
            return df

        excel_header_row = header_row + 1  # 0-indexed → 1-indexed
        first_data_row = excel_header_row + 1

        # 列字母 → DataFrame 列索引映射
        col_names = list(df.columns)
        letter_to_idx: dict[str, int] = {}
        for i in range(len(col_names)):
            letter_to_idx[get_column_letter(i + 1)] = i

        cell_ref_pattern = re.compile(r'([A-Z]+)(\d+)')
        data_row_str = str(first_data_row)

        resolved_cols: set[str] = set()
        unresolved: dict[str, str] = {}

        for col_idx in nan_cols_idx:
            col_name = str(col_names[col_idx])
            cell = ws.cell(row=first_data_row, column=col_idx + 1)
            formula = cell.value
            if not isinstance(formula, str) or not formula.startswith('='):
                continue

            formula_body = formula[1:]

            # 提取所有单元格引用
            refs = cell_ref_pattern.findall(formula_body)
            if not refs:
                unresolved[col_name] = "不支持的公式结构（无单元格引用）"
                continue

            # 仅处理同行引用
            if not all(row_num == data_row_str for _, row_num in refs):
                unresolved[col_name] = "不支持跨行/跨区引用公式"
                continue

            # 检查所有引用列是否存在
            all_valid = True
            for letter, _ in refs:
                if letter not in letter_to_idx:
                    all_valid = False
                    break
            if not all_valid:
                unresolved[col_name] = "公式引用超出可读列范围"
                continue

            # 构建求值表达式：将单元格引用替换为变量名
            expr = formula_body
            namespace: dict[str, Any] = {}
            # 按字母长度降序替换，避免 A 替换 AA 的子串问题
            sorted_refs = sorted(set(refs), key=lambda r: (-len(r[0]), r[0]))
            for letter, row_num in sorted_refs:
                ref_idx = letter_to_idx[letter]
                var_name = f'_c{ref_idx}'
                namespace[var_name] = df.iloc[:, ref_idx]
                expr = expr.replace(f'{letter}{row_num}', var_name)

            try:
                result = eval(expr, {"__builtins__": {}}, namespace)  # noqa: S307
                if isinstance(result, pd.Series):
                    df.iloc[:, col_idx] = result
                    resolved_cols.add(col_name)
                    unresolved.pop(col_name, None)
                    logger.info(
                        "公式列 '%s' 求值成功 (formula=%s)",
                        col_names[col_idx], formula,
                    )
                else:
                    unresolved[col_name] = "公式求值未返回可用序列"
            except Exception:
                unresolved[col_name] = "公式求值失败"
                logger.debug(
                    "公式列 '%s' 求值失败 (formula=%s), 已跳过",
                    col_names[col_idx], formula,
                )
                continue
    finally:
        wb.close()

    formula_meta["resolved_columns"] = sorted(resolved_cols)
    formula_meta["unresolved_columns"] = sorted(unresolved)
    formula_meta["unresolved_details"] = unresolved
    df.attrs["formula_resolution"] = formula_meta

    return df


def _get_sheet_total_rows(safe_path: Any, sheet_name: str | None) -> int | None:
    """快速获取 sheet/CSV 总行数（不加载全部数据）。"""
    if _is_csv_file(safe_path):
        try:
            with open(safe_path, encoding="utf-8", errors="replace") as f:
                return max(sum(1 for _ in f) - 1, 0)  # 减去 header 行
        except Exception:
            return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(safe_path, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            return ws.max_row or 0
        finally:
            wb.close()
    except Exception:
        return None


def _read_csv_df(
    safe_path: Any,
    max_rows: int | None = None,
    header_row: int | None = None,
) -> tuple[pd.DataFrame, int]:
    """读取 CSV/TSV 文件为 DataFrame。"""
    from pathlib import Path

    p = Path(safe_path) if not isinstance(safe_path, Path) else safe_path
    sep = "\t" if p.suffix.lower() == ".tsv" else ","
    kwargs: dict[str, Any] = {"filepath_or_buffer": safe_path, "sep": sep}
    if header_row is not None:
        kwargs["header"] = header_row
    if max_rows is not None:
        kwargs["nrows"] = max_rows
    effective_header = header_row if header_row is not None else 0
    df = pd.read_csv(**kwargs)
    return df, effective_header


def _read_df(
    safe_path: Any,
    sheet_name: str | None,
    max_rows: int | None = None,
    header_row: int | None = None,
) -> tuple[pd.DataFrame, int]:
    """统一读取 Excel/CSV 为 DataFrame，含 header 自动检测 + 公式列求值。

    当自动检测的 header_row 导致超过 50% 列名为 Unnamed 时，
    自动向下尝试最多 5 行寻找更合理的表头。

    当检测为表单类文档时（header_row=-1），使用 header=None 读取全部数据，
    不执行 Unnamed 回退逻辑。

    Returns:
        (DataFrame, effective_header_row) 元组。
        effective_header 为 -1 表示表单类文档，None 表示使用默认 header=0。
    """
    # CSV/TSV 走专用路径
    if _is_csv_file(safe_path):
        return _read_csv_df(safe_path, max_rows=max_rows, header_row=header_row)

    kwargs = _build_read_kwargs(safe_path, sheet_name, max_rows=max_rows, header_row=header_row)
    # 注意=None 时 kwargs.get("header") 返回 None，不是 0
    # 我们需要区分：用户指定 header=None（不使用header）和表单类文档（header=None）
    # 通过检查 kwargs 中是否有特殊的标记来区分
    effective_header = kwargs.get("header")
    if effective_header is None:
        # header=None 可能是用户指定，也可能是表单类文档
        # 需要检查是否是被自动检测为表单类文档
        if kwargs.get("_form_type_document"):
            effective_header = -1
        else:
            effective_header = 0  # 用户显式指定 header=None，使用默认值
    # 移除内部标记，避免传给 pd.read_excel
    kwargs.pop("_form_type_document", None)
    df = pd.read_excel(**kwargs)

    # 表单类文档（header=-1）不使用 Unnamed 回退逻辑
    # 仅在自动检测模式下（用户未显式指定 header_row）执行 Unnamed 回退
    if header_row is None and effective_header != -1:
        unnamed_ratio = (
            sum(1 for c in df.columns if str(c).startswith("Unnamed"))
            / max(len(df.columns), 1)
        )
        if unnamed_ratio > 0.5:
            logger.info(
                "自动检测 header_row=%d 产生 %.0f%% Unnamed 列名，尝试回退",
                effective_header, unnamed_ratio * 100,
            )
            for try_header in range(effective_header + 1, min(effective_header + 6, 30)):
                retry_kwargs = {**kwargs, "header": try_header}
                try:
                    df_retry = pd.read_excel(**retry_kwargs)
                except Exception:
                    break
                if df_retry.empty:
                    break
                retry_unnamed = sum(
                    1 for c in df_retry.columns if str(c).startswith("Unnamed")
                )
                if retry_unnamed / max(len(df_retry.columns), 1) < 0.3:
                    logger.info("回退成功：header_row=%d → %d", effective_header, try_header)
                    df = df_retry
                    effective_header = try_header
                    break

    df = _resolve_formula_columns(df, safe_path, sheet_name, effective_header)
    return df, effective_header


# ── 工具函数 ──────────────────────────────────────────────


def _read_range_direct(
    safe_path: Any,
    sheet_name: str | None,
    cell_range: str,
) -> dict[str, Any]:
    """用 openpyxl read_only 模式读取指定坐标范围的原始单元格值。"""
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(safe_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            resolved = resolve_sheet_name(sheet_name, wb.sheetnames)
            ws = wb[resolved] if resolved else wb.active
        else:
            ws = wb.active
        if ws is None:
            return {"error": "无法打开工作表"}

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        rows: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        ):
            rows.append(_trim_trailing_nulls_generic([_serialize_cell_value(c) for c in row]))

        return {
            "range": cell_range,
            "start_row": min_row,
            "end_row": max_row,
            "rows_count": len(rows),
            "columns_count": max_col - min_col + 1,
            "data": rows,
        }
    finally:
        wb.close()


def read_excel(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int | None = None,
    include_style_summary: bool = False,
    header_row: int | None = None,
    include: list[str] | None = None,
    max_style_scan_rows: int = 200,
    range: str | None = None,
    offset: int | None = None,
    sample_rows: int | None = None,
) -> str:
    """读取 Excel/CSV 文件并返回数据摘要，可通过 include 按需附加额外维度。

    Args:
        file_path: Excel/CSV 文件路径（相对或绝对）。支持 .xlsx/.xls/.xlsm/.xlsb/.csv/.tsv。
        sheet_name: 工作表名称，默认读取第一个（CSV 时忽略）。
        max_rows: 最大读取行数，默认全部读取。
        include_style_summary: 是否附带样式概览（已废弃，请用 include=["styles"]）。
        header_row: 列头所在行号（从0开始），默认自动检测。
            当工作表有合并标题行时，需指定真正的列头行号。
        include: 按需请求的额外维度列表。可选值：
            styles — 压缩样式类（Style Classes + cell_style_map + merged_ranges）
            charts — 嵌入图表元信息
            images — 嵌入图片元信息
            freeze_panes — 冻结窗格位置
            conditional_formatting — 条件格式规则
            data_validation — 数据验证规则
            print_settings — 打印设置
            column_widths — 非默认列宽
            formulas — 含公式的单元格
            categorical_summary — 分类列的 value_counts（unique 值 < 阈值的列）
            summary — 每列数据质量概要（null 率、unique 数、min/max、高频值）
            vba — VBA 宏信息（仅 .xlsm 文件有效，含模块列表及可选源码）
        max_style_scan_rows: styles/formulas 维度扫描的最大行数，默认 200。
        range: Excel 坐标范围（如 "A1:F20"、"B100:D200"），指定后进入精确读取模式，
            绕过 pandas 直接用 openpyxl 读取指定区域，大文件友好。不支持 CSV。
        offset: 数据行偏移（从0开始，header 之后起算），与 max_rows 组合实现分页。
        sample_rows: 等距采样行数，用于了解大表数据分布。

    Returns:
        JSON 格式的数据摘要字符串。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    not_found = check_file_exists(safe_path, file_path, guard)
    if not_found is not None:
        return not_found

    # .xls/.xlsb → 透明转换为 xlsx（后续 openpyxl 调用统一走 xlsx）
    from excelmanus.tools._helpers import ensure_openpyxl_compatible, check_sheet_name
    safe_path = ensure_openpyxl_compatible(safe_path)

    # ── sheet 名验证：提前拦截无效 sheet 名，附带可用列表 ──
    if sheet_name is not None and not _is_csv_file(safe_path):
        resolved_sheet, sheet_err = check_sheet_name(safe_path, sheet_name)
        if sheet_err is not None:
            return sheet_err
        sheet_name = resolved_sheet  # 可能经过 case-insensitive 修正

    # ── range 模式：精确读取指定坐标范围 ──
    if range is not None:
        if _is_csv_file(safe_path):
            return json.dumps(
                {"status": "error", "message": "range 参数不支持 CSV 文件，请使用 offset + max_rows 分页读取"},
                ensure_ascii=False,
            )
        result = _read_range_direct(safe_path, sheet_name, range)
        result["file"] = str(safe_path.name)
        return json.dumps(result, ensure_ascii=False, indent=2, default=str)

    # ── 标准模式 ──
    # offset 调整：读取 offset + max_rows 行再切片
    effective_max_rows = max_rows
    if offset is not None and offset > 0 and max_rows is not None:
        effective_max_rows = offset + max_rows

    df, effective_header = _read_df(safe_path, sheet_name, max_rows=effective_max_rows, header_row=header_row)

    # 应用 offset 切片
    if offset is not None and offset > 0:
        df = df.iloc[offset:].reset_index(drop=True)

    # 当 max_rows/offset 限制了读取范围时，获取 sheet 实际总行数
    total_rows_in_sheet: int | None = None
    if max_rows is not None or (offset is not None and offset > 0):
        total_rows_in_sheet = _get_sheet_total_rows(safe_path, sheet_name)

    # 构建摘要信息
    summary: dict[str, Any] = {
        "file": str(safe_path.name),
        "shape": {"rows": df.shape[0], "columns": df.shape[1]},
    }

    # 数据完整性指示：截断元数据放在 columns/preview 之前，确保即使被引擎层截断也能保留
    if total_rows_in_sheet is not None and total_rows_in_sheet > df.shape[0]:
        completeness = build_completeness_meta(
            total_available=total_rows_in_sheet,
            returned=df.shape[0],
        )
        summary["total_rows_in_sheet"] = total_rows_in_sheet
        summary["is_truncated"] = completeness.get("is_truncated", False)
        summary["truncation_note"] = completeness.get("truncation_note", "")

    summary["columns"] = [str(c) for c in df.columns]
    summary["dtypes"] = {str(col): str(dtype) for col, dtype in df.dtypes.items()}
    _null_info = _build_null_info(df)
    if _null_info:
        summary["null_info"] = _null_info
    summary["preview"] = _df_to_compact_records(df.head(10))

    # 自动 tail 预览：表格 > 20 行时附加最后 5 行
    if df.shape[0] > 20:
        tail_start = df.shape[0] - 5
        summary["tail_preview"] = _df_to_compact_records(df.tail(5))
        summary["tail_note"] = f"显示最后 5 行（第 {tail_start + 1}~{df.shape[0]} 行）"

    # 等距采样：sample_rows 指定时附加采样数据
    if sample_rows is not None and sample_rows > 0 and len(df) > sample_rows:
        step = max(1, len(df) // sample_rows)
        indices = list(_builtin_range(0, len(df), step))[:sample_rows]
        sampled_df = df.iloc[indices]
        summary["sample_preview"] = _df_to_compact_records(sampled_df)
        summary["sample_note"] = f"等距采样 {len(indices)} 行（共 {len(df)} 行，间隔 {step}）"

    formula_meta = df.attrs.get("formula_resolution")
    if isinstance(formula_meta, dict):
        if formula_meta.get("resolved_columns") or formula_meta.get("unresolved_columns"):
            summary["formula_resolution"] = formula_meta

    # 当 header_row 被自动检测时，告知 LLM 实际使用的行号
    # effective_header=-1 表示检测为表单类文档，使用 header=None
    if header_row is None and effective_header != 0:
        if effective_header == -1:
            summary["detected_form_type"] = True
            summary["detected_header_row"] = "form_type_document"
        else:
            summary["detected_header_row"] = effective_header

    # Unnamed 列名警告：提醒 LLM 列名不可靠，建议指定 header_row
    # 表单类文档不使用此警告
    unnamed_cols = [str(c) for c in df.columns if str(c).startswith("Unnamed")]
    if unnamed_cols and effective_header != -1:
        summary["unnamed_columns_warning"] = (
            f"检测到 {len(unnamed_cols)} 个 Unnamed 列名（共 {len(df.columns)} 列），"
            f"可能是合并标题行导致。建议使用 header_row 参数指定真正的列头行号重新读取。"
        )

    # 合并单元格警告：高合并率时提醒 LLM 注意值传播
    if not _is_csv_file(safe_path):
        try:
            from openpyxl import load_workbook as _lw
            _wb_mc = _lw(safe_path, read_only=False, data_only=True)
            try:
                _ws_mc = (
                    _wb_mc[sheet_name]
                    if sheet_name and sheet_name in _wb_mc.sheetnames
                    else _wb_mc.active
                )
                if _ws_mc is not None:
                    _mc_summary = _collect_merged_cell_summary(_ws_mc)
                    if _mc_summary:
                        summary["merged_cell_summary"] = _mc_summary
            finally:
                _wb_mc.close()
        except Exception:
            pass

    # 向后兼容：include_style_summary=True 映射为 include=["styles"]
    include_set: set[str] = set()
    if include:
        include_set.update(include)
    if include_style_summary and "styles" not in include_set:
        include_set.add("styles")

    # 校验 include 维度
    invalid_dims = include_set - set(INCLUDE_DIMENSIONS)
    if invalid_dims:
        summary["include_warning"] = f"未知的 include 维度已忽略: {sorted(invalid_dims)}"
        include_set -= invalid_dims

    # 分发基于 DataFrame 的 include 维度（不需要 openpyxl）
    if "categorical_summary" in include_set:
        summary["categorical_summary"] = _collect_categorical_summary(df)
        include_set.discard("categorical_summary")

    if "summary" in include_set:
        summary["data_summary"] = _collect_data_summary(df)
        include_set.discard("summary")

    # vba 维度：基于文件级别，不依赖 worksheet
    if "vba" in include_set:
        summary["vba"] = _collect_vba_info(safe_path)
        include_set.discard("vba")

    # 分发 include 维度采集（需要用 openpyxl 打开，CSV 不支持）
    if include_set and not _is_csv_file(safe_path):
        from openpyxl import load_workbook

        # styles/charts/images/freeze_panes 等需要非 data_only 模式
        # formulas 也需要非 data_only 模式以读取公式文本
        needs_formulas = "formulas" in include_set
        wb_include = load_workbook(safe_path, data_only=not needs_formulas)
        try:
            ws_include = (
                wb_include[sheet_name]
                if sheet_name and sheet_name in wb_include.sheetnames
                else wb_include.active
            )
            extra = _dispatch_include_dimensions(ws_include, include_set, max_style_scan_rows)
            summary.update(extra)
        finally:
            wb_include.close()

    return json.dumps(summary, ensure_ascii=False, separators=(',', ':'), default=str)



def write_excel(file_path: str, data: list[dict], sheet_name: str = "Sheet1") -> str:
    """将数据写入 Excel 文件。

    当目标文件已存在时，仅写入/替换指定工作表，保留其他工作表。
    当目标文件不存在时，创建新文件。

    注意：仅推荐用于覆盖写入整表或简单追加。跨表匹配、复杂清洗转换、
    超过3行的条件更新等场景，请优先使用 run_code 工具（pandas）。

    Args:
        file_path: 目标 Excel 文件路径。
        data: 要写入的数据，每个字典代表一行。
        sheet_name: 工作表名称，默认 Sheet1。

    Returns:
        操作结果描述。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    df = pd.DataFrame(data)

    if safe_path.exists() and safe_path.suffix.lower() in (".xlsx", ".xlsm"):
        # 已有文件：使用 append 模式，仅替换指定 sheet，保留其他 sheet
        writer_kwargs: dict[str, Any] = {
            "engine": "openpyxl",
            "mode": "a",
            "if_sheet_exists": "replace",
        }
        if safe_path.suffix.lower() == ".xlsm":
            writer_kwargs["engine_kwargs"] = {"keep_vba": True}
        with pd.ExcelWriter(safe_path, **writer_kwargs) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # 新文件：直接写入
        df.to_excel(safe_path, sheet_name=sheet_name, index=False)

    return json.dumps(
        {"status": "success", "file": str(safe_path.name), "rows": len(df), "columns": len(df.columns)},
        ensure_ascii=False,
    )



def analyze_data(
    file_path: str,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> str:
    """对 Excel 数据进行基本统计分析。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称，默认第一个。
        header_row: 列头所在行号（从0开始），默认自动检测。

    Returns:
        JSON 格式的统计分析结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    df, _ = _read_df(safe_path, sheet_name, header_row=header_row)

    # 基本统计信息
    result: dict[str, Any] = {
        "file": str(safe_path.name),
        "shape": {"rows": df.shape[0], "columns": df.shape[1]},
        "columns": [str(c) for c in df.columns],
        "missing_values": {str(col): int(count) for col, count in df.isnull().sum().items() if count > 0},
    }

    # 数值列统计
    numeric_df = df.select_dtypes(include=["number"])
    if not numeric_df.empty:
        stats = numeric_df.describe().to_dict()
        # 将 numpy 类型转为 Python 原生类型
        result["numeric_stats"] = {
            str(col): {k: float(v) for k, v in col_stats.items()}
            for col, col_stats in stats.items()
        }

    return json.dumps(result, ensure_ascii=False, indent=2, default=str)




def filter_data(
    file_path: str,
    column: str | None = None,
    operator: str | None = None,
    value: Any = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
    columns: list[str] | None = None,
    conditions: list[dict[str, Any]] | None = None,
    logic: str = "and",
    max_rows: int | None = None,
    sort_by: str | None = None,
    ascending: bool = True,
    limit: int | None = None,
) -> str:
    """根据条件过滤 Excel 数据行并可选排序，支持单条件和多条件 AND/OR 组合。

    Args:
        file_path: Excel 文件路径。
        column: 要过滤的列名（单条件模式）。
        operator: 比较运算符（单条件模式）。支持：
            eq/ne/gt/ge/lt/le/contains/in/not_in/between/isnull/notnull/startswith/endswith
        value: 比较值（单条件模式）。
        sheet_name: 工作表名称，默认第一个。
        header_row: 列头所在行号（从0开始），默认自动检测。
        columns: 只返回指定列（投影），默认返回全部列。
        conditions: 多条件数组，每个元素为 {"column": str, "operator": str, "value": Any}。
        logic: 多条件组合方式，"and"（默认）或 "or"。
        max_rows: 最多返回的数据行数，默认返回全部。
        sort_by: 排序列名，默认不排序。
        ascending: 排序方向，默认升序。
        limit: 排序后限制返回行数，默认返回全部。

    Returns:
        JSON 格式的过滤结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    df, _ = _read_df(safe_path, sheet_name, header_row=header_row)

    ops = {
        "eq": lambda s, v: s == v,
        "ne": lambda s, v: s != v,
        "gt": lambda s, v: s > v,
        "ge": lambda s, v: s >= v,
        "lt": lambda s, v: s < v,
        "le": lambda s, v: s <= v,
        "contains": lambda s, v: s.astype(str).str.contains(str(v), na=False),
        "in": lambda s, v: s.isin(v if isinstance(v, list) else [v]),
        "not_in": lambda s, v: ~s.isin(v if isinstance(v, list) else [v]),
        "between": lambda s, v: s.between(v[0], v[1]) if isinstance(v, list) and len(v) >= 2 else pd.Series([False] * len(s), index=s.index),
        "isnull": lambda s, v: s.isna(),
        "notnull": lambda s, v: s.notna(),
        "startswith": lambda s, v: s.astype(str).str.startswith(str(v), na=False),
        "endswith": lambda s, v: s.astype(str).str.endswith(str(v), na=False),
    }

    # 构建条件列表：兼容单条件和多条件
    if conditions:
        cond_list = conditions
    elif column is not None and operator is not None:
        cond_list = [{"column": column, "operator": operator, "value": value}]
    else:
        return json.dumps(
            {"error": "请提供 column/operator/value 单条件，或 conditions 多条件数组"},
            ensure_ascii=False,
        )

    if logic not in ("and", "or"):
        return json.dumps(
            {"error": f"不支持的逻辑运算符 '{logic}'，支持: and, or"},
            ensure_ascii=False,
        )

    # 逐条件构建 mask
    masks = []
    for cond in cond_list:
        col = cond.get("column")
        op = cond.get("operator")
        val = cond.get("value")
        if col not in df.columns:
            return json.dumps(
                {"error": f"列 '{col}' 不存在，可用列: {[str(c) for c in df.columns]}"},
                ensure_ascii=False,
                default=str,
            )
        if op not in ops:
            return json.dumps(
                {"error": f"不支持的运算符 '{op}'，支持: {list(ops.keys())}"},
                ensure_ascii=False,
            )
        masks.append(ops[op](df[col], val))

    # 组合 mask
    if logic == "and":
        combined_mask = functools.reduce(lambda a, b: a & b, masks)
    else:
        combined_mask = functools.reduce(lambda a, b: a | b, masks)

    filtered = df[combined_mask]

    # 投影：只保留指定列
    if columns:
        valid_cols = [c for c in columns if c in filtered.columns]
        missing_cols = [c for c in columns if c not in filtered.columns]
        filtered = filtered[valid_cols]
    else:
        missing_cols = []

    # 排序
    if sort_by is not None:
        if sort_by not in filtered.columns:
            return json.dumps(
                {"error": f"排序列 '{sort_by}' 不存在，可用列: {[str(c) for c in filtered.columns]}"},
                ensure_ascii=False,
                default=str,
            )
        # 对排序列做数值转换以支持文本型数值的正确排序
        sort_key = _coerce_numeric(filtered[sort_by])
        filtered = filtered.assign(**{"__sort_key__": sort_key}).sort_values(
            by="__sort_key__", ascending=ascending, na_position="last", kind="mergesort"
        ).drop(columns=["__sort_key__"])

    # 排序后限制返回行数（limit）
    total_filtered = len(filtered)
    if limit is not None and limit > 0:
        filtered = filtered.head(limit)

    # max_rows 兜底限制
    if max_rows is not None and max_rows > 0:
        filtered = filtered.head(max_rows)

    result: dict[str, Any] = {
        "file": str(safe_path.name),
        "filters": cond_list,
        "logic": logic,
        "original_rows": len(df),
        "filtered_rows": total_filtered,
        "returned_rows": len(filtered),
        "columns": [str(c) for c in filtered.columns],
        "data": _df_to_compact_records(filtered),
    }
    if total_filtered > len(filtered):
        result["truncated"] = True
        result["note"] = f"结果已截断，共 {total_filtered} 条匹配，返回前 {len(filtered)} 条"
    if missing_cols:
        result["missing_columns"] = missing_cols

    return json.dumps(result, ensure_ascii=False, separators=(',', ':'), default=str)




def transform_data(
    file_path: str,
    operations: list[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
    header_row: int | None = None,
) -> str:
    """对 Excel 数据执行转换操作。

    支持的操作类型：
    - rename: 重命名列，参数 {"columns": {"旧名": "新名"}}
    - add_column: 添加新列，参数 {"name": "列名", "value": 值或表达式}
    - drop_columns: 删除列，参数 {"columns": ["列名1", "列名2"]}
    - sort: 排序，参数 {"by": "列名", "ascending": true/false}

    Args:
        file_path: 源 Excel 文件路径。
        operations: 转换操作列表，每项包含 type 和对应参数。
        sheet_name: 工作表名称，默认第一个。
        output_path: 输出文件路径，默认覆盖源文件。
        header_row: 列头所在行号（从0开始），默认自动检测。

    Returns:
        JSON 格式的转换结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    df, _ = _read_df(safe_path, sheet_name, header_row=header_row)

    applied: list[str] = []
    for op in operations:
        op_type = op.get("type", "")

        if op_type == "rename":
            columns_map = op.get("columns", {})
            df = df.rename(columns=columns_map)
            applied.append(f"rename: {columns_map}")

        elif op_type == "add_column":
            col_name = op.get("name", "")
            col_value = op.get("value", None)
            df[col_name] = col_value
            applied.append(f"add_column: {col_name}")

        elif op_type == "drop_columns":
            cols = op.get("columns", [])
            existing = [c for c in cols if c in df.columns]
            df = df.drop(columns=existing)
            applied.append(f"drop_columns: {existing}")

        elif op_type == "sort":
            by = op.get("by", "")
            ascending = op.get("ascending", True)
            if by in df.columns:
                df = df.sort_values(by=by, ascending=ascending)
                applied.append(f"sort: {by} {'asc' if ascending else 'desc'}")

        else:
            applied.append(f"unknown_op: {op_type}")

    def _resolve_target_sheet(default_path: Any, desired_sheet: str | None) -> str:
        if desired_sheet:
            return desired_sheet
        try:
            from openpyxl import load_workbook

            wb = load_workbook(default_path, read_only=True, data_only=True)
            try:
                if wb.sheetnames:
                    return wb.sheetnames[0]
            finally:
                wb.close()
        except Exception:
            pass
        return "Sheet1"

    # 写入输出文件
    if output_path is not None:
        out_safe = guard.resolve_and_validate(output_path)
    else:
        out_safe = safe_path

    target_sheet = _resolve_target_sheet(safe_path, sheet_name)

    source_ext = safe_path.suffix.lower()
    output_ext = out_safe.suffix.lower()
    can_preserve_other_sheets = (
        source_ext in {".xlsx", ".xlsm"}
        and output_ext in {".xlsx", ".xlsm"}
    )

    if can_preserve_other_sheets:
        # 输出文件不存在时，先复制源工作簿，再仅替换目标 sheet。
        if output_path is not None and not out_safe.exists():
            out_safe.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(safe_path, out_safe)

        writer_kwargs: dict[str, Any] = {
            "engine": "openpyxl",
            "mode": "a" if out_safe.exists() else "w",
        }
        if writer_kwargs["mode"] == "a":
            writer_kwargs["if_sheet_exists"] = "replace"
        if output_ext == ".xlsm":
            writer_kwargs["engine_kwargs"] = {"keep_vba": True}

        with pd.ExcelWriter(out_safe, **writer_kwargs) as writer:
            df.to_excel(writer, index=False, sheet_name=target_sheet)
    else:
        df.to_excel(out_safe, index=False, sheet_name=target_sheet)

    result = {
        "status": "success",
        "file": str(out_safe.name),
        "sheet": target_sheet,
        "operations_applied": applied,
        "shape": {"rows": df.shape[0], "columns": df.shape[1]},
    }
    return json.dumps(result, ensure_ascii=False, indent=2)



# inspect_excel_files 可用的 include 维度
_SCAN_FILES_DIMENSIONS = (
    "freeze_panes",
    "charts",
    "images",
    "conditional_formatting",
    "column_widths",
    "vba",
)

# 递归扫描时跳过的噪音目录
_SCAN_SKIP_DIRS: frozenset[str] = frozenset({
    ".git", ".venv", "node_modules", "__pycache__",
    ".worktrees", "dist", "build", "outputs",
})


def inspect_excel_files(
    directory: str = ".",
    max_files: int = 20,
    preview_rows: int = 3,
    max_columns: int = 15,
    include: list[str] | None = None,
    recursive: bool = True,
    search: str | None = None,
    sheet_name: str | None = None,
) -> str:
    """批量扫描目录下所有 Excel 文件，返回轻量级概览，可按需附加额外维度。

    使用 openpyxl 只读模式，仅读取 sheet 元信息和少量预览行，
    避免加载完整 DataFrame，适合快速了解工作区全貌。

    Args:
        directory: 扫描目录（相对于工作目录），默认当前目录。
        max_files: 最多扫描文件数，默认 20。
        preview_rows: 每个 sheet 预览行数，默认 3。
        max_columns: header/preview 最多展示列数，默认 15。
        include: 按需请求的额外维度列表。可选值：
            freeze_panes, charts, images, conditional_formatting, column_widths。
        recursive: 是否递归扫描子目录，默认 True。
        search: 模糊搜索关键词，匹配文件名或 sheet 名称。
        sheet_name: 按 sheet 名称精确搜索，返回包含该 sheet 的文件。

    Returns:
        JSON 格式的批量概览结果。
    """
    from datetime import datetime, timezone
    from pathlib import Path

    from openpyxl import load_workbook
    from excelmanus.tools._helpers import ensure_openpyxl_compatible as _compat

    include_set: set[str] = set(include) if include else set()
    invalid_dims = include_set - set(_SCAN_FILES_DIMENSIONS)
    include_set -= invalid_dims
    needs_full = bool(include_set)

    guard = _get_guard()
    safe_dir = guard.resolve_and_validate(directory)

    if not safe_dir.is_dir():
        return json.dumps(
            {"error": f"路径 '{directory}' 不是一个有效的目录"},
            ensure_ascii=False,
        )

    # 收集 Excel 文件（.xlsx / .xlsm），跳过隐藏文件和临时文件
    # 先收集全部再排序，确保结果确定性（glob 返回顺序依赖文件系统，不可靠）
    glob_method = safe_dir.rglob if recursive else safe_dir.glob
    excel_paths: list[Path] = []
    for ext in ("*.xlsx", "*.xlsm", "*.xls", "*.xlsb"):
        for p in glob_method(ext):
            if p.name.startswith((".", "~$")):
                continue
            # 递归模式下跳过噪音目录
            if recursive:
                rel_parts = p.relative_to(safe_dir).parts[:-1]  # 不含文件名
                if any(part in _SCAN_SKIP_DIRS for part in rel_parts):
                    continue
            excel_paths.append(p)
    excel_paths.sort(key=lambda p: str(p.relative_to(safe_dir)).lower())

    # ── 搜索过滤：按文件名 / sheet 名匹配 ──
    if search or sheet_name:
        search_lower = (search or "").lower()
        sheet_lower = (sheet_name or "").lower()
        matched: list[Path] = []
        # 第一轮：按文件名快速过滤（无需打开文件）
        remaining: list[Path] = []
        for fp in excel_paths:
            if search_lower and search_lower in fp.name.lower():
                matched.append(fp)
            else:
                remaining.append(fp)
        # 第二轮：需要读取 sheet names 的文件
        for fp in remaining:
            if len(matched) >= max_files:
                break
            try:
                wb_peek = load_workbook(_compat(fp), read_only=True, data_only=True)
                try:
                    for sn in wb_peek.sheetnames:
                        sn_lower = sn.lower()
                        if sheet_lower and sheet_lower in sn_lower:
                            matched.append(fp)
                            break
                        if search_lower and search_lower in sn_lower:
                            matched.append(fp)
                            break
                finally:
                    wb_peek.close()
            except Exception:  # noqa: BLE001
                continue
        excel_paths = matched

    excel_paths = excel_paths[:max_files]

    files_summary: list[dict[str, Any]] = []
    for fp in excel_paths:
        stat = fp.stat()
        file_info: dict[str, Any] = {
            "file": fp.name,
            "path": str(fp.relative_to(guard.workspace_root)),
            "size": _format_size(stat.st_size),
            "modified": datetime.fromtimestamp(
                stat.st_mtime, tz=timezone.utc
            ).strftime("%Y-%m-%d"),
        }

        sheets_info: list[dict[str, Any]] = []
        try:
            wb = load_workbook(_compat(fp), read_only=not needs_full, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                total_cols = ws.max_column or 0
                sheet_data: dict[str, Any] = {
                    "name": sn,
                    "rows": ws.max_row or 0,
                    "columns": total_cols,
                }

                # 读取抽样行，用于表头识别与预览
                scan_rows = max(8, preview_rows + 8)
                scan_cols = max(1, min(total_cols if total_cols > 0 else _HEADER_SCAN_COLS, _HEADER_SCAN_COLS))
                rows_raw: list[list[Any]] = []
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=scan_rows,
                    min_col=1,
                    max_col=scan_cols,
                    values_only=True,
                ):
                    rows_raw.append([_normalize_cell(c) for c in row])

                if rows_raw:
                    header_idx = _guess_header_row_from_rows(rows_raw, max_scan=scan_rows)
                    if header_idx is None:
                        header_idx = 0

                    header_raw = rows_raw[header_idx] if header_idx < len(rows_raw) else []
                    header = _trim_trailing_nulls([_cell_to_str(c) for c in header_raw])

                    preview_raw = rows_raw[header_idx + 1:header_idx + 1 + preview_rows]
                    preview = [_trim_trailing_nulls([_cell_to_str(c) for c in r]) for r in preview_raw]

                    # 仅对 preview 数据行限宽，header 完整保留以确保 agent 理解全部列语义
                    if any(len(r) > max_columns for r in preview):
                        preview = [r[:max_columns] for r in preview]
                        sheet_data["preview_columns_truncated"] = max_columns

                    sheet_data["header_row_hint"] = header_idx
                    sheet_data["business_columns"] = len(header)
                    sheet_data["header"] = header
                    sheet_data["preview"] = preview

                # 按需采集额外维度
                if needs_full and include_set:
                    if "freeze_panes" in include_set:
                        sheet_data["freeze_panes"] = _collect_freeze_panes(ws)
                    if "charts" in include_set:
                        sheet_data["charts"] = _collect_charts(ws)
                    if "images" in include_set:
                        sheet_data["images"] = _collect_images(ws)
                    if "conditional_formatting" in include_set:
                        sheet_data["conditional_formatting"] = _collect_conditional_formatting(ws)
                    if "column_widths" in include_set:
                        sheet_data["column_widths"] = _collect_column_widths(ws)

                sheets_info.append(sheet_data)
            wb.close()
        except Exception as exc:  # noqa: BLE001
            file_info["error"] = f"无法读取: {exc}"

        file_info["sheets"] = sheets_info
        # vba 维度：文件级别，不依赖 sheet
        if "vba" in include_set:
            file_info["vba"] = _collect_vba_info(fp, extract_source=False)
        files_summary.append(file_info)

    # 紧凑文件清单放在最前，即使详细信息被截断也能保留完整文件列表
    file_list = [
        {"file": fp.name, "size": _format_size(fp.stat().st_size)}
        for fp in excel_paths
    ]

    result: dict[str, Any] = {
        "directory": directory,
        "excel_files_found": len(excel_paths),
        "truncated": len(excel_paths) >= max_files,
        "file_list": file_list,
        "files": files_summary,
    }
    if invalid_dims:
        result["include_warning"] = f"未知的 include 维度已忽略: {sorted(invalid_dims)}"
    return json.dumps(result, ensure_ascii=False, separators=(',', ':'), default=str)


def _cell_to_str(value: Any) -> str | None:
    """将单元格值转换为紧凑字符串，None 保持为 None。"""
    if value is None:
        return None
    return str(value)


def _trim_trailing_nulls(row: list[Any]) -> list[Any]:
    """去除列表尾部连续的 None 值，减少 JSON 体积。"""
    end = len(row)
    while end > 0 and row[end - 1] is None:
        end -= 1
    return row[:end]


def _df_to_compact_records(df: "pd.DataFrame") -> list[dict[str, Any]]:
    """将 DataFrame 转为紧凑记录：去除 null/NaN 键，大幅减少 token 浪费。

    原理：等效于 JSON 版 Markdown-KV——每个值显式关联其键名，无 null 噪音。
    研究表明 KV 格式 LLM 理解度最高（60.7% vs JSON 53.7% vs CSV 44.3%）。
    合并单元格导致的 NaN 由 merged_cell_summary 独立解释，不依赖数据中的 null。
    """
    records: list[dict[str, Any]] = []
    cols = [str(c) for c in df.columns]
    for row in df.itertuples(index=False):
        d: dict[str, Any] = {}
        for col_name, val in zip(cols, row):
            if pd.notna(val):
                # 序列化 datetime 类型
                if isinstance(val, (date, datetime)):
                    d[col_name] = val.isoformat()
                else:
                    d[col_name] = val
        records.append(d)
    return records


def _build_null_info(df: "pd.DataFrame") -> dict[str, Any] | None:
    """生成空值摘要：一行代替 N×M 个 null token。

    返回 None 表示无显著空值。
    """
    if df.empty:
        return None
    null_rates = df.isnull().mean()
    all_null_cols = [str(c) for c in null_rates[null_rates == 1.0].index]
    high_null_cols = [str(c) for c in null_rates[(null_rates >= 0.6) & (null_rates < 1.0)].index]
    if not all_null_cols and not high_null_cols:
        return None
    info: dict[str, Any] = {}
    if all_null_cols:
        info["完全为空"] = all_null_cols
    if high_null_cols:
        info["高空值率(≥60%)"] = high_null_cols
    return info


def _format_size(size_bytes: int) -> str:
    """将字节数格式化为可读字符串。"""
    for unit in ("B", "KB", "MB", "GB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f}{unit}" if unit != "B" else f"{size_bytes}{unit}"
        size_bytes /= 1024  # type: ignore[assignment]
    return f"{size_bytes:.1f}TB"


# ── include 维度采集函数 ──────────────────────────────────

# include 参数所有合法维度
INCLUDE_DIMENSIONS = (
    "data_preview",
    "styles",
    "charts",
    "images",
    "freeze_panes",
    "conditional_formatting",
    "data_validation",
    "print_settings",
    "column_widths",
    "formulas",
    "categorical_summary",
    "summary",
    "vba",
)

# categorical_summary 默认阈值：unique 值数量低于此值的列视为分类列
_CATEGORICAL_UNIQUE_THRESHOLD = 20


def _collect_categorical_summary(
    df: "pd.DataFrame",
    threshold: int = _CATEGORICAL_UNIQUE_THRESHOLD,
) -> dict[str, Any]:
    """对分类列（unique 值 < threshold）计算 value_counts，返回摘要字典。

    Returns:
        {"threshold": int, "columns": {col: {val: count, ...}, ...}}
    """
    result: dict[str, dict[str, int]] = {}
    for col in df.columns:
        series = df[col].dropna()
        if series.empty:
            continue
        n_unique = series.nunique()
        if 0 < n_unique <= threshold:
            vc = series.value_counts(dropna=True)
            result[str(col)] = {str(k): int(v) for k, v in vc.items()}
    return {"threshold": threshold, "columns": result}


def _collect_data_summary(df: "pd.DataFrame") -> dict[str, Any]:
    """计算每列数据质量概要：null 率、unique 数、min/max（数值列）、top_values（分类列）。"""
    result: dict[str, Any] = {}
    for col in df.columns:
        col_str = str(col)
        series = df[col]
        info: dict[str, Any] = {
            "null_rate": round(float(series.isna().mean()), 4),
            "unique": int(series.nunique()),
        }
        if pd.api.types.is_numeric_dtype(series):
            desc = series.describe()
            info["min"] = desc.get("min")
            info["max"] = desc.get("max")
            info["mean"] = round(float(desc.get("mean", 0)), 2)
        else:
            vc = series.dropna().value_counts().head(3)
            if not vc.empty:
                info["top_values"] = [str(v) for v in vc.index.tolist()]
        result[col_str] = info
    return result


def _color_to_hex_short(color: Any) -> str | None:
    """将 openpyxl Color 对象转为 6 位十六进制字符串，无效或默认色返回 None。"""
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb:
        rgb = str(color.rgb)
        if rgb in ("00000000", "FFFFFFFF"):
            return None
        return rgb[2:] if len(rgb) == 8 else rgb
    if hasattr(color, "theme") and color.theme is not None:
        return f"theme:{color.theme}"
    return None


def _extract_style_tuple(cell: Any) -> tuple | None:
    """从单元格提取样式关键属性元组，全默认样式返回 None。"""
    parts: list[Any] = []
    has_custom = False

    # 字体
    f = cell.font
    if f:
        font_info: dict[str, Any] = {}
        if f.name and f.name != "Calibri":
            font_info["name"] = f.name
        if f.size and f.size != 11:
            font_info["size"] = f.size
        if f.bold:
            font_info["bold"] = True
        if f.italic:
            font_info["italic"] = True
        if f.underline and f.underline != "none":
            font_info["underline"] = f.underline
        if f.strike:
            font_info["strike"] = True
        c = _color_to_hex_short(f.color)
        if c and c != "000000":
            font_info["color"] = c
        if font_info:
            has_custom = True
        parts.append(tuple(sorted(font_info.items())) if font_info else ())
    else:
        parts.append(())

    # 填充
    fl = cell.fill
    if fl:
        fill_type = fl.fill_type or fl.patternType
        if fill_type and fill_type != "none":
            fg = _color_to_hex_short(fl.fgColor)
            parts.append(("fill", fill_type, fg))
            has_custom = True
        else:
            parts.append(())
    else:
        parts.append(())

    # 边框
    b = cell.border
    if b:
        border_parts: list[tuple[str, str]] = []
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(b, side_name, None)
            if side and side.style and side.style != "none":
                border_parts.append((side_name, side.style))
        if border_parts:
            has_custom = True
        parts.append(tuple(border_parts))
    else:
        parts.append(())

    # 对齐
    a = cell.alignment
    if a:
        align_info: dict[str, Any] = {}
        if a.horizontal and a.horizontal != "general":
            align_info["horizontal"] = a.horizontal
        if a.vertical and a.vertical != "bottom":
            align_info["vertical"] = a.vertical
        if a.wrap_text:
            align_info["wrap_text"] = True
        if align_info:
            has_custom = True
        parts.append(tuple(sorted(align_info.items())) if align_info else ())
    else:
        parts.append(())

    # 数字格式
    nf = cell.number_format
    if nf and nf != "General":
        parts.append(nf)
        has_custom = True
    else:
        parts.append("")

    if not has_custom:
        return None
    return tuple(parts)


def _style_tuple_to_dict(st: tuple) -> dict[str, Any]:
    """将样式元组还原为可读字典。"""
    result: dict[str, Any] = {}
    font_parts, fill_parts, border_parts, align_parts, num_fmt = st

    if font_parts:
        result["font"] = dict(font_parts)
    if fill_parts:
        _, fill_type, fg = fill_parts
        info: dict[str, Any] = {"type": fill_type}
        if fg:
            info["color"] = fg
        result["fill"] = info
    if border_parts:
        result["border"] = {side: style for side, style in border_parts}
    if align_parts:
        result["alignment"] = dict(align_parts)
    if num_fmt:
        result["number_format"] = num_fmt
    return result


def _collect_styles_compressed(
    ws: Any,
    max_rows: int = 200,
) -> dict[str, Any]:
    """扫描工作表，以 Style Classes 压缩方式返回样式信息。

    算法：
    1. 逐单元格提取样式元组
    2. 为唯一组合分配 sN ID
    3. 按列扫描合并连续相同样式的单元格为范围

    Returns:
        包含 style_classes, cell_style_map, merged_ranges 的字典。
    """
    from openpyxl.utils import get_column_letter

    scan_rows = min(ws.max_row or 0, max_rows)
    scan_cols = ws.max_column or 0
    if scan_rows == 0 or scan_cols == 0:
        return {
            "style_classes": {},
            "cell_style_map": {},
            "merged_ranges": [str(mr) for mr in ws.merged_cells.ranges],
            "rows_scanned": scan_rows,
        }

    # 第一遍：收集所有样式元组，分配 ID
    style_to_id: dict[tuple, str] = {}
    # cell_map[col_idx][row_idx] = style_id
    cell_map: dict[int, dict[int, str]] = {}
    id_counter = 0

    for row in ws.iter_rows(min_row=1, max_row=scan_rows, min_col=1, max_col=scan_cols):
        for cell in row:
            st = _extract_style_tuple(cell)
            if st is None:
                continue
            if st not in style_to_id:
                style_to_id[st] = f"s{id_counter}"
                id_counter += 1
            sid = style_to_id[st]
            col_idx = cell.column
            row_idx = cell.row
            if col_idx not in cell_map:
                cell_map[col_idx] = {}
            cell_map[col_idx][row_idx] = sid

    # 构建 style_classes 字典
    style_classes = {sid: _style_tuple_to_dict(st) for st, sid in style_to_id.items()}

    # 第二遍：按列合并连续相同 style_id 为范围
    range_map: dict[str, str] = {}  # "A1:A10" -> "s0"

    for col_idx in sorted(cell_map.keys()):
        col_letter = get_column_letter(col_idx)
        rows_dict = cell_map[col_idx]
        sorted_rows = sorted(rows_dict.keys())
        if not sorted_rows:
            continue

        # 合并连续行
        start_row = sorted_rows[0]
        current_sid = rows_dict[start_row]
        prev_row = start_row

        for r in sorted_rows[1:]:
            sid = rows_dict[r]
            if sid == current_sid and r == prev_row + 1:
                # 连续且相同
                prev_row = r
            else:
                # 输出前一段
                if start_row == prev_row:
                    range_map[f"{col_letter}{start_row}"] = current_sid
                else:
                    range_map[f"{col_letter}{start_row}:{col_letter}{prev_row}"] = current_sid
                start_row = r
                current_sid = sid
                prev_row = r

        # 输出最后一段
        if start_row == prev_row:
            range_map[f"{col_letter}{start_row}"] = current_sid
        else:
            range_map[f"{col_letter}{start_row}:{col_letter}{prev_row}"] = current_sid

    merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    return {
        "style_classes": style_classes,
        "cell_style_map": range_map,
        "merged_ranges": merged_ranges,
        "rows_scanned": scan_rows,
    }


def _collect_charts(ws: Any) -> list[dict[str, Any]]:
    """检测工作表中嵌入的图表，返回元信息列表。"""
    charts_info: list[dict[str, Any]] = []
    chart_list = getattr(ws, "_charts", [])
    for chart in chart_list:
        info: dict[str, Any] = {}
        # 图表类型
        type_name = type(chart).__name__.replace("Chart", "").lower()
        info["type"] = type_name
        if hasattr(chart, "title") and chart.title:
            title = chart.title
            if isinstance(title, str):
                info["title"] = title
            else:
                # openpyxl Title/Text object: drill into rich text paragraphs
                text_obj = title.text if hasattr(title, "text") else title
                rich = getattr(text_obj, "rich", None)
                if rich is not None:
                    parts: list[str] = []
                    for p in getattr(rich, "p", []):
                        for r in (getattr(p, "r", None) or []):
                            if getattr(r, "t", None):
                                parts.append(r.t)
                    if parts:
                        info["title"] = "".join(parts)
        info["series_count"] = len(chart.series) if hasattr(chart, "series") else 0
        # 锚点位置
        if hasattr(chart, "anchor") and chart.anchor:
            anchor = chart.anchor
            if hasattr(anchor, "_from") and anchor._from:
                f = anchor._from
                from openpyxl.utils import get_column_letter as gcl
                info["anchor_cell"] = f"{gcl(f.col + 1)}{f.row + 1}"
        charts_info.append(info)
    return charts_info


def _collect_images(ws: Any) -> list[dict[str, Any]]:
    """检测工作表中嵌入的图片，返回元信息列表。"""
    images_info: list[dict[str, Any]] = []
    image_list = getattr(ws, "_images", [])
    for img in image_list:
        info: dict[str, Any] = {}
        if hasattr(img, "width") and img.width:
            info["width_px"] = img.width
        if hasattr(img, "height") and img.height:
            info["height_px"] = img.height
        # 图片格式
        if hasattr(img, "format"):
            info["format"] = img.format
        elif hasattr(img, "path") and img.path:
            ext = str(img.path).rsplit(".", 1)[-1] if "." in str(img.path) else "unknown"
            info["format"] = ext
        # 锚点位置
        if hasattr(img, "anchor") and img.anchor:
            anchor = img.anchor
            if isinstance(anchor, str):
                info["anchor_cell"] = anchor
            elif hasattr(anchor, "_from") and anchor._from:
                f = anchor._from
                from openpyxl.utils import get_column_letter as gcl
                info["anchor_cell"] = f"{gcl(f.col + 1)}{f.row + 1}"
        images_info.append(info)
    return images_info


def _collect_freeze_panes(ws: Any) -> str | None:
    """返回冻结窗格位置（如 'A4'），未冻结返回 None。"""
    fp = ws.freeze_panes
    return str(fp) if fp else None


def _collect_conditional_formatting(ws: Any) -> list[dict[str, Any]]:
    """收集条件格式规则列表。"""
    rules_info: list[dict[str, Any]] = []
    for cf in ws.conditional_formatting:
        ranges_str = str(cf)
        for rule in cf.rules:
            info: dict[str, Any] = {"range": ranges_str}
            if hasattr(rule, "type") and rule.type:
                info["type"] = rule.type
            if hasattr(rule, "priority") and rule.priority is not None:
                info["priority"] = rule.priority
            if hasattr(rule, "formula") and rule.formula:
                info["formula"] = list(rule.formula) if not isinstance(rule.formula, str) else [rule.formula]
            if hasattr(rule, "operator") and rule.operator:
                info["operator"] = rule.operator
            rules_info.append(info)
    return rules_info


def _collect_data_validation(ws: Any) -> list[dict[str, Any]]:
    """收集数据验证规则列表。"""
    validations: list[dict[str, Any]] = []
    dv_list = getattr(ws, "data_validations", None)
    if dv_list is None:
        return validations
    dv_items = getattr(dv_list, "dataValidation", [])
    for dv in dv_items:
        info: dict[str, Any] = {}
        if hasattr(dv, "sqref") and dv.sqref:
            info["range"] = str(dv.sqref)
        if hasattr(dv, "type") and dv.type:
            info["type"] = dv.type
        if hasattr(dv, "formula1") and dv.formula1:
            info["formula1"] = str(dv.formula1)
        if hasattr(dv, "formula2") and dv.formula2:
            info["formula2"] = str(dv.formula2)
        if hasattr(dv, "allow_blank") and dv.allow_blank is not None:
            info["allow_blank"] = bool(dv.allow_blank)
        if hasattr(dv, "showDropDown") and dv.showDropDown is not None:
            info["show_dropdown"] = bool(dv.showDropDown)
        validations.append(info)
    return validations


def _collect_print_settings(ws: Any) -> dict[str, Any]:
    """收集打印设置信息。"""
    info: dict[str, Any] = {}
    if ws.print_area:
        info["print_area"] = ws.print_area
    ps = ws.page_setup
    if ps:
        if ps.orientation:
            info["orientation"] = ps.orientation
        if ps.paperSize is not None:
            info["paper_size"] = ps.paperSize
        if ps.fitToWidth is not None:
            info["fit_to_width"] = ps.fitToWidth
        if ps.fitToHeight is not None:
            info["fit_to_height"] = ps.fitToHeight
        if ps.scale is not None:
            info["scale"] = ps.scale
    if ws.print_title_rows:
        info["repeat_rows"] = ws.print_title_rows
    if ws.print_title_cols:
        info["repeat_columns"] = ws.print_title_cols
    return info


def _collect_column_widths(ws: Any) -> dict[str, float]:
    """收集非默认列宽映射。"""
    widths: dict[str, float] = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None and dim.width != 8.0:
            widths[col_letter] = round(dim.width, 2)
    return widths


def _collect_formulas(ws: Any, max_rows: int = 200) -> list[dict[str, str]]:
    """收集含公式的单元格位置和公式内容。"""
    from openpyxl.utils import get_column_letter

    formulas: list[dict[str, str]] = []
    scan_rows = min(ws.max_row or 0, max_rows)
    scan_cols = ws.max_column or 0
    if scan_rows == 0 or scan_cols == 0:
        return formulas

    for row in ws.iter_rows(min_row=1, max_row=scan_rows, min_col=1, max_col=scan_cols):
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                coord = f"{get_column_letter(cell.column)}{cell.row}"
                formulas.append({"cell": coord, "formula": val})
    return formulas


def _dispatch_include_dimensions(
    ws_for_include: Any,
    include_set: set[str],
    max_style_scan_rows: int,
) -> dict[str, Any]:
    """根据 include 集合分发各维度采集，返回合并字典。"""
    extra: dict[str, Any] = {}

    if "styles" in include_set:
        extra["styles"] = _collect_styles_compressed(ws_for_include, max_rows=max_style_scan_rows)

    if "charts" in include_set:
        extra["charts"] = _collect_charts(ws_for_include)

    if "images" in include_set:
        extra["images"] = _collect_images(ws_for_include)

    if "freeze_panes" in include_set:
        extra["freeze_panes"] = _collect_freeze_panes(ws_for_include)

    if "conditional_formatting" in include_set:
        extra["conditional_formatting"] = _collect_conditional_formatting(ws_for_include)

    if "data_validation" in include_set:
        extra["data_validation"] = _collect_data_validation(ws_for_include)

    if "print_settings" in include_set:
        extra["print_settings"] = _collect_print_settings(ws_for_include)

    if "column_widths" in include_set:
        extra["column_widths"] = _collect_column_widths(ws_for_include)

    if "formulas" in include_set:
        extra["formulas"] = _collect_formulas(ws_for_include, max_rows=max_style_scan_rows)

    # vba 维度在调用方单独处理（需要 file path，不依赖 worksheet）

    return extra


# ── VBA 信息提取 ──────────────────────────────────────────


def _collect_vba_info(file_path: Any, *, extract_source: bool = True) -> dict[str, Any]:
    """提取 .xlsm 文件中的 VBA 宏信息。

    使用 zipfile 读取 vbaProject.bin 的存在性和模块列表。
    当 oletools 可用且 extract_source=True 时，提取完整 VBA 源代码。

    Args:
        file_path: Excel 文件路径。
        extract_source: 是否尝试提取 VBA 源代码（需要 oletools）。

    Returns:
        VBA 信息字典，包含 has_vba、modules 及可选的 source。
    """
    import zipfile
    from pathlib import Path

    path = Path(file_path) if not isinstance(file_path, Path) else file_path
    result: dict[str, Any] = {"has_vba": False, "modules": []}

    if path.suffix.lower() not in (".xlsm", ".xlsb"):
        return result

    # 从 ZIP 结构检测 VBA 项目
    try:
        with zipfile.ZipFile(path, "r") as zf:
            vba_names = [n for n in zf.namelist() if n.startswith("xl/vbaProject")]
            if not vba_names:
                return result
            result["has_vba"] = True
            # 提取模块名列表（从 ZIP 中的 VBA 相关条目）
            macro_entries = [
                n for n in zf.namelist()
                if n.startswith("xl/") and (
                    n.endswith(".bin") and "vba" in n.lower()
                )
            ]
            result["vba_archive_entries"] = macro_entries
    except (zipfile.BadZipFile, Exception):
        result["error"] = "无法读取 ZIP 结构"
        return result

    # 尝试用 oletools 提取 VBA 源码
    if extract_source:
        try:
            from oletools.olevba import VBA_Parser  # type: ignore[import-untyped]

            vba_parser = VBA_Parser(str(path))
            if vba_parser.detect_vba_macros():
                modules: list[dict[str, str]] = []
                for (_, _, vba_filename, vba_code) in vba_parser.extract_macros():
                    modules.append({
                        "name": vba_filename,
                        "code": vba_code,
                    })
                result["modules"] = modules
                result["module_count"] = len(modules)

                # 安全分析摘要
                analysis = list(vba_parser.analyze_macros())
                if analysis:
                    suspicious = [
                        {"type": str(a[0]), "keyword": str(a[1]), "description": str(a[2])}
                        for a in analysis
                    ]
                    result["security_analysis"] = suspicious
            vba_parser.close()
        except ImportError:
            result["source_note"] = (
                "VBA 宏已检测到，但未安装 oletools 库，无法提取源代码。"
                "安装方式: pip install oletools"
            )
        except Exception as exc:
            result["source_error"] = f"VBA 源码提取失败: {exc}"

    return result


# ── 数值强制转换辅助 ──────────────────────────────────────


def _coerce_numeric(series: pd.Series) -> pd.Series:
    """尝试将含文本格式的数值列转换为 float。

    处理常见格式：千分位逗号 "1,234.56"、带单位后缀 "1,234.56元"、
    百分号 "16.36%"。无法转换的值保留 NaN。
    """
    if pd.api.types.is_numeric_dtype(series):
        return series

    cleaned = series.astype(str).str.strip()
    # 移除常见中文单位后缀
    cleaned = cleaned.str.replace(r'[元万亿份个台件套]$', '', regex=True)
    # 移除百分号并标记
    is_pct = cleaned.str.endswith('%')
    cleaned = cleaned.str.replace('%', '', regex=False)
    # 移除千分位逗号
    cleaned = cleaned.str.replace(',', '', regex=False)
    # 转换为数值
    result = pd.to_numeric(cleaned, errors='coerce')
    # 百分比列除以 100
    if is_pct.any() and not is_pct.all():
        # 混合格式，不做百分比转换
        pass
    elif is_pct.all():
        result = result / 100
    return result


def group_aggregate(
    file_path: str,
    group_by: str | list[str],
    aggregations: dict[str, str | list[str]] | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
    sort_by: str | None = None,
    ascending: bool = True,
    limit: int | None = None,
) -> str:
    """按指定列分组并执行聚合统计。

    Args:
        file_path: Excel 文件路径。
        group_by: 分组列名（单个字符串或列表）。
        aggregations: 聚合配置，键为列名，值为聚合函数名或列表。
            支持的聚合函数：count, sum, mean, min, max, median, std, nunique, first, last。
            特殊值 "*" 作为键表示对所有行计数（等价于 COUNT(*)）。
        sheet_name: 工作表名称，默认第一个。
        header_row: 列头所在行号（从0开始），默认自动检测。
        sort_by: 结果排序列名，默认不排序。
        ascending: 排序方向，默认升序。
        limit: 限制返回行数，默认全部返回。

    Returns:
        JSON 格式的聚合结果。
    """
    if aggregations is None:
        return json.dumps(
            {"error": "缺少必需参数 'aggregations'，请指定聚合配置，例如: {\"销售额\": \"sum\"} 或 {\"*\": \"count\"}"},
            ensure_ascii=False,
        )

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    df, _ = _read_df(safe_path, sheet_name, header_row=header_row)

    # 规范化 group_by 为列表
    if isinstance(group_by, str):
        group_by_cols = [group_by]
    else:
        group_by_cols = list(group_by)

    # 校验分组列
    missing_group = [c for c in group_by_cols if c not in df.columns]
    if missing_group:
        return json.dumps(
            {"error": f"分组列不存在: {missing_group}，可用列: {[str(c) for c in df.columns]}"},
            ensure_ascii=False,
            default=str,
        )

    _VALID_AGGS = {"count", "sum", "mean", "min", "max", "median", "std", "nunique", "first", "last"}
    numeric_funcs = {"sum", "mean", "min", "max", "median", "std"}
    formula_meta = df.attrs.get("formula_resolution", {})
    unresolved_formula_cols = set(formula_meta.get("unresolved_columns", [])) if isinstance(formula_meta, dict) else set()

    # 构建 pandas 聚合字典
    agg_dict: dict[str, list[str]] = {}
    has_star_count = False
    risky_formula_cols: set[str] = set()
    for col, funcs in aggregations.items():
        if col == "*":
            has_star_count = True
            continue
        if col not in df.columns:
            return json.dumps(
                {"error": f"聚合列 '{col}' 不存在，可用列: {[str(c) for c in df.columns]}"},
                ensure_ascii=False,
                default=str,
            )
        func_list = [funcs] if isinstance(funcs, str) else list(funcs)
        invalid = [f for f in func_list if f not in _VALID_AGGS]
        if invalid:
            return json.dumps(
                {"error": f"不支持的聚合函数: {invalid}，支持: {sorted(_VALID_AGGS)}"},
                ensure_ascii=False,
            )
        if col in unresolved_formula_cols and any(f in numeric_funcs for f in func_list):
            risky_formula_cols.add(col)
        # 对需要数值的聚合函数，尝试强制转换列类型
        if any(f in numeric_funcs for f in func_list):
            df[col] = _coerce_numeric(df[col])
        agg_dict[col] = func_list

    if risky_formula_cols:
        unresolved_details = formula_meta.get("unresolved_details", {}) if isinstance(formula_meta, dict) else {}
        blocked_details = {c: unresolved_details.get(c, "未解析公式列") for c in sorted(risky_formula_cols)}
        return json.dumps(
            {
                "error": "检测到未解析公式列，已阻止高风险数值聚合（避免返回误导性结果）",
                "blocked_columns": sorted(risky_formula_cols),
                "unresolved_formula_details": blocked_details,
                "suggestion": "请先在 Excel 重算并保存，或改用脚本直接计算这些列后再聚合",
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        )

    if not agg_dict and not has_star_count:
        return json.dumps(
            {"error": "aggregations 不能为空，至少指定一个聚合列或 '*' 进行计数"},
            ensure_ascii=False,
        )

    grouped = df.groupby(group_by_cols, dropna=False)

    if agg_dict:
        result_df = grouped.agg(agg_dict)
        # 扁平化多级列名
        result_df.columns = [
            f"{col}_{func}" if len(funcs) > 1 or has_star_count else f"{col}_{func}"
            for col, funcs in agg_dict.items()
            for func in funcs
        ]
        result_df = result_df.reset_index()
    else:
        result_df = pd.DataFrame({c: [] for c in group_by_cols})

    # COUNT(*) 行计数
    if has_star_count:
        count_series = grouped.size().reset_index(name="count")
        if result_df.empty or len(result_df) == 0:
            result_df = count_series
        else:
            result_df = result_df.merge(count_series, on=group_by_cols, how="left")

    # 排序
    if sort_by and sort_by in result_df.columns:
        result_df = result_df.sort_values(by=sort_by, ascending=ascending)
    elif sort_by:
        # sort_by 可能是原始列名，尝试匹配生成的列名
        candidates = [c for c in result_df.columns if c.startswith(sort_by)]
        if candidates:
            result_df = result_df.sort_values(by=candidates[0], ascending=ascending)

    # 限制行数
    total_before_limit = len(result_df)
    if limit is not None and limit > 0:
        result_df = result_df.head(limit)

    result: dict[str, Any] = {
        "file": str(safe_path.name),
        "group_by": group_by_cols,
        "aggregations": {k: v if isinstance(v, list) else [v] for k, v in aggregations.items()},
        "total_groups": int(grouped.ngroups),
        "rows_returned": len(result_df),
        "columns": [str(c) for c in result_df.columns],
        "data": _df_to_compact_records(result_df),
    }
    if limit is not None and limit > 0 and total_before_limit > limit:
        completeness = build_completeness_meta(
            total_available=total_before_limit,
            returned=len(result_df),
            entity_name="组",
        )
        result["is_truncated"] = completeness.get("is_truncated", False)
        result["truncation_note"] = completeness.get("truncation_note", "")

    return json.dumps(result, ensure_ascii=False, separators=(',', ':'), default=str)


def _normalize_mapping_keys(series: pd.Series) -> pd.Series:
    """标准化映射键：转字符串、去首尾空格、移除空值。"""
    normalized = series.astype(str).str.strip()
    normalized = normalized[normalized.notna()]
    normalized = normalized[normalized != ""]
    normalized = normalized[normalized.str.lower() != "nan"]
    return normalized


def analyze_sheet_mapping(
    file_path: str,
    left_sheet: str,
    left_key: str,
    right_sheet: str,
    right_key: str | None = None,
    right_key_candidates: list[str] | None = None,
    left_header_row: int | None = None,
    right_header_row: int | None = None,
    max_unmatched_samples: int = 20,
) -> str:
    """分析两个工作表键字段的可映射性，输出覆盖率与未匹配样本。

    用于跨表写回前先确认映射口径是否可靠，避免误写。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    left_df, _ = _read_df(safe_path, left_sheet, header_row=left_header_row)
    if left_key not in left_df.columns:
        return json.dumps(
            {"error": f"左表字段 '{left_key}' 不存在，可用列: {[str(c) for c in left_df.columns]}"},
            ensure_ascii=False,
            default=str,
        )

    candidates: list[str] = []
    if right_key:
        candidates.append(right_key)
    if right_key_candidates:
        candidates.extend(right_key_candidates)
    # 去重且保序
    seen: set[str] = set()
    unique_candidates: list[str] = []
    for col in candidates:
        c = str(col).strip()
        if not c or c in seen:
            continue
        seen.add(c)
        unique_candidates.append(c)

    if not unique_candidates:
        return json.dumps(
            {"error": "right_key 或 right_key_candidates 至少提供一个候选字段"},
            ensure_ascii=False,
        )

    right_df, _ = _read_df(safe_path, right_sheet, header_row=right_header_row)
    left_keys = _normalize_mapping_keys(left_df[left_key])
    left_set = set(left_keys.unique().tolist())
    if not left_set:
        return json.dumps(
            {"error": f"左表字段 '{left_key}' 没有可用键值"},
            ensure_ascii=False,
        )

    analyses: list[dict[str, Any]] = []
    for candidate in unique_candidates:
        if candidate not in right_df.columns:
            analyses.append(
                {
                    "right_key": candidate,
                    "error": f"字段不存在，可用列: {[str(c) for c in right_df.columns]}",
                }
            )
            continue

        right_keys = _normalize_mapping_keys(right_df[candidate])
        right_set = set(right_keys.unique().tolist())
        matched = left_set & right_set
        unmatched_left = sorted(left_set - right_set)
        unmatched_right = sorted(right_set - left_set)

        left_cov = len(matched) / max(len(left_set), 1)
        right_cov = len(matched) / max(len(right_set), 1) if right_set else 0.0

        analyses.append(
            {
                "right_key": candidate,
                "left_unique_count": len(left_set),
                "right_unique_count": len(right_set),
                "matched_count": len(matched),
                "left_coverage": round(left_cov, 4),
                "right_coverage": round(right_cov, 4),
                "unmatched_left_samples": unmatched_left[:max_unmatched_samples],
                "unmatched_right_samples": unmatched_right[:max_unmatched_samples],
            }
        )

    valid_analyses = [a for a in analyses if "error" not in a]
    best = None
    if valid_analyses:
        best = max(
            valid_analyses,
            key=lambda x: (x["left_coverage"], x["matched_count"], -x["right_unique_count"]),
        )

    result: dict[str, Any] = {
        "file": str(safe_path.name),
        "left_sheet": left_sheet,
        "left_key": left_key,
        "right_sheet": right_sheet,
        "right_candidates": unique_candidates,
        "analysis": analyses,
    }
    if best is not None:
        result["best_candidate"] = best["right_key"]
        result["best_left_coverage"] = best["left_coverage"]
        result["mapping_recommendation"] = (
            "可自动映射"
            if best["left_coverage"] >= 0.8
            else "映射覆盖不足，建议人工确认口径"
        )
    else:
        result["mapping_recommendation"] = "候选字段均不可用"

    return json.dumps(result, ensure_ascii=False, indent=2, default=str)


# ── Excel 对比工具 ─────────────────────────────────────────


def _load_sheet_as_df(
    safe_path: Any,
    sheet_name: str | None,
    header_row: int | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """加载一个 sheet 为 DataFrame，返回 (df, sheet_names)。"""
    from openpyxl import load_workbook

    if _is_csv_file(safe_path):
        df, _ = _read_csv_df(safe_path, max_rows=None, header_row=header_row)
        return df, ["Sheet1"]

    wb = load_workbook(safe_path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    df, _ = _read_df(safe_path, sheet_name, max_rows=None, header_row=header_row)
    return df, sheet_names


def compare_excel(
    file_a: str,
    file_b: str,
    sheet_a: str = "",
    sheet_b: str = "",
    ignore_style: bool = True,
    key_columns: list[str] | None = None,
    max_diffs: int = 500,
) -> str:
    """对比两个 Excel 文件（或同一文件的两个 Sheet），返回结构化差异报告。

    支持两种对比模式：
    - 行号对齐模式（默认）：逐行逐列对比
    - 关键列匹配模式（指定 key_columns 时）：按关键列 join 后对比

    Args:
        file_a: 基准文件路径
        file_b: 对比文件路径（与 file_a 相同时用于跨 Sheet 对比）
        sheet_a: file_a 的工作表名（空字符串=第一个）
        sheet_b: file_b 的工作表名（空字符串=第一个）
        ignore_style: 是否忽略样式差异（默认 True）
        key_columns: 关键列名列表（用于行匹配，为空则按行号对齐）
        max_diffs: 最大差异数量（超出截断）

    Returns:
        JSON 格式的差异报告。
    """
    guard = _get_guard()

    # ── 1. 解析与校验路径 ──
    from excelmanus.tools._helpers import ensure_openpyxl_compatible as _compat2
    try:
        safe_a = _compat2(guard.resolve_and_validate(file_a))
    except Exception as e:
        return json.dumps({"error": f"文件 A 路径无效: {e}"}, ensure_ascii=False)
    try:
        safe_b = _compat2(guard.resolve_and_validate(file_b))
    except Exception as e:
        return json.dumps({"error": f"文件 B 路径无效: {e}"}, ensure_ascii=False)

    not_found_a = check_file_exists(safe_a, file_a, guard)
    if not_found_a is not None:
        return not_found_a
    not_found_b = check_file_exists(safe_b, file_b, guard)
    if not_found_b is not None:
        return not_found_b

    # ── 2. 加载数据 ──
    # 预读 sheet 列表，用于错误提示
    _sheets_a_hint: list[str] = []
    _sheets_b_hint: list[str] = []
    if not _is_csv_file(safe_a):
        try:
            from openpyxl import load_workbook as _lw
            _wb = _lw(safe_a, read_only=True, data_only=True)
            _sheets_a_hint = list(_wb.sheetnames)
            _wb.close()
        except Exception:
            pass
    if not _is_csv_file(safe_b):
        try:
            from openpyxl import load_workbook as _lw2
            _wb2 = _lw2(safe_b, read_only=True, data_only=True)
            _sheets_b_hint = list(_wb2.sheetnames)
            _wb2.close()
        except Exception:
            pass

    try:
        df_a, sheets_a = _load_sheet_as_df(safe_a, sheet_a or None)
    except Exception as e:
        return json.dumps(
            {"error": f"无法读取文件 A: {e}", "available_sheets": _sheets_a_hint},
            ensure_ascii=False,
        )
    try:
        df_b, sheets_b = _load_sheet_as_df(safe_b, sheet_b or None)
    except Exception as e:
        return json.dumps(
            {"error": f"无法读取文件 B: {e}", "available_sheets": _sheets_b_hint},
            ensure_ascii=False,
        )

    # ── 3. 结构对比 ──
    # 构建 str→原始列名 映射，用于 .at[] 访问
    col_map_a: dict[str, Any] = {str(c): c for c in df_a.columns}
    col_map_b: dict[str, Any] = {str(c): c for c in df_b.columns}
    cols_a = set(col_map_a.keys())
    cols_b = set(col_map_b.keys())
    columns_added = sorted(cols_b - cols_a)
    columns_deleted = sorted(cols_a - cols_b)
    str_cols_a_list = list(col_map_a.keys())
    common_cols = sorted(cols_a & cols_b, key=lambda c: str_cols_a_list.index(c) if c in str_cols_a_list else 0)

    sheets_only_a = sorted(set(sheets_a) - set(sheets_b)) if str(safe_a) != str(safe_b) else []
    sheets_only_b = sorted(set(sheets_b) - set(sheets_a)) if str(safe_a) != str(safe_b) else []

    # ── 4. 数据对比 ──
    cell_diffs: list[dict[str, Any]] = []
    rows_added = 0
    rows_deleted = 0
    rows_modified = 0
    total_cells_compared = 0

    if key_columns and all(k in cols_a and k in cols_b for k in key_columns):
        # ── 4B. 关键列匹配模式 ──
        str_cols_a = {str(c): c for c in df_a.columns}
        str_cols_b = {str(c): c for c in df_b.columns}

        # 将 key_columns 映射回原始列名
        key_a = [str_cols_a[k] for k in key_columns]
        key_b = [str_cols_b[k] for k in key_columns]

        df_a_keyed = df_a.set_index(key_a)
        df_b_keyed = df_b.set_index(key_b)

        all_keys = set(df_a_keyed.index.tolist()) | set(df_b_keyed.index.tolist())

        for key_val in all_keys:
            if len(cell_diffs) >= max_diffs:
                break
            key_label = str(key_val)
            in_a = key_val in df_a_keyed.index
            in_b = key_val in df_b_keyed.index

            if in_a and not in_b:
                rows_deleted += 1
                continue
            if not in_a and in_b:
                rows_added += 1
                continue

            # 两边都有 → 对比 common_cols 中非 key 的列
            row_a = df_a_keyed.loc[key_val]
            row_b = df_b_keyed.loc[key_val]
            # 处理 duplicate key 的情况：取第一行
            if isinstance(row_a, pd.DataFrame):
                row_a = row_a.iloc[0]
            if isinstance(row_b, pd.DataFrame):
                row_b = row_b.iloc[0]

            row_changed = False
            for col in common_cols:
                if col in key_columns:
                    continue
                total_cells_compared += 1
                val_a = _serialize_cell_value(row_a.get(col))
                val_b = _serialize_cell_value(row_b.get(col))
                if str(val_a) != str(val_b):
                    row_changed = True
                    if len(cell_diffs) < max_diffs:
                        cell_diffs.append({
                            "key": key_label,
                            "column": col,
                            "old": val_a,
                            "new": val_b,
                        })
            if row_changed:
                rows_modified += 1

    else:
        # ── 4A. 行号对齐模式 ──
        rows_added = max(0, len(df_b) - len(df_a))
        rows_deleted = max(0, len(df_a) - len(df_b))
        min_rows = min(len(df_a), len(df_b))

        # 大文件优化：先做行级 hash 快速过滤
        use_hash_filter = min_rows > 10000
        diff_row_indices: set[int] | None = None

        if use_hash_filter and common_cols:
            hash_a = df_a[list(df_a.columns)].iloc[:min_rows].astype(str).apply(
                lambda r: hash(tuple(r)), axis=1
            )
            hash_b = df_b[list(df_b.columns)].iloc[:min_rows].astype(str).apply(
                lambda r: hash(tuple(r)), axis=1
            )
            diff_row_indices = set((hash_a != hash_b).to_numpy().nonzero()[0])
            logger.info(
                "行级 hash 过滤：%d/%d 行有差异",
                len(diff_row_indices), min_rows,
            )

        for row_idx in _builtin_range(min_rows):
            if len(cell_diffs) >= max_diffs:
                break
            if diff_row_indices is not None and row_idx not in diff_row_indices:
                total_cells_compared += len(common_cols)
                continue

            row_changed = False
            for col in common_cols:
                total_cells_compared += 1
                orig_col_a = col_map_a.get(col)
                orig_col_b = col_map_b.get(col)
                val_a = _serialize_cell_value(
                    df_a.at[row_idx, orig_col_a] if orig_col_a is not None else None
                )
                val_b = _serialize_cell_value(
                    df_b.at[row_idx, orig_col_b] if orig_col_b is not None else None
                )
                if str(val_a) != str(val_b):
                    row_changed = True
                    if len(cell_diffs) < max_diffs:
                        from openpyxl.utils import get_column_letter
                        try:
                            col_idx = list(str(c) for c in df_a.columns).index(col)
                            cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 2}"
                        except (ValueError, IndexError):
                            cell_ref = f"R{row_idx + 2}:{col}"
                        cell_diffs.append({
                            "cell": cell_ref,
                            "old": val_a,
                            "new": val_b,
                        })
            if row_changed:
                rows_modified += 1

        # 统计超出部分行的额外单元格
        if len(df_b) > len(df_a):
            for extra_idx in _builtin_range(len(df_a), min(len(df_b), len(df_a) + max_diffs)):
                if len(cell_diffs) >= max_diffs:
                    break
                for col in cols_b:
                    total_cells_compared += 1
                    orig_col = col_map_b.get(col)
                    val = _serialize_cell_value(df_b.at[extra_idx, orig_col] if orig_col is not None else None)
                    if val is not None and str(val) != "":
                        cell_diffs.append({
                            "cell": f"R{extra_idx + 2}:{col}",
                            "old": None,
                            "new": val,
                        })

    # ── 5. 构建结果 ──
    truncated = len(cell_diffs) >= max_diffs
    cells_different = len(cell_diffs)

    # 选取前 10 个 sample_diffs 供 LLM 参考
    sample_diffs = cell_diffs[:10]

    is_same_file = str(safe_a) == str(safe_b)
    diff_mode = "cross_sheet" if is_same_file and (sheet_a or sheet_b) else "cross_file"

    result: dict[str, Any] = {
        "status": "ok",
        "diff_mode": diff_mode,
        "file_a": str(safe_a.name),
        "file_b": str(safe_b.name),
        "sheet_a": sheet_a or "(默认)",
        "sheet_b": sheet_b or "(默认)",
        "summary": {
            "total_cells_compared": total_cells_compared,
            "cells_different": cells_different,
            "rows_added": rows_added,
            "rows_deleted": rows_deleted,
            "rows_modified": rows_modified,
            "columns_added": columns_added,
            "columns_deleted": columns_deleted,
            "sheets_only_in_a": sheets_only_a,
            "sheets_only_in_b": sheets_only_b,
        },
        "sample_diffs": sample_diffs,
        "truncated": truncated,
    }

    if cells_different == 0 and rows_added == 0 and rows_deleted == 0:
        result["hint"] = "两个文件（或 Sheet）的数据完全相同。"
    else:
        parts = []
        if cells_different > 0:
            parts.append(f"{cells_different} 处单元格差异")
        if rows_added > 0:
            parts.append(f"{rows_added} 行新增")
        if rows_deleted > 0:
            parts.append(f"{rows_deleted} 行删除")
        if columns_added:
            parts.append(f"新增列: {', '.join(columns_added)}")
        if columns_deleted:
            parts.append(f"删除列: {', '.join(columns_deleted)}")
        result["hint"] = f"共发现 {'、'.join(parts)}。完整 diff 已通过前端展示。"

    return json.dumps(result, ensure_ascii=False, indent=2, default=str)


# ── scan_excel_snapshot ──────────────────────────────────────

_SNAPSHOT_MAX_SHEETS = 10
_SNAPSHOT_MAX_SAMPLE_VALUES = 5
_SNAPSHOT_MAX_TOP_VALUES = 5
_SNAPSHOT_MAX_SIGNALS = 20
_SNAPSHOT_CATEGORICAL_THRESHOLD = 20


def _infer_column_type(series: "pd.Series") -> str:
    """根据实际值分布推断列类型（不只看 dtype）。

    Returns:
        "numeric" | "string" | "date" | "boolean" | "mixed" | "empty"
    """
    non_null = series.dropna()
    if non_null.empty:
        return "empty"

    type_counts: dict[str, int] = {}
    for val in non_null.head(200):
        if isinstance(val, bool):
            type_counts["boolean"] = type_counts.get("boolean", 0) + 1
        elif isinstance(val, (int, float)):
            type_counts["numeric"] = type_counts.get("numeric", 0) + 1
        elif isinstance(val, str):
            type_counts["string"] = type_counts.get("string", 0) + 1
        else:
            t = type(val).__name__
            if "date" in t.lower() or "time" in t.lower():
                type_counts["date"] = type_counts.get("date", 0) + 1
            else:
                type_counts["other"] = type_counts.get("other", 0) + 1

    if not type_counts:
        return "empty"

    dominant = max(type_counts, key=type_counts.get)  # type: ignore[arg-type]
    total = sum(type_counts.values())
    dominant_ratio = type_counts[dominant] / total

    if len(type_counts) == 1:
        return dominant
    if dominant_ratio >= 0.9:
        return dominant
    return "mixed"


def _detect_mixed_types(series: "pd.Series") -> dict[str, int] | None:
    """检测同列中不同 Python 类型的分布。仅在存在混合时返回。"""
    non_null = series.dropna()
    if non_null.empty:
        return None

    type_counts: dict[str, int] = {}
    for val in non_null.head(500):
        if isinstance(val, bool):
            t = "bool"
        elif isinstance(val, int):
            t = "int"
        elif isinstance(val, float):
            t = "float"
        elif isinstance(val, str):
            t = "str"
        else:
            t = type(val).__name__
        type_counts[t] = type_counts.get(t, 0) + 1

    if len(type_counts) <= 1:
        return None
    # int + float 不算混合
    if set(type_counts.keys()) <= {"int", "float"}:
        return None
    return type_counts


def _detect_outliers_iqr(series: "pd.Series") -> int:
    """使用 IQR 方法检测数值列的异常值个数。"""
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if len(numeric) < 4:
        return 0
    q1 = float(numeric.quantile(0.25))
    q3 = float(numeric.quantile(0.75))
    iqr = q3 - q1
    if iqr == 0:
        return 0
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    return int(((numeric < lower) | (numeric > upper)).sum())


def _compute_column_stats(
    series: "pd.Series",
    inferred_type: str,
    col_name: str,
) -> dict[str, Any]:
    """计算单列的统计信息。"""
    total = len(series)
    null_count = int(series.isna().sum())
    null_rate = round(null_count / total, 4) if total > 0 else 0.0
    non_null = series.dropna()
    unique_count = int(non_null.nunique()) if not non_null.empty else 0

    stats: dict[str, Any] = {
        "name": col_name,
        "dtype": str(series.dtype),
        "inferred_type": inferred_type,
        "null_count": null_count,
        "null_rate": null_rate,
        "unique_count": unique_count,
    }

    # 样本值
    sample = non_null.head(_SNAPSHOT_MAX_SAMPLE_VALUES).tolist()
    stats["sample_values"] = [
        str(v) if not isinstance(v, (int, float, bool)) else v
        for v in sample
    ]

    # 数值列特有统计
    if inferred_type == "numeric" and not non_null.empty:
        numeric = pd.to_numeric(non_null, errors="coerce").dropna()
        if not numeric.empty:
            stats["min"] = round(float(numeric.min()), 2)
            stats["max"] = round(float(numeric.max()), 2)
            stats["mean"] = round(float(numeric.mean()), 2)
            stats["median"] = round(float(numeric.median()), 2)
            if len(numeric) >= 4:
                stats["q1"] = round(float(numeric.quantile(0.25)), 2)
                stats["q3"] = round(float(numeric.quantile(0.75)), 2)
            stats["outlier_count"] = _detect_outliers_iqr(numeric)

    # 分类列（低基数 string/mixed）特有统计
    if unique_count > 0 and unique_count <= _SNAPSHOT_CATEGORICAL_THRESHOLD:
        vc = non_null.value_counts(dropna=True).head(_SNAPSHOT_MAX_TOP_VALUES)
        stats["top_values"] = [
            {"value": str(k), "count": int(v)} for k, v in vc.items()
        ]

    # 类型混杂检测
    mixed = _detect_mixed_types(series)
    if mixed is not None:
        stats["mixed_type_counts"] = mixed

    return stats


def _detect_cross_sheet_relationships(
    sheet_columns: dict[str, list[str]],
    sheet_dfs: dict[str, "pd.DataFrame"],
) -> list[dict[str, Any]]:
    """检测跨 Sheet 关联：共享列名 + 值重叠率。"""
    relationships: list[dict[str, Any]] = []
    sheet_names = list(sheet_columns.keys())

    # 共享列名检测
    if len(sheet_names) >= 2:
        from itertools import combinations
        for s1, s2 in combinations(sheet_names, 2):
            shared = set(sheet_columns[s1]) & set(sheet_columns[s2])
            if shared:
                relationships.append({
                    "type": "shared_column_name",
                    "columns": sorted(shared),
                    "sheets": [s1, s2],
                })

    # 候选外键检测（值重叠率）
    for rel in list(relationships):
        if rel["type"] != "shared_column_name":
            continue
        s1, s2 = rel["sheets"]
        df1, df2 = sheet_dfs.get(s1), sheet_dfs.get(s2)
        if df1 is None or df2 is None:
            continue
        for col in rel["columns"]:
            if col not in df1.columns or col not in df2.columns:
                continue
            vals1 = set(df1[col].dropna().astype(str).tolist()[:500])
            vals2 = set(df2[col].dropna().astype(str).tolist()[:500])
            if not vals1 or not vals2:
                continue
            overlap = len(vals1 & vals2)
            smaller = min(len(vals1), len(vals2))
            rate = round(overlap / smaller, 2) if smaller > 0 else 0
            if rate >= 0.3:
                relationships.append({
                    "type": "candidate_foreign_key",
                    "source": {"sheet": s2, "column": col},
                    "target": {"sheet": s1, "column": col},
                    "overlap_rate": rate,
                })

    return relationships


def _generate_quality_signals(
    sheets_data: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """根据扫描结果生成阈值化质量信号。"""
    signals: list[dict[str, Any]] = []

    for sheet in sheets_data:
        sheet_name = sheet["name"]
        total_rows = sheet.get("rows", 0)

        for col in sheet.get("columns", []):
            col_name = col["name"]
            null_rate = col.get("null_rate", 0)

            # empty_column
            if null_rate >= 1.0:
                signals.append({
                    "severity": "high",
                    "type": "empty_column",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": "该列全部为空值",
                })
                continue

            # missing_data
            if null_rate > 0.3:
                signals.append({
                    "severity": "high",
                    "type": "missing_data",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": f"{col['null_count']} 个空值 ({null_rate:.1%})",
                })
            elif null_rate > 0.05:
                signals.append({
                    "severity": "medium",
                    "type": "missing_data",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": f"{col['null_count']} 个空值 ({null_rate:.1%})",
                })

            # type_mixed
            if col.get("mixed_type_counts"):
                counts = col["mixed_type_counts"]
                desc = " + ".join(f"{k}({v})" for k, v in counts.items())
                signals.append({
                    "severity": "high",
                    "type": "type_mixed",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": f"混合了 {desc}",
                })

            # outliers
            outlier_count = col.get("outlier_count", 0)
            if outlier_count > 0:
                signals.append({
                    "severity": "medium",
                    "type": "outliers",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": f"{outlier_count} 个异常值 (IQR 方法)",
                })

            # constant_column
            if col.get("unique_count") == 1 and col.get("null_count", 0) == 0:
                signals.append({
                    "severity": "low",
                    "type": "constant_column",
                    "sheet": sheet_name,
                    "column": col_name,
                    "detail": "仅 1 个唯一值",
                })

            # high_cardinality
            if (
                col.get("inferred_type") == "string"
                and total_rows > 10
                and col.get("unique_count", 0) > 0
            ):
                unique_rate = col["unique_count"] / max(total_rows, 1)
                if unique_rate > 0.9:
                    signals.append({
                        "severity": "info",
                        "type": "high_cardinality",
                        "sheet": sheet_name,
                        "column": col_name,
                        "detail": f"唯一率 {unique_rate:.1%}，疑似 ID 列",
                    })

        # duplicate_rows
        dup_count = sheet.get("duplicate_row_count")
        if dup_count is not None and dup_count > 0:
            dup_rate = dup_count / max(total_rows, 1)
            sev = "high" if dup_rate > 0.05 else "medium"
            signals.append({
                "severity": sev,
                "type": "duplicate_rows",
                "sheet": sheet_name,
                "detail": f"{dup_count} 行完全重复 ({dup_rate:.1%})",
            })

        # high_merge_ratio — 合并单元格占比高，数据读取可能产生大量 NaN
        merged_summary = sheet.get("merged_cell_summary")
        if merged_summary:
            ratio_str = merged_summary.get("merged_cell_ratio", "0%")
            data_merged = merged_summary.get("data_merged_ranges", 0)
            try:
                ratio_val = float(ratio_str.rstrip("%")) / 100
            except (ValueError, AttributeError):
                ratio_val = 0.0
            if ratio_val > _FORM_MERGED_CELL_RATIO_THRESHOLD:
                detail_parts = [f"合并单元格占比 {ratio_str}"]
                if data_merged > 0:
                    detail_parts.append(
                        f"其中 {data_merged} 处数据区跨行合并（pandas 读取会产生 NaN）"
                    )
                col_groups = merged_summary.get("column_group_spans", [])
                if col_groups:
                    detail_parts.append(
                        f"列组标头: {', '.join(col_groups[:5])}"
                    )
                signals.append({
                    "severity": "high",
                    "type": "high_merge_ratio",
                    "sheet": sheet_name,
                    "detail": "；".join(detail_parts),
                })

    return signals[:_SNAPSHOT_MAX_SIGNALS]


def scan_excel_snapshot(
    file_path: str,
    max_sample_rows: int = 500,
    include_relationships: bool = True,
) -> str:
    """一次性扫描 Excel 文件，返回所有 Sheet 的 schema、列统计、数据质量信号。

    Args:
        file_path: Excel/CSV 文件路径（相对或绝对）。
        max_sample_rows: 大表采样行数上限（默认 500）。
        include_relationships: 是否检测跨 Sheet 关联。

    Returns:
        JSON 格式的完整扫描报告。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    not_found = check_file_exists(safe_path, file_path, guard)
    if not_found is not None:
        return not_found

    file_size = safe_path.stat().st_size
    size_str = _format_size(file_size)

    # CSV 特殊处理
    if _is_csv_file(safe_path):
        return _scan_csv_snapshot(safe_path, size_str, max_sample_rows)

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    from openpyxl import load_workbook

    # 元数据扫描（read_only=True，快速获取行列数/公式/合并）
    wb_meta = load_workbook(safe_path, read_only=True, data_only=True)
    sheet_metas: list[dict[str, Any]] = []
    for ws in wb_meta.worksheets[:_SNAPSHOT_MAX_SHEETS]:
        meta: dict[str, Any] = {
            "name": ws.title,
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
        }
        sheet_metas.append(meta)
    wb_meta.close()

    # 检测公式和合并单元格（需要非 read_only 模式，但只读前几行）
    try:
        wb_full = load_workbook(safe_path, read_only=False, data_only=False)
        for i, ws in enumerate(wb_full.worksheets[:_SNAPSHOT_MAX_SHEETS]):
            if i < len(sheet_metas):
                has_merged = len(ws.merged_cells.ranges) > 0
                sheet_metas[i]["has_merged_cells"] = has_merged
                # 合并单元格摘要：语义分类 + 合并率 + 处理建议
                if has_merged:
                    merged_summary = _collect_merged_cell_summary(ws)
                    if merged_summary:
                        sheet_metas[i]["merged_cell_summary"] = merged_summary
                # 检测公式：扫描前 20 行
                has_formulas = False
                for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row or 0), values_only=False):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            has_formulas = True
                            break
                    if has_formulas:
                        break
                sheet_metas[i]["has_formulas"] = has_formulas
        wb_full.close()
    except Exception:
        for meta in sheet_metas:
            meta.setdefault("has_formulas", False)
            meta.setdefault("has_merged_cells", False)

    # 逐 Sheet 统计
    sheets_data: list[dict[str, Any]] = []
    sheet_columns: dict[str, list[str]] = {}
    sheet_dfs: dict[str, pd.DataFrame] = {}

    for meta in sheet_metas:
        sheet_name = meta["name"]
        total_rows = meta["rows"]
        data_rows = max(0, total_rows - 1)  # 减去 header
        sampled = data_rows > max_sample_rows

        try:
            read_kwargs = _build_read_kwargs(safe_path, sheet_name, max_rows=max_sample_rows if sampled else None)
            form_type = read_kwargs.pop("_form_type_document", False)
            df = pd.read_excel(**read_kwargs)

            if form_type:
                df.columns = [f"Col_{i}" for i in range(len(df.columns))]
        except Exception as exc:
            sheets_data.append({
                **meta,
                "error": f"读取失败: {exc}",
                "columns": [],
            })
            continue

        # 列统计
        columns_stats: list[dict[str, Any]] = []
        col_names: list[str] = []
        for col in df.columns:
            col_str = str(col)
            col_names.append(col_str)
            inferred = _infer_column_type(df[col])
            stats = _compute_column_stats(df[col], inferred, col_str)
            columns_stats.append(stats)

        # 重复行检测
        dup_count: int | None = None
        if data_rows <= 10000:
            try:
                dup_count = int(df.duplicated().sum())
            except Exception:
                dup_count = None

        sheet_data: dict[str, Any] = {
            **meta,
            "duplicate_row_count": dup_count,
            "columns": columns_stats,
        }
        if sampled:
            sheet_data["sampled"] = True
            sheet_data["sample_size"] = len(df)

        sheets_data.append(sheet_data)
        sheet_columns[sheet_name] = col_names
        sheet_dfs[sheet_name] = df

    # 跨 Sheet 关联
    relationships: list[dict[str, Any]] = []
    if include_relationships and len(sheet_columns) >= 2:
        relationships = _detect_cross_sheet_relationships(sheet_columns, sheet_dfs)

    # 质量信号
    quality_signals = _generate_quality_signals(sheets_data)

    result: dict[str, Any] = {
        "file": safe_path.name,
        "size": size_str,
        "sheet_count": len(sheet_metas),
        "sheets": sheets_data,
        "relationships": relationships,
        "quality_signals": quality_signals,
    }

    if len(wb_meta.sheetnames if hasattr(wb_meta, 'sheetnames') else sheet_metas) > _SNAPSHOT_MAX_SHEETS:
        result["truncated"] = True
        result["truncated_note"] = f"仅扫描前 {_SNAPSHOT_MAX_SHEETS} 个 Sheet"

    return json.dumps(result, ensure_ascii=False, separators=(",", ":"), default=str)


def _scan_csv_snapshot(
    safe_path: Any,
    size_str: str,
    max_sample_rows: int,
) -> str:
    """CSV 文件的 scan_excel_snapshot 简化实现。"""
    try:
        df, _ = _read_csv_df(safe_path, max_rows=max_sample_rows)
    except Exception as exc:
        return json.dumps({"error": f"CSV 读取失败: {exc}"}, ensure_ascii=False)

    total_rows = len(df)
    sampled = False
    # 检查实际行数是否超过采样
    try:
        with open(safe_path, "r", encoding="utf-8", errors="ignore") as f:
            actual_lines = sum(1 for _ in f) - 1  # 减去 header
        if actual_lines > max_sample_rows:
            sampled = True
            total_rows = actual_lines
    except Exception:
        pass

    columns_stats = []
    for col in df.columns:
        inferred = _infer_column_type(df[col])
        stats = _compute_column_stats(df[col], inferred, str(col))
        columns_stats.append(stats)

    dup_count = int(df.duplicated().sum()) if len(df) <= 10000 else None

    sheet_data: dict[str, Any] = {
        "name": "Sheet1",
        "rows": total_rows,
        "cols": len(df.columns),
        "header_row": 0,
        "duplicate_row_count": dup_count,
        "has_formulas": False,
        "has_merged_cells": False,
        "columns": columns_stats,
    }
    if sampled:
        sheet_data["sampled"] = True
        sheet_data["sample_size"] = len(df)

    quality_signals = _generate_quality_signals([sheet_data])

    result: dict[str, Any] = {
        "file": safe_path.name,
        "size": size_str,
        "sheet_count": 1,
        "sheets": [sheet_data],
        "relationships": [],
        "quality_signals": quality_signals,
    }
    return json.dumps(result, ensure_ascii=False, separators=(",", ":"), default=str)


# ── search_excel_values ──────────────────────────────────────


def search_excel_values(
    file_path: str = "",
    query: str = "",
    match_mode: str = "contains",
    sheets: list[str] | None = None,
    columns: list[str] | None = None,
    max_results: int = 50,
    case_sensitive: bool = False,
    file_paths: list[str] | None = None,
) -> str:
    """跨 Sheet 搜索 Excel 单元格值，类似 ripgrep。

    Args:
        file_path: Excel 文件路径（相对或绝对）。
        query: 搜索字符串或正则表达式。
        match_mode: 匹配模式："contains"（默认）| "exact" | "regex" | "startswith"。
        sheets: 限定搜索的 Sheet 列表（默认全部）。
        columns: 限定搜索的列名列表（默认全部）。
        max_results: 最大返回匹配数（默认 50）。
        case_sensitive: 是否区分大小写（默认 False）。
        file_paths: 多文件搜索路径列表（与 file_path 互补，传入此参数时跨文件搜索）。

    Returns:
        JSON 格式的搜索结果。
    """
    # 跨文件搜索：当 file_paths 包含多个文件时，逐文件搜索并合并结果
    if file_paths and len(file_paths) > 1:
        all_matches: list[dict[str, Any]] = []
        all_summary: list[dict[str, Any]] = []
        total = 0
        files_searched = 0
        per_file_max = max(10, max_results // len(file_paths))
        for fp in file_paths[:10]:  # 最多搜索 10 个文件
            sub_result_str = search_excel_values(
                file_path=fp, query=query, match_mode=match_mode,
                sheets=sheets, columns=columns, max_results=per_file_max,
                case_sensitive=case_sensitive,
            )
            try:
                sub = json.loads(sub_result_str)
                if "error" in sub:
                    continue
                files_searched += 1
                sub_total = sub.get("total_matches", 0)
                total += sub_total
                for m in sub.get("matches", []):
                    m["file"] = fp
                    all_matches.append(m)
                for s in sub.get("summary_by_sheet", []):
                    s["file"] = fp
                    all_summary.append(s)
            except (json.JSONDecodeError, ValueError):
                continue
        # 按文件+sheet排序，截断到 max_results
        all_matches = all_matches[:max_results]
        return json.dumps({
            "query": query, "match_mode": match_mode,
            "total_matches": total, "returned": len(all_matches),
            "truncated": total > len(all_matches),
            "files_searched": files_searched,
            "matches": all_matches,
            "summary_by_sheet": all_summary,
        }, ensure_ascii=False, separators=(",", ":"), default=str)

    # 单文件搜索：兼容 file_paths=[单个文件] 的情况
    if not file_path and file_paths:
        file_path = file_paths[0]
    if not file_path:
        return json.dumps(
            {"error": "必须提供 file_path 或 file_paths 参数"},
            ensure_ascii=False,
        )
    import re as _re

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    not_found = check_file_exists(safe_path, file_path, guard)
    if not_found is not None:
        return not_found

    if not query:
        return json.dumps(
            {"total_matches": 0, "returned": 0, "truncated": False,
             "sheets_searched": 0, "matches": [], "summary_by_sheet": []},
            ensure_ascii=False,
        )

    # 编译匹配函数
    if match_mode == "fuzzy":
        # 模糊匹配：将 query 拆分为子串 token，单元格值包含所有 token 即匹配
        # 拆分规则：按空格/标点分割，同时将连续中文与连续数字/字母分离
        # 额外在中文数字边界处拆分（如 '电子一班' → ['电子', '一班']）
        _cjk_re = _re.compile(r'[\u4e00-\u9fff]+|[a-zA-Z]+|\d+', _re.UNICODE)
        _CN_DIGITS_SET = set("一二三四五六七八九十")
        _pre_tokens = _cjk_re.findall(query)
        _raw_tokens: list[str] = []
        for _pt in _pre_tokens:
            # 对纯中文 token，在中文数字与非中文数字字符之间拆分
            if all('\u4e00' <= c <= '\u9fff' for c in _pt) and any(c in _CN_DIGITS_SET for c in _pt):
                _sub_re = _re.compile(r'(?<=[^\u4e00\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u5341])(?=[一二三四五六七八九十])|(?<=[一二三四五六七八九十])(?=[^\u4e00\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u5341])')
                _parts = _sub_re.split(_pt)
                _raw_tokens.extend(p for p in _parts if p)
            else:
                _raw_tokens.append(_pt)
        # 中文数字 → 阿拉伯数字等价替换，合并两套 token
        _CN_DIGIT = {"一": "1", "二": "2", "三": "3", "四": "4", "五": "5",
                     "六": "6", "七": "7", "八": "8", "九": "9", "十": "10"}
        _normalized_tokens: list[str] = []
        for _tok in _raw_tokens:
            _alt = _tok
            for _cn, _ar in _CN_DIGIT.items():
                _alt = _alt.replace(_cn, _ar)
            if not case_sensitive:
                _tok = _tok.lower()
                _alt = _alt.lower()
            _normalized_tokens.append(_tok)
            if _alt != _tok:
                _normalized_tokens.append(_alt)
        # 去重并过滤空串
        _fuzzy_tokens = list(dict.fromkeys(t for t in _normalized_tokens if t))
        if not _fuzzy_tokens:
            # 无有效 token 时回退到 contains
            if case_sensitive:
                def _match(cell_str: str) -> bool:
                    return query in cell_str
            else:
                _q_lower = query.lower()
                def _match(cell_str: str) -> bool:
                    return _q_lower in cell_str.lower()
        else:
            # 每个原始 token 至少有一个变体命中即可（原始或数字替换版本）
            # 构造 token 组：每组内的 token 是同一原始词的变体，组内 OR，组间 AND
            _token_groups: list[list[str]] = []
            for _tok in _raw_tokens:
                _group = [_tok.lower() if not case_sensitive else _tok]
                _alt = _tok
                for _cn, _ar in _CN_DIGIT.items():
                    _alt = _alt.replace(_cn, _ar)
                if _alt != _tok:
                    _group.append(_alt.lower() if not case_sensitive else _alt)
                _token_groups.append(_group)

            def _match(cell_str: str) -> bool:
                _s = cell_str if case_sensitive else cell_str.lower()
                return all(any(v in _s for v in grp) for grp in _token_groups)
    elif match_mode == "regex":
        try:
            flags = 0 if case_sensitive else _re.IGNORECASE
            pattern = _re.compile(query, flags)
        except _re.error as exc:
            return json.dumps(
                {"error": f"正则表达式无效: {exc}"},
                ensure_ascii=False,
            )
        def _match(cell_str: str) -> bool:
            return bool(pattern.search(cell_str))
    elif match_mode == "exact":
        if case_sensitive:
            def _match(cell_str: str) -> bool:
                return cell_str == query
        else:
            _q_lower = query.lower()
            def _match(cell_str: str) -> bool:
                return cell_str.lower() == _q_lower
    elif match_mode == "startswith":
        if case_sensitive:
            def _match(cell_str: str) -> bool:
                return cell_str.startswith(query)
        else:
            _q_lower = query.lower()
            def _match(cell_str: str) -> bool:
                return cell_str.lower().startswith(_q_lower)
    else:  # contains
        if case_sensitive:
            def _match(cell_str: str) -> bool:
                return query in cell_str
        else:
            _q_lower = query.lower()
            def _match(cell_str: str) -> bool:
                return _q_lower in cell_str.lower()

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    safe_path = ensure_openpyxl_compatible(safe_path)

    wb = load_workbook(safe_path, read_only=True, data_only=True)
    matches: list[dict[str, Any]] = []
    sheet_match_counts: dict[str, int] = {}
    total_matches = 0
    sheets_searched = 0

    target_sheets = set(s.lower() for s in sheets) if sheets else None

    for ws in wb.worksheets:
        if target_sheets and ws.title.lower() not in target_sheets:
            continue
        sheets_searched += 1

        # 读取 header 行（仅用于列名标注，不再跳过 row 1 的搜索）
        header: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(row)]
            break

        target_col_indices: set[int] | None = None
        if columns:
            col_set_lower = {c.lower() for c in columns}
            target_col_indices = {
                i for i, h in enumerate(header) if h.lower() in col_set_lower
            }
            if not target_col_indices:
                continue

        # 逐行扫描（从第 1 行开始，包含 header 行——表单类文档的数据可能从第 1 行起）
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True), start=1
        ):
            if len(matches) >= max_results:
                break
            row_list = list(row)
            for col_idx, cell_val in enumerate(row_list):
                if cell_val is None:
                    continue
                if target_col_indices is not None and col_idx not in target_col_indices:
                    continue
                cell_str = str(cell_val)
                if not _match(cell_str):
                    continue

                total_matches += 1
                sheet_match_counts[ws.title] = sheet_match_counts.get(ws.title, 0) + 1

                if len(matches) < max_results:
                    col_name = header[col_idx] if col_idx < len(header) else f"Col_{col_idx}"
                    cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx}"

                    # context: 同行其他列的值（最多 5 列）
                    context: dict[str, str] = {}
                    ctx_count = 0
                    for ci, cv in enumerate(row_list):
                        if ci == col_idx or cv is None:
                            continue
                        h = header[ci] if ci < len(header) else f"Col_{ci}"
                        context[h] = str(cv)[:100]
                        ctx_count += 1
                        if ctx_count >= 5:
                            break

                    matches.append({
                        "sheet": ws.title,
                        "row": row_idx,
                        "column": col_name,
                        "value": cell_str[:200],
                        "cell_ref": cell_ref,
                        "context": context,
                    })

        if len(matches) >= max_results:
            # 继续统计剩余 sheet 的总匹配数（但不收集详情）
            continue

    wb.close()

    truncated = total_matches > len(matches)
    summary_by_sheet = [
        {"sheet": s, "matches": c}
        for s, c in sorted(sheet_match_counts.items(), key=lambda x: -x[1])
    ]

    result: dict[str, Any] = {
        "query": query,
        "match_mode": match_mode,
        "total_matches": total_matches,
        "returned": len(matches),
        "truncated": truncated,
        "sheets_searched": sheets_searched,
        "matches": matches,
        "summary_by_sheet": summary_by_sheet,
    }

    # 0 结果时添加智能提示，引导 LLM 优化搜索策略
    if total_matches == 0 and match_mode in ("contains", "exact", "startswith"):
        hints: list[str] = []
        if len(query) > 2:
            hints.append(f"缩短搜索词（如只搜索 '{query[:2]}' 或其中某个关键词）")
        # 检测中文数字，提示可能的阿拉伯数字等价
        _CN_DIGIT_MAP = {"一": "1", "二": "2", "三": "3", "四": "4", "五": "5",
                         "六": "6", "七": "7", "八": "8", "九": "9", "十": "10"}
        _has_cn_digit = any(c in query for c in _CN_DIGIT_MAP)
        if _has_cn_digit:
            _alt = query
            for cn, ar in _CN_DIGIT_MAP.items():
                _alt = _alt.replace(cn, ar)
            hints.append(f"查询含中文数字，可尝试阿拉伯数字版本: '{_alt}'")
        hints.append("尝试 match_mode='fuzzy' 进行分词模糊匹配（自动拆分关键词 + 中文数字转换）")
        hints.append("尝试 match_mode='regex' 用正则灵活匹配")
        result["search_hints"] = hints

    return json.dumps(result, ensure_ascii=False, separators=(",", ":"), default=str)


# ── 跨文件关系发现 ──────────────────────────────────────


# 列名归一化映射：中英文常见同义词（小写 key → 归一化标准形式）
_COLUMN_NAME_SYNONYMS: dict[str, str] = {
    "customerid": "客户id",
    "customer_id": "客户id",
    "客户编号": "客户id",
    "客户号": "客户id",
    "cust_id": "客户id",
    "employeeid": "员工id",
    "employee_id": "员工id",
    "员工编号": "员工id",
    "工号": "员工id",
    "emp_id": "员工id",
    "productid": "产品id",
    "product_id": "产品id",
    "产品编号": "产品id",
    "商品编号": "产品id",
    "prod_id": "产品id",
    "orderid": "订单id",
    "order_id": "订单id",
    "订单编号": "订单id",
    "order_no": "订单id",
    "订单号": "订单id",
    # 注意：不映射 name/姓名/名称、date/日期/时间 等泛化列名，
    # 因为它们语义过于宽泛（产品名称 vs 客户姓名），会导致大量误匹配。
    # 这些列只能通过精确列名匹配来发现关联。
}

# 列名后缀去除列表（归一化时尝试剥离）
# 注意：compact 形式已去除下划线，所以后缀列表中不含下划线前缀
_COLUMN_SUFFIX_STRIP = ("id", "编号", "号", "编码", "代码", "code", "no", "num")


def _normalize_column_name(name: str) -> str:
    """将列名归一化为标准形式，用于跨文件列匹配。

    策略：
    1. strip + 统一小写 + 去除空格/下划线差异
    2. 查同义词表精确映射
    3. 去常见后缀后二次查表
    """
    raw = str(name).strip()
    if not raw:
        return ""
    # 统一小写、去首尾空白
    lower = raw.lower().strip()
    # 去除空格和下划线差异
    compact = lower.replace(" ", "").replace("_", "").replace("-", "")

    # 查同义词表（精确匹配原始小写形式）
    if lower in _COLUMN_NAME_SYNONYMS:
        return _COLUMN_NAME_SYNONYMS[lower]
    if compact in _COLUMN_NAME_SYNONYMS:
        return _COLUMN_NAME_SYNONYMS[compact]

    # 去后缀后二次查表
    for suffix in _COLUMN_SUFFIX_STRIP:
        if compact.endswith(suffix) and len(compact) > len(suffix):
            stripped = compact[: -len(suffix)]
            if stripped in _COLUMN_NAME_SYNONYMS:
                return _COLUMN_NAME_SYNONYMS[stripped]

    return compact


def _detect_cross_file_relationships(
    file_columns: dict[str, dict[str, list[str]]],
    file_dfs: dict[str, dict[str, "pd.DataFrame"]],
    *,
    overlap_threshold: float = 0.5,
    max_sample: int = 500,
) -> list[dict[str, Any]]:
    """检测跨文件列关联：精确列名 + 归一化列名 + 值重叠率 + 方向性 + 类型兼容性。

    每个匹配列对额外输出：
    - unique_ratio_a/b: 各列唯一值占比（用于判断一对多方向）
    - relationship: "one_to_many" | "many_to_one" | "one_to_one" | "many_to_many"
    - suggested_join: 推荐的 pandas merge 策略
    - type_a/type_b: 推断列类型
    - type_compatible: 类型是否兼容

    Args:
        file_columns: {file_path: {sheet_name: [col_names]}}
        file_dfs: {file_path: {sheet_name: DataFrame}}
        overlap_threshold: 值重叠率阈值（默认 0.5）
        max_sample: 值重叠检测的最大采样数

    Returns:
        file_pairs 列表
    """
    from itertools import combinations
    from typing import NamedTuple

    # 展开为 (file, sheet, col_name) 四元组
    class _ColRef(NamedTuple):
        file: str
        sheet: str
        col: str
        normalized: str

    all_cols: list[_ColRef] = []
    for fp, sheets in file_columns.items():
        for sheet_name, cols in sheets.items():
            for col in cols:
                if str(col).startswith("Unnamed"):
                    continue
                norm = _normalize_column_name(col)
                if norm:
                    all_cols.append(_ColRef(file=fp, sheet=sheet_name, col=col, normalized=norm))

    # 类型兼容性矩阵（对称）
    _COMPATIBLE_TYPES: set[frozenset[str]] = {
        frozenset({"numeric", "numeric"}),
        frozenset({"string", "string"}),
        frozenset({"date", "date"}),
        frozenset({"numeric", "string"}),  # 数字可能以字符串形式存储
        frozenset({"string", "mixed"}),
        frozenset({"numeric", "mixed"}),
    }

    def _types_compatible(t1: str, t2: str) -> bool:
        if t1 == t2:
            return True
        return frozenset({t1, t2}) in _COMPATIBLE_TYPES

    # 按文件对分组比较
    file_paths_sorted = sorted(file_columns.keys())
    file_pairs: list[dict[str, Any]] = []

    for fp_a, fp_b in combinations(file_paths_sorted, 2):
        cols_a = [c for c in all_cols if c.file == fp_a]
        cols_b = [c for c in all_cols if c.file == fp_b]
        if not cols_a or not cols_b:
            continue

        shared_columns: list[dict[str, Any]] = []
        seen_pairs: set[tuple[str, str, str, str]] = set()  # 去重

        for ca in cols_a:
            for cb in cols_b:
                pair_key = (ca.sheet, ca.col, cb.sheet, cb.col)
                if pair_key in seen_pairs:
                    continue

                # 判断匹配类型
                match_type: str | None = None
                if ca.col.lower().strip() == cb.col.lower().strip():
                    match_type = "exact"
                elif ca.normalized == cb.normalized:
                    match_type = "normalized"
                else:
                    continue

                seen_pairs.add(pair_key)

                # 值重叠检测 + 方向性分析
                df_a = file_dfs.get(fp_a, {}).get(ca.sheet)
                df_b = file_dfs.get(fp_b, {}).get(cb.sheet)
                overlap_ratio = 0.0
                sample_overlap: list[str] = []
                unique_ratio_a = 0.0
                unique_ratio_b = 0.0
                type_a = "unknown"
                type_b = "unknown"

                if df_a is not None and df_b is not None:
                    if ca.col in df_a.columns and cb.col in df_b.columns:
                        series_a = df_a[ca.col].dropna()
                        series_b = df_b[cb.col].dropna()
                        vals_a = set(series_a.astype(str).tolist()[:max_sample])
                        vals_b = set(series_b.astype(str).tolist()[:max_sample])

                        if vals_a and vals_b:
                            overlap = vals_a & vals_b
                            smaller = min(len(vals_a), len(vals_b))
                            overlap_ratio = round(len(overlap) / smaller, 2) if smaller > 0 else 0.0
                            sample_overlap = sorted(overlap)[:5]

                        # 唯一值占比（方向性信号）
                        len_a = len(series_a)
                        len_b = len(series_b)
                        if len_a > 0:
                            unique_ratio_a = round(series_a.nunique() / len_a, 2)
                        if len_b > 0:
                            unique_ratio_b = round(series_b.nunique() / len_b, 2)

                        # 列类型推断
                        type_a = _infer_column_type(df_a[ca.col])
                        type_b = _infer_column_type(df_b[cb.col])

                if overlap_ratio >= overlap_threshold or match_type == "exact":
                    # 方向性判断
                    relationship = "many_to_many"
                    if unique_ratio_a >= 0.95 and unique_ratio_b >= 0.95:
                        relationship = "one_to_one"
                    elif unique_ratio_a >= 0.95:
                        relationship = "one_to_many"  # A 是主表（唯一键），B 是多端
                    elif unique_ratio_b >= 0.95:
                        relationship = "many_to_one"  # B 是主表，A 是多端

                    # 合并策略建议
                    suggested_join = "inner"  # 默认内连接
                    if overlap_ratio >= 0.8:
                        if relationship in ("one_to_many", "one_to_one"):
                            suggested_join = "left"   # A 为主表左连接
                        elif relationship == "many_to_one":
                            suggested_join = "right"  # B 为主表
                        else:
                            suggested_join = "inner"
                    elif overlap_ratio >= 0.5:
                        suggested_join = "left"  # 中等重叠用 left 保留主表全量
                    else:
                        suggested_join = "outer"  # 低重叠用 outer 避免丢数据

                    entry: dict[str, Any] = {
                        "col_a": ca.col,
                        "sheet_a": ca.sheet,
                        "col_b": cb.col,
                        "sheet_b": cb.sheet,
                        "match_type": match_type,
                        "overlap_ratio": overlap_ratio,
                        "unique_ratio_a": unique_ratio_a,
                        "unique_ratio_b": unique_ratio_b,
                        "relationship": relationship,
                        "suggested_join": suggested_join,
                        "type_a": type_a,
                        "type_b": type_b,
                        "type_compatible": _types_compatible(type_a, type_b),
                    }
                    if sample_overlap:
                        entry["sample_overlap"] = sample_overlap
                    if not _types_compatible(type_a, type_b):
                        entry["type_warning"] = (
                            f"列类型不一致（{type_a} vs {type_b}），"
                            "合并前可能需要类型转换"
                        )
                    shared_columns.append(entry)

        if shared_columns:
            file_pairs.append({
                "file_a": fp_a,
                "file_b": fp_b,
                "shared_columns": shared_columns,
            })

    return file_pairs


def discover_file_relationships(
    file_paths: list[str] | None = None,
    directory: str = ".",
    max_files: int = 5,
    sample_rows: int = 200,
) -> str:
    """发现多个 Excel 文件之间的列关联关系（共享列名、疑似外键）。

    对指定文件（或目录内所有 Excel 文件）提取各 sheet 的列名和样本值，
    跨文件交叉比对（精确匹配 + 归一化匹配 + 值重叠检测）。

    Args:
        file_paths: 要分析的文件路径列表。为空时扫描 directory。
        directory: 扫描目录（相对于工作目录），默认当前目录。
        max_files: 最多分析的文件数，默认 5。
        sample_rows: 每个 sheet 采样的行数，默认 200。

    Returns:
        JSON 格式的跨文件关系报告。
    """
    from pathlib import Path

    guard = _get_guard()

    # ── 收集文件路径 ──
    paths: list[Path] = []
    if file_paths:
        for fp in file_paths[:max_files]:
            try:
                safe = guard.resolve_and_validate(fp)
                if safe.exists() and safe.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".xlsb", ".csv", ".tsv"}:
                    paths.append(safe)
            except Exception:
                continue
    else:
        safe_dir = guard.resolve_and_validate(directory)
        if safe_dir.is_dir():
            for ext in ("*.xlsx", "*.xlsm", "*.xls", "*.xlsb"):
                for p in safe_dir.rglob(ext):
                    if p.name.startswith((".", "~$")):
                        continue
                    # 跳过噪音目录（与 inspect_excel_files 一致）
                    try:
                        rel_parts = p.relative_to(safe_dir).parts[:-1]
                        if any(part in _SCAN_SKIP_DIRS for part in rel_parts):
                            continue
                    except ValueError:
                        pass
                    paths.append(p)
                    if len(paths) >= max_files:
                        break
                if len(paths) >= max_files:
                    break
            paths.sort(key=lambda p: str(p).lower())
            paths = paths[:max_files]

    if len(paths) < 2:
        return json.dumps({
            "files_analyzed": len(paths),
            "file_pairs": [],
            "summary": "需要至少 2 个文件才能分析跨文件关系" if len(paths) < 2 else "",
        }, ensure_ascii=False)

    # ── 提取列信息和样本数据 ──
    from excelmanus.tools._helpers import ensure_openpyxl_compatible

    file_columns: dict[str, dict[str, list[str]]] = {}  # rel_path → {sheet → [cols]}
    file_dfs: dict[str, dict[str, pd.DataFrame]] = {}  # rel_path → {sheet → df}
    file_display: dict[str, str] = {}  # rel_path → display_name

    for fp in paths:
        rel_path = str(fp.relative_to(guard.workspace_root)) if fp.is_relative_to(guard.workspace_root) else str(fp)
        file_display[rel_path] = fp.name

        try:
            if _is_csv_file(fp):
                df, _ = _read_csv_df(fp, max_rows=sample_rows)
                cols = [str(c) for c in df.columns if not str(c).startswith("Unnamed")]
                file_columns[rel_path] = {"Sheet1": cols}
                file_dfs[rel_path] = {"Sheet1": df}
                continue

            compat_path = ensure_openpyxl_compatible(fp)
            from openpyxl import load_workbook
            wb = load_workbook(compat_path, read_only=True, data_only=True)
            sheets_cols: dict[str, list[str]] = {}
            sheets_dfs: dict[str, pd.DataFrame] = {}
            try:
                for ws in wb.worksheets[:8]:  # 最多 8 个 sheet
                    sheet_name = ws.title
                    try:
                        read_kwargs = _build_read_kwargs(compat_path, sheet_name, max_rows=sample_rows)
                        read_kwargs.pop("_form_type_document", None)
                        df = pd.read_excel(**read_kwargs)
                        cols = [str(c) for c in df.columns if not str(c).startswith("Unnamed")]
                        if cols:
                            sheets_cols[sheet_name] = cols
                            sheets_dfs[sheet_name] = df
                    except Exception:
                        continue
            finally:
                wb.close()
            if sheets_cols:
                file_columns[rel_path] = sheets_cols
                file_dfs[rel_path] = sheets_dfs
        except Exception as exc:
            logger.debug("跨文件关系发现：读取 %s 失败: %s", fp, exc)
            continue

    if len(file_columns) < 2:
        return json.dumps({
            "files_analyzed": len(file_columns),
            "file_pairs": [],
            "summary": "可读取的文件不足 2 个，无法分析跨文件关系",
        }, ensure_ascii=False)

    # ── 跨文件关系检测 ──
    file_pairs = _detect_cross_file_relationships(file_columns, file_dfs)

    # ── 构建摘要 + 合并提示 ──
    summary_parts: list[str] = []
    merge_hints: list[dict[str, Any]] = []
    type_warnings: list[str] = []
    related_file_set: set[str] = set()  # 用于 suggested_groups

    for pair in file_pairs:
        fa = file_display.get(pair["file_a"], pair["file_a"])
        fb = file_display.get(pair["file_b"], pair["file_b"])
        related_file_set.add(pair["file_a"])
        related_file_set.add(pair["file_b"])
        cols = pair["shared_columns"]

        col_descs: list[str] = []
        for c in cols[:3]:
            if c["col_a"] == c["col_b"]:
                col_descs.append(f"{c['col_a']}(重叠{c['overlap_ratio']:.0%})")
            else:
                col_descs.append(f"{c['col_a']}↔{c['col_b']}(重叠{c['overlap_ratio']:.0%})")
            # 收集类型告警
            if c.get("type_warning"):
                type_warnings.append(f"{fa}.{c['col_a']} vs {fb}.{c['col_b']}: {c['type_warning']}")
        desc = "、".join(col_descs)
        if len(cols) > 3:
            desc += f" 等{len(cols)}对"
        summary_parts.append(f"{fa} ↔ {fb}: {desc}")

        # 合并提示（取第一个最强关联列的建议）
        best_col = max(cols, key=lambda x: x.get("overlap_ratio", 0))
        _JOIN_LABELS = {
            "left": "左连接（保留左表全量）",
            "right": "右连接（保留右表全量）",
            "inner": "内连接（仅保留匹配行）",
            "outer": "全外连接（保留两表全量）",
        }
        _REL_LABELS = {
            "one_to_one": "一对一",
            "one_to_many": f"{fa} 为主表，{fb} 为多端",
            "many_to_one": f"{fb} 为主表，{fa} 为多端",
            "many_to_many": "多对多",
        }
        merge_hints.append({
            "file_a": fa,
            "file_b": fb,
            "key_column_a": best_col["col_a"],
            "key_column_b": best_col["col_b"],
            "relationship": _REL_LABELS.get(best_col.get("relationship", ""), ""),
            "suggested_join": best_col.get("suggested_join", "inner"),
            "suggested_join_label": _JOIN_LABELS.get(best_col.get("suggested_join", "inner"), ""),
            "pandas_hint": (
                f"pd.merge(df_a, df_b, "
                f"left_on='{best_col['col_a']}', right_on='{best_col['col_b']}', "
                f"how='{best_col.get('suggested_join', 'inner')}')"
            ),
        })

    summary = ""
    if summary_parts:
        summary = "跨文件关联发现：" + "；".join(summary_parts)
    else:
        analyzed_names = [file_display.get(k, k) for k in file_columns]
        summary = f"在 {', '.join(analyzed_names)} 之间未发现明显的列关联"

    result: dict[str, Any] = {
        "files_analyzed": len(file_columns),
        "file_pairs": file_pairs,
        "summary": summary,
    }

    # 合并操作提示（供 agent 直接参考的可操作建议）
    if merge_hints:
        result["merge_hints"] = merge_hints

    # 类型兼容性告警
    if type_warnings:
        result["type_warnings"] = type_warnings

    # 方面4联动：文件组建议（当 ≥2 个文件有强关联时）
    if len(related_file_set) >= 2:
        group_files = [
            {"path": fp, "display_name": file_display.get(fp, fp)}
            for fp in sorted(related_file_set)
        ]
        result["suggested_groups"] = [{
            "name": "关联文件组",
            "reason": summary,
            "files": group_files,
        }]

    return json.dumps(result, ensure_ascii=False, separators=(",", ":"), default=str)


# ── get_tools() 导出 ──────────────────────────────────────



def get_tools() -> list[ToolDef]:
    """返回数据操作 Skill 的所有工具定义。"""
    return [
        ToolDef(
            name="read_excel",
            description=(
                "读取 Excel/CSV 数据摘要（形状、列名、类型、前10行+后5行预览），通过 include 按需附加样式/图表/公式/数据概要等维度。"
                "支持 range 精确读取指定坐标区域、offset+max_rows 分页、sample_rows 等距采样。"
                "适用场景：探查文件结构、确认列名与数据类型、查看任意区域数据、了解数据分布。"
                "不适用：批量数据处理或跨表操作（改用 run_code + pandas）。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel/CSV 文件路径（相对于工作目录），支持 .xlsx/.xls/.xlsm/.xlsb/.csv/.tsv",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认读取第一个 sheet（CSV 时忽略）",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "最大读取行数，默认全部；小数据集（<=100行）建议不限制",
                        "minimum": 1,
                    },
                    "header_row": {
                        "type": "integer",
                        "description": "列头行号（0-indexed，即第1行=0），默认自动检测",
                        "minimum": 0,
                    },
                    "range": {
                        "type": "string",
                        "description": "Excel 坐标范围（如 'A1:F20'、'B100:D200'），指定后精确读取该区域，大文件友好。不支持 CSV。格式：列字母+行号[:列字母+行号]",
                        "pattern": "^[A-Za-z]{1,3}\\d{1,7}(:[A-Za-z]{1,3}\\d{1,7})?$",
                    },
                    "offset": {
                        "type": "integer",
                        "description": "数据行偏移（从0开始，header 之后起算），与 max_rows 组合实现分页",
                        "minimum": 0,
                    },
                    "sample_rows": {
                        "type": "integer",
                        "description": "等距采样行数，用于了解大表数据分布",
                        "minimum": 1,
                    },
                    "include": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": [
                                "styles",
                                "charts",
                                "images",
                                "freeze_panes",
                                "conditional_formatting",
                                "data_validation",
                                "print_settings",
                                "column_widths",
                                "formulas",
                                "categorical_summary",
                                "summary",
                            ],
                        },
                        "description": "按需附加的额外维度列表",
                    },
                    "max_style_scan_rows": {
                        "type": "integer",
                        "description": "styles/formulas 维度扫描的最大行数，默认 200",
                        "default": 200,
                        "minimum": 1,
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=read_excel,
            max_result_chars=10000,
            write_effect="none",
        ),
        # write_excel: Batch 1 精简
        # analyze_data: Batch 4 精简，由 run_code + pandas describe() 替代
        ToolDef(
            name="filter_data",
            description=(
                "按条件筛选 Excel 数据行，支持单条件或多条件 AND/OR 组合、列投影、排序和 Top-N。"
                "适用场景：按条件查找特定数据行、排序取 Top-N、按列投影裁剪输出。"
                "不适用：需要聚合统计（改用 run_code + pandas groupby）。"
                "参数模式：单条件用 column/operator/value；多条件用 conditions 数组（二者互斥）。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径（相对于工作目录）",
                    },
                    "column": {
                        "type": "string",
                        "description": "筛选列名（单条件模式，与 conditions 互斥）",
                    },
                    "operator": {
                        "type": "string",
                        "enum": ["eq", "ne", "gt", "ge", "lt", "le", "contains", "in", "not_in", "between", "isnull", "notnull", "startswith", "endswith"],
                        "description": "比较运算符（单条件模式）",
                    },
                    "value": {
                        "description": "比较值（单条件模式）：in/not_in 传数组，between 传 [min,max]，isnull/notnull 可不传",
                    },
                    "conditions": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "column": {"type": "string"},
                                "operator": {
                                    "type": "string",
                                    "enum": ["eq", "ne", "gt", "ge", "lt", "le", "contains", "in", "not_in", "between", "isnull", "notnull", "startswith", "endswith"],
                                },
                                "value": {},
                            },
                            "required": ["column", "operator", "value"],
                        },
                        "description": "多条件数组（与单条件互斥）",
                    },
                    "logic": {
                        "type": "string",
                        "enum": ["and", "or"],
                        "description": "多条件组合逻辑，默认 and",
                        "default": "and",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认第一个",
                    },
                    "header_row": {
                        "type": "integer",
                        "description": "列头行号（0-indexed，即第1行=0），默认自动检测",
                        "minimum": 0,
                    },
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "只返回指定列（投影）",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "最多返回行数",
                        "minimum": 1,
                    },
                    "sort_by": {
                        "type": "string",
                        "description": "排序列名",
                    },
                    "ascending": {
                        "type": "boolean",
                        "description": "排序方向，默认升序",
                        "default": True,
                    },
                    "limit": {
                        "type": "integer",
                        "description": "排序后限制返回行数（Top-N）",
                        "minimum": 1,
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=filter_data,
            max_result_chars=8000,
            write_effect="none",
        ),
        ToolDef(
            name="inspect_excel_files",
            description=(
                "批量扫描目录下所有 Excel 文件概况（sheet/行列/列名/预览），支持按文件名或 sheet 名模糊搜索定位。"
                "适用场景：工作区有多个 Excel 文件时快速了解全貌、按关键词定位目标文件。"
                "不适用：已知文件路径时直接用 read_excel。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "directory": {
                        "type": "string",
                        "description": "扫描目录（相对于工作目录），默认当前目录",
                        "default": ".",
                    },
                    "max_files": {
                        "type": "integer",
                        "description": "最多扫描文件数，默认 20",
                        "default": 20,
                        "minimum": 1,
                    },
                    "preview_rows": {
                        "type": "integer",
                        "description": "每个 sheet 预览数据行数（不含标题行），默认 3",
                        "default": 3,
                        "minimum": 0,
                    },
                    "max_columns": {
                        "type": "integer",
                        "description": "header/preview 最多展示列数，默认 15",
                        "default": 15,
                        "minimum": 1,
                    },
                    "include": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": [
                                "freeze_panes",
                                "charts",
                                "images",
                                "conditional_formatting",
                                "column_widths",
                            ],
                        },
                        "description": "按需附加的额外维度",
                    },
                    "recursive": {
                        "type": "boolean",
                        "description": "是否递归扫描子目录，默认 true",
                        "default": True,
                    },
                    "search": {
                        "type": "string",
                        "description": "模糊搜索关键词，匹配文件名或 sheet 名（如 '学生花名册'、'销售汇总'）",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "按 sheet 名称搜索，返回包含该 sheet 的文件",
                    },
                },
                "required": [],
                "additionalProperties": False,
            },
            func=inspect_excel_files,
            max_result_chars=0,
            write_effect="none",
        ),
        ToolDef(
            name="compare_excel",
            description=(
                "对比两个 Excel 文件或同一文件的两个工作表，返回结构化差异报告。"
                "支持按行/列/单元格级别展示新增、删除和修改。"
                "适用场景：两个版本文件对比、模板 vs 填充后对比、同文件跨 Sheet 对比。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_a": {
                        "type": "string",
                        "description": "第一个文件路径（基准文件）",
                    },
                    "file_b": {
                        "type": "string",
                        "description": "第二个文件路径（对比文件）。与 file_a 相同时用于跨 Sheet 对比",
                    },
                    "sheet_a": {
                        "type": "string",
                        "description": "file_a 的工作表名（可选，默认第一个）",
                        "default": "",
                    },
                    "sheet_b": {
                        "type": "string",
                        "description": "file_b 的工作表名（可选，默认第一个）",
                        "default": "",
                    },
                    "ignore_style": {
                        "type": "boolean",
                        "description": "是否忽略样式差异，仅对比值",
                        "default": True,
                    },
                    "key_columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "用于行匹配的关键列名（可选，默认按行号对齐）",
                        "default": [],
                    },
                    "max_diffs": {
                        "type": "integer",
                        "description": "最大差异数量（超出截断）",
                        "default": 500,
                        "minimum": 1,
                    },
                },
                "required": ["file_a", "file_b"],
                "additionalProperties": False,
            },
            func=compare_excel,
            write_effect="none",
        ),
        # transform_data: Batch 1 精简
        # group_aggregate: Batch 4 精简，由 run_code + pandas groupby() 替代
        # analyze_sheet_mapping: Batch 4 精简，由 run_code + pandas merge 分析替代
        ToolDef(
            name="scan_excel_snapshot",
            description=(
                "一次性扫描 Excel 文件全貌：所有 Sheet 的列 schema、数据类型、统计概要（空值/唯一值/min/max/分位数/异常值）、"
                "重复行检测、类型混杂检测、跨 Sheet 关联发现、数据质量信号汇总。"
                "适用场景：首次接触文件时快速了解全貌，替代多次 read_excel + run_code 的组合。"
                "不适用：已知具体需求时直接用 read_excel 定向读取。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径（相对于工作目录），支持 .xlsx/.xls/.xlsm/.xlsb/.csv/.tsv",
                    },
                    "max_sample_rows": {
                        "type": "integer",
                        "description": "大表采样行数上限（默认 500），行数 ≤ 此值时全量计算",
                        "default": 500,
                        "minimum": 10,
                    },
                    "include_relationships": {
                        "type": "boolean",
                        "description": "是否检测跨 Sheet 关联（共享列名、疑似外键），默认 true",
                        "default": True,
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=scan_excel_snapshot,
            max_result_chars=15000,
            write_effect="none",
        ),
        ToolDef(
            name="search_excel_values",
            description=(
                "跨 Sheet/跨文件搜索 Excel 单元格值（类似 ripgrep），支持包含/精确/正则/前缀/模糊匹配。"
                "返回匹配的 sheet/行/列/值/单元格引用及同行上下文。"
                "适用场景：在单个或多个文件中查找特定值或模式、定位数据出现位置。"
                "跨文件搜索：传入 file_paths 列表即可一次搜索多个文件。"
                "模糊匹配（fuzzy）：自动拆分关键词并逐个子串匹配，适合用户口语与实际数据不完全一致的场景"
                "（如搜索'电子一班'可匹配'24级电子信息科学与技术1班'）。"
                "不适用：条件筛选和排序（改用 filter_data）。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径（相对于工作目录），单文件搜索时使用",
                    },
                    "file_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "多文件搜索路径列表（跨文件搜索时使用，最多 10 个文件），与 file_path 二选一",
                    },
                    "query": {
                        "type": "string",
                        "description": "搜索字符串或正则表达式",
                    },
                    "match_mode": {
                        "type": "string",
                        "enum": ["contains", "exact", "regex", "startswith", "fuzzy"],
                        "description": "匹配模式：contains（默认）| exact | regex | startswith | fuzzy（分词模糊匹配，适合口语化搜索）",
                        "default": "contains",
                    },
                    "sheets": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "限定搜索的 Sheet 名称列表（默认全部）",
                    },
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "限定搜索的列名列表（默认全部）",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "最大返回匹配数，默认 50",
                        "default": 50,
                        "minimum": 1,
                    },
                    "case_sensitive": {
                        "type": "boolean",
                        "description": "是否区分大小写，默认 false",
                        "default": False,
                    },
                },
                "required": ["query"],
                "additionalProperties": False,
            },
            func=search_excel_values,
            max_result_chars=8000,
            write_effect="none",
        ),
        ToolDef(
            name="discover_file_relationships",
            description=(
                "发现多个 Excel 文件之间的列关联关系（共享列名、疑似外键、归一化匹配）。"
                "对指定文件（或目录内所有 Excel）提取列名+样本值，跨文件交叉比对。"
                "适用场景：多文件合并/匹配前的关系探查、确定哪些列可用作 JOIN 键。"
                "不适用：单文件跨 Sheet 关系（改用 scan_excel_snapshot 的 include_relationships）。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "要分析的文件路径列表（最多 5 个）。为空时扫描 directory 下所有 Excel 文件",
                    },
                    "directory": {
                        "type": "string",
                        "description": "扫描目录（相对于工作目录），默认当前目录。仅在 file_paths 为空时生效",
                        "default": ".",
                    },
                    "max_files": {
                        "type": "integer",
                        "description": "最多分析的文件数，默认 5",
                        "default": 5,
                        "minimum": 2,
                        "maximum": 10,
                    },
                    "sample_rows": {
                        "type": "integer",
                        "description": "每个 sheet 采样的行数（用于值重叠检测），默认 200",
                        "default": 200,
                        "minimum": 10,
                    },
                },
                "required": [],
                "additionalProperties": False,
            },
            func=discover_file_relationships,
            max_result_chars=10000,
            write_effect="none",
        ),
    ]


