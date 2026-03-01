"""图片工具模块：read_image + 流水线工具（rebuild / verify）。"""

from __future__ import annotations

import base64
import json
from pathlib import Path
from typing import Any

from excelmanus.security import FileAccessGuard, SecurityViolationError
from excelmanus.tools._guard_ctx import get_guard as _get_ctx_guard
from excelmanus.tools.registry import ToolDef

_guard: FileAccessGuard | None = None
_SUPPORTED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
_MAX_SIZE_BYTES = 20_000_000


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard（优先 per-session contextvar）。"""
    ctx_guard = _get_ctx_guard()
    if ctx_guard is not None:
        return ctx_guard
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    global _guard
    _guard = FileAccessGuard(workspace_root)


def _resolve_path(user_path: str) -> Path:
    """解析路径；仅在显式初始化 guard 后执行工作区校验。"""
    if _guard is None:
        return Path(user_path)
    return _get_guard().resolve_and_validate(user_path)


def read_image(*, file_path: str, detail: str = "auto") -> str:
    """读取本地图片文件并返回元数据 + base64 注入标记。"""
    try:
        path = _resolve_path(file_path)
    except SecurityViolationError as exc:
        return json.dumps(
            {"status": "error", "message": f"路径校验失败: {exc}"},
            ensure_ascii=False,
        )
    if not path.is_file():
        return json.dumps(
            {"status": "error", "message": f"文件不存在: {file_path}"},
            ensure_ascii=False,
        )
    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        return json.dumps(
            {"status": "error", "message": f"不支持的图片格式: {path.suffix}"},
            ensure_ascii=False,
        )
    size = path.stat().st_size
    if size > _MAX_SIZE_BYTES:
        return json.dumps(
            {
                "status": "error",
                "message": f"文件大小超限: {size} bytes > {_MAX_SIZE_BYTES}",
            },
            ensure_ascii=False,
        )

    raw = path.read_bytes()
    b64 = base64.b64encode(raw).decode("ascii")
    ext = path.suffix.lower()
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".webp": "image/webp",
    }
    mime = mime_map.get(ext, "image/png")

    return json.dumps(
        {
            "status": "ok",
            "mime_type": mime,
            "size_bytes": size,
            "file_path": str(path),
            "hint": "图片已加载到视觉上下文，你现在可以看到这张图片。",
            "__tool_result_image__": {
                "base64": b64,
                "mime_type": mime,
                "detail": detail,
            },
        },
        ensure_ascii=False,
    )


def _infer_number_format(display_text: str) -> str | None:
    """从 display_text 推断 Excel number_format。

    委托给共享模块 ``excelmanus.format_utils``。
    """
    from excelmanus.format_utils import infer_number_format
    return infer_number_format(display_text)


def rebuild_excel_from_spec(*, spec_path: str, output_path: str = "outputs/draft.xlsx") -> str:
    """从 ReplicaSpec JSON 确定性编译为 Excel 文件。"""
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from excelmanus.replica_spec import ReplicaSpec

    try:
        spec_file = _resolve_path(spec_path)
    except SecurityViolationError as exc:
        return json.dumps({"status": "error", "message": f"路径校验失败: {exc}"}, ensure_ascii=False)

    if not spec_file.is_file():
        return json.dumps({"status": "error", "message": f"Spec 文件不存在: {spec_path}"}, ensure_ascii=False)

    try:
        spec = ReplicaSpec.model_validate_json(spec_file.read_text(encoding="utf-8"))
    except Exception as exc:
        return json.dumps({"status": "error", "message": f"Spec 解析失败: {exc}"}, ensure_ascii=False)

    wb = Workbook()
    # 删除默认 sheet
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    # 应用 WorkbookSpec 全局默认字体（替换 openpyxl 内置的 Calibri 11pt）
    if spec.workbook and spec.workbook.default_font:
        df = spec.workbook.default_font
        wb._fonts[0] = Font(
            name=df.name or "Calibri",
            size=df.size or 11,
            bold=df.bold or False,
            italic=df.italic or False,
            color=df.color.lstrip("#") if df.color else None,
        )

    cells_written = 0
    styles_applied = 0
    merges_applied = 0
    formulas_written = 0
    skipped_items: list[str] = []

    for sheet_spec in spec.sheets:
        ws = wb.create_sheet(title=sheet_spec.name)

        # 构建 openpyxl 样式对象缓存
        style_cache: dict[str, dict[str, Any]] = {}
        for style_id, style_class in sheet_spec.styles.items():
            s: dict[str, Any] = {}
            if style_class.font:
                f = style_class.font
                s["font"] = Font(
                    name=f.name,
                    size=f.size,
                    bold=f.bold,
                    italic=f.italic,
                    color=f.color.lstrip("#") if f.color else None,
                )
            if style_class.fill and style_class.fill.color:
                s["fill"] = PatternFill(
                    patternType="solid",
                    fgColor=style_class.fill.color.lstrip("#"),
                )
            if style_class.border:
                b = style_class.border
                # 优先使用四边独立样式，回退到统一样式
                def _make_side(side_spec, fallback_style=None, fallback_color=None):
                    st = side_spec.style if side_spec else fallback_style
                    cl = side_spec.color if side_spec else fallback_color
                    if not st:
                        return Side()
                    return Side(
                        style=st,
                        color=cl.lstrip("#") if cl else None,
                    )
                if b.top or b.bottom or b.left or b.right:
                    # 四边独立模式
                    s["border"] = Border(
                        top=_make_side(b.top, b.style, b.color),
                        bottom=_make_side(b.bottom, b.style, b.color),
                        left=_make_side(b.left, b.style, b.color),
                        right=_make_side(b.right, b.style, b.color),
                    )
                elif b.style:
                    # 统一样式（向后兼容）
                    side = Side(
                        style=b.style,
                        color=b.color.lstrip("#") if b.color else None,
                    )
                    s["border"] = Border(left=side, right=side, top=side, bottom=side)
            if style_class.alignment:
                a = style_class.alignment
                s["alignment"] = Alignment(
                    horizontal=a.horizontal,
                    vertical=a.vertical,
                    wrap_text=a.wrap_text,
                )
            style_cache[style_id] = s

        # 写入 cells
        for cell_spec in sheet_spec.cells:
            cell = ws[cell_spec.address]
            # 值类型转换
            if cell_spec.value_type == "formula" and cell_spec.value and str(cell_spec.value).startswith("="):
                cell.value = str(cell_spec.value)
                formulas_written += 1
            elif cell_spec.formula_candidate and cell_spec.confidence >= 0.8:
                cell.value = cell_spec.formula_candidate
                formulas_written += 1
            elif cell_spec.value_type == "number" and cell_spec.value is not None:
                try:
                    cell.value = float(cell_spec.value) if "." in str(cell_spec.value) else int(cell_spec.value)
                except (ValueError, TypeError):
                    cell.value = cell_spec.value
            elif cell_spec.value_type == "boolean" and cell_spec.value is not None:
                cell.value = bool(cell_spec.value)
            elif cell_spec.value_type == "date" and cell_spec.value is not None:
                try:
                    cell.value = datetime.fromisoformat(str(cell_spec.value))
                except (ValueError, TypeError):
                    cell.value = cell_spec.value
            elif cell_spec.value_type == "empty":
                cell.value = None
            else:
                cell.value = cell_spec.value

            # number_format：优先使用显式指定，否则从 display_text 推断
            if cell_spec.number_format:
                cell.number_format = cell_spec.number_format
            elif cell_spec.display_text and cell_spec.value_type == "number" and cell_spec.value is not None:
                inferred_fmt = _infer_number_format(cell_spec.display_text)
                if inferred_fmt:
                    cell.number_format = inferred_fmt

            # 应用样式
            if cell_spec.style_id and cell_spec.style_id in style_cache:
                s = style_cache[cell_spec.style_id]
                if "font" in s:
                    cell.font = s["font"]
                if "fill" in s:
                    cell.fill = s["fill"]
                if "border" in s:
                    cell.border = s["border"]
                if "alignment" in s:
                    cell.alignment = s["alignment"]
                styles_applied += 1
            else:
                # 无显式样式时，根据 value_type 推断默认对齐
                inferred_h = None
                if cell_spec.value_type == "number":
                    inferred_h = "right"
                elif cell_spec.value_type == "date":
                    inferred_h = "center"
                elif cell_spec.value_type == "string" and cell_spec.value:
                    inferred_h = "left"
                if inferred_h:
                    cell.alignment = Alignment(
                        horizontal=inferred_h, vertical="center",
                    )

            cells_written += 1

        # 合并单元格（安全模式：检测非锚点位置的数据冲突）
        # 构建 spec 中有值的 cell 地址集合
        from openpyxl.utils import range_boundaries

        valued_cells: dict[str, str] = {}  # 地址 → 值的 repr
        for cs in sheet_spec.cells:
            if cs.value is not None and cs.value_type != "empty":
                valued_cells[cs.address.upper()] = repr(cs.value)

        for mr in sheet_spec.merged_ranges:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(mr.range)
                anchor = f"{get_column_letter(min_col)}{min_row}"
                # 检测非锚点位置是否有 spec 定义的值
                conflict_cells: list[str] = []
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        addr = f"{get_column_letter(c)}{r}"
                        if addr.upper() != anchor.upper() and addr.upper() in valued_cells:
                            conflict_cells.append(addr)
                if conflict_cells:
                    skipped_items.append(
                        f"merge {mr.range} 跳过: 非锚点单元格 {', '.join(conflict_cells)} "
                        f"含有值，合并会导致数据丢失"
                    )
                    continue
                ws.merge_cells(mr.range)
                merges_applied += 1
                # 合并单元格锚点强制居中对齐
                anchor_cell = ws[anchor]
                if not anchor_cell.alignment or (
                    anchor_cell.alignment.horizontal in (None, "general")
                ):
                    anchor_cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                    )
            except Exception as exc:
                skipped_items.append(f"merge {mr.range}: {exc}")

        # 列宽（容错处理：数组长度与实际列数不匹配时做截断/忽略）
        max_col_used = ws.max_column or 1
        for i, width in enumerate(sheet_spec.column_widths):
            if i >= max_col_used + 10:  # 允许少量溢出，超过太多截断
                break
            try:
                col_letter = get_column_letter(i + 1)
                if isinstance(width, (int, float)) and width > 0:
                    ws.column_dimensions[col_letter].width = width
            except (ValueError, TypeError):
                skipped_items.append(f"column_width[{i}]: 无效值 {width!r}")

        # 行高
        for row_str, height in sheet_spec.row_heights.items():
            try:
                ws.row_dimensions[int(row_str)].height = height
            except (ValueError, TypeError):
                pass

        # 冻结窗格
        if sheet_spec.freeze_panes:
            ws.freeze_panes = sheet_spec.freeze_panes

    # ── auto_fit 收尾：对列宽/行高缺失或不完整的 sheet 自动适配 ──
    from excelmanus.tools.format_tools import (
        _COL_PADDING,
        _MAX_COL_WIDTH,
        _MIN_COL_WIDTH,
        _estimate_display_width,
        _estimate_row_height,
    )
    from openpyxl.cell.cell import MergedCell as _MergedCell

    for sheet_spec in spec.sheets:
        ws = wb[sheet_spec.name]
        # 列宽 auto_fit：当 spec 未提供足够的列宽时补充
        has_col_widths = bool(sheet_spec.column_widths)
        if not has_col_widths:
            merged_non_anchor: set[str] = set()
            for mr in ws.merged_cells.ranges:
                a = f"{get_column_letter(mr.min_col)}{mr.min_row}"
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        addr = f"{get_column_letter(c)}{r}"
                        if addr != a:
                            merged_non_anchor.add(addr)
            for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row or 1):
                max_w = 0.0
                cl = get_column_letter(col_cells[0].column)
                for c in col_cells:
                    coord = f"{cl}{c.row}"
                    if coord in merged_non_anchor or c.value is None:
                        continue
                    fs = 11.0
                    ib = False
                    if c.font:
                        if c.font.size:
                            fs = float(c.font.size)
                        if c.font.bold:
                            ib = True
                    nf = c.number_format if c.number_format != "General" else None
                    w = _estimate_display_width(c.value, fs, ib, nf)
                    if w > max_w:
                        max_w = w
                width = max(_MIN_COL_WIDTH, min(max_w + _COL_PADDING, _MAX_COL_WIDTH))
                ws.column_dimensions[cl].width = width

        # 行高 auto_fit：当 spec 未提供行高时补充
        has_row_heights = bool(sheet_spec.row_heights)
        if not has_row_heights:
            col_widths_map: dict[str, float] = {}
            for cl_letter, dim in ws.column_dimensions.items():
                if dim.width is not None:
                    col_widths_map[cl_letter] = dim.width
            for row_idx in range(1, (ws.max_row or 0) + 1):
                row_cells = [c for c in ws[row_idx] if not isinstance(c, _MergedCell)]
                if not row_cells:
                    continue
                h = _estimate_row_height(row_cells, col_widths_map)
                ws.row_dimensions[row_idx].height = h

    # 保存
    try:
        out = _resolve_path(output_path)
    except SecurityViolationError as exc:
        return json.dumps({"status": "error", "message": f"路径校验失败: {exc}"}, ensure_ascii=False)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return json.dumps({
        "status": "ok",
        "output_path": str(out),
        "build_summary": {
            "cells_written": cells_written,
            "styles_applied": styles_applied,
            "merges_applied": merges_applied,
            "formulas_written": formulas_written,
            "skipped_items": skipped_items,
        },
    }, ensure_ascii=False)


def verify_excel_replica(
    *, spec_path: str, excel_path: str, report_path: str = "outputs/replica_diff_report.md",
) -> str:
    """验证 Excel 文件与 ReplicaSpec 的一致性，生成差异报告。"""
    from openpyxl import load_workbook

    from excelmanus.replica_spec import ReplicaSpec

    try:
        spec_file = _resolve_path(spec_path)
        excel_file = _resolve_path(excel_path)
        rp = _resolve_path(report_path)
    except SecurityViolationError as exc:
        return json.dumps({"status": "error", "message": f"路径校验失败: {exc}"}, ensure_ascii=False)

    if not spec_file.is_file():
        return json.dumps({"status": "error", "message": f"Spec 文件不存在: {spec_path}"}, ensure_ascii=False)
    if not excel_file.is_file():
        return json.dumps({"status": "error", "message": f"Excel 文件不存在: {excel_path}"}, ensure_ascii=False)

    # .xls/.xlsb → 透明转换为 xlsx
    from excelmanus.tools._helpers import ensure_openpyxl_compatible
    excel_file = ensure_openpyxl_compatible(excel_file)

    try:
        spec = ReplicaSpec.model_validate_json(spec_file.read_text(encoding="utf-8"))
    except Exception as exc:
        return json.dumps({"status": "error", "message": f"Spec 解析失败: {exc}"}, ensure_ascii=False)

    try:
        wb = load_workbook(str(excel_file))
    except Exception as exc:
        return json.dumps({"status": "error", "message": f"Excel 加载失败: {exc}"}, ensure_ascii=False)

    matches = 0
    mismatches: list[str] = []
    merge_conflicts: list[str] = []
    missing: list[str] = []
    low_confidence: list[str] = []
    style_diffs: list[str] = []
    total_cells = 0

    for sheet_spec in spec.sheets:
        if sheet_spec.name not in wb.sheetnames:
            missing.append(f"Sheet '{sheet_spec.name}' 不存在于 Excel 中")
            continue
        ws = wb[sheet_spec.name]

        # 构建合并区域查找表：addr → (anchor_addr, merge_range_str)
        merge_lookup: dict[str, tuple[str, str]] = {}
        from openpyxl.utils import get_column_letter
        for merged_range in ws.merged_cells.ranges:
            anchor_addr = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    addr = f"{get_column_letter(c)}{r}"
                    if addr.upper() != anchor_addr.upper():
                        merge_lookup[addr.upper()] = (anchor_addr, str(merged_range))

        # 值比对
        for cell_spec in sheet_spec.cells:
            total_cells += 1
            try:
                actual = ws[cell_spec.address].value
            except Exception:
                missing.append(f"{sheet_spec.name}!{cell_spec.address}: 无法读取")
                continue

            expected = cell_spec.value
            # 类型感知比较
            if _values_match(expected, actual):
                matches += 1
            elif actual is None and cell_spec.address.upper() in merge_lookup:
                # 该 cell 在合并区域的非锚点位置，值被合并操作清零
                anchor, mr_str = merge_lookup[cell_spec.address.upper()]
                merge_conflicts.append(
                    f"{sheet_spec.name}!{cell_spec.address}: 期望={expected!r} "
                    f"但该单元格在合并区域 {mr_str} 内（锚点={anchor}），值被合并覆盖"
                )
                # 仍计为匹配（数据在锚点可读，这是 merge 的预期行为）
                matches += 1
            else:
                mismatches.append(
                    f"{sheet_spec.name}!{cell_spec.address}: 期望={expected!r} 实际={actual!r}"
                )

        # 合并比对
        actual_merges = {str(m) for m in ws.merged_cells.ranges}
        for mr in sheet_spec.merged_ranges:
            if mr.range not in actual_merges:
                mismatches.append(f"{sheet_spec.name}: 合并范围 {mr.range} 缺失")

        # ── 样式维度验证 ──────────────────────────────────

        # 对齐比对
        for cell_spec in sheet_spec.cells:
            if not cell_spec.style_id or cell_spec.style_id not in sheet_spec.styles:
                continue
            style_cls = sheet_spec.styles[cell_spec.style_id]
            if not style_cls.alignment:
                continue
            try:
                actual_cell = ws[cell_spec.address]
                actual_align = actual_cell.alignment
            except Exception:
                continue
            spec_h = style_cls.alignment.horizontal
            spec_v = style_cls.alignment.vertical
            actual_h = actual_align.horizontal if actual_align else None
            actual_v = actual_align.vertical if actual_align else None
            # 'general' 和 None 视为等价
            norm_h = lambda x: None if x in (None, "general") else x
            if norm_h(spec_h) and norm_h(spec_h) != norm_h(actual_h):
                style_diffs.append(
                    f"{sheet_spec.name}!{cell_spec.address}: "
                    f"水平对齐 期望={spec_h} 实际={actual_h}"
                )
            if spec_v and spec_v != (actual_v or "bottom"):
                style_diffs.append(
                    f"{sheet_spec.name}!{cell_spec.address}: "
                    f"垂直对齐 期望={spec_v} 实际={actual_v}"
                )

        # 列宽偏差（容差 ±2）
        max_col_used = ws.max_column or 1
        for i, expected_w in enumerate(sheet_spec.column_widths):
            if i >= max_col_used + 10:
                break
            col_letter = get_column_letter(i + 1)
            actual_dim = ws.column_dimensions.get(col_letter)
            actual_w = actual_dim.width if actual_dim and actual_dim.width else 8.0
            if abs(expected_w - actual_w) > 2.0:
                style_diffs.append(
                    f"{sheet_spec.name} 列{col_letter}: "
                    f"列宽 期望={expected_w:.1f} 实际={actual_w:.1f}"
                )

        # 行高偏差（容差 ±3）
        for row_str, expected_h in sheet_spec.row_heights.items():
            try:
                row_num = int(row_str)
            except ValueError:
                continue
            actual_dim = ws.row_dimensions.get(row_num)
            actual_h = actual_dim.height if actual_dim and actual_dim.height else 15.0
            if abs(expected_h - actual_h) > 3.0:
                style_diffs.append(
                    f"{sheet_spec.name} 行{row_num}: "
                    f"行高 期望={expected_h:.1f} 实际={actual_h:.1f}"
                )

    # 收集低置信项
    for u in spec.uncertainties:
        low_confidence.append(f"{u.location}: {u.reason} (置信度={u.confidence:.0%})")

    match_rate = matches / total_cells if total_cells > 0 else 1.0

    # 生成 Markdown 报告
    report_lines = [
        "# ReplicaSpec 验证报告\n",
        f"**匹配率**: {match_rate:.1%} ({matches}/{total_cells})\n",
    ]
    if not mismatches and not missing and not merge_conflicts and not style_diffs:
        report_lines.append("## ✅ 全部匹配\n")
    if mismatches:
        report_lines.append(f"## ❌ 不匹配项 ({len(mismatches)})\n")
        for m in mismatches:
            report_lines.append(f"- {m}")
        report_lines.append("")
    if merge_conflicts:
        report_lines.append(f"## 🔀 合并单元格冲突 ({len(merge_conflicts)})\n")
        report_lines.append("以下单元格在合并区域非锚点位置，值已被合并覆盖（不影响匹配率）：\n")
        for mc in merge_conflicts:
            report_lines.append(f"- {mc}")
        report_lines.append("")
    if missing:
        report_lines.append(f"## ⚠️ 缺失项 ({len(missing)})\n")
        for m in missing:
            report_lines.append(f"- {m}")
        report_lines.append("")
    if style_diffs:
        report_lines.append(f"## 📐 样式偏差 ({len(style_diffs)})\n")
        report_lines.append("以下样式属性与 Spec 不一致（对齐/列宽/行高）：\n")
        for sd in style_diffs:
            report_lines.append(f"- {sd}")
        report_lines.append("")
    if low_confidence:
        report_lines.append(f"## 🔍 低置信项 ({len(low_confidence)})\n")
        report_lines.append("以下项目建议人工确认：\n")
        for lc in low_confidence:
            report_lines.append(f"- {lc}")
        report_lines.append("")

    report_text = "\n".join(report_lines)
    rp.parent.mkdir(parents=True, exist_ok=True)
    rp.write_text(report_text, encoding="utf-8")

    return json.dumps({
        "status": "ok",
        "report_path": str(rp),
        "match_rate": round(match_rate, 4),
        "issues": {
            "missing": len(missing),
            "conflict": len(mismatches),
            "merge_conflicts": len(merge_conflicts),
            "style_diffs": len(style_diffs),
            "low_confidence": len(low_confidence),
            "total": len(missing) + len(mismatches) + len(style_diffs) + len(low_confidence),
        },
    }, ensure_ascii=False)


def _values_match(expected: Any, actual: Any) -> bool:
    """类型感知的值比较。"""
    if expected is None and actual is None:
        return True
    if expected is None or actual is None:
        return False
    # 数值比较（int/float 互通）
    try:
        if isinstance(expected, (int, float)) or isinstance(actual, (int, float)):
            return abs(float(expected) - float(actual)) < 1e-9
    except (ValueError, TypeError):
        pass
    # 日期归一化比较
    d_expected = _normalize_to_date(expected)
    d_actual = _normalize_to_date(actual)
    if d_expected is not None and d_actual is not None:
        return d_expected == d_actual
    return str(expected) == str(actual)

def _normalize_to_date(val: Any) -> "date | None":
    """尝试将值归一化为 date 对象。"""
    from datetime import date, datetime

    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None



def get_tools() -> list[ToolDef]:
    return [
        ToolDef(
            name="read_image",
            description=(
                "读取本地图片文件并加载到视觉上下文（png/jpg/gif/bmp/webp）。"
                "适用场景：查看图片内容、分析截图中的文字或数据、确认图表样式。"
                "不适用：需要将图片中的表格还原为 Excel（改用 extract_table_spec → rebuild_excel_from_spec 工具链）。"
                "相关工具：extract_table_spec（从图片提取表格结构）、rebuild_excel_from_spec（编译为 Excel）。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "图片文件的绝对或相对路径",
                    },
                    "detail": {
                        "type": "string",
                        "enum": ["auto", "low", "high"],
                        "default": "auto",
                        "description": "图片分析精度",
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=read_image,
            max_result_chars=2000,  # 注入后 base64 已移除，仅剩元数据
            write_effect="none",
        ),
        ToolDef(
            name="rebuild_excel_from_spec",
            description=(
                "从 ReplicaSpec JSON 确定性编译为 Excel 文件。"
                "适用场景：将 extract_table_spec 提取的表格规格编译为真实 Excel 文件。"
                "不适用：从零创建 Excel（改用 run_code + openpyxl）。"
                "工具链：read_image → extract_table_spec → rebuild_excel_from_spec → verify_excel_replica。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "spec_path": {
                        "type": "string",
                        "description": "ReplicaSpec JSON 文件路径",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "输出 Excel 文件路径",
                        "default": "outputs/draft.xlsx",
                    },
                },
                "required": ["spec_path"],
                "additionalProperties": False,
            },
            func=rebuild_excel_from_spec,
            write_effect="workspace_write",
        ),
        ToolDef(
            name="verify_excel_replica",
            description=(
                "验证 Excel 文件与 ReplicaSpec 的一致性，生成差异报告。"
                "适用场景：rebuild_excel_from_spec 后校验产出物是否与规格一致。"
                "不适用：普通数据校验（改用 run_code + pandas 比较）。"
                "工具链：rebuild_excel_from_spec 之后调用本工具验证。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "spec_path": {
                        "type": "string",
                        "description": "ReplicaSpec JSON 文件路径",
                    },
                    "excel_path": {
                        "type": "string",
                        "description": "要验证的 Excel 文件路径",
                    },
                    "report_path": {
                        "type": "string",
                        "description": "差异报告输出路径",
                        "default": "outputs/replica_diff_report.md",
                    },
                },
                "required": ["spec_path", "excel_path"],
                "additionalProperties": False,
            },
            func=verify_excel_replica,
            write_effect="workspace_write",
        ),
        ToolDef(
            name="extract_table_spec",
            description=(
                "从图片自动提取表格结构和样式，生成 ReplicaSpec JSON。"
                "支持多表格检测，采用 4 阶段渐进式 VLM 提取（骨架 → 数据 → 样式 → 校验）。"
                "支持批量处理多个图片文件，大幅减少 VLM API 调用次数。"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "单张图片文件路径（与 file_paths 二选一）",
                    },
                    "file_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "多张图片文件路径数组，批量处理时自动使用优化的批量管线",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "输出 ReplicaSpec JSON 路径（批量模式时为目录前缀）",
                        "default": "outputs/replica_spec.json",
                    },
                    "skip_style": {
                        "type": "boolean",
                        "description": "跳过样式提取（仅提取数据结构，速度更快）",
                        "default": False,
                    },
                },
                "additionalProperties": False,
            },
            func=lambda **kw: json.dumps({"__extract_pending__": True}),
            write_effect="workspace_write",
            max_result_chars=5000,
        ),
    ]
