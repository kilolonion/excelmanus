"""工具层共享辅助函数。"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from difflib import SequenceMatcher
from typing import Any, Sequence

_logger = logging.getLogger(__name__)

# 文件存在性检查时，最多列出的可用文件数量
_MAX_SUGGESTION_FILES = 15
_EXCEL_SUFFIXES: frozenset[str] = frozenset({".xlsx", ".xls", ".xlsm", ".xlsb"})


def check_file_exists(safe_path: Path, user_path: str, guard: Any) -> str | None:
    """检查文件是否存在，不存在时返回结构化错误 JSON（含可用文件提示）。

    Args:
        safe_path: 经 guard.resolve_and_validate 后的绝对路径。
        user_path: 用户/LLM 原始传入的 file_path 字符串。
        guard: FileAccessGuard 实例（用于获取 workspace_root）。

    Returns:
        文件存在返回 None；不存在返回 JSON 错误字符串。
    """
    if safe_path.is_file():
        return None

    workspace_root: Path = guard.workspace_root
    suggestions: list[str] = []

    # 1) 列出目标目录下的 Excel 文件
    parent = safe_path.parent
    if parent.is_dir():
        try:
            for f in sorted(parent.iterdir()):
                if f.is_file() and f.suffix.lower() in _EXCEL_SUFFIXES:
                    try:
                        suggestions.append(str(f.relative_to(workspace_root)))
                    except ValueError:
                        suggestions.append(f.name)
        except OSError:
            pass

    # 2) 若目标目录下没找到，扫描工作区根目录（仅第一层）
    if not suggestions:
        try:
            for f in sorted(workspace_root.iterdir()):
                if f.is_file() and f.suffix.lower() in _EXCEL_SUFFIXES:
                    suggestions.append(f.name)
                if len(suggestions) >= _MAX_SUGGESTION_FILES:
                    break
        except OSError:
            pass

    # 3) 若仍为空，递归扫描工作区（限深度 2）
    if not suggestions:
        try:
            for f in sorted(workspace_root.rglob("*")):
                if f.is_file() and f.suffix.lower() in _EXCEL_SUFFIXES:
                    try:
                        suggestions.append(str(f.relative_to(workspace_root)))
                    except ValueError:
                        suggestions.append(str(f))
                    if len(suggestions) >= _MAX_SUGGESTION_FILES:
                        break
        except OSError:
            pass

    payload: dict[str, Any] = {
        "error": f"文件不存在: {user_path}",
        "hint": "请检查文件路径是否正确，或使用 inspect_excel_files / list_directory 确认可用文件。",
    }
    if suggestions:
        payload["available_excel_files"] = suggestions[:_MAX_SUGGESTION_FILES]

    return json.dumps(payload, ensure_ascii=False)


# fuzzy matching 相似度阈值：≥ 此值时自动纠正
_FUZZY_MATCH_THRESHOLD: float = 0.6


def resolve_sheet_name(
    requested: str | None,
    available: Sequence[str],
) -> str | None:
    """对 sheet 名做三级模糊匹配。

    优先精确匹配；若失败则尝试忽略大小写匹配；
    若仍失败则尝试 SequenceMatcher fuzzy 匹配（阈值 0.6）。
    匹配成功返回 *实际* sheet 名（保留原始大小写），
    匹配失败返回 ``None``。

    Args:
        requested: LLM / 用户传入的 sheet 名，可为 None。
        available: workbook 中实际存在的 sheet 名列表。

    Returns:
        匹配到的实际 sheet 名，或 None。
    """
    if requested is None:
        return None

    # 精确匹配
    if requested in available:
        return requested

    # 大小写不敏感回退匹配
    lower = requested.lower()
    for name in available:
        if name.lower() == lower:
            return name

    # 第三级：fuzzy matching（SequenceMatcher）
    best_name, best_ratio = _find_closest_sheet_name(requested, available)
    if best_name is not None and best_ratio >= _FUZZY_MATCH_THRESHOLD:
        _logger.info(
            "sheet 名 fuzzy 纠正: '%s' → '%s' (相似度 %.2f)",
            requested, best_name, best_ratio,
        )
        return best_name

    return None


def _find_closest_sheet_name(
    requested: str,
    available: Sequence[str],
) -> tuple[str | None, float]:
    """在 available 中找到与 requested 最接近的 sheet 名。

    Returns:
        (best_match_name, best_ratio)，无候选时返回 (None, 0.0)。
    """
    if not available:
        return None, 0.0
    req_lower = requested.lower()
    best_name: str | None = None
    best_ratio: float = 0.0
    for name in available:
        ratio = SequenceMatcher(None, req_lower, name.lower()).ratio()
        if ratio > best_ratio:
            best_name = name
            best_ratio = ratio
    return best_name, best_ratio


def get_worksheet(wb: Any, sheet_name: str | None) -> Any:
    """从 workbook 中获取工作表，支持 case-insensitive fallback。

    若 sheet_name 为 None，返回 active sheet。
    若 sheet_name 非空但无法匹配，抛出 ValueError 并附带可用 sheet 列表。

    Args:
        wb: openpyxl Workbook 对象。
        sheet_name: 请求的 sheet 名。

    Returns:
        匹配到的 Worksheet 对象。

    Raises:
        ValueError: sheet_name 非空但在 workbook 中未找到匹配项。
    """
    if not sheet_name:
        return wb.active
    resolved = resolve_sheet_name(sheet_name, wb.sheetnames)
    if resolved is not None:
        return wb[resolved]
    raise ValueError(
        f"工作表 '{sheet_name}' 不存在。"
        f"该文件包含以下工作表: {wb.sheetnames}。"
        f"请使用正确的工作表名称重试。"
    )


def check_sheet_name(safe_path: Path, sheet_name: str | None) -> tuple[str | None, str | None]:
    """验证 sheet_name 是否存在于 Excel 文件中，不存在时返回结构化错误。

    优先精确匹配，然后 case-insensitive 回退。
    匹配成功返回 (resolved_name, None)；
    匹配失败返回 (None, error_json_str)。
    sheet_name 为 None 时直接返回 (None, None)（使用默认 sheet）。

    Args:
        safe_path: 经 guard.resolve_and_validate 后的绝对路径。
        sheet_name: LLM / 用户传入的 sheet 名。

    Returns:
        (resolved_sheet_name, error_json_or_none) 元组。
    """
    if sheet_name is None:
        return None, None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(safe_path, read_only=True, data_only=True)
        try:
            available = wb.sheetnames
            resolved = resolve_sheet_name(sheet_name, available)
            if resolved is not None:
                # fuzzy 纠正时附带提示（非精确匹配才提示）
                if resolved != sheet_name and resolved.lower() != sheet_name.lower():
                    _logger.info(
                        "check_sheet_name: fuzzy 纠正 '%s' → '%s'",
                        sheet_name, resolved,
                    )
                return resolved, None
            # 未找到：构造含可用 sheet 列表的结构化错误
            payload: dict[str, Any] = {
                "error": f"工作表 '{sheet_name}' 不存在",
                "available_sheets": available,
            }
            # 附加最相似的 sheet 名建议（即使低于自动纠正阈值）
            closest, ratio = _find_closest_sheet_name(sheet_name, available)
            if closest is not None and ratio > 0.3:
                payload["closest_match"] = closest
                payload["similarity"] = round(ratio, 2)
                payload["hint"] = (
                    f"该文件包含以下工作表: {available}。"
                    f"最接近的是 '{closest}'（相似度 {ratio:.0%}），请确认后重试。"
                )
            else:
                payload["hint"] = f"该文件包含以下工作表: {available}。请使用正确的工作表名称重试。"
            return None, json.dumps(payload, ensure_ascii=False)
        finally:
            wb.close()
    except Exception as exc:
        _logger.warning("check_sheet_name 异常: %s", exc)
        return sheet_name, None  # 异常时放行，让下游处理


def ensure_openpyxl_compatible(safe_path: Path) -> Path:
    """确保路径指向 openpyxl 可操作的文件格式（.xlsx/.xlsm）。

    若为 .xls/.xlsb，透明转换为同目录 .xlsx 并返回新路径。
    转换结果会被缓存（同名 .xlsx 已存在时跳过转换）。
    CSV 文件原样返回（由调用方处理）。

    Args:
        safe_path: 经 guard.resolve_and_validate 后的绝对路径。

    Returns:
        openpyxl 可直接打开的文件路径。
    """
    from excelmanus.xls_converter import needs_conversion, ensure_xlsx

    if not needs_conversion(safe_path):
        return safe_path

    try:
        xlsx_path, converted = ensure_xlsx(safe_path)
        if converted:
            _logger.info("工具层自动转换: %s → %s", safe_path.name, xlsx_path.name)
        return xlsx_path
    except Exception as exc:
        _logger.warning("工具层 xls 转换失败，返回原路径: %s (%s)", safe_path.name, exc)
        return safe_path
