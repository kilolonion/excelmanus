"""图表工具：支持柱状图、折线图、饼图、散点图和雷达图。"""

from __future__ import annotations

import json
from functools import lru_cache
import warnings
from typing import Any

import matplotlib
matplotlib.use("Agg")  # 非交互式后端，必须在导入 pyplot 之前设置

from matplotlib import font_manager
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from openpyxl import load_workbook

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools import data_tools
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.chart")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "chart"
SKILL_DESCRIPTION = "可视化工具集：支持柱状图、折线图、饼图、散点图和雷达图"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None

SUPPORTED_CHART_TYPES = ("bar", "line", "pie", "scatter", "radar")
CJK_FONT_CANDIDATES = (
    "Noto Sans CJK SC",
    "Microsoft YaHei",
    "PingFang SC",
    "WenQuanYi Zen Hei",
    "Arial Unicode MS",
    "STHeiti",
    "Songti SC",
    "SimHei",
)


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard 单例。"""
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


@lru_cache(maxsize=1)
def _select_cjk_font() -> str | None:
    """选择一个可用的中文字体，避免中文字符渲染警告。"""
    available = {font.name for font in font_manager.fontManager.ttflist}
    for font_name in CJK_FONT_CANDIDATES:
        if font_name in available:
            return font_name
    return None


def _contains_cjk(text: str) -> bool:
    """判断文本是否包含 CJK 统一表意文字（常见中文字符范围）。"""
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def _configure_plot_fonts(labels: pd.Series, title: str) -> None:
    """根据数据内容配置字体，优先选择可用中文字体。"""
    sample_texts = [title]
    sample_texts.extend(str(v) for v in labels.head(20).tolist())
    needs_cjk = any(_contains_cjk(text) for text in sample_texts if text)

    if needs_cjk:
        cjk_font = _select_cjk_font()
        if cjk_font:
            plt.rcParams["font.sans-serif"] = [cjk_font, "DejaVu Sans"]
        else:
            plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
            logger.warning("未检测到可用中文字体，图表中文可能显示异常。")
    else:
        plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]

    plt.rcParams["axes.unicode_minus"] = False


# ── 工具函数 ──────────────────────────────────────────────


def create_chart(
    file_path: str,
    chart_type: str,
    x_column: str,
    y_column: str,
    output_path: str,
    title: str | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> str:
    """从 Excel 数据生成图表并保存为图片文件。

    Args:
        file_path: 数据源 Excel 文件路径。
        chart_type: 图表类型，支持 bar/line/pie/scatter/radar。
        x_column: X 轴（或标签）列名。
        y_column: Y 轴（或数值）列名。
        output_path: 输出图片文件路径（如 output.png）。
        title: 图表标题，默认自动生成。
        sheet_name: 工作表名称，默认读取第一个。
        header_row: 表头行号（从0开始），默认自动检测。

    Returns:
        JSON 格式的操作结果。
    """
    if chart_type not in SUPPORTED_CHART_TYPES:
        return json.dumps(
            {"status": "error", "message": f"不支持的图表类型 '{chart_type}'，支持: {list(SUPPORTED_CHART_TYPES)}"},
            ensure_ascii=False,
        )

    guard = _get_guard()
    safe_input = guard.resolve_and_validate(file_path)
    safe_output = guard.resolve_and_validate(output_path)

    # 读取数据
    kwargs: dict[str, Any] = {"io": safe_input}
    detected_header_row: int | None = None
    if sheet_name is not None:
        kwargs["sheet_name"] = sheet_name
    if header_row is not None:
        kwargs["header"] = header_row
    else:
        detected = data_tools._detect_header_row(safe_input, sheet_name)
        if detected is not None and detected > 0:
            kwargs["header"] = detected
            detected_header_row = detected

    df = pd.read_excel(**kwargs)

    # 校验列名
    for col_name, col_label in [(x_column, "x_column"), (y_column, "y_column")]:
        if col_name not in df.columns:
            return json.dumps(
                {"status": "error", "message": f"{col_label} '{col_name}' 不存在，可用列: {list(df.columns)}"},
                ensure_ascii=False,
            )

    plot_df = df[[x_column, y_column]].dropna()
    if plot_df.empty:
        return json.dumps(
            {"status": "error", "message": f"列 '{x_column}' 与 '{y_column}' 没有可绘图的数据"},
            ensure_ascii=False,
        )

    if chart_type == "radar" and len(plot_df) < 3:
        return json.dumps(
            {"status": "error", "message": "雷达图至少需要 3 条有效数据"},
            ensure_ascii=False,
        )

    x_data = plot_df[x_column]
    y_data = plot_df[y_column]
    chart_title = title or f"{y_column} by {x_column}"

    _configure_plot_fonts(x_data, chart_title)

    fig, ax = plt.subplots(figsize=(10, 6))

    try:
        # 部分环境缺少完整 CJK 字体时，抑制 matplotlib 的缺字 warning，避免污染日志。
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                message=r"Glyph .* missing from font\(s\) .*",
            )

            if chart_type == "bar":
                ax.bar(x_data.astype(str), y_data)
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)
                plt.xticks(rotation=45, ha="right")

            elif chart_type == "line":
                ax.plot(x_data, y_data, marker="o")
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)

            elif chart_type == "pie":
                ax.pie(y_data, labels=x_data.astype(str), autopct="%1.1f%%")
                ax.set_aspect("equal")

            elif chart_type == "scatter":
                ax.scatter(x_data, y_data)
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)

            elif chart_type == "radar":
                # 雷达图需要特殊处理
                plt.close(fig)
                fig, ax = _draw_radar(x_data, y_data, chart_title)

            ax.set_title(chart_title)
            fig.tight_layout()
            fig.savefig(safe_output, dpi=150, bbox_inches="tight")

    finally:
        plt.close(fig)

    logger.info("已生成 %s 图表 -> %s", chart_type, safe_output.name)

    result: dict[str, Any] = {"status": "success", "chart_type": chart_type, "output_file": str(safe_output.name)}
    if header_row is None and detected_header_row is not None:
        result["detected_header_row"] = detected_header_row
    return json.dumps(result, ensure_ascii=False)


def _draw_radar(
    labels: pd.Series, values: pd.Series, title: str
) -> tuple[plt.Figure, plt.Axes]:
    """绘制雷达图。

    Args:
        labels: 各维度标签。
        values: 各维度数值。
        title: 图表标题。

    Returns:
        (fig, ax) 元组。
    """
    categories = labels.astype(str).tolist()
    vals = values.tolist()
    n = len(categories)
    if n == 0:
        raise ValueError("雷达图数据为空")
    if n < 3:
        raise ValueError("雷达图至少需要 3 个维度")

    # 计算角度（均匀分布）
    angles = [i / n * 2 * np.pi for i in range(n)]
    # 闭合多边形
    vals_closed = vals + [vals[0]]
    angles_closed = angles + [angles[0]]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw={"projection": "polar"})
    ax.plot(angles_closed, vals_closed, "o-", linewidth=2)
    ax.fill(angles_closed, vals_closed, alpha=0.25)
    ax.set_xticks(angles)
    ax.set_xticklabels(categories)
    ax.set_title(title, y=1.08)

    return fig, ax


EXCEL_CHART_TYPES = ("bar", "line", "pie", "scatter", "area")


def create_excel_chart(
    file_path: str,
    chart_type: str,
    data_range: str,
    categories_range: str | None = None,
    sheet_name: str | None = None,
    target_cell: str = "A1",
    target_sheet: str | None = None,
    title: str | None = None,
    x_title: str | None = None,
    y_title: str | None = None,
    style: int | None = None,
    width: float = 15.0,
    height: float = 10.0,
    from_rows: bool = False,
) -> str:
    """在 Excel 工作表中插入原生图表对象（嵌入式图表，非图片）。

    Args:
        file_path: Excel 文件路径。
        chart_type: 图表类型，支持 bar/line/pie/scatter/area。
        data_range: 数据区域引用（如 "B1:B20"），包含数值数据。
            多系列时可指定多列范围（如 "B1:D20"）。
        categories_range: 分类轴标签区域（如 "A2:A20"），可选。
        sheet_name: 数据源工作表名称，默认活动工作表。
        target_cell: 图表放置位置的锚点单元格（如 "E1"），默认 "A1"。
        target_sheet: 图表放置的目标工作表，默认与数据源相同。
        title: 图表标题。
        x_title: X 轴标题。
        y_title: Y 轴标题。
        style: Excel 图表样式编号（1-48），可选。
        width: 图表宽度（厘米），默认 15。
        height: 图表高度（厘米），默认 10。
        from_rows: 是否按行读取数据（默认按列），True 则每一行是一个系列。

    Returns:
        JSON 格式的操作结果。
    """
    from openpyxl.chart import (
        AreaChart,
        BarChart,
        LineChart,
        PieChart,
        Reference,
        ScatterChart,
    )

    if chart_type not in EXCEL_CHART_TYPES:
        return json.dumps(
            {"status": "error", "message": f"不支持的图表类型 '{chart_type}'，支持: {list(EXCEL_CHART_TYPES)}"},
            ensure_ascii=False,
        )

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # 创建图表对象
    chart_class_map = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }
    chart = chart_class_map[chart_type]()

    if title:
        chart.title = title
    if style is not None:
        chart.style = style
    chart.width = width
    chart.height = height

    # 非饼图/散点图设置轴标题
    if chart_type not in ("pie",):
        if x_title:
            chart.x_axis.title = x_title
        if y_title:
            chart.y_axis.title = y_title

    # 解析数据范围
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(data_range)

    data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

    # 分类轴
    cats_ref = None
    if categories_range:
        c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(categories_range)
        cats_ref = Reference(ws, min_col=c_min_col, min_row=c_min_row, max_col=c_max_col, max_row=c_max_row)

    if chart_type == "scatter":
        # 散点图需要特殊处理：X 值引用 + Y 值引用
        from openpyxl.chart import Series as ChartSeries
        if cats_ref is not None:
            x_values = cats_ref
        else:
            # 默认取第一列作为 X
            x_values = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        # 数据列从第二列开始
        for col_idx in range(min_col if cats_ref else min_col + 1, max_col + 1):
            y_values = Reference(ws, min_col=col_idx, min_row=min_row + 1, max_row=max_row)
            series = ChartSeries(y_values, xvalues=x_values, title_from_data=False)
            chart.series.append(series)
    else:
        chart.add_data(data_ref, titles_from_data=True, from_rows=from_rows)
        if cats_ref is not None:
            chart.set_categories(cats_ref)

    # 放置图表到目标工作表
    target_ws = ws
    if target_sheet and target_sheet in wb.sheetnames:
        target_ws = wb[target_sheet]
    elif target_sheet and target_sheet not in wb.sheetnames:
        target_ws = wb.create_sheet(title=target_sheet)

    target_ws.add_chart(chart, target_cell)
    wb.save(safe_path)
    wb.close()

    logger.info(
        "create_excel_chart: %s[%s] %s at %s",
        safe_path.name, target_ws.title, chart_type, target_cell,
    )
    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "chart_type": chart_type,
            "data_range": data_range,
            "target_sheet": target_ws.title,
            "target_cell": target_cell,
        },
        ensure_ascii=False,
    )


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回可视化 Skill 的所有工具定义。"""
    return [
        ToolDef(
            name="create_chart",
            description="从 Excel 数据生成图表（柱状图、折线图、饼图、散点图、雷达图）并保存为图片",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "数据源 Excel 文件路径",
                    },
                    "chart_type": {
                        "type": "string",
                        "enum": ["bar", "line", "pie", "scatter", "radar"],
                        "description": "图表类型",
                    },
                    "x_column": {
                        "type": "string",
                        "description": "X 轴（或标签）列名",
                    },
                    "y_column": {
                        "type": "string",
                        "description": "Y 轴（或数值）列名",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "输出图片文件路径（如 chart.png）",
                    },
                    "title": {
                        "type": "string",
                        "description": "图表标题，默认自动生成",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认读取第一个",
                    },
                    "header_row": {
                        "type": "integer",
                        "description": "表头行号（从0开始），默认自动检测",
                    },
                },
                "required": ["file_path", "chart_type", "x_column", "y_column", "output_path"],
                "additionalProperties": False,
            },
            func=create_chart,
        ),
        ToolDef(
            name="create_excel_chart",
            description=(
                "在 Excel 工作表中插入原生图表对象（嵌入式，非图片）。"
                "支持 bar/line/pie/scatter/area。"
                "通过 data_range 指定数值数据区域，categories_range 指定分类标签"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "chart_type": {
                        "type": "string",
                        "enum": ["bar", "line", "pie", "scatter", "area"],
                        "description": "图表类型",
                    },
                    "data_range": {
                        "type": "string",
                        "description": "数值数据区域（如 'B1:B20' 或多列 'B1:D20'），第一行作为系列名",
                    },
                    "categories_range": {
                        "type": "string",
                        "description": "分类轴标签区域（如 'A2:A20'），可选",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "数据源工作表名称，默认活动工作表",
                    },
                    "target_cell": {
                        "type": "string",
                        "description": "图表放置位置（如 'E1'），默认 A1",
                        "default": "A1",
                    },
                    "target_sheet": {
                        "type": "string",
                        "description": "图表放置的目标工作表名，默认与数据源相同。不存在时自动创建",
                    },
                    "title": {
                        "type": "string",
                        "description": "图表标题",
                    },
                    "x_title": {
                        "type": "string",
                        "description": "X 轴标题",
                    },
                    "y_title": {
                        "type": "string",
                        "description": "Y 轴标题",
                    },
                    "style": {
                        "type": "integer",
                        "description": "Excel 图表样式编号（1-48）",
                    },
                    "width": {
                        "type": "number",
                        "description": "图表宽度（厘米），默认 15",
                        "default": 15.0,
                    },
                    "height": {
                        "type": "number",
                        "description": "图表高度（厘米），默认 10",
                        "default": 10.0,
                    },
                    "from_rows": {
                        "type": "boolean",
                        "description": "是否按行读取数据系列（默认按列）",
                        "default": False,
                    },
                },
                "required": ["file_path", "chart_type", "data_range"],
                "additionalProperties": False,
            },
            func=create_excel_chart,
        ),
    ]
