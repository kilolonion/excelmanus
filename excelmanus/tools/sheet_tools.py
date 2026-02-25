"""工作表管理工具：提供工作表级别的查看、创建、复制、重命名、删除和跨表数据传输能力。"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools._guard_ctx import get_guard as _get_ctx_guard
from excelmanus.tools._helpers import resolve_sheet_name
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.sheet")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "sheet"
SKILL_DESCRIPTION = "工作表管理工具集：列表、创建、复制、重命名、删除和跨表数据传输"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None
_MAX_LIST_PAGE_SIZE = 500


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard（优先 per-session contextvar）。"""
    ctx_guard = _get_ctx_guard()
    if ctx_guard is not None:
        return ctx_guard
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


def _validate_pagination(offset: int, limit: int, *, max_limit: int = _MAX_LIST_PAGE_SIZE) -> str | None:
    """校验分页参数，返回错误信息或 None。"""
    if offset < 0:
        return "offset 必须大于或等于 0"
    if limit <= 0:
        return "limit 必须为正整数"
    if limit > max_limit:
        return f"limit 不能超过 {max_limit}"
    return None


# ── 工具函数 ──────────────────────────────────────────────


# list_sheets 可用的 include 维度
_LIST_SHEETS_DIMENSIONS = (
    "columns",
    "dtypes",
    "freeze_panes",
    "preview",
    "charts",
    "images",
    "conditional_formatting",
    "column_widths",
)


def list_sheets(
    file_path: str,
    offset: int = 0,
    limit: int = 100,
    include: list[str] | None = None,
    max_preview_rows: int = 5,
) -> str:
    """列出 Excel 文件中所有工作表的名称和基本信息，可按需附加额外维度。

    Args:
        file_path: Excel 文件路径。
        offset: 分页起始偏移（从 0 开始），默认 0。
        limit: 分页大小，默认 100，最大 500。
        include: 按需请求的额外维度列表。可选值：
            columns — 列名列表
            dtypes — 列数据类型（需用 pandas 读取）
            freeze_panes — 冻结窗格位置
            preview — 前 N 行数据预览
            charts — 嵌入图表元信息
            images — 嵌入图片元信息
            conditional_formatting — 条件格式规则
            column_widths — 非默认列宽
        max_preview_rows: preview 维度的预览行数，默认 5。

    Returns:
        JSON 格式的工作表列表，包含名称、行列数和是否为活动表。
    """
    paging_error = _validate_pagination(offset, limit)
    if paging_error is not None:
        return json.dumps({"error": paging_error}, ensure_ascii=False)

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    include_set: set[str] = set(include) if include else set()
    invalid_dims = include_set - set(_LIST_SHEETS_DIMENSIONS)
    include_set -= invalid_dims

    # 是否需要非 read_only 模式（charts/images/freeze_panes 等需要）
    needs_full = bool(include_set - {"columns", "dtypes", "preview"})

    wb = load_workbook(safe_path, read_only=not needs_full, data_only=True)
    try:
        active_name = wb.active.title if wb.active else None
        sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            info: dict[str, Any] = {
                "name": ws.title,
                "rows": ws.max_row or 0,
                "columns": ws.max_column or 0,
                "is_active": ws.title == active_name,
            }

            # columns 维度：读取第一行作为列名
            if "columns" in include_set:
                header_row = list(ws.iter_rows(
                    min_row=1, max_row=1, values_only=True,
                ))
                if header_row and header_row[0]:
                    info["column_names"] = [
                        str(c) if c is not None else None
                        for c in header_row[0]
                        if c is not None
                    ]

            # preview 维度：前 N 行数据
            if "preview" in include_set:
                preview_rows: list[list[Any]] = []
                for row in ws.iter_rows(
                    min_row=2, max_row=1 + max_preview_rows, values_only=True,
                ):
                    preview_rows.append([
                        str(c) if c is not None else None for c in row
                    ])
                info["preview"] = preview_rows

            # 以下维度通过 data_tools 的采集函数实现
            if needs_full and include_set:
                from excelmanus.tools.data_tools import (
                    _collect_charts,
                    _collect_column_widths,
                    _collect_conditional_formatting,
                    _collect_freeze_panes,
                    _collect_images,
                )

                if "freeze_panes" in include_set:
                    info["freeze_panes"] = _collect_freeze_panes(ws)
                if "charts" in include_set:
                    info["charts"] = _collect_charts(ws)
                if "images" in include_set:
                    info["images"] = _collect_images(ws)
                if "conditional_formatting" in include_set:
                    info["conditional_formatting"] = _collect_conditional_formatting(ws)
                if "column_widths" in include_set:
                    info["column_widths"] = _collect_column_widths(ws)

            sheets.append(info)
    finally:
        wb.close()

    total = len(sheets)
    end = offset + limit
    paged_sheets = sheets[offset:end]
    has_more = end < total

    result: dict[str, Any] = {
        "file": safe_path.name,
        "sheet_count": total,
        "offset": offset,
        "limit": limit,
        "returned": len(paged_sheets),
        "has_more": has_more,
        "sheets": paged_sheets,
    }
    if invalid_dims:
        result["include_warning"] = f"未知的 include 维度已忽略: {sorted(invalid_dims)}"

    return json.dumps(result, ensure_ascii=False, indent=2, default=str)


def create_sheet(
    file_path: str,
    sheet_name: str,
    position: int | None = None,
) -> str:
    """在已有 Excel 文件中新建空白工作表。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 新工作表名称。
        position: 插入位置索引（0 表示最前面），默认追加到最后。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)

    if sheet_name in wb.sheetnames:
        wb.close()
        return json.dumps(
            {"status": "error", "message": f"工作表 '{sheet_name}' 已存在"},
            ensure_ascii=False,
        )

    kwargs: dict[str, Any] = {"title": sheet_name}
    if position is not None:
        kwargs["index"] = position

    wb.create_sheet(**kwargs)
    wb.save(safe_path)
    wb.close()

    logger.info("已在 %s 中创建工作表 '%s'", safe_path.name, sheet_name)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "created_sheet": sheet_name,
            "total_sheets": len(wb.sheetnames),
            "all_sheets": wb.sheetnames,
        },
        ensure_ascii=False,
        indent=2,
    )


def copy_sheet(
    file_path: str,
    source_sheet: str,
    new_name: str | None = None,
) -> str:
    """复制工作表（同文件内）。

    Args:
        file_path: Excel 文件路径。
        source_sheet: 要复制的源工作表名称。
        new_name: 新工作表名称，默认自动生成（如 "Sheet1 Copy"）。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)

    if source_sheet not in wb.sheetnames:
        resolved = resolve_sheet_name(source_sheet, wb.sheetnames)
        if resolved is None:
            wb.close()
            return json.dumps(
                {
                    "status": "error",
                    "message": f"源工作表 '{source_sheet}' 不存在，可用: {wb.sheetnames}",
                },
                ensure_ascii=False,
            )
        source_sheet = resolved

    if new_name and new_name in wb.sheetnames:
        wb.close()
        return json.dumps(
            {"status": "error", "message": f"工作表 '{new_name}' 已存在"},
            ensure_ascii=False,
        )

    source_ws = wb[source_sheet]
    copied_ws = wb.copy_worksheet(source_ws)

    if new_name:
        copied_ws.title = new_name

    wb.save(safe_path)
    final_name = copied_ws.title
    wb.close()

    logger.info("已在 %s 中复制工作表 '%s' -> '%s'", safe_path.name, source_sheet, final_name)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "source_sheet": source_sheet,
            "new_sheet": final_name,
            "all_sheets": wb.sheetnames,
        },
        ensure_ascii=False,
        indent=2,
    )


def rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str,
) -> str:
    """重命名工作表。

    Args:
        file_path: Excel 文件路径。
        old_name: 当前工作表名称。
        new_name: 新名称。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)

    if old_name not in wb.sheetnames:
        resolved = resolve_sheet_name(old_name, wb.sheetnames)
        if resolved is None:
            wb.close()
            return json.dumps(
                {
                    "status": "error",
                    "message": f"工作表 '{old_name}' 不存在，可用: {wb.sheetnames}",
                },
                ensure_ascii=False,
            )
        old_name = resolved

    if new_name in wb.sheetnames:
        wb.close()
        return json.dumps(
            {"status": "error", "message": f"工作表 '{new_name}' 已存在"},
            ensure_ascii=False,
        )

    wb[old_name].title = new_name
    wb.save(safe_path)
    wb.close()

    logger.info("已在 %s 中重命名工作表 '%s' -> '%s'", safe_path.name, old_name, new_name)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "old_name": old_name,
            "new_name": new_name,
            "all_sheets": wb.sheetnames,
        },
        ensure_ascii=False,
        indent=2,
    )


def delete_sheet(
    file_path: str,
    sheet_name: str,
    confirm: bool = False,
) -> str:
    """删除工作表（需二次确认）。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 要删除的工作表名称。
        confirm: 是否确认删除，必须为 True 才执行。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)

    if sheet_name not in wb.sheetnames:
        resolved = resolve_sheet_name(sheet_name, wb.sheetnames)
        if resolved is None:
            wb.close()
            return json.dumps(
                {
                    "status": "error",
                    "message": f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}",
                },
                ensure_ascii=False,
            )
        sheet_name = resolved

    if len(wb.sheetnames) <= 1:
        wb.close()
        return json.dumps(
            {"status": "error", "message": "不能删除唯一的工作表"},
            ensure_ascii=False,
        )

    if not confirm:
        # 返回待删除工作表信息，供 LLM 二次确认
        ws = wb[sheet_name]
        wb.close()
        return json.dumps(
            {
                "status": "pending_confirmation",
                "message": "请将 confirm 设为 true 以确认删除",
                "sheet_name": sheet_name,
                "rows": ws.max_row or 0,
                "columns": ws.max_column or 0,
            },
            ensure_ascii=False,
            indent=2,
        )

    del wb[sheet_name]
    wb.save(safe_path)
    remaining = wb.sheetnames
    wb.close()

    logger.info("已在 %s 中删除工作表 '%s'", safe_path.name, sheet_name)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "deleted_sheet": sheet_name,
            "remaining_sheets": remaining,
        },
        ensure_ascii=False,
        indent=2,
    )


def copy_range_between_sheets(
    source_file: str,
    source_sheet: str,
    source_range: str,
    target_file: str | None = None,
    target_sheet: str = "Sheet1",
    target_start: str = "A1",
) -> str:
    """从源工作表复制指定范围的数据到目标工作表，支持跨文件。

    Args:
        source_file: 源 Excel 文件路径。
        source_sheet: 源工作表名称。
        source_range: 源数据范围，如 "A1:D10"。
        target_file: 目标 Excel 文件路径，默认与源文件相同。
        target_sheet: 目标工作表名称，默认 Sheet1。
        target_start: 目标起始单元格，默认 A1。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_source = guard.resolve_and_validate(source_file)
    same_file = target_file is None or str(guard.resolve_and_validate(target_file)) == str(safe_source)
    safe_target = safe_source if same_file else guard.resolve_and_validate(target_file)  # type: ignore[arg-type]

    # 读取源数据
    src_wb = load_workbook(safe_source, data_only=True)
    resolved_source = resolve_sheet_name(source_sheet, src_wb.sheetnames)
    if resolved_source is None:
        src_wb.close()
        return json.dumps(
            {
                "status": "error",
                "message": f"源工作表 '{source_sheet}' 不存在，可用: {src_wb.sheetnames}",
            },
            ensure_ascii=False,
        )

    src_ws = src_wb[resolved_source]

    # 解析源范围，提取值
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(source_range)

    data: list[list[Any]] = []
    for row in src_ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
    ):
        data.append(list(row))
    src_wb.close()

    if not data:
        return json.dumps(
            {"status": "error", "message": f"源范围 '{source_range}' 无数据"},
            ensure_ascii=False,
        )

    # 写入目标
    if safe_target.exists():
        tgt_wb = load_workbook(safe_target)
    else:
        tgt_wb = Workbook()

    if target_sheet not in tgt_wb.sheetnames:
        resolved_target = resolve_sheet_name(target_sheet, tgt_wb.sheetnames)
        if resolved_target is not None:
            target_sheet = resolved_target
        else:
            tgt_wb.create_sheet(title=target_sheet)
            # 如果是新建的 Workbook 且自带默认 Sheet，可移除
            if "Sheet" in tgt_wb.sheetnames and target_sheet != "Sheet" and len(tgt_wb.sheetnames) > 1:
                del tgt_wb["Sheet"]

    tgt_ws = tgt_wb[target_sheet]

    # 解析目标起始位置
    from openpyxl.utils.cell import coordinate_to_tuple
    start_row, start_col = coordinate_to_tuple(target_start.upper())

    # 写入数据
    cells_written = 0
    for r_idx, row_data in enumerate(data):
        for c_idx, value in enumerate(row_data):
            tgt_ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
            cells_written += 1

    tgt_wb.save(safe_target)
    tgt_wb.close()

    logger.info(
        "已从 %s[%s]%s 复制 %d 个单元格到 %s[%s]%s",
        safe_source.name, source_sheet, source_range,
        cells_written,
        safe_target.name, target_sheet, target_start,
    )

    return json.dumps(
        {
            "status": "success",
            "source": {
                "file": safe_source.name,
                "sheet": source_sheet,
                "range": source_range,
            },
            "target": {
                "file": safe_target.name,
                "sheet": target_sheet,
                "start_cell": target_start,
            },
            "rows_copied": len(data),
            "cells_written": cells_written,
        },
        ensure_ascii=False,
        indent=2,
    )


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回工作表管理工具的所有工具定义。"""
    return [
        ToolDef(
            name="list_sheets",
            description=(
                "列出 Excel 文件中所有工作表的名称、行列数和是否为活动表，"
                "通过 include 按需附加 columns/preview/charts 等维度"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "offset": {
                        "type": "integer",
                        "description": "分页起始偏移（从 0 开始），默认 0",
                        "default": 0,
                    },
                    "limit": {
                        "type": "integer",
                        "description": "分页大小，默认 100，最大 500",
                        "default": 100,
                    },
                    "include": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": [
                                "columns",
                                "freeze_panes",
                                "preview",
                                "charts",
                                "images",
                                "conditional_formatting",
                                "column_widths",
                            ],
                        },
                        "description": "按需附加的额外维度",
                    },
                    "max_preview_rows": {
                        "type": "integer",
                        "description": "preview 维度的预览行数，默认 5",
                        "default": 5,
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=list_sheets,
            max_result_chars=0,
            write_effect="none",
        ),
        # Batch 3 精简：create_sheet/copy_sheet/rename_sheet/delete_sheet/copy_range_between_sheets 已删除，由 run_code 替代
    ]


