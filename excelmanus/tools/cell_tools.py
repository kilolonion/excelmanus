"""单元格级操作工具：提供原地写入值/公式、插入/删除行列的能力。

与 data_tools 的 DataFrame 范式不同，本模块直接使用 openpyxl 操作单元格，
适用于需要保留工作表其他区域数据不变的场景。
"""

from __future__ import annotations

import json
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.tools._helpers import get_worksheet
from excelmanus.tools.registry import ToolDef

logger = get_logger("tools.cell")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "cell"
SKILL_DESCRIPTION = "单元格级操作工具集：原地写入值/公式、插入行列、删除行列"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard 单例。"""
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


# ── 内部辅助 ──────────────────────────────────────────────


def _parse_single_cell(ref: str) -> tuple[int, int]:
    """解析单个单元格引用（如 'A1'）为 (row, col) 元组（1-indexed）。"""
    return coordinate_to_tuple(ref.upper())


def _coerce_value(raw: Any) -> Any:
    """尝试将字符串值自动转换为合适的 Python 类型。

    - 以 '=' 开头的字符串保留为公式
    - 纯数字字符串转为 int/float
    - 其他保留为字符串
    """
    if not isinstance(raw, str):
        return raw
    stripped = raw.strip()
    if not stripped:
        return stripped
    # 公式原样保留
    if stripped.startswith("="):
        return stripped
    # 尝试数值转换
    try:
        if "." in stripped or "e" in stripped.lower():
            return float(stripped)
        return int(stripped)
    except (ValueError, OverflowError):
        pass
    return raw

def _resolve_merged_cell(ws: Any, row: int, col: int) -> tuple[int, int, bool]:
    """检测 (row, col) 是否在合并区域内，若是则返回主单元格坐标。

    Returns:
        (actual_row, actual_col, redirected)
        redirected 为 True 表示目标是合并区域的从属单元格，已重定向到主单元格。
    """
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            if row == merged_range.min_row and col == merged_range.min_col:
                return row, col, False  # 就是主单元格，无需重定向
            return merged_range.min_row, merged_range.min_col, True
    return row, col, False



def _capture_range_snapshot(
    ws: Any,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> list[dict[str, Any]]:
    """捕获指定范围内所有非空单元格的值快照。

    Returns:
        列表，每项为 {"cell": "A1", "value": ...}
    """
    snapshot: list[dict[str, Any]] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell_obj = ws.cell(row=r, column=c)
            ref = f"{get_column_letter(c)}{r}"
            val = cell_obj.value
            snapshot.append({"cell": ref, "value": val})
    return snapshot


def _compute_cell_diff(
    before: list[dict[str, Any]],
    after: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """对比写入前后快照，返回变化的单元格列表。

    Returns:
        列表，每项为 {"cell": "A1", "old": ..., "new": ...}
    """
    before_map = {item["cell"]: item["value"] for item in before}
    after_map = {item["cell"]: item["value"] for item in after}
    all_cells = sorted(set(before_map) | set(after_map))
    changes: list[dict[str, Any]] = []
    for cell_ref in all_cells:
        old_val = before_map.get(cell_ref)
        new_val = after_map.get(cell_ref)
        if old_val != new_val:
            changes.append({
                "cell": cell_ref,
                "old": _serialize_cell_value(old_val),
                "new": _serialize_cell_value(new_val),
            })
    return changes


def _serialize_cell_value(val: Any) -> Any:
    """将单元格值序列化为 JSON 安全类型。"""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


# ── 工具函数 ──────────────────────────────────────────────


def write_cells(
    file_path: str,
    sheet_name: str | None = None,
    cell: str | None = None,
    value: Any = None,
    cell_range: str | None = None,
    values: list[list[Any]] | None = None,
    return_preview: bool = False,
) -> str:
    """向指定单元格或范围写入值/公式，不影响工作表其他区域的数据。

    注意：仅适用于少量数据（<=3行）的精确写入。超过3行的批量写入、条件更新、
    复杂数据流转必须使用 run_code 工具。

    两种模式（互斥）：
    - **单元格模式**：传 cell + value，写入单个单元格。
    - **范围模式**：传 cell_range + values，批量写入二维数据。
      cell_range 可以只指定起始单元格（如 "A1"），values 的行列数决定实际写入范围。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称，默认活动工作表。
        cell: 目标单元格引用（如 "A1"），单元格模式。
        value: 要写入的值（数字、字符串、公式），单元格模式。
        cell_range: 目标范围起始位置或完整范围（如 "A1" 或 "A1:C3"），范围模式。
        values: 二维数组，范围模式。每个内层列表代表一行。

    Returns:
        JSON 格式的操作结果。
    """
    # 参数校验
    single_mode = cell is not None
    range_mode = cell_range is not None or values is not None

    if single_mode and range_mode:
        return json.dumps(
            {"error": "cell/value 与 cell_range/values 互斥，请只使用其中一种模式"},
            ensure_ascii=False,
        )
    if not single_mode and not range_mode:
        return json.dumps(
            {"error": "必须指定 cell+value（单元格模式）或 cell_range+values（范围模式）"},
            ensure_ascii=False,
        )

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    # 计算受影响范围，用于写入前快照
    _snap_min_row = _snap_min_col = _snap_max_row = _snap_max_col = 0
    try:
        if single_mode:
            row, col = _parse_single_cell(cell)  # type: ignore[arg-type]
            _snap_min_row, _snap_min_col = row, col
            _snap_max_row, _snap_max_col = row, col
        else:
            if values is None or not values:
                return json.dumps(
                    {"error": "范围模式下 values 不能为空"},
                    ensure_ascii=False,
                )
            start_ref = cell_range or "A1"
            if ":" in start_ref:
                start_ref = start_ref.split(":")[0]
            _snap_min_row, _snap_min_col = _parse_single_cell(start_ref)
            _snap_max_row = _snap_min_row + len(values) - 1
            _max_cols = max(len(r) if isinstance(r, list) else 1 for r in values)
            _snap_max_col = _snap_min_col + _max_cols - 1

        # 写入前快照
        before_snapshot = _capture_range_snapshot(
            ws, _snap_min_row, _snap_min_col, _snap_max_row, _snap_max_col,
        )
    except Exception:
        before_snapshot = []

    try:
        if single_mode:
            # 单元格模式
            row, col = _parse_single_cell(cell)  # type: ignore[arg-type]
            actual_row, actual_col, redirected = _resolve_merged_cell(ws, row, col)
            ws.cell(row=actual_row, column=actual_col, value=_coerce_value(value))
            wb.save(safe_path)
            actual_ref = f"{get_column_letter(actual_col)}{actual_row}"
            result: dict[str, Any] = {
                "status": "success",
                "file": safe_path.name,
                "cell": actual_ref if redirected else cell,
                "value_written": value,
                "cells_written": 1,
            }
            if redirected:
                result["note"] = (
                    f"{cell} 是合并区域的从属单元格，已自动写入主单元格 {actual_ref}"
                )
        else:
            # 范围模式
            start_ref = cell_range or "A1"
            if ":" in start_ref:
                start_ref = start_ref.split(":")[0]
            start_row, start_col = _parse_single_cell(start_ref)

            cells_written = 0
            skipped_merged = 0
            for r_idx, row_data in enumerate(values):
                if not isinstance(row_data, list):
                    row_data = [row_data]
                for c_idx, val in enumerate(row_data):
                    target_row = start_row + r_idx
                    target_col = start_col + c_idx
                    actual_row, actual_col, redirected = _resolve_merged_cell(
                        ws, target_row, target_col
                    )
                    if redirected:
                        # 从属单元格：跳过，值由主单元格决定
                        skipped_merged += 1
                        continue
                    ws.cell(
                        row=actual_row,
                        column=actual_col,
                        value=_coerce_value(val),
                    )
                    cells_written += 1

            wb.save(safe_path)
            end_row = start_row + len(values) - 1
            max_cols = max(len(r) if isinstance(r, list) else 1 for r in values)
            end_col = start_col + max_cols - 1
            actual_range = (
                f"{get_column_letter(start_col)}{start_row}"
                f":{get_column_letter(end_col)}{end_row}"
            )
            result: dict[str, Any] = {
                "status": "success",
                "file": safe_path.name,
                "range": actual_range,
                "rows_written": len(values),
                "cells_written": cells_written,
            }
            if skipped_merged:
                result["skipped_merged_cells"] = skipped_merged
    finally:
        wb.close()

    # 写入后快照并计算 diff
    try:
        wb_after = load_workbook(safe_path)
        ws_after = get_worksheet(wb_after, sheet_name)
        after_snapshot = _capture_range_snapshot(
            ws_after, _snap_min_row, _snap_min_col, _snap_max_row, _snap_max_col,
        )
        wb_after.close()
        excel_diff = _compute_cell_diff(before_snapshot, after_snapshot)
        if excel_diff:
            affected_range = (
                f"{get_column_letter(_snap_min_col)}{_snap_min_row}"
                f":{get_column_letter(_snap_max_col)}{_snap_max_row}"
            )
            result["_excel_diff"] = {
                "file_path": file_path,
                "sheet": sheet_name or "",
                "affected_range": affected_range,
                "changes": excel_diff,
            }
    except Exception as exc:
        logger.debug("写入后 diff 捕获失败: %s", exc)

    # 检测是否写入了公式
    has_formulas = False
    if single_mode:
        if isinstance(value, str) and str(value).strip().startswith("="):
            has_formulas = True
    elif values:
        for row_data in values:
            if isinstance(row_data, list):
                for v in row_data:
                    if isinstance(v, str) and v.strip().startswith("="):
                        has_formulas = True
                        break
            if has_formulas:
                break

    # 写入后返回受影响区域的预览
    if return_preview:
        from openpyxl import load_workbook as _lw

        wb2 = _lw(safe_path, data_only=True)
        try:
            ws2 = get_worksheet(wb2, sheet_name)
            if single_mode:
                r, c = _parse_single_cell(cell)  # type: ignore[arg-type]
                val = ws2.cell(row=r, column=c).value
                result["preview_after"] = [[str(val) if val is not None else None]]
            else:
                preview_rows: list[list[Any]] = []
                for row_cells in ws2.iter_rows(
                    min_row=start_row,
                    max_row=end_row,
                    min_col=start_col,
                    max_col=end_col,
                    values_only=True,
                ):
                    preview_rows.append([
                        str(c) if c is not None else None for c in row_cells
                    ])
                result["preview_after"] = preview_rows
        finally:
            wb2.close()

    # 公式写入警告: openpyxl 写入的公式无缓存计算值，外部读取时会显示为空
    if has_formulas:
        result["formula_warning"] = (
            "写入的公式无缓存计算值，仅在 Excel 打开时才会计算。"
            "若需确保值立即可读，建议改用计算后的具体值写入。"
        )

    logger.info("write_cells: %s", result)
    return json.dumps(result, ensure_ascii=False, indent=2, default=str)


def insert_rows(
    file_path: str,
    row: int,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """在指定行号前插入空行，已有数据自动下移。

    Args:
        file_path: Excel 文件路径。
        row: 插入位置行号（1-indexed），新行将插入到此行之前。
        count: 插入行数，默认 1。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    if row < 1:
        return json.dumps({"error": "row 必须 >= 1"}, ensure_ascii=False)
    if count < 1:
        return json.dumps({"error": "count 必须 >= 1"}, ensure_ascii=False)
    if count > 10000:
        return json.dumps({"error": "单次插入行数不能超过 10000"}, ensure_ascii=False)

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    rows_before = ws.max_row or 0
    ws.insert_rows(row, amount=count)
    wb.save(safe_path)
    wb.close()

    logger.info("insert_rows: %s 行 %d 处插入 %d 行", safe_path.name, row, count)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "inserted_at_row": row,
            "count": count,
            "rows_before": rows_before,
            "rows_after": rows_before + count,
        },
        ensure_ascii=False,
        indent=2,
    )


def insert_columns(
    file_path: str,
    column: int | str,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """在指定列前插入空列，已有数据自动右移。

    Args:
        file_path: Excel 文件路径。
        column: 插入位置，可以是列号（1-indexed 整数）或列字母（如 "C"）。
        count: 插入列数，默认 1。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    # 解析列号
    if isinstance(column, str):
        column_str = column.strip().upper()
        if column_str.isdigit():
            col_idx = int(column_str)
        else:
            # 字母转数字：A=1, B=2, ..., Z=26, AA=27, ...
            col_idx = 0
            for ch in column_str:
                if not ch.isalpha():
                    return json.dumps(
                        {"error": f"无效的列标识: '{column}'，应为列字母（如 'C'）或数字"},
                        ensure_ascii=False,
                    )
                col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    else:
        col_idx = int(column)

    if col_idx < 1:
        return json.dumps({"error": "column 必须 >= 1"}, ensure_ascii=False)
    if count < 1:
        return json.dumps({"error": "count 必须 >= 1"}, ensure_ascii=False)
    if count > 1000:
        return json.dumps({"error": "单次插入列数不能超过 1000"}, ensure_ascii=False)

    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = get_worksheet(wb, sheet_name)

    cols_before = ws.max_column or 0
    ws.insert_cols(col_idx, amount=count)
    wb.save(safe_path)
    wb.close()

    col_letter = get_column_letter(col_idx)
    logger.info("insert_columns: %s 列 %s(%d) 处插入 %d 列", safe_path.name, col_letter, col_idx, count)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "inserted_at_column": col_letter,
            "inserted_at_column_index": col_idx,
            "count": count,
            "columns_before": cols_before,
            "columns_after": cols_before + count,
        },
        ensure_ascii=False,
        indent=2,
    )


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回单元格级操作工具的所有工具定义。

    Batch 1 精简：write_cells/insert_rows/insert_columns 已删除，由 run_code 替代。
    函数实现保留以支持内部引用和未来可能的恢复。
    """
    return []
