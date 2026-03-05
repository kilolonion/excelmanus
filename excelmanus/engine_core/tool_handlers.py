"""ToolHandler 策略实现 — 从 _dispatch_tool_execution if-elif 提取的独立处理器。

每个 Handler 负责一类工具的执行逻辑，通过 can_handle / handle 接口
与 ToolDispatcher 的策略表对接。
"""

from __future__ import annotations

import asyncio
import json
import time
from collections.abc import Sequence
from typing import TYPE_CHECKING, Any

from excelmanus.logger import get_logger, log_tool_call

if TYPE_CHECKING:
    from pathlib import Path

    from excelmanus.engine import AgentEngine
    from excelmanus.engine_core.tool_dispatcher import ToolDispatcher, _ToolExecOutcome
    from excelmanus.pipeline.models import PipelineConfig
    from excelmanus.replica_spec import ReplicaSpec

logger = get_logger("tool_handlers")


# ---------------------------------------------------------------------------
# 基类
# ---------------------------------------------------------------------------

class BaseToolHandler:
    """所有 handler 的基类，持有 engine 和 dispatcher 引用（双轨兼容）。"""

    def __init__(self, engine: AgentEngine, dispatcher: ToolDispatcher) -> None:
        self._engine = engine
        self._dispatcher = dispatcher

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        raise NotImplementedError

    async def handle(
        self,
        tool_name: str,
        tool_call_id: str,
        arguments: dict[str, Any],
        *,
        tool_scope: Sequence[str] | None = None,
        on_event: Any = None,
        iteration: int = 0,
        route_result: Any = None,
    ) -> _ToolExecOutcome:
        raise NotImplementedError


# ---------------------------------------------------------------------------
# 技能激活处理器（SkillActivationHandler）
# ---------------------------------------------------------------------------

class SkillActivationHandler(BaseToolHandler):
    """处理 activate_skill 工具调用。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "activate_skill"

    async def handle(self, tool_name, tool_call_id, arguments, **kwargs):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        selected_name = arguments.get("skill_name")
        if not isinstance(selected_name, str) or not selected_name.strip():
            result_str = "工具参数错误: skill_name 必须为非空字符串。"
            log_tool_call(logger, tool_name, arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        result_str = await e.handle_activate_skill(selected_name.strip())
        success = result_str.startswith("OK")
        error = None if success else result_str
        log_tool_call(logger, tool_name, arguments, result=result_str if success else None, error=error if not success else None)
        return _ToolExecOutcome(result_str=result_str, success=success, error=error)


# ---------------------------------------------------------------------------
# 技能管理处理器（SkillManagementHandler）
# ---------------------------------------------------------------------------

class SkillManagementHandler(BaseToolHandler):
    """处理 manage_skills 工具调用：搜索/安装/卸载/查看技能。"""

    _VERSION_CACHE_TTL = 300  # 5 分钟

    def __init__(self, engine: "AgentEngine", dispatcher: "ToolDispatcher") -> None:
        super().__init__(engine, dispatcher)
        self._version_cache: dict[str, tuple[str, float]] = {}  # slug → (version, timestamp)

    def _cache_version(self, slug: str, version: str | None) -> None:
        """缓存 slug→version 映射。"""
        if slug and version:
            self._version_cache[slug] = (version, time.monotonic())

    def _get_cached_version(self, slug: str) -> str | None:
        """获取缓存的版本号，过期返回 None。"""
        entry = self._version_cache.get(slug)
        if entry is None:
            return None
        version, ts = entry
        if time.monotonic() - ts > self._VERSION_CACHE_TTL:
            self._version_cache.pop(slug, None)
            return None
        return version

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "manage_skills"

    async def handle(self, tool_name, tool_call_id, arguments, **kwargs):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        action = str(arguments.get("action", "")).strip()
        dispatch = {
            "search": self._handle_search,
            "detail": self._handle_detail,
            "install": self._handle_install,
            "list": self._handle_list,
            "uninstall": self._handle_uninstall,
            "update": self._handle_update,
        }
        handler_fn = dispatch.get(action)
        if handler_fn is None:
            result_str = f"不支持的操作: {action}（支持 search/install/detail/list/uninstall/update）"
            log_tool_call(logger, tool_name, arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        return await handler_fn(arguments)

    # ── search ────────────────────────────────────────────

    async def _handle_search(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        query = str(arguments.get("query", "")).strip()
        if not query:
            result_str = "参数错误: search 操作需要提供 query 参数。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        try:
            results = await manager.clawhub_search(query, limit=10)
        except Exception as exc:
            result_str = f"ClawHub 搜索失败: {exc}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        if not results:
            result_str = f"未找到与 '{query}' 相关的技能。"
            log_tool_call(logger, "manage_skills", arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True)

        lines = [f"找到 {len(results)} 个相关技能：\n"]
        for r in results:
            slug = r.get("slug", "")
            display_name = r.get("display_name", slug)
            summary = r.get("summary", "")
            version = r.get("version", "")
            line = f"  - {display_name} (slug={slug}, v{version})"
            if summary:
                line += f" — {summary}"
            lines.append(line)
        # P2: 缓存搜索结果中的版本号，供后续 install 跳过版本解析
        for r in results:
            self._cache_version(r.get("slug", ""), r.get("version"))

        lines.append("\n可使用 action=install, slug=<slug> 安装感兴趣的技能。")
        result_str = "\n".join(lines)
        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── detail ────────────────────────────────────────────

    async def _handle_detail(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        slug = str(arguments.get("slug", "")).strip()
        if not slug:
            result_str = "参数错误: detail 操作需要提供 slug 参数。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        try:
            detail = await manager.clawhub_skill_detail(slug)
        except Exception as exc:
            result_str = f"获取技能详情失败: {exc}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        parts = [
            f"技能: {detail.get('display_name', slug)}",
            f"标识: {detail.get('slug', slug)}",
            f"版本: {detail.get('latest_version', '未知')}",
        ]
        summary = detail.get("summary", "")
        if summary:
            parts.append(f"简介: {summary}")
        tags = detail.get("tags")
        if tags:
            parts.append(f"标签: {', '.join(tags)}")
        owner = detail.get("owner_display_name") or detail.get("owner_handle")
        if owner:
            parts.append(f"作者: {owner}")
        changelog = detail.get("latest_changelog", "")
        if changelog:
            parts.append(f"更新日志: {changelog}")
        stats = detail.get("stats")
        if isinstance(stats, dict):
            dl = stats.get("downloads")
            if dl is not None:
                parts.append(f"下载量: {dl}")

        result_str = "\n".join(parts)
        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── install ───────────────────────────────────────────

    async def _handle_install(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        slug = str(arguments.get("slug", "")).strip()
        if not slug:
            result_str = "参数错误: install 操作需要提供 slug 参数（ClawHub 标识符或 GitHub URL）。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 安全门控
        e = self._engine
        if getattr(getattr(e, "_config", None), "external_safe_mode", True):
            result_str = "安全模式已开启，禁止安装技能。请关闭 external_safe_mode 后重试。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        overwrite = bool(arguments.get("overwrite", False))

        # 自动检测来源：GitHub URL vs ClawHub slug
        source = "github_url" if slug.startswith("http") else "clawhub"

        # P2: 从缓存获取版本号，跳过版本解析
        cached_version = self._get_cached_version(slug) if source == "clawhub" else None

        try:
            result = await manager.import_skillpack_async(
                source=source, value=slug, actor="agent", overwrite=overwrite,
                version=cached_version,
            )
        except Exception as exc:
            exc_str = str(exc)
            # 对冲突错误提供更友好的提示
            if "已存在" in exc_str or "conflict" in exc_str.lower():
                result_str = f"技能已存在: {exc_str}\n如需覆盖安装，请传入 overwrite=true。"
            else:
                result_str = f"安装失败: {exc_str}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 安装成功 → 失效工具缓存，使新技能出现在 activate_skill enum 中
        e._tools_cache = None

        name = result.get("name", slug)
        desc = result.get("description", "")
        version = result.get("version", "")
        parts = [f"OK 技能 '{name}' 安装成功。"]
        if version:
            parts[0] = f"OK 技能 '{name}' (v{version}) 安装成功。"
        if desc:
            parts.append(f"描述: {desc}")
        parts.append("现在可以通过 activate_skill 激活使用此技能。")
        result_str = "\n".join(parts)
        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── list ──────────────────────────────────────────────

    async def _handle_list(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        skills = manager.list_skillpacks()
        if not skills:
            result_str = "当前没有已安装的技能。"
            log_tool_call(logger, "manage_skills", arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True)

        loaded_names: set[str] = set()
        try:
            loaded = self._engine._skill_resolver.get_loaded_skillpacks()
            if loaded:
                loaded_names = set(loaded.keys())
        except Exception:
            pass

        lines = [f"已安装 {len(skills)} 个技能：\n"]
        for s in skills:
            name = s.get("name", "")
            desc = s.get("description", "")
            version = s.get("version", "")
            line = f"  - {name}"
            if version:
                line += f" (v{version})"
            if name in loaded_names:
                line += " [已加载]"
            if desc:
                line += f" — {desc}"
            lines.append(line)
        result_str = "\n".join(lines)
        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── uninstall ─────────────────────────────────────────

    async def _handle_uninstall(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        slug = str(arguments.get("slug", "")).strip()
        if not slug:
            result_str = "参数错误: uninstall 操作需要提供 slug 参数（技能名称）。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 安全门控
        e = self._engine
        if getattr(getattr(e, "_config", None), "external_safe_mode", True):
            result_str = "安全模式已开启，禁止卸载技能。请关闭 external_safe_mode 后重试。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        try:
            result = manager.delete_skillpack(name=slug, actor="agent")
        except Exception as exc:
            result_str = f"卸载失败: {exc}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 卸载成功 → 清理运行时状态
        e._tools_cache = None
        # 移除已激活的该技能（防止过期引用）
        deleted_name = result.get("name", slug)
        if hasattr(e, "_active_skills"):
            e._active_skills = [
                s for s in e._active_skills if s.name != deleted_name
            ]
        if hasattr(e, "_loaded_skill_names"):
            e._loaded_skill_names.pop(deleted_name, None)
        # 清理 ClawHub lockfile 残留条目
        lockfile = getattr(manager, "_clawhub_lockfile", None)
        if lockfile is not None:
            try:
                lockfile.remove(deleted_name)
            except Exception:
                logger.debug("清理 ClawHub lockfile 条目失败: %s", deleted_name, exc_info=True)

        result_str = f"OK 技能 '{deleted_name}' 已卸载。"
        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── update ─────────────────────────────────────────────

    async def _handle_update(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        # 安全门控
        e = self._engine
        if getattr(getattr(e, "_config", None), "external_safe_mode", True):
            result_str = "安全模式已开启，禁止更新技能。请关闭 external_safe_mode 后重试。"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        manager = self._get_manager()
        if manager is None:
            return self._manager_unavailable(arguments)

        slug = str(arguments.get("slug", "")).strip()

        # 无 slug → 检查可用更新
        if not slug:
            try:
                updates = await manager.clawhub_check_updates()
            except Exception as exc:
                result_str = f"检查更新失败: {exc}"
                log_tool_call(logger, "manage_skills", arguments, error=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

            if not updates:
                result_str = "所有已安装的 ClawHub 技能均为最新版本。"
                log_tool_call(logger, "manage_skills", arguments, result=result_str)
                return _ToolExecOutcome(result_str=result_str, success=True)

            available = [u for u in updates if u.get("update_available")]
            if not available:
                result_str = "所有已安装的 ClawHub 技能均为最新版本。"
                log_tool_call(logger, "manage_skills", arguments, result=result_str)
                return _ToolExecOutcome(result_str=result_str, success=True)

            lines = [f"发现 {len(available)} 个可更新技能：\n"]
            for u in available:
                lines.append(
                    f"  - {u.get('slug')} : {u.get('installed_version', '?')} → {u.get('latest_version', '?')}"
                )
            lines.append("\n可使用 action=update, slug=<slug> 更新指定技能。")
            result_str = "\n".join(lines)
            log_tool_call(logger, "manage_skills", arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True)

        # 有 slug → 执行更新
        try:
            results = await manager.clawhub_update(slug=slug)
        except Exception as exc:
            result_str = f"更新失败: {exc}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        e._tools_cache = None

        if results and results[0].get("success"):
            version = results[0].get("version", "")
            result_str = f"OK 技能 '{slug}' 已更新到 v{version}。"
        else:
            error = results[0].get("error", "未知错误") if results else "未知错误"
            result_str = f"更新失败: {error}"
            log_tool_call(logger, "manage_skills", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        log_tool_call(logger, "manage_skills", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    # ── helpers ───────────────────────────────────────────

    def _get_manager(self):
        """获取 SkillpackManager，不可用时返回 None。"""
        try:
            return self._engine._require_skillpack_manager()
        except RuntimeError:
            return None

    def _manager_unavailable(self, arguments: dict[str, Any]):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        result_str = "技能管理器不可用。请检查技能系统是否已正确配置。"
        log_tool_call(logger, "manage_skills", arguments, error=result_str)
        return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)


# ---------------------------------------------------------------------------
# 委托处理器（DelegationHandler）
# ---------------------------------------------------------------------------

class DelegationHandler(BaseToolHandler):
    """处理 delegate / delegate_to_subagent（兼容） / list_subagents / parallel_delegate（兼容）。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name in ("delegate", "delegate_to_subagent", "list_subagents", "parallel_delegate")

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):

        if tool_name == "list_subagents":
            return self._handle_list(arguments)

        # delegate / delegate_to_subagent / parallel_delegate 统一处理
        # 判断是并行还是单任务模式
        tasks_value = arguments.get("tasks")
        if tool_name == "parallel_delegate" or (isinstance(tasks_value, list) and len(tasks_value) >= 2):
            return await self._handle_parallel(arguments, on_event=on_event)
        else:
            return await self._handle_delegate(tool_call_id, arguments, on_event=on_event, iteration=iteration)

    async def _handle_delegate(self, tool_call_id, arguments, *, on_event, iteration):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        task_value = arguments.get("task")
        task_brief = arguments.get("task_brief")
        if isinstance(task_brief, dict) and task_brief.get("title"):
            task_value = e.render_task_brief(task_brief)
        if not isinstance(task_value, str) or not task_value.strip():
            result_str = "工具参数错误: task、task_brief 或 tasks 必须提供其一。"
            log_tool_call(logger, "delegate", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        agent_name_value = arguments.get("agent_name")
        if agent_name_value is not None and not isinstance(agent_name_value, str):
            result_str = "工具参数错误: agent_name 必须为字符串。"
            log_tool_call(logger, "delegate", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        raw_file_paths = arguments.get("file_paths")
        if raw_file_paths is not None and not isinstance(raw_file_paths, list):
            result_str = "工具参数错误: file_paths 必须为字符串数组。"
            log_tool_call(logger, "delegate", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        delegate_outcome = await e.delegate_to_subagent(
            task=task_value.strip(),
            agent_name=agent_name_value.strip() if isinstance(agent_name_value, str) else None,
            file_paths=raw_file_paths,
            on_event=on_event,
        )
        result_str = delegate_outcome.reply
        success = delegate_outcome.success
        error = None if success else result_str

        # 写入传播
        sub_result = delegate_outcome.subagent_result
        if success and sub_result is not None and sub_result.structured_changes:
            e.record_workspace_write_action()

        # 子代理审批问题：阻塞等待用户决策
        if (
            not success
            and sub_result is not None
            and sub_result.pending_approval_id is not None
        ):
            import asyncio

            pending = e.approval.pending
            approval_id_value = sub_result.pending_approval_id
            high_risk_tool = (
                pending.tool_name
                if pending is not None and pending.approval_id == approval_id_value
                else "高风险工具"
            )
            question = e.enqueue_subagent_approval_question(
                approval_id=approval_id_value,
                tool_name=high_risk_tool,
                picked_agent=delegate_outcome.picked_agent or "subagent",
                task_text=delegate_outcome.task_text,
                normalized_paths=delegate_outcome.normalized_paths,
                tool_call_id=tool_call_id,
                on_event=on_event,
                iteration=iteration,
            )
            # 阻塞等待用户回答（支持 question_resolver / InteractionRegistry）
            try:
                payload = await e.await_question_answer(question)
            except (asyncio.TimeoutError, asyncio.CancelledError):
                e._question_flow.pop_current()
                e._interaction_registry.cleanup_done()
                result_str = "子代理审批问题超时/取消。"
                log_tool_call(logger, "delegate", arguments, result=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

            e._question_flow.pop_current()
            e._interaction_registry.cleanup_done()

            # 处理子代理审批回答（accept/fullaccess-retry/reject）
            if isinstance(payload, dict):
                result_str, success = await e.process_subagent_approval_inline(
                    payload=payload,
                    approval_id=approval_id_value,
                    picked_agent=delegate_outcome.picked_agent or "subagent",
                    task_text=delegate_outcome.task_text,
                    normalized_paths=delegate_outcome.normalized_paths,
                    on_event=on_event,
                )
                error = None if success else result_str
            else:
                result_str = str(payload)
                success = True
                error = None

        log_tool_call(logger, "delegate", arguments, result=result_str if success else None, error=error if not success else None)
        return _ToolExecOutcome(
            result_str=result_str, success=success, error=error,
        )

    def _handle_list(self, arguments):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        result_str = self._engine.handle_list_subagents()
        log_tool_call(logger, "list_subagents", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    async def _handle_parallel(self, arguments, *, on_event):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        raw_tasks = arguments.get("tasks")
        if not isinstance(raw_tasks, list) or len(raw_tasks) < 2:
            result_str = "工具参数错误: tasks 必须为包含至少 2 个子任务的数组。"
            log_tool_call(logger, "delegate", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        try:
            pd_outcome = await e.parallel_delegate_to_subagents(tasks=raw_tasks, on_event=on_event)
            result_str = pd_outcome.reply
            success = pd_outcome.success
            error = None if success else result_str

            for pd_sub_outcome in pd_outcome.outcomes:
                sub_result = pd_sub_outcome.subagent_result
                if pd_sub_outcome.success and sub_result is not None and sub_result.structured_changes:
                    e.record_workspace_write_action()
        except Exception as exc:
            result_str = f"parallel_delegate 执行异常: {exc}"
            success = False
            error = str(exc)

        log_tool_call(logger, "delegate", arguments, result=result_str if success else None, error=error if not success else None)
        return _ToolExecOutcome(result_str=result_str, success=success, error=error)


# ---------------------------------------------------------------------------
# 完成任务处理器（FinishTaskHandler）
# ---------------------------------------------------------------------------

class FinishTaskHandler(BaseToolHandler):
    """处理 finish_task 工具调用。"""

    # ── 分级验证强度 ──────────────────────────────────────
    # skip:     不触发 verifier
    # advisory: verifier 结果仅追加提示，不阻塞 finish
    # blocking: verifier fail+high 阻塞 finish，支持 fix-verify 循环
    _VERIFIER_LEVEL_BY_TAG: dict[str, str] = {
        "cross_sheet": "blocking",
        "large_data": "blocking",
        "formula": "blocking",
        "multi_file": "blocking",
        "simple": "advisory",
    }
    _DEFAULT_VERIFIER_LEVEL: str = "advisory"
    # blocking 模式最大重试次数（之后降为 advisory）
    _MAX_BLOCKING_ATTEMPTS: int = 2

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "finish_task"

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome, _render_finish_task_report

        e = self._engine
        report = arguments.get("report")
        summary = arguments.get("summary", "")
        rendered = _render_finish_task_report(report, summary)
        _has_write = getattr(e, "_has_write_tool_call", False)
        _hint = getattr(e, "_current_write_hint", "unknown")
        _guard_mode = getattr(getattr(e, "_config", None), "guard_mode", "off")
        finish_accepted = False

        if _has_write:
            result_str = f"✅ 任务完成\n\n{rendered}" if rendered else "✓ 任务完成。"
            success = True
            finish_accepted = True
        elif _guard_mode == "off" or _hint in ("read_only", "unknown"):
            _no_write_suffix = "（无写入）" if _hint != "unknown" else ""
            result_str = f"✅ 任务完成{_no_write_suffix}\n\n{rendered}" if rendered else f"✓ 任务完成{_no_write_suffix}。"
            success = True
            finish_accepted = True
        elif getattr(e, "_finish_task_warned", False):
            _no_write_suffix = "（无写入）" if _hint == "read_only" else ""
            result_str = f"✅ 任务完成{_no_write_suffix}\n\n{rendered}" if rendered else f"✓ 任务完成{_no_write_suffix}。"
            success = True
            finish_accepted = True
        else:
            result_str = (
                "⚠️ 未检测到写入类工具的成功调用。"
                "如果确实不需要写入，请再次调用 finish_task 并在 summary 中说明原因。"
                "否则请先执行写入操作。"
            )
            e._finish_task_warned = True
            success = True
            finish_accepted = False

        # ── 输出空值率前置检查 ────────────────────────────
        if finish_accepted and _has_write:
            import asyncio as _asyncio
            null_warning = await _asyncio.to_thread(self._check_output_null_rate, e)
            if null_warning:
                finish_accepted = False
                result_str = null_warning
                success = True

        # ── Verifier 接线 ──────────────────────────────────
        if finish_accepted:
            _report_dict = report if isinstance(report, dict) else None
            verifier_text = await self._run_verifier_if_needed(
                e, report=_report_dict, summary=summary, on_event=on_event,
            )
            if verifier_text is not None:
                if verifier_text.startswith("BLOCK:"):
                    finish_accepted = False
                    result_str = verifier_text[len("BLOCK:"):]
                    success = True
                else:
                    result_str += verifier_text

        _report_for_event = report if isinstance(report, dict) else None
        if not _report_for_event:
            top_files = arguments.get("affected_files")
            if top_files and isinstance(top_files, list):
                _report_for_event = {"affected_files": top_files}
        self._dispatcher._emit_files_changed_from_report(e, on_event, tool_call_id, _report_for_event, iteration)

        # 任务完成时，如果有 pending staged 文件，发射 staging_updated 提示
        if finish_accepted and on_event is not None:
            try:
                tx = getattr(e, "transaction", None)
                if tx is not None:
                    staged = tx.list_staged()
                    if staged:
                        from excelmanus.events import EventType, ToolCallEvent as _TCEvent
                        staging_files = []
                        for s in staged:
                            staging_files.append({
                                "original_path": tx.to_relative(s["original"]),
                                "backup_path": tx.to_relative(s["backup"]),
                            })
                        on_event(_TCEvent(
                            event_type=EventType.STAGING_UPDATED,
                            tool_call_id=tool_call_id,
                            staging_action="finish_hint",
                            staging_files=staging_files,
                            staging_pending_count=len(staged),
                            iteration=iteration,
                        ))
            except Exception:
                pass

        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=success, finish_accepted=finish_accepted)

    # ── 输出空值率检测 ──────────────────────────────────
    _NULL_RATE_THRESHOLD: float = 0.5  # >50% 空值即告警
    _NULL_CHECK_MAX_ROWS: int = 200    # 最多扫描行数（性能保护）

    @staticmethod
    def _check_output_null_rate(engine: Any) -> str | None:
        """扫描写入操作涉及的文件，检测输出是否存在异常高空值率。

        若写入的表格 >50% 单元格为空（且文件总行数 > 1），返回警告字符串阻断 finish_task。
        任何异常静默跳过（fail-open）。
        """
        import logging as _logging

        _state = getattr(engine, "_state", None)
        if _state is None:
            return None
        write_log: list[dict] = getattr(_state, "write_operations_log", [])
        if not write_log:
            return None

        # 收集写入涉及的 (file_path, sheet) 对
        guard = getattr(engine, "_file_access_guard", None)
        if guard is None:
            return None

        checked: set[tuple[str, str]] = set()
        alerts: list[str] = []

        for entry in write_log:
            file_path = entry.get("file_path", "")
            sheet = entry.get("sheet", "")
            if not file_path:
                continue
            key = (file_path, sheet)
            if key in checked:
                continue
            checked.add(key)

            try:
                from pathlib import Path
                safe_path = guard.resolve_and_validate(file_path)
                if not Path(safe_path).is_file():
                    continue
                suffix = Path(safe_path).suffix.lower()
                if suffix not in (".xlsx", ".xlsm", ".xls", ".xlsb"):
                    continue

                import openpyxl
                wb = openpyxl.load_workbook(safe_path, read_only=True, data_only=True)
                try:
                    target_sheets = [sheet] if sheet and sheet in wb.sheetnames else wb.sheetnames[:3]
                    for sn in target_sheets:
                        ws = wb[sn]
                        total_cells = 0
                        null_cells = 0
                        row_count = 0
                        for row in ws.iter_rows(min_row=2, max_row=FinishTaskHandler._NULL_CHECK_MAX_ROWS + 1):
                            row_count += 1
                            for cell in row:
                                total_cells += 1
                                if cell.value is None:
                                    null_cells += 1
                        if total_cells > 0 and row_count > 1:
                            null_rate = null_cells / total_cells
                            if null_rate > FinishTaskHandler._NULL_RATE_THRESHOLD:
                                pct = int(null_rate * 100)
                                alerts.append(
                                    f"文件 `{file_path}` Sheet `{sn}`: "
                                    f"{pct}% 单元格为空值（{null_cells}/{total_cells}）"
                                )
                finally:
                    wb.close()
            except Exception:  # noqa: BLE001
                _logging.getLogger(__name__).debug(
                    "null rate check skipped for %s: %s", file_path, exc_info=True,
                )
                continue

        if not alerts:
            return None

        detail = "\n".join(f"- {a}" for a in alerts[:3])
        return (
            f"⚠️ 输出文件空值率异常，疑似数据未正确写入：\n{detail}\n\n"
            "请先用 `read_excel` 或 `run_code` 检查写入结果是否正确，"
            "确认无误后再次调用 finish_task。"
        )

    def _resolve_verifier_level(
        self, task_tags: tuple[str, ...], has_write: bool, write_hint: str,
    ) -> str:
        """根据 task_tags 和写入状态决定验证级别。

        Returns: "skip" | "advisory" | "blocking"
        """
        # 无写入 + read_only hint → 跳过
        if not has_write and write_hint == "read_only":
            return "skip"

        # 从 tag 映射中取最高级别
        level = self._DEFAULT_VERIFIER_LEVEL
        for tag in task_tags:
            tag_level = self._VERIFIER_LEVEL_BY_TAG.get(tag)
            if tag_level == "blocking":
                level = "blocking"
                break  # blocking 已是最高级别
        return level

    async def _run_verifier_if_needed(
        self,
        engine: Any,
        *,
        report: dict | None,
        summary: str,
        on_event: Any,
    ) -> str | None:
        """根据 task_tags 决定 verifier 模式并执行。

        返回值含义：
        - None: 跳过或 fail-open（advisory 后台模式也返回 None，结果通过 _pending_verifier_task 异步获取）
        - "BLOCK:..." : blocking 模式下验证失败
        - 其他字符串: advisory 追加文本（仅 blocking 降级时同步返回）
        """
        last_route = getattr(engine, "_last_route_result", None)
        task_tags: tuple[str, ...] = ()
        if last_route is not None:
            task_tags = tuple(getattr(last_route, "task_tags", ()) or ())

        has_write = getattr(engine, "_has_write_tool_call", False)
        write_hint = getattr(engine, "_current_write_hint", "unknown")
        level = self._resolve_verifier_level(task_tags, has_write, write_hint)

        if level == "skip":
            return None

        attempt_count = getattr(engine, "_verification_attempt_count", 0)

        if level == "blocking" and attempt_count < self._MAX_BLOCKING_ATTEMPTS:
            engine._verification_attempt_count = attempt_count + 1
            return await engine._run_finish_verifier_advisory(
                report=report, summary=summary, on_event=on_event, blocking=True,
            )
        else:
            # advisory / blocking 降级 → 后台并行执行，不阻塞 finish 回复
            task = asyncio.create_task(
                engine._run_finish_verifier_advisory(
                    report=report, summary=summary, on_event=on_event, blocking=False,
                ),
                name="verifier_advisory",
            )
            engine._pending_verifier_task = task
            return None


# ---------------------------------------------------------------------------
# 询问用户处理器（AskUserHandler）
# ---------------------------------------------------------------------------

class AskUserHandler(BaseToolHandler):
    """处理 ask_user 工具调用。

    阻塞模式：await 用户回答（通过 InteractionRegistry Future），
    返回回答内容作为 tool result，循环不中断。
    """

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "ask_user"

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        result_str = await self._engine.handle_ask_user_blocking(
            arguments=arguments, tool_call_id=tool_call_id, on_event=on_event, iteration=iteration,
        )
        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(
            result_str=result_str, success=True,
            pending_question=False, question_id=None, defer_tool_result=False,
        )


# ---------------------------------------------------------------------------
# 建议模式切换处理器（SuggestModeSwitchHandler）
# ---------------------------------------------------------------------------

class SuggestModeSwitchHandler(BaseToolHandler):
    """处理 suggest_mode_switch 工具调用。

    阻塞模式：await 用户选择后返回结果。
    """

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "suggest_mode_switch"

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        import asyncio
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        target_mode = str(arguments.get("target_mode", "write")).strip()
        reason = str(arguments.get("reason", "")).strip()
        mode_labels = {"write": "写入", "read": "读取", "plan": "计划"}
        target_label = mode_labels.get(target_mode, target_mode)

        question_payload = {
            "header": "建议切换模式",
            "text": f"{reason}\n\n是否切换到「{target_label}」模式？",
            "options": [
                {"label": f"切换到{target_label}", "description": f"切换到{target_label}模式继续"},
                {"label": "保持当前模式", "description": "不切换，继续当前模式"},
            ],
            "multiSelect": False,
        }

        pending_q = e._question_flow.enqueue(
            question_payload=question_payload,
            tool_call_id=tool_call_id,
        )
        e._interaction_handler.emit_user_question_event(
            question=pending_q,
            on_event=on_event,
            iteration=iteration,
        )

        # 阻塞等待用户回答（支持 question_resolver / InteractionRegistry）
        try:
            payload = await e.await_question_answer(pending_q)
        except (asyncio.TimeoutError, asyncio.CancelledError):
            e._question_flow.pop_current()
            e._interaction_registry.cleanup_done()
            result_str = "用户未回答模式切换建议（超时/取消）。"
            log_tool_call(logger, tool_name, arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True)

        e._question_flow.pop_current()
        e._interaction_registry.cleanup_done()

        # 解析用户选择并执行实际模式切换
        accepted = False
        if isinstance(payload, dict):
            selected_options = payload.get("selected_options", [])
            if selected_options:
                first_label = str(selected_options[0].get("label", "")).strip()
                # 第一个选项 = 确认切换
                if first_label.startswith("切换到"):
                    accepted = True

        old_mode = getattr(e, "_current_chat_mode", "write")
        if accepted and target_mode in ("write", "read", "plan"):
            e._current_chat_mode = target_mode
            e._tools_cache = None  # 失效工具缓存，下轮重建时反映新模式
            result_str = (
                f"用户已确认切换。模式已从「{mode_labels.get(old_mode, old_mode)}」"
                f"切换到「{target_label}」。请按新模式继续执行任务。"
            )
        else:
            result_str = (
                f"用户选择保持当前「{mode_labels.get(old_mode, old_mode)}」模式。"
                "请在当前模式下继续。"
            )

        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)


# ---------------------------------------------------------------------------
# 仅审计处理器（AuditOnlyHandler）
# ---------------------------------------------------------------------------

class AuditOnlyHandler(BaseToolHandler):
    """处理 audit-only 工具（低风险但需审计）。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return self._engine.approval.is_audit_only_tool(tool_name)

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        result_value, audit_record = await e.execute_tool_with_audit(
            tool_name=tool_name, arguments=arguments, tool_scope=tool_scope,
            approval_id=e.approval.new_approval_id(), created_at_utc=e.approval.utc_now(),
            undoable=e.approval.is_undoable_tool(tool_name),
        )
        result_str = str(result_value)
        raw_result_str = result_str
        tool_def = getattr(e.registry, "get_tool", lambda _: None)(tool_name)
        if tool_def is not None:
            result_str = tool_def.truncate_result(result_str)
        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True, audit_record=audit_record, raw_result_str=raw_result_str)


# ---------------------------------------------------------------------------
# 高风险审批处理器（HighRiskApprovalHandler）
# ---------------------------------------------------------------------------

class HighRiskApprovalHandler(BaseToolHandler):
    """处理高风险工具（需审批或 fullaccess 直接执行）。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return self._engine.approval.is_high_risk_tool(tool_name)

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None, skip_high_risk_approval_by_hook=False):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        e = self._engine
        if not e.full_access_enabled and not skip_high_risk_approval_by_hook:
            pending = e.approval.create_pending(tool_name=tool_name, arguments=arguments, tool_scope=tool_scope)
            e.emit_pending_approval_event(pending=pending, on_event=on_event, iteration=iteration, tool_call_id=tool_call_id)
            result_str = e.format_pending_prompt(pending)
            log_tool_call(logger, tool_name, arguments, result=result_str)
            return _ToolExecOutcome(
                result_str=result_str, success=True,
                pending_approval=True, approval_id=pending.approval_id,
            )
        elif e.approval.is_mcp_tool(tool_name):
            probe_before, probe_before_partial = self._dispatcher._capture_unknown_write_probe(tool_name)
            result_value = await self._dispatcher.call_registry_tool(tool_name=tool_name, arguments=arguments, tool_scope=tool_scope)
            self._dispatcher._apply_unknown_write_probe(tool_name=tool_name, before_snapshot=probe_before, before_partial=probe_before_partial)
            result_str = str(result_value)
            raw_result_str = getattr(self._dispatcher, '_last_call_raw_result', result_str)
            log_tool_call(logger, tool_name, arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True, raw_result_str=raw_result_str)
        else:
            result_value, audit_record = await e.execute_tool_with_audit(
                tool_name=tool_name, arguments=arguments, tool_scope=tool_scope,
                approval_id=e.approval.new_approval_id(), created_at_utc=e.approval.utc_now(),
                undoable=e.approval.is_undoable_tool(tool_name),
            )
            result_str = str(result_value)
            raw_result_str = result_str
            tool_def = getattr(e.registry, "get_tool", lambda _: None)(tool_name)
            if tool_def is not None:
                result_str = tool_def.truncate_result(result_str)
            log_tool_call(logger, tool_name, arguments, result=result_str)
            return _ToolExecOutcome(result_str=result_str, success=True, audit_record=audit_record, raw_result_str=raw_result_str)


# ---------------------------------------------------------------------------
# 默认工具处理器（DefaultToolHandler）
# ---------------------------------------------------------------------------

class DefaultToolHandler(BaseToolHandler):
    """兜底：普通 registry 工具直接调用。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return True  # 兜底，总是匹配

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        probe_before, probe_before_partial = self._dispatcher._capture_unknown_write_probe(tool_name)
        result_value = await self._dispatcher.call_registry_tool(tool_name=tool_name, arguments=arguments, tool_scope=tool_scope)
        self._dispatcher._apply_unknown_write_probe(tool_name=tool_name, before_snapshot=probe_before, before_partial=probe_before_partial)
        result_str = str(result_value)
        raw_result_str = getattr(self._dispatcher, '_last_call_raw_result', result_str)
        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True, raw_result_str=raw_result_str)


# ---------------------------------------------------------------------------
# 提取表结构处理器（ExtractTableSpecHandler）
# ---------------------------------------------------------------------------

_SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
_MAX_IMAGE_SIZE_BYTES = 20_000_000


class ExtractTableSpecHandler(BaseToolHandler):
    """处理 extract_table_spec 工具：4 阶段渐进式 VLM 提取 → ReplicaSpec JSON。
    
    支持两种模式：
    - 单个文件：使用 ProgressivePipeline
    - 多个文件：使用 ProgressivePipelineBatch（批量优化，减少 VLM 调用）
    """

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "extract_table_spec"

    async def handle(
        self, tool_name, tool_call_id, arguments, *,
        tool_scope=None, on_event=None, iteration=0, route_result=None,
    ):

        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        file_path = arguments.get("file_path", "")
        file_paths = arguments.get("file_paths", [])
        output_path = arguments.get("output_path", "outputs/replica_spec.json")
        skip_style = arguments.get("skip_style", False)

        # 统一处理：单个文件转为列表
        if file_path and not file_paths:
            file_paths = [file_path]
        
        if not file_paths:
            result_str = json.dumps(
                {"status": "error", "message": "必须提供 file_path 或 file_paths 参数"},
                ensure_ascii=False,
            )
            log_tool_call(logger, tool_name, arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 包装 on_event 回调，为 PIPELINE_PROGRESS 事件自动注入 tool_call_id
        def _on_event_with_tool_id(event):
            if on_event is not None:
                from excelmanus.events import EventType
                if hasattr(event, "event_type") and event.event_type == EventType.PIPELINE_PROGRESS:
                    event.tool_call_id = tool_call_id
                on_event(event)

        wrapped_on_event = _on_event_with_tool_id if on_event else None

        # 判断是否使用批量模式
        use_batch = len(file_paths) > 1

        if use_batch:
            return await self._handle_batch(
                file_paths=file_paths,
                output_path=output_path,
                skip_style=skip_style,
                on_event=wrapped_on_event,
                arguments=arguments,
            )
        else:
            return await self._handle_single(
                file_path=file_paths[0],
                output_path=output_path,
                skip_style=skip_style,
                on_event=wrapped_on_event,
                arguments=arguments,
            )

    async def _handle_single(
        self,
        file_path: str,
        output_path: str,
        skip_style: bool,
        on_event,
        arguments: dict,
    ) -> _ToolExecOutcome:
        """处理单个文件模式。"""
        from pathlib import Path
        from datetime import datetime, timezone

        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher, _ToolExecOutcome
        from excelmanus.pipeline import PipelineConfig, PipelinePauseError, ProgressivePipeline

        # ── 校验文件（基于 workspace_root 解析相对路径） ──
        from excelmanus.security import FileAccessGuard, SecurityViolationError
        workspace_root = self._engine.config.workspace_root
        guard = FileAccessGuard(workspace_root)
        try:
            path = guard.resolve_and_validate(file_path)
        except SecurityViolationError as exc:
            result_str = json.dumps(
                {"status": "error", "message": f"路径校验失败: {exc}"},
                ensure_ascii=False,
            )
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)
        if not path.is_file():
            result_str = json.dumps(
                {"status": "error", "message": f"文件不存在: {file_path}"},
                ensure_ascii=False,
            )
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        if path.suffix.lower() not in _SUPPORTED_IMAGE_EXTENSIONS:
            result_str = json.dumps(
                {"status": "error", "message": f"不支持的图片格式: {path.suffix}"},
                ensure_ascii=False,
            )
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        size = path.stat().st_size
        if size > _MAX_IMAGE_SIZE_BYTES:
            result_str = json.dumps(
                {"status": "error", "message": f"文件过大: {size} > {_MAX_IMAGE_SIZE_BYTES}"},
                ensure_ascii=False,
            )
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # ── 读取图片 ──
        raw_bytes = path.read_bytes()
        ext = path.suffix.lower()
        mime_map = {
            ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".bmp": "image/bmp", ".webp": "image/webp",
        }
        mime = mime_map.get(ext, "image/png")

        # ── 构建适配器回调 ──
        e = self._engine
        dispatcher = self._dispatcher

        async def _vlm_caller(
            messages: list[dict], phase_label: str, response_format: dict | None,
        ) -> tuple[str | None, Exception | None]:
            raw_text, error, _fr = await dispatcher._call_vlm_with_retry(
                messages=messages,
                vlm_client=e.vlm_client,
                vlm_model=e.vlm_model,
                vlm_timeout=e.config.vlm_timeout_seconds,
                vlm_max_retries=e.config.vlm_max_retries,
                vlm_base_delay=e.config.vlm_retry_base_delay_seconds,
                phase_label=phase_label,
                response_format=response_format,
                max_tokens=e.config.vlm_max_tokens,
            )
            return raw_text, error

        def _image_preparer(raw: bytes, mode: str) -> tuple[bytes, str]:
            return ToolDispatcher._prepare_image_for_vlm(
                raw,
                max_long_edge=e.config.vlm_image_max_long_edge,
                jpeg_quality=e.config.vlm_image_jpeg_quality,
                mode=mode,
            )

        # ── 公共参数 ──
        from excelmanus.engine_core.tool_dispatcher import _image_content_hash
        image_hash = f"sha256:{_image_content_hash(raw_bytes)}"
        provenance = {
            "source_image_hash": image_hash,
            "model": e.vlm_model,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        out_path = Path(output_path)
        output_dir = str(out_path.parent) if out_path.parent != Path(".") else "outputs"
        output_basename = out_path.stem

        # ── 策略选择：单轮提取 vs 4 阶段 Pipeline ──
        strategy = self._select_extraction_strategy(e, skip_style)
        if strategy == "single_pass":
            result = await self._handle_single_pass(
                raw_bytes=raw_bytes,
                mime=mime,
                path=path,
                provenance=provenance,
                output_dir=output_dir,
                output_basename=output_basename,
                skip_style=skip_style,
                on_event=on_event,
                arguments=arguments,
                _vlm_caller=_vlm_caller,
                _image_preparer=_image_preparer,
            )
            if result is not None:
                return result
            # 单轮提取失败 → 回退到 Pipeline
            logger.warning("单轮提取失败，回退到 4 阶段 Pipeline")

        # ── 4 阶段 Pipeline 路径 ──
        pipeline_config = PipelineConfig(
            skip_style=skip_style,
            uncertainty_pause_threshold=e.config.vlm_pipeline_uncertainty_threshold,
            uncertainty_confidence_floor=e.config.vlm_pipeline_uncertainty_confidence_floor,
            chunk_cell_threshold=e.config.vlm_pipeline_chunk_cell_threshold,
        )

        pipeline = ProgressivePipeline(
            image_bytes=raw_bytes,
            mime=mime,
            file_path=str(path),
            output_dir=output_dir,
            output_basename=output_basename,
            config=pipeline_config,
            vlm_caller=_vlm_caller,
            image_preparer=_image_preparer,
            provenance=provenance,
            on_event=on_event,
        )

        # 注入 B 通道描述缓存到 Pipeline（若图片匹配）
        cached_desc = dispatcher._last_vlm_description
        cached_hash = dispatcher._last_vlm_description_image_hash
        if cached_desc and cached_hash:
            current_hash = _image_content_hash(raw_bytes)
            if current_hash == cached_hash:
                pipeline._b_channel_description = cached_desc
                logger.info("B 通道描述已注入 Pipeline 结构阶段 (hash=%s)", cached_hash)

        try:
            spec, spec_path = await pipeline.run()
        except PipelinePauseError as pause:
            result_str = json.dumps({
                "status": "paused",
                "message": f"管线在 {pause.phase.value} 阶段暂停：{len(pause.uncertainties)} 个不确定项",
                "spec_path": pause.spec_path,
                "checkpoint": pause.checkpoint,
                "uncertainties": [
                    {"location": u.location, "reason": u.reason, "confidence": u.confidence}
                    for u in pause.uncertainties[:10]
                ],
                "hint": "请确认不确定项后，使用 resume_from_phase 继续管线。",
            }, ensure_ascii=False)
            log_tool_call(logger, "extract_table_spec", arguments, result=result_str)
            self._engine.record_write_action()
            return _ToolExecOutcome(result_str=result_str, success=True)
        except RuntimeError as exc:
            result_str = ToolDispatcher._build_vlm_failure_result(exc, 1, str(path))
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)
        except Exception as exc:
            result_str = json.dumps({
                "status": "error",
                "message": f"管线执行失败: {exc}",
            }, ensure_ascii=False)
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        self._engine.record_write_action()
        return self._build_spec_result(spec, spec_path, arguments)

    @staticmethod
    def _select_extraction_strategy(engine: "AgentEngine", skip_style: bool) -> str:
        """选择提取策略：single_pass 或 pipeline。

        策略矩阵：
        ┌────────────┬─────────┬──────────┐
        │ model_tier │ simple  │ complex  │
        ├────────────┼─────────┼──────────┤
        │ strong     │ single  │ single   │  ← Gemini 2.5 Pro / GPT-5 级别
        │ standard   │ single  │ pipeline │  ← GPT-4.1 / Claude Sonnet 4 级别
        │ weak       │ pipeline│ pipeline │  ← 小模型
        └────────────┴─────────┴──────────┘
        """
        tier = getattr(engine.config, "vlm_extraction_tier", "auto")

        if tier == "auto":
            model_name = (engine.vlm_model or "").lower()
            if any(k in model_name for k in [
                "gemini-2.5", "gemini-3", "gpt-5",
            ]):
                tier = "strong"
            elif any(k in model_name for k in [
                "gpt-4o", "gpt-4.1", "claude-4",
                "claude-sonnet", "claude-opus",
                "qwen-vl-max", "qwen2.5-vl-72b",
            ]):
                tier = "standard"
            else:
                tier = "weak"

        if tier == "strong":
            logger.info("提取策略: single_pass (model_tier=strong)")
            return "single_pass"
        if tier == "weak":
            logger.info("提取策略: pipeline (model_tier=weak)")
            return "pipeline"
        # standard: skip_style 时走单轮（更快），否则走 pipeline（样式需要多阶段）
        if skip_style:
            logger.info("提取策略: single_pass (model_tier=standard, skip_style=True)")
            return "single_pass"
        logger.info("提取策略: pipeline (model_tier=standard)")
        return "pipeline"

    async def _handle_single_pass(
        self,
        *,
        raw_bytes: bytes,
        mime: str,
        path: "Path",
        provenance: dict,
        output_dir: str,
        output_basename: str,
        skip_style: bool,
        on_event,
        arguments: dict,
        _vlm_caller,
        _image_preparer,
    ) -> "_ToolExecOutcome | None":
        """单轮合并提取：一次 VLM 调用完成结构+数据+样式。

        Returns:
            _ToolExecOutcome 或 None（失败，应回退到 Pipeline）。
        """
        import base64
        from excelmanus.pipeline.single_pass import (
            SINGLE_PASS_PROMPT,
            SINGLE_PASS_NO_STYLE_PROMPT,
            parse_single_pass_result,
        )
        from excelmanus.events import EventType, ToolCallEvent

        e = self._engine
        _dispatcher = self._dispatcher

        # 发射进度事件
        if on_event:
            e.emit(on_event, ToolCallEvent(
                event_type=EventType.PIPELINE_PROGRESS,
                pipeline_stage="single_pass",
                pipeline_message="正在执行单轮提取...",
                pipeline_phase_index=0,
                pipeline_total_phases=1,
            ))

        # 图片预处理
        compressed, c_mime = _image_preparer(raw_bytes, "data")
        b64 = base64.b64encode(compressed).decode("ascii")
        prompt = SINGLE_PASS_NO_STYLE_PROMPT if skip_style else SINGLE_PASS_PROMPT

        messages = [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {
                    "url": f"data:{c_mime};base64,{b64}", "detail": "high",
                }},
                {"type": "text", "text": prompt},
            ],
        }]

        raw_text, error = await _vlm_caller(messages, "单轮提取", {"type": "json_object"})

        if raw_text is None:
            logger.warning("单轮提取 VLM 调用失败: %s", error)
            return None

        spec = parse_single_pass_result(raw_text, provenance)
        if spec is None:
            logger.warning("单轮提取结果解析失败")
            return None

        # 保存 spec
        from pathlib import Path as _Path
        spec_path = _Path(output_dir) / f"{output_basename}.json"
        spec_path.parent.mkdir(parents=True, exist_ok=True)
        spec_path.write_text(
            spec.model_dump_json(indent=2, exclude_none=True), encoding="utf-8",
        )

        # 发射完成事件
        if on_event:
            e.emit(on_event, ToolCallEvent(
                event_type=EventType.PIPELINE_PROGRESS,
                pipeline_stage="single_pass_done",
                pipeline_message="单轮提取完成",
                pipeline_phase_index=0,
                pipeline_total_phases=1,
                pipeline_spec_path=str(spec_path),
            ))

        self._engine.record_write_action()
        return self._build_spec_result(spec, str(spec_path), arguments)

    @staticmethod
    def _build_spec_result(
        spec: "ReplicaSpec", spec_path: str, arguments: dict,
    ) -> "_ToolExecOutcome":
        """构建 extract_table_spec 成功的 result。"""
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome

        total_cells = sum(len(s.cells) for s in spec.sheets)
        has_styles = any(bool(s.styles) for s in spec.sheets)
        result_str = json.dumps({
            "status": "ok",
            "output_path": spec_path,
            "table_count": len(spec.sheets),
            "cell_count": total_cells,
            "uncertainties_count": len(spec.uncertainties),
            "has_styles": has_styles,
            "hint": (
                f"已生成 ReplicaSpec ({len(spec.sheets)} 个表格, {total_cells} 个单元格)。"
                "下一步请调用 rebuild_excel_from_spec 编译为 Excel 文件。"
            ),
        }, ensure_ascii=False)
        log_tool_call(logger, "extract_table_spec", arguments, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True)

    async def _handle_batch(
        self,
        file_paths: list[str],
        output_path: str,
        skip_style: bool,
        on_event,
        arguments: dict,
    ) -> _ToolExecOutcome:
        """处理批量文件模式 - 使用批量优化的管线。"""
        from pathlib import Path
        from datetime import datetime, timezone

        from excelmanus.engine_core.tool_dispatcher import ToolDispatcher, _ToolExecOutcome
        from excelmanus.pipeline.batch import ProgressivePipelineBatch, BatchPipelineConfig

        # ── 校验所有文件 ──
        from excelmanus.security import FileAccessGuard, SecurityViolationError
        workspace_root = self._engine.config.workspace_root
        guard = FileAccessGuard(workspace_root)

        mime_map = {
            ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".bmp": "image/bmp", ".webp": "image/webp",
        }

        items = []
        for fp in file_paths:
            try:
                path = guard.resolve_and_validate(fp)
            except SecurityViolationError as exc:
                result_str = json.dumps(
                    {"status": "error", "message": f"路径校验失败: {exc}"},
                    ensure_ascii=False,
                )
                log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)
            
            if not path.is_file():
                result_str = json.dumps(
                    {"status": "error", "message": f"文件不存在: {fp}"},
                    ensure_ascii=False,
                )
                log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

            if path.suffix.lower() not in _SUPPORTED_IMAGE_EXTENSIONS:
                result_str = json.dumps(
                    {"status": "error", "message": f"不支持的图片格式: {path.suffix}"},
                    ensure_ascii=False,
                )
                log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

            size = path.stat().st_size
            if size > _MAX_IMAGE_SIZE_BYTES:
                result_str = json.dumps(
                    {"status": "error", "message": f"文件过大: {size} > {_MAX_IMAGE_SIZE_BYTES}"},
                    ensure_ascii=False,
                )
                log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
                return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

            # 读取图片
            raw_bytes = path.read_bytes()
            ext = path.suffix.lower()
            mime = mime_map.get(ext, "image/png")

            from excelmanus.engine_core.tool_dispatcher import _image_content_hash
            image_hash = f"sha256:{_image_content_hash(raw_bytes)}"
            provenance = {
                "source_image_hash": image_hash,
                "model": self._engine.vlm_model,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }

            items.append({
                "image_bytes": raw_bytes,
                "mime": mime,
                "file_path": str(path),
                "output_dir": "outputs",
                "output_basename": path.stem,
                "provenance": provenance,
            })

        # ── 构建适配器回调 ──
        e = self._engine
        dispatcher = self._dispatcher

        async def _vlm_caller(
            messages: list[dict], phase_label: str, response_format: dict | None,
        ) -> tuple[str | None, Exception | None]:
            raw_text, error, _fr = await dispatcher._call_vlm_with_retry(
                messages=messages,
                vlm_client=e.vlm_client,
                vlm_model=e.vlm_model,
                vlm_timeout=e.config.vlm_timeout_seconds,
                vlm_max_retries=e.config.vlm_max_retries,
                vlm_base_delay=e.config.vlm_retry_base_delay_seconds,
                phase_label=phase_label,
                response_format=response_format,
                max_tokens=e.config.vlm_max_tokens,
            )
            return raw_text, error

        def _image_preparer(raw: bytes, mode: str) -> tuple[bytes, str]:
            return ToolDispatcher._prepare_image_for_vlm(
                raw,
                max_long_edge=e.config.vlm_image_max_long_edge,
                jpeg_quality=e.config.vlm_image_jpeg_quality,
                mode=mode,
            )

        # 输出目录
        out_path = Path(output_path)
        _output_dir = str(out_path.parent) if out_path.parent != Path(".") else "outputs"

        # 批量管线配置
        batch_config = BatchPipelineConfig.from_pipeline_config(PipelineConfig(
            skip_style=skip_style,
            uncertainty_pause_threshold=e.config.vlm_pipeline_uncertainty_threshold,
            uncertainty_confidence_floor=e.config.vlm_pipeline_uncertainty_confidence_floor,
            chunk_cell_threshold=e.config.vlm_pipeline_chunk_cell_threshold,
        ))

        # 创建批量管线
        pipeline = ProgressivePipelineBatch(
            items=items,
            config=batch_config,
            vlm_caller=_vlm_caller,
            image_preparer=_image_preparer,
            on_event=on_event,
        )

        try:
            results = await pipeline.run()
        except Exception as exc:
            result_str = json.dumps({
                "status": "error",
                "message": f"批量管线执行失败: {exc}",
            }, ensure_ascii=False)
            log_tool_call(logger, "extract_table_spec", arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)

        # 统计结果
        success_count = sum(1 for spec, _ in results if spec is not None)
        total_count = len(results)
        
        output_files = [path for _, path in results if path]
        
        result_str = json.dumps({
            "status": "ok",
            "total_files": total_count,
            "success_count": success_count,
            "output_paths": output_files,
            "hint": (
                f"批量提取完成：{success_count}/{total_count} 个文件成功。"
                "下一步请调用 rebuild_excel_from_spec 编译为 Excel 文件。"
            ),
        }, ensure_ascii=False)
        log_tool_call(logger, "extract_table_spec", arguments, result=result_str)
        self._engine.record_write_action()
        return _ToolExecOutcome(result_str=result_str, success=True)


# ---------------------------------------------------------------------------
# 代码策略处理器（CodePolicyHandler）
# ---------------------------------------------------------------------------

class CodePolicyHandler(BaseToolHandler):
    """处理 run_code 工具（代码策略引擎路由）。"""

    def can_handle(self, tool_name: str, **kwargs: Any) -> bool:
        return tool_name == "run_code" and self._engine.config.code_policy_enabled

    async def handle(self, tool_name, tool_call_id, arguments, *, tool_scope=None, on_event=None, iteration=0, route_result=None):
        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome
        from excelmanus.security.code_policy import CodePolicyEngine, CodeRiskTier, strip_exit_calls

        e = self._engine
        _code_arg = arguments.get("code") or ""
        _cp_engine = CodePolicyEngine(
            extra_safe_modules=e.config.code_policy_extra_safe_modules,
            extra_blocked_modules=e.config.code_policy_extra_blocked_modules,
        )
        _analysis = _cp_engine.analyze(_code_arg)
        _auto_green = _analysis.tier == CodeRiskTier.GREEN and e.config.code_policy_green_auto_approve
        _auto_yellow = _analysis.tier == CodeRiskTier.YELLOW and e.config.code_policy_yellow_auto_approve

        if _auto_green or _auto_yellow or e.full_access_enabled:
            return await self._execute_code_with_policy(
                code=_code_arg, arguments=arguments, analysis=_analysis,
                tool_name=tool_name, tool_call_id=tool_call_id, tool_scope=tool_scope,
                on_event=on_event, iteration=iteration,
            )

        # 风险等级为 RED 或配置不允许自动执行 → 尝试清洗降级
        _sanitized_code = strip_exit_calls(_code_arg) if _analysis.tier == CodeRiskTier.RED else None
        if _sanitized_code is not None:
            _re_analysis = _cp_engine.analyze(_sanitized_code)
            _re_auto_green = _re_analysis.tier == CodeRiskTier.GREEN and e.config.code_policy_green_auto_approve
            _re_auto_yellow = _re_analysis.tier == CodeRiskTier.YELLOW and e.config.code_policy_yellow_auto_approve
            if _re_auto_green or _re_auto_yellow:
                logger.info(
                    "run_code 自动清洗: %s → %s (移除退出调用)",
                    _analysis.tier.value, _re_analysis.tier.value,
                )
                _sanitized_args = {**arguments, "code": _sanitized_code}
                return await self._execute_code_with_policy(
                    code=_sanitized_code, arguments=_sanitized_args, analysis=_re_analysis,
                    tool_name=tool_name, tool_call_id=tool_call_id, tool_scope=tool_scope,
                    on_event=on_event, iteration=iteration, label_suffix="(清洗后)",
                )

        # 无法降级 → /accept 审批流程
        _caps_detail = ", ".join(sorted(_analysis.capabilities))
        _details_text = "; ".join(_analysis.details[:3])
        pending = e.approval.create_pending(tool_name=tool_name, arguments=arguments, tool_scope=tool_scope)
        result_str = (
            f"⚠️ 代码包含高风险操作，需要人工确认：\n"
            f"- 风险等级: {_analysis.tier.value}\n"
            f"- 检测到: {_caps_detail}\n"
            f"- 详情: {_details_text}\n"
            f"{e.format_pending_prompt(pending)}"
        )
        e.emit_pending_approval_event(
            pending=pending, on_event=on_event, iteration=iteration, tool_call_id=tool_call_id,
        )
        logger.info("run_code 策略引擎: tier=%s → pending approval %s", _analysis.tier.value, pending.approval_id)
        log_tool_call(logger, tool_name, arguments, result=result_str)
        return _ToolExecOutcome(
            result_str=result_str, success=True,
            pending_approval=True, approval_id=pending.approval_id,
        )

    async def _execute_code_with_policy(
        self,
        *,
        code: str,
        arguments: dict[str, Any],
        analysis: Any,
        tool_name: str,
        tool_call_id: str,
        tool_scope: Sequence[str] | None,
        on_event: Any,
        iteration: int,
        label_suffix: str = "",
    ) -> _ToolExecOutcome:
        """统一的代码策略执行路径（GREEN/YELLOW/降级后均走此方法）。

        消除原先 GREEN/YELLOW 路径与 RED→降级路径的 ~100 行重复代码。
        """
        import json as _json

        from excelmanus.engine_core.tool_dispatcher import _ToolExecOutcome
        from excelmanus.security.code_policy import extract_excel_targets

        e = self._engine
        dispatcher = self._dispatcher

        _sandbox_tier = analysis.tier.value
        _augmented_args = {**arguments, "sandbox_tier": _sandbox_tier}

        # ── run_code 前: 对可能被修改的 Excel 文件做快照 ──
        _excel_targets = [
            t.file_path for t in extract_excel_targets(code)
            if t.operation in ("write", "unknown")
        ]
        _before_snap = dispatcher._snapshot_excel_for_diff(
            _excel_targets, e.config.workspace_root,
        ) if _excel_targets else {}
        # uploads 目录快照，用于检测新建/变更文件
        _uploads_before = dispatcher._snapshot_uploads_dir(e.config.workspace_root)

        result_value, audit_record = await e.execute_tool_with_audit(
            tool_name=tool_name, arguments=_augmented_args, tool_scope=tool_scope,
            approval_id=e.approval.new_approval_id(), created_at_utc=e.approval.utc_now(),
            undoable=False,
        )
        result_str = str(result_value)
        tool_def = getattr(e.registry, "get_tool", lambda _: None)(tool_name)
        if tool_def is not None:
            result_str = tool_def.truncate_result(result_str)

        # ── 写入追踪 ──
        _rc_json: dict | None = None
        try:
            _rc_json = _json.loads(result_str)
            if not isinstance(_rc_json, dict):
                _rc_json = None
        except (_json.JSONDecodeError, TypeError):
            pass
        _has_cow = bool(_rc_json and _rc_json.get("cow_mapping"))
        _has_ast_write = any(t.operation == "write" for t in extract_excel_targets(code))
        if (audit_record is not None and audit_record.changes) or _has_cow or _has_ast_write:
            e.record_write_action()
            # 写入操作日志（供 verifier playbook 注入）
            _state = getattr(e, "_state", None)
            if _state is not None:
                _cow_paths = ""
                if _has_cow and _rc_json:
                    _cow_map = _rc_json.get("cow_mapping")
                    if isinstance(_cow_map, dict):
                        _cow_paths = ", ".join(
                            str(v) for v in _cow_map.values() if isinstance(v, str) and v.strip()
                        )
                _ast_paths = ", ".join(
                    t.file_path for t in extract_excel_targets(code)
                    if t.operation == "write" and t.file_path != "<variable>"
                ) if _has_ast_write else ""
                _file_path = _cow_paths or _ast_paths
                _state.record_write_operation(
                    tool_name="run_code",
                    file_path=_file_path,
                    summary=dispatcher._extract_run_code_write_summary(result_str),
                )

        # ── window 感知桥接 ──
        _stdout_tail = ""
        if _rc_json is not None:
            _stdout_tail = _rc_json.get("stdout_tail", "")
        if audit_record is not None and e.window_perception is not None:
            e.window_perception.observe_code_execution(
                code=code,
                audit_changes=audit_record.changes if audit_record else None,
                stdout_tail=_stdout_tail,
                iteration=iteration,
            )
            e._context_builder.mark_window_notice_dirty()

        # ── files_changed 事件 ──
        _uploads_after = dispatcher._snapshot_uploads_dir(e.config.workspace_root)
        _uploads_changed = dispatcher._diff_uploads_snapshots(_uploads_before, _uploads_after)
        dispatcher._emit_files_changed_from_audit(
            e, on_event, tool_call_id, code,
            audit_record.changes if audit_record else None,
            iteration,
            extra_changed_paths=_uploads_changed or None,
        )

        # ── Excel diff ──
        if _excel_targets and on_event is not None:
            try:
                _after_snap = dispatcher._snapshot_excel_for_diff(
                    _excel_targets, e.config.workspace_root,
                )
                _diffs = dispatcher._compute_snapshot_diffs(_before_snap, _after_snap)
                from excelmanus.events import EventType, ToolCallEvent
                for _rd in _diffs:
                    _rd_old_merges: list[dict[str, int]] = _rd.get("old_merge_ranges", [])
                    _rd_new_merges: list[dict[str, int]] = _rd.get("new_merge_ranges", [])
                    _rd_hints: list[str] = []
                    try:
                        _, _rd_hints = dispatcher._extract_sheet_metadata(
                            _rd["file_path"], _rd["sheet"] or None,
                            e.config.workspace_root,
                        )
                    except Exception:
                        pass
                    e.emit(
                        on_event,
                        ToolCallEvent(
                            event_type=EventType.EXCEL_DIFF,
                            tool_call_id=tool_call_id,
                            excel_file_path=_rd["file_path"],
                            excel_sheet=_rd["sheet"],
                            excel_affected_range=_rd["affected_range"],
                            excel_changes=_rd["changes"],
                            excel_merge_ranges=_rd_new_merges,
                            excel_old_merge_ranges=_rd_old_merges,
                            excel_metadata_hints=_rd_hints,
                        ),
                    )
            except Exception:
                logger.debug("run_code%s Excel diff 计算失败", label_suffix, exc_info=True)

        logger.info(
            "run_code 策略引擎: tier=%s%s auto_approved=True caps=%s",
            analysis.tier.value, label_suffix, sorted(analysis.capabilities),
        )
        log_tool_call(logger, tool_name, _augmented_args, result=result_str)
        return _ToolExecOutcome(result_str=result_str, success=True, audit_record=audit_record)
