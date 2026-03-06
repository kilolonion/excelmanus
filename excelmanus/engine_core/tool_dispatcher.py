"""ToolDispatcher — 从 AgentEngine 解耦的工具调度组件。

负责管理：
- 工具参数解析（JSON string / dict / None）
- 普通工具的 registry 调用（含线程池执行）
- 单个工具调用的完整执行流程（execute）
- 工具结果截断
"""

from __future__ import annotations

import asyncio
import json
import re
import threading
import time
from collections.abc import Sequence
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

from excelmanus.engine_core.tool_errors import (
    DEFAULT_RETRY_POLICY,
    classify_tool_error,
    compact_error,
)
from excelmanus.engine_core.workspace_probe import (
    collect_workspace_mtime_index,
    has_workspace_mtime_changes,
)
from excelmanus.hooks import HookDecision, HookEvent
from excelmanus.logger import get_logger, log_tool_call
from excelmanus.tools.registry import ToolNotAllowedError


@dataclass
class _ToolExecOutcome:
    """特殊工具 handler 的结构化返回，收敛副作用信号。"""

    result_str: str
    success: bool
    error: str | None = None
    error_kind: str | None = None  # ToolErrorKind.value: retryable/permanent/needs_human/overflow
    pending_approval: bool = False
    approval_id: str | None = None
    audit_record: Any = None
    pending_question: bool = False
    question_id: str | None = None
    defer_tool_result: bool = False
    finish_accepted: bool = False
    raw_result_str: str | None = None  # 截断前的原始结果，供窗口感知解析使用

if TYPE_CHECKING:
    from pathlib import Path

    from excelmanus.engine import AgentEngine
    from excelmanus.events import EventCallback
    from excelmanus.stores.tool_call_store import ToolCallStore

logger = get_logger("tool_dispatcher")


def _render_finish_task_report(
    report: dict[str, Any] | None,
    summary: str,
) -> str:
    """将 finish_task 的参数渲染为用户可读文本。

    新格式只有 summary + affected_files；兼容旧格式的 report dict。
    """
    # 新格式：直接使用 summary 自然语言
    if not report or not isinstance(report, dict):
        return summary.strip() if summary else ""

    # 旧格式兼容：将 report dict 的各字段拼接为自然段落
    parts: list[str] = []
    for key in ("operations", "key_findings", "explanation", "suggestions"):
        text = (report.get(key) or "").strip()
        if text:
            parts.append(text)

    affected_files = report.get("affected_files")
    if affected_files and isinstance(affected_files, list):
        file_lines = [f"- {f}" for f in affected_files if isinstance(f, str) and f.strip()]
        if file_lines:
            parts.append("涉及文件：\n" + "\n".join(file_lines))

    if not parts:
        return summary.strip() if summary else ""

    return "\n\n".join(parts)


# JSON 代码块提取用正则（复用 small_model 的逻辑，避免跨模块依赖）
_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(.*?)```", re.DOTALL | re.IGNORECASE)

# 被视为"输出被截断"的 finish_reason 集合（覆盖不同 VLM 提供商的命名差异）
_TRUNCATION_FINISH_REASONS = {"length", "max_tokens"}


def _is_likely_truncated(raw_text: str, finish_reason: str | None) -> bool:
    """启发式检测 VLM 输出是否被截断。

    检测信号：
    1. finish_reason 明确指示截断
    2. 文本包含 JSON 开头 '{' 但不以 '}' 结尾（启发式）
    """
    if finish_reason and finish_reason.lower() in _TRUNCATION_FINISH_REASONS:
        return True
    stripped = (raw_text or "").rstrip()
    if stripped and '{' in stripped and not stripped.endswith('}'):
        return True
    return False


def _parse_vlm_json(text: str, *, try_repair: bool = False) -> dict[str, Any] | None:
    """从 VLM 输出中提取 JSON dict，支持 fence 包裹、前后缀污染和截断修复。

    Args:
        text: VLM 原始输出文本。
        try_repair: 为 True 时表示已知输出可能被截断（如 finish_reason=length），
                    用于日志提示。无论此参数为何值，解析失败时都会尝试修复。
    """
    content = (text or "").strip()
    if not content:
        return None
    candidates = [content]
    for match in _JSON_FENCE_RE.finditer(content):
        body = (match.group(1) or "").strip()
        if body:
            candidates.append(body)
    left = content.find("{")
    right = content.rfind("}")
    if left >= 0 and right > left:
        candidates.append(content[left: right + 1].strip())
    for candidate in candidates:
        try:
            data = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict):
            return data
    # ── 截断修复：解析失败时始终尝试，try_repair 仅影响日志级别 ──
    if left >= 0:
        log_fn = logger.info if try_repair else logger.debug
        log_fn(
            "JSON 直接解析失败（%d 字符），尝试截断修复",
            len(content),
        )
        return _repair_truncated_json(content[left:])
    return None


def _repair_truncated_json(fragment: str) -> dict[str, Any] | None:
    """尝试修复被截断的 JSON（如 VLM 输出因 max_tokens 被截断）。

    策略：
    1. 收集所有可能的回退切点（逆序扫描，跳过字符串内部）
    2. 从最靠近末尾的切点开始尝试：截断 → 补全未闭合括号 → json.loads
    3. 某个切点修复成功则返回，全部失败返回 None
    """
    if not fragment or fragment[0] != '{':
        return None

    # ── 收集候选切点（从后往前，字符串外的分隔符位置）──
    cut_points: list[int] = []
    in_str = False
    esc = False
    for i, ch in enumerate(fragment):
        if esc:
            esc = False
            continue
        if ch == '\\' and in_str:
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch in (',', '[', '{', '}', ']'):
            # , → 截到它之前（移除尾部不完整元素）
            # 其它 → 截到它之后（保留该括号）
            cut_points.append(i if ch == ',' else i + 1)

    # 从最靠近末尾的切点开始尝试（优先保留更多数据）
    for cut in reversed(cut_points):
        trimmed = fragment[:cut]
        # 统计未闭合括号
        stack: list[str] = []
        s_in_str = False
        s_esc = False
        for ch in trimmed:
            if s_esc:
                s_esc = False
                continue
            if ch == '\\' and s_in_str:
                s_esc = True
                continue
            if ch == '"':
                s_in_str = not s_in_str
                continue
            if s_in_str:
                continue
            if ch in ('{', '['):
                stack.append(ch)
            elif ch == '}' and stack and stack[-1] == '{':
                stack.pop()
            elif ch == ']' and stack and stack[-1] == '[':
                stack.pop()
        # 如果仍在字符串内，说明切点在引号中间——跳过
        if s_in_str:
            continue
        closers = {'[': ']', '{': '}'}
        suffix = ''.join(closers.get(b, '') for b in reversed(stack))
        repaired = trimmed + suffix
        try:
            data = json.loads(repaired)
            if isinstance(data, dict):
                lost = len(fragment) - cut
                logger.info(
                    "截断 JSON 修复成功（回退 %d 字符，补全 %d 个括号）",
                    lost, len(stack),
                )
                return data
        except json.JSONDecodeError:
            continue
    return None


def _image_content_hash(raw_bytes: bytes) -> str:
    """计算图片内容的稳定 hash（全文 sha256，截取前 16 hex）。

    所有图片去重 / B 通道缓存 / provenance 均应使用此函数，
    确保同一张图片在不同代码路径产生相同 hash。
    """
    import hashlib
    return hashlib.sha256(raw_bytes).hexdigest()[:16]


def _image_content_hash_b64(b64_str: str) -> str:
    """从 base64 编码字符串计算图片内容 hash（先解码为原始字节）。

    如果 base64 解码失败（如数据不完整），回退到直接 hash 字符串字节。
    """
    import base64 as _b64
    try:
        raw = _b64.b64decode(b64_str, validate=True)
    except Exception:
        # 容错：无法解码时直接 hash 原始字符串
        raw = b64_str.encode("utf-8") if isinstance(b64_str, str) else b64_str
    return _image_content_hash(raw)


class ToolDispatcher:
    """工具调度器：参数解析、分支路由、执行、审计。"""

    def __init__(self, engine: "AgentEngine") -> None:
        self._engine = engine
        self._pending_vlm_image: dict | None = None
        self._deferred_image_injections: list[dict[str, Any]] = []
        # 已注入图片的 hash 集合（用于去重）
        self._injected_image_hashes: set[str] = set()
        # B 通道最后一次 VLM 描述缓存（供 Pipeline 结构阶段复用）
        self._last_vlm_description: str | None = None
        self._last_vlm_description_image_hash: str | None = None
        # 每会话 sleep 取消事件（abort 时中断正在执行的 sleep 工具）
        self._sleep_cancel_event = threading.Event()
        # 最近一次工具调用的截断前原始结果（供窗口感知解析）
        self._last_call_raw_result: str = ""

        self._tool_call_store: "ToolCallStore | None" = None
        db = getattr(engine, "_database", None)
        if db is not None:
            try:
                from excelmanus.stores.tool_call_store import ToolCallStore as _TCS
                self._tool_call_store = _TCS(db)
            except Exception:
                logger.debug("工具调用审计日志初始化失败", exc_info=True)

        # ── 策略处理器表 ──
        from excelmanus.engine_core.tool_handlers import (
            AskUserHandler,
            AuditOnlyHandler,
            CodePolicyHandler,
            DefaultToolHandler,
            DelegationHandler,
            ExtractTableSpecHandler,
            FinishTaskHandler,
            HighRiskApprovalHandler,
            SkillActivationHandler,
            SkillManagementHandler,
            SuggestModeSwitchHandler,
        )
        # T2: 按工具名建立 O(1) 索引，跳过需动态判断的 handler
        # 每个 handler 只实例化一次，specific 和 generic 复用同一对象
        _specific: dict[str, Any] = {}
        _skill = SkillActivationHandler(engine, self)
        _specific["activate_skill"] = _skill
        _specific["manage_skills"] = SkillManagementHandler(engine, self)
        _deleg = DelegationHandler(engine, self)
        for _dn in ("delegate", "delegate_to_subagent", "list_subagents", "parallel_delegate"):
            _specific[_dn] = _deleg
        _specific["finish_task"] = FinishTaskHandler(engine, self)
        _specific["ask_user"] = AskUserHandler(engine, self)
        _specific["suggest_mode_switch"] = SuggestModeSwitchHandler(engine, self)
        _specific["extract_table_spec"] = ExtractTableSpecHandler(engine, self)
        self._specific_handlers: dict[str, Any] = _specific
        # 动态/条件 handler + 兜底（保持原有顺序）
        _code_policy = CodePolicyHandler(engine, self)
        _audit_only = AuditOnlyHandler(engine, self)
        _high_risk = HighRiskApprovalHandler(engine, self)
        _default = DefaultToolHandler(engine, self)
        self._generic_handlers = [_code_policy, _audit_only, _high_risk, _default]

    @property
    def _registry(self) -> Any:
        return self._engine.registry

    @property
    def _persistent_memory(self) -> Any:
        return self._engine._persistent_memory

    def _capture_unknown_write_probe(self, tool_name: str) -> tuple[dict[str, tuple[int, int]] | None, bool]:
        """为 unknown 写入语义工具采集执行前快照。"""
        e = self._engine
        if e.get_tool_write_effect(tool_name) != "unknown":
            return None, False
        try:
            return collect_workspace_mtime_index(e.config.workspace_root)
        except Exception:
            logger.debug("unknown 写入探针前置快照失败", exc_info=True)
            return None, False

    def _apply_unknown_write_probe(
        self,
        *,
        tool_name: str,
        before_snapshot: dict[str, tuple[int, int]] | None,
        before_partial: bool,
    ) -> None:
        """对 unknown 写入语义工具执行后做 mtime 兜底检测。"""
        if before_snapshot is None:
            return
        e = self._engine
        try:
            after_snapshot, after_partial = collect_workspace_mtime_index(e.config.workspace_root)
        except Exception:
            logger.debug("unknown 写入探针后置快照失败", exc_info=True)
            return

        if has_workspace_mtime_changes(before_snapshot, after_snapshot):
            e.record_workspace_write_action()
            logger.info(
                "unknown 写入探针命中: tool=%s partial_before=%s partial_after=%s",
                tool_name,
                before_partial,
                after_partial,
            )

    # ── 结构化结果提取（统一 JSON 解析） ──────────────────────

    def _extract_structured_result(self, result_str: str) -> tuple[str, dict[str, str] | None]:
        """从工具结果 JSON 中统一提取结构化字段（单次 json.loads）。

        处理：
        - ``__tool_result_image__``: 图片注入（B+C 通道路由）
        - ``cow_mapping``: CoW 路径映射注册

        Returns:
            (cleaned_result_str, cow_mapping_or_none)
        """
        try:
            parsed = json.loads(result_str)
            if not isinstance(parsed, dict):
                return result_str, None
        except (json.JSONDecodeError, TypeError):
            return result_str, None

        mutated = False

        # ── CoW 映射提取 ──
        cow_mapping: dict[str, str] | None = None
        raw_cow = parsed.get("cow_mapping")
        if raw_cow and isinstance(raw_cow, dict):
            cow_mapping = raw_cow
            self._engine.state.register_cow_mappings(cow_mapping)
            tx = self._engine.transaction
            if tx is not None:
                tx.register_cow_mappings(cow_mapping)

        # ── 图片注入提取（延迟注入，避免破坏 tool_calls→tool_responses 序列） ──
        if "__tool_result_image__" in parsed:
            injection = parsed.pop("__tool_result_image__")
            mutated = True
            e = self._engine

            # C 通道：主模型支持视觉 → 延迟注入图片到对话 memory
            # 不能在此处直接调用 add_image_message，否则会在 assistant(tool_calls)
            # 和 tool(responses) 之间插入 user 消息，导致 API 400 错误。
            if e.is_vision_capable:
                # 图片去重：检测同一图片是否已注入过
                _img_hash = _image_content_hash_b64(injection["base64"])
                if _img_hash in self._injected_image_hashes:
                    logger.info("C 通道: 图片已在上下文中 (hash=%s)，跳过重复注入", _img_hash)
                    parsed["hint"] = "图片已在视觉上下文中，无需重复注入。"
                else:
                    self._deferred_image_injections.append({
                        "base64": injection["base64"],
                        "mime_type": injection.get("mime_type", "image/png"),
                        "detail": injection.get("detail", "auto"),
                    })
                    self._injected_image_hashes.add(_img_hash)
                    logger.info("C 通道: 图片已缓存待注入 (hash=%s, mime=%s)", _img_hash, injection.get("mime_type"))
                    parsed["hint"] = "图片已加载到视觉上下文，你现在可以看到这张图片。"
            else:
                logger.info("主模型无视觉能力，跳过图片注入")
                parsed["hint"] = "当前主模型不支持视觉输入，图片未注入。"

            # B 通道：缓存图片数据供异步 VLM 描述
            # 当主模型有视觉能力时，C 通道已直接注入图片，跳过 B 通道以避免
            # 额外的 VLM API 调用延迟（10-30s），主模型直接看图效果已足够。
            if e.vlm_enhance_available and not e.is_vision_capable:
                self._pending_vlm_image = injection
                parsed["vlm_enhance"] = "VLM 增强描述将自动生成并追加到下方。"
            elif not e.is_vision_capable and not e.vlm_enhance_available:
                parsed["hint"] += "且未配置 VLM 增强，无法分析图片内容。建议配置 EXCELMANUS_VLM_* 环境变量。"

        cleaned = json.dumps(parsed, ensure_ascii=False) if mutated else result_str
        return cleaned, cow_mapping

    def flush_deferred_images(self) -> int:
        """将延迟的图片注入实际写入 memory。

        必须在当前 assistant tool_calls 对应的所有 tool result 写入 memory 之后调用，
        否则 user 角色的图片消息会破坏 tool_calls → tool_responses 的消息序列，
        导致 OpenAI 兼容 API 返回 400 错误。

        Returns:
            注入的图片数量。
        """
        if not self._deferred_image_injections:
            return 0
        e = self._engine
        count = 0
        for inj in self._deferred_image_injections:
            e.memory.add_image_message(
                base64_data=inj["base64"],
                mime_type=inj.get("mime_type", "image/png"),
                detail=inj.get("detail", "auto"),
            )
            count += 1
        logger.info("C 通道: 已注入 %d 张延迟图片到 memory", count)
        self._deferred_image_injections.clear()
        return count

    # 向后兼容别名（测试中可能直接调用）
    def _try_inject_image(self, result_str: str) -> str:
        """向后兼容：提取图片注入，返回清理后的 result_str。"""
        cleaned, _ = self._extract_structured_result(result_str)
        return cleaned

    def _extract_and_register_cow_mapping(self, result_str: str) -> dict[str, str] | None:
        """向后兼容：提取 cow_mapping。"""
        _, cow = self._extract_structured_result(result_str)
        return cow

    async def _run_vlm_describe(self) -> str | None:
        """B 通道：调用小 VLM 生成图片的 Markdown 描述。

        读取 _pending_vlm_image 中缓存的图片数据，调用 VLM，返回描述文本。
        调用后清除缓存。返回 None 表示失败或无待处理图片。
        """
        import base64

        from excelmanus.vision_extractor import build_describe_prompt

        injection = self._pending_vlm_image
        if injection is None:
            return None
        self._pending_vlm_image = None

        e = self._engine
        vlm_client = e.vlm_client
        vlm_model = e.vlm_model

        # 预处理图片（data 模式：增强文字可读性）
        raw_bytes = base64.b64decode(injection["base64"])
        compressed, mime = self._prepare_image_for_vlm(
            raw_bytes,
            max_long_edge=e.config.vlm_image_max_long_edge,
            jpeg_quality=e.config.vlm_image_jpeg_quality,
            mode="data",
        )
        b64 = base64.b64encode(compressed).decode("ascii")
        image_content = {
            "type": "image_url",
            "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"},
        }

        prompt = build_describe_prompt()
        messages = [
            {"role": "user", "content": [
                image_content,
                {"type": "text", "text": prompt},
            ]},
        ]

        raw_text, last_error, _fr = await self._call_vlm_with_retry(
            messages=messages,
            vlm_client=vlm_client,
            vlm_model=vlm_model,
            vlm_timeout=e.config.vlm_timeout_seconds,
            vlm_max_retries=e.config.vlm_max_retries,
            vlm_base_delay=e.config.vlm_retry_base_delay_seconds,
            phase_label="B通道描述",
        )

        if raw_text is None:
            sanitized = self._sanitize_vlm_error(last_error) if last_error else "未知错误"
            logger.warning("B 通道 VLM 描述失败: %s", sanitized)
            return None

        logger.info("B 通道 VLM 描述完成: %d 字符", len(raw_text))
        # 缓存 B 通道描述，供 Pipeline 结构阶段复用
        self._last_vlm_description = raw_text
        _b64_str = injection.get("base64", "")
        self._last_vlm_description_image_hash = (
            _image_content_hash_b64(_b64_str) if _b64_str else None
        )
        return raw_text

    async def _call_vlm_with_retry(
        self,
        *,
        messages: list[dict],
        vlm_client: Any,
        vlm_model: str,
        vlm_timeout: int,
        vlm_max_retries: int,
        vlm_base_delay: float,
        phase_label: str = "",
        response_format: dict | None = None,
        max_tokens: int | None = None,
    ) -> tuple[str | None, Exception | None, str | None]:
        """共享的 VLM 调用逻辑（带超时+网络错误重试）。

        返回 (raw_text, last_error, finish_reason)。raw_text 为 None 表示全部失败。
        """
        import asyncio

        raw_text: str | None = None
        last_error: Exception | None = None
        finish_reason: str | None = None
        label = f" [{phase_label}]" if phase_label else ""

        create_kwargs: dict[str, Any] = {
            "model": vlm_model,
            "messages": messages,
            "temperature": 0.0,
        }
        if response_format is not None:
            create_kwargs["response_format"] = response_format
        if max_tokens is not None:
            create_kwargs["max_tokens"] = max_tokens

        for attempt in range(vlm_max_retries + 1):
            try:
                response = await asyncio.wait_for(
                    vlm_client.chat.completions.create(**create_kwargs),
                    timeout=vlm_timeout,
                )
                raw_text = response.choices[0].message.content or ""
                finish_reason = getattr(response.choices[0], "finish_reason", None)
                if finish_reason == "length":
                    logger.warning(
                        "VLM%s 输出被截断（finish_reason=length），"
                        "输出长度 %d 字符，考虑增大 EXCELMANUS_VLM_MAX_TOKENS",
                        label, len(raw_text),
                    )
                break
            except asyncio.TimeoutError:
                last_error = TimeoutError(f"VLM 调用超时（{vlm_timeout}s）")
                logger.warning("VLM%s 超时（%ds），不重试", label, vlm_timeout)
                break
            except Exception as exc:
                last_error = exc
                sanitized = self._sanitize_vlm_error(exc)
                logger.warning(
                    "VLM%s 失败（attempt %d/%d）: %s",
                    label, attempt + 1, vlm_max_retries + 1, sanitized,
                )
                if attempt < vlm_max_retries:
                    delay = vlm_base_delay * (2 ** attempt)
                    await asyncio.sleep(delay)

        return raw_text, last_error, finish_reason

    @staticmethod
    def _prepare_image_for_vlm(
        raw: bytes,
        *,
        max_long_edge: int = 2048,
        jpeg_quality: int = 92,
        mode: str = "data",  # "data" | "style"
    ) -> tuple[bytes, str]:
        """自适应预处理图片以提升 VLM 表格识别质量。

        根据图片特征自动选择处理策略：
        1. 长边超限时等比缩放（保留文字细节）
        2. 灰色背景检测 → 白底替换（消除表格灰底干扰）
        3. 自适应对比度增强（低对比度图片加强，高对比度跳过）
        4. 扫描件/复印件自动二值化（基于直方图双峰检测）
        5. 智能锐化（仅对模糊图片应用，避免过度锐化）
        6. 转为高质量 JPEG
        - 返回 (processed_bytes, mime_type)
        """
        import io

        try:
            from PIL import Image, ImageFilter, ImageOps, ImageStat
        except ImportError:
            return raw, "image/png"

        try:
            img = Image.open(io.BytesIO(raw))
        except Exception:
            return raw, "image/png"

        # ── 1. 缩放（仅在超限时） ──
        w, h = img.size
        long_edge = max(w, h)
        if long_edge > max_long_edge:
            scale = max_long_edge / long_edge
            new_w, new_h = int(w * scale), int(h * scale)
            img = img.resize((new_w, new_h), Image.LANCZOS)

        # ── 2. 转 RGB ──
        if img.mode in ("RGBA", "LA", "P"):
            background = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            background.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        # ── style 模式：仅缩放+RGB转换，保留原始颜色 ──
        if mode == "style":
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
            compressed = buf.getvalue()
            if len(compressed) < len(raw):
                return compressed, "image/jpeg"
            return raw, "image/png"

        # ── 3. 分析图片特征 ──
        try:
            gray = img.convert("L")
            stat = ImageStat.Stat(gray)
            mean_brightness = stat.mean[0]  # 0-255
            stddev = stat.stddev[0]
            hist = gray.histogram()  # 256 bins
        except Exception:
            mean_brightness, stddev, hist = 128.0, 50.0, [0] * 256

        # ── 4. 灰色背景检测与去除 ──
        # 如果背景偏灰（中值亮度在 180-230 之间），将灰底白化
        try:
            if 180 <= mean_brightness <= 230 and stddev < 60:
                # 用灰度图判断哪些像素属于背景，生成 mask，
                # 再对 RGB 图统一白化，避免逐通道独立比较导致颜色失真
                thresh = int(mean_brightness - 20)
                bg_mask = gray.point(lambda p: 255 if p > thresh else 0, "1")
                white = Image.new("RGB", img.size, (255, 255, 255))
                img = Image.composite(white, img, bg_mask)
                logger.debug("图片预处理: 检测到灰色背景，已白化")
        except (ValueError, OSError, RuntimeError):
            logger.debug("图片预处理: 灰色背景白化失败", exc_info=True)

        # ── 5. 自适应对比度增强 ──
        try:
            if stddev < 40:
                # 低对比度：强力增强
                img = ImageOps.autocontrast(img, cutoff=1)
                logger.debug("图片预处理: 低对比度(stddev=%.1f)，强力增强", stddev)
            elif stddev < 70:
                # 中等对比度：适度增强
                img = ImageOps.autocontrast(img, cutoff=0.5)
            # stddev >= 70：高对比度图片，跳过对比度增强
        except (ValueError, OSError, RuntimeError):
            logger.debug("图片预处理: 对比度增强失败", exc_info=True)

        # ── 6. 扫描件二值化检测 ──
        # 通过直方图分析：如果亮度分布呈双峰（文字+背景），适用二值化
        try:
            if stddev > 30:
                # 计算直方图暗区(0-128)和亮区(128-255)的占比
                dark_ratio = sum(hist[:128]) / max(sum(hist), 1)
                light_ratio = sum(hist[128:]) / max(sum(hist), 1)
                # 双峰特征：暗区和亮区各占 10-90%
                is_bimodal = 0.05 < dark_ratio < 0.50 and 0.50 < light_ratio < 0.95
                # 对于扫描件（高对比度双峰），应用轻度阈值化增强
                if is_bimodal and stddev > 80:
                    gray_for_thresh = img.convert("L")
                    # 类 Otsu 简化：用均值作为阈值
                    threshold = int(mean_brightness * 0.85)
                    binary = gray_for_thresh.point(lambda p: 255 if p > threshold else 0, "L")
                    img = binary.convert("RGB")
                    logger.debug("图片预处理: 扫描件特征，已二值化(阈值=%d)", threshold)
        except (ValueError, OSError, RuntimeError):
            logger.debug("图片预处理: 扫描件二值化失败", exc_info=True)

        # ── 7. 智能锐化（仅对模糊图片） ──
        try:
            # 通过边缘检测评估清晰度
            edges = gray.filter(ImageFilter.FIND_EDGES)
            edge_stat = ImageStat.Stat(edges)
            edge_mean = edge_stat.mean[0]
            if edge_mean < 15:
                # 模糊图片：应用锐化
                img = img.filter(ImageFilter.SHARPEN)
                logger.debug("图片预处理: 模糊图片(edge_mean=%.1f)，已锐化", edge_mean)
            elif edge_mean < 30:
                # 中等清晰度：轻度锐化
                img = img.filter(ImageFilter.DETAIL)
        except (ValueError, OSError, RuntimeError):
            logger.debug("图片预处理: 智能锐化失败", exc_info=True)

        # ── 8. 输出 JPEG ──
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
        compressed = buf.getvalue()
        if len(compressed) < len(raw):
            return compressed, "image/jpeg"
        return raw, "image/png"

    @staticmethod
    def _sanitize_vlm_error(exc: Exception) -> str:
        """净化 VLM 错误消息：移除 HTML 响应体，提取关键信息。"""
        import re as _re
        msg = str(exc)
        if "<html" in msg.lower() or "<!doctype" in msg.lower():
            code_match = _re.search(r"(\d{3})[:\s]", msg)
            code = code_match.group(1) if code_match else "unknown"
            title_match = _re.search(r"<title[^>]*>(.*?)</title>", msg, _re.IGNORECASE)
            title = title_match.group(1).strip() if title_match else "Gateway error"
            return f"HTTP {code}: {title}"
        return msg[:500]

    @staticmethod
    def _build_vlm_failure_result(
        exc: Exception | None, attempts: int, file_path: str,
    ) -> str:
        """构建 VLM 失败时的结构化降级引导结果。"""
        error_msg = ToolDispatcher._sanitize_vlm_error(exc) if exc else "未知错误"
        return json.dumps({
            "status": "error",
            "error_code": "VLM_CALL_FAILED",
            "message": f"VLM 提取在 {attempts} 次尝试后失败: {error_msg}",
            "fallback_hint": (
                "建议降级方案：1) 使用 read_image 查看图片，由主模型直接描述表格内容；"
                "2) 用 run_code + openpyxl 根据描述手动构建 Excel 文件。"
                "VLM 上游 API 可能暂时不可用。"
            ),
            "file_path": file_path,
        }, ensure_ascii=False)

    def _redirect_cow_paths(
        self,
        tool_name: str,
        arguments: dict[str, Any],
    ) -> tuple[dict[str, Any], list[str]]:
        """检查工具参数中的文件路径是否命中 CoW 注册表，自动重定向。

        返回 (可能修改过的 arguments, 重定向提醒消息列表)。
        """
        from excelmanus.tools.policy import (
            AUDIT_TARGET_ARG_RULES_ALL,
            AUDIT_TARGET_ARG_RULES_FIRST,
            READ_ONLY_SAFE_TOOLS,
        )

        # 仅从 FileRegistry 查询 CoW 重定向
        _file_reg = self._engine.file_registry
        _use_file_reg = bool(
            _file_reg is not None
            and getattr(_file_reg, "has_versions", False)
            and hasattr(_file_reg, "lookup_cow_redirect")
        )
        if not _use_file_reg:
            return arguments, []

        path_fields: list[str] = []
        all_fields = AUDIT_TARGET_ARG_RULES_ALL.get(tool_name)
        if all_fields is not None:
            path_fields.extend(all_fields)
        else:
            first_fields = AUDIT_TARGET_ARG_RULES_FIRST.get(tool_name)
            if first_fields is not None:
                path_fields.extend(first_fields)

        if tool_name in READ_ONLY_SAFE_TOOLS:
            for key in ("file_path", "path", "directory"):
                if key in arguments and key not in path_fields:
                    path_fields.append(key)

        # run_code 的 code 参数中的路径由沙盒层 sandbox_hook 处理，此处不拦截
        if not path_fields:
            return arguments, []

        workspace_root = self._engine.config.workspace_root
        redirected = dict(arguments)
        reminders: list[str] = []
        for field_name in path_fields:
            raw = arguments.get(field_name)
            if raw is None:
                continue
            raw_str = str(raw).strip()
            if not raw_str:
                continue
            # 尝试匹配：直接匹配相对路径，或去掉 workspace_root 前缀后匹配
            rel_path = raw_str
            if workspace_root and raw_str.startswith(workspace_root):
                rel_path = raw_str[len(workspace_root):].lstrip("/")
            redirect = _file_reg.lookup_cow_redirect(rel_path)
            if isinstance(redirect, str):
                # 保持原始路径格式（绝对/相对）
                if raw_str.startswith(workspace_root):
                    new_path = f"{workspace_root}/{redirect}"
                else:
                    new_path = redirect
                redirected[field_name] = new_path
                reminders.append(
                    f"⚠️ 路径 `{raw_str}` 是受保护的原始文件，"
                    f"已自动重定向到副本 `{new_path}`。"
                    f"请在后续操作中直接使用副本路径。"
                )
                logger.info(
                    "CoW 路径拦截: tool=%s field=%s %s → %s",
                    tool_name, field_name, raw_str, new_path,
                )
        return redirected, reminders

    def parse_arguments(self, raw_args: Any) -> tuple[dict[str, Any], str | None]:
        """解析工具调用参数，返回 (arguments, error)。

        error 为 None 表示解析成功。
        """
        if raw_args is None or raw_args == "":
            return {}, None
        if isinstance(raw_args, dict):
            return raw_args, None
        if isinstance(raw_args, str):
            try:
                parsed = json.loads(raw_args)
                if not isinstance(parsed, dict):
                    return {}, f"参数必须为 JSON 对象，当前类型: {type(parsed).__name__}"
                return parsed, None
            except (json.JSONDecodeError, TypeError) as exc:
                return {}, f"JSON 解析失败: {exc}"
        return {}, f"参数类型无效: {type(raw_args).__name__}"

    def cancel_active_sleep(self) -> None:
        """中断当前会话正在执行的 sleep 工具调用。"""
        self._sleep_cancel_event.set()

    async def call_registry_tool(
        self,
        *,
        tool_name: str,
        arguments: dict[str, Any],
        tool_scope: Sequence[str] | None = None,
    ) -> str:
        """调用工具，返回截断后的结果字符串。

        MCP 工具（具有 async_func）直接 await，避免线程池 + asyncio.run 开销。
        普通工具仍走 asyncio.to_thread 线程池路径。
        """
        from excelmanus.tools import memory_tools
        from excelmanus.tools.sleep_tools import set_cancel_event, reset_cancel_event

        registry = self._registry

        # 检测是否有异步快速路径（MCP 工具）
        tool_def = registry.get_tool(tool_name)
        _has_async = (
            tool_def is not None
            and getattr(tool_def, "async_func", None) is not None
            and callable(tool_def.async_func)
            and asyncio.iscoroutinefunction(tool_def.async_func)
        )
        if _has_async:
            # MCP 异步快速路径：直接 await，不经线程池
            result_value = await registry.call_tool_async(
                tool_name,
                arguments,
                tool_scope=tool_scope,
            )
        else:
            # 普通工具：走线程池路径
            persistent_memory = self._persistent_memory
            sleep_cancel_event = self._sleep_cancel_event

            # 将每会话的 sleep 取消事件注入 contextvar，
            # asyncio.to_thread 会自动拷贝到工作线程。
            _sleep_token = set_cancel_event(sleep_cancel_event)

            def _call() -> Any:
                with memory_tools.bind_memory_context(persistent_memory):
                    return registry.call_tool(
                        tool_name,
                        arguments,
                        tool_scope=tool_scope,
                    )

            try:
                result_value = await asyncio.to_thread(_call)
            finally:
                reset_cancel_event(_sleep_token)

        result_str = str(result_value)

        # 先处理图片注入（移除 base64 载荷），再做截断，
        # 避免截断破坏 JSON 导致注入失败。
        if result_str:
            result_str = self._try_inject_image(result_str)

        # 保存截断前的原始结果，供窗口感知解析使用
        self._last_call_raw_result: str = result_str

        # 工具结果截断
        tool_def = getattr(registry, "get_tool", lambda _: None)(tool_name)
        if tool_def is not None:
            result_str = tool_def.truncate_result(result_str)

        return result_str

    # ── 核心执行方法：从 AgentEngine._execute_tool_call 搬迁 ──

    async def execute(
        self,
        tc: Any,
        tool_scope: Sequence[str] | None,
        on_event: "EventCallback | None",
        iteration: int,
        route_result: Any | None = None,
        skip_start_event: bool = False,
    ) -> Any:
        """单个工具调用：参数解析 → 执行 → 事件发射 → 返回结果。

        从 AgentEngine._execute_tool_call 整体搬迁，通过 self._engine
        引用回调 AgentEngine 上的基础设施方法。
        """
        from excelmanus.tools.code_tools import set_sandbox_env as _set_sandbox_env
        from excelmanus.tools._guard_ctx import set_guard as _set_guard, reset_guard as _reset_guard

        e = self._engine  # 引擎快捷引用

        # 注入每会话的沙盒环境和 FileAccessGuard 到 contextvars。
        _sandbox_token = _set_sandbox_env(e.sandbox_env)
        _guard_token = _set_guard(e.file_access_guard)
        try:
            return await self._execute_inner(
                tc, tool_scope, on_event, iteration, route_result, skip_start_event,
                _sandbox_token,
            )
        finally:
            from excelmanus.tools.code_tools import _current_sandbox_env
            _current_sandbox_env.reset(_sandbox_token)
            _reset_guard(_guard_token)

    async def _execute_inner(
        self,
        tc: Any,
        tool_scope: Sequence[str] | None,
        on_event: "EventCallback | None",
        iteration: int,
        route_result: Any | None,
        skip_start_event: bool,
        _sandbox_token: Any,
    ) -> Any:
        from excelmanus.engine import ToolCallResult
        from excelmanus.events import EventType, ToolCallEvent

        e = self._engine
        _t0 = time.monotonic()

        function = getattr(tc, "function", None)
        tool_name = getattr(function, "name", "")
        raw_args = getattr(function, "arguments", None)
        tool_call_id = getattr(tc, "id", "") or f"call_{int(time.time() * 1000)}"

        # 参数解析
        arguments, parse_error = self.parse_arguments(raw_args)

        # 发射 TOOL_CALL_START 事件（并行路径已预发射，跳过避免重复）
        if not skip_start_event:
            e.emit(
                on_event,
                ToolCallEvent(
                    event_type=EventType.TOOL_CALL_START,
                    tool_call_id=tool_call_id,
                    tool_name=tool_name,
                    arguments=arguments,
                    iteration=iteration,
                ),
            )

        # /tools 开启时额外发射简要工具调用通知
        if getattr(e, "_show_tool_calls", False):
            e.emit(
                on_event,
                ToolCallEvent(
                    event_type=EventType.TOOL_CALL_NOTICE,
                    tool_call_id=tool_call_id,
                    tool_name=tool_name,
                    arguments=arguments,
                    iteration=iteration,
                ),
            )

        pending_approval = False
        approval_id: str | None = None
        audit_record = None
        pending_question = False
        question_id: str | None = None
        defer_tool_result = False
        finish_accepted = False
        error_kind: str | None = None
        _cow_reminders: list[str] = []
        _raw_result_str: str | None = None

        # 执行工具调用
        hook_skill = e.pick_route_skill(route_result)
        if parse_error is not None:
            result_str = f"工具参数解析错误: {parse_error}"
            success = False
            error = result_str
            log_tool_call(
                logger,
                tool_name,
                {"_raw_arguments": raw_args},
                error=error,
            )
        else:
            # ── 备份沙盒模式：重定向文件路径 ──
            arguments = e.redirect_backup_paths(tool_name, arguments)

            # ── CoW 路径拦截：将原始保护路径重定向到 outputs/ 副本 ──
            arguments, _cow_reminders = self._redirect_cow_paths(tool_name, arguments)

            pre_hook_raw = e.run_skill_hook(
                skill=hook_skill,
                event=HookEvent.PRE_TOOL_USE,
                payload={
                    "tool_name": tool_name,
                    "arguments": dict(arguments),
                    "iteration": iteration,
                },
                tool_name=tool_name,
            )
            pre_hook = await e.resolve_hook_result(
                event=HookEvent.PRE_TOOL_USE,
                hook_result=pre_hook_raw,
                on_event=on_event,
            )
            if pre_hook is not None and isinstance(pre_hook.updated_input, dict):
                arguments = dict(pre_hook.updated_input)
            skip_high_risk_approval_by_hook = (
                pre_hook is not None and pre_hook.decision == HookDecision.ALLOW
            )
            if skip_high_risk_approval_by_hook:
                logger.info(
                    "Hook ALLOW 已生效，跳过确认门禁：tool=%s iteration=%s",
                    tool_name,
                    iteration,
                )

            if pre_hook is not None and pre_hook.decision == HookDecision.DENY:
                reason = pre_hook.reason or "Hook 拒绝执行该工具。"
                result_str = f"工具调用被 Hook 拒绝：{reason}"
                success = False
                error = result_str
                log_tool_call(logger, tool_name, arguments, error=error)
            elif pre_hook is not None and pre_hook.decision == HookDecision.ASK:
                try:
                    pending = e.approval.create_pending(
                        tool_name=tool_name,
                        arguments=arguments,
                        tool_scope=tool_scope,
                    )
                    pending_approval = True
                    approval_id = pending.approval_id
                    result_str = e.format_pending_prompt(pending)
                    success = True
                    error = None
                    e.emit_pending_approval_event(
                        pending=pending, on_event=on_event, iteration=iteration,
                        tool_call_id=tool_call_id,
                    )
                    log_tool_call(logger, tool_name, arguments, result=result_str)
                except ValueError:
                    result_str = e.approval.pending_block_message()
                    success = False
                    error = result_str
                    log_tool_call(logger, tool_name, arguments, error=error)
            else:
                outcome = await self._dispatch_via_handlers(
                    tool_name=tool_name,
                    tool_call_id=tool_call_id,
                    arguments=arguments,
                    tool_scope=tool_scope,
                    on_event=on_event,
                    iteration=iteration,
                    route_result=route_result,
                    skip_high_risk_approval_by_hook=skip_high_risk_approval_by_hook,
                )
                result_str = outcome.result_str
                success = outcome.success
                error = outcome.error
                error_kind = outcome.error_kind
                pending_approval = outcome.pending_approval
                approval_id = outcome.approval_id
                audit_record = outcome.audit_record
                pending_question = outcome.pending_question
                question_id = outcome.question_id
                defer_tool_result = outcome.defer_tool_result
                finish_accepted = outcome.finish_accepted
                _raw_result_str = outcome.raw_result_str

            # ── 检测 registry 层返回的结构化错误 JSON ──
            if success and e.registry.is_error_result(result_str):
                success = False
                try:
                    _err = json.loads(result_str)
                    error = _err.get("message") or _err.get("error") or result_str
                except Exception:
                    error = result_str

            post_hook_event = HookEvent.POST_TOOL_USE if success else HookEvent.POST_TOOL_USE_FAILURE
            post_hook_raw = e.run_skill_hook(
                skill=hook_skill,
                event=post_hook_event,
                payload={
                    "tool_name": tool_name,
                    "arguments": dict(arguments),
                    "success": success,
                    "result": result_str,
                    "error": error,
                    "iteration": iteration,
                },
                tool_name=tool_name,
            )
            post_hook = await e.resolve_hook_result(
                event=post_hook_event,
                hook_result=post_hook_raw,
                on_event=on_event,
            )
            if post_hook is not None:
                if post_hook.additional_context:
                    result_str = f"{result_str}\n[Hook] {post_hook.additional_context}"
                if post_hook.decision == HookDecision.DENY:
                    reason = post_hook.reason or "post hook 拒绝"
                    success = False
                    error = reason
                    result_str = f"{result_str}\n[Hook 拒绝] {reason}"

        # ── 后处理流水线 ──
        result_str, success, error = await self._postprocess_result(
            tool_name=tool_name,
            tool_call_id=tool_call_id,
            arguments=arguments,
            result_str=result_str,
            success=success,
            error=error,
            iteration=iteration,
            on_event=on_event,
            cow_reminders=_cow_reminders,
            start_time=_t0,
            raw_result_str=_raw_result_str,
            error_kind=error_kind,
        )

        return ToolCallResult(
            tool_name=tool_name,
            arguments=arguments,
            result=result_str,
            success=success,
            error=error,
            error_kind=error_kind,
            pending_approval=pending_approval,
            approval_id=approval_id,
            audit_record=audit_record,
            pending_question=pending_question,
            question_id=question_id,
            defer_tool_result=defer_tool_result,
            finish_accepted=finish_accepted,
        )

    async def _dispatch_via_handlers(
        self,
        tool_name: str,
        tool_call_id: str,
        arguments: dict[str, Any],
        *,
        tool_scope: Sequence[str] | None = None,
        on_event: "EventCallback | None" = None,
        iteration: int = 0,
        route_result: Any = None,
        skip_high_risk_approval_by_hook: bool = False,
    ) -> "_ToolExecOutcome":
        """通过策略处理器表分发工具执行。

        先查 _specific_handlers O(1) 索引，未命中则遍历 _generic_handlers。
        对 RETRYABLE 错误自动重试（指数退避，不消耗 Agent 迭代预算）。
        """
        policy = DEFAULT_RETRY_POLICY
        last_outcome: _ToolExecOutcome | None = None

        for attempt in range(policy.max_retries + 1):
            outcome = await self._dispatch_single_attempt(
                tool_name=tool_name,
                tool_call_id=tool_call_id,
                arguments=arguments,
                tool_scope=tool_scope,
                on_event=on_event,
                iteration=iteration,
                route_result=route_result,
                skip_high_risk_approval_by_hook=skip_high_risk_approval_by_hook,
            )
            # 成功或非错误结果直接返回
            if outcome.success:
                return outcome

            # 对失败结果做错误分类
            tool_error = classify_tool_error(
                outcome.error or outcome.result_str,
                tool_name=tool_name,
            )

            # 非可重试错误：压缩后直接返回
            if not tool_error.retryable:
                compacted = compact_error(outcome.error, tool_error=tool_error)
                return _ToolExecOutcome(
                    result_str=compacted,
                    success=False,
                    error=compacted,
                    error_kind=tool_error.kind.value,
                    audit_record=outcome.audit_record,
                )

            last_outcome = outcome
            # 最后一次重试也失败了
            if attempt >= policy.max_retries:
                break

            # 可重试：等待后重试
            delay = policy.delay_for_attempt(attempt)
            logger.info(
                "工具 %s 可重试错误（%s），第 %d/%d 次重试，等待 %.1fs",
                tool_name, tool_error.summary[:80],
                attempt + 1, policy.max_retries, delay,
            )
            await asyncio.sleep(delay)

        # 所有重试均失败
        final_error = classify_tool_error(
            last_outcome.error or last_outcome.result_str if last_outcome else "unknown",
            tool_name=tool_name,
        )
        compacted = compact_error(
            last_outcome.error if last_outcome else "unknown",
            tool_error=final_error,
        )
        retried_msg = f"{compacted}\n[已自动重试 {policy.max_retries} 次仍失败]"
        return _ToolExecOutcome(
            result_str=retried_msg,
            success=False,
            error=retried_msg,
            error_kind=final_error.kind.value,
            audit_record=last_outcome.audit_record if last_outcome else None,
        )

    async def _dispatch_single_attempt(
        self,
        tool_name: str,
        tool_call_id: str,
        arguments: dict[str, Any],
        *,
        tool_scope: Sequence[str] | None = None,
        on_event: "EventCallback | None" = None,
        iteration: int = 0,
        route_result: Any = None,
        skip_high_risk_approval_by_hook: bool = False,
    ) -> "_ToolExecOutcome":
        """单次工具执行尝试（从 _dispatch_via_handlers 提取）。"""
        from excelmanus.engine import _AuditedExecutionError

        try:
            # T2: O(1) 索引查找特定工具 handler，未命中时走动态/兜底链
            handler = self._specific_handlers.get(tool_name)
            if handler is None:
                for candidate in self._generic_handlers:
                    if candidate.can_handle(tool_name):
                        handler = candidate
                        break
                else:
                    raise RuntimeError(f"No handler found for tool: {tool_name}")

            handler_kwargs: dict[str, Any] = {
                "tool_scope": tool_scope,
                "on_event": on_event,
                "iteration": iteration,
                "route_result": route_result,
            }
            if handler.__class__.__name__ == "HighRiskApprovalHandler":
                handler_kwargs["skip_high_risk_approval_by_hook"] = skip_high_risk_approval_by_hook

            return await handler.handle(
                tool_name,
                tool_call_id,
                arguments,
                **handler_kwargs,
            )
        except ValueError as exc:
            result_str = str(exc)
            log_tool_call(logger, tool_name, arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)
        except ToolNotAllowedError:
            permission_error = {
                "error_code": "TOOL_NOT_ALLOWED",
                "tool": tool_name,
                "message": f"工具 '{tool_name}' 不在当前授权范围内。",
            }
            result_str = json.dumps(permission_error, ensure_ascii=False)
            log_tool_call(logger, tool_name, arguments, error=result_str)
            return _ToolExecOutcome(result_str=result_str, success=False, error=result_str)
        except Exception as exc:
            root_exc: Exception = exc
            audit_record = None
            if isinstance(exc, _AuditedExecutionError):
                audit_record = exc.record
                root_exc = exc.cause
            result_str = f"工具执行错误: {root_exc}"
            log_tool_call(logger, tool_name, arguments, error=str(root_exc))
            return _ToolExecOutcome(
                result_str=result_str,
                success=False,
                error=str(root_exc),
                audit_record=audit_record,
            )

    async def _postprocess_result(
        self,
        *,
        tool_name: str,
        tool_call_id: str,
        arguments: dict[str, Any],
        result_str: str,
        success: bool,
        error: str | None,
        iteration: int,
        on_event: "EventCallback | None",
        cow_reminders: list[str],
        start_time: float = 0.0,
        raw_result_str: str | None = None,
        error_kind: str | None = None,
    ) -> tuple[str, bool, str | None]:
        """后处理流水线：CoW/备份/图片/VLM/窗口感知/硬截断/事件/审计/任务清单。

        返回 (result_str, success, error)，其中 success/error 可能被
        结构化错误检测修改。
        """
        from excelmanus.events import EventType, ToolCallEvent

        e = self._engine

        # ── 保留原始 JSON 结果用于 Excel 事件提取 ──
        # 后续 enrichment 步骤会在 result_str 上追加非 JSON 文本（CoW 提醒、
        # 备份通知、VLM 描述、窗口感知等），导致 json.loads 失败。
        # 必须在 enrichment 之前保存原始结果供 _emit_excel_events 使用。
        _raw_result_for_excel_events = result_str

        # ── 通用结构化字段提取（CoW 映射 + 图片注入，单次 JSON 解析） ──
        if success and result_str:
            result_str, _cow_extracted = self._extract_structured_result(result_str)
            if _cow_extracted:
                logger.info(
                    "CoW 映射已注册: tool=%s mappings=%s", tool_name, _cow_extracted,
                )

        # ── CoW 路径拦截提醒：追加到工具结果中 ──
        if cow_reminders:
            result_str = result_str + "\n" + "\n".join(cow_reminders)

        # ── 备份沙盒提醒：首次写入成功后追加备份文件路径 ──
        tx = e.transaction
        if (
            success
            and e.workspace.transaction_enabled
            and tx is not None
            and not e.state.backup_write_notice_shown
        ):
            from pathlib import Path as _Path

            from excelmanus.tools.policy import READ_ONLY_SAFE_TOOLS as _RO_TOOLS

            if tool_name not in _RO_TOOLS:
                backups = tx.list_staged()
                if backups:
                    backup_dir = str(tx.staging_dir)
                    file_names = [
                        _Path(b["backup"]).name
                        for b in backups
                        if b.get("exists") == "True"
                    ]
                    files_str = "、".join(file_names) if file_names else ""
                    notice_parts = [
                        f"\n[备份提示] 修改已保存到备份副本目录 `{backup_dir}/`",
                    ]
                    if files_str:
                        notice_parts.append(f"（当前备份文件：{files_str}）")
                    notice_parts.append(
                        "。请在回复中告知用户备份文件位置，"
                        "用户可通过 `/backup apply` 将修改应用到原文件。"
                    )
                    result_str = result_str + "".join(notice_parts)
                    e.state.backup_write_notice_shown = True

        # ── Post-Write Inline Checkpoint（零 LLM 调用回读验证）──
        if success and tool_name in self._EXCEL_WRITE_TOOLS:
            _ws_root = getattr(getattr(e, "_config", None), "workspace_root", "")
            _ckpt = self._post_write_checkpoint(tool_name, arguments, _ws_root)
            if _ckpt:
                result_str = result_str + _ckpt

        # ── B 通道：异步 VLM 描述追加 ──
        # 当主模型有视觉能力时跳过（C 通道已直接注入图片，无需额外 VLM 描述）
        if success and self._pending_vlm_image is not None and not e.is_vision_capable:
            e.emit(
                on_event,
                ToolCallEvent(
                    event_type=EventType.PIPELINE_PROGRESS,
                    tool_call_id=tool_call_id,
                    pipeline_stage="vlm_describe",
                    pipeline_message="正在调用 VLM 生成图片描述...",
                ),
            )
            vlm_desc = await self._run_vlm_describe()
            if vlm_desc:
                result_str = (
                    result_str
                    + "\n\n--- VLM 增强描述（B 通道） ---\n"
                    + vlm_desc
                )
                logger.info("B 通道描述已追加到 tool result")
            else:
                result_str = (
                    result_str
                    + "\n\n[VLM 增强描述失败，请直接基于图片或已有信息操作]"
                )

        result_str = e._enrich_tool_result_with_window_perception(
            tool_name=tool_name,
            arguments=arguments,
            result_text=result_str,
            success=success,
            raw_result_text=raw_result_str,
        )
        result_str = e._apply_tool_result_hard_cap(result_str)
        if error:
            error = e._apply_tool_result_hard_cap(str(error))

        # 发射 TOOL_CALL_END 事件
        e.emit(
            on_event,
            ToolCallEvent(
                event_type=EventType.TOOL_CALL_END,
                tool_call_id=tool_call_id,
                tool_name=tool_name,
                arguments=arguments,
                result=result_str,
                success=success,
                error=error,
                iteration=iteration,
            ),
        )

        # 工具执行失败时发出结构化失败引导（仅限基础设施级错误）
        if not success and error:
            _err_lower = str(error).lower()
            _emit_tool_guidance = any(kw in _err_lower for kw in (
                "permission denied", "errno 13",
                "no space left", "disk full", "errno 28",
                "codec can't decode", "codec can't encode",
                "unicodedecodeerror", "unicodeencodeerror",
                "invalid start byte", "invalid continuation byte",
            ))
            if _emit_tool_guidance:
                try:
                    from excelmanus.error_guidance import classify_failure as _clf
                    _tool_guidance = _clf(
                        RuntimeError(str(error)),
                        stage="tool_execution",
                        provider="",
                        model="",
                    )
                    e.emit(
                        on_event,
                        ToolCallEvent(
                            event_type=EventType.FAILURE_GUIDANCE,
                            fg_category=_tool_guidance.category,
                            fg_code=_tool_guidance.code,
                            fg_title=_tool_guidance.title,
                            fg_message=_tool_guidance.message,
                            fg_stage=_tool_guidance.stage,
                            fg_retryable=_tool_guidance.retryable,
                            fg_diagnostic_id=_tool_guidance.diagnostic_id,
                            fg_actions=_tool_guidance.actions,
                            fg_provider=_tool_guidance.provider,
                            fg_model=_tool_guidance.model,
                        ),
                    )
                except Exception:
                    pass

        # 工具调用审计日志
        if self._tool_call_store is not None:
            try:
                _session_id = getattr(e, "_session_id", None)
                self._tool_call_store.log(
                    session_id=_session_id,
                    turn=e.state.session_turn,
                    iteration=iteration,
                    tool_name=tool_name,
                    arguments_hash=e.state._args_fingerprint(arguments) if arguments else None,
                    success=success,
                    duration_ms=(time.monotonic() - start_time) * 1000 if start_time else 0.0,
                    result_chars=len(result_str) if result_str else 0,
                    error_type=error_kind if error_kind else (error[:50] if error else None),
                    error_preview=str(error)[:200] if error else None,
                )
            except Exception:
                pass

        # Excel 预览/Diff 事件（使用 enrichment 之前的原始结果，确保 JSON 可解析）
        if success and _raw_result_for_excel_events:
            self._emit_excel_events(
                e, on_event, tool_call_id, tool_name, arguments,
                _raw_result_for_excel_events, iteration,
            )

        # 写入类工具 → files_changed 事件（补充 _excel_diff / _text_diff 未覆盖的场景）
        if success and on_event is not None:
            _emit_fc = False
            _fc_files: list[str] = []
            if tool_name in self._EXCEL_WRITE_TOOLS:
                _fp = (arguments.get("file_path") or "").strip()
                if _fp:
                    _fc_files.append(_fp)
                    _emit_fc = True
            elif e.get_tool_write_effect(tool_name) == "workspace_write":
                for _pk_fc in ("file_path", "output_path", "path", "target_path"):
                    _pv_fc = (arguments.get(_pk_fc) or "").strip()
                    if _pv_fc:
                        _fc_files.append(_pv_fc)
                        _emit_fc = True
            if _emit_fc and _fc_files:
                from excelmanus.events import EventType, ToolCallEvent
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.FILES_CHANGED,
                        tool_call_id=tool_call_id,
                        iteration=iteration,
                        changed_files=_fc_files,
                    ),
                )

        # ── 自动追踪 affected_files + write_operations_log ──
        if success:
            _state = getattr(e, "_state", None)
            if _state is not None:
                if tool_name in self._EXCEL_WRITE_TOOLS:
                    _afp = (arguments.get("file_path") or "").strip()
                    if _afp:
                        _state.record_affected_file(_afp)
                    # 写入操作日志（供 verifier delta 注入）
                    _state.record_write_operation(
                        tool_name=tool_name,
                        file_path=_afp,
                        sheet=(arguments.get("sheet") or "").strip(),
                        cell_range=(arguments.get("range") or "").strip(),
                        summary=self._extract_write_summary(tool_name, arguments, result_str),
                    )
                elif tool_name == "run_code" and _raw_result_for_excel_events:
                    try:
                        import json as _json
                        _parsed = _json.loads(_raw_result_for_excel_events.strip())
                        if isinstance(_parsed, dict):
                            _cow = _parsed.get("cow_mapping")
                            _cow_paths = ""
                            if isinstance(_cow, dict):
                                for _v in _cow.values():
                                    if isinstance(_v, str) and _v.strip():
                                        _state.record_affected_file(_v)
                                _cow_paths = ", ".join(
                                    str(v) for v in _cow.values() if isinstance(v, str) and v.strip()
                                )
                            # 即使无 cow_mapping，只要 has_write_tool_call 已被标记
                            # （由 CodePolicyHandler 或 legacy 路径设置），也应记录
                            if _cow_paths or _state.has_write_tool_call:
                                # 避免与 CodePolicyHandler 重复记录
                                _already_logged = any(
                                    e.get("tool_name") == "run_code"
                                    for e in _state.write_operations_log
                                )
                                if not _already_logged:
                                    _state.record_write_operation(
                                        tool_name="run_code",
                                        file_path=_cow_paths,
                                        summary=self._extract_run_code_write_summary(result_str),
                                    )
                    except Exception:
                        pass
                elif e.get_tool_write_effect(tool_name) == "workspace_write":
                    for _pk in ("file_path", "output_path", "path", "target_path",
                                "source", "destination"):
                        _pv = (arguments.get(_pk) or "").strip()
                        if _pv:
                            _state.record_affected_file(_pv)
                    # 通用写入工具日志
                    _first_path = next(
                        ((arguments.get(k) or "").strip() for k in ("file_path", "output_path", "path", "target_path",
                                                                     "source", "destination")
                         if (arguments.get(k) or "").strip()),
                        "",
                    )
                    _state.record_write_operation(
                        tool_name=tool_name,
                        file_path=_first_path,
                    )

        # ── ErrorSolutionStore：记录错误/解决方案 + 检索历史方案 ──
        # 配对策略：用 tool_name 作为 pending key（而非 tool_call_id），
        # 因为同一工具的下一次成功调用自然对应上一次失败的解决方案。
        _error_store = getattr(e, "_error_solution_store", None)
        if _error_store is not None:
            try:
                if not success and error:
                    # 记录工具执行错误（等待后续同名工具成功时配对）
                    await _error_store.record_error(tool_name, tool_name, error)
                    # 语义检索历史类似错误的解决方案，追加到错误结果中
                    _guidance = await _error_store.get_guidance_text(error)
                    if _guidance:
                        result_str = result_str + "\n\n" + _guidance
                elif success and tool_name:
                    # 工具成功 → 尝试配对之前同名工具的错误，记录解决方案
                    _solution_text = self._extract_write_summary(tool_name, arguments, result_str)
                    if _solution_text:
                        await _error_store.record_solution(
                            tool_name, tool_name, _solution_text, success=True,
                        )
            except Exception:
                logger.debug("ErrorSolutionStore 操作失败", exc_info=True)

        # 写后事件记录到 FileRegistry
        if success:
            _freg = e.file_registry
            if _freg is not None:
                try:
                    # rename_file 特殊处理：原子迁移路径，保留 file_id / provenance
                    if tool_name == "rename_file":
                        _src = (arguments.get("source") or "").strip()
                        _dst = (arguments.get("destination") or "").strip()
                        if _src and _dst:
                            _freg.rename_entry(
                                _src, _dst,
                                session_id=getattr(e, "session_id", None),
                                turn=e.state.session_turn,
                            )

                    _write_paths: list[str] = []
                    if tool_name in self._EXCEL_WRITE_TOOLS:
                        _wp = (arguments.get("file_path") or "").strip()
                        if _wp:
                            _write_paths.append(_wp)
                    elif e.get_tool_write_effect(tool_name) == "workspace_write":
                        for _pk2 in ("file_path", "output_path", "path", "target_path",
                                     "source", "destination"):
                            _pv2 = (arguments.get(_pk2) or "").strip()
                            if _pv2:
                                _write_paths.append(_pv2)
                    for _wpath in _write_paths:
                        _entry = _freg.get_by_path(_wpath)
                        if _entry is not None:
                            _freg.record_event(
                                _entry.id,
                                "tool_write",
                                tool_name=tool_name,
                                turn=e.state.session_turn,
                            )
                except Exception:
                    logger.debug("FileRegistry 写后事件记录失败", exc_info=True)

        # 任务清单事件：成功执行 task_create/task_update/write_plan 后发射对应事件
        if success and tool_name == "write_plan":
            task_list = e._task_store.current
            if task_list is not None:
                plan_path = e._task_store.plan_file_path or ""
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.PLAN_CREATED,
                        plan_file_path=plan_path,
                        plan_title=task_list.title,
                        plan_task_count=len(task_list.items),
                    ),
                )
                # 同时发射 TASK_LIST_CREATED 以复用前端任务清单渲染
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.TASK_LIST_CREATED,
                        task_list_data=task_list.to_dict(),
                    ),
                )
            # write_plan 返回纯文本（非 JSON），_emit_excel_events 无法检测 _text_diff，
            # 因此在此处直接计算 diff 并发射 TEXT_DIFF 事件。
            _plan_content = (arguments.get("content") or "").strip()
            _plan_path = e._task_store.plan_file_path or ""
            if _plan_content and _plan_path and on_event is not None:
                from excelmanus.tools.code_tools import _generate_text_diff
                _td = _generate_text_diff("", _plan_content, _plan_path)
                if _td is not None:
                    e.emit(
                        on_event,
                        ToolCallEvent(
                            event_type=EventType.TEXT_DIFF,
                            tool_call_id=tool_call_id,
                            text_diff_file_path=_td.get("file_path", ""),
                            text_diff_hunks=_td.get("hunks", [])[:300],
                            text_diff_additions=_td.get("additions", 0),
                            text_diff_deletions=_td.get("deletions", 0),
                            text_diff_truncated=_td.get("truncated", False),
                        ),
                    )
        elif success and tool_name == "task_create":
            task_list = e._task_store.current
            if task_list is not None:
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.TASK_LIST_CREATED,
                        task_list_data=task_list.to_dict(),
                    ),
                )
        elif success and tool_name == "task_update":
            task_list = e._task_store.current
            if task_list is not None:
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.TASK_ITEM_UPDATED,
                        task_index=arguments.get("task_index"),
                        task_status=arguments.get("status", ""),
                        task_result=arguments.get("result"),
                        task_list_data=task_list.to_dict(),
                    ),
                )

        return result_str, success, error

    # ── uploads 目录快照（检测 run_code 新建/变更文件）────────

    @staticmethod
    def _snapshot_uploads_dir(workspace_root: str) -> dict[str, float] | None:
        """对 uploads/ 目录做轻量 mtime 快照，返回 {rel_path: mtime}。"""
        import os
        from pathlib import Path as _P
        uploads = _P(workspace_root) / "uploads"
        if not uploads.is_dir():
            return None
        snap: dict[str, float] = {}
        try:
            for root, _dirs, files in os.walk(uploads):
                _dirs[:] = [d for d in _dirs if not d.startswith(".")]
                for fname in files:
                    if fname.startswith("."):
                        continue
                    full = os.path.join(root, fname)
                    try:
                        snap[os.path.relpath(full, uploads)] = os.path.getmtime(full)
                    except OSError:
                        continue
        except OSError:
            return None
        return snap

    @staticmethod
    def _diff_uploads_snapshots(
        before: dict[str, float] | None,
        after: dict[str, float] | None,
    ) -> list[str]:
        """对比 uploads 快照，返回新建或修改的相对路径列表。"""
        if before is None or after is None:
            return []
        changed: list[str] = []
        for rel_path, mtime in after.items():
            if rel_path not in before or before[rel_path] != mtime:
                changed.append(rel_path)
        return changed

    # ── Post-Write Inline Checkpoint ────────────────────────

    @staticmethod
    def _post_write_checkpoint(
        tool_name: str,
        arguments: dict,
        workspace_root: str,
    ) -> str:
        """写入工具成功后执行轻量级回读验证（零 LLM 调用）。

        返回一行简洁的 checkpoint 结果，如 "✓ 回读确认: Sheet1, 100行×3列"。
        任何异常静默返回空字符串（不影响主流程）。
        """
        from pathlib import Path as _P

        file_path = (arguments.get("file_path") or "").strip()
        if not file_path:
            return ""

        abs_path = _P(file_path) if _P(file_path).is_absolute() else _P(workspace_root) / file_path
        abs_path = abs_path.resolve()

        # .xls/.xlsb → 工具层已转换为 .xlsx，checkpoint 需要打开转换后的文件
        from excelmanus.tools._helpers import ensure_openpyxl_compatible
        abs_path = ensure_openpyxl_compatible(abs_path)

        if not abs_path.is_file():
            return ""

        try:
            if tool_name == "write_cells":
                return ToolDispatcher._checkpoint_write_cells(abs_path, arguments)
            elif tool_name == "create_sheet":
                return ToolDispatcher._checkpoint_create_sheet(abs_path, arguments)
            elif tool_name == "delete_sheet":
                return ToolDispatcher._checkpoint_delete_sheet(abs_path, arguments)
            elif tool_name in ("insert_rows", "insert_columns"):
                return ToolDispatcher._checkpoint_insert(abs_path, arguments, tool_name)
        except Exception:
            return ""
        return ""

    @staticmethod
    def _checkpoint_write_cells(abs_path: "Path", arguments: dict) -> str:
        """write_cells 后回读验证：检查写入范围的行列数。"""
        import openpyxl as _opx

        sheet_name = arguments.get("sheet_name") or arguments.get("sheet")
        cell = arguments.get("cell")
        cell_range = arguments.get("cell_range")
        values = arguments.get("values")

        # read_only=False 必须：ws[cell] 随机访问在 read_only 模式下不支持
        wb = _opx.load_workbook(str(abs_path), read_only=False, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            if ws is None:
                return ""

            # 单元格模式
            if cell and not cell_range:
                val = ws[cell].value
                display = repr(val)[:60] if val is not None else "None"
                return f"\n✓ 回读确认: {ws.title}!{cell} = {display}"

            # 范围模式：检查写入区域行列数
            if values and isinstance(values, list):
                expected_rows = len(values)
                expected_cols = max((len(r) if isinstance(r, list) else 1) for r in values)
                # 读取实际写入区域的行列范围
                start_ref = (cell_range or "A1").split(":")[0]
                # 简单验证：读取目标区域的第一个和最后一个单元格
                first_val = ws[start_ref].value
                actual_rows = ws.max_row
                return (
                    f"\n✓ 回读确认: {ws.title}, "
                    f"写入 {expected_rows}行×{expected_cols}列, "
                    f"首格={repr(first_val)[:40]}, "
                    f"sheet总行数={actual_rows}"
                )

            return f"\n✓ 回读确认: {ws.title}, max_row={ws.max_row}"
        finally:
            wb.close()

    @staticmethod
    def _checkpoint_create_sheet(abs_path: "Path", arguments: dict) -> str:
        """create_sheet 后验证：确认 sheet 存在。"""
        import openpyxl as _opx

        target = arguments.get("sheet_name") or arguments.get("name", "")
        if not target:
            return ""
        wb = _opx.load_workbook(str(abs_path), read_only=True)
        try:
            if target in wb.sheetnames:
                return f"\n✓ 回读确认: sheet「{target}」已创建"
            else:
                return f"\n⚠ 回读异常: sheet「{target}」未找到"
        finally:
            wb.close()

    @staticmethod
    def _checkpoint_delete_sheet(abs_path: "Path", arguments: dict) -> str:
        """delete_sheet 后验证：确认 sheet 已删除。"""
        import openpyxl as _opx

        target = arguments.get("sheet_name") or arguments.get("name", "")
        if not target:
            return ""
        wb = _opx.load_workbook(str(abs_path), read_only=True)
        try:
            if target not in wb.sheetnames:
                return f"\n✓ 回读确认: sheet「{target}」已删除"
            else:
                return f"\n⚠ 回读异常: sheet「{target}」仍存在"
        finally:
            wb.close()

    @staticmethod
    def _checkpoint_insert(abs_path: "Path", arguments: dict, tool_name: str) -> str:
        """insert_rows/insert_columns 后验证：报告当前维度。"""
        import openpyxl as _opx

        sheet_name = arguments.get("sheet_name") or arguments.get("sheet")
        wb = _opx.load_workbook(str(abs_path), read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            if ws is None:
                return ""
            if tool_name == "insert_rows":
                return f"\n✓ 回读确认: {ws.title}, 当前总行数={ws.max_row}"
            else:
                return f"\n✓ 回读确认: {ws.title}, 当前总列数={ws.max_column}"
        finally:
            wb.close()

    # ── 写入操作日志辅助（供 verifier delta 注入）────────────

    @staticmethod
    def _extract_write_summary(tool_name: str, arguments: dict, result_str: str) -> str:
        """从 Excel 写入工具的参数/结果中提取简洁摘要。"""
        if tool_name == "write_cells":
            values = arguments.get("values")
            if isinstance(values, list):
                row_count = len(values)
                col_count = len(values[0]) if values and isinstance(values[0], list) else 1
                return f"写入 {row_count} 行 × {col_count} 列"
            return "写入数据"
        elif tool_name == "create_sheet":
            name = arguments.get("sheet_name") or arguments.get("name", "")
            return f"创建 sheet「{name}」" if name else "创建 sheet"
        elif tool_name == "delete_sheet":
            name = arguments.get("sheet_name") or arguments.get("name", "")
            return f"删除 sheet「{name}」" if name else "删除 sheet"
        elif tool_name == "insert_rows":
            count = arguments.get("count", 1)
            return f"插入 {count} 行"
        elif tool_name == "insert_columns":
            count = arguments.get("count", 1)
            return f"插入 {count} 列"
        return ""

    @staticmethod
    def _extract_run_code_write_summary(result_str: str) -> str:
        """从 run_code 的 stdout 中提取写入摘要（取首行非空输出）。"""
        if not result_str:
            return "run_code 写入"
        for line in result_str.split("\n"):
            stripped = line.strip()
            if stripped and not stripped.startswith("{") and not stripped.startswith("["):
                return stripped[:120]
        return "run_code 写入"

    # ── Excel 预览/Diff 事件辅助 ────────────────────────────

    _EXCEL_READ_TOOLS = {"read_excel"}
    _EXCEL_WRITE_TOOLS = {"write_to_sheet", "format_range"}

    @staticmethod
    def _extract_preview_styles(
        file_path: str, sheet_name: str | None, num_rows: int, num_cols: int,
        workspace_root: str,
    ) -> list[list]:
        """Best-effort: 提取 preview 区域的单元格样式（header + data rows）。"""
        from pathlib import Path
        from excelmanus.tools._style_extract import extract_cell_style
        import openpyxl

        abs_path = Path(file_path) if Path(file_path).is_absolute() else Path(workspace_root) / file_path
        abs_path = abs_path.resolve()
        if not abs_path.is_file() or abs_path.suffix.lower() not in (".xlsx", ".xlsm", ".xls", ".xlsb"):
            return []

        # .xls/.xlsb → 透明转换为 xlsx
        from excelmanus.tools._helpers import ensure_openpyxl_compatible
        abs_path = ensure_openpyxl_compatible(abs_path)

        wb = openpyxl.load_workbook(str(abs_path), read_only=False, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            if ws is None:
                return []
            styles: list[list] = []
            for r in range(1, num_rows + 2):  # +1 header row, +1 for 1-based
                row_styles: list = []
                for c in range(1, num_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    row_styles.append(extract_cell_style(cell))
                styles.append(row_styles)
            return styles
        finally:
            wb.close()

    @staticmethod
    def _extract_file_merge_ranges(
        file_path: str, sheet_name: str | None, workspace_root: str,
    ) -> list[dict[str, int]]:
        """Best-effort: 提取指定工作表的合并单元格区域。"""
        merges, _ = ToolDispatcher._extract_sheet_metadata(file_path, sheet_name, workspace_root)
        return merges

    @staticmethod
    def _extract_sheet_metadata(
        file_path: str, sheet_name: str | None, workspace_root: str,
    ) -> tuple[list[dict[str, int]], list[str]]:
        """Best-effort: 一次打开文件，提取合并区域 + 元数据提示。"""
        from pathlib import Path
        from excelmanus.tools._style_extract import extract_merge_ranges, extract_worksheet_hints
        import openpyxl

        abs_path = Path(file_path) if Path(file_path).is_absolute() else Path(workspace_root) / file_path
        abs_path = abs_path.resolve()
        if not abs_path.is_file() or abs_path.suffix.lower() not in (".xlsx", ".xlsm", ".xls", ".xlsb"):
            return [], []

        # .xls/.xlsb → 透明转换为 xlsx
        from excelmanus.tools._helpers import ensure_openpyxl_compatible
        abs_path = ensure_openpyxl_compatible(abs_path)

        wb = openpyxl.load_workbook(str(abs_path), read_only=False, data_only=False)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            if ws is None:
                return [], []
            return extract_merge_ranges(ws), extract_worksheet_hints(ws)
        finally:
            wb.close()

    @staticmethod
    def _collect_actual_excel_paths(
        ast_targets: list[str],
        result_json: dict | None,
        audit_changes: list | None,
        workspace_root: str,
    ) -> list[str]:
        """从 AST 目标、cow_mapping、audit_changes 中收集实际 Excel 文件路径。

        过滤掉 AST 无法解析的 ``<variable>`` 占位符，并从执行结果中
        补充实际被修改的文件路径，确保 diff 快照能正确工作。
        """
        from excelmanus.window_perception.extractor import is_excel_path, normalize_path

        seen: set[str] = set()
        result: list[str] = []

        # 1. AST 提取的字面量路径（过滤 <variable> 占位符）
        for p in ast_targets:
            if p and p != "<variable>" and p not in seen:
                seen.add(p)
                result.append(p)

        # 2. cow_mapping 中的实际路径（CoW 模式下最可靠）
        if result_json:
            cow = result_json.get("cow_mapping")
            if isinstance(cow, dict):
                for orig, copy_path in cow.items():
                    for cp in (orig, copy_path):
                        if isinstance(cp, str) and cp.strip():
                            norm = normalize_path(cp)
                            if norm and is_excel_path(norm) and norm not in seen:
                                seen.add(norm)
                                result.append(norm)

        # 3. audit_changes 中记录的实际文件变更
        if audit_changes:
            for change in audit_changes:
                path = getattr(change, "path", None) or ""
                if path:
                    norm = normalize_path(path)
                    if norm and is_excel_path(norm) and norm not in seen:
                        seen.add(norm)
                        result.append(norm)

        return result

    @staticmethod
    def _snapshot_excel_for_diff(
        file_paths: list[str], workspace_root: str,
    ) -> dict[str, list[tuple[str, list[dict], list[dict]]]]:
        """对指定 Excel 文件做轻量快照，返回 {file_path: [(sheet, cells, merges)]}。

        文件不存在时记录空列表（tombstone），以便 diff 能检测"从无到有"的新建场景。
        """
        from pathlib import Path
        snapshots: dict[str, list[tuple[str, list[dict], list[dict]]]] = {}
        for fp in file_paths:
            try:
                abs_path = Path(fp) if Path(fp).is_absolute() else Path(workspace_root) / fp
                abs_path = abs_path.resolve()
                if not abs_path.is_file():
                    # 文件不存在 → 记录空快照（tombstone），支持新建文件 diff
                    snapshots[fp] = []
                    continue
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                from excelmanus.tools._style_extract import extract_cell_style, extract_merge_ranges
                # .xls/.xlsb → 透明转换为 xlsx
                from excelmanus.tools._helpers import ensure_openpyxl_compatible
                abs_path = ensure_openpyxl_compatible(abs_path)
                wb = load_workbook(str(abs_path), data_only=False, read_only=False)
                file_snaps: list[tuple[str, list[dict], list[dict]]] = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cells: list[dict] = []
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 500),
                                            max_col=min(ws.max_column or 0, 50)):
                        for cell in row:
                            if cell.value is not None:
                                ref = f"{get_column_letter(cell.column)}{cell.row}"
                                val = cell.value
                                if isinstance(val, (int, float, bool, str)):
                                    entry: dict = {"cell": ref, "value": val}
                                else:
                                    entry = {"cell": ref, "value": str(val)}
                                style = extract_cell_style(cell)
                                if style:
                                    entry["style"] = style
                                cells.append(entry)
                    merges = extract_merge_ranges(ws)
                    file_snaps.append((sheet_name, cells, merges))
                wb.close()
                snapshots[fp] = file_snaps
            except Exception:
                pass
        return snapshots

    @staticmethod
    def _compute_snapshot_diffs(
        before: dict[str, list[tuple[str, list[dict], list[dict]]]],
        after: dict[str, list[tuple[str, list[dict], list[dict]]]],
    ) -> list[dict]:
        """对比前后快照，返回 [{file_path, sheet, affected_range, changes, old_merge_ranges, new_merge_ranges}]。"""
        results: list[dict] = []
        all_files = set(before) | set(after)
        for fp in sorted(all_files):
            # 兼容 2-tuple (旧格式) 和 3-tuple (新格式含 merges)
            def _unpack(items: list) -> dict[str, tuple[list[dict], list[dict]]]:
                out: dict[str, tuple[list[dict], list[dict]]] = {}
                for item in items:
                    if len(item) >= 3:
                        out[item[0]] = (item[1], item[2])
                    else:
                        out[item[0]] = (item[1], [])
                return out

            before_sheets = _unpack(before.get(fp, []))
            after_sheets = _unpack(after.get(fp, []))
            all_sheets = set(before_sheets) | set(after_sheets)
            for sheet in sorted(all_sheets):
                b_data, b_merges = before_sheets.get(sheet, ([], []))
                a_data, a_merges = after_sheets.get(sheet, ([], []))
                from excelmanus.tools.cell_tools import _compute_cell_diff
                changes = _compute_cell_diff(b_data, a_data)
                if changes:
                    from openpyxl.utils import get_column_letter as _gcl
                    from openpyxl.utils.cell import coordinate_to_tuple as _ctt
                    min_r = min_c = float("inf")
                    max_r = max_c = 0
                    for ch in changes:
                        try:
                            r, c = _ctt(ch["cell"].upper())
                            if r < min_r: min_r = r
                            if r > max_r: max_r = r
                            if c < min_c: min_c = c
                            if c > max_c: max_c = c
                        except Exception:
                            pass
                    if min_r != float("inf"):
                        _range = f"{_gcl(min_c)}{min_r}:{_gcl(max_c)}{max_r}" if (min_r, min_c) != (max_r, max_c) else f"{_gcl(min_c)}{min_r}"
                    else:
                        _range = changes[0]["cell"]
                    results.append({
                        "file_path": fp,
                        "sheet": sheet,
                        "affected_range": _range,
                        "changes": changes[:200],
                        "old_merge_ranges": b_merges,
                        "new_merge_ranges": a_merges,
                    })
        return results

    def _emit_excel_events(
        self,
        e: Any,
        on_event: Any,
        tool_call_id: str,
        tool_name: str,
        arguments: dict,
        result_str: str,
        iteration: int,
    ) -> None:
        """在工具调用成功后，检测 Excel 相关结果并发射预览/Diff 事件。"""
        import json as _json
        from excelmanus.events import EventType, ToolCallEvent

        try:
            parsed = _json.loads(result_str)
        except (ValueError, TypeError):
            return
        if not isinstance(parsed, dict):
            return

        # 工具 read_excel 对应事件 EXCEL_PREVIEW
        if tool_name in self._EXCEL_READ_TOOLS:
            columns = parsed.get("columns", [])
            preview = parsed.get("preview", [])
            if columns and preview:
                rows_data = []
                for record in preview[:50]:
                    if isinstance(record, dict):
                        rows_data.append([record.get(c) for c in columns])
                    elif isinstance(record, list):
                        rows_data.append(record)
                total_rows = parsed.get("total_rows_in_sheet") or parsed.get("shape", {}).get("rows", 0)
                # Best-effort: 提取预览单元格样式
                cell_styles: list[list] = []
                try:
                    cell_styles = self._extract_preview_styles(
                        arguments.get("file_path", ""),
                        arguments.get("sheet_name") or None,
                        len(rows_data),
                        len(columns),
                        e.config.workspace_root,
                    )
                except Exception:
                    logger.debug("提取预览单元格样式失败", exc_info=True)
                merge_ranges: list[dict[str, int]] = []
                metadata_hints: list[str] = []
                try:
                    merge_ranges, metadata_hints = self._extract_sheet_metadata(
                        arguments.get("file_path", ""),
                        arguments.get("sheet_name") or None,
                        e.config.workspace_root,
                    )
                except Exception:
                    logger.debug("提取工作表元数据失败", exc_info=True)
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.EXCEL_PREVIEW,
                        tool_call_id=tool_call_id,
                        excel_file_path=arguments.get("file_path", ""),
                        excel_sheet=arguments.get("sheet_name", ""),
                        excel_columns=columns[:100],
                        excel_rows=rows_data[:50],
                        excel_total_rows=int(total_rows) if total_rows else 0,
                        excel_truncated=bool(parsed.get("is_truncated", False)),
                        excel_cell_styles=cell_styles,
                        excel_merge_ranges=merge_ranges,
                        excel_metadata_hints=metadata_hints,
                    ),
                )

        # _excel_diff 对应 EXCEL_DIFF（写入工具在结果中附带）
        diff_data = parsed.get("_excel_diff")
        if isinstance(diff_data, dict):
            changes = diff_data.get("changes", [])
            if changes:
                # 优先使用 diff_data 自带的 merge ranges（写入前后各自捕获）
                diff_old_merges: list[dict[str, int]] = diff_data.get("old_merge_ranges", [])
                diff_new_merges: list[dict[str, int]] = diff_data.get("new_merge_ranges", [])
                diff_hints: list[str] = []
                if not diff_new_merges:
                    try:
                        diff_new_merges, diff_hints = self._extract_sheet_metadata(
                            diff_data.get("file_path", ""),
                            diff_data.get("sheet") or None,
                            e.config.workspace_root,
                        )
                    except Exception:
                        logger.debug("提取 diff 工作表元数据失败", exc_info=True)
                else:
                    try:
                        _, diff_hints = self._extract_sheet_metadata(
                            diff_data.get("file_path", ""),
                            diff_data.get("sheet") or None,
                            e.config.workspace_root,
                        )
                    except Exception:
                        pass
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.EXCEL_DIFF,
                        tool_call_id=tool_call_id,
                        excel_file_path=diff_data.get("file_path", ""),
                        excel_sheet=diff_data.get("sheet", ""),
                        excel_affected_range=diff_data.get("affected_range", ""),
                        excel_changes=changes[:200],
                        excel_merge_ranges=diff_new_merges,
                        excel_old_merge_ranges=diff_old_merges,
                        excel_metadata_hints=diff_hints,
                    ),
                )

        # compare_excel 跨文件/跨 Sheet 对比 → EXCEL_DIFF（带扩展字段）
        if tool_name == "compare_excel" and parsed.get("status") == "ok":
            diff_mode = parsed.get("diff_mode", "cross_file")
            sample_diffs = parsed.get("sample_diffs", [])
            # 将 sample_diffs 转为 excel_changes 格式
            cross_changes: list[dict] = []
            for sd in sample_diffs[:200]:
                cross_changes.append({
                    "cell": sd.get("cell", sd.get("column", "")),
                    "key": sd.get("key", ""),
                    "old": sd.get("old"),
                    "new": sd.get("new"),
                })
            e.emit(
                on_event,
                ToolCallEvent(
                    event_type=EventType.EXCEL_DIFF,
                    tool_call_id=tool_call_id,
                    excel_file_path=arguments.get("file_a", ""),
                    excel_sheet=arguments.get("sheet_a", ""),
                    excel_changes=cross_changes,
                    excel_diff_mode=diff_mode,
                    excel_file_b=arguments.get("file_b", ""),
                    excel_sheet_b=arguments.get("sheet_b", ""),
                    excel_diff_summary=parsed.get("summary"),
                ),
            )

        # _text_diff 对应 TEXT_DIFF（write_text_file / edit_text_file 在结果中附带）
        text_diff_data = parsed.get("_text_diff")
        if isinstance(text_diff_data, dict):
            hunks = text_diff_data.get("hunks", [])
            if hunks:
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.TEXT_DIFF,
                        tool_call_id=tool_call_id,
                        text_diff_file_path=text_diff_data.get("file_path", ""),
                        text_diff_hunks=hunks[:300],
                        text_diff_additions=text_diff_data.get("additions", 0),
                        text_diff_deletions=text_diff_data.get("deletions", 0),
                        text_diff_truncated=text_diff_data.get("truncated", False),
                    ),
                )

        # _text_preview 对应 TEXT_PREVIEW（read_text_file 在结果中附带）
        text_preview_data = parsed.get("_text_preview")
        if isinstance(text_preview_data, dict):
            preview_content = text_preview_data.get("content", "")
            if preview_content:
                e.emit(
                    on_event,
                    ToolCallEvent(
                        event_type=EventType.TEXT_PREVIEW,
                        tool_call_id=tool_call_id,
                        text_preview_file_path=text_preview_data.get("file_path", ""),
                        text_preview_content=preview_content[:20000],
                        text_preview_line_count=text_preview_data.get("line_count", 0),
                        text_preview_truncated=text_preview_data.get("truncated", False),
                    ),
                )

        # _file_download 对应 FILE_DOWNLOAD（offer_download 工具在结果中附带）
        dl_data = parsed.get("_file_download")
        if isinstance(dl_data, dict) and dl_data.get("file_path"):
            e.emit(
                on_event,
                ToolCallEvent(
                    event_type=EventType.FILE_DOWNLOAD,
                    tool_call_id=tool_call_id,
                    download_file_path=dl_data.get("file_path", ""),
                    download_filename=dl_data.get("filename", ""),
                    download_description=dl_data.get("description", ""),
                ),
            )

    def _emit_files_changed_from_report(
        self,
        e: Any,
        on_event: Any,
        tool_call_id: str,
        report: dict | None,
        iteration: int,
    ) -> None:
        """finish_task 完成后，从 report['affected_files'] 提取受影响文件并发射 FILES_CHANGED 事件。"""
        if not report or on_event is None:
            return
        from excelmanus.events import EventType, ToolCallEvent
        from excelmanus.window_perception.extractor import is_excel_path, normalize_path

        affected: set[str] = set()
        for f in report.get("affected_files", []):
            if isinstance(f, str) and f:
                norm = normalize_path(f)
                if norm and is_excel_path(norm):
                    affected.add(norm)
        if not affected:
            return
        e.emit(
            on_event,
            ToolCallEvent(
                event_type=EventType.FILES_CHANGED,
                tool_call_id=tool_call_id,
                iteration=iteration,
                changed_files=sorted(affected),
            ),
        )

    def _emit_files_changed_from_audit(
        self,
        e: Any,
        on_event: Any,
        tool_call_id: str,
        code: str,
        audit_changes: list[Any] | None,
        iteration: int,
        extra_changed_paths: list[str] | None = None,
        cow_mapping: dict[str, str] | None = None,
    ) -> None:
        """run_code 执行后，从审计、AST、cow_mapping 和 mtime 探针中提取受影响文件并发射 FILES_CHANGED 事件。"""
        from excelmanus.events import EventType, ToolCallEvent
        from excelmanus.security.code_policy import extract_excel_targets
        from excelmanus.window_perception.extractor import is_excel_path, normalize_path

        affected: set[str] = set()

        if audit_changes:
            for change in audit_changes:
                path = getattr(change, "path", None) or ""
                if path:
                    norm = normalize_path(path)
                    if norm and is_excel_path(norm):
                        affected.add(norm)

        for target in extract_excel_targets(code or ""):
            if target.operation in ("write", "unknown") and target.file_path != "<variable>":
                norm = normalize_path(target.file_path)
                if norm and is_excel_path(norm):
                    affected.add(norm)

        # cow_mapping 中的实际路径（AST 无法解析变量时的可靠回退）
        if cow_mapping:
            for orig, copy_path in cow_mapping.items():
                for cp in (orig, copy_path):
                    if isinstance(cp, str) and cp.strip():
                        norm = normalize_path(cp)
                        if norm and is_excel_path(norm):
                            affected.add(norm)

        # mtime 探针检测到的新建/变更文件（不限文件类型）
        if extra_changed_paths:
            for p in extra_changed_paths:
                if p:
                    affected.add(p)

        if not affected or on_event is None:
            return

        e.emit(
            on_event,
            ToolCallEvent(
                event_type=EventType.FILES_CHANGED,
                tool_call_id=tool_call_id,
                iteration=iteration,
                changed_files=sorted(affected),
            ),
        )

