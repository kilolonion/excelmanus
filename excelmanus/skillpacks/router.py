"""Skillpack 路由器：斜杠直连与 fallback。"""

from __future__ import annotations

import asyncio
from collections import OrderedDict
from dataclasses import replace
from pathlib import Path
import re
from typing import Any

from excelmanus.config import ExcelManusConfig
from excelmanus.events import EventCallback, EventType, ToolCallEvent
from excelmanus.logger import get_logger
from excelmanus.skillpacks.arguments import parse_arguments, substitute
from excelmanus.skillpacks.context_builder import build_contexts_with_budget
from excelmanus.skillpacks.loader import SkillpackLoader
from excelmanus.skillpacks.models import SkillMatchResult, Skillpack
from excelmanus.tools.policy import TOOL_CATEGORIES

logger = get_logger("skillpacks.router")

_EXCEL_PATH_PATTERN = re.compile(
    r"""(?:"([^"]+\.(?:xlsx|xlsm|xls))"|'([^']+\.(?:xlsx|xlsm|xls))'|([^\s"'，。！？；：]+?\.(?:xlsx|xlsm|xls)))""",
    re.IGNORECASE,
)
_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
_MAY_WRITE_HINT_RE = re.compile(
    r"(创建|修改|写入|删除|替换|填充|插入|更新|设置|格式化|高亮|加粗|标红|美化|条件格式|"
    r"画图|生成图表|柱状图|饼图|折线图|雷达图|散点图|排序|合并|转置|拆分|"
    r"write|update|format|chart|"
    r"fill|match|replace|insert|sort|merge|solve|fix|address|apply|"
    r"\bset\b|\bput\b|\badd\b|create|make|generate|compute|calculate|populate|assign)",
    re.IGNORECASE,
)
_READ_ONLY_HINT_RE = re.compile(
    r"(查看|列出|读取|扫描|分析|统计|对比|检查|预览|找出|筛选|汇总|排名|占比|read|scan|analyz|inspect|list)",
    re.IGNORECASE,
)

# 纯问候/闲聊/身份问答检测：短消息且仅含问候词或元问题时跳过 LLM 分类
_CHITCHAT_RE = re.compile(
    r"^\s*(?:"
    # 问候
    r"你好|您好|hi|hello|hey|嗨|哈喽|早上好|下午好|晚上好|good\s*(?:morning|afternoon|evening)"
    r"|在吗|在不在|谢谢|thanks|thank\s*you|好的|ok|okay"
    # 身份/元问题（你是谁、介绍一下你自己、what are you 等）
    r"|你是谁|你是什么|你叫什么|介绍一下你?自己|你能做什么|你会什么|你有什么功能"
    r"|who\s*are\s*you|what\s*are\s*you|what\s*can\s*you\s*do|introduce\s*yourself"
    r"|what\s*(?:is|are)\s*(?:your|ur)\s*(?:name|capabilities)"
    # 通用短问答（怎么用、什么是、如何使用、帮助）
    r"|怎么用|如何使用|help|帮助|使用说明|使用方法"
    r")[?？!！。.\s]*$",
    re.IGNORECASE,
)

# 任务标签词法规则：(tag_name, pattern)
_TASK_TAG_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("cross_sheet", re.compile(
        r"(跨[表页sheet]|从.*[表页sheet].*到|Sheet\s*\d.*Sheet\s*\d|VLOOKUP|INDEX.*MATCH|lookup|匹配.*填[入充])",
        re.IGNORECASE,
    )),
    ("formatting", re.compile(
        r"(格式化|高亮|加粗|标红|美化|条件格式|字体|颜色|边框|填充色|行高|列宽|对齐)",
        re.IGNORECASE,
    )),
    ("chart", re.compile(
        r"(图表|柱状图|折线图|饼图|散点图|雷达图|chart|画图|生成图)",
        re.IGNORECASE,
    )),
    ("data_fill", re.compile(
        r"(填[入充写]|补[充全齐]|populate|fill)",
        re.IGNORECASE,
    )),
    ("large_data", re.compile(
        r"(批量|全[部量表]|所有行|每一行|大量|bulk|batch)",
        re.IGNORECASE,
    )),
]

class SkillRouter:
    """简化后的技能路由器：仅负责斜杠直连路由和技能目录生成。"""

    def __init__(self, config: ExcelManusConfig, loader: SkillpackLoader) -> None:
        self._config = config
        self._loader = loader
        self._router_client: Any = None
        self._fallback_client: Any = None
        self._file_structure_cache: OrderedDict[
            tuple[str, int, int], tuple[list[str], int, int]
        ] = OrderedDict()
        self._file_structure_cache_limit = 64

    def _get_router_client(self) -> Any:
        """懒加载 AUX 端点客户端，首次调用时创建并缓存。"""
        if self._router_client is None:
            from excelmanus.providers import create_client
            self._router_client = create_client(
                api_key=self._config.aux_api_key or self._config.api_key,
                base_url=self._config.aux_base_url or self._config.base_url,
            )
        return self._router_client

    def _get_fallback_client(self) -> Any:
        """懒加载主模型降级客户端，首次调用时创建并缓存。"""
        if self._fallback_client is None:
            from excelmanus.providers import create_client
            self._fallback_client = create_client(
                api_key=self._config.api_key,
                base_url=self._config.base_url,
            )
        return self._fallback_client

    async def route(
        self,
        user_message: str,
        *,
        slash_command: str | None = None,
        raw_args: str | None = None,
        file_paths: list[str] | None = None,
        blocked_skillpacks: set[str] | None = None,
        write_hint: str | None = None,
        on_event: EventCallback | None = None,
    ) -> SkillMatchResult:
        """执行路由：斜杠命令按技能直连；非斜杠默认全量工具。"""
        skillpacks = self._loader.get_skillpacks()
        if not skillpacks:
            skillpacks = self._loader.load_all()

        if not skillpacks:
            return await self._build_fallback_result(
                selected=[],
                route_mode="no_skillpack",
                all_skillpacks={},
            )

        blocked = set(blocked_skillpacks or [])
        # 斜杠直连路由使用排除被限制技能后的集合
        available_skillpacks = (
            {
                name: skill
                for name, skill in skillpacks.items()
                if name not in blocked
            }
            if blocked
            else skillpacks
        )

        # ── 0. 斜杠命令直连路由（参数化）──
        candidate_file_paths = self._collect_candidate_file_paths(
            user_message=user_message,
            file_paths=file_paths or [],
            raw_args=raw_args or "",
        )
        if slash_command and slash_command.strip():
            # 斜杠路径同步词法分类 write_hint，避免 LLM 调用延迟
            # 分类文本优先用 raw_args（任务描述），其次 user_message
            _hint_text = (raw_args or "").strip() or user_message
            slash_write_hint: str = (
                write_hint
                or self._classify_write_hint_lexical(_hint_text)
                or "unknown"
            )
            direct_skill = self._find_skill_by_name(
                skillpacks=available_skillpacks,
                name=slash_command.strip(),
            )
            if direct_skill is None:
                return await self._build_fallback_result(
                    selected=[],
                    route_mode="slash_not_found",
                    all_skillpacks=skillpacks,
                    user_message=user_message,
                    candidate_file_paths=candidate_file_paths,
                    blocked_skillpacks=blocked_skillpacks,
                    write_hint=slash_write_hint,
                )
            if not direct_skill.user_invocable:
                return await self._build_fallback_result(
                    selected=[],
                    route_mode="slash_not_user_invocable",
                    all_skillpacks=skillpacks,
                    user_message=user_message,
                    candidate_file_paths=candidate_file_paths,
                    blocked_skillpacks=blocked_skillpacks,
                    write_hint=slash_write_hint,
                )
            return await self._build_parameterized_result(
                skill=direct_skill,
                raw_args=raw_args or "",
                user_message=user_message,
                candidate_file_paths=candidate_file_paths,
                write_hint=slash_write_hint,
            )

        # ── 1. 纯问候/闲聊短路：跳过 LLM 分类，零延迟返回 ──
        if _CHITCHAT_RE.match(user_message.strip()) and not candidate_file_paths:
            logger.debug("chitchat 短路: %s", user_message[:30])
            return await self._build_all_tools_result(
                user_message=user_message,
                candidate_file_paths=candidate_file_paths,
                write_hint="read_only",
                task_tags=(),
            )

        # ── 2. 非斜杠消息：默认全量工具（tool_scope 置空，由引擎补全）──
        # 并行调用小模型判断 write_hint + task_tags
        classified_hint, classified_tags = await self._classify_task(
            user_message, write_hint=write_hint, on_event=on_event,
        )
        return await self._build_all_tools_result(
            user_message=user_message,
            candidate_file_paths=candidate_file_paths,
            write_hint=classified_hint,
            task_tags=classified_tags,
        )

    def build_skill_catalog(
        self,
        blocked_skillpacks: set[str] | None = None,
    ) -> tuple[str, list[str]]:
        """生成技能目录摘要和技能名称列表。

        被限制的技能仍会出现在目录中（标注需要 fullAccess），
        以便 LLM 知道其存在并在用户需要时给出权限提示。

        返回:
            (catalog_text, skill_names) 元组
        """
        skillpacks = self._loader.get_skillpacks()
        if not skillpacks:
            skillpacks = self._loader.load_all()

        if not skillpacks:
            return ("", [])

        blocked = set(blocked_skillpacks or [])
        visible_pairs = sorted(
            (
                (name, skill)
                for name, skill in skillpacks.items()
                if not skill.disable_model_invocation
            ),
            key=lambda item: item[0].lower(),
        )
        skill_names = [name for name, _ in visible_pairs]
        lines = ["可用技能：\n"]
        for name, skill in visible_pairs:
            if name in blocked:
                lines.append(
                    f"- {name}：{skill.description} "
                    f"[⚠️ 需要 fullAccess 权限，使用 /fullAccess on 开启]"
                )
            else:
                lines.append(f"- {name}：{skill.description}")
        catalog_text = "\n".join(lines)
        return (catalog_text, skill_names)



    def _build_result(
        self,
        selected: list[Skillpack],
        route_mode: str,
        *,
        parameterized: bool = False,
        write_hint: str = "unknown",
    ) -> SkillMatchResult:
        """构建路由结果。"""
        skills_used = [skill.name for skill in selected]
        contexts = build_contexts_with_budget(
            selected, self._config.skills_context_char_budget
        )
        return SkillMatchResult(
            skills_used=skills_used,
            route_mode=route_mode,
            system_contexts=contexts,
            parameterized=parameterized,
            write_hint=write_hint,
        )

    async def _build_fallback_result(
        self,
        selected: list[Skillpack],
        route_mode: str,
        all_skillpacks: dict[str, Skillpack],
        *,
        user_message: str = "",
        candidate_file_paths: list[str] | None = None,
        blocked_skillpacks: set[str] | None = None,
        write_hint: str = "unknown",
    ) -> SkillMatchResult:
        """构建 fallback 路由结果。

        技能目录通过 activate_skill 元工具的 description 传递。
        """
        result = self._build_result(
            selected=selected,
            route_mode=route_mode,
            write_hint=write_hint,
        )

        system_contexts = list(result.system_contexts)
        large_file_context = self._build_large_file_context(
            user_message=user_message,
            candidate_file_paths=candidate_file_paths,
        )
        if large_file_context:
            system_contexts.append(large_file_context)

        file_structure_context, sheet_count, max_total_rows = await self._build_file_structure_context(
            candidate_file_paths=candidate_file_paths,
        )
        if file_structure_context:
            system_contexts.append(file_structure_context)

        return SkillMatchResult(
            skills_used=result.skills_used,
            route_mode=result.route_mode,
            system_contexts=system_contexts,
            parameterized=result.parameterized,
            write_hint=result.write_hint,
            sheet_count=sheet_count,
            max_total_rows=max_total_rows,
        )

    async def _build_all_tools_result(
        self,
        *,
        user_message: str,
        candidate_file_paths: list[str] | None,
        write_hint: str = "unknown",
        task_tags: tuple[str, ...] = (),
    ) -> SkillMatchResult:
        """构建非斜杠默认路由：tool_scope 为空表示由引擎放开全量工具。"""
        system_contexts: list[str] = []
        large_file_context = self._build_large_file_context(
            user_message=user_message,
            candidate_file_paths=candidate_file_paths,
        )
        if large_file_context:
            system_contexts.append(large_file_context)

        file_structure_context, sheet_count, max_total_rows = await self._build_file_structure_context(
            candidate_file_paths=candidate_file_paths,
        )
        if file_structure_context:
            system_contexts.append(file_structure_context)

        return SkillMatchResult(
            skills_used=[],
            route_mode="all_tools",
            system_contexts=system_contexts,
            parameterized=False,
            write_hint=write_hint,
            sheet_count=sheet_count,
            max_total_rows=max_total_rows,
            task_tags=task_tags,
        )

    @staticmethod
    def _emit_pipeline(
        on_event: EventCallback | None,
        stage: str,
        message: str,
    ) -> None:
        if on_event is not None:
            on_event(ToolCallEvent(
                event_type=EventType.PIPELINE_PROGRESS,
                pipeline_stage=stage,
                pipeline_message=message,
            ))

    async def _classify_task(
        self,
        user_message: str,
        *,
        write_hint: str | None = None,
        on_event: EventCallback | None = None,
    ) -> tuple[str, tuple[str, ...]]:
        """综合分类：write_hint + task_tags。

        三级策略：
        1. 词法规则（零延迟）：关键词/正则能确定时直接返回
        2. 同步 LLM 分类（带超时）：词法不确定时调用小模型，
           阻塞等待结果（超时 ~2s），避免异步竞态导致下游误判
        3. 保守默认 ``may_write``：LLM 超时/失败时的最终兜底

        Returns:
            (write_hint, task_tags)
        """
        lexical_hint = write_hint or self._classify_write_hint_lexical(user_message)
        lexical_tags = self._classify_task_tags_lexical(user_message)

        if lexical_hint:
            return lexical_hint, tuple(lexical_tags)

        # 词法无法判断：同步调用 LLM 分类（带超时），避免异步竞态
        try:
            llm_hint, llm_tags = await self._classify_task_llm(user_message)
            if llm_hint:
                merged_tags = list(set(lexical_tags + llm_tags))
                return llm_hint, tuple(merged_tags)
        except Exception:
            logger.debug("同步 LLM 分类失败，回退到 may_write", exc_info=True)

        # LLM 也无法判断：返回 may_write 作为保守默认值
        return "may_write", tuple(lexical_tags)

    async def _classify_task_llm(
        self,
        user_message: str,
    ) -> tuple[str | None, list[str]]:
        """调用小模型判断 write_hint 和 task_tags。

        返回 (write_hint | None, task_tags)。失败时返回 (None, [])。
        """
        import asyncio
        import json as _json

        system_prompt = (
            "你是任务分类器。判断用户请求的意图和任务类型。"
            '只输出 JSON: {"write_hint": "may_write"|"read_only", "task_tags": [...]}\n'
            "write_hint 判断：\n"
            "- may_write：创建/修改/写入/删除/替换/填充/格式化/图表/排序/合并/转置/"
            "fill/match/replace/insert/sort/solve/fix/apply/set/add/create/generate/compute/calculate\n"
            "- read_only：消息仅涉及查看/列出/读取/分析/统计/对比/检查/预览/筛选/找出/汇总/排名/占比，且不涉及任何数据变更时\n"
            "- read_only 典型示例：'筛选出所有XX并统计每组数量'、'找出哪个XX最多'、'分析数据趋势'、'对比两个表的差异'\n"
            "- 当消息中同时包含文件路径和数据变更描述（如 fill column D、match invoice numbers）时，优先 may_write\n"
            "- 仅当完全确定不涉及写入时才判定 read_only\n\n"
            "task_tags 可选值（选择所有适用的）：\n"
            "- cross_sheet：涉及跨工作表查找或数据传输\n"
            "- formatting：涉及样式/颜色/字体/边框等格式化\n"
            "- chart：涉及图表生成\n"
            "- data_fill：涉及数据填充或批量写入\n"
            "- large_data：涉及大量数据处理（>100行）\n"
            "- 无明确标签时返回空数组 []"
        )
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message[:500]},
        ]

        async def _try_classify(
            client: Any, model: str, timeout: float,
        ) -> tuple[str | None, list[str]] | None:
            try:
                response = await asyncio.wait_for(
                    client.chat.completions.create(
                        model=model,
                        messages=messages,
                        max_tokens=60,
                        temperature=0,
                    ),
                    timeout=timeout,
                )
                text = (response.choices[0].message.content or "").strip()
                left = text.find("{")
                right = text.rfind("}")
                if left >= 0 and right > left:
                    text = text[left:right + 1]
                try:
                    data = _json.loads(text)
                    hint = str(data.get("write_hint", "")).strip().lower()
                    hint = hint if hint in ("may_write", "read_only") else None
                    raw_tags = data.get("task_tags", [])
                    tags = [
                        str(t).strip().lower()
                        for t in (raw_tags if isinstance(raw_tags, list) else [])
                        if str(t).strip()
                    ]
                    return hint, tags
                except (_json.JSONDecodeError, TypeError, AttributeError):
                    return None
            except Exception:
                return None

        timeout_seconds = 10.0

        # 1) AUX 端点
        result = await _try_classify(
            self._get_router_client(), self._config.aux_model, timeout_seconds,
        )
        if result is not None:
            return result

        # 2) 降级到主模型
        result = await _try_classify(
            self._get_fallback_client(), self._config.model, timeout_seconds,
        )
        if result is not None:
            return result

        return None, []

    @staticmethod
    def _classify_task_tags_lexical(user_message: str) -> list[str]:
        """词法快速推断 task_tags。"""
        text = str(user_message or "").strip()
        if not text:
            return []
        tags: list[str] = []
        for tag_name, pattern in _TASK_TAG_PATTERNS:
            if pattern.search(text):
                tags.append(tag_name)
        return tags

    @staticmethod
    def _classify_write_hint_lexical(user_message: str) -> str | None:
        """本地词法兜底：优先识别写入/格式化/图表等明确写意图。

        返回 None 表示词法无法判断，交由 LLM 分类。
        """
        text = str(user_message or "").strip()
        if not text:
            return None
        if _MAY_WRITE_HINT_RE.search(text):
            return "may_write"
        if _READ_ONLY_HINT_RE.search(text):
            return "read_only"
        # 消息中包含 Excel 文件路径引用 → 大概率要操作文件，保守归为 may_write
        if _EXCEL_PATH_PATTERN.search(text):
            return "may_write"
        return None

    @staticmethod
    def _normalize_skill_name(name: str) -> str:
        return name.strip().lower().replace("-", "").replace("_", "")

    def _find_skill_by_name(
        self,
        *,
        skillpacks: dict[str, Skillpack],
        name: str,
    ) -> Skillpack | None:
        direct = skillpacks.get(name)
        if direct is not None:
            return direct

        by_lower = {skill_name.lower(): skill for skill_name, skill in skillpacks.items()}
        direct_lower = by_lower.get(name.lower())
        if direct_lower is not None:
            return direct_lower

        normalized = self._normalize_skill_name(name)
        normalized_map = {
            self._normalize_skill_name(skill_name): skill
            for skill_name, skill in skillpacks.items()
        }
        return normalized_map.get(normalized)

    async def _build_parameterized_result(
        self,
        *,
        skill: Skillpack,
        raw_args: str,
        user_message: str = "",
        candidate_file_paths: list[str] | None = None,
        write_hint: str = "unknown",
    ) -> SkillMatchResult:
        args = parse_arguments(raw_args)
        rendered_instructions = substitute(skill.instructions, args)
        parameterized_skill = replace(skill, instructions=rendered_instructions)
        result = self._build_result(
            selected=[parameterized_skill],
            route_mode="slash_direct",
            parameterized=True,
            write_hint=write_hint,
        )
        system_contexts = list(result.system_contexts)
        large_file_context = self._build_large_file_context(
            user_message=user_message,
            candidate_file_paths=candidate_file_paths,
        )
        if large_file_context:
            system_contexts.append(large_file_context)
        _, sheet_count, max_total_rows = await self._build_file_structure_context(
            candidate_file_paths=candidate_file_paths,
        )
        return SkillMatchResult(
            skills_used=result.skills_used,
            route_mode=result.route_mode,
            system_contexts=system_contexts,
            parameterized=result.parameterized,
            write_hint=result.write_hint,
            sheet_count=sheet_count,
            max_total_rows=max_total_rows,
        )

    def _collect_candidate_file_paths(
        self,
        *,
        user_message: str,
        file_paths: list[str],
        raw_args: str,
    ) -> list[str]:
        merged: list[str] = []
        merged.extend(file_paths)
        merged.extend(self._extract_excel_paths(user_message))
        merged.extend(self._extract_excel_paths(raw_args))
        deduped: list[str] = []
        seen: set[str] = set()
        for path in merged:
            normalized = path.strip()
            if not normalized or normalized in seen:
                continue
            deduped.append(normalized)
            seen.add(normalized)
        return deduped

    def _build_file_structure_context_sync(
        self,
        *,
        candidate_file_paths: list[str] | None,
        max_files: int = 3,
        max_sheets: int = 5,
        scan_rows: int = 12,
    ) -> tuple[str, int, int]:
        """预读候选 Excel 文件的 sheet 结构，注入路由上下文。

        用 openpyxl 只读模式仅读前几行，帮助 LLM 确定 header_row
        和可用列名，避免盲猜导致的多轮重试。

        Returns:
            (context_text, sheet_count, max_total_rows)
        """
        if not candidate_file_paths:
            return "", 0, 0

        file_sections: list[str] = []
        processed = 0
        seen: set[str] = set()
        all_sheet_count = 0
        all_max_total_rows = 0

        for raw_path in candidate_file_paths:
            if processed >= max_files:
                break
            normalized = (raw_path or "").strip()
            if not normalized:
                continue
            try:
                path_obj = Path(normalized).expanduser()
                if not path_obj.is_absolute():
                    path_obj = Path.cwd() / path_obj
                resolved = path_obj.resolve(strict=False)
            except OSError:
                continue

            resolved_str = str(resolved)
            if resolved_str in seen:
                continue
            seen.add(resolved_str)

            if resolved.suffix.lower() not in _EXCEL_SUFFIXES:
                continue
            try:
                if not resolved.is_file():
                    continue
                stat_result = resolved.stat()
            except OSError:
                continue

            cache_key = (resolved_str, stat_result.st_mtime_ns, stat_result.st_size)
            cached_entry = self._file_structure_cache.get(cache_key)
            if cached_entry is not None:
                cached_sheet_lines, cached_sheet_count, cached_max_total_rows = cached_entry
                all_sheet_count += cached_sheet_count
                all_max_total_rows = max(all_max_total_rows, cached_max_total_rows)
                self._file_structure_cache.move_to_end(cache_key)
                if cached_sheet_lines:
                    file_sections.append(
                        f"文件: {normalized}\n" + "\n".join(cached_sheet_lines)
                    )
                    processed += 1
                continue

            try:
                from openpyxl import load_workbook
                wb = load_workbook(resolved, read_only=True, data_only=True)
            except Exception:
                continue

            sheet_lines: list[str] = []
            file_sheet_count = 0
            file_max_total_rows = 0
            try:
                file_sheet_count = len(wb.sheetnames)
                for idx, sn in enumerate(wb.sheetnames):
                    ws = wb[sn]
                    if idx >= max_sheets:
                        # 超出详细预览数量的 sheet：仅输出摘要行（名称+行列数）
                        summary_rows = ws.max_row or 0
                        summary_cols = ws.max_column or 0
                        file_max_total_rows = max(file_max_total_rows, summary_rows)
                        sheet_lines.append(f"  [{sn}] {summary_rows}行×{summary_cols}列")
                        continue

                    rows_data: list[list[Any]] = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= scan_rows:
                            break
                        normalized_row = []
                        for c in row:
                            if isinstance(c, str):
                                text = c.strip()
                                normalized_row.append(text if text else None)
                            else:
                                normalized_row.append(c)
                        rows_data.append(normalized_row)

                    total_rows = ws.max_row or 0
                    total_cols = ws.max_column or 0
                    file_max_total_rows = max(file_max_total_rows, total_rows)
                    sheet_lines.append(f"  [{sn}] {total_rows}行×{total_cols}列")

                    for ri, row_vals in enumerate(rows_data):
                        # 过滤尾部 None
                        trimmed = row_vals
                        while trimmed and trimmed[-1] is None:
                            trimmed = trimmed[:-1]
                        display = [str(v) if v is not None else None for v in trimmed]
                        sheet_lines.append(f"    第{ri}行: {display}")

                    # 启发式 header_row 建议
                    header_hint = self._guess_header_row(rows_data)
                    if header_hint is not None:
                        sheet_lines.append(f"    → 建议 header_row={header_hint}")
            finally:
                wb.close()

            all_sheet_count += file_sheet_count
            all_max_total_rows = max(all_max_total_rows, file_max_total_rows)
            self._set_file_structure_cache(
                cache_key=cache_key,
                sheet_lines=sheet_lines,
                sheet_count=file_sheet_count,
                max_total_rows=file_max_total_rows,
            )

            if sheet_lines:
                file_sections.append(f"文件: {normalized}\n" + "\n".join(sheet_lines))
                processed += 1

        if not file_sections:
            return "", all_sheet_count, all_max_total_rows

        header = (
            "[文件结构预览] 以下是用户提及的 Excel 文件结构，请据此确定正确的 header_row 和列名。\n"
            "请基于以上预览直接调用工具执行用户请求，不要重复描述文件结构。"
        )
        return header + "\n" + "\n".join(file_sections), all_sheet_count, all_max_total_rows

    async def _build_file_structure_context(
        self,
        *,
        candidate_file_paths: list[str] | None,
        max_files: int = 3,
        max_sheets: int = 5,
        scan_rows: int = 12,
    ) -> tuple[str, int, int]:
        """异步包装：将同步 openpyxl I/O 卸载到线程池，避免阻塞事件循环。"""
        return await asyncio.to_thread(
            self._build_file_structure_context_sync,
            candidate_file_paths=candidate_file_paths,
            max_files=max_files,
            max_sheets=max_sheets,
            scan_rows=scan_rows,
        )

    def _set_file_structure_cache(
        self,
        *,
        cache_key: tuple[str, int, int],
        sheet_lines: list[str],
        sheet_count: int,
        max_total_rows: int,
    ) -> None:
        """写入文件结构缓存并维护 LRU 容量。"""
        target_path = cache_key[0]
        stale_keys = [
            key
            for key in self._file_structure_cache.keys()
            if key[0] == target_path and key != cache_key
        ]
        for stale_key in stale_keys:
            self._file_structure_cache.pop(stale_key, None)

        self._file_structure_cache[cache_key] = (
            list(sheet_lines),
            sheet_count,
            max_total_rows,
        )
        self._file_structure_cache.move_to_end(cache_key)

        while len(self._file_structure_cache) > self._file_structure_cache_limit:
            self._file_structure_cache.popitem(last=False)

    @staticmethod
    def _guess_header_row(rows: list[list[Any]], max_scan: int = 12) -> int | None:
        """启发式猜测 header 行号（0-indexed）。

        支持多段结构：通过非空数量、文本占比、关键字命中和下一行数据特征综合评分。
        """
        if not rows:
            return None

        keywords = ("月份", "日期", "城市", "产品", "部门", "姓名", "工号", "金额", "数量", "状态", "营收", "利润")

        def _trim(row: list[Any]) -> list[Any]:
            trimmed = list(row)
            while trimmed and trimmed[-1] is None:
                trimmed.pop()
            return trimmed

        def _score(row: list[Any], row_idx: int, next_row: list[Any] | None) -> float:
            row = _trim(row)
            non_empty = [v for v in row if v is not None]
            if len(non_empty) < 3:
                return float("-inf")

            text_values = [str(v).strip() for v in non_empty if isinstance(v, str) and str(v).strip()]
            string_count = len(text_values)
            numeric_count = sum(1 for v in non_empty if isinstance(v, (int, float)))
            unique_ratio = len(set(map(str, non_empty))) / max(len(non_empty), 1)
            keyword_hits = sum(1 for text in text_values if any(k in text for k in keywords))

            score = 0.0
            score += len(non_empty) * 2.0
            score += string_count * 1.4
            score -= numeric_count * 1.2
            score += unique_ratio * 2.0
            score += keyword_hits * 2.5
            score -= row_idx * 0.03

            first = non_empty[0]
            if isinstance(first, str) and any(token in first for token in ("生成时间", "报表", "分析", "仪表盘", "──")):
                score -= 5.0

            if next_row:
                nn = [v for v in _trim(next_row) if v is not None]
                if nn:
                    next_numeric_ratio = sum(1 for v in nn if isinstance(v, (int, float))) / len(nn)
                    score += next_numeric_ratio * 1.5
            return score

        upper = min(max_scan, len(rows))
        best_idx: int | None = None
        best_score = float("-inf")
        for idx in range(upper):
            nxt = rows[idx + 1] if idx + 1 < upper else None
            score = _score(rows[idx], idx, nxt)
            if score > best_score:
                best_score = score
                best_idx = idx

        if best_idx is None or best_score == float("-inf"):
            return None
        return best_idx

    def _build_large_file_context(
        self,
        *,
        user_message: str,
        candidate_file_paths: list[str] | None,
    ) -> str:
        large_files = self._detect_large_excel_files(candidate_file_paths)
        if not large_files:
            return ""

        threshold = self._config.large_excel_threshold_bytes
        normalized_message = (user_message or "").strip()
        if normalized_message:
            user_summary = normalized_message[:120]
            if len(normalized_message) > 120:
                user_summary += "..."
        else:
            user_summary = "(空)"

        lines = [
            "[路由提示] 检测到大文件 Excel，优先采用代码方式分步处理。"
            "请直接调用推荐的工具开始处理，不要先输出处理计划。",
            f"- 用户请求：{user_summary}",
            f"- 大文件阈值：{self._format_bytes(threshold)}（{threshold} bytes）",
            "- 命中文件：",
        ]
        for file_path, file_size in large_files:
            lines.append(f"  - {file_path}（{self._format_bytes(file_size)}）")
        lines.append("- 建议优先选择 `excel_code_runner`，先抽样探查后再全量处理。")
        return "\n".join(lines)

    def _detect_large_excel_files(
        self,
        candidate_file_paths: list[str] | None,
    ) -> list[tuple[str, int]]:
        if not candidate_file_paths:
            return []

        threshold = self._config.large_excel_threshold_bytes
        if threshold <= 0:
            return []

        large_files: list[tuple[str, int]] = []
        seen_paths: set[str] = set()
        for raw_path in candidate_file_paths:
            normalized = (raw_path or "").strip()
            if not normalized:
                continue
            try:
                path_obj = Path(normalized).expanduser()
                if not path_obj.is_absolute():
                    path_obj = Path.cwd() / path_obj
                resolved = path_obj.resolve(strict=False)
            except OSError:
                continue

            normalized_resolved = str(resolved)
            if normalized_resolved in seen_paths:
                continue
            seen_paths.add(normalized_resolved)

            if resolved.suffix.lower() not in _EXCEL_SUFFIXES:
                continue

            try:
                if not resolved.is_file():
                    continue
                file_size = resolved.stat().st_size
            except OSError:
                continue

            if file_size >= threshold:
                large_files.append((normalized_resolved, file_size))
        return large_files

    @staticmethod
    def _format_bytes(size: int) -> str:
        if size < 1024:
            return f"{size} B"
        value = float(size)
        units = ["KB", "MB", "GB", "TB"]
        for unit in units:
            value /= 1024
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}"
        return f"{size} B"

    @staticmethod
    def _extract_excel_paths(text: str) -> list[str]:
        if not text:
            return []
        paths: list[str] = []
        for match in _EXCEL_PATH_PATTERN.finditer(text):
            value = next((group for group in match.groups() if group), "")
            candidate = value.strip().strip("，。！？；：,.;:()[]{}")
            if candidate:
                paths.append(candidate)
        return paths
