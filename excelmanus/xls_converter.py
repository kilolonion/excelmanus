"""XLS/XLSB → XLSX 透明转换层。

.xls (BIFF8, 1997-2003) 和 .xlsb (Excel Binary) 格式不被 openpyxl 支持。
本模块在文件入口处将它们一次性转换为 .xlsx，后续全链路统一走 openpyxl。

转换策略：
- .xls → 用 xlrd 读取 → openpyxl 写出 .xlsx（保留数据、sheet 结构、基础格式）
- .xlsb → 用 pyxlsb 读取 → openpyxl 写出 .xlsx（仅数据 + sheet 结构）
- 转换后的 .xlsx 保存在原文件同目录，文件名后缀替换
- 原始文件保留不删除
- FileRegistry 记录 original_name 保留原始文件名

限制：
- .xls 格式的图表、VBA 宏、条件格式、数据验证不会被转换
- .xlsb 仅转换数据和 sheet 结构
- 转换过程可能丢失部分复杂样式（渐变填充、自定义数字格式等）
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 需要转换的扩展名
CONVERTIBLE_EXTENSIONS = frozenset({".xls", ".xlsb"})


def needs_conversion(path: str | Path) -> bool:
    """判断文件是否需要转换为 xlsx。"""
    return Path(path).suffix.lower() in CONVERTIBLE_EXTENSIONS


def converted_xlsx_path(original: str | Path) -> Path:
    """计算转换后的 xlsx 文件路径（同目录，后缀替换）。"""
    p = Path(original)
    return p.with_suffix(".xlsx")


def convert_to_xlsx(
    src: str | Path,
    dst: str | Path | None = None,
    *,
    overwrite: bool = False,
) -> Path:
    """将 .xls/.xlsb 文件转换为 .xlsx。

    Args:
        src: 源文件路径。
        dst: 目标 .xlsx 路径，默认同目录后缀替换。
        overwrite: 是否覆盖已存在的目标文件。

    Returns:
        转换后的 .xlsx 文件路径。

    Raises:
        FileNotFoundError: 源文件不存在。
        ValueError: 不支持的文件格式。
        ConversionError: 转换过程中出错。
    """
    src = Path(src)
    if not src.exists():
        raise FileNotFoundError(f"源文件不存在: {src}")

    ext = src.suffix.lower()
    if ext not in CONVERTIBLE_EXTENSIONS:
        raise ValueError(f"不需要转换的格式: {ext}")

    if dst is None:
        dst = converted_xlsx_path(src)
    else:
        dst = Path(dst)

    if dst.exists() and not overwrite:
        logger.info("转换目标已存在，跳过: %s", dst)
        return dst

    dst.parent.mkdir(parents=True, exist_ok=True)

    if ext == ".xls":
        return _convert_xls(src, dst)
    elif ext == ".xlsb":
        return _convert_xlsb(src, dst)
    else:
        raise ValueError(f"不支持的格式: {ext}")


class ConversionError(Exception):
    """转换过程中的错误。"""
    pass


# ── .xls 转换 ──────────────────────────────────────────────


def _convert_xls(src: Path, dst: Path) -> Path:
    """用 xlrd 读取 .xls，openpyxl 写出 .xlsx。

    保留：数据、sheet 结构、基础字体/填充/对齐/边框、列宽、行高、合并单元格。
    不保留：图表、VBA、条件格式、数据验证、超链接、图片。
    """
    try:
        import xlrd
    except ImportError:
        raise ConversionError(
            "需要 xlrd 库来读取 .xls 文件。请运行: uv pip install xlrd"
        )

    from openpyxl import Workbook

    try:
        xls_wb = xlrd.open_workbook(str(src), formatting_info=True)
    except xlrd.XLRDError as e:
        raise ConversionError(f"读取 .xls 文件失败: {e}") from e

    xlsx_wb = Workbook()
    # 删除默认创建的 Sheet
    if xlsx_wb.sheetnames:
        xlsx_wb.remove(xlsx_wb.active)

    xf_list = xls_wb.xf_list
    font_list = xls_wb.font_list
    format_map = xls_wb.format_map

    for si in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(si)
        xlsx_ws = xlsx_wb.create_sheet(title=xls_ws.name)

        # 数据 + 基础格式
        for row_idx in range(xls_ws.nrows):
            for col_idx in range(xls_ws.ncols):
                cell_value = xls_ws.cell_value(row_idx, col_idx)
                cell_type = xls_ws.cell_type(row_idx, col_idx)
                xlsx_cell = xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1)

                # 类型转换
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(cell_value, xls_wb.datemode)
                        from datetime import datetime, date, time
                        if dt_tuple[0] == 0 and dt_tuple[1] == 0 and dt_tuple[2] == 0:
                            xlsx_cell.value = time(*dt_tuple[3:])
                        elif dt_tuple[3] == 0 and dt_tuple[4] == 0 and dt_tuple[5] == 0:
                            xlsx_cell.value = date(*dt_tuple[:3])
                        else:
                            xlsx_cell.value = datetime(*dt_tuple)
                    except Exception:
                        xlsx_cell.value = cell_value
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    xlsx_cell.value = bool(cell_value)
                elif cell_type == xlrd.XL_CELL_ERROR:
                    xlsx_cell.value = None
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    continue
                else:
                    xlsx_cell.value = cell_value

                # 样式迁移
                try:
                    xf_idx = xls_ws.cell_xf_index(row_idx, col_idx)
                    _apply_xls_style(
                        xlsx_cell, xf_idx,
                        xf_list, font_list, format_map, xls_wb,
                    )
                except Exception:
                    pass  # 样式迁移失败静默跳过

        # 合并单元格
        for crange in xls_ws.merged_cells:
            rlo, rhi, clo, chi = crange
            from openpyxl.utils import get_column_letter
            start = f"{get_column_letter(clo + 1)}{rlo + 1}"
            end = f"{get_column_letter(chi)}{rhi}"
            try:
                xlsx_ws.merge_cells(f"{start}:{end}")
            except Exception:
                pass

        # 列宽
        for col_idx in range(xls_ws.ncols):
            try:
                from openpyxl.utils import get_column_letter as _gcl
                col_letter = _gcl(col_idx + 1)
                # xlrd 不直接提供列宽，尝试从 colinfo_map 获取
                if hasattr(xls_ws, 'colinfo_map') and col_idx in xls_ws.colinfo_map:
                    col_info = xls_ws.colinfo_map[col_idx]
                    # xlrd 列宽单位是 1/256 字符宽度
                    width = col_info.width / 256.0
                    if width > 0:
                        xlsx_ws.column_dimensions[col_letter].width = width
            except Exception:
                pass

        # 行高
        for row_idx in range(xls_ws.nrows):
            try:
                if hasattr(xls_ws, 'rowinfo_map') and row_idx in xls_ws.rowinfo_map:
                    row_info = xls_ws.rowinfo_map[row_idx]
                    # xlrd 行高单位是 twips (1/20 point)
                    height = (row_info.height or 0) / 20.0
                    if height > 0:
                        xlsx_ws.row_dimensions[row_idx + 1].height = height
            except Exception:
                pass

    try:
        xlsx_wb.save(str(dst))
    except Exception as e:
        raise ConversionError(f"保存 .xlsx 失败: {e}") from e
    finally:
        xlsx_wb.close()

    logger.info("XLS → XLSX 转换完成: %s → %s (%d sheets)", src.name, dst.name, xls_wb.nsheets)
    return dst


def _apply_xls_style(
    xlsx_cell: Any,
    xf_idx: int,
    xf_list: list,
    font_list: list,
    format_map: dict,
    xls_wb: Any,
) -> None:
    """将 xlrd XF 记录的样式应用到 openpyxl cell。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    if xf_idx >= len(xf_list):
        return
    xf = xf_list[xf_idx]

    # 字体
    if xf.font_index < len(font_list):
        xls_font = font_list[xf.font_index]
        xlsx_cell.font = Font(
            name=xls_font.name,
            size=xls_font.height / 20 if xls_font.height else None,
            bold=xls_font.bold,
            italic=xls_font.italic,
            underline="single" if xls_font.underline_type else None,
            strike=xls_font.struck_out,
            color=_xlrd_color_to_hex(xls_font.colour_index, xls_wb),
        )

    # 对齐
    try:
        horz_map = {0: None, 1: "left", 2: "center", 3: "right", 4: "fill",
                    5: "justify", 6: "centerContinuous", 7: "distributed"}
        vert_map = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}
        xlsx_cell.alignment = Alignment(
            horizontal=horz_map.get(xf.alignment.hor_align),
            vertical=vert_map.get(xf.alignment.vert_align),
            wrap_text=bool(xf.alignment.text_wrapped),
        )
    except Exception:
        pass

    # 填充（背景色）
    try:
        bg_color = _xlrd_color_to_hex(xf.background.pattern_colour_index, xls_wb)
        if bg_color and bg_color != "000000":
            xlsx_cell.fill = PatternFill(
                start_color=bg_color,
                end_color=bg_color,
                fill_type="solid",
            )
    except Exception:
        pass

    # 数字格式
    try:
        fmt_key = xf.format_key
        if fmt_key in format_map:
            fmt_str = format_map[fmt_key].format_str
            if fmt_str and fmt_str != "General":
                xlsx_cell.number_format = fmt_str
    except Exception:
        pass


def _xlrd_color_to_hex(colour_index: int, xls_wb: Any) -> str | None:
    """xlrd colour_index → hex 颜色字符串。"""
    if colour_index is None or colour_index < 0:
        return None
    # xlrd 的标准颜色表
    try:
        colour_map = xls_wb.colour_map
        if colour_index in colour_map:
            rgb = colour_map[colour_index]
            if rgb and rgb != (0, 0, 0):
                return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
    except Exception:
        pass
    return None


# ── .xlsb 转换 ──────────────────────────────────────────────


def _convert_xlsb(src: Path, dst: Path) -> Path:
    """用 pyxlsb 读取 .xlsb，openpyxl 写出 .xlsx。

    仅转换数据和 sheet 结构，不保留样式。
    """
    try:
        from pyxlsb import open_workbook as open_xlsb
    except ImportError:
        raise ConversionError(
            "需要 pyxlsb 库来读取 .xlsb 文件。请运行: uv pip install pyxlsb"
        )

    from openpyxl import Workbook

    try:
        xlsb_wb = open_xlsb(str(src))
    except Exception as e:
        raise ConversionError(f"读取 .xlsb 文件失败: {e}") from e

    xlsx_wb = Workbook()
    if xlsx_wb.sheetnames:
        xlsx_wb.remove(xlsx_wb.active)

    try:
        for sheet_name in xlsb_wb.sheets:
            xlsx_ws = xlsx_wb.create_sheet(title=sheet_name)
            with xlsb_wb.get_sheet(sheet_name) as xlsb_ws:
                for row in xlsb_ws.rows():
                    for cell in row:
                        if cell.v is not None:
                            xlsx_ws.cell(
                                row=cell.r + 1,
                                column=cell.c + 1,
                                value=cell.v,
                            )
    except Exception as e:
        raise ConversionError(f"转换 .xlsb 内容失败: {e}") from e
    finally:
        xlsb_wb.close()

    # 确保至少有一个 sheet
    if not xlsx_wb.sheetnames:
        xlsx_wb.create_sheet(title="Sheet1")

    try:
        xlsx_wb.save(str(dst))
    except Exception as e:
        raise ConversionError(f"保存 .xlsx 失败: {e}") from e
    finally:
        xlsx_wb.close()

    logger.info("XLSB → XLSX 转换完成: %s → %s", src.name, dst.name)
    return dst


# ── 便捷入口 ────────────────────────────────────────────────


def ensure_xlsx(path: str | Path, *, overwrite: bool = False) -> tuple[Path, bool]:
    """确保路径指向 xlsx 文件。如果是 xls/xlsb 则自动转换。

    Returns:
        (实际可用的 xlsx 路径, 是否发生了转换)
    """
    p = Path(path)
    if not needs_conversion(p):
        return p, False

    xlsx_path = convert_to_xlsx(p, overwrite=overwrite)
    return xlsx_path, True
