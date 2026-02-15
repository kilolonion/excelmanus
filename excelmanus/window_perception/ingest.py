"""WURM ingest 逻辑：数据提取、合并、写入与筛选。"""

from __future__ import annotations

from collections.abc import Iterable
from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from .models import CachedRange, ChangeRecord, ColumnDef, DetailLevel, WindowState


def extract_data_rows(result_json: dict[str, Any] | None, tool_name: str) -> list[dict[str, Any]]:
    """从工具结果提取行级数据。"""
    if not isinstance(result_json, dict):
        return []

    candidates = (
        result_json.get("data"),
        result_json.get("preview"),
        result_json.get("preview_after"),
    )
    for candidate in candidates:
        rows = _normalize_rows(candidate)
        if rows:
            return rows

    # metadata-only 工具可不返回数据行
    _ = tool_name
    return []


def extract_columns(result_json: dict[str, Any] | None, rows: list[dict[str, Any]]) -> list[ColumnDef]:
    """提取并推断列信息。"""
    raw_columns = result_json.get("columns") if isinstance(result_json, dict) else None
    names: list[str] = []

    if isinstance(raw_columns, list):
        names = [str(item).strip() for item in raw_columns if str(item).strip()]
    elif rows:
        names = [str(key) for key in rows[0].keys()]

    if not names:
        return []

    inferred: list[ColumnDef] = []
    for idx, name in enumerate(names):
        values = [row.get(name) for row in rows[:50]]
        inferred.append(ColumnDef(name=name, inferred_type=_infer_type(values, fallback_idx=idx)))
    return inferred


def ingest_read_result(
    window: WindowState,
    *,
    new_range: str,
    new_rows: list[dict[str, Any]],
    iteration: int,
) -> list[int]:
    """将读取结果写入窗口并执行范围块合并。"""
    if not new_range:
        new_range = window.viewport_range or "A1:A1"

    to_merge = _collect_merge_indices(window.cached_ranges, new_range)
    if not to_merge:
        for cached in window.cached_ranges:
            cached.is_current_viewport = False
        window.cached_ranges.append(
            CachedRange(
                range_ref=new_range,
                rows=list(new_rows),
                is_current_viewport=True,
                added_at_iteration=iteration,
            )
        )
    else:
        merged_range = new_range
        merged_rows = list(new_rows)
        for idx in sorted(to_merge):
            cached = window.cached_ranges[idx]
            merged_range = union_range(merged_range, cached.range_ref)
            merged_rows = deduplicated_merge(
                cached.rows,
                merged_rows,
                existing_range=cached.range_ref,
                incoming_range=merged_range,
            )

        # 删除参与合并的旧块，追加新块，保证缓存结构稳定。
        window.cached_ranges = [
            cached
            for idx, cached in enumerate(window.cached_ranges)
            if idx not in to_merge
        ]
        for cached in window.cached_ranges:
            cached.is_current_viewport = False
        window.cached_ranges.append(
            CachedRange(
                range_ref=merged_range,
                rows=merged_rows,
                is_current_viewport=True,
                added_at_iteration=iteration,
            )
        )

    _trim_cached_ranges(window)
    window.viewport_range = new_range
    _rebuild_data_buffer(window)
    window.stale_hint = None
    window.detail_level = DetailLevel.FULL
    window.idle_turns = 0

    if not new_rows:
        return []
    start = max(0, len(window.data_buffer) - len(new_rows))
    return list(range(start, len(window.data_buffer)))


def ingest_write_result(
    window: WindowState,
    *,
    target_range: str,
    result_json: dict[str, Any] | None,
    iteration: int,
) -> list[int]:
    """处理写入结果：可映射即原地 patch，无法映射才设置 stale。"""
    updated_rows: list[int] = []
    preview_after = result_json.get("preview_after") if isinstance(result_json, dict) else None
    normalized_preview = _to_matrix(preview_after)
    column_defs = window.columns or window.schema
    if normalized_preview and column_defs:
        updated_rows = _apply_preview_patch(window, target_range=target_range, matrix=normalized_preview)

    if updated_rows:
        window.stale_hint = None
    else:
        hint_range = target_range or window.viewport_range or "当前视口"
        window.stale_hint = f"{hint_range} 已修改，依赖此区域的公式值可能已变化"
    window.detail_level = DetailLevel.FULL
    window.current_iteration = iteration
    return updated_rows


def ingest_filter_result(
    window: WindowState,
    *,
    filter_condition: dict[str, Any],
    filtered_rows: list[dict[str, Any]],
    iteration: int,
) -> list[int]:
    """处理筛选结果。"""
    if window.unfiltered_buffer is None:
        window.unfiltered_buffer = list(window.data_buffer)

    window.data_buffer = list(filtered_rows)
    window.filter_state = dict(filter_condition)
    window.cached_ranges = [
        CachedRange(
            range_ref=window.viewport_range or "A1:A1",
            rows=list(filtered_rows),
            is_current_viewport=True,
            added_at_iteration=iteration,
        )
    ]
    window.detail_level = DetailLevel.FULL
    window.current_iteration = iteration
    if not filtered_rows:
        return []
    return list(range(len(filtered_rows)))


def is_adjacent_or_overlapping(range_a: str, range_b: str) -> bool:
    """判断两个矩形范围是否连续或重叠。"""
    try:
        a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(range_a)
        b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(range_b)
    except ValueError:
        return False

    col_gap = max(0, max(a_min_col, b_min_col) - min(a_max_col, b_max_col) - 1)
    row_gap = max(0, max(a_min_row, b_min_row) - min(a_max_row, b_max_row) - 1)
    return col_gap == 0 and row_gap == 0


def union_range(range_a: str, range_b: str) -> str:
    """合并两个范围为最小包围框。"""
    a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(range_a)
    b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(range_b)
    min_col = min(a_min_col, b_min_col)
    min_row = min(a_min_row, b_min_row)
    max_col = max(a_max_col, b_max_col)
    max_row = max(a_max_row, b_max_row)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def deduplicated_merge(
    existing_rows: list[dict[str, Any]],
    incoming_rows: list[dict[str, Any]],
    *,
    existing_range: str = "",
    incoming_range: str = "",
) -> list[dict[str, Any]]:
    """按“几何+行位置”优先合并，退化到主键合并。"""
    geometry_merged = _merge_rows_by_geometry(
        existing_rows=existing_rows,
        existing_range=existing_range,
        incoming_rows=incoming_rows,
        incoming_range=incoming_range,
    )
    if geometry_merged is not None:
        return geometry_merged
    return _merge_rows_by_primary_key(
        existing_rows=existing_rows,
        incoming_rows=incoming_rows,
    )


def make_change_record(
    *,
    operation: str,
    tool_summary: str,
    affected_range: str,
    change_type: str,
    iteration: int,
    affected_row_indices: list[int],
) -> ChangeRecord:
    """构造变更记录。"""
    return ChangeRecord(
        operation=operation,
        tool_summary=tool_summary,
        affected_range=affected_range,
        change_type=change_type,
        iteration=iteration,
        affected_row_indices=affected_row_indices,
    )


def _normalize_rows(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    rows: list[dict[str, Any]] = []
    for idx, item in enumerate(value):
        if isinstance(item, dict):
            rows.append({str(k): v for k, v in item.items()})
            continue
        if isinstance(item, list):
            rows.append({f"col_{col_idx + 1}": cell for col_idx, cell in enumerate(item)})
            continue
        rows.append({"value": item, "_idx": idx})
    return rows


def _infer_type(values: Iterable[Any], *, fallback_idx: int) -> str:
    _ = fallback_idx
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            return "text"
        if isinstance(value, (int, float)):
            return "number"
        text = str(value)
        if "-" in text and len(text) >= 8:
            return "date"
        return "text"
    return "unknown"


def _trim_cached_ranges(window: WindowState) -> None:
    total_rows = sum(len(cached.rows) for cached in window.cached_ranges)
    max_rows = max(1, int(window.max_cached_rows))
    while total_rows > max_rows:
        candidates = [item for item in window.cached_ranges if not item.is_current_viewport]
        if not candidates:
            break
        oldest = min(candidates, key=lambda item: item.added_at_iteration)
        total_rows -= len(oldest.rows)
        window.cached_ranges.remove(oldest)


def _rebuild_data_buffer(window: WindowState) -> None:
    buffer: list[dict[str, Any]] = []
    for cached in window.cached_ranges:
        buffer.extend(cached.rows)
    window.data_buffer = buffer


def _to_matrix(value: Any) -> list[list[Any]]:
    if not isinstance(value, list) or not value:
        return []
    matrix: list[list[Any]] = []
    for row in value:
        if isinstance(row, list):
            matrix.append(list(row))
        else:
            matrix.append([row])
    return matrix


def _apply_preview_patch(window: WindowState, *, target_range: str, matrix: list[list[Any]]) -> list[int]:
    column_defs = window.columns or window.schema
    if not column_defs:
        return []

    touched_rows: set[int] = set()
    patched_any = False
    # 优先尝试 patch 到缓存范围，确保后续 rebuild 不丢失写入更新。
    for cached in window.cached_ranges:
        touched = _patch_matrix_to_cached_range(
            cached=cached,
            target_range=target_range,
            matrix=matrix,
            column_defs=column_defs,
        )
        if touched:
            patched_any = True
            touched_rows.update(touched)

    if patched_any:
        _rebuild_data_buffer(window)
        return [
            idx
            for idx, row in enumerate(window.data_buffer)
            if id(row) in touched_rows
        ]

    # 没有缓存映射命中时，回退到“当前视口 + data_buffer”映射。
    try:
        min_col, min_row, _, _ = range_boundaries(target_range)
        viewport_min_col, viewport_min_row, _, _ = range_boundaries(window.viewport_range or target_range)
    except ValueError:
        return []

    if not window.data_buffer:
        return []

    updated: list[int] = []
    column_names = [col.name for col in column_defs]
    row_base = min_row - viewport_min_row
    col_base = min_col - viewport_min_col

    for r_idx, row_vals in enumerate(matrix):
        buffer_index = row_base + r_idx
        if buffer_index < 0 or buffer_index >= len(window.data_buffer):
            continue
        row_obj = window.data_buffer[buffer_index]
        for c_idx, value in enumerate(row_vals):
            col_index = col_base + c_idx
            if col_index < 0 or col_index >= len(column_names):
                continue
            row_obj[column_names[col_index]] = value
        updated.append(buffer_index)
    return updated


def _collect_merge_indices(cached_ranges: list[CachedRange], new_range: str) -> set[int]:
    """收集与 new_range 连通（重叠或相邻）的缓存块索引。"""
    merged_indices: set[int] = set()
    current_union = new_range
    changed = True
    while changed:
        changed = False
        for idx, cached in enumerate(cached_ranges):
            if idx in merged_indices:
                continue
            if is_adjacent_or_overlapping(cached.range_ref, current_union):
                merged_indices.add(idx)
                current_union = union_range(current_union, cached.range_ref)
                changed = True
    return merged_indices


def _merge_rows_by_geometry(
    *,
    existing_rows: list[dict[str, Any]],
    existing_range: str,
    incoming_rows: list[dict[str, Any]],
    incoming_range: str,
) -> list[dict[str, Any]] | None:
    """按绝对行号合并两批行。"""
    if not existing_range or not incoming_range:
        return None
    try:
        _, existing_min_row, _, _ = range_boundaries(existing_range)
        _, incoming_min_row, _, _ = range_boundaries(incoming_range)
    except ValueError:
        return None

    merged_by_row: dict[int, dict[str, Any]] = {}
    order: list[int] = []
    for idx, row in enumerate(existing_rows):
        row_no = existing_min_row + idx
        merged_by_row[row_no] = dict(row)
        order.append(row_no)
    for idx, row in enumerate(incoming_rows):
        row_no = incoming_min_row + idx
        if row_no not in merged_by_row:
            merged_by_row[row_no] = dict(row)
            order.append(row_no)
        else:
            merged_by_row[row_no].update(row)
    return [merged_by_row[row_no] for row_no in sorted(set(order))]


def _merge_rows_by_primary_key(
    *,
    existing_rows: list[dict[str, Any]],
    incoming_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """按主键合并；无法识别主键时按顺序拼接。"""
    if not existing_rows:
        return list(incoming_rows)
    if not incoming_rows:
        return list(existing_rows)

    key_name = _detect_primary_key(existing_rows, incoming_rows)
    if not key_name:
        # 不做内容签名去重，避免“同值不同行”被误合并。
        return [*existing_rows, *incoming_rows]

    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in [*existing_rows, *incoming_rows]:
        raw_key = row.get(key_name)
        key = str(raw_key) if raw_key is not None else ""
        if not key:
            continue
        if key not in merged:
            order.append(key)
            merged[key] = dict(row)
        else:
            merged[key].update(row)

    # 保留无主键行，按顺序附加在末尾。
    keyless_rows = [
        dict(row)
        for row in [*existing_rows, *incoming_rows]
        if row.get(key_name) in {None, ""}
    ]
    return [merged[item] for item in order] + keyless_rows


def _detect_primary_key(
    existing_rows: list[dict[str, Any]],
    incoming_rows: list[dict[str, Any]],
) -> str:
    candidates = ("id", "ID", "Id", "row_id", "key", "主键")
    existing_keys = set().union(*(row.keys() for row in existing_rows if isinstance(row, dict)))
    incoming_keys = set().union(*(row.keys() for row in incoming_rows if isinstance(row, dict)))
    shared_keys = existing_keys & incoming_keys
    for candidate in candidates:
        if candidate in shared_keys:
            return candidate
    return ""


def _patch_matrix_to_cached_range(
    *,
    cached: CachedRange,
    target_range: str,
    matrix: list[list[Any]],
    column_defs: list[ColumnDef],
) -> set[int]:
    try:
        target_min_col, target_min_row, _, _ = range_boundaries(target_range)
        cached_min_col, cached_min_row, cached_max_col, cached_max_row = range_boundaries(cached.range_ref)
    except ValueError:
        return set()

    column_names = [col.name for col in column_defs]
    if not column_names or not cached.rows:
        return set()

    touched_ids: set[int] = set()
    for r_idx, row_vals in enumerate(matrix):
        abs_row = target_min_row + r_idx
        if abs_row < cached_min_row or abs_row > cached_max_row:
            continue
        cached_row_idx = abs_row - cached_min_row
        if cached_row_idx < 0 or cached_row_idx >= len(cached.rows):
            continue
        row_obj = cached.rows[cached_row_idx]
        row_touched = False
        for c_idx, value in enumerate(row_vals):
            abs_col = target_min_col + c_idx
            if abs_col < cached_min_col or abs_col > cached_max_col:
                continue
            col_index = abs_col - cached_min_col
            if col_index < 0 or col_index >= len(column_names):
                continue
            row_obj[column_names[col_index]] = value
            row_touched = True
        if row_touched:
            touched_ids.add(id(row_obj))

    return touched_ids
