"""窗口感知层管理器。"""

from __future__ import annotations

import asyncio
from copy import deepcopy
import logging
import re
from typing import Any, Awaitable, Callable, Literal

from openpyxl.utils.cell import get_column_letter, range_boundaries

from excelmanus.logger import get_logger
from excelmanus.memory import TokenCounter

from .adaptive import AdaptiveModeSelector
from .apply import apply_delta
from .advisor import HybridAdvisor, LifecyclePlan, RuleBasedAdvisor, WindowLifecycleAdvisor
from .advisor_context import AdvisorContext
from .budget import WindowBudgetAllocator
from .confirmation import build_confirmation_record, serialize_confirmation
from .delta import ExplorerDelta, FieldAppendDelta, FieldSetDelta, SheetReadDelta
from .domain import ExplorerWindow, SheetWindow, Window
from .extractor import (
    compute_scroll_position,
    extract_column_widths,
    extract_conditional_effects,
    extract_directory,
    extract_explorer_entries,
    extract_file_path,
    extract_freeze_panes,
    extract_merged_range_delta,
    extract_merged_ranges,
    extract_preview_rows,
    extract_range_ref,
    extract_row_heights,
    extract_shape,
    extract_sheet_dimensions,
    extract_sheet_name,
    extract_sheet_tabs,
    extract_status_bar,
    extract_style_summary,
    extract_viewport_geometry,
    is_excel_path,
    normalize_path,
    parse_json_payload,
)
from .ingest import (
    extract_columns,
    extract_data_rows,
    ingest_filter_result,
    ingest_read_result,
    ingest_write_result,
    make_change_record,
    summarize_shape,
)
from .identity import ExplorerIdentity, SheetIdentity
from .locator import LocatorReject, WindowLocator
from .models import (
    CachedRange,
    DetailLevel,
    IntentTag,
    OpEntry,
    PerceptionBudget,
    Viewport,
    WindowRenderAction,
    WindowType,
)
from .renderer import (
    build_tool_perception_payload,
    render_system_notice,
    render_tool_perception_block,
    render_window_background,
    render_window_keep,
    render_window_minimized,
)
from .repeat_detector import RepeatDetector
from .rule_registry import (
    classify_tool_meta,
    is_read_like_tool,
    is_write_like_tool,
    repeat_threshold as v2_repeat_threshold,
    resolve_intent_decision,
    task_type_from_intent,
)
from .rules import classify_tool
from .strategies import get_strategy

AsyncAdvisorRunner = Callable[
    [list[Window], str | None, PerceptionBudget, AdvisorContext],
    Awaitable[LifecyclePlan | None],
]

_RANGE_RE = re.compile(r"^[A-Za-z]+\d+(?::[A-Za-z]+\d+)?$")
_FORMULA_HINT_RE = re.compile(
    r"(=|SUMIFS\s*\(|VLOOKUP\s*\(|XLOOKUP\s*\(|INDEX\s*\(|MATCH\s*\(|IF\s*\()",
    re.IGNORECASE,
)
_REPEAT_READ_TOOLS = {
    "read_excel",
    "read_sheet",
    "filter_data",
    "focus_window_refill",
}
_WRITE_LIKE_TOOLS = {
    # Batch 1/2/3 精简：全部内置写入工具已删除，仅保留 MCP 工具名
    "write_to_sheet",
    "format_range",
}
_INTENT_USER_KEYWORDS: dict[IntentTag, tuple[str, ...]] = {
    IntentTag.AGGREGATE: ("汇总", "总计", "求和", "平均", "同比", "环比", "统计", "销量", "占比"),
    IntentTag.FORMAT: ("格式", "样式", "粗体", "颜色", "字体", "列宽", "行高", "边框", "合并", "条件格式"),
    IntentTag.VALIDATE: ("空值", "缺失", "异常", "重复", "校验", "完整性", "一致性", "脏数据"),
    IntentTag.FORMULA: ("公式", "函数", "引用", "计算错误", "VLOOKUP", "XLOOKUP", "SUMIFS"),
    IntentTag.ENTRY: ("写入", "录入", "填充", "更新", "覆盖", "新增"),
}
_INTENT_TO_TASK_TYPE: dict[IntentTag, str] = {
    IntentTag.AGGREGATE: "DATA_COMPARISON",
    IntentTag.FORMAT: "FORMAT_CHECK",
    IntentTag.VALIDATE: "ANOMALY_SEARCH",
    IntentTag.FORMULA: "FORMULA_DEBUG",
    IntentTag.ENTRY: "DATA_ENTRY",
    IntentTag.GENERAL: "GENERAL_BROWSE",
}
_INTENT_FORMAT_TOOLS = {
    # Batch 2/3 精简：仅保留 MCP 工具名
    "format_range",
}
_INTENT_AGGREGATE_TOOLS: set[str] = set()  # analyze_data: Batch 4 精简
_INTENT_VALIDATE_TOOLS = {"filter_data"}
_INTENT_ENTRY_TOOLS = {"write_to_sheet"}  # write_excel, write_cells: Batch 1 精简

logger = get_logger("window_perception.manager")


class WindowPerceptionManager:
    """维护窗口状态并生成上下文注入。"""

    _MAX_DORMANT_WINDOWS = 10

    def __init__(
        self,
        *,
        enabled: bool,
        budget: PerceptionBudget,
        adaptive_model_mode_overrides: dict[str, str] | None = None,
        advisor_mode: Literal["rules", "hybrid"] = "hybrid",
        advisor_trigger_window_count: int = 3,
        advisor_trigger_turn: int = 4,
        advisor_plan_ttl_turns: int = 2,
        intent_enabled: bool = True,
        intent_sticky_turns: int = 3,
        intent_repeat_warn_threshold: int = 2,
        intent_repeat_trip_threshold: int = 3,
        rule_engine_version: str = "v1",
    ) -> None:
        self._enabled = enabled
        self._budget = budget
        self._advisor_mode: Literal["rules", "hybrid"] = (
            advisor_mode if advisor_mode in {"rules", "hybrid"} else "hybrid"
        )
        self._advisor: WindowLifecycleAdvisor = (
            RuleBasedAdvisor() if self._advisor_mode == "rules" else HybridAdvisor()
        )
        self._windows: dict[str, Window] = {}
        self._locator = WindowLocator()
        self._active_window_id: str | None = None
        self._seq: int = 0
        self._operation_seq: int = 0
        self._notice_turn: int = 0
        self._last_notice_operation_seq: int = 0
        self._last_window_count: int = 0
        self._turn_hint_is_new_task: bool = False
        self._turn_hint_user_intent_summary: str = ""
        self._turn_hint_agent_recent_output: str = ""
        self._advisor_runner: AsyncAdvisorRunner | None = None
        self._advisor_task: asyncio.Task[None] | None = None
        self._cached_small_model_plan: LifecyclePlan | None = None
        self._advisor_trigger_window_count = max(1, int(advisor_trigger_window_count))
        self._advisor_trigger_turn = max(1, int(advisor_trigger_turn))
        self._advisor_plan_ttl_turns = max(0, int(advisor_plan_ttl_turns))
        self._intent_enabled = bool(intent_enabled)
        self._intent_sticky_turns = max(1, int(intent_sticky_turns))
        self._intent_repeat_warn_threshold = max(1, int(intent_repeat_warn_threshold))
        self._intent_repeat_trip_threshold = max(
            self._intent_repeat_warn_threshold + 1,
            int(intent_repeat_trip_threshold),
        )
        normalized_rule_version = str(rule_engine_version or "v1").strip().lower()
        self._rule_engine_version = (
            normalized_rule_version
            if normalized_rule_version in {"v1", "v2"}
            else "v1"
        )
        self._repeat_detector = RepeatDetector()
        self._adaptive_selector = AdaptiveModeSelector(
            model_mode_overrides=adaptive_model_mode_overrides or {},
        )
        self._turn_hint_intent_hint: str = ""
        self._turn_hint_task_tags: tuple[str, ...] = ()
        self._last_identity_reject_code: str | None = None

    @property
    def enabled(self) -> bool:
        """是否启用窗口感知层。"""
        return self._enabled

    @property
    def last_identity_reject_code(self) -> str | None:
        """最近一次 identity/locator 显式拒绝码。"""
        return self._last_identity_reject_code

    def bind_async_advisor_runner(self, runner: AsyncAdvisorRunner | None) -> None:
        """绑定异步小模型顾问回调。"""
        self._advisor_runner = runner

    def set_turn_hints(
        self,
        *,
        is_new_task: bool,
        user_intent_summary: str = "",
        agent_recent_output: str = "",
        turn_intent_hint: str = "",
        task_tags: tuple[str, ...] | None = None,
    ) -> None:
        """设置当前轮次提示信息。"""
        self._turn_hint_is_new_task = bool(is_new_task)
        self._turn_hint_user_intent_summary = self._normalize_hint(user_intent_summary, max_chars=200)
        self._turn_hint_agent_recent_output = self._normalize_hint(agent_recent_output, max_chars=200)
        self._turn_hint_intent_hint = self._normalize_hint(turn_intent_hint, max_chars=200)
        if task_tags is not None:
            self._turn_hint_task_tags = self._normalize_task_tags(task_tags)
        elif is_new_task:
            self._turn_hint_task_tags = ()

    def reset(self) -> None:
        """重置状态。"""
        self._cancel_advisor_task()
        self._windows.clear()
        self._locator = WindowLocator()
        self._active_window_id = None
        self._seq = 0
        self._operation_seq = 0
        self._notice_turn = 0
        self._last_notice_operation_seq = 0
        self._last_window_count = 0
        self._turn_hint_is_new_task = False
        self._turn_hint_user_intent_summary = ""
        self._turn_hint_agent_recent_output = ""
        self._turn_hint_intent_hint = ""
        self._turn_hint_task_tags = ()
        self._cached_small_model_plan = None
        self._repeat_detector = RepeatDetector()
        self._adaptive_selector.reset()
        self._last_identity_reject_code = None

    def observe_subagent_context(
        self,
        *,
        candidate_paths: list[str],
        subagent_name: str,
        task: str,
    ) -> None:
        """记录子代理观察到的文件。"""
        if not self._enabled:
            return
        self._operation_seq += 1
        clean_task = " ".join(task.strip().split())
        summary = f"由{subagent_name}观察: {clean_task}" if clean_task else f"由{subagent_name}观察"
        for raw in candidate_paths:
            normalized = normalize_path(raw)
            if not normalized or not is_excel_path(normalized):
                continue
            window = self._find_sheet_window(
                file_path=normalized,
                sheet_name="",
            )
            if window is None:
                window = SheetWindow.new(
                    id=self._new_id("sheet"),
                    title=normalized,
                    file_path=normalized,
                    sheet_name="",
                )
                self._set_window_field(window, "summary", summary)
                self._windows[window.id] = window
                self._register_window_identity(window)
            self._wake_window(window)
            self._set_window_field(window, "summary", summary)
            self._touch(window)

    def observe_subagent_writes(
        self,
        *,
        structured_changes: list[Any],
        subagent_name: str,
        task: str,
        iteration: int = 0,
    ) -> None:
        """子代理写入后，将影响同步到 window 系统。

        与 observe_code_execution 类似，标记 stale + 清缓存 + 追加 change_log。
        structured_changes 为 SubagentFileChange 列表（duck-typed：需有 path/tool_name/change_type/sheets_affected）。
        """
        if not self._enabled or not structured_changes:
            return
        self._operation_seq += 1
        clean_task = " ".join(task.strip().split()) if task else ""

        affected_count = 0
        for change in structured_changes:
            raw_path = getattr(change, "path", "") or ""
            normalized = normalize_path(raw_path)
            if not normalized or not is_excel_path(normalized):
                continue

            tool_name = getattr(change, "tool_name", "") or "subagent"
            change_type = getattr(change, "change_type", "write") or "write"

            stale_reason = (
                f"subagent({subagent_name}) 通过 {tool_name} 修改此文件，缓存已清空。"
                "请调用 read_excel 刷新数据。"
            )

            window = self._find_sheet_window(file_path=normalized, sheet_name="")
            if window is None:
                for win in self._windows.values():
                    if win.type != WindowType.SHEET:
                        continue
                    if normalize_path(win.file_path or "") == normalized:
                        window = win
                        break

            if window is None:
                window = SheetWindow.new(
                    id=self._new_id("sheet"),
                    title=normalized,
                    file_path=normalized,
                    sheet_name="",
                )
                self._windows[window.id] = window
                self._register_window_identity(window)

            # 标记 stale + 清空数据缓存
            self._set_window_field(window, "stale_hint", stale_reason)
            self._set_window_field(window, "data_buffer", [])
            self._set_window_field(window, "preview_rows", [])
            self._set_window_field(window, "cached_ranges", [])
            self._set_window_field(window, "unfiltered_buffer", None)

            # 追加 change_log
            tool_summary = f"subagent({subagent_name}).{tool_name}"
            if clean_task:
                tool_summary = f"{tool_summary}: {clean_task}"
            self._append_change(
                window,
                make_change_record(
                    operation="subagent_write",
                    tool_summary=tool_summary,
                    affected_range="-",
                    change_type=change_type,
                    iteration=iteration,
                ),
            )

            self._set_window_field(window, "summary", stale_reason)
            self._repeat_detector.reset_for_file(normalized)
            self._wake_window(window)
            self._touch(window)
            affected_count += 1

        if affected_count:
            logger.info(
                "observe_subagent_writes: %d files affected by subagent %s",
                affected_count,
                subagent_name,
            )

    def observe_code_execution(
        self,
        *,
        code: str,
        audit_changes: list[Any] | None,
        stdout_tail: str,
        iteration: int,
    ) -> None:
        """run_code 执行后，将影响同步到 window 系统。

        三层信息提取：
        - Layer A: AST 提取脚本中的 Excel 目标文件
        - Layer B: 审计 diff 提取实际变化的文件（ground truth）
        - Layer C: stdout 解析提取操作摘要
        """
        if not self._enabled:
            return

        # ── Layer A: AST 提取 ──
        from excelmanus.security.code_policy import extract_excel_targets
        ast_targets = extract_excel_targets(code or "")
        ast_write_files: set[str] = set()
        for target in ast_targets:
            if target.operation in ("write", "unknown"):
                normalized = normalize_path(target.file_path)
                if normalized and is_excel_path(normalized):
                    ast_write_files.add(normalized)

        # ── Layer B: 审计 diff ──
        audit_excel_files: set[str] = set()
        if audit_changes:
            for change in audit_changes:
                path = getattr(change, "path", None) or ""
                if not path:
                    continue
                normalized = normalize_path(path)
                if normalized and is_excel_path(normalized):
                    audit_excel_files.add(normalized)

        # ── Layer C: stdout 摘要 ──
        execution_summary = self._parse_code_execution_summary(stdout_tail or "")

        # ── 合并受影响文件 ──
        affected_files = audit_excel_files | ast_write_files
        if not affected_files:
            return

        self._operation_seq += 1

        for file_path in affected_files:
            in_audit = file_path in audit_excel_files
            stale_reason = (
                f"run_code {'已' if in_audit else '可能'}修改此文件，缓存已清空。"
                "请调用 read_excel 刷新数据。"
            )
            if execution_summary:
                stale_reason = f"{stale_reason}（{execution_summary}）"

            # 查找或创建 SheetWindow
            window = self._find_sheet_window(file_path=file_path, sheet_name="")
            if window is None:
                # 按 file_path 模糊查找（忽略 sheet_name）
                for win in self._windows.values():
                    if win.type != WindowType.SHEET:
                        continue
                    if normalize_path(win.file_path or "") == file_path:
                        window = win
                        break

            if window is None and in_audit:
                # 仅对审计确认变化的文件创建新 window
                window = SheetWindow.new(
                    id=self._new_id("sheet"),
                    title=file_path,
                    file_path=file_path,
                    sheet_name="",
                )
                self._windows[window.id] = window
                self._register_window_identity(window)

            if window is None:
                continue

            # 标记 stale + 清空数据缓存
            self._set_window_field(window, "stale_hint", stale_reason)
            self._set_window_field(window, "data_buffer", [])
            self._set_window_field(window, "preview_rows", [])
            self._set_window_field(window, "cached_ranges", [])
            self._set_window_field(window, "unfiltered_buffer", None)

            # 追加 change_log
            self._append_change(
                window,
                make_change_record(
                    operation="code_execution",
                    tool_summary=f"run_code: {execution_summary or '代码执行'}",
                    affected_range="-",
                    change_type="code_modified",
                    iteration=iteration,
                ),
            )

            # 更新 summary
            self._set_window_field(window, "summary", stale_reason)

            # 重置 repeat detector
            self._repeat_detector.reset_for_file(file_path)

            # 唤醒 + touch
            self._wake_window(window)
            self._touch(window)

        logger.info(
            "observe_code_execution: %d files affected (%d from audit, %d from AST)",
            len(affected_files),
            len(audit_excel_files),
            len(ast_write_files),
        )

    def observe_write_tool_call(self, *, tool_name: str, arguments: dict[str, Any]) -> None:
        """写工具兜底：标记关联窗口 stale 并清空缓存。"""
        if not self._enabled:
            return

        normalized_tool = str(tool_name or "").strip()
        if not normalized_tool:
            return

        window = self._locate_sheet_window_for_write(arguments)
        if window is None:
            return

        target_range = str(
            arguments.get("range")
            or arguments.get("cell_range")
            or arguments.get("cell")
            or window.viewport_range
            or "当前视口"
        ).strip()
        stale_reason = f"{normalized_tool} 已写入 {target_range}，缓存已清空。请调用 read_excel 刷新数据。"

        self._set_window_field(window, "stale_hint", stale_reason)
        self._set_window_field(window, "data_buffer", [])
        self._set_window_field(window, "preview_rows", [])
        self._set_window_field(window, "cached_ranges", [])
        self._set_window_field(window, "unfiltered_buffer", None)
        self._set_window_field(window, "last_op_kind", "write")
        self._set_window_field(window, "last_write_range", target_range)
        self._set_window_field(window, "summary", stale_reason)
        self._reset_repeat_counter_after_write(window)
        self._wake_window(window)
        self._touch(window)

    def _locate_sheet_window_for_write(self, arguments: dict[str, Any]) -> Window | None:
        file_path = normalize_path(extract_file_path(arguments, None))
        sheet_name = str(extract_sheet_name(arguments, None) or "").strip()
        lowered_sheet_name = sheet_name.lower()

        if file_path and sheet_name:
            matched = self._find_sheet_window(file_path=file_path, sheet_name=sheet_name)
            if matched is not None:
                return matched

        if self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and active.type == WindowType.SHEET and not active.dormant:
                active_file = normalize_path(active.file_path or "")
                active_sheet = str(active.sheet_name or "").strip().lower()
                file_ok = (not file_path) or (active_file == file_path)
                sheet_ok = (not sheet_name) or (active_sheet == lowered_sheet_name)
                if file_ok and sheet_ok:
                    return active

        if not file_path:
            return None

        candidates: list[Window] = []
        for item in self._windows.values():
            if item.type != WindowType.SHEET or item.dormant:
                continue
            if normalize_path(item.file_path or "") != file_path:
                continue
            if sheet_name and str(item.sheet_name or "").strip().lower() != lowered_sheet_name:
                continue
            candidates.append(item)

        if not candidates:
            return None
        return sorted(candidates, key=lambda item: item.last_access_seq, reverse=True)[0]

    @staticmethod
    def _parse_code_execution_summary(stdout_tail: str) -> str:
        """从 stdout 中提取操作摘要（best-effort）。"""
        if not stdout_tail or not stdout_tail.strip():
            return ""
        import re as _re
        # 尝试提取行数信息
        row_match = _re.search(r"(\d+)\s*(?:rows?|行)", stdout_tail)
        row_hint = f"{row_match.group(1)}行" if row_match else ""
        # 取最后一行非空文本作为摘要
        lines = [line.strip() for line in stdout_tail.strip().splitlines() if line.strip()]
        last_line = lines[-1] if lines else ""
        if len(last_line) > 100:
            last_line = last_line[:100] + "..."
        parts = [p for p in [row_hint, last_line] if p]
        return "; ".join(parts) if parts else ""

    def build_system_notice(self, *, mode: str = "enriched", model_id: str = "") -> str:
        """构建系统注入窗口快照。"""
        if not self._enabled:
            return ""
        effective_mode = self.resolve_effective_mode(
            requested_mode=mode,
            model_id=model_id,
        )
        self._notice_turn += 1
        self._age_windows()
        self._refresh_active_window()
        self._recycle_idle_windows()

        active_windows = [item for item in self._windows.values() if not item.dormant]
        if not active_windows:
            self._last_notice_operation_seq = self._operation_seq
            self._last_window_count = 0
            return ""

        is_new_task = self._turn_hint_is_new_task
        self._turn_hint_is_new_task = False
        context = AdvisorContext(
            turn_number=self._notice_turn,
            is_new_task=is_new_task,
            window_count_changed=len(active_windows) != self._last_window_count,
            user_intent_summary=self._turn_hint_user_intent_summary,
            agent_recent_output=self._turn_hint_agent_recent_output,
            task_type=self._task_type_from_windows(active_windows),
            task_tags=self._turn_hint_task_tags,
        )
        cached_small_model_plan = self._get_fresh_small_model_plan(current_turn=context.turn_number)
        lifecycle_plan = self._advisor.advise(
            windows=active_windows,
            active_window_id=self._active_window_id,
            budget=self._budget,
            context=context,
            small_model_plan=cached_small_model_plan,
            plan_ttl_turns=self._advisor_plan_ttl_turns,
        )
        self._log_lifecycle_reason_codes(lifecycle_plan)

        allocator = WindowBudgetAllocator(self._budget)
        full_rows = allocator.compute_window_full_max_rows(len(active_windows))
        snapshots = allocator.allocate(
            windows=active_windows,
            active_window_id=self._active_window_id,
            render_keep=lambda window: render_window_keep(
                window,
                detail_level=window.detail_level,
                mode=effective_mode,
                max_rows=full_rows,
                current_iteration=window.current_iteration,
                intent_profile=self._build_intent_profile(window, level="full"),
            ),
            render_background=lambda window: render_window_background(
                window,
                intent_profile=self._build_intent_profile(window, level="summary"),
            ),
            render_minimized=lambda window: render_window_minimized(
                window,
                intent_profile=self._build_intent_profile(window, level="icon"),
            ),
            lifecycle_plan=lifecycle_plan,
        )
        visible = [item for item in snapshots if item.action != WindowRenderAction.CLOSE]

        for item in snapshots:
            if item.action != WindowRenderAction.CLOSE:
                continue
            self._mark_window_dormant(item.window_id)
        self._evict_dormant_windows()

        self._last_notice_operation_seq = self._operation_seq
        final_active_windows = [item for item in self._windows.values() if not item.dormant]
        # 用 evict 后的最终窗口数修正 window_count_changed，保持与 active_windows 一致
        context.window_count_changed = len(final_active_windows) != self._last_window_count
        self._last_window_count = len(final_active_windows)
        self._schedule_async_advisor(
            active_windows=final_active_windows,
            context=context,
        )

        return render_system_notice(visible, mode=effective_mode)

    def resolve_effective_mode(self, *, requested_mode: str, model_id: str = "") -> str:
        """解析本次调用应使用的实际 mode。"""
        normalized = str(requested_mode or "enriched").strip().lower()
        if normalized == "adaptive":
            return self._adaptive_selector.select_mode(
                model_id=model_id,
                requested_mode=normalized,
            )
        if normalized in {"unified", "anchored", "enriched"}:
            return normalized
        return "enriched"

    def enrich_tool_result(
        self,
        *,
        tool_name: str,
        arguments: dict[str, Any],
        result_text: str,
        success: bool,
        mode: str = "enriched",
        model_id: str = "",
    ) -> str:
        """增强工具返回。"""
        if not self._enabled or not success:
            return result_text

        effective_mode = self.resolve_effective_mode(
            requested_mode=mode,
            model_id=model_id,
        )

        if effective_mode in {"anchored", "unified"}:
            return self.ingest_and_confirm(
                tool_name=tool_name,
                arguments=arguments,
                result_text=result_text,
                success=success,
                mode=effective_mode,
                requested_mode=mode,
                model_id=model_id,
            )

        payload = self.update_from_tool_call(
            tool_name=tool_name,
            arguments=arguments,
            result_text=result_text,
        )
        if payload is None:
            return result_text

        block = render_tool_perception_block(payload)
        if not block:
            return result_text
        block = self._truncate_tool_append(block)
        if not block.strip():
            return result_text
        return f"{result_text}\n\n{block}" if result_text.strip() else block

    def ingest_and_confirm(
        self,
        *,
        tool_name: str,
        arguments: dict[str, Any],
        result_text: str,
        success: bool,
        mode: str = "anchored",
        requested_mode: str = "anchored",
        model_id: str = "",
    ) -> str:
        """WURM 路径：ingest + anchored 确认，异常时原子回退 enriched。"""
        if not self._enabled or not success:
            return result_text

        classification = self._classify_tool(tool_name)
        if classification.window_type is None:
            return result_text

        parsed = parse_json_payload(result_text)
        result_json = parsed if isinstance(parsed, dict) else None
        repeat_warning = False
        canonical_name = classification.canonical_name or tool_name
        is_adaptive_requested = str(requested_mode or "").strip().lower() == "adaptive"
        payload: dict[str, Any] | None = None
        window: Window | None = None

        try:
            payload = self.update_from_tool_call(
                tool_name=tool_name,
                arguments=arguments,
                result_text=result_text,
            )
            if payload is None:
                return result_text

            window = self._locate_window_by_identity(
                window_type=classification.window_type,
                arguments=arguments,
                result_json=result_json,
            )
            if window is None:
                return self._enriched_fallback(
                    tool_name=tool_name,
                    arguments=arguments,
                    result_text=result_text,
                    success=success,
                    payload=payload,
                )

            intent_tag = window.intent_tag

            if self._is_repeat_read_tool(canonical_name):
                try:
                    repeat_identity = self._extract_repeat_identity(
                        arguments=arguments,
                        result_json=result_json,
                        window=window,
                    )
                    if repeat_identity is not None:
                        self._repeat_detector.record_read(
                            *repeat_identity,
                            intent_tag=intent_tag.value,
                        )
                except Exception:
                    pass
                # 仅记录用于审计，不阻断或降级。
                # 重新读取是 agent 验证写入结果、消除幻觉的主要机制，不应干预。

            self._apply_delta_pipeline(
                window=window,
                canonical_tool_name=canonical_name,
                arguments=arguments,
                result_json=result_json,
            )
            if self._is_write_like_tool(canonical_name):
                self._reset_repeat_counter_after_write(window)

        except Exception:
            if is_adaptive_requested:
                self._adaptive_selector.mark_ingest_failure()
            logger.warning(
                "window.ingest_and_confirm ingest failed: tool=%s window_type=%s adaptive=%s",
                tool_name,
                classification.window_type.value if classification.window_type else "-",
                is_adaptive_requested,
                exc_info=True,
            )
            return self._enriched_fallback(
                tool_name=tool_name,
                arguments=arguments,
                result_text=result_text,
                success=success,
                payload=payload,
            )

        if is_adaptive_requested:
            self._adaptive_selector.mark_ingest_success()

        try:
            # 策略分发：仅 explorer 使用策略的 inline confirmation
            strategy = get_strategy(classification.window_type)
            if (
                strategy is not None
                and classification.window_type == WindowType.EXPLORER
                and strategy.should_replace_result()
                and window is not None
            ):
                return strategy.build_inline_confirmation(
                    window=window,
                    tool_name=canonical_name,
                    result_json=result_json,
                )

            if window is None:
                return result_text

            return self.generate_confirmation(
                window=window,
                tool_name=canonical_name,
                mode=mode,
                repeat_warning=repeat_warning,
            )
        except Exception:
            logger.warning(
                "window.ingest_and_confirm confirmation failed, fallback to enriched: "
                "tool=%s window_type=%s mode=%s",
                tool_name,
                classification.window_type.value if classification.window_type else "-",
                mode,
                exc_info=True,
            )
            return self._enriched_fallback(
                tool_name=tool_name,
                arguments=arguments,
                result_text=result_text,
                success=success,
                payload=payload,
            )

    def update_from_tool_call(
        self,
        *,
        tool_name: str,
        arguments: dict[str, Any],
        result_text: str,
    ) -> dict[str, Any] | None:
        """根据工具调用更新窗口状态并返回感知 payload。"""
        if not self._enabled:
            return None

        classification = self._classify_tool(tool_name)
        if classification.window_type is None:
            return None

        parsed = parse_json_payload(result_text)
        result_json = parsed if isinstance(parsed, dict) else None

        self._operation_seq += 1

        if classification.window_type == WindowType.EXPLORER:
            window = self._update_explorer_window(
                arguments=arguments,
                result_json=result_json,
            )
            return build_tool_perception_payload(window)

        window = self._update_sheet_window(
            canonical_tool_name=classification.canonical_name,
            arguments=arguments,
            result_json=result_json,
        )
        intent_decision = self._resolve_window_intent(
            window=window,
            canonical_tool_name=classification.canonical_name or tool_name,
            arguments=arguments,
            result_json=result_json,
        )
        self._apply_window_intent(
            window=window,
            tag=intent_decision["tag"],
            confidence=intent_decision["confidence"],
            source=intent_decision["source"],
            force=intent_decision["force"],
        )
        return build_tool_perception_payload(window)

    def generate_confirmation(
        self,
        *,
        window: Window,
        tool_name: str,
        mode: str = "anchored",
        repeat_warning: bool = False,
    ) -> str:
        """生成 anchored/unified 模式工具确认文本。"""
        record = build_confirmation_record(
            window=window,
            tool_name=tool_name,
            repeat_warning=repeat_warning,
        )
        return serialize_confirmation(record, mode=mode)

    def _enriched_fallback(
        self,
        *,
        tool_name: str,
        arguments: dict[str, Any],
        result_text: str,
        success: bool,
        payload: dict[str, Any] | None = None,
    ) -> str:
        """ingest 失败时回退到 enriched 逻辑。"""
        if payload is None:
            payload = self.update_from_tool_call(
                tool_name=tool_name,
                arguments=arguments,
                result_text=result_text,
            )
        if payload is None:
            return result_text
        block = render_tool_perception_block(payload)
        if not block:
            return result_text
        block = self._truncate_tool_append(block)
        if not block.strip():
            return result_text
        return f"{result_text}\n\n{block}" if result_text.strip() else block

    def _locate_window_by_identity(
        self,
        *,
        window_type: WindowType,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> Window | None:
        """Locate window by strong identity, then minimal active fallback."""

        if window_type == WindowType.EXPLORER:
            directory = normalize_path(extract_directory(arguments, result_json)) or "."
            identity = ExplorerIdentity(directory_norm=directory)
            try:
                window_id = self._locator.find(identity, expected_kind=WindowType.EXPLORER.value)
            except LocatorReject as exc:
                self._last_identity_reject_code = exc.code
                return None
            if window_id:
                return self._windows.get(window_id)
            return self._resolve_target_window(window_type, arguments, result_json)

        file_path = normalize_path(extract_file_path(arguments, result_json))
        sheet_name = str(extract_sheet_name(arguments, result_json) or "").strip()
        if not file_path and self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and active.type == WindowType.SHEET:
                file_path = active.file_path or ""
                if not sheet_name:
                    sheet_name = active.sheet_name or ""

        if file_path and sheet_name:
            identity = SheetIdentity(
                file_path_norm=file_path,
                sheet_name_norm=sheet_name.lower(),
            )
            try:
                window_id = self._locator.find(identity, expected_kind=WindowType.SHEET.value)
            except LocatorReject as exc:
                self._last_identity_reject_code = exc.code
                return None
            if window_id:
                return self._windows.get(window_id)
        return self._resolve_target_window(window_type, arguments, result_json)

    def _resolve_target_window(
        self,
        window_type: WindowType,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> Window | None:
        if window_type == WindowType.EXPLORER:
            directory = normalize_path(extract_directory(arguments, result_json)) or "."
            window = self._find_explorer_window(directory=directory)
            if window is not None:
                return window
            for wid, win in self._windows.items():
                if win.type == WindowType.EXPLORER and not win.dormant:
                    return win
            return None

        file_path = normalize_path(extract_file_path(arguments, result_json))
        sheet_name = str(extract_sheet_name(arguments, result_json) or "").strip()
        if not file_path and self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and active.type == WindowType.SHEET:
                file_path = active.file_path or ""
                if not sheet_name:
                    sheet_name = active.sheet_name or ""
        window = self._find_sheet_window(file_path=file_path, sheet_name=sheet_name)
        if window is not None:
            return window
        if self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and active.type == WindowType.SHEET:
                return active
        return None

    def _find_explorer_window(self, *, directory: str) -> ExplorerWindow | None:
        normalized = normalize_path(directory) or "."
        for window in self._windows.values():
            if window.type != WindowType.EXPLORER:
                continue
            if normalize_path(window.directory or ".") == normalized:
                return window  # type: ignore[return-value]
        return None

    def _find_sheet_window(self, *, file_path: str, sheet_name: str) -> SheetWindow | None:
        normalized_file = normalize_path(file_path or "")
        normalized_sheet = str(sheet_name or "").strip().lower()
        for window in self._windows.values():
            if window.type != WindowType.SHEET:
                continue
            if normalize_path(window.file_path or "") != normalized_file:
                continue
            if str(window.sheet_name or "").strip().lower() != normalized_sheet:
                continue
            return window  # type: ignore[return-value]
        return None

    def _apply_delta_pipeline(
        self,
        *,
        window: Window,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> None:
        """Apply kind-guarded delta contract then ingest mutable state."""

        delta = self._build_window_delta(
            window=window,
            canonical_tool_name=canonical_tool_name,
            arguments=arguments,
            result_json=result_json,
        )
        apply_delta(window, delta)
        self._apply_ingest(
            window=window,
            canonical_tool_name=canonical_tool_name,
            arguments=arguments,
            result_json=result_json,
        )

    def _build_window_delta(
        self,
        *,
        window: Window,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> ExplorerDelta | SheetReadDelta:
        if window.type == WindowType.EXPLORER:
            directory = normalize_path(extract_directory(arguments, result_json)) or (window.directory or ".")
            return ExplorerDelta(directory=directory)

        rows = extract_data_rows(result_json, canonical_tool_name)
        columns = extract_columns(result_json, rows)
        explicit_rows, explicit_cols = extract_shape(result_json)
        row_count, col_count = summarize_shape(
            rows,
            columns,
            explicit_rows=explicit_rows,
            explicit_cols=explicit_cols,
        )
        return SheetReadDelta(
            range_ref=extract_range_ref(
                arguments,
                default_rows=self._budget.default_rows,
                default_cols=self._budget.default_cols,
            ),
            rows=row_count,
            cols=col_count,
            change_summary=canonical_tool_name,
        )

    @staticmethod
    def _window_kind(window: Window) -> Literal["explorer", "sheet"]:
        return "explorer" if window.type == WindowType.EXPLORER else "sheet"

    def _set_window_field(self, window: Window, field: str, value: Any) -> None:
        apply_delta(
            window,
            FieldSetDelta(
                kind=self._window_kind(window),
                field=field,
                value=value,
            ),
        )

    def _append_window_field(self, window: Window, field: str, value: Any) -> None:
        apply_delta(
            window,
            FieldAppendDelta(
                kind=self._window_kind(window),
                field=field,
                value=value,
            ),
        )

    def _register_window_identity(self, window: Window) -> None:
        if window.type == WindowType.EXPLORER:
            directory = normalize_path(window.directory or "") or "."
            try:
                self._locator.register(window.id, ExplorerIdentity(directory_norm=directory))
            except LocatorReject as exc:
                self._last_identity_reject_code = exc.code
            return
        file_path = normalize_path(window.file_path or "")
        sheet_name = str(window.sheet_name or "").strip().lower()
        if file_path and sheet_name:
            try:
                self._locator.register(
                    window.id,
                    SheetIdentity(file_path_norm=file_path, sheet_name_norm=sheet_name),
                )
            except LocatorReject as exc:
                self._last_identity_reject_code = exc.code

    def _apply_ingest(
        self,
        *,
        window: Window,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> None:
        """根据工具类别将结果写入 WURM 数据容器。"""
        iteration = self._operation_seq
        self._set_window_field(window, "current_iteration", iteration)

        if window.type == WindowType.EXPLORER:
            # 委托给 ExplorerStrategy
            strategy = get_strategy(WindowType.EXPLORER)
            if strategy is not None:
                strategy.apply_ingest(
                    window=window,
                    tool_name=canonical_tool_name,
                    arguments=arguments,
                    result_json=result_json,
                    iteration=iteration,
                )
                return
            # 回退：旧逻辑
            self._append_operation(window, canonical_tool_name, arguments, True)
            self._append_change(
                window,
                make_change_record(
                    operation="enrich",
                    tool_summary=f"{canonical_tool_name}",
                    affected_range="-",
                    change_type="enriched",
                    iteration=iteration,
                    affected_row_indices=[],
                ),
            )
            return

        rows = extract_data_rows(result_json, canonical_tool_name)
        columns = extract_columns(result_json, rows)
        if columns:
            self._set_window_field(window, "columns", columns)
            self._set_window_field(window, "schema", list(columns))
        self._sync_window_schema_columns(window)

        range_ref = extract_range_ref(
            arguments,
            default_rows=self._budget.default_rows,
            default_cols=self._budget.default_cols,
        )
        self._set_window_field(window, "viewport_range", range_ref)
        self._set_window_field(
            window,
            "max_cached_rows",
            max(1, int(self._budget.window_data_buffer_max_rows)),
        )

        if canonical_tool_name in {"filter_data"}:
            # 适配多条件模式：优先使用 conditions 数组
            if arguments.get("conditions"):
                filter_condition = {
                    "conditions": arguments["conditions"],
                    "logic": arguments.get("logic", "and"),
                }
            else:
                filter_condition = {
                    "column": arguments.get("column"),
                    "operator": arguments.get("operator"),
                    "value": arguments.get("value"),
                }
            affected = ingest_filter_result(
                window,
                filter_condition=filter_condition,
                filtered_rows=rows,
                iteration=iteration,
            )
            change = make_change_record(
                operation="filter",
                tool_summary=f"{canonical_tool_name}({filter_condition})",
                affected_range=range_ref,
                change_type="filtered",
                iteration=iteration,
                affected_row_indices=affected,
            )
        elif canonical_tool_name in {
            # Batch 1/2/3 精简：仅保留 MCP 工具名
            "write_to_sheet",
            "format_range",
        }:
            target_range = str(
                arguments.get("range")
                or arguments.get("cell_range")
                or arguments.get("cell")
                or range_ref
            ).strip()
            affected = ingest_write_result(
                window,
                target_range=target_range,
                result_json=result_json,
                iteration=iteration,
            )
            change = make_change_record(
                operation="write",
                tool_summary=f"{canonical_tool_name}({target_range})",
                affected_range=target_range,
                change_type="modified",
                iteration=iteration,
                affected_row_indices=affected,
            )
        else:
            affected = ingest_read_result(
                window,
                new_range=range_ref,
                new_rows=rows,
                iteration=iteration,
            )
            change = make_change_record(
                operation="read",
                tool_summary=f"{canonical_tool_name}({range_ref})",
                affected_range=range_ref,
                change_type="added" if rows else "enriched",
                iteration=iteration,
                affected_row_indices=affected,
            )

        if window.viewport is not None:
            self._set_window_field(window, "total_rows", window.viewport.total_rows)
            self._set_window_field(window, "total_cols", window.viewport.total_cols)
        if window.total_rows <= 0:
            self._set_window_field(window, "total_rows", len(window.data_buffer))
        if window.total_cols <= 0:
            self._set_window_field(window, "total_cols", len(window.columns or window.schema))
        # 设置操作类型标记，供渲染器聚焦渲染使用
        if change.operation == "write":
            self._set_window_field(window, "last_op_kind", "write")
            self._set_window_field(window, "last_write_range", change.affected_range)
        elif change.operation == "filter":
            self._set_window_field(window, "last_op_kind", "filter")
            self._set_window_field(window, "last_write_range", None)
        else:
            self._set_window_field(window, "last_op_kind", "read")
            self._set_window_field(window, "last_write_range", None)

        self._set_window_field(window, "detail_level", DetailLevel.FULL)
        self._append_operation(window, canonical_tool_name, arguments, True)
        self._append_change(window, change)

    def _append_operation(
        self,
        window: Window,
        tool_name: str,
        arguments: dict[str, Any],
        success: bool,
    ) -> None:
        self._append_window_field(
            window,
            "operation_history",
            OpEntry(
                tool_name=tool_name,
                arguments=dict(arguments),
                iteration=window.current_iteration,
                success=success,
            ),
        )
        max_entries = max(1, int(window.max_history_entries))
        if len(window.operation_history) > max_entries:
            self._set_window_field(window, "operation_history", window.operation_history[-max_entries:])

    def _append_change(self, window: Window, record) -> None:
        self._append_window_field(window, "change_log", record)
        max_entries = max(1, int(window.max_change_records))
        if len(window.change_log) > max_entries:
            self._set_window_field(window, "change_log", window.change_log[-max_entries:])

    def _update_explorer_window(
        self,
        *,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> Window:
        directory = normalize_path(extract_directory(arguments, result_json)) or "."
        entries = extract_explorer_entries(result_json)
        identity = ExplorerIdentity(directory_norm=directory)
        window_id: str | None = None
        try:
            window_id = self._locator.find(identity, expected_kind=WindowType.EXPLORER.value)
        except LocatorReject as exc:
            self._last_identity_reject_code = exc.code
        window = self._windows.get(window_id) if window_id else None
        if window is None:
            window = self._find_explorer_window(directory=directory)
        if window is None:
            window = ExplorerWindow.new(
                id=self._new_id("explorer"),
                title="资源管理器",
                directory=directory,
            )
            self._windows[window.id] = window

        self._wake_window(window)
        self._set_window_field(window, "directory", directory)
        self._set_window_field(window, "entries", [str(item) for item in entries])
        self._set_window_field(window, "summary", f"{len(entries)} 个可见项" if entries else "目录视图")
        self._register_window_identity(window)
        self._touch(window)
        self._active_window_id = window.id
        return window

    def _update_sheet_window(
        self,
        *,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> Window:
        file_path = normalize_path(extract_file_path(arguments, result_json))
        sheet_name = extract_sheet_name(arguments, result_json)

        if not file_path and self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and active.type == WindowType.SHEET:
                file_path = active.file_path or ""
                if not sheet_name:
                    sheet_name = active.sheet_name or ""

        normalized_sheet = str(sheet_name or "").strip()
        window_id: str | None = None
        if file_path and normalized_sheet:
            identity = SheetIdentity(
                file_path_norm=file_path,
                sheet_name_norm=normalized_sheet.lower(),
            )
            try:
                window_id = self._locator.find(identity, expected_kind=WindowType.SHEET.value)
            except LocatorReject as exc:
                self._last_identity_reject_code = exc.code
        window = self._windows.get(window_id) if window_id else None
        if window is None:
            window = self._find_sheet_window(file_path=file_path, sheet_name=normalized_sheet)
        if window is None:
            window = SheetWindow.new(
                id=self._new_id("sheet"),
                title=f"{file_path}/{normalized_sheet}" if file_path or normalized_sheet else "表格窗口",
                file_path=file_path or "",
                sheet_name=normalized_sheet or "",
            )
            self._windows[window.id] = window

        self._wake_window(window)

        # 读取工具访问 stale window 时自动清除过时标记
        if window.stale_hint and is_read_like_tool(canonical_tool_name):
            self._set_window_field(window, "stale_hint", None)

        tabs = extract_sheet_tabs(result_json)
        if tabs:
            self._set_window_field(window, "sheet_tabs", tabs)
        sheet_dims = extract_sheet_dimensions(result_json)
        if sheet_dims:
            self._set_window_field(window, "sheet_dimensions", sheet_dims)
        if normalized_sheet:
            self._set_window_field(window, "sheet_name", normalized_sheet)
        if file_path:
            self._set_window_field(window, "file_path", file_path)

        total_rows, total_cols = extract_shape(result_json)
        range_ref = extract_range_ref(
            arguments,
            default_rows=self._budget.default_rows,
            default_cols=self._budget.default_cols,
        )
        geometry = extract_viewport_geometry(
            range_ref,
            default_rows=self._budget.default_rows,
            default_cols=self._budget.default_cols,
        )
        viewport = window.viewport or Viewport()
        viewport.range_ref = range_ref
        viewport.visible_rows = geometry["visible_rows"]
        viewport.visible_cols = geometry["visible_cols"]
        if total_rows > 0:
            viewport.total_rows = total_rows
        if total_cols > 0:
            viewport.total_cols = total_cols
        self._set_window_field(window, "viewport", viewport)
        self._set_window_field(window, "viewport_range", range_ref)
        self._set_window_field(window, "total_rows", viewport.total_rows)
        self._set_window_field(window, "total_cols", viewport.total_cols)
        self._set_window_field(
            window,
            "max_cached_rows",
            max(1, int(self._budget.window_data_buffer_max_rows)),
        )

        scroll_position = compute_scroll_position(
            geometry,
            total_rows=viewport.total_rows,
            total_cols=viewport.total_cols,
        )
        self._set_window_field(window, "scroll_position", scroll_position)

        preview = extract_preview_rows(result_json)
        if preview:
            self._set_window_field(window, "preview_rows", preview)
            if not window.data_buffer:
                normalized_preview = extract_data_rows({"preview": preview}, "read_excel")
                if normalized_preview:
                    self._set_window_field(window, "data_buffer", normalized_preview)
            if not (window.columns or window.schema):
                inferred_columns = extract_columns({"preview": preview}, window.data_buffer)
                if inferred_columns:
                    self._set_window_field(window, "columns", inferred_columns)
                    self._set_window_field(window, "schema", list(inferred_columns))
        self._sync_window_schema_columns(window)

        status_bar = extract_status_bar(
            window.preview_rows,
            columns=window.columns or window.schema or None,
        )
        if status_bar:
            self._set_window_field(window, "status_bar", status_bar)

        freeze = extract_freeze_panes(result_json)
        if freeze:
            self._set_window_field(window, "freeze_panes", freeze)

        column_widths = extract_column_widths(result_json, sheet_name=window.sheet_name or "")
        if column_widths:
            self._set_window_field(window, "column_widths", column_widths)

        row_heights = extract_row_heights(result_json)
        if row_heights:
            self._set_window_field(window, "row_heights", row_heights)

        merged_ranges = extract_merged_ranges(result_json)
        if merged_ranges:
            self._set_window_field(
                window,
                "merged_ranges",
                [str(item).strip().upper() for item in merged_ranges if str(item).strip()],
            )
        add_ranges, remove_ranges = extract_merged_range_delta(result_json)
        if add_ranges or remove_ranges:
            existing = {
                str(item).strip().upper()
                for item in window.merged_ranges
                if str(item).strip()
            }
            existing.update(add_ranges)
            for removed in remove_ranges:
                existing.discard(removed)
            self._set_window_field(window, "merged_ranges", sorted(existing))

        conditional_effects = extract_conditional_effects(result_json)
        if conditional_effects:
            self._set_window_field(window, "conditional_effects", [str(item) for item in conditional_effects])

        style_summary = extract_style_summary(result_json)
        if style_summary:
            self._set_window_field(window, "style_summary", style_summary)

        if canonical_tool_name in {"write_to_sheet", "format_range"}:
            target_range = str(arguments.get("range") or arguments.get("cell_range") or arguments.get("cell") or "").strip()
            if target_range:
                self._set_window_field(window, "summary", f"最近修改区域: {target_range}")
        elif canonical_tool_name in {"list_sheets", "describe_sheets"}:
            self._set_window_field(window, "summary", "工作表元信息已更新")

        self._register_window_identity(window)
        self._touch(window)
        self._active_window_id = window.id
        return window

    def focus_window_action(
        self,
        window_id: str,
        action: str,
        range_ref: str | None = None,
        rows: int | None = None,
    ) -> dict[str, Any]:
        """执行 focus_window 的状态操作。"""
        if not self._enabled:
            return {"status": "error", "message": "窗口感知未启用"}

        window = self._windows.get(window_id)
        if window is None or window.type != WindowType.SHEET:
            available_windows = sorted(
                wid
                for wid, item in self._windows.items()
                if item.type == WindowType.SHEET and not item.dormant
            )
            return {
                "status": "error",
                "message": f"窗口不存在或类型不支持: {window_id}",
                "available_windows": available_windows,
            }

        normalized_action = str(action or "").strip().lower()
        previous_active_id = self._active_window_id if self._active_window_id != window.id else None
        self._operation_seq += 1
        self._set_window_field(window, "current_iteration", self._operation_seq)
        self._wake_window(window)
        self._touch(window)
        self._active_window_id = window.id
        if previous_active_id:
            self._downgrade_previous_focus(previous_active_id)

        if normalized_action == "restore":
            self._set_window_field(window, "detail_level", DetailLevel.FULL)
            self._append_operation(window, "focus_window", {"action": normalized_action}, True)
            return {
                "status": "ok",
                "action": normalized_action,
                "window_id": window.id,
                "active_window_id": self._active_window_id,
            }

        if normalized_action == "clear_filter":
            restored = bool(window.unfiltered_buffer is not None)
            if restored:
                restored_rows = list(window.unfiltered_buffer or [])
                self._set_window_field(window, "data_buffer", restored_rows)
                self._set_window_field(window, "filter_state", None)
                self._set_window_field(window, "unfiltered_buffer", None)
                active_range = window.viewport_range or "A1:A1"
                self._set_window_field(
                    window,
                    "cached_ranges",
                    [
                    CachedRange(
                        range_ref=active_range,
                        rows=restored_rows[: max(1, int(window.max_cached_rows))],
                        is_current_viewport=True,
                        added_at_iteration=window.current_iteration,
                    )
                    ],
                )
            self._set_window_field(window, "stale_hint", None)
            self._set_window_field(window, "detail_level", DetailLevel.FULL)
            self._apply_window_intent(
                window=window,
                tag=IntentTag.VALIDATE,
                confidence=0.95,
                source="tool_rule",
                force=True,
            )
            self._append_operation(window, "focus_window", {"action": normalized_action}, True)
            self._append_change(
                window,
                make_change_record(
                    operation="focus",
                    tool_summary="focus_window(clear_filter)",
                    affected_range=window.viewport_range or "-",
                    change_type="filtered_cleared" if restored else "noop",
                    iteration=window.current_iteration,
                    affected_row_indices=[],
                ),
            )
            return {
                "status": "ok",
                "action": normalized_action,
                "window_id": window.id,
                "active_window_id": self._active_window_id,
                "restored": restored,
                "rows": len(window.data_buffer),
            }

        if normalized_action in {"scroll", "expand"}:
            target_range = self._normalize_focus_range(range_ref or window.viewport_range or "")
            if normalized_action == "expand":
                grow_rows = max(1, int(rows or self._budget.default_rows))
                target_range = self._expand_range_down(target_range, grow_rows)
            if not target_range:
                return {"status": "error", "message": "缺少有效 range 参数"}

            cache_hit = any(
                self._range_contains(cached.range_ref, target_range)
                for cached in window.cached_ranges
            )
            self._set_window_field(window, "viewport_range", target_range)
            self._set_current_viewport_range(window, target_range)
            # focus 补读/滚动继承窗口意图，不主动改写 intent。
            self._append_operation(
                window,
                "focus_window",
                {"action": normalized_action, "range": target_range, "rows": rows},
                True,
            )
            self._append_change(
                window,
                make_change_record(
                    operation="focus",
                    tool_summary=f"focus_window({normalized_action})",
                    affected_range=target_range,
                    change_type="focus_hit" if cache_hit else "focus_refill",
                    iteration=window.current_iteration,
                    affected_row_indices=[],
                ),
            )
            return {
                "status": "ok" if cache_hit else "needs_refill",
                "action": normalized_action,
                "window_id": window.id,
                "active_window_id": self._active_window_id,
                "range": target_range,
                "cache_hit": cache_hit,
                "file_path": window.file_path,
                "sheet_name": window.sheet_name,
            }

        return {"status": "error", "message": f"不支持的 action: {action}"}

    def _downgrade_previous_focus(self, previous_window_id: str) -> None:
        previous = self._windows.get(previous_window_id)
        if previous is None or previous.type != WindowType.SHEET or previous.dormant:
            return
        if self._budget.system_budget_tokens < max(80, self._budget.minimized_tokens * 2):
            self._set_window_field(previous, "detail_level", DetailLevel.ICON)
        else:
            self._set_window_field(previous, "detail_level", DetailLevel.SUMMARY)

    def ingest_focus_read_result(
        self,
        *,
        window_id: str,
        range_ref: str,
        result_text: str,
        tool_name: str,
        arguments: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """将 focus_window 自动补读结果写回窗口。"""
        window = self._windows.get(window_id)
        if window is None or window.type != WindowType.SHEET:
            return {"status": "error", "message": f"窗口不存在或类型不支持: {window_id}"}

        parsed = parse_json_payload(result_text)
        result_json = parsed if isinstance(parsed, dict) else None
        rows = extract_data_rows(result_json, tool_name)
        columns = extract_columns(result_json, rows)

        self._operation_seq += 1
        self._set_window_field(window, "current_iteration", self._operation_seq)
        if columns:
            self._set_window_field(window, "columns", columns)
            self._set_window_field(window, "schema", list(columns))
        self._sync_window_schema_columns(window)

        affected = ingest_read_result(
            window,
            new_range=self._normalize_focus_range(range_ref) or (window.viewport_range or "A1:A1"),
            new_rows=rows,
            iteration=window.current_iteration,
        )
        if window.viewport is not None:
            self._set_window_field(window, "total_rows", window.viewport.total_rows)
            self._set_window_field(window, "total_cols", window.viewport.total_cols)
        if window.total_rows <= 0:
            self._set_window_field(window, "total_rows", len(window.data_buffer))
        if window.total_cols <= 0:
            self._set_window_field(window, "total_cols", len(window.columns or window.schema))
        self._set_window_field(window, "detail_level", DetailLevel.FULL)
        self._touch(window)
        self._active_window_id = window.id

        op_args = dict(arguments or {})
        op_args["range"] = range_ref
        self._append_operation(window, "focus_window_refill", op_args, True)
        self._append_change(
            window,
            make_change_record(
                operation="read",
                tool_summary=f"{tool_name}({range_ref})",
                affected_range=range_ref,
                change_type="added" if rows else "enriched",
                iteration=window.current_iteration,
                affected_row_indices=affected,
            ),
        )
        return {
            "status": "ok",
            "window_id": window.id,
            "range": range_ref,
            "rows": len(rows),
            "cols": len(window.columns or window.schema),
        }

    def _classify_tool(self, tool_name: str):
        if self._rule_engine_version == "v2":
            return classify_tool_meta(tool_name)
        return classify_tool(tool_name)

    def _is_repeat_read_tool(self, tool_name: str) -> bool:
        normalized = str(tool_name or "").strip().lower()
        if self._rule_engine_version == "v2":
            return is_read_like_tool(normalized)
        return normalized in _REPEAT_READ_TOOLS

    def _is_write_like_tool(self, tool_name: str) -> bool:
        normalized = str(tool_name or "").strip().lower()
        if self._rule_engine_version == "v2":
            return is_write_like_tool(normalized)
        return normalized in _WRITE_LIKE_TOOLS

    def _reset_repeat_counter_after_write(self, window: Window) -> None:
        try:
            file_path = str(window.file_path or "").strip()
            sheet_name = str(window.sheet_name or "").strip()
            if file_path and sheet_name:
                self._repeat_detector.record_write(file_path, sheet_name)
        except Exception:
            return

    def _extract_repeat_identity(
        self,
        *,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
        window: Window | None = None,
    ) -> tuple[str, str, str] | None:
        file_path = normalize_path(extract_file_path(arguments, result_json))
        sheet_name = str(extract_sheet_name(arguments, result_json) or "").strip()
        if (not file_path or not sheet_name) and window is not None:
            file_path = file_path or str(window.file_path or "").strip()
            sheet_name = sheet_name or str(window.sheet_name or "").strip()
        range_ref = self._extract_explicit_range_ref(arguments, result_json)
        if not range_ref and window is not None:
            range_ref = str(window.viewport_range or "").strip().upper()
        if not file_path or not sheet_name or not range_ref:
            return None
        return file_path, sheet_name, range_ref

    @staticmethod
    def _extract_explicit_range_ref(
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> str:
        for key in ("range", "cell_range", "source_range", "cell"):
            value = arguments.get(key)
            if isinstance(value, str) and _RANGE_RE.match(value.strip()):
                normalized = value.strip().upper()
                return normalized if ":" in normalized else f"{normalized}:{normalized}"
        if isinstance(result_json, dict):
            for key in ("range", "cell_range", "source_range"):
                value = result_json.get(key)
                if isinstance(value, str) and _RANGE_RE.match(value.strip()):
                    normalized = value.strip().upper()
                    return normalized if ":" in normalized else f"{normalized}:{normalized}"
        return ""

    @staticmethod
    def _normalize_focus_range(value: str) -> str:
        candidate = str(value or "").strip().upper()
        if not candidate:
            return ""
        if _RANGE_RE.match(candidate):
            if ":" in candidate:
                return candidate
            return f"{candidate}:{candidate}"
        return ""

    @staticmethod
    def _expand_range_down(range_ref: str, rows: int) -> str:
        if not range_ref:
            return ""
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        except ValueError:
            return ""
        grow = max(1, int(rows))
        new_max_row = max_row + grow
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"

    @staticmethod
    def _range_contains(cached_range: str, target_range: str) -> bool:
        try:
            c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(cached_range)
            t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(target_range)
        except ValueError:
            return False
        return (
            c_min_col <= t_min_col <= t_max_col <= c_max_col
            and c_min_row <= t_min_row <= t_max_row <= c_max_row
        )

    def _set_current_viewport_range(self, window: Window, target_range: str) -> None:
        selected = False
        for cached in window.cached_ranges:
            matches = (
                cached.range_ref == target_range
                or WindowPerceptionManager._range_contains(cached.range_ref, target_range)
            )
            cached.is_current_viewport = matches
            selected = selected or matches
        if not selected:
            self._append_window_field(
                window,
                "cached_ranges",
                CachedRange(
                    range_ref=target_range,
                    rows=[],
                    is_current_viewport=True,
                    added_at_iteration=window.current_iteration,
                ),
            )

    def _repeat_thresholds_for_intent(self, intent_tag: IntentTag) -> tuple[int, int]:
        if self._rule_engine_version == "v2":
            return v2_repeat_threshold(
                intent_tag,
                base_warn=self._intent_repeat_warn_threshold,
                base_trip=self._intent_repeat_trip_threshold,
            )
        warn = self._intent_repeat_warn_threshold
        trip = self._intent_repeat_trip_threshold
        if intent_tag in {IntentTag.AGGREGATE, IntentTag.VALIDATE, IntentTag.FORMULA}:
            return warn, max(warn + 1, trip)
        relaxed_warn = max(3, warn + 1)
        relaxed_trip = max(relaxed_warn + 1, trip + 1, 4)
        return relaxed_warn, relaxed_trip

    def _resolve_window_intent(
        self,
        *,
        window: Window | None,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
    ) -> dict[str, Any]:
        current_tag = window.intent_tag if window is not None else IntentTag.GENERAL
        if self._rule_engine_version == "v2":
            decision = resolve_intent_decision(
                current_tag=current_tag,
                current_confidence=float(window.intent_confidence if window else 0.0),
                current_lock_until_turn=int(window.intent_lock_until_turn if window else 0),
                current_turn=self._current_turn(),
                intent_enabled=self._intent_enabled,
                sticky_turns=self._intent_sticky_turns,
                user_intent_text=(self._turn_hint_intent_hint or self._turn_hint_user_intent_summary),
                canonical_tool_name=canonical_tool_name,
                arguments=arguments,
                result_json=result_json,
            )
            logger.debug(
                "window.intent_decision: window=%s tool=%s tag=%s source=%s rule_id=%s",
                getattr(window, "id", "-"),
                canonical_tool_name,
                decision.tag.value,
                decision.source,
                decision.rule_id,
            )
            return decision.to_dict()

        if not self._intent_enabled:
            return {
                "tag": current_tag,
                "confidence": float(window.intent_confidence if window else 0.0),
                "source": "carry" if window is not None else "default",
                "force": False,
            }

        user_text = self._turn_hint_intent_hint or self._turn_hint_user_intent_summary
        user_tag, user_conf = self._intent_from_user(user_text)
        tool_tag, tool_conf = self._intent_from_tool(
            canonical_tool_name=canonical_tool_name,
            arguments=arguments,
            result_json=result_json,
            current_tag=current_tag,
        )

        force_switch = user_tag != IntentTag.GENERAL and user_conf >= 0.75
        tag = current_tag
        confidence = float(window.intent_confidence if window else 0.0)
        source = "carry" if window is not None else "default"

        if force_switch:
            tag = user_tag
            confidence = user_conf
            source = "user_rule"
        elif user_tag != IntentTag.GENERAL and user_conf >= 0.5:
            tag = user_tag
            confidence = user_conf
            source = "user_rule"
        elif tool_tag != IntentTag.GENERAL:
            tag = tool_tag
            confidence = tool_conf
            source = "tool_rule"
        elif window is None:
            tag = IntentTag.GENERAL
            confidence = 0.0
            source = "default"
        elif window.intent_tag == IntentTag.GENERAL:
            tag = IntentTag.GENERAL
            confidence = 0.0
            source = "default"

        if (
            window is not None
            and not force_switch
            and window.intent_lock_until_turn >= self._current_turn()
            and tag != window.intent_tag
        ):
            return {
                "tag": window.intent_tag,
                "confidence": max(0.0, float(window.intent_confidence)),
                "source": "carry",
                "force": False,
            }

        return {
            "tag": tag,
            "confidence": max(0.0, min(1.0, float(confidence))),
            "source": source,
            "force": force_switch,
        }

    def _apply_window_intent(
        self,
        *,
        window: Window,
        tag: IntentTag,
        confidence: float,
        source: str,
        force: bool = False,
    ) -> None:
        if not self._intent_enabled:
            return

        current_turn = self._current_turn()
        normalized_source = source if source in {"user_rule", "tool_rule", "carry", "default"} else "default"
        normalized_conf = max(0.0, min(1.0, float(confidence)))
        if (
            not force
            and window.intent_lock_until_turn >= current_turn
            and tag != window.intent_tag
        ):
            return

        if tag == window.intent_tag and not force:
            self._set_window_field(
                window,
                "intent_confidence",
                max(window.intent_confidence, normalized_conf),
            )
            if normalized_source != "default":
                self._set_window_field(window, "intent_source", normalized_source)
            return

        self._set_window_field(window, "intent_tag", tag)
        self._set_window_field(window, "intent_confidence", normalized_conf)
        self._set_window_field(window, "intent_source", normalized_source)
        self._set_window_field(window, "intent_updated_turn", current_turn)
        self._set_window_field(
            window,
            "intent_lock_until_turn",
            current_turn + max(0, self._intent_sticky_turns - 1),
        )

    def _intent_from_user(self, text: str) -> tuple[IntentTag, float]:
        normalized = str(text or "").strip()
        if not normalized:
            return IntentTag.GENERAL, 0.0

        lower = normalized.lower()
        tag_scores: dict[IntentTag, float] = {}
        for tag, keywords in _INTENT_USER_KEYWORDS.items():
            hit_count = sum(1 for keyword in keywords if keyword in normalized or keyword.lower() in lower)
            if hit_count <= 0:
                continue
            tag_scores[tag] = min(0.95, 0.55 + 0.15 * hit_count)

        if not tag_scores:
            explicit = lower.strip()
            try:
                tag = IntentTag(explicit)
            except ValueError:
                return IntentTag.GENERAL, 0.0
            return tag, 0.8

        selected = max(tag_scores.items(), key=lambda item: item[1])
        return selected[0], selected[1]

    def _intent_from_tool(
        self,
        *,
        canonical_tool_name: str,
        arguments: dict[str, Any],
        result_json: dict[str, Any] | None,
        current_tag: IntentTag,
    ) -> tuple[IntentTag, float]:
        tool = str(canonical_tool_name or "").strip().lower()
        if tool in _INTENT_FORMAT_TOOLS:
            return IntentTag.FORMAT, 0.88
        if tool in _INTENT_AGGREGATE_TOOLS:
            return IntentTag.AGGREGATE, 0.84
        if tool in _INTENT_VALIDATE_TOOLS:
            return IntentTag.VALIDATE, 0.9
        if tool in _INTENT_ENTRY_TOOLS:
            if self._has_formula_signal(arguments=arguments, result_json=result_json):
                return IntentTag.FORMULA, 0.9
            return IntentTag.ENTRY, 0.84
        if tool in {"read_excel", "read_sheet", "focus_window_refill"}:
            if current_tag != IntentTag.GENERAL:
                return current_tag, 0.7
            return IntentTag.AGGREGATE, 0.62
        return IntentTag.GENERAL, 0.0

    def _has_formula_signal(self, *, arguments: dict[str, Any], result_json: dict[str, Any] | None) -> bool:
        for candidate in self._iter_text_values(arguments):
            if _FORMULA_HINT_RE.search(candidate):
                return True
        if isinstance(result_json, dict):
            for candidate in self._iter_text_values(result_json):
                if _FORMULA_HINT_RE.search(candidate):
                    return True
        return False

    def _iter_text_values(self, payload: Any) -> list[str]:
        if payload is None:
            return []
        if isinstance(payload, str):
            return [payload]
        if isinstance(payload, dict):
            results: list[str] = []
            for value in payload.values():
                results.extend(self._iter_text_values(value))
            return results
        if isinstance(payload, (list, tuple)):
            results: list[str] = []
            for value in payload:
                results.extend(self._iter_text_values(value))
            return results
        if isinstance(payload, (int, float, bool)):
            return [str(payload)]
        return []

    def _task_type_from_windows(self, windows: list[Window]) -> str:
        hint_tag, hint_conf = self._intent_from_user(self._turn_hint_intent_hint or self._turn_hint_user_intent_summary)
        if hint_tag != IntentTag.GENERAL and hint_conf >= 0.5:
            if self._rule_engine_version == "v2":
                return task_type_from_intent(hint_tag)
            return _INTENT_TO_TASK_TYPE.get(hint_tag, "GENERAL_BROWSE")

        if self._active_window_id:
            active = self._windows.get(self._active_window_id)
            if active is not None and not active.dormant:
                if self._rule_engine_version == "v2":
                    return task_type_from_intent(active.intent_tag)
                return _INTENT_TO_TASK_TYPE.get(active.intent_tag, "GENERAL_BROWSE")

        for window in sorted(windows, key=lambda item: item.last_access_seq, reverse=True):
            if window.intent_tag != IntentTag.GENERAL:
                if self._rule_engine_version == "v2":
                    return task_type_from_intent(window.intent_tag)
                return _INTENT_TO_TASK_TYPE.get(window.intent_tag, "GENERAL_BROWSE")
        return "GENERAL_BROWSE"

    def _build_intent_profile(self, window: Window, *, level: str) -> dict[str, Any]:
        intent = window.intent_tag
        profile: dict[str, Any] = {
            "intent": intent.value,
            "label": intent.value,
            "max_rows": max(1, int(self._budget.window_full_max_rows)),
            "show_style": False,
            "show_quality": False,
            "show_formula": False,
            "show_change": False,
            "focus_text": "",
            "level": level,
        }
        if intent == IntentTag.AGGREGATE:
            profile["focus_text"] = "统计优先"
        elif intent == IntentTag.FORMAT:
            profile.update({
                "max_rows": 3 if level == "full" else 1,
                "show_style": True,
                "show_change": True,
                "focus_text": "样式优先",
            })
        elif intent == IntentTag.VALIDATE:
            profile.update({
                "max_rows": 5 if level == "full" else 2,
                "show_quality": True,
                "focus_text": "质量校验优先",
            })
        elif intent == IntentTag.FORMULA:
            profile.update({
                "max_rows": 5 if level == "full" else 2,
                "show_formula": True,
                "show_change": True,
                "focus_text": "公式排查优先",
            })
        elif intent == IntentTag.ENTRY:
            profile.update({
                "max_rows": 4 if level == "full" else 2,
                "show_change": True,
                "focus_text": "写入变更优先",
            })
        else:
            profile["focus_text"] = "通用浏览"
        return profile

    def _current_turn(self) -> int:
        return max(1, int(self._notice_turn))

    def _sync_window_schema_columns(self, window: Window) -> None:
        if not isinstance(window, SheetWindow):
            return
        if window.schema and not window.columns:
            self._set_window_field(window, "columns", list(window.schema))
        elif window.columns and not window.schema:
            self._set_window_field(window, "schema", list(window.columns))

    def _log_lifecycle_reason_codes(self, plan: LifecyclePlan) -> None:
        if not logger.isEnabledFor(logging.DEBUG):
            return
        if not plan.advices:
            return
        summary = ", ".join(
            f"{advice.window_id}:{advice.tier}:{(advice.reason_code or 'n/a')}"
            for advice in plan.advices
        )
        logger.debug(
            "window.lifecycle_plan: source=%s task=%s reasons=%s",
            plan.source,
            plan.task_type,
            summary,
        )

    def _age_windows(self) -> None:
        """窗口老化：未在本轮前后被新访问的窗口递增 idle。"""
        for window in self._windows.values():
            if window.dormant:
                continue
            if window.last_access_seq > self._last_notice_operation_seq:
                continue
            self._set_window_field(window, "idle_turns", window.idle_turns + 1)

    def _recycle_idle_windows(self) -> None:
        """按 idle 阈值自动回收窗口（标记 dormant）。"""
        terminate_after = max(1, int(self._budget.terminate_after_idle))
        for window in list(self._windows.values()):
            if window.dormant:
                continue
            if window.idle_turns < terminate_after:
                continue
            self._mark_window_dormant(window.id)

    def _refresh_active_window(self) -> None:
        """当前活动窗口只在“本轮有新访问”时保持激活。"""
        if self._active_window_id is None:
            return
        active = self._windows.get(self._active_window_id)
        if active is None or active.dormant:
            self._active_window_id = None
            return
        if active.last_access_seq <= self._last_notice_operation_seq:
            self._active_window_id = None

    def _touch(self, window: Window) -> None:
        self._set_window_field(window, "last_access_seq", self._operation_seq)
        self._set_window_field(window, "idle_turns", 0)

    def _new_id(self, prefix: str) -> str:
        self._seq += 1
        return f"{prefix}_{self._seq}"

    def _drop_window(self, window_id: str) -> None:
        if self._windows.pop(window_id, None) is None:
            return

        if self._active_window_id == window_id:
            self._active_window_id = None

    def _mark_window_dormant(self, window_id: str) -> None:
        window = self._windows.get(window_id)
        if window is None:
            return
        self._set_window_field(window, "dormant", True)
        self._set_window_field(window, "detail_level", DetailLevel.NONE)
        if self._active_window_id == window_id:
            self._active_window_id = None

    def _wake_window(self, window: Window) -> None:
        self._set_window_field(window, "dormant", False)
        if window.detail_level == DetailLevel.NONE:
            self._set_window_field(window, "detail_level", DetailLevel.FULL)

    def _evict_dormant_windows(self) -> None:
        dormant = [item for item in self._windows.values() if item.dormant]
        overflow = len(dormant) - self._MAX_DORMANT_WINDOWS
        if overflow <= 0:
            return
        to_drop = sorted(dormant, key=lambda item: item.last_access_seq)[:overflow]
        for item in to_drop:
            self._drop_window(item.id)

    def _schedule_async_advisor(
        self,
        *,
        active_windows: list[Window],
        context: AdvisorContext,
    ) -> None:
        if self._advisor_mode != "hybrid":
            return
        if self._advisor_runner is None:
            return
        if not active_windows:
            return
        if not self._should_invoke_small_model(context=context, active_windows=active_windows):
            return
        if self._advisor_task is not None and not self._advisor_task.done():
            return

        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            return

        windows_snapshot = [deepcopy(item) for item in active_windows]
        context_snapshot = AdvisorContext(
            turn_number=context.turn_number,
            is_new_task=context.is_new_task,
            window_count_changed=context.window_count_changed,
            user_intent_summary=context.user_intent_summary,
            agent_recent_output=context.agent_recent_output,
            task_type=context.task_type,
            task_tags=context.task_tags,
        )
        task = loop.create_task(
            self._run_async_advisor(
                windows=windows_snapshot,
                active_window_id=self._active_window_id,
                context=context_snapshot,
            )
        )
        self._advisor_task = task
        task.add_done_callback(self._on_advisor_task_done)

    async def _run_async_advisor(
        self,
        *,
        windows: list[Window],
        active_window_id: str | None,
        context: AdvisorContext,
    ) -> None:
        runner = self._advisor_runner
        if runner is None:
            return

        try:
            plan = await runner(windows, active_window_id, self._budget, context)
        except asyncio.CancelledError:
            return
        except Exception:
            return

        if plan is None:
            return
        if plan.generated_turn <= 0:
            plan.generated_turn = context.turn_number
        # 安全性说明：此赋值发生在 asyncio 协程中，与读取方（build_system_notice 等）
        # 同处于单线程事件循环，切换仅在 await 点发生，赋值本身是原子操作，无竞态风险。
        # 若未来引入 ThreadPoolExecutor 执行此协程，需改用 asyncio.Lock 保护。
        self._cached_small_model_plan = plan

    def _on_advisor_task_done(self, task: asyncio.Task[None]) -> None:
        if self._advisor_task is task:
            self._advisor_task = None
        if task.cancelled():
            return
        try:
            _ = task.exception()
        except Exception:
            return

    def _cancel_advisor_task(self) -> None:
        task = self._advisor_task
        self._advisor_task = None
        if task is not None and not task.done():
            task.cancel()

    def _get_fresh_small_model_plan(self, *, current_turn: int) -> LifecyclePlan | None:
        plan = self._cached_small_model_plan
        if plan is None:
            return None
        if int(plan.generated_turn) <= 0:
            return None
        if current_turn - int(plan.generated_turn) > self._advisor_plan_ttl_turns:
            return None
        return plan

    def _should_invoke_small_model(
        self,
        *,
        context: AdvisorContext,
        active_windows: list[Window],
    ) -> bool:
        if context.is_new_task:
            return True
        if context.window_count_changed:
            return True
        if len(active_windows) >= self._advisor_trigger_window_count:
            return True
        return context.turn_number >= self._advisor_trigger_turn

    @staticmethod
    def _normalize_hint(text: str, *, max_chars: int) -> str:
        normalized = " ".join(str(text or "").split())
        if len(normalized) <= max_chars:
            return normalized
        return normalized[:max_chars]

    @staticmethod
    def _normalize_task_tags(task_tags: tuple[str, ...] | list[str]) -> tuple[str, ...]:
        normalized: list[str] = []
        for raw_tag in task_tags:
            tag = str(raw_tag or "").strip().lower()
            if not tag or tag in normalized:
                continue
            normalized.append(tag)
        return tuple(normalized)

    def _truncate_tool_append(self, text: str) -> str:
        max_tokens = max(0, int(self._budget.tool_append_tokens))
        if max_tokens <= 0:
            return ""
        if self._estimate_tokens(text) <= max_tokens:
            return text

        lines = text.splitlines()
        while lines and self._estimate_tokens("\n".join(lines)) > max_tokens:
            lines.pop()
        return "\n".join(lines)

    @staticmethod
    def _estimate_tokens(text: str) -> int:
        return TokenCounter.count_message({"role": "system", "content": text})
