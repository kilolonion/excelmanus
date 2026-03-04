"""API 服务模块：基于 FastAPI 的 REST API 服务。

端点：
- POST   /api/v1/chat                        对话接口（完整 JSON）
- POST   /api/v1/chat/stream                  对话接口（SSE 流式）
- POST   /api/v1/chat/subscribe                 SSE 重连（页面刷新后接入正在执行的任务）
- POST   /api/v1/chat/abort                   终止活跃聊天任务
- GET    /api/v1/skills                      列出 Skillpack 摘要
- GET    /api/v1/skills/{name}               查询 Skillpack 详情
- POST   /api/v1/skills                      创建 project Skillpack
- PATCH  /api/v1/skills/{name}               更新 project Skillpack
- DELETE /api/v1/skills/{name}               软删除 project Skillpack
- POST   /api/v1/skills/import               从本地路径或 GitHub URL 导入 Skillpack
- GET    /api/v1/mcp/servers                 列出 MCP Server 配置+状态
- POST   /api/v1/mcp/servers                 新增 MCP Server
- PUT    /api/v1/mcp/servers/{name}          更新 MCP Server 配置
- DELETE /api/v1/mcp/servers/{name}          删除 MCP Server
- POST   /api/v1/mcp/reload                  热重载所有 MCP 连接
- POST   /api/v1/mcp/servers/{name}/test     测试单个 MCP Server 连接
- GET    /api/v1/files/excel                  返回 xlsx 文件二进制流（Univer 加载）
- GET    /api/v1/files/excel/snapshot         返回 Excel 轻量 JSON 快照（聊天内嵌预览）
- POST   /api/v1/files/excel/write            侧边面板编辑回写单元格
- DELETE /api/v1/sessions/{session_id}        删除会话
- GET    /api/v1/sessions/{sid}/operations     操作历史时间线列表
- GET    /api/v1/sessions/{sid}/operations/{id} 操作详情（含 diff）
- POST   /api/v1/sessions/{sid}/operations/{id}/undo 回滚指定操作
- GET    /api/v1/health                       健康检查
- GET    /api/v1/channels                     渠道协同启动状态
"""

from __future__ import annotations

# ── Web 依赖守卫 ─────────────────────────────────────────
_web_missing: list[str] = []
try:
    import fastapi as _fastapi_check  # noqa: F401
except ImportError:
    _web_missing.append("fastapi")
try:
    import uvicorn as _uvicorn_check  # noqa: F401
except ImportError:
    _web_missing.append("uvicorn")
if _web_missing:
    raise ImportError(
        f"Web API 模式缺少依赖: {', '.join(_web_missing)}。"
        f"\n请运行: pip install excelmanus[web]"
    )
del _web_missing
# ─────────────────────────────────────────────────────────

import asyncio
import json
import os
import re
import uuid
from contextlib import asynccontextmanager
from pathlib import Path
from typing import TYPE_CHECKING, Annotated, Any, AsyncIterator, Literal

import uvicorn
from fastapi import APIRouter, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response, StreamingResponse
from pydantic import AliasChoices, BaseModel, ConfigDict, Field, StringConstraints

import excelmanus
from excelmanus.config import (
    ConfigError,
    ExcelManusConfig,
    ModelProfile,
    load_config,
    load_cors_allow_origins,
)
from excelmanus.engine import ChatResult, ToolCallResult
from excelmanus.events import EventType, ToolCallEvent
from excelmanus.logger import get_logger, setup_logging
from excelmanus.mentions import MentionParser, MentionResolver
from excelmanus.mentions.parser import ResolvedMention
from excelmanus.mcp.manager import MCPManager
from excelmanus.output_guard import (
    guard_public_reply,
    sanitize_external_data,
    sanitize_external_text,
)
from excelmanus.session import (
    SessionBusyError,
    SessionLimitExceededError,
    SessionManager,
    SessionNotFoundError,
)
from excelmanus.skillpacks import (
    SkillpackConflictError,
    SkillpackInputError,
    SkillpackLoader,
    SkillpackNotFoundError,
    SkillRouter,
)
from excelmanus.skillpacks.user_skill_service import UserSkillService
from excelmanus.skillpacks.clawhub import ClawHubError, ClawHubNotFoundError
from excelmanus.skillpacks.importer import SkillImportError
from excelmanus.tools import ToolRegistry
from excelmanus.api_sse import (
    SessionStreamState as _SessionStreamState,
    sse_event_to_sse as _sse_event_to_sse_impl,
    sse_format as _sse_format,
)
from excelmanus.error_guidance import FailureGuidance, classify_failure, classify_workspace_full

if TYPE_CHECKING:
    from excelmanus.channels.rate_limit import RateLimitConfig
    from excelmanus.engine import AgentEngine
    from excelmanus.skillpacks import SkillpackManager
    from excelmanus.workspace import IsolatedWorkspace

logger = get_logger("api")

# ── 请求 / 响应模型 ──────────────────────────────────────


class ImageAttachment(BaseModel):
    """图片附件。"""

    data: str  # base64 编码
    media_type: str = "image/png"
    detail: Literal["auto", "low", "high"] = "auto"


class ChatRequest(BaseModel):
    """对话请求体。"""

    model_config = ConfigDict(extra="forbid")

    message: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1)
    ]
    session_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ] | None = None
    chat_mode: Literal["write", "read", "plan"] = "write"
    channel: str | None = None
    images: list[ImageAttachment] = Field(default_factory=list)


class ChatResponse(BaseModel):
    """对话响应体。"""

    session_id: str
    reply: str
    skills_used: list[str]
    tool_scope: list[str] = Field(default_factory=list)
    route_mode: str
    iterations: int = 0
    truncated: bool = False
    tool_calls: list[dict] = Field(default_factory=list)
    # token 使用统计
    prompt_tokens: int = 0
    completion_tokens: int = 0
    total_tokens: int = 0
    # 自动生成的会话标题（仅首轮返回）
    title: str | None = None


class ErrorResponse(BaseModel):
    """错误响应体（不暴露内部堆栈）。"""

    error: str
    error_id: str


class SkillpackSummaryResponse(BaseModel):
    """Skillpack 摘要响应。"""

    model_config = ConfigDict(populate_by_name=True)

    name: str
    description: str
    source: str
    writable: bool
    argument_hint: str = Field(
        default="",
        validation_alias=AliasChoices("argument_hint", "argument-hint"),
        serialization_alias="argument-hint",
    )


class SkillpackDetailResponse(SkillpackSummaryResponse):
    """Skillpack 详情响应。"""

    file_patterns: list[str] = Field(
        default_factory=list,
        validation_alias=AliasChoices("file_patterns", "file-patterns"),
        serialization_alias="file-patterns",
    )
    resources: list[str]
    version: str
    disable_model_invocation: bool = Field(
        default=False,
        validation_alias=AliasChoices(
            "disable_model_invocation",
            "disable-model-invocation",
        ),
        serialization_alias="disable-model-invocation",
    )
    user_invocable: bool = Field(
        default=True,
        validation_alias=AliasChoices("user_invocable", "user-invocable"),
        serialization_alias="user-invocable",
    )
    instructions: str
    resource_contents: dict[str, str]
    hooks: dict[str, Any] = Field(default_factory=dict)
    model: str | None = None
    metadata: dict[str, Any] = Field(default_factory=dict)
    command_dispatch: str = Field(
        default="none",
        validation_alias=AliasChoices("command_dispatch", "command-dispatch"),
        serialization_alias="command-dispatch",
    )
    command_tool: str | None = Field(
        default=None,
        validation_alias=AliasChoices("command_tool", "command-tool"),
        serialization_alias="command-tool",
    )
    required_mcp_servers: list[str] = Field(
        default_factory=list,
        validation_alias=AliasChoices(
            "required_mcp_servers",
            "required-mcp-servers",
        ),
        serialization_alias="required-mcp-servers",
    )
    required_mcp_tools: list[str] = Field(
        default_factory=list,
        validation_alias=AliasChoices(
            "required_mcp_tools",
            "required-mcp-tools",
        ),
        serialization_alias="required-mcp-tools",
    )
    extensions: dict[str, Any] = Field(default_factory=dict)


class SkillpackImportRequest(BaseModel):
    """导入 skillpack 请求体。"""

    model_config = ConfigDict(extra="forbid")
    source: Literal["local_path", "github_url", "clawhub"]
    value: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=2048)
    ]
    overwrite: bool = False


class SkillpackCreateRequest(BaseModel):
    """创建 skillpack 请求体。"""

    model_config = ConfigDict(extra="forbid")
    name: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=255)
    ]
    payload: dict[str, Any]


class SkillpackPatchRequest(BaseModel):
    """更新 skillpack 请求体。"""

    model_config = ConfigDict(extra="forbid")
    payload: dict[str, Any]


class SkillpackMutationResponse(BaseModel):
    """写操作响应。"""

    status: str
    name: str
    detail: dict[str, Any] | None = None


# ── 全局状态（由 lifespan 初始化） ────────────────────────

_session_manager: SessionManager | None = None
_tool_registry: ToolRegistry | None = None
_skillpack_loader: SkillpackLoader | None = None
_skill_router: SkillRouter | None = None
_skillpack_manager: "SkillpackManager | None" = None
_user_skill_service: UserSkillService | None = None
_config: ExcelManusConfig | None = None
_config_incomplete: bool = False  # True when essential config (API key/base_url/model) is missing
_active_chat_tasks: dict[str, asyncio.Task[Any]] = {}
_draining: bool = False  # True during graceful shutdown, health returns "draining"
_restart_reason: str = ""  # 重启原因，draining 期间通过 health 传递给前端
_channel_launcher: Any = None  # 类型：ChannelLauncher | None（渠道协同启动器）


_session_stream_states: dict[str, _SessionStreamState] = {}
_rules_manager: Any = None  # 类型：RulesManager | None
_api_persistent_memory: Any = None  # 类型：PersistentMemory | None（API 层共享）
_database: Any = None  # 类型：Database | None
_config_store: Any = None  # 类型：ConfigStore | None
_file_registries: dict[str, Any] = {}  # workspace_root → FileRegistry（上传接入）
_router = APIRouter()


def _get_file_registry(workspace_root: str, *, user_id: str | None = None) -> Any:
    """获取或懒创建指定工作区的 FileRegistry 实例。

    ISO-3: 当 user_id 非空时使用 ScopedDatabase 创建 FileRegistry，
    确保 SQLite 模式下读写用户独立的 data.db。
    """
    if _database is None:
        return None
    cache_key = f"{workspace_root}:{user_id or ''}"
    reg = _file_registries.get(cache_key)
    if reg is not None:
        return reg
    try:
        from excelmanus.file_registry import FileRegistry
        db_conn = _database
        if user_id is not None and _config is not None:
            from excelmanus.user_scope import UserScope
            scope = UserScope.create(user_id, _database, _config.workspace_root, data_root=_config.data_root)
            db_conn = scope.scoped_db
        reg = FileRegistry(db_conn, workspace_root)
        _file_registries[cache_key] = reg
        return reg
    except Exception:
        logger.debug("FileRegistry 创建失败 (%s)", workspace_root, exc_info=True)
        return None


def _build_bootstrap_config() -> tuple[ExcelManusConfig, ConfigError | None]:
    """构建应用启动配置：优先完整配置，失败时保留错误并回退到仅供导入期使用的占位配置。"""
    try:
        return load_config(), None
    except ConfigError as exc:
        fallback = ExcelManusConfig(
            api_key="",
            base_url="https://example.invalid/v1",
            model="",
            cors_allow_origins=tuple(load_cors_allow_origins()),
        )
        return fallback, exc


def _is_external_safe_mode() -> bool:
    """是否启用对外安全模式（默认开启）。"""
    if _config is None:
        return True
    return bool(_config.external_safe_mode)


def _get_isolation_user_id(request: Request) -> str | None:
    """Auth 启用时返回当前用户 ID，否则返回 None。

    Auth 即隔离：不再需要额外的 session_isolation_enabled 开关。
    """
    if not getattr(request.app.state, "auth_enabled", False):
        return None
    from excelmanus.auth.dependencies import extract_user_id
    return extract_user_id(request)


def _resolve_workspace(request: Request) -> "IsolatedWorkspace":
    """解析当前请求对应的隔离工作区。"""
    assert _config is not None
    from excelmanus.workspace import IsolatedWorkspace, SandboxConfig
    user_id = _get_isolation_user_id(request)
    auth_enabled = getattr(request.app.state, "auth_enabled", False)
    docker_enabled = getattr(request.app.state, "docker_sandbox_enabled", False)
    return IsolatedWorkspace.resolve(
        _config.workspace_root,
        user_id=user_id,
        auth_enabled=auth_enabled,
        sandbox_config=SandboxConfig(docker_enabled=docker_enabled),
        transaction_enabled=_config.backup_enabled,
        data_root=_config.data_root,
    )


def _resolve_workspace_root(request: Request) -> str:
    """返回当前请求对应的工作区根目录路径字符串。"""
    return str(_resolve_workspace(request).root_dir)


async def _has_session_access(session_id: str, request: Request) -> bool:
    """会话存在且属于当前用户时返回 True（不存在/无权均返回 False）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    try:
        await _session_manager.get_session_detail(session_id, user_id=user_id)
    except SessionNotFoundError:
        return False
    return True


async def _require_admin_if_auth_enabled(request: Request) -> JSONResponse | None:
    """认证启用时要求管理员权限；未启用认证时跳过校验。"""
    if not getattr(request.app.state, "auth_enabled", False):
        return None
    from excelmanus.auth.dependencies import get_current_user_from_request

    user = await get_current_user_from_request(request)
    if user.role != "admin":
        return _error_json_response(403, "需要管理员权限。")
    return None


def _get_user_allowed_models(request: Request) -> list[str]:
    """返回当前用户的 allowed_models 列表。空列表表示不限制。

    认证未启用时始终返回空列表（不限制）。
    """
    if not getattr(request.app.state, "auth_enabled", False):
        return []
    user_store = getattr(request.app.state, "user_store", None)
    if user_store is None:
        return []
    user_id = getattr(getattr(request, "state", None), "user_id", None)
    if not user_id:
        return []
    user = user_store.get_by_id(user_id)
    if user is None:
        return []
    import json as _json
    raw = getattr(user, "allowed_models", None)
    if not raw:
        return []
    if isinstance(raw, list):
        return raw
    try:
        parsed = _json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _list_available_model_names_for_user(request: Request, user_id: str | None) -> list[str]:
    """返回当前用户可切换的模型名称列表（含 default）。"""
    names: list[str] = ["default"]
    if _config_store is not None:
        names.extend((p.get("name") or "") for p in _config_store.list_profiles())
    # 去重 + 去空
    dedup: list[str] = []
    seen: set[str] = set()
    for n in names:
        n = n.strip()
        if not n or n in seen:
            continue
        seen.add(n)
        dedup.append(n)
    return dedup


def _public_route_fields(route_mode: str, skills_used: list[str], tool_scope: list[str]) -> tuple[str, list[str], list[str]]:
    """根据安全模式裁剪路由元信息。"""
    if _is_external_safe_mode():
        return "hidden", [], []
    return route_mode, skills_used, tool_scope


def _build_reply_sse(chat_result: Any, engine: Any) -> str:
    """构建 reply SSE 事件文本（chat_stream 与 chat_subscribe 共用）。"""
    normalized_reply = guard_public_reply((chat_result.reply or "").strip())
    route = engine.last_route_result
    route_mode, skills_used, tool_scope = _public_route_fields(
        route.route_mode,
        route.skills_used,
        route.tool_scope,
    )
    return _sse_format("reply", {
        "content": normalized_reply,
        "skills_used": skills_used,
        "tool_scope": tool_scope,
        "route_mode": route_mode,
        "iterations": chat_result.iterations,
        "truncated": chat_result.truncated,
        "prompt_tokens": chat_result.prompt_tokens,
        "completion_tokens": chat_result.completion_tokens,
        "total_tokens": chat_result.total_tokens,
    })


def _public_excel_path(path: str, *, safe_mode: bool) -> str:
    """将 Excel 路径规范化为前端可直接回传的形式。

    - 优先返回 workspace 相对路径（`./subdir/file.xlsx`），避免泄露绝对路径且保持可用。
    - 已被脱敏为 `<path>/name.xlsx` 的历史值，降级为 `./name.xlsx` 以兼容旧数据。
    - 对工作区外绝对路径，safe_mode 下仍保留脱敏行为。
    - 处理用户隔离架构的路径（`./users/{user_id}/xxx` → `./xxx`）。
    """
    raw = str(path or "").strip()
    if not raw:
        return ""

    if raw.startswith("<path>/"):
        basename = raw.removeprefix("<path>/").strip()
        return f"./{basename}" if basename else ""

    from pathlib import Path

    # 处理 ./users/{user_id}/ 格式的路径（用户隔离架构）
    if raw.startswith("./users/"):
        parts = raw.split("/", 3)  # ['.', 'users', '{user_id}', 'rest']
        if len(parts) >= 4:
            # 提取用户相对路径部分
            user_relative = parts[3]
            return f"./{user_relative}" if user_relative else ""

    if _config is not None:
        workspace = Path(_config.workspace_root).resolve()
        candidate = Path(raw)
        if candidate.is_absolute():
            try:
                rel = candidate.resolve().relative_to(workspace)
                rel_str = rel.as_posix()
                # 如果是 users/{user_id}/xxx 格式，提取用户相对路径
                if rel_str.startswith("users/"):
                    parts = rel_str.split("/", 2)
                    if len(parts) >= 3:
                        return f"./{parts[2]}"
                return f"./{rel_str}"
            except Exception:
                return sanitize_external_text(raw, max_len=500) if safe_mode else raw

    normalized = raw.replace("\\", "/")
    if normalized.startswith("./"):
        return normalized
    if normalized.startswith("/"):
        return sanitize_external_text(normalized, max_len=500) if safe_mode else normalized
    return f"./{normalized}"




def _persist_excel_event(session_id: str, event: ToolCallEvent) -> None:
    """将 EXCEL_DIFF / EXCEL_PREVIEW / FILES_CHANGED 事件持久化到 SQLite。"""
    if _session_manager is None or _session_manager.chat_history is None:
        return
    ch = _session_manager.chat_history
    try:
        if event.event_type == EventType.EXCEL_DIFF:
            pub_path = _public_excel_path(event.excel_file_path, safe_mode=False)
            ch.save_excel_diff(
                session_id=session_id,
                tool_call_id=event.tool_call_id or "",
                file_path=pub_path,
                sheet=event.excel_sheet or "",
                affected_range=event.excel_affected_range or "",
                changes=list(event.excel_changes or [])[:200],
            )
            ch.save_affected_file(session_id, pub_path)
        elif event.event_type == EventType.EXCEL_PREVIEW:
            pub_path = _public_excel_path(event.excel_file_path, safe_mode=False)
            ch.save_excel_preview(
                session_id=session_id,
                tool_call_id=event.tool_call_id or "",
                file_path=pub_path,
                sheet=event.excel_sheet or "",
                columns=list(event.excel_columns or [])[:100],
                rows=list(event.excel_rows or [])[:50],
                total_rows=event.excel_total_rows or 0,
                truncated=bool(event.excel_truncated),
                cell_styles=list(event.excel_cell_styles or [])[:51],
            )
            ch.save_affected_file(session_id, pub_path)
        elif event.event_type == EventType.FILES_CHANGED:
            for f in (event.changed_files or [])[:50]:
                pub = _public_excel_path(f, safe_mode=False)
                if pub:
                    ch.save_affected_file(session_id, pub)
    except Exception:
        logger.debug("持久化 Excel 事件失败", exc_info=True)


def _fire_and_forget(coro: Any, *, name: str = "bridge_notify") -> None:
    """安全地 fire-and-forget 一个协程，捕获异常避免 'Task exception was never retrieved' 警告。"""
    from excelmanus.engine_utils import fire_and_forget
    fire_and_forget(coro, name=name)


def _error_json_response(status_code: int, message: str) -> JSONResponse:
    """构建统一错误响应。"""
    error_id = str(uuid.uuid4())
    body = ErrorResponse(error=message, error_id=error_id)
    return JSONResponse(status_code=status_code, content=body.model_dump())


def _make_content_disposition(filename: str) -> str:
    """构建兼容非 ASCII 文件名的 Content-Disposition 头部值（RFC 5987）。"""
    from urllib.parse import quote

    try:
        filename.encode("ascii")
        return f'attachment; filename="{filename}"'
    except UnicodeEncodeError:
        ascii_fallback = filename.encode("ascii", "replace").decode("ascii")
        encoded = quote(filename)
        return (
            f'attachment; filename="{ascii_fallback}"; '
            f"filename*=UTF-8''{encoded}"
        )


def _require_skillpack_manager():
    """获取模块级 SkillpackManager 实例（无需创建 AgentEngine）。"""
    if _skillpack_manager is None:
        raise HTTPException(status_code=503, detail="SkillpackManager 未初始化")
    return _skillpack_manager


def _require_user_skill_manager(user_id: str | None = None) -> "SkillpackManager":
    """获取 per-user SkillpackManager 实例（技能隔离）。

    优先使用 UserSkillService 返回 per-user 管理器，
    回退到全局 _skillpack_manager（单用户/未初始化场景）。
    """
    if _user_skill_service is not None:
        return _user_skill_service.get_manager(user_id)
    return _require_skillpack_manager()


def _to_skill_summary(detail: dict[str, Any]) -> SkillpackSummaryResponse:
    """从详情字典提取摘要响应。"""
    return SkillpackSummaryResponse(
        name=str(detail.get("name", "")),
        description=str(detail.get("description", "")),
        source=str(detail.get("source", "")),
        writable=bool(detail.get("writable", False)),
        argument_hint=str(
            detail.get("argument-hint", detail.get("argument_hint", "")) or ""
        ),
    )


def _to_skill_detail(detail: dict[str, Any]) -> SkillpackDetailResponse:
    """详情字典转换为响应模型。"""
    return SkillpackDetailResponse(
        name=str(detail.get("name", "")),
        description=str(detail.get("description", "")),
        source=str(detail.get("source", "")),
        writable=bool(detail.get("writable", False)),
        argument_hint=str(
            detail.get("argument-hint", detail.get("argument_hint", "")) or ""
        ),
        file_patterns=list(
            detail.get("file-patterns", detail.get("file_patterns", [])) or []
        ),
        resources=list(detail.get("resources", []) or []),
        version=str(detail.get("version", "1.0.0")),
        disable_model_invocation=bool(
            detail.get(
                "disable-model-invocation",
                detail.get("disable_model_invocation", False),
            )
        ),
        user_invocable=bool(
            detail.get("user-invocable", detail.get("user_invocable", True))
        ),
        instructions=str(detail.get("instructions", "") or ""),
        resource_contents=dict(detail.get("resource_contents", {}) or {}),
        hooks=dict(detail.get("hooks", {}) or {}),
        model=(
            str(detail.get("model")).strip()
            if detail.get("model") is not None and str(detail.get("model")).strip()
            else None
        ),
        metadata=dict(detail.get("metadata", {}) or {}),
        command_dispatch=str(
            detail.get("command-dispatch", detail.get("command_dispatch", "none"))
            or "none"
        ),
        command_tool=(
            str(
                detail.get("command-tool", detail.get("command_tool"))
            ).strip()
            if detail.get("command-tool", detail.get("command_tool")) is not None
            and str(detail.get("command-tool", detail.get("command_tool"))).strip()
            else None
        ),
        required_mcp_servers=list(
            detail.get(
                "required-mcp-servers",
                detail.get("required_mcp_servers", []),
            )
            or []
        ),
        required_mcp_tools=list(
            detail.get(
                "required-mcp-tools",
                detail.get("required_mcp_tools", []),
            )
            or []
        ),
        extensions=dict(detail.get("extensions", {}) or {}),
    )


def _to_standard_skill_detail_dict(detail: dict[str, Any]) -> dict[str, Any]:
    """将技能详情标准化为 API 输出字段。"""
    return _to_skill_detail(detail).model_dump(by_alias=True, exclude_none=False)


def _normalize_chat_result(value: ChatResult | str | None) -> ChatResult:
    """兼容测试桩仍返回 str 的场景。"""
    if isinstance(value, ChatResult):
        return value
    return ChatResult(reply="" if value is None else str(value))


def _serialize_images(images: list[ImageAttachment]) -> list[dict[str, str]]:
    """将请求中的图片附件标准化为引擎可消费的字典列表。"""
    return [img.model_dump() for img in images]


def _public_tool_calls(tool_calls: list[ToolCallResult]) -> list[dict]:
    """根据安全模式裁剪工具调用明细。"""
    if _is_external_safe_mode():
        return []

    rows: list[dict] = []
    for item in tool_calls:
        rows.append({
            "tool_name": item.tool_name,
            "arguments": sanitize_external_data(
                item.arguments if isinstance(item.arguments, dict) else {},
                max_len=1000,
            ),
            "result": sanitize_external_text(item.result or "", max_len=3000),
            "success": bool(item.success),
            "error": (
                sanitize_external_text(item.error, max_len=1000)
                if item.error
                else None
            ),
            "pending_approval": bool(item.pending_approval),
            "approval_id": item.approval_id,
            "pending_question": bool(item.pending_question),
            "question_id": item.question_id,


        })
    return rows


# ── Lifespan ──────────────────────────────────────────────


@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncIterator[None]:
    """应用生命周期：初始化配置、注册 Skill、启动清理任务。"""
    global _session_manager, _tool_registry, _skillpack_loader, _skill_router, _skillpack_manager, _user_skill_service, _config, _database, _config_incomplete, _channel_launcher

    # create_app 已在构建应用时确定启动配置；lifespan 不再二次加载。
    bootstrap_error: ConfigError | None = app.state.bootstrap_config_error
    if bootstrap_error is not None:
        _config_incomplete = True
        logger.warning(
            "\n"
            "╔══════════════════════════════════════════════════════════════╗\n"
            "║  ⚠️  模型配置缺失，服务以降级模式启动                      ║\n"
            "║                                                            ║\n"
            "║  %s\n"
            "║                                                            ║\n"
            "║  你可以通过以下任一方式完成配置：                          ║\n"
            "║    1. 打开浏览器访问前端页面，按引导填写 API Key           ║\n"
            "║    2. 编辑项目根目录下的 .env 文件，设置以下必填项：       ║\n"
            "║       EXCELMANUS_API_KEY=sk-xxx                            ║\n"
            "║       EXCELMANUS_BASE_URL=https://api.openai.com/v1        ║\n"
            "║       EXCELMANUS_MODEL=gpt-4o                              ║\n"
            "║    3. 使用 EXCELMANUS_MODELS 环境变量配置多模型            ║\n"
            "║                                                            ║\n"
            "║  配置完成后，通过前端设置页保存即可生效（无需重启）。      ║\n"
            "╚══════════════════════════════════════════════════════════════╝",
            str(bootstrap_error).ljust(58)[:58] + "║",
        )

    _config = app.state.bootstrap_config
    setup_logging(_config.log_level)

    # ── 集中数据管理：注册安装 + 首次迁移 ──────────────
    from excelmanus.data_home import (
        register_installation,
        migrate_project_env,
        ensure_data_dirs,
        has_project_local_data,
        migrate_data_from_project,
        is_data_centralized,
        scan_once,
    )
    project_root = Path(__file__).resolve().parent.parent
    logger.info("部署模式: %s", _config.deploy_mode)
    try:
        register_installation(project_root)
        # 服务器/Docker 模式跳过桌面目录扫描（无 GUI 环境）
        scan_once(skip_desktop_scan=_config.is_server)
        migrate_project_env(project_root)
        if _config.data_root:
            ensure_data_dirs()
            # 仅 standalone 模式执行自动迁移；服务器模式由管理员手动触发
            if _config.is_standalone:
                if not is_data_centralized() and has_project_local_data(project_root):
                    stats = migrate_data_from_project(project_root)
                    if stats:
                        logger.info("首次运行数据迁移完成: %s", stats)
    except Exception:
        logger.debug("集中数据管理初始化失败（非致命）", exc_info=True)

    # ── 首次启动自动创建桌面快捷方式 ──────────────────────
    # 服务器/Docker 模式跳过自动创建（服务器无桌面环境，用户可通过 API 手动创建浏览器书签）
    if _config.is_standalone:
        try:
            from excelmanus.shortcuts import get_shortcut_info, create_desktop_shortcut
            si = get_shortcut_info()
            if not si.get("exists"):
                sc_path = create_desktop_shortcut(project_root)
                if sc_path:
                    logger.info("已自动创建桌面快捷方式: %s", sc_path)
        except Exception:
            logger.debug("自动创建桌面快捷方式失败（非致命）", exc_info=True)

    # 初始化工具层
    _tool_registry = ToolRegistry()
    _tool_registry.register_builtin_tools(_config.workspace_root)

    # 初始化 Skillpack 层
    _skillpack_loader = SkillpackLoader(_config, _tool_registry)
    _skillpack_loader.load_all()
    _skill_router = SkillRouter(_config, _skillpack_loader)
    from excelmanus.skillpacks import SkillpackManager
    _skillpack_manager = SkillpackManager(_config, _skillpack_loader)

    # 初始化 per-user 技能服务（多用户隔离）
    _user_skill_service = UserSkillService(_config, _tool_registry)

    # 初始化统一数据库
    from excelmanus.database import Database

    _database = None
    chat_history = None
    auth_enabled = os.environ.get("EXCELMANUS_AUTH_ENABLED", "").strip().lower() in ("1", "true", "yes")
    need_database = _config.chat_history_enabled or auth_enabled
    if need_database:
        resolved_db_path = os.path.expanduser(
            _config.chat_history_db_path or _config.db_path
        )
        if _config.database_url:
            _database = Database(database_url=_config.database_url)
            logger.info("统一数据库已启用 (PostgreSQL)")
        else:
            _database = Database(resolved_db_path)
            logger.info("统一数据库已启用: %s", resolved_db_path)
        if _config.chat_history_enabled:
            from excelmanus.chat_history import ChatHistoryStore
            chat_history = ChatHistoryStore(_database)

    # 初始化 ConfigStore 并从 .env 迁移已有 profiles
    global _config_store
    if _database is not None:
        from excelmanus.stores.config_store import ConfigStore
        _config_store = ConfigStore(_database)
        existing = _config_store.list_profiles()
        if not existing:
            env_models_raw = os.environ.get("EXCELMANUS_MODELS", "")
            if env_models_raw:
                n = _config_store.import_profiles_from_env(
                    env_models_raw, _config.api_key, _config.base_url,
                )
                if n:
                    logger.info("已从 EXCELMANUS_MODELS 迁移 %d 个模型 profile 到数据库", n)
                    # 迁移完成后清除 .env 中的 EXCELMANUS_MODELS，避免双数据源
                    try:
                        env_path = _find_env_file()
                        lines = _read_env_file(env_path)
                        cleaned = [
                            ln for ln in lines
                            if not ln.strip().startswith("EXCELMANUS_MODELS=")
                        ]
                        if len(cleaned) != len(lines):
                            _write_env_file(env_path, cleaned)
                            logger.info("已从 .env 清除 EXCELMANUS_MODELS（数据库为唯一来源）")
                    except Exception:
                        logger.debug("清除 .env 中 EXCELMANUS_MODELS 失败", exc_info=True)
                    os.environ.pop("EXCELMANUS_MODELS", None)
        _sync_config_profiles_from_db()
        app.state.config_store = _config_store
        logger.info("ConfigStore 已初始化")

    # 初始化会话管理器
    # 始终创建共享 MCP 管理器并在启动时自动连接
    shared_mcp_manager = MCPManager(_config.workspace_root)
    try:
        await shared_mcp_manager.initialize(_tool_registry)
        mcp_info = shared_mcp_manager.get_server_info()
        mcp_ready = sum(1 for s in mcp_info if s["status"] == "ready")
        if mcp_info:
            logger.info(
                "MCP 启动自动连接完成: %d/%d 个 Server 就绪",
                mcp_ready, len(mcp_info),
            )
    except Exception:
        logger.warning("MCP 启动自动连接失败", exc_info=True)

    _session_manager = SessionManager(
        max_sessions=_config.max_sessions,
        ttl_seconds=_config.session_ttl_seconds,
        config=_config,
        registry=_tool_registry,
        skill_router=_skill_router,
        shared_mcp_manager=shared_mcp_manager,
        chat_history=chat_history,
        database=_database,
        config_store=_config_store,
        user_store=None,  # 延迟设置，等 UserStore 初始化完成后注入
        user_skill_service=_user_skill_service,
    )
    await _session_manager.start_background_cleanup()

    # 初始化全局 RulesManager + 共享 PersistentMemory（供 API 层直接使用）
    global _rules_manager, _api_persistent_memory
    try:
        from excelmanus.rules import RulesManager as _RM
        from excelmanus.stores.rules_store import RulesStore as _RS
        _rules_db_store = _RS(_database) if _database is not None else None
        _rules_manager = _RM(db_store=_rules_db_store)
    except Exception:
        logger.debug("RulesManager 初始化失败", exc_info=True)

    if _config.memory_enabled:
        try:
            from excelmanus.persistent_memory import PersistentMemory as _PM
            if _database is not None:
                from excelmanus.stores.memory_store import MemoryStore as _MS
                _mem_backend = _MS(_database)
            else:
                from excelmanus.stores.file_memory_backend import FileMemoryBackend as _FMB
                _mem_backend = _FMB(
                    memory_dir=_config.memory_dir,
                    auto_load_lines=_config.memory_auto_load_lines,
                )
            _api_persistent_memory = _PM(
                backend=_mem_backend,
                auto_load_lines=_config.memory_auto_load_lines,
            )
        except Exception:
            logger.debug("API PersistentMemory 初始化失败", exc_info=True)

    loaded_skillpacks = (
        sorted(_skillpack_loader.get_skillpacks().keys())
        if _skillpack_loader is not None
        else []
    )
    tool_names = (
        sorted(_tool_registry.get_tool_names())
        if _tool_registry is not None
        else []
    )
    # 初始化认证模块（UserStore）+ 速率限制器
    _user_store = None
    if _database is not None:
        try:
            from excelmanus.auth.store import UserStore
            _user_store = UserStore(_database)
            app.state.user_store = _user_store
            app.state.auth_enabled = auth_enabled
            app.state.workspace_root = _config.workspace_root
            app.state.data_root = _config.data_root
            # Auth 即隔离：auth_enabled 时自动启用会话隔离，无需额外开关。
            app.state.session_isolation_enabled = auth_enabled
            # 将 UserStore 注入 SessionManager，使其能读取用户自定义 LLM 配置
            if _session_manager is not None:
                _session_manager.set_user_store(_user_store)
            # 初始化 CredentialStore（订阅凭证管理）
            try:
                from excelmanus.auth.providers.credential_store import CredentialStore as _CredStore
                _cred_store = _CredStore(_database.conn)
                app.state.credential_store = _cred_store
                if _session_manager is not None:
                    _session_manager.set_credential_store(_cred_store)
                # 初始化 CredentialResolver（运行时凭证解析 + 自动刷新）
                try:
                    from excelmanus.auth.providers.resolver import CredentialResolver as _CredResolver
                    _cred_resolver = _CredResolver(
                        credential_store=_cred_store,
                        user_store=_user_store,
                    )
                    app.state.credential_resolver = _cred_resolver
                    if _session_manager is not None:
                        _session_manager.set_credential_resolver(_cred_resolver)
                    logger.info("CredentialResolver 已初始化")
                except Exception:
                    logger.debug("CredentialResolver 初始化失败", exc_info=True)
                    app.state.credential_resolver = None
            except Exception:
                logger.debug("CredentialStore 初始化失败", exc_info=True)
                app.state.credential_store = None
                app.state.credential_resolver = None
            if auth_enabled:
                logger.info("认证系统已启用")
            else:
                logger.info("认证系统已初始化（未强制启用，设置 EXCELMANUS_AUTH_ENABLED=true 启用）")
        except Exception:
            logger.warning("认证模块初始化失败", exc_info=True)

    if auth_enabled:
        try:
            from excelmanus.auth.rate_limit import RateLimiter
            app.state.rate_limiter = RateLimiter()
            logger.info("速率限制器已启用")
        except Exception:
            logger.warning("速率限制器初始化失败", exc_info=True)

    # Docker 沙盒开关：默认关闭，管理员可通过 config_kv 或环境变量开启
    _docker_env = os.environ.get(
        "EXCELMANUS_DOCKER_SANDBOX", ""
    ).strip().lower() in ("1", "true", "yes")
    if _config_store is not None:
        _docker_db = _config_store.get("docker_sandbox_enabled")
        if _docker_db:
            _docker_env = _docker_db.lower() in ("1", "true", "yes")
    if _docker_env:
        from excelmanus.security.docker_sandbox import is_docker_available, is_sandbox_image_ready
        if not is_docker_available():
            logger.warning("Docker 沙盒已开启但 Docker daemon 不可用，已自动关闭")
            _docker_env = False
        elif not is_sandbox_image_ready():
            logger.warning(
                "Docker 沙盒已开启但镜像 excelmanus-sandbox:latest 未找到，"
                "请运行 docker build -t excelmanus-sandbox:latest -f Dockerfile.sandbox ."
            )
        else:
            logger.info("Docker 沙盒已启用")
    app.state.docker_sandbox_enabled = _docker_env
    from excelmanus.tools.code_tools import init_docker_sandbox
    init_docker_sandbox(_docker_env)
    # 将 Docker 沙盒状态传播到会话管理器，用于按工作区注入。
    if _session_manager is not None:
        await _session_manager.set_sandbox_docker_enabled(_docker_env)

    logger.info(
        "API 服务启动完成，已加载 %d 个工具、%d 个 Skillpack",
        len(tool_names),
        len(loaded_skillpacks),
    )

    # ── 后台静默检查更新（非阻塞） ──────────────────────
    async def _background_update_check() -> None:
        try:
            import asyncio
            from functools import partial
            from excelmanus.updater import check_for_updates, get_current_version
            loop = asyncio.get_running_loop()
            info = await loop.run_in_executor(None, partial(check_for_updates, project_root, force=False))
            if info.has_update:
                logger.info(
                    "发现新版本: %s → %s（%d 个新提交）。"
                    "可在设置页「版本管理」中执行更新。",
                    get_current_version(project_root), info.latest, info.commits_behind,
                )
        except Exception:
            logger.debug("启动时后台版本检查失败（非致命）", exc_info=True)

    _fire_and_forget(_background_update_check(), name="update_check")

    # ── 渠道绑定管理器 + 服务令牌 ──────────────────────────
    _bind_manager = None
    _service_token: str | None = None
    if auth_enabled and _user_store is not None:
        try:
            from excelmanus.auth.channel_bind import ChannelBindManager
            _bind_manager = ChannelBindManager(user_store=_user_store)
            app.state.bind_manager = _bind_manager
            logger.info("渠道绑定管理器已初始化")
        except Exception:
            logger.debug("渠道绑定管理器初始化失败", exc_info=True)
        try:
            from excelmanus.auth.security import get_or_create_service_token
            _service_token = get_or_create_service_token()
            logger.info("服务令牌已就绪")
        except Exception:
            logger.debug("服务令牌获取失败", exc_info=True)
    app.state.service_token = _service_token

    # ── EventBridge（跨渠道实时事件推送） ──────────────────
    _event_bridge = None
    try:
        from excelmanus.channels.event_bridge import EventBridge
        _event_bridge = EventBridge()
        app.state.event_bridge = _event_bridge
        logger.info("EventBridge 已初始化")
    except Exception:
        logger.debug("EventBridge 初始化失败", exc_info=True)

    # ── 渠道协同启动（可选） ──────────────────────────────
    _channels_config: list[str] = getattr(app.state, "channels", None) or []
    if not _channels_config:
        from excelmanus.channels.launcher import parse_channels_config
        _channels_config = parse_channels_config()

    # 始终创建 launcher（即使环境变量未配置渠道），以支持前端热启动
    from excelmanus.channels.launcher import ChannelLauncher
    api_port = int(os.environ.get("EXCELMANUS_API_PORT", "8000"))
    _channel_launcher = ChannelLauncher(
        _channels_config,
        api_port=api_port,
        bind_manager=_bind_manager,
        service_token=_service_token,
        event_bridge=_event_bridge,
        config_store=_config_store,
    )
    app.state.channel_launcher = _channel_launcher
    # 合并环境变量/CLI + 持久化配置，统一启动渠道（避免无凭证先失败再重试）
    _channels_to_start: dict[str, dict[str, str] | None] = {}
    for _ch_name in _channels_config:
        _channels_to_start[_ch_name] = None

    if _config_store is not None:
        try:
            from excelmanus.channels.config_store import ChannelConfigStore
            _ccs = ChannelConfigStore(_config_store)
            for _ch_name, _ch_cfg in _ccs.load_all().items():
                if _ch_cfg.enabled and _ch_cfg.has_required_credentials():
                    _channels_to_start[_ch_name] = _ch_cfg.credentials
        except Exception:
            logger.debug("加载持久化渠道配置失败", exc_info=True)

    for _ch_name, _ch_creds in _channels_to_start.items():
        ok, msg = await _channel_launcher.start_channel(
            _ch_name, credentials=_ch_creds,
        )
        if ok:
            if _ch_creds is not None:
                logger.info("从持久化配置自动启动渠道: %s", _ch_name)
            else:
                logger.info("渠道 %s 已启动", _ch_name)
        else:
            logger.warning("渠道 %s 启动失败: %s", _ch_name, msg)

    yield

    # ── Graceful Shutdown: 标记 draining 并等待活跃连接排空 ──
    global _draining
    _draining = True
    logger.info("API 服务进入 draining 状态，等待活跃连接排空...")

    if _session_manager is not None:
        drain_timeout = 30  # 最长等待 30 秒
        for _drain_i in range(drain_timeout):
            active = await _session_manager.get_active_count()
            if active == 0:
                logger.info("所有活跃连接已排空")
                break
            if _drain_i % 5 == 0:
                logger.info("等待 %d 个活跃连接排空... (%d/%ds)", active, _drain_i, drain_timeout)
            await asyncio.sleep(1)
        else:
            active = await _session_manager.get_active_count()
            if active > 0:
                logger.warning("排空超时，仍有 %d 个活跃连接，强制关闭", active)

    # 停止渠道协同 Bot
    if _channel_launcher is not None:
        await _channel_launcher.stop()

    # 关闭所有会话与 MCP 连接
    if _session_manager is not None:
        await _session_manager.shutdown()

    # 关闭统一数据库
    if _database is not None:
        _database.close()

    logger.info("API 服务已关闭")


# ── 全局异常处理 ──────────────────────────────────────────


# 错误消息映射表：将内部错误类型/消息映射为友好用户消息
_ERROR_MESSAGE_MAP: dict[str, str] = {
    # 会话相关错误
    "SessionNotFoundError": "会话已过期或不存在，请刷新页面重新开始。",
    "SessionLimitExceededError": "系统繁忙，会话数量已达上限，请稍后再试。",
    "SessionBusyError": "会话正在处理中，请稍等片刻再提交新请求。",
    # 配置相关错误
    "ConfigError": "配置错误，请检查设置后重试。",
    # 工具执行错误
    "ToolNotFoundError": "请求的工具不可用，请刷新页面后重试。",
    "ToolExecutionError": "工具执行失败，请稍后重试。",
    "ToolNotAllowedError": "当前操作不被允许，请尝试其他方式。",
    # Skillpack 错误
    "SkillpackNotFoundError": "技能包不存在，请刷新页面后重试。",
    "SkillpackManagerError": "技能包加载失败，请稍后重试。",
    # 安全相关错误
    "SecurityViolationError": "安全检查未通过，请检查输入后重试。",
    # 数据库错误
    "database": "数据库操作失败，请稍后重试。",
    "sqlite": "数据存储失败，请稍后重试。",
    # 网络/API 错误
    "timeout": "请求超时，请稍后重试。",
    "TimeoutError": "请求超时，请稍后重试。",
    "ConnectionError": "网络连接失败，请检查网络后重试。",
    "ConnectionRefusedError": "服务暂不可用，请稍后重试。",
    # 认证错误
    "AuthenticationError": "认证失败，请重新登录。",
    "UnauthorizedError": "登录已过期，请重新登录。",
    "PermissionError": "权限不足，无法执行此操作。",
    # 文件相关错误
    "FileNotFoundError": "文件不存在，请检查路径后重试。",
    "PermissionError": "没有文件访问权限，请检查权限设置。",
    # 内存/资源错误
    "MemoryError": "内存不足，请减少操作范围后重试。",
    "out of memory": "内存不足，请减少操作范围后重试。",
    # 默认内部错误（当无法映射时）
    "internal_error": "服务内部错误，请联系管理员。",
}


def _get_friendly_error_message(
    exc: Exception, friendly_enabled: bool
) -> str:
    """获取友好的错误消息。

    如果 friendly_error_messages 功能开启，则尝试将内部错误映射为友好消息；
    否则返回原始错误消息或通用内部错误消息。
    """
    if not friendly_enabled:
        # 功能关闭时，返回原始错误消息
        return str(exc)

    # 尝试匹配错误类型名
    exc_type_name = type(exc).__name__
    if exc_type_name in _ERROR_MESSAGE_MAP:
        return _ERROR_MESSAGE_MAP[exc_type_name]

    # 尝试匹配错误消息中的关键词
    exc_message = str(exc).lower()
    for key, friendly_msg in _ERROR_MESSAGE_MAP.items():
        if key in exc_message:
            return friendly_msg

    # 无法映射时返回通用友好消息
    return "服务处理出现异常，请稍后重试。如问题持续，请联系管理员。"


def _extract_provider_from_base_url(base_url: str | None) -> str:
    """从 base_url 提取 provider 名（如 openai/deepseek/anthropic）。"""
    if not base_url:
        return ""
    try:
        from urllib.parse import urlparse
        hostname = urlparse(base_url).hostname or ""
        parts = hostname.split(".")
        if len(parts) >= 2:
            return parts[-2]
        return parts[0] if parts else ""
    except Exception:
        return ""


def _failure_guidance_sse(guidance: FailureGuidance) -> str:
    """将 FailureGuidance 格式化为单个 failure_guidance SSE 文本。"""
    return _sse_format("failure_guidance", guidance.to_dict())


def _failure_guidance_text(guidance: FailureGuidance) -> str:
    """将结构化失败引导转换为可持久化的 assistant 文本。"""
    lines: list[str] = []
    if guidance.title:
        lines.append(f"⚠️ {guidance.title}")
    if guidance.message:
        lines.append(guidance.message)
    if guidance.diagnostic_id:
        lines.append(f"诊断 ID: {guidance.diagnostic_id}")
    return "\n".join(lines).strip() or "服务处理出现异常，请稍后重试。"


def _persist_failure_guidance_message(
    *,
    session_id: str | None,
    engine: Any | None,
    guidance: FailureGuidance,
) -> None:
    """将失败引导以 assistant 消息持久化，防止前端刷新后只剩用户消息。"""
    if not session_id or engine is None or _session_manager is None:
        return
    try:
        raw_messages = getattr(engine, "raw_messages", None)
        if (
            isinstance(raw_messages, list)
            and raw_messages
            and isinstance(raw_messages[-1], dict)
            and raw_messages[-1].get("role") == "assistant"
        ):
            return
        engine.memory.add_assistant_message(_failure_guidance_text(guidance))
        _session_manager.flush_messages_sync(session_id)
    except Exception:
        logger.debug("会话 %s 失败引导消息持久化失败", session_id, exc_info=True)


async def _handle_session_not_found(
    request: Request, exc: SessionNotFoundError
) -> JSONResponse:
    """会话不存在 → 404。"""
    error_id = str(uuid.uuid4())
    friendly_enabled = _config.friendly_error_messages if _config else False
    error_msg = _get_friendly_error_message(exc, friendly_enabled)
    body = ErrorResponse(error=error_msg, error_id=error_id)
    return JSONResponse(status_code=404, content=body.model_dump())


async def _handle_session_limit(
    request: Request, exc: SessionLimitExceededError
) -> JSONResponse:
    """会话数量超限 → 429。"""
    error_id = str(uuid.uuid4())
    friendly_enabled = _config.friendly_error_messages if _config else False
    error_msg = _get_friendly_error_message(exc, friendly_enabled)
    body = ErrorResponse(error=error_msg, error_id=error_id)
    return JSONResponse(status_code=429, content=body.model_dump())


async def _handle_session_busy(
    request: Request, exc: SessionBusyError
) -> JSONResponse:
    """会话正在处理中 → 409。"""
    error_id = str(uuid.uuid4())
    friendly_enabled = _config.friendly_error_messages if _config else False
    error_msg = _get_friendly_error_message(exc, friendly_enabled)
    body = ErrorResponse(error=error_msg, error_id=error_id)
    return JSONResponse(status_code=409, content=body.model_dump())


async def _handle_unexpected(
    request: Request, exc: Exception
) -> JSONResponse:
    """未预期异常 → 500，返回 error_id，不暴露堆栈。"""
    error_id = str(uuid.uuid4())
    logger.error(
        "未预期异常 [error_id=%s]: %s", error_id, exc, exc_info=True
    )
    friendly_enabled = _config.friendly_error_messages if _config else False
    error_msg = _get_friendly_error_message(exc, friendly_enabled)
    body = ErrorResponse(error=error_msg, error_id=error_id)
    return JSONResponse(status_code=500, content=body.model_dump())


def _register_exception_handlers(application: FastAPI) -> None:
    """注册全局异常处理器。"""
    application.add_exception_handler(SessionNotFoundError, _handle_session_not_found)
    application.add_exception_handler(SessionLimitExceededError, _handle_session_limit)
    application.add_exception_handler(SessionBusyError, _handle_session_busy)
    application.add_exception_handler(Exception, _handle_unexpected)


def create_app(
    config: ExcelManusConfig | None = None,
    *,
    channels: list[str] | None = None,
) -> FastAPI:
    """创建 FastAPI 应用，CORS 与运行期配置共享同一来源。

    Args:
        config: 预构建的配置对象；为 None 时自动从环境加载。
        channels: 要协同启动的渠道列表（如 ["qq"]）；
                  为 None 时从 EXCELMANUS_CHANNELS 环境变量读取。
    """
    bootstrap_error: ConfigError | None = None
    bootstrap_config = config
    if bootstrap_config is None:
        bootstrap_config, bootstrap_error = _build_bootstrap_config()
    assert bootstrap_config is not None

    application = FastAPI(
        title="ExcelManus API",
        version=excelmanus.__version__,
        lifespan=lifespan,
    )
    application.state.bootstrap_config = bootstrap_config
    application.state.bootstrap_config_error = bootstrap_error
    application.state.channels = channels

    # 构建 CORS 允许来源列表：除了显式配置的来源外，自动添加本机 LAN IP
    # 的前端端口来源，以便浏览器直连后端的 SSE 流式请求不被 CORS 拦截。
    cors_origins = set(bootstrap_config.cors_allow_origins)
    lan_ips: set[str] = set()
    # 方法1: gethostname + getaddrinfo（部分系统可用）
    try:
        import socket
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None, socket.AF_INET):
            ip = info[4][0]
            if ip and not ip.startswith("127."):
                lan_ips.add(ip)
    except Exception:
        pass
    # 方法2: UDP connect trick（不实际发送数据，最可靠的主 IP 获取方式）
    try:
        import socket as _sock
        s = _sock.socket(_sock.AF_INET, _sock.SOCK_DGRAM)
        s.settimeout(0)
        s.connect(("10.254.254.254", 1))
        ip = s.getsockname()[0]
        s.close()
        if ip and not ip.startswith("127."):
            lan_ips.add(ip)
    except Exception:
        pass
    # 从环境变量读取前端端口，支持多端口（逗号分隔）和自定义端口部署（默认 3000）
    _frontend_ports_raw = os.environ.get("EXCELMANUS_FRONTEND_PORT", "3000").strip()
    _frontend_ports = [p.strip() for p in _frontend_ports_raw.split(",") if p.strip()]
    for ip in lan_ips:
        for _fp in _frontend_ports:
            cors_origins.add(f"http://{ip}:{_fp}")

    application.add_middleware(
        CORSMiddleware,
        allow_origins=list(cors_origins),
        allow_credentials=True,
        allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
        allow_headers=[
            "Content-Type",
            "Authorization",
            "Accept",
            "X-Requested-With",
            "Cache-Control",
        ],
        expose_headers=["X-Request-Id"],
    )

    # 认证中间件（在 CORS 之后运行，确保预检 OPTIONS 请求正常通过）
    from excelmanus.auth.middleware import AuthMiddleware
    application.add_middleware(AuthMiddleware)

    _register_exception_handlers(application)

    # 注册认证路由
    from excelmanus.auth.router import router as auth_router
    application.include_router(auth_router)

    # 注册子路由模块
    from excelmanus.api_routes_mcp import router as mcp_router
    application.include_router(mcp_router)
    from excelmanus.api_routes_rules import router as rules_router
    application.include_router(rules_router)
    from excelmanus.api_routes_version import router as version_router
    application.include_router(version_router)

    application.include_router(_router)
    return application


# ── 端点 ──────────────────────────────────────────────────

# OpenAPI 错误响应 schema（供路由 responses 参数引用）
_error_responses: dict = {
    403: {"model": ErrorResponse, "description": "无权限执行"},
    404: {"model": ErrorResponse, "description": "会话不存在"},
    409: {"model": ErrorResponse, "description": "会话正在处理中"},
    422: {"model": ErrorResponse, "description": "请求参数错误"},
    429: {"model": ErrorResponse, "description": "会话数量超限"},
    500: {"model": ErrorResponse, "description": "服务内部错误"},
}


def _is_save_command(message: str) -> bool:
    """判断消息是否为 /save 命令（含 /save 或 /save <路径>）。

    支持带文件上传前缀的消息格式（如 "[已上传文件: x]\\n\\n/save"）。
    """
    stripped = message.strip()
    # 若有 \\n\\n，取最后一段作为用户输入（网页端上传文件时会在前加前缀）
    if "\n\n" in stripped:
        cmd_part = stripped.split("\n\n")[-1].strip()
    else:
        cmd_part = stripped
    if not cmd_part.lower().startswith("/save"):
        return False
    parts = cmd_part.split(None, 1)
    return parts[0].lower() == "/save"


async def _handle_save_command(
    session_manager: SessionManager,
    session_id: str,
    engine: Any,
    message: str,
) -> str:
    """执行 /save：将对话导出为 JSON 文件，返回保存路径作为回复。"""
    path: str | None = None
    stripped = message.strip()
    cmd_part = stripped.split("\n\n")[-1].strip() if "\n\n" in stripped else stripped
    parts = cmd_part.split(None, 1)
    if len(parts) > 1:
        path = parts[1].strip() or None

    engine.memory.add_user_message(message)
    if hasattr(engine, "state") and engine.state is not None:
        engine._state.increment_turn()

    saved_path = engine.save_conversation(path)
    if saved_path:
        reply = f"对话已保存至：`{saved_path}`"
    else:
        reply = "当前对话为空，未生成文件。"

    engine.memory.add_assistant_message(reply)
    await session_manager.release_for_chat(session_id)
    return reply


async def _resolve_mentions(
    message: str,
    engine: Any,
) -> tuple[str, list[ResolvedMention] | None]:
    """解析用户消息中的 @ 提及标记，返回 (display_text, mention_contexts)。

    display_text 将 ``@file:name`` 替换为 ``name``，确保 LLM 看到的文本不含 @ 前缀。
    """
    try:
        parse_result = MentionParser.parse(message)
        if not parse_result.mentions:
            return message, None

        from excelmanus.security.guard import FileAccessGuard

        guard = FileAccessGuard(engine._config.workspace_root)
        skill_loader = getattr(engine, "_skill_loader", None)
        if skill_loader is None:
            _router = getattr(engine, "_skill_router", None)
            if _router is not None:
                skill_loader = getattr(_router, "_loader", None)
        mcp_manager = getattr(engine, "_mcp_manager", None)
        resolver = MentionResolver(
            workspace_root=engine._config.workspace_root,
            guard=guard,
            skill_loader=skill_loader,
            mcp_manager=mcp_manager,
        )
        mention_contexts = await resolver.resolve(list(parse_result.mentions))
        return parse_result.display_text, mention_contexts
    except Exception:
        logger.debug("API 层 @ 提及解析失败，回退到原始消息", exc_info=True)
        return message, None


@_router.post("/api/v1/chat", response_model=ChatResponse, responses=_error_responses)
async def chat(request: ChatRequest, raw_request: Request) -> ChatResponse:
    """对话接口：创建或复用会话，将消息传递给 AgentEngine。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if _config_incomplete:
        return _error_json_response(
            503,
            "模型尚未配置，请先在设置页面或 .env 文件中配置 API Key、Base URL 和 Model。",
        )

    from excelmanus.auth.dependencies import extract_user_id
    auth_user_id = extract_user_id(raw_request)
    isolation_user_id = _get_isolation_user_id(raw_request)

    rate_limiter = getattr(raw_request.app.state, "rate_limiter", None)
    if rate_limiter is not None and auth_user_id:
        rate_limiter.check_chat(auth_user_id)

    session_id, engine = await _session_manager.acquire_for_chat(
        request.session_id, user_id=isolation_user_id,
    )

    if _is_save_command(request.message):
        try:
            reply_text = await _handle_save_command(
                _session_manager, session_id, engine, request.message
            )
        finally:
            await _session_manager.release_for_chat(session_id)
        return ChatResponse(
            session_id=session_id,
            reply=guard_public_reply(reply_text),
            skills_used=[],
            route_mode="control_command",
        )

    try:
        display_text, mention_contexts = await _resolve_mentions(
            request.message, engine,
        )

        def _on_event_sync(event: ToolCallEvent) -> None:
            _persist_excel_event(session_id, event)

        chat_result = _normalize_chat_result(
            await engine.chat(
                display_text,
                on_event=_on_event_sync,
                mention_contexts=mention_contexts,
                images=_serialize_images(request.images),
                chat_mode=request.chat_mode,
                channel=request.channel,
            )
        )
    finally:
        await _session_manager.release_for_chat(session_id)

    normalized_reply = guard_public_reply(chat_result.reply.strip())
    # ── 自动标题（仅首轮）：截取用户消息 + 后台 AI 生成 ──
    generated_title: str | None = None
    if engine.session_turn == 1:
        generated_title = _truncate_user_message_as_title(display_text)
        if generated_title:
            _ch = _session_manager.chat_history if _session_manager else None
            if _ch is not None:
                try:
                    _ch.update_session(session_id, title=generated_title, title_source="truncated")
                except Exception:
                    logger.debug("截取标题写入失败", exc_info=True)
        _fire_and_forget(_generate_session_title_background(
            session_id=session_id,
            user_message=display_text,
            assistant_reply=normalized_reply,
        ), name="session_title")
    route = engine.last_route_result
    route_mode, skills_used, tool_scope = _public_route_fields(
        route.route_mode,
        route.skills_used,
        route.tool_scope,
    )
    return ChatResponse(
        session_id=session_id,
        reply=normalized_reply,
        skills_used=skills_used,
        tool_scope=tool_scope,
        route_mode=route_mode,
        iterations=chat_result.iterations,
        truncated=chat_result.truncated,
        tool_calls=_public_tool_calls(chat_result.tool_calls),
        prompt_tokens=chat_result.prompt_tokens,
        completion_tokens=chat_result.completion_tokens,
        total_tokens=chat_result.total_tokens,
        title=generated_title,
    )


@_router.post("/api/v1/chat/stream", responses=_error_responses)
async def chat_stream(request: ChatRequest, raw_request: Request) -> StreamingResponse:
    """SSE 流式对话接口：实时推送思考过程、工具调用、最终回复。

    延迟初始化架构：SSE 连接立即建立并推送进度事件，
    会话获取、配额检查、@引用解析等阻塞操作在流内部执行，
    实现毫秒级首次视觉反馈。
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if _config_incomplete:
        return _error_json_response(
            503,
            "模型尚未配置，请先在设置页面或 .env 文件中配置 API Key、Base URL 和 Model。",
        )
    if not (request.message or "").strip() and not request.images:
        return _error_json_response(400, "消息内容不能为空。")

    # ── 仅保留纯内存操作在流外部（微秒级） ──
    from excelmanus.auth.dependencies import extract_user_id
    auth_user_id = extract_user_id(raw_request)
    isolation_user_id = _get_isolation_user_id(raw_request)
    _is_service = getattr(raw_request.state, "is_service_token", False)
    _bridge = getattr(raw_request.app.state, "event_bridge", None)

    rate_limiter = getattr(raw_request.app.state, "rate_limiter", None)
    if rate_limiter is not None and auth_user_id:
        rate_limiter.check_chat(auth_user_id)

    async def _event_generator() -> AsyncIterator[str]:
        """SSE 事件生成器：所有阻塞操作在首个 yield 之后执行。

        延迟初始化架构：SSE 连接立即建立并推送进度事件，
        会话获取、配额检查、@引用解析等阻塞操作在流内部执行，
        实现毫秒级首次视觉反馈。
        """
        safe_mode = _is_external_safe_mode()
        _is_channel_request = bool(request.channel)

        # ── 所有可能在 finally 中引用的变量预初始化 ──
        session_id: str | None = None
        engine: AgentEngine | None = None
        acquired = False
        stream_state: _SessionStreamState | None = None
        chat_task: asyncio.Task[Any] | None = None
        queue_get_task: asyncio.Task[ToolCallEvent | None] | None = None

        async def _cancel_task(task: asyncio.Task[Any] | None) -> None:
            if task is None or task.done():
                return
            task.cancel()
            try:
                await task
            except asyncio.CancelledError:
                pass

        _last_pipeline_stage = "initializing"

        try:
            # ── 立即推送进度，让前端在任何阻塞操作前就有视觉反馈 ──
            yield _sse_format("pipeline_progress", {
                "stage": "initializing",
                "message": "正在初始化...",
            })

            # ── 延迟初始化：会话获取（可能创建 Engine + MCP sync） ──
            try:
                session_id, engine = await _session_manager.acquire_for_chat(
                    request.session_id, user_id=isolation_user_id,
                )
                acquired = True
            except Exception as exc:
                _sid = request.session_id or "unknown"
                logger.error(
                    "会话获取失败 [session=%s]: %s",
                    _sid, exc, exc_info=True,
                )
                _guidance = classify_failure(exc, stage="initializing")
                yield _failure_guidance_sse(_guidance)
                yield _sse_format("done", {})
                return

            assert session_id is not None and engine is not None
            yield _sse_format("session_init", {"session_id": session_id})

            # EventBridge: 通知其他渠道 chat 已开始
            _origin_channel = request.channel or "web"
            if _bridge is not None and isolation_user_id:
                _fire_and_forget(
                    _bridge.notify(isolation_user_id, "chat_started", {
                        "session_id": session_id,
                        "origin_channel": _origin_channel,
                        "message_preview": (request.message or "")[:80],
                    }),
                    name="bridge_chat_started",
                )

            # ── 工作区配额检查 ──
            _auth_enabled = getattr(raw_request.app.state, "auth_enabled", False)
            if _auth_enabled:
                try:
                    _ws = _resolve_workspace(raw_request)
                    _usage = _ws.get_usage()
                    if _usage.over_files or _usage.over_size:
                        _parts: list[str] = []
                        if _usage.over_files:
                            _parts.append(f"文件数 {_usage.file_count}/{_usage.max_files}")
                        if _usage.over_size:
                            _parts.append(f"存储 {_usage.size_mb} MB/{_usage.max_size_mb} MB")
                        _detail = "、".join(_parts)
                        _wf_guidance = classify_workspace_full(
                            stage="initializing", detail=_detail,
                        )
                        yield _failure_guidance_sse(_wf_guidance)
                        yield _sse_format("done", {})
                        return
                except Exception:
                    logger.debug("配额检查异常，跳过", exc_info=True)

            # ── 保存命令快速路径 ──
            if _is_save_command(request.message):
                try:
                    reply_text = await _handle_save_command(
                        _session_manager, session_id, engine, request.message
                    )
                    yield _sse_format("reply", {
                        "content": guard_public_reply(reply_text),
                        "skills_used": [],
                        "tool_scope": [],
                        "route_mode": "control_command",
                        "iterations": 0,
                        "truncated": False,
                        "prompt_tokens": 0,
                        "completion_tokens": 0,
                        "total_tokens": 0,
                    })
                except Exception as exc:
                    logger.error("SSE /save 流异常: %s", exc, exc_info=True)
                    _save_provider = _extract_provider_from_base_url(
                        getattr(engine, "active_base_url", None) or (getattr(_config, "base_url", None) if _config else None)
                    )
                    _save_model = getattr(engine, "current_model", None) or (getattr(_config, "model", "") if _config else "")
                    _save_guidance = classify_failure(exc, stage="save_command", provider=_save_provider, model=_save_model)
                    yield _failure_guidance_sse(_save_guidance)
                else:
                    yield _sse_format("done", {})
                return

            # ── @引用解析 + 图片日志 ──
            display_text, mention_contexts = await _resolve_mentions(
                request.message, engine,
            )

            if request.images:
                logger.info(
                    "chat_stream 收到 %d 张图片附件 (media_types=%s, data_lens=%s)",
                    len(request.images),
                    [img.media_type for img in request.images],
                    [len(img.data) for img in request.images],
                )

            # ── 设置事件流管道 ──
            stream_state = _SessionStreamState()
            _session_stream_states[session_id] = stream_state
            event_queue = stream_state.attach()

            _sse_event_count = 0

            _flush_scheduled = False

            def _schedule_flush() -> None:
                """将同步 flush 调度为异步任务，避免在事件回调中阻塞事件循环。

                同一轮 tool batch 中多个 TOOL_CALL_END 事件只触发一次 flush：
                首个 TOOL_CALL_END 设置标记并调度，后续的跳过。
                标记在 flush 完成后重置，为下一轮 batch 做准备。
                """
                nonlocal _flush_scheduled
                if _flush_scheduled:
                    return
                _flush_scheduled = True

                async def _do_flush() -> None:
                    nonlocal _flush_scheduled
                    try:
                        if _session_manager is not None:
                            await asyncio.to_thread(
                                _session_manager.flush_messages_sync, session_id,
                            )
                    except Exception:
                        logger.debug("异步消息持久化失败", exc_info=True)
                    finally:
                        _flush_scheduled = False

                _fire_and_forget(_do_flush(), name="flush_messages")

            def _on_event(event: ToolCallEvent) -> None:
                """引擎事件回调：通过 stream_state 投递，同时持久化 Excel 事件。

                消息持久化通过 asyncio.to_thread 异步执行，不阻塞事件循环。
                """
                nonlocal _sse_event_count
                _sse_event_count += 1
                has_subscriber = stream_state.subscriber_queue is not None
                logger.debug(
                    "SSE 事件投递 #%d [%s] subscriber=%s qsize=%s",
                    _sse_event_count,
                    event.event_type.value if hasattr(event.event_type, 'value') else event.event_type,
                    has_subscriber,
                    stream_state.subscriber_queue.qsize() if has_subscriber else "N/A",
                )
                stream_state.deliver(event)
                _persist_excel_event(session_id, event)
                if (
                    event.event_type == EventType.TOOL_CALL_END
                    and _session_manager is not None
                ):
                    _schedule_flush()
                # EventBridge: 跨渠道实时事件推送（Web↔Bot 双向）
                # origin_channel 用于接收方过滤自身发出的事件，防止回声
                if (
                    _bridge is not None
                    and isolation_user_id
                    and event.event_type in (
                        EventType.PENDING_APPROVAL,
                        EventType.USER_QUESTION,
                        EventType.APPROVAL_RESOLVED,
                    )
                ):
                    _bridge_data: dict[str, Any] = {}
                    _bridge_evt = ""
                    _origin = request.channel or "web"
                    if event.event_type == EventType.PENDING_APPROVAL:
                        _bridge_evt = "approval"
                        _bridge_data = {
                            "approval_id": event.approval_id,
                            "approval_tool_name": event.approval_tool_name,
                            "risk_level": event.approval_risk_level,
                            "args_summary": event.approval_args_summary or {},
                            "session_id": session_id,
                            "origin_channel": _origin,
                        }
                    elif event.event_type == EventType.USER_QUESTION:
                        _bridge_evt = "question"
                        _bridge_data = {
                            "id": event.question_id,
                            "header": event.question_header,
                            "text": event.question_text,
                            "options": event.question_options or [],
                            "session_id": session_id,
                            "origin_channel": _origin,
                        }
                    elif event.event_type == EventType.APPROVAL_RESOLVED:
                        _bridge_evt = "approval_resolved"
                        _bridge_data = {
                            "approval_id": event.approval_id,
                            "session_id": session_id,
                            "origin_channel": _origin,
                        }
                    if _bridge_evt:
                        _fire_and_forget(
                            _bridge.notify(isolation_user_id, _bridge_evt, _bridge_data),
                            name=f"bridge_{_bridge_evt}",
                        )

            async def _run_chat_inner() -> ChatResult:
                """后台执行 engine.chat，完成后释放会话锁。"""
                try:
                    result = await engine.chat(
                        display_text,
                        on_event=_on_event,
                        mention_contexts=mention_contexts,
                        images=_serialize_images(request.images),
                        chat_mode=request.chat_mode,
                        channel=request.channel,
                    )
                    return _normalize_chat_result(result)
                finally:
                    await _session_manager.release_for_chat(session_id)
                    nonlocal acquired
                    acquired = False

            # ── 启动 chat 任务 ──
            chat_task = asyncio.create_task(_run_chat_inner())
            _active_chat_tasks[session_id] = chat_task

            def _cleanup_active_chat_task(done_task: asyncio.Task[Any]) -> None:
                """后台 chat 任务完成后清理活跃任务映射与流状态。"""
                if _active_chat_tasks.get(session_id) is done_task:
                    _active_chat_tasks.pop(session_id, None)
                ss = _session_stream_states.get(session_id)
                if ss is not None and ss.subscriber_queue is None:
                    _session_stream_states.pop(session_id, None)
                try:
                    done_task.result()
                except asyncio.CancelledError:
                    pass
                except Exception:
                    logger.warning(
                        "会话 %s 的后台聊天任务异常结束",
                        session_id,
                        exc_info=True,
                    )

            chat_task.add_done_callback(_cleanup_active_chat_task)
            queue_get_task = asyncio.create_task(event_queue.get())

            # ── 事件消费循环 ──
            while True:
                assert queue_get_task is not None
                done, _ = await asyncio.wait(
                    [queue_get_task, chat_task],
                    return_when=asyncio.FIRST_COMPLETED,
                )

                if queue_get_task in done:
                    event = queue_get_task.result()
                    if event is not None:
                        if event.event_type == EventType.PIPELINE_PROGRESS and event.pipeline_stage:
                            _last_pipeline_stage = event.pipeline_stage
                        sse = _sse_event_to_sse(event, safe_mode=safe_mode, is_channel=_is_channel_request)
                        if sse is not None:
                            yield sse
                    if chat_task.done():
                        queue_get_task = None
                    else:
                        queue_get_task = asyncio.create_task(event_queue.get())

                if chat_task in done:
                    # 排空队列中剩余事件
                    while True:
                        try:
                            event = event_queue.get_nowait()
                        except asyncio.QueueEmpty:
                            break
                        if event is not None:
                            if event.event_type == EventType.PIPELINE_PROGRESS and event.pipeline_stage:
                                _last_pipeline_stage = event.pipeline_stage
                            sse = _sse_event_to_sse(event, safe_mode=safe_mode, is_channel=_is_channel_request)
                            if sse is not None:
                                yield sse
                    break

            # chat 任务完成：获取结果
            chat_result = chat_task.result()
            normalized_reply = guard_public_reply((chat_result.reply or "").strip())
            yield _build_reply_sse(chat_result, engine)

            # EventBridge: 通知其他渠道 chat 已完成
            if _bridge is not None and isolation_user_id:
                _fire_and_forget(
                    _bridge.notify(isolation_user_id, "chat_completed", {
                        "session_id": session_id,
                        "origin_channel": _origin_channel,
                        "reply_summary": normalized_reply[:300] if normalized_reply else "",
                        "tool_count": len(chat_result.tool_calls),
                        "has_error": any(not tc.success for tc in chat_result.tool_calls),
                    }),
                    name="bridge_chat_completed",
                )

            # 记录已认证用户的 token 使用量
            if auth_user_id and chat_result.total_tokens > 0:
                try:
                    _user_store = getattr(raw_request.app.state, "user_store", None)
                    if _user_store is not None:
                        _user_store.record_token_usage(auth_user_id, chat_result.total_tokens)
                except Exception:
                    logger.debug("记录用户 token 用量失败", exc_info=True)
            # ── 自动标题（仅首轮）：立即截取用户消息，后台异步 AI 生成 ──
            if engine.session_turn == 1:
                _instant_title = _truncate_user_message_as_title(display_text)
                if _instant_title:
                    yield _sse_format("session_title", {
                        "session_id": session_id,
                        "title": _instant_title,
                    })
                    # 同步写入 DB，确保 poll 能立即拿到
                    _ch = _session_manager.chat_history if _session_manager else None
                    if _ch is not None:
                        try:
                            _ch.update_session(session_id, title=_instant_title, title_source="truncated")
                        except Exception:
                            logger.debug("截取标题写入失败", exc_info=True)
                # fire-and-forget：后台 AI 生成更好的标题，前端通过 SessionSync poll 获取
                _fire_and_forget(
                    _generate_session_title_background(
                        session_id=session_id,
                        user_message=display_text,
                        assistant_reply=normalized_reply,
                    ),
                    name="generate_session_title",
                )
            yield _sse_format("done", {})

        except (asyncio.CancelledError, GeneratorExit):
            # 客户端断开（如页面刷新）时允许 chat 在后台继续。
            if stream_state is not None:
                stream_state.detach()
            if session_id is not None:
                logger.info("会话 %s 的流式连接已断开，后台任务继续执行", session_id)
        except Exception as exc:
            logger.error(
                "SSE 流异常: %s", exc, exc_info=True
            )
            _provider = _extract_provider_from_base_url(
                getattr(engine, "active_base_url", None) or (getattr(_config, "base_url", None) if _config else None)
            )
            _model_name = getattr(engine, "current_model", None) or (getattr(_config, "model", "") if _config else "")
            _top_guidance = classify_failure(
                exc,
                stage=_last_pipeline_stage,
                provider=_provider,
                model=_model_name,
            )
            _persist_failure_guidance_message(
                session_id=session_id,
                engine=engine,
                guidance=_top_guidance,
            )
            yield _failure_guidance_sse(_top_guidance)
            yield _sse_format("done", {})
            # 确保 chat 任务被取消
            if chat_task is not None and not chat_task.done():
                chat_task.cancel()
                try:
                    await chat_task
                except (asyncio.CancelledError, Exception):
                    pass
        finally:
            if stream_state is not None:
                stream_state.detach()
            if chat_task is not None and chat_task.done():
                if _active_chat_tasks.get(session_id) is chat_task:
                    _active_chat_tasks.pop(session_id, None)
                _session_stream_states.pop(session_id, None)
            await _cancel_task(queue_get_task)
            # 安全网：确保 in_flight 锁被释放（正常路径已在 _run_chat_inner 中释放）
            if acquired and session_id is not None:
                try:
                    await _session_manager.release_for_chat(session_id)
                except Exception:
                    logger.debug("安全网 release_for_chat 异常", exc_info=True)

    return StreamingResponse(
        _event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


class AbortRequest(BaseModel):
    """终止请求体。"""

    model_config = ConfigDict(extra="forbid")

    session_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]


class GuideRequest(BaseModel):
    """引导消息请求体：向运行中的会话注入追加指令。"""

    model_config = ConfigDict(extra="forbid")

    message: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=4096)
    ]


class AnswerQuestionRequest(BaseModel):
    """回答问题请求体（阻塞式 ask_user）。"""

    model_config = ConfigDict(extra="forbid")

    question_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]
    answer: str


class ApproveRequest(BaseModel):
    """审批决策请求体（阻塞式审批）。"""

    model_config = ConfigDict(extra="forbid")

    approval_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]
    decision: Literal["accept", "reject", "fullaccess"]


class BackupApplyRequest(BaseModel):
    """将备份文件应用回原文件。"""
    session_id: str
    files: list[str] | None = Field(default=None, description="要应用的原始文件路径列表。为空或 null 时应用全部。")


class BackupDiscardRequest(BaseModel):
    """丢弃备份文件。"""
    session_id: str
    files: list[str] | None = Field(default=None, description="要丢弃的原始文件路径列表。为空或 null 时丢弃全部。")


class BackupUndoRequest(BaseModel):
    """撤销已应用的备份。"""
    session_id: str
    original_path: str = Field(description="原始文件路径")
    undo_path: str = Field(description="undo 备份文件路径")


class RollbackRequest(BaseModel):
    """对话回退请求体。"""

    model_config = ConfigDict(extra="forbid")

    session_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]
    turn_index: int = Field(..., ge=0, description="目标用户轮次索引（0-indexed）")
    rollback_files: bool = Field(default=False, description="是否同时回滚文件变更")
    new_message: str | None = Field(default=None, description="替换该轮用户消息内容（可选）")
    resend_mode: bool = Field(default=False, description="重发模式：移除目标用户消息，调用方随后通过 /chat/stream 重新发送")


@_router.get("/api/v1/backup/list")
async def backup_list(session_id: str, request: Request) -> JSONResponse:
    """列出指定会话的待应用备份文件。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)
    if engine is None:
        # 区分: session 属于其他用户 → 404（安全隔离）; session 不存在 → 200 空默认值（local-first）
        if user_id and _session_manager.get_engine(session_id) is not None:
            return _error_json_response(404, f"会话 '{session_id}' 不存在或未加载。")
        return JSONResponse(status_code=200, content={"files": [], "backup_enabled": False})
    tx = engine.transaction
    if tx is None:
        return JSONResponse(status_code=200, content={"files": [], "backup_enabled": False})
    staged = tx.list_staged()
    files = []
    for b in staged:
        bp = Path(b["backup"])
        file_info: dict = {
            "original_path": tx.to_relative(b["original"]),
            "backup_path": tx.to_relative(b["backup"]),
            "exists": b["exists"] == "True",
            "modified_at": bp.stat().st_mtime if bp.exists() else None,
        }
        # 附加轻量变更摘要
        try:
            summary = tx.diff_staged_summary(b["original"])
            if summary:
                file_info["summary"] = summary
        except Exception:
            pass
        files.append(file_info)
    # 检查 agent 是否活跃（用于前端 in-flight 提示）
    in_flight = False
    if _session_manager is not None:
        try:
            in_flight = await _session_manager.is_session_in_flight(session_id)
        except Exception:
            pass
    return JSONResponse(status_code=200, content={
        "files": files,
        "backup_enabled": True,
        "in_flight": in_flight,
    })


@_router.post("/api/v1/backup/apply")
async def backup_apply(request: BackupApplyRequest, raw_request: Request) -> JSONResponse:
    """将备份副本应用回原始文件。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    # B3: 原子化检查 in_flight + 获取 engine，消除 TOCTOU 竞态
    try:
        engine = await _session_manager.get_engine_if_idle(request.session_id, user_id=user_id)
    except SessionBusyError:
        return _error_json_response(409, "会话正在处理中，请等待完成后再应用备份。")
    if engine is None:
        return _error_json_response(404, f"会话 '{request.session_id}' 不存在或未加载。")
    tx = engine.transaction
    if tx is None:
        return _error_json_response(400, "该会话未启用备份模式。")
    if request.files:
        applied = []
        original_rel_paths: set[str] = set()
        for fp in request.files:
            result = tx.commit_one(fp)
            if result:
                item: dict = {
                    "original": tx.to_relative(result["original"]),
                    "backup": tx.to_relative(result["backup"]),
                }
                if result.get("undo_path"):
                    item["undo_path"] = result["undo_path"]
                applied.append(item)
                original_rel_paths.add(tx.to_relative(result["original"]))
        if original_rel_paths:
            engine._approval.mark_non_undoable_for_paths(original_rel_paths)
        remaining = len(tx.list_staged())
        return JSONResponse(status_code=200, content={
            "status": "ok", "applied": applied, "count": len(applied),
            "pending_count": remaining,
        })
    else:
        raw_applied = tx.commit_all()
        applied = []
        original_rel_paths_all: set[str] = set()
        for a in raw_applied:
            item_all: dict = {
                "original": tx.to_relative(a["original"]),
                "backup": tx.to_relative(a["backup"]),
            }
            if a.get("undo_path"):
                item_all["undo_path"] = a["undo_path"]
            applied.append(item_all)
            original_rel_paths_all.add(tx.to_relative(a["original"]))
        if original_rel_paths_all:
            engine._approval.mark_non_undoable_for_paths(original_rel_paths_all)
        return JSONResponse(status_code=200, content={
            "status": "ok", "applied": applied, "count": len(applied),
            "pending_count": 0,
        })


@_router.post("/api/v1/backup/discard")
async def backup_discard(request: BackupDiscardRequest, raw_request: Request) -> JSONResponse:
    """丢弃备份映射。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    # B3: 原子化检查 in_flight + 获取 engine，消除 TOCTOU 竞态
    try:
        engine = await _session_manager.get_engine_if_idle(request.session_id, user_id=user_id)
    except SessionBusyError:
        return _error_json_response(409, "会话正在处理中，请等待完成后再丢弃备份。")
    if engine is None:
        return _error_json_response(404, f"会话 '{request.session_id}' 不存在或未加载。")
    tx = engine.transaction
    if tx is None:
        return _error_json_response(400, "该会话未启用备份模式。")
    if request.files:
        count = 0
        for fp in request.files:
            if tx.rollback_one(fp):
                count += 1
        remaining = len(tx.list_staged())
        return JSONResponse(status_code=200, content={"status": "ok", "discarded": count, "pending_count": remaining})
    else:
        tx.rollback_all()
        return JSONResponse(status_code=200, content={"status": "ok", "discarded": "all", "pending_count": 0})


@_router.post("/api/v1/backup/undo")
async def backup_undo(request: BackupUndoRequest, raw_request: Request) -> JSONResponse:
    """撤销已应用的备份，将原始文件恢复到应用前的状态。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    engine = _session_manager.get_engine(request.session_id, user_id=user_id)
    if engine is None:
        return _error_json_response(404, f"会话 '{request.session_id}' 不存在或未加载。")
    tx = engine.transaction
    if tx is None:
        return _error_json_response(400, "该会话未启用备份模式。")
    ok = tx.undo_commit(request.original_path, request.undo_path)
    if not ok:
        return _error_json_response(400, "撤销失败：undo 备份文件不存在或已过期。")
    return JSONResponse(status_code=200, content={"status": "ok", "undone": request.original_path})


# ── 工作区事务别名（规范名称） ────────
# 委托到 backup_* 处理器以保持向后兼容。
# /backup/* 路径保留为废弃别名。

@_router.get("/api/v1/workspace/staged")
async def workspace_staged(session_id: str, request: Request) -> JSONResponse:
    return await backup_list(session_id, request)


@_router.post("/api/v1/workspace/commit")
async def workspace_commit(request: BackupApplyRequest, raw_request: Request) -> JSONResponse:
    return await backup_apply(request, raw_request)


@_router.post("/api/v1/workspace/rollback")
async def workspace_rollback(request: BackupDiscardRequest, raw_request: Request) -> JSONResponse:
    return await backup_discard(request, raw_request)


# ── 轮次 Checkpoint API ──────────────────────────────────


@_router.get("/api/v1/checkpoint/list")
async def checkpoint_list(session_id: str, request: Request) -> JSONResponse:
    """列出指定会话的轮次 checkpoint 时间线。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)
    if engine is None:
        return _error_json_response(404, f"会话 '{session_id}' 不存在或未加载。")
    if not engine.checkpoint_enabled:
        return JSONResponse(status_code=200, content={
            "checkpoints": [], "checkpoint_enabled": False,
        })
    _reg = engine.file_registry
    if _reg is None or not getattr(_reg, 'has_versions', False):
        return JSONResponse(status_code=200, content={
            "checkpoints": [], "checkpoint_enabled": True,
            "error": "FileRegistry not available",
        })
    cps = _reg.list_turn_checkpoints()
    items = []
    for cp in cps:
        items.append({
            "turn_number": cp.turn_number,
            "created_at": cp.created_at,
            "files_modified": cp.files_modified,
            "tool_names": cp.tool_names,
            "version_count": len(cp.version_ids),
        })
    return JSONResponse(status_code=200, content={
        "checkpoints": items, "checkpoint_enabled": True,
    })


class CheckpointRollbackRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    session_id: str
    turn_number: int


@_router.post("/api/v1/checkpoint/rollback")
async def checkpoint_rollback(
    request: CheckpointRollbackRequest, raw_request: Request,
) -> JSONResponse:
    """回退到指定轮次之前的文件状态。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    # B3: 原子化检查 in_flight + 获取 engine，消除 TOCTOU 竞态
    try:
        engine = await _session_manager.get_engine_if_idle(request.session_id, user_id=user_id)
    except SessionBusyError:
        return _error_json_response(409, "会话正在处理中，请等待完成后再回退。")
    if engine is None:
        return _error_json_response(404, f"会话 '{request.session_id}' 不存在或未加载。")
    if not engine.checkpoint_enabled:
        return _error_json_response(400, "该会话未启用 checkpoint 模式。")
    _reg = engine.file_registry
    if _reg is None or not getattr(_reg, 'has_versions', False):
        return _error_json_response(400, "FileRegistry not available for rollback.")
    restored = _reg.rollback_to_turn(request.turn_number)
    return JSONResponse(status_code=200, content={
        "status": "ok",
        "turn_number": request.turn_number,
        "restored_files": restored,
        "count": len(restored),
    })


class RollbackPreviewRequest(BaseModel):
    """回滚预览请求体。"""
    model_config = ConfigDict(extra="forbid")
    session_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]
    turn_index: int = Field(..., ge=0, description="目标用户轮次索引（0-indexed）")


@_router.post("/api/v1/chat/rollback/preview")
async def chat_rollback_preview(
    request: RollbackPreviewRequest, raw_request: Request,
) -> JSONResponse:
    """预览回滚到指定用户轮次后会影响的文件变更（不实际执行）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    engine = _session_manager.get_engine(request.session_id, user_id=user_id)
    if engine is None:
        return _error_json_response(404, f"会话 '{request.session_id}' 不存在或未加载。")
    try:
        preview = engine.rollback_preview(request.turn_index)
    except IndexError as exc:
        return _error_json_response(400, str(exc))
    return JSONResponse(status_code=200, content=preview)


@_router.post("/api/v1/chat/rollback")
async def chat_rollback(request: RollbackRequest, raw_request: Request) -> JSONResponse:
    """回退对话到指定用户轮次，可选回滚文件变更。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)

    try:
        result = await _session_manager.rollback_session(
            request.session_id,
            request.turn_index,
            rollback_files=request.rollback_files,
            new_message=request.new_message,
            resend_mode=request.resend_mode,
            user_id=user_id,
        )
    except SessionNotFoundError as exc:
        return _error_json_response(404, str(exc))
    except SessionBusyError:
        return _error_json_response(409, "会话正在处理中，请等待完成后再回退。")
    except IndexError as exc:
        return _error_json_response(400, str(exc))

    return JSONResponse(
        status_code=200,
        content={
            "status": "ok",
            "removed_messages": result["removed_messages"],
            "file_rollback_results": result["file_rollback_results"],
            "turn_index": result["turn_index"],
        },
    )


@_router.get("/api/v1/chat/turns")
async def chat_turns(session_id: str, request: Request) -> JSONResponse:
    """列出指定会话的用户轮次摘要。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)
    if engine is None:
        return _error_json_response(404, f"会话 '{session_id}' 不存在或未加载。")

    turns = engine.list_user_turns()
    return JSONResponse(
        status_code=200,
        content={"turns": turns},
    )


class _SubscribeRequest(BaseModel):
    """SSE 重连请求体。"""

    model_config = ConfigDict(extra="forbid")

    session_id: Annotated[
        str, StringConstraints(strip_whitespace=True, min_length=1, max_length=128)
    ]
    skip_replay: bool = Field(
        default=False,
        description="跳过缓冲事件重放，仅接收实时新事件。"
        "前端已从后端加载了当前消息时使用，避免重复 block。",
    )


@_router.post("/api/v1/chat/subscribe", responses=_error_responses)
async def chat_subscribe(request: _SubscribeRequest, raw_request: Request) -> StreamingResponse:
    """SSE 重连端点：页面刷新后重新接入正在执行的聊天任务事件流。

    先重放断连期间缓冲的事件，再实时推送后续事件，直到任务完成。
    若任务已结束，返回 done 事件后关闭。
    """
    session_id = request.session_id

    if not await _has_session_access(session_id, raw_request):
        return _error_json_response(404, f"会话 '{session_id}' 不存在。")  # type: ignore[return-value]

    chat_task = _active_chat_tasks.get(session_id)
    stream_state = _session_stream_states.get(session_id)

    # 没有活跃任务或 stream_state → 返回即时 done
    if chat_task is None or chat_task.done() or stream_state is None:
        async def _done_stream() -> AsyncIterator[str]:
            yield _sse_format("session_init", {"session_id": session_id})
            yield _sse_format("subscribe_resume", {"status": "completed"})
            yield _sse_format("done", {})

        return StreamingResponse(
            _done_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )

    # 附着新的订阅者队列
    event_queue = stream_state.attach()
    # skip_replay=True 时仅保留 SSE-only 事件（thinking/iteration 等不进 SQLite，
    # 前端无法从后端消息恢复），丢弃可从后端持久化消息恢复的事件。
    _SSE_ONLY_EVENT_TYPES = {
        EventType.THINKING, EventType.THINKING_DELTA,
        EventType.ITERATION_START, EventType.RETRACT_THINKING,
    }
    if request.skip_replay:
        all_buffered = stream_state.drain_buffer()
        buffered_events = [e for e in all_buffered if e.event_type in _SSE_ONLY_EVENT_TYPES]
    else:
        buffered_events = stream_state.drain_buffer()

    safe_mode = _is_external_safe_mode()

    async def _subscribe_generator() -> AsyncIterator[str]:
        yield _sse_format("session_init", {"session_id": session_id})
        yield _sse_format("subscribe_resume", {
            "status": "reconnected",
            "buffered_count": len(buffered_events),
        })

        # 重放缓冲事件（skip_replay 时为空）
        for event in buffered_events:
            sse = _sse_event_to_sse(event, safe_mode=safe_mode)
            if sse is not None:
                yield sse

        # 实时消费后续事件
        queue_get_task: asyncio.Task[ToolCallEvent | None] | None = asyncio.create_task(
            event_queue.get()
        )

        async def _cancel_task(task: asyncio.Task[Any] | None) -> None:
            if task is None or task.done():
                return
            task.cancel()
            try:
                await task
            except asyncio.CancelledError:
                pass

        try:
            while True:
                assert queue_get_task is not None
                wait_set: set[asyncio.Task[Any]] = {queue_get_task}
                if not chat_task.done():
                    wait_set.add(chat_task)
                done_set, _ = await asyncio.wait(
                    wait_set,
                    return_when=asyncio.FIRST_COMPLETED,
                )

                if queue_get_task in done_set:
                    event = queue_get_task.result()
                    if event is not None:
                        sse = _sse_event_to_sse(event, safe_mode=safe_mode)
                        if sse is not None:
                            yield sse
                    if chat_task.done():
                        queue_get_task = None
                    else:
                        queue_get_task = asyncio.create_task(event_queue.get())

                if chat_task.done() and (queue_get_task is None or queue_get_task not in done_set):
                    # 排空队列中剩余事件
                    while True:
                        try:
                            event = event_queue.get_nowait()
                        except asyncio.QueueEmpty:
                            break
                        if event is not None:
                            sse = _sse_event_to_sse(event, safe_mode=safe_mode)
                            if sse is not None:
                                yield sse
                    break

            # chat 任务完成：获取结果并发送 reply + done
            if not chat_task.cancelled():
                try:
                    chat_result = chat_task.result()
                    if _session_manager is None:
                        return  # 服务未初始化，静默退出
                    engine = _session_manager.get_engine(session_id)
                    if engine is not None:
                        normalized_reply = guard_public_reply((chat_result.reply or "").strip())
                        yield _build_reply_sse(chat_result, engine)
                        # B2: 首轮标题（与 chat_stream 保持一致：截取 + 后台 AI）
                        if engine.session_turn == 1:
                            _user_msg = ""
                            for _m in engine.raw_messages:
                                if _m.get("role") == "user":
                                    _user_msg = str(_m.get("content", ""))[:200]
                                    break
                            _instant_title = _truncate_user_message_as_title(_user_msg)
                            if _instant_title:
                                yield _sse_format("session_title", {
                                    "session_id": session_id,
                                    "title": _instant_title,
                                })
                                _ch = _session_manager.chat_history if _session_manager else None
                                if _ch is not None:
                                    try:
                                        _ch.update_session(session_id, title=_instant_title, title_source="truncated")
                                    except Exception:
                                        pass
                            _fire_and_forget(_generate_session_title_background(
                                session_id=session_id,
                                user_message=_user_msg,
                                assistant_reply=normalized_reply,
                            ), name="session_title")
                except asyncio.CancelledError:
                    pass
                except Exception as exc:
                    logger.warning(
                        "subscribe 获取 chat 结果失败: %s",
                        exc, exc_info=True,
                    )
                    _sub_guidance = classify_failure(exc, stage="subscribe_resume")
                    yield _failure_guidance_sse(_sub_guidance)
            yield _sse_format("done", {})

        except (asyncio.CancelledError, GeneratorExit):
            stream_state.detach()
            logger.info("会话 %s 的重连流式连接再次断开", session_id)
        except Exception as exc:
            logger.error(
                "SSE subscribe 流异常: %s", exc, exc_info=True
            )
            _sub_top_guidance = classify_failure(exc, stage="subscribe_resume")
            yield _failure_guidance_sse(_sub_top_guidance)
            yield _sse_format("done", {})
        finally:
            stream_state.detach()
            if chat_task.done():
                _session_stream_states.pop(session_id, None)
            await _cancel_task(queue_get_task)

    return StreamingResponse(
        _subscribe_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@_router.post("/api/v1/chat/abort")
async def chat_abort(request: AbortRequest, raw_request: Request) -> JSONResponse:
    """终止指定会话的活跃聊天任务。"""
    if not await _has_session_access(request.session_id, raw_request):
        return JSONResponse(
            status_code=200,
            content={"status": "no_active_task"},
        )
    task = _active_chat_tasks.get(request.session_id)
    if task is None or task.done():
        return JSONResponse(
            status_code=200,
            content={"status": "no_active_task"},
        )
    # 中断当前会话正在执行的 sleep 工具（会话隔离，不影响其他会话）
    if _session_manager is not None:
        engine = _session_manager.get_engine(request.session_id)
        if engine is not None and engine._tool_dispatcher is not None:
            engine._tool_dispatcher.cancel_active_sleep()

    task.cancel()
    logger.info("通过 abort 端点取消会话 %s 的聊天任务", request.session_id)
    return JSONResponse(
        status_code=200,
        content={"status": "cancelled"},
    )


@_router.post("/api/v1/chat/{session_id}/guide", responses=_error_responses)
async def chat_guide(
    session_id: str,
    request: GuideRequest,
    raw_request: Request,
) -> JSONResponse:
    """向运行中的会话注入引导消息（不启动新 chat）。

    消息存入 engine 内部队列，agent 在下次 LLM 迭代时自动看到。
    即使当前没有 in-flight 任务也可投递，下次 chat() 时生效。
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, raw_request):
        return JSONResponse(status_code=403, content={"error": "无权访问此会话"})

    engine = _session_manager.get_engine(session_id)
    if engine is None:
        return _error_json_response(404, f"会话 '{session_id}' 不存在或未加载")

    engine.push_guide_message(request.message)
    in_flight = session_id in _active_chat_tasks and not _active_chat_tasks[session_id].done()
    logger.info(
        "Guide 消息已投递: session=%s, in_flight=%s, len=%d",
        session_id[:8], in_flight, len(request.message),
    )
    return JSONResponse(
        status_code=200,
        content={"status": "delivered", "in_flight": in_flight},
    )


@_router.post("/api/v1/chat/{session_id}/answer", responses=_error_responses)
async def chat_answer(
    session_id: str,
    request: AnswerQuestionRequest,
    raw_request: Request,
) -> JSONResponse:
    """提交 ask_user 问题的回答，resolve 阻塞中的 Future。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, raw_request):
        return JSONResponse(status_code=403, content={"error": "无权访问此会话"})

    engine = _session_manager.get_engine(session_id)
    if engine is None:
        return JSONResponse(status_code=404, content={"error": "会话不存在或未激活"})

    registry = engine.interaction_registry
    # 构造 payload：兼容 QuestionFlowManager 的 parse_answer 格式
    payload = {"raw_input": request.answer, "question_id": request.question_id}

    # 尝试解析选项（如果 question_flow 中有对应问题）
    try:
        pending_q = engine._question_flow.current()
        if pending_q is not None and pending_q.question_id == request.question_id:
            parsed = engine._question_flow.parse_answer(request.answer, pending_q)
            payload = parsed.to_tool_result()
    except Exception:
        logger.debug("解析回答失败，使用原始文本", exc_info=True)

    ok = registry.resolve(request.question_id, payload)
    if not ok:
        return JSONResponse(
            status_code=404,
            content={"error": f"问题 {request.question_id} 不存在或已回答"},
        )
    logger.info("问题已回答: session=%s question=%s", session_id, request.question_id)
    return JSONResponse(status_code=200, content={"status": "answered"})


@_router.post("/api/v1/chat/{session_id}/approve", responses=_error_responses)
async def chat_approve(
    session_id: str,
    request: ApproveRequest,
    raw_request: Request,
) -> JSONResponse:
    """提交审批决策，resolve 阻塞中的 Future。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, raw_request):
        return JSONResponse(status_code=403, content={"error": "无权访问此会话"})

    engine = _session_manager.get_engine(session_id)
    if engine is None:
        return JSONResponse(status_code=404, content={"error": "会话不存在或未激活"})

    registry = engine.interaction_registry
    payload = {"decision": request.decision, "approval_id": request.approval_id}
    ok = registry.resolve(request.approval_id, payload)
    if not ok:
        return JSONResponse(
            status_code=404,
            content={"error": f"审批 {request.approval_id} 不存在或已处理"},
        )
    logger.info(
        "审批已决策: session=%s approval=%s decision=%s",
        session_id, request.approval_id, request.decision,
    )
    return JSONResponse(status_code=200, content={"status": "resolved"})



def _truncate_user_message_as_title(user_message: str, max_len: int = 60) -> str | None:
    """从用户消息截取会话标题（去除文件通知前缀，取前 max_len 字符）。"""
    if not user_message:
        return None
    # 去除 [已上传文件: ...] / [已上传图片: ...] 等前缀
    import re
    cleaned = re.sub(r"\[已上传(?:文件|图片): [^\]]*\]\s*", "", user_message).strip()
    if not cleaned:
        return None
    # 截取并添加省略号
    if len(cleaned) > max_len:
        return cleaned[:max_len].rstrip() + "…"
    return cleaned


async def _generate_session_title_background(
    *,
    session_id: str,
    user_message: str,
    assistant_reply: str,
) -> None:
    """后台 fire-and-forget：用 AUX 模型生成更好的会话标题并更新 DB。

    前端通过 SessionSync 的 list_sessions 轮询自动获取更新后的标题。
    """
    try:
        title = await _generate_session_title_with_timeout(
            session_id=session_id,
            user_message=user_message,
            assistant_reply=assistant_reply,
            timeout=8.0,
        )
        if title:
            logger.info("会话 %s 后台 AI 标题已更新: %s", session_id, title)
    except Exception:
        logger.debug("会话 %s 后台标题生成失败", session_id, exc_info=True)


async def _generate_session_title_with_timeout(
    *,
    session_id: str,
    user_message: str,
    assistant_reply: str,
    timeout: float = 5.0,
) -> str | None:
    """用 AUX 模型生成会话标题，超时返回 None。"""
    if _config is None or _session_manager is None:
        return None
    ch = _session_manager.chat_history
    if ch is None:
        return None

    _aux_effective = _config.aux_enabled and bool(_config.aux_model)
    if not _aux_effective:
        return None

    try:
        from excelmanus.providers import create_client as _create_client
        from excelmanus.session_title import generate_session_title

        client = _create_client(
            api_key=_config.aux_api_key or _config.api_key,
            base_url=_config.aux_base_url or _config.base_url,
            protocol=_config.aux_protocol,
        )
        title = await asyncio.wait_for(
            generate_session_title(
                user_message=user_message,
                assistant_reply=assistant_reply,
                client=client,
                model=_config.aux_model,
            ),
            timeout=timeout,
        )
        if title:
            ch.update_session(session_id, title=title, title_source="auto")
            logger.info("会话 %s 自动标题: %s", session_id, title)
        return title
    except asyncio.TimeoutError:
        logger.info("会话 %s 标题生成超时 (%.1fs)", session_id, timeout)
        return None
    except Exception:
        logger.warning("会话 %s 标题生成失败", session_id, exc_info=True)
        return None


def _sse_event_to_sse(
    event: ToolCallEvent,
    *,
    safe_mode: bool,
    is_channel: bool = False,
) -> str | None:
    """将 ToolCallEvent 转换为 SSE 文本（委托到 api/sse.py）。"""
    return _sse_event_to_sse_impl(
        event,
        safe_mode=safe_mode,
        is_channel=is_channel,
        public_path_fn=lambda path, sm: _public_excel_path(path, safe_mode=sm),
    )


@_router.get(
    "/api/v1/skills",
    response_model=list[SkillpackSummaryResponse],
    responses={
        500: _error_responses[500],
    },
)
async def list_skills(raw_request: Request) -> list[SkillpackSummaryResponse] | JSONResponse:
    """列出全部已加载 skillpack 摘要（per-user 隔离）。"""
    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    details = manager.list_skillpacks()
    return [_to_skill_summary(detail) for detail in details]


@_router.get(
    "/api/v1/skills/{name}",
    response_model=SkillpackDetailResponse | SkillpackSummaryResponse,
    responses={
        404: _error_responses[404],
        422: _error_responses[422],
        500: _error_responses[500],
    },
)
async def get_skill(name: str, raw_request: Request) -> SkillpackDetailResponse | SkillpackSummaryResponse | JSONResponse:
    """查询单个 skillpack（per-user 隔离）。"""
    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    try:
        detail = manager.get_skillpack(name)
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except SkillpackNotFoundError as exc:
        return _error_json_response(404, str(exc))

    if _is_external_safe_mode():
        return _to_skill_summary(detail)
    return _to_skill_detail(detail)


@_router.post(
    "/api/v1/skills",
    status_code=201,
    response_model=SkillpackMutationResponse,
    responses={
        403: _error_responses[403],
        409: _error_responses[409],
        422: _error_responses[422],
        500: _error_responses[500],
    },
)
async def create_skill(
    request: SkillpackCreateRequest,
    raw_request: Request,
) -> SkillpackMutationResponse | JSONResponse:
    """创建 skillpack（per-user 隔离）。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止写入 skillpack。")

    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    try:
        detail = manager.create_skillpack(
            name=request.name,
            payload=request.payload,
            actor=user_id or "api",
        )
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except SkillpackConflictError as exc:
        return _error_json_response(409, str(exc))

    if _user_skill_service is not None:
        _user_skill_service.invalidate(user_id)
    return SkillpackMutationResponse(
        status="created",
        name=str(detail.get("name", request.name)),
        detail=_to_standard_skill_detail_dict(detail),
    )


@_router.patch(
    "/api/v1/skills/{name}",
    response_model=SkillpackMutationResponse,
    responses={
        403: _error_responses[403],
        404: _error_responses[404],
        409: _error_responses[409],
        422: _error_responses[422],
        500: _error_responses[500],
    },
)
async def patch_skill(
    name: str,
    request: SkillpackPatchRequest,
    raw_request: Request,
) -> SkillpackMutationResponse | JSONResponse:
    """更新 skillpack（per-user 隔离）。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止写入 skillpack。")

    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    try:
        detail = manager.patch_skillpack(
            name=name,
            payload=request.payload,
            actor=user_id or "api",
        )
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except SkillpackNotFoundError as exc:
        return _error_json_response(404, str(exc))
    except SkillpackConflictError as exc:
        return _error_json_response(409, str(exc))

    if _user_skill_service is not None:
        _user_skill_service.invalidate(user_id)
    return SkillpackMutationResponse(
        status="updated",
        name=str(detail.get("name", name)),
        detail=_to_standard_skill_detail_dict(detail),
    )


@_router.delete(
    "/api/v1/skills/{name}",
    response_model=SkillpackMutationResponse,
    responses={
        403: _error_responses[403],
        404: _error_responses[404],
        409: _error_responses[409],
        422: _error_responses[422],
        500: _error_responses[500],
    },
)
async def delete_skill(
    name: str,
    raw_request: Request,
    reason: str = "",
) -> SkillpackMutationResponse | JSONResponse:
    """软删除 skillpack（per-user 隔离）。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止写入 skillpack。")

    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    try:
        detail = manager.delete_skillpack(
            name=name,
            actor=user_id or "api",
            reason=reason,
        )
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except SkillpackNotFoundError as exc:
        return _error_json_response(404, str(exc))
    except SkillpackConflictError as exc:
        return _error_json_response(409, str(exc))

    if _user_skill_service is not None:
        _user_skill_service.invalidate(user_id)
    return SkillpackMutationResponse(
        status="deleted",
        name=str(detail.get("name", name)),
        detail=detail,
    )


@_router.post(
    "/api/v1/skills/import",
    status_code=201,
    response_model=SkillpackMutationResponse,
    responses={
        403: _error_responses[403],
        409: _error_responses[409],
        422: _error_responses[422],
        500: _error_responses[500],
    },
)
async def import_skill(
    request: SkillpackImportRequest,
    raw_request: Request,
) -> SkillpackMutationResponse | JSONResponse:
    """从本地路径或 GitHub URL 导入 SKILL.md 及附属资源（per-user 隔离）。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止写入 skillpack。")

    user_id = _get_isolation_user_id(raw_request)
    manager = _require_user_skill_manager(user_id)
    try:
        result = await manager.import_skillpack_async(
            source=request.source,
            value=request.value,
            actor=user_id or "api",
            overwrite=request.overwrite,
        )
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except SkillpackConflictError as exc:
        return _error_json_response(409, str(exc))
    except SkillImportError as exc:
        return _error_json_response(422, str(exc))

    if _user_skill_service is not None:
        _user_skill_service.invalidate(user_id)
    return SkillpackMutationResponse(
        status="imported",
        name=str(result.get("name", "")),
        detail=result,
    )


# ── ClawHub API ──────────────────────────────────────────


@_router.get("/api/v1/clawhub/search")
async def clawhub_search(
    q: str = "",
    limit: int = 15,
) -> dict[str, Any]:
    """搜索 ClawHub 技能市场。"""
    if not q.strip():
        return {"results": []}
    manager = _require_skillpack_manager()
    try:
        results = await manager.clawhub_search(q.strip(), limit=limit)
    except ClawHubError as exc:
        return _error_json_response(502, f"ClawHub 请求失败：{exc}")
    return {"results": results}


@_router.get("/api/v1/clawhub/skill/{slug}")
async def clawhub_skill_detail(slug: str) -> dict[str, Any]:
    """获取 ClawHub 技能详情。"""
    manager = _require_skillpack_manager()
    try:
        detail = await manager.clawhub_skill_detail(slug)
    except ClawHubNotFoundError as exc:
        return _error_json_response(404, str(exc))
    except ClawHubError as exc:
        return _error_json_response(502, f"ClawHub 请求失败：{exc}")
    return detail


class ClawHubInstallRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    slug: str
    version: str | None = None
    overwrite: bool = False


@_router.post("/api/v1/clawhub/install", status_code=201)
async def clawhub_install(
    request: ClawHubInstallRequest,
) -> dict[str, Any]:
    """从 ClawHub 安装技能。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止安装。")
    manager = _require_skillpack_manager()
    try:
        result = await manager.import_skillpack_async(
            source="clawhub",
            value=request.slug,
            actor="api",
            overwrite=request.overwrite,
        )
    except ClawHubNotFoundError as exc:
        return _error_json_response(404, str(exc))
    except ClawHubError as exc:
        logger.warning("ClawHub 安装失败 slug=%s: %s", request.slug, exc, exc_info=True)
        return _error_json_response(502, f"ClawHub 安装失败：{exc}")
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    return {"status": "installed", **result}


@_router.get("/api/v1/clawhub/updates")
async def clawhub_check_updates() -> dict[str, Any]:
    """检查已安装 ClawHub 技能的可用更新。"""
    manager = _require_skillpack_manager()
    try:
        updates = await manager.clawhub_check_updates()
    except ClawHubError as exc:
        return _error_json_response(502, f"ClawHub 请求失败：{exc}")
    return {"updates": updates}


class ClawHubUpdateRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    slug: str | None = None
    version: str | None = None
    all: bool = False


@_router.post("/api/v1/clawhub/update")
async def clawhub_update(
    request: ClawHubUpdateRequest,
) -> dict[str, Any]:
    """更新 ClawHub 技能。"""
    if _is_external_safe_mode():
        return _error_json_response(403, "external_safe_mode 开启时禁止更新。")
    manager = _require_skillpack_manager()
    try:
        results = await manager.clawhub_update(
            slug=request.slug,
            version=request.version,
            update_all=request.all,
        )
    except SkillpackInputError as exc:
        return _error_json_response(422, str(exc))
    except ClawHubError as exc:
        return _error_json_response(502, f"ClawHub 更新失败：{exc}")
    return {"results": results}


@_router.get("/api/v1/clawhub/installed")
async def clawhub_list_installed() -> dict[str, Any]:
    """列出已安装的 ClawHub 技能。"""
    manager = _require_skillpack_manager()
    try:
        installed = await manager.clawhub_list_installed()
    except ClawHubError as exc:
        return _error_json_response(502, str(exc))
    return {"installed": installed}


@_router.delete("/api/v1/sessions/{session_id}", responses={
    409: _error_responses[409],
    404: _error_responses[404],
    500: _error_responses[500],
})
async def delete_session(session_id: str, request: Request) -> dict:
    """删除指定会话并释放资源。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    deleted = await _session_manager.delete(session_id, user_id=user_id)
    if not deleted:
        raise SessionNotFoundError(f"会话 '{session_id}' 不存在。")
    return {"status": "ok", "session_id": session_id}


@_router.patch("/api/v1/sessions/{session_id}/archive", responses={
    404: _error_responses[404],
    500: _error_responses[500],
})
async def archive_session(session_id: str, request: Request) -> dict:
    """归档或取消归档会话。

    请求体: {"archive": true}  归档
    请求体: {"archive": false} 取消归档
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")

    body = await request.json()
    archive = body.get("archive", True)
    user_id = _get_isolation_user_id(request)

    updated = await _session_manager.archive_session(
        session_id, archive=archive, user_id=user_id
    )
    if not updated:
        raise SessionNotFoundError(f"会话 '{session_id}' 不存在。")
    return {
        "status": "ok",
        "session_id": session_id,
        "archived": archive,
    }


@_router.patch("/api/v1/sessions/{session_id}/title", responses={
    404: _error_responses[404],
    500: _error_responses[500],
})
async def update_session_title_api(session_id: str, request: Request) -> JSONResponse:
    """用户手动更新会话标题。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    body = await request.json()
    title = (body.get("title") or "").strip()
    if not title:
        return _error_json_response(400, "标题不能为空")
    if len(title) > 100:
        return _error_json_response(400, "标题长度不能超过 100 字符")
    user_id = _get_isolation_user_id(request)
    ok = await _session_manager.update_session_title(
        session_id, title, user_id=user_id
    )
    if not ok:
        return _error_json_response(404, "会话不存在")
    return JSONResponse(content={"status": "ok", "title": title})


# ── Approvals / Undo ─────────────────────────────────────


@_router.get("/api/v1/approvals")
async def list_approvals(request: Request) -> JSONResponse:
    """列出已执行的审批记录（支持分页与筛选）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")

    limit = int(request.query_params.get("limit", "50"))
    undoable_only = request.query_params.get("undoable_only", "false").lower() == "true"
    session_id = request.query_params.get("session_id")
    if not session_id:
        return _error_json_response(400, "缺少 session_id 参数。")
    if not await _has_session_access(session_id, request):
        return _error_json_response(404, "会话不存在。")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)

    if engine is None:
        return JSONResponse(content={"approvals": []})

    records = engine._approval.list_applied(limit=limit, undoable_only=undoable_only)
    items = []
    for rec in records:
        item: dict = {
            "id": rec.approval_id,
            "tool_name": rec.tool_name,
            "created_at_utc": rec.created_at_utc,
            "applied_at_utc": rec.applied_at_utc,
            "execution_status": rec.execution_status,
            "undoable": rec.undoable,
            "result_preview": sanitize_external_text(rec.result_preview or "", max_len=200),
        }
        if not _is_external_safe_mode():
            item["arguments"] = sanitize_external_data(rec.arguments)
            item["changes"] = [
                {"path": c.path, "before_exists": c.before_exists, "after_exists": c.after_exists}
                for c in (rec.changes or [])
            ]
        items.append(item)
    return JSONResponse(content={"approvals": items})


@_router.post("/api/v1/approvals/{approval_id}/undo")
async def undo_approval(approval_id: str, request: Request) -> JSONResponse:
    """回滚指定审批操作。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")

    session_id = request.query_params.get("session_id")
    if not session_id:
        return _error_json_response(400, "缺少 session_id 参数。")
    if not await _has_session_access(session_id, request):
        return _error_json_response(404, "会话不存在。")
    user_id = _get_isolation_user_id(request)
    engine = await _session_manager.get_or_restore_engine(session_id, user_id=user_id)

    if engine is None:
        return _error_json_response(404, "没有活跃会话。")

    result_msg = engine._approval.undo(approval_id)
    success = "已回滚" in result_msg
    return JSONResponse(content={
        "status": "ok" if success else "error",
        "message": result_msg,
        "approval_id": approval_id,
    })


# ── 操作历史时间线 API ────────────────────────────────────


def _change_type(before_exists: bool, after_exists: bool) -> str:
    """从 before/after 存在状态推导变更类型。"""
    if not before_exists and after_exists:
        return "added"
    if before_exists and not after_exists:
        return "deleted"
    return "modified"


@_router.get("/api/v1/sessions/{session_id}/operations")
async def list_operations(
    session_id: str,
    request: Request,
    limit: int = 50,
    offset: int = 0,
) -> JSONResponse:
    """列出指定会话的写入操作历史（时间线）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, request):
        return _error_json_response(404, f"会话 '{session_id}' 不存在。")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)

    if engine is None:
        return JSONResponse(content={"operations": [], "total": 0, "has_more": False})

    from excelmanus.tools.policy import sanitize_approval_args_summary

    # 获取比请求多 1 条以判断 has_more
    records = engine._approval.list_applied(
        limit=offset + limit + 1,
        session_id=session_id,
    )
    total = len(records)
    has_more = total > offset + limit
    page = records[offset : offset + limit]

    items = []
    safe_mode = _is_external_safe_mode()
    for rec in page:
        changes = []
        for c in rec.changes or []:
            changes.append({
                "path": c.path,
                "change_type": _change_type(c.before_exists, c.after_exists),
                "before_size": c.before_size,
                "after_size": c.after_size,
                "is_binary": c.is_binary,
            })
        item: dict[str, Any] = {
            "approval_id": rec.approval_id,
            "tool_name": rec.tool_name,
            "arguments_summary": (
                sanitize_approval_args_summary(rec.arguments)
                if not safe_mode else {}
            ),
            "session_turn": rec.session_turn,
            "created_at_utc": rec.created_at_utc,
            "applied_at_utc": rec.applied_at_utc,
            "execution_status": rec.execution_status,
            "undoable": rec.undoable,
            "changes": changes,
            "result_preview": sanitize_external_text(
                rec.result_preview or "", max_len=300,
            ),
        }
        items.append(item)

    return JSONResponse(content={
        "operations": items,
        "total": total,
        "has_more": has_more,
    })


@_router.get("/api/v1/sessions/{session_id}/operations/{approval_id}")
async def get_operation_detail(
    session_id: str,
    approval_id: str,
    request: Request,
) -> JSONResponse:
    """获取单条操作的详情（含 diff 内容）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, request):
        return _error_json_response(404, f"会话 '{session_id}' 不存在。")
    user_id = _get_isolation_user_id(request)
    engine = _session_manager.get_engine(session_id, user_id=user_id)

    if engine is None:
        return _error_json_response(404, "没有活跃会话。")

    rec = engine._approval.get_applied(approval_id)
    if rec is None:
        return _error_json_response(404, f"操作 '{approval_id}' 不存在。")
    if rec.session_id and rec.session_id != session_id:
        return _error_json_response(404, f"操作 '{approval_id}' 不属于此会话。")

    from excelmanus.tools.policy import sanitize_approval_args_summary

    safe_mode = _is_external_safe_mode()

    changes = []
    for c in rec.changes or []:
        changes.append({
            "path": c.path,
            "change_type": _change_type(c.before_exists, c.after_exists),
            "before_size": c.before_size,
            "after_size": c.after_size,
            "is_binary": c.is_binary,
        })

    # 读取 patch 文件内容（如果存在）
    patch_content: str | None = None
    if rec.patch_file and not safe_mode:
        patch_path = Path(engine._config.workspace_root) / rec.patch_file
        if patch_path.is_file():
            try:
                patch_content = patch_path.read_text(encoding="utf-8")[:50000]
            except OSError:
                pass

    result: dict[str, Any] = {
        "approval_id": rec.approval_id,
        "tool_name": rec.tool_name,
        "arguments_summary": (
            sanitize_approval_args_summary(rec.arguments)
            if not safe_mode else {}
        ),
        "arguments": (
            sanitize_external_data(rec.arguments) if not safe_mode else {}
        ),
        "session_turn": rec.session_turn,
        "created_at_utc": rec.created_at_utc,
        "applied_at_utc": rec.applied_at_utc,
        "execution_status": rec.execution_status,
        "undoable": rec.undoable,
        "changes": changes,
        "result_preview": sanitize_external_text(
            rec.result_preview or "", max_len=1000,
        ),
        "patch_content": patch_content,
        "error_type": rec.error_type,
        "error_message": rec.error_message,
    }
    return JSONResponse(content=result)


@_router.post("/api/v1/sessions/{session_id}/operations/{approval_id}/undo")
async def undo_operation(
    session_id: str,
    approval_id: str,
    request: Request,
) -> JSONResponse:
    """回滚指定操作。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    if not await _has_session_access(session_id, request):
        return _error_json_response(404, f"会话 '{session_id}' 不存在。")
    user_id = _get_isolation_user_id(request)
    engine = await _session_manager.get_or_restore_engine(
        session_id, user_id=user_id,
    )

    if engine is None:
        return _error_json_response(404, "没有活跃会话。")

    # 验证操作属于此会话
    rec = engine._approval.get_applied(approval_id)
    if rec is None:
        return _error_json_response(404, f"操作 '{approval_id}' 不存在。")
    if rec.session_id and rec.session_id != session_id:
        return _error_json_response(404, f"操作 '{approval_id}' 不属于此会话。")

    result_msg = engine._approval.undo(approval_id)
    success = "已回滚" in result_msg
    return JSONResponse(content={
        "status": "ok" if success else "error",
        "message": result_msg,
        "approval_id": approval_id,
    })


# ── Excel 预览 API ────────────────────────────────────────


def _resolve_excel_path(
    path: str,
    session_id: str | None = None,
    *,
    workspace_root: str | None = None,
    user_id: str | None = None,
) -> str | None:
    """将相对/绝对路径解析为安全的绝对路径。

    当提供 workspace_root 时基于该目录解析（多用户隔离），否则回退到全局配置。
    当提供 session_id 且该会话启用了备份模式时，自动将路径重定向到备份副本。

    支持的路径形式：
    - 相对路径: ``./foo.xlsx``, ``foo.xlsx``, ``subdir/foo.xlsx``
    - 绝对路径: 仅当位于 workspace 内时允许
    """
    if _config is None:
        return None
    from pathlib import Path

    ws_root = workspace_root or _config.workspace_root
    workspace = Path(ws_root).resolve()
    workspace_str = str(workspace)

    resolved: str | None = None

    candidate = Path(path)
    if candidate.is_absolute():
        abs_resolved = candidate.resolve()
        if str(abs_resolved).startswith(workspace_str) and abs_resolved.is_file():
            resolved = str(abs_resolved)
    else:
        target = (workspace / path).resolve()
        if str(target).startswith(workspace_str) and target.is_file():
            resolved = str(target)
        else:
            # Fallback: 裸文件名可能位于 outputs/ 或 scripts/ 等子目录
            for _subdir in ("outputs", "scripts", "uploads"):
                fallback = (workspace / _subdir / path).resolve()
                if str(fallback).startswith(workspace_str) and fallback.is_file():
                    resolved = str(fallback)
                    break

    if resolved is None:
        return None

    if session_id and _session_manager is not None:
        engine = _session_manager.get_engine(session_id, user_id=user_id)
        if engine is not None and engine.backup_enabled:
            tx = engine.transaction
            if tx is not None:
                try:
                    staged_resolved = tx.resolve_read(resolved)
                    if staged_resolved != resolved and Path(staged_resolved).is_file():
                        return staged_resolved
                except ValueError:
                    pass

    return resolved


@_router.get("/api/v1/files/excel/list")
async def list_excel_files(request: Request) -> JSONResponse:
    """扫描当前用户 workspace 中所有 Excel 文件，返回路径列表。"""
    assert _config is not None, "服务未初始化"
    from pathlib import Path as _Path

    workspace = _Path(_resolve_workspace_root(request)).resolve()
    excel_exts = {".xlsx", ".xls", ".xlsm", ".xlsb", ".csv"}
    skip_dirs = {"node_modules", "__pycache__", ".venv", ".git", ".next"}
    results: list[dict[str, Any]] = []

    import re
    _upload_prefix_re = re.compile(r"^[0-9a-f]{8}_")
    uploads_dir = str(workspace / "uploads")
    backups_dir = str((workspace / "outputs" / "backups").resolve())
    audits_dir = str((workspace / "outputs" / "audits").resolve())

    for root, dirs, filenames in os.walk(workspace):
        dirs[:] = [d for d in dirs if d not in skip_dirs and not d.startswith(".")]
        root_resolved = str(_Path(root).resolve())
        if root_resolved.startswith(backups_dir) or root_resolved.startswith(audits_dir):
            dirs.clear()
            continue
        for fname in filenames:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in excel_exts:
                continue
            full = os.path.join(root, fname)
            try:
                rel = os.path.relpath(full, workspace)
                mtime = os.path.getmtime(full)
            except OSError:
                continue
            display_name = fname
            if root == uploads_dir and _upload_prefix_re.match(fname):
                display_name = fname[9:]
            results.append({
                "path": f"./{rel}",
                "filename": display_name,
                "modified_at": mtime,
            })

    results.sort(key=lambda x: x["modified_at"], reverse=True)
    return JSONResponse(content={"files": results})


_MAX_WORKSPACE_FILES = 2000

# 工作区内部目录，不向前端暴露
_WORKSPACE_HIDDEN_DIRS = frozenset({
    ".tmp", ".staging", ".versions", ".git", ".venv",
    "__pycache__", "node_modules",
})

# uploads/ 下 hash 前缀正则（8位hex + 下划线）
_UPLOAD_PREFIX_RE = re.compile(r"^[0-9a-f]{8}_")


@_router.get("/api/v1/files/workspace/list")
async def list_workspace_files(request: Request) -> JSONResponse:
    """扫描用户工作区根目录中的文件与文件夹（用于文件树视图）。

    扫描范围：完整工作区根目录（包含 agent 创建的文件、uploads/ 等）。
    排除：隐藏目录/文件、内部目录（.tmp/.staging/scripts/temp 等）。
    """
    assert _config is not None, "服务未初始化"

    ws = _resolve_workspace(request)
    ws_root = str(ws.root_dir)  # 完整工作区根目录

    results: list[dict[str, Any]] = []
    for root, dirs, filenames in os.walk(ws_root):
        # 统一使用 / 分隔符（兼容 Windows）
        rel_root = os.path.relpath(root, ws_root).replace("\\", "/")

        # 过滤隐藏目录与内部目录
        dirs[:] = sorted(
            d for d in dirs
            if not d.startswith(".")
            and d not in _WORKSPACE_HIDDEN_DIRS
            # scripts/temp 是 run_code 临时脚本目录，不展示
            and not (rel_root == "scripts" and d == "temp")
            # outputs/backups 和 outputs/audits 是内部备份/审计目录
            and not (rel_root == "outputs" and d in ("backups", "audits"))
        )

        for dname in dirs:
            full = os.path.join(root, dname)
            try:
                mtime = os.path.getmtime(full)
            except OSError:
                continue
            rel = f"{rel_root}/{dname}" if rel_root != "." else dname
            results.append({
                "path": rel,
                "filename": dname,
                "modified_at": mtime,
                "is_dir": True,
            })
        for fname in sorted(filenames):
            if fname.startswith(".") or fname.startswith("_rc_") or fname.startswith("_sw_"):
                continue
            full = os.path.join(root, fname)
            try:
                mtime = os.path.getmtime(full)
            except OSError:
                continue
            display_name = fname
            # uploads/ 下的 hash 前缀文件显示原始名
            if rel_root == "uploads" or (rel_root != "." and rel_root.startswith("uploads/")):
                if _UPLOAD_PREFIX_RE.match(fname):
                    display_name = fname[9:]
            rel = f"{rel_root}/{fname}" if rel_root != "." else fname
            results.append({
                "path": rel,
                "filename": display_name,
                "modified_at": mtime,
                "is_dir": False,
            })
            if len(results) >= _MAX_WORKSPACE_FILES:
                break
        if len(results) >= _MAX_WORKSPACE_FILES:
            break

    results.sort(key=lambda x: (not x["is_dir"], x["path"].lower()))
    return JSONResponse(content={"files": results, "truncated": len(results) >= _MAX_WORKSPACE_FILES})


@_router.get("/api/v1/files/workspace/storage")
async def get_workspace_storage(request: Request) -> JSONResponse:
    """返回当前用户工作区的存储用量概览（用于前端进度条展示）。"""
    assert _config is not None, "服务未初始化"
    ws = _resolve_workspace(request)
    usage = ws.get_usage()
    return JSONResponse(content={
        "total_bytes": usage.total_bytes,
        "size_mb": usage.size_mb,
        "max_bytes": usage.max_bytes,
        "max_size_mb": usage.max_size_mb,
        "file_count": usage.file_count,
        "max_files": usage.max_files,
        "over_size": usage.over_size,
        "over_files": usage.over_files,
    })


@_router.get("/api/v1/files/registry")
async def get_file_registry(request: Request) -> JSONResponse:
    """返回 FileRegistry 全量文件列表 + 可选事件历史。

    Query params:
      - include_deleted: bool (default false) — 是否包含已软删除的文件
      - include_events: bool (default false) — 是否附带每个文件的事件历史
      - file_id: str (optional) — 仅返回指定文件及其事件/谱系
    """
    assert _config is not None, "服务未初始化"

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return JSONResponse(content={"files": [], "total": 0})

    include_deleted = request.query_params.get("include_deleted", "").lower() in ("1", "true")
    include_events = request.query_params.get("include_events", "").lower() in ("1", "true")
    file_id = request.query_params.get("file_id", "").strip()

    def _event_to_dict(evt: Any) -> dict[str, Any]:
        return {
            "id": evt.id,
            "file_id": evt.file_id,
            "event_type": evt.event_type,
            "session_id": evt.session_id,
            "turn": evt.turn,
            "tool_name": evt.tool_name,
            "details": evt.details,
            "created_at": evt.created_at,
        }

    # 单文件查询模式
    if file_id:
        entry = registry.get_by_id(file_id)
        if entry is None:
            # 退而求其次：按路径查
            entry = registry.get_by_path(file_id)
        if entry is None:
            return _error_json_response(404, f"文件未找到: {file_id}")
        file_dict = entry.to_dict()
        if include_events:
            file_dict["events"] = [_event_to_dict(e) for e in registry.get_events(entry.id)]
        children = registry.get_children(entry.id)
        file_dict["children"] = [c.to_dict() for c in children]
        lineage = registry.get_lineage(entry.id)
        file_dict["lineage"] = [a.to_dict() for a in lineage]
        return JSONResponse(content={"file": file_dict})

    # 全量列表模式
    entries = registry.list_all(include_deleted=include_deleted)
    files: list[dict[str, Any]] = []
    for entry in entries:
        d = entry.to_dict()
        if include_events:
            d["events"] = [_event_to_dict(e) for e in registry.get_events(entry.id)]
        files.append(d)

    return JSONResponse(content={"files": files, "total": len(files)})


# ── File Groups API ──────────────────────────────────────────


@_router.get("/api/v1/files/groups")
async def list_file_groups(request: Request) -> JSONResponse:
    """列出当前工作区的所有文件组（附成员摘要）。"""
    assert _config is not None, "服务未初始化"
    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return JSONResponse(content={"groups": []})

    try:
        groups = registry.list_groups()
        result = []
        for g in groups:
            members = registry.get_group_files(g.id)
            result.append({
                **g.to_dict(),
                "members": members,
            })
        return JSONResponse(content={"groups": result})
    except Exception:
        return JSONResponse(content={"groups": []})


@_router.post("/api/v1/files/groups")
async def create_file_group(request: Request) -> JSONResponse:
    """创建文件组。

    Body: {name: str, description?: str, file_ids?: [{id: str, role?: str}]}
    """
    assert _config is not None, "服务未初始化"
    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return _error_json_response(500, "FileRegistry 不可用")

    body = await request.json()
    name = body.get("name", "").strip()
    if not name:
        return _error_json_response(400, "缺少文件组名称")

    description = body.get("description", "")
    file_ids_raw = body.get("file_ids", [])

    # 提取纯 file_id 列表用于创建
    plain_ids = []
    role_map: dict[str, str] = {}
    for item in file_ids_raw:
        if isinstance(item, dict):
            fid = item.get("id", "")
            role = item.get("role", "member")
        else:
            fid = str(item)
            role = "member"
        if fid:
            plain_ids.append(fid)
            role_map[fid] = role

    try:
        group = registry.create_group(name, file_ids=plain_ids, description=description)
        # 设置角色（create_group 默认 member，需更新非默认角色）
        for fid, role in role_map.items():
            if role != "member":
                registry.add_to_group(group.id, fid, role)
        members = registry.get_group_files(group.id)
        return JSONResponse(status_code=201, content={**group.to_dict(), "members": members})
    except Exception as exc:
        return _error_json_response(500, f"创建文件组失败: {exc}")


@_router.put("/api/v1/files/groups/{group_id}")
async def update_file_group(group_id: str, request: Request) -> JSONResponse:
    """更新文件组名称/描述。

    Body: {name?: str, description?: str}
    """
    assert _config is not None, "服务未初始化"
    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return _error_json_response(500, "FileRegistry 不可用")

    body = await request.json()
    name = body.get("name")
    description = body.get("description")

    group = registry.update_group(group_id, name=name, description=description)
    if group is None:
        return _error_json_response(404, f"文件组未找到: {group_id}")

    members = registry.get_group_files(group.id)
    return JSONResponse(content={**group.to_dict(), "members": members})


@_router.delete("/api/v1/files/groups/{group_id}")
async def delete_file_group(group_id: str, request: Request) -> JSONResponse:
    """删除文件组。"""
    assert _config is not None, "服务未初始化"
    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return _error_json_response(500, "FileRegistry 不可用")

    ok = registry.delete_group(group_id)
    if not ok:
        return _error_json_response(404, f"文件组未找到: {group_id}")
    return JSONResponse(content={"status": "deleted", "group_id": group_id})


@_router.put("/api/v1/files/groups/{group_id}/members")
async def update_file_group_members(group_id: str, request: Request) -> JSONResponse:
    """管理文件组成员。

    Body: {add?: [{file_id: str, role?: str}], remove?: [str]}
    """
    assert _config is not None, "服务未初始化"
    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    registry = _get_file_registry(ws_root, user_id=_iso_uid)
    if registry is None:
        return _error_json_response(500, "FileRegistry 不可用")

    group = registry.get_group(group_id)
    if group is None:
        return _error_json_response(404, f"文件组未找到: {group_id}")

    body = await request.json()
    to_add = body.get("add", [])
    to_remove = body.get("remove", [])

    for item in to_add:
        if isinstance(item, dict):
            fid = item.get("file_id", "")
            role = item.get("role", "member")
        else:
            fid = str(item)
            role = "member"
        if fid:
            registry.add_to_group(group_id, fid, role)

    for fid in to_remove:
        if fid:
            registry.remove_from_group(group_id, str(fid))

    members = registry.get_group_files(group_id)
    return JSONResponse(content={**group.to_dict(), "members": members})


@_router.get("/api/v1/files/excel")
async def get_excel_file(request: Request) -> StreamingResponse:
    """返回 xlsx 文件二进制流供前端 Univer 加载。"""
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    session_id = request.query_params.get("session_id")
    if not path:
        return _error_json_response(400, "缺少 path 参数")  # type: ignore[return-value]

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    resolved = _resolve_excel_path(path, session_id, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"文件不存在或路径非法: {path}")  # type: ignore[return-value]

    from pathlib import Path as _Path

    file_path = _Path(resolved)
    suffix = file_path.suffix.lower()
    if suffix not in {".xlsx", ".xls", ".xlsb", ".xlsm", ".csv"}:
        return _error_json_response(400, f"不支持的文件格式: {suffix}")  # type: ignore[return-value]

    # .xls/.xlsb → 透明转换为 xlsx 供前端 Univer 加载
    actual_file = resolved
    from excelmanus.xls_converter import needs_conversion as _nc2, ensure_xlsx as _ensure2
    if _nc2(resolved):
        try:
            _xlsx_p2, _ = _ensure2(resolved)
            actual_file = str(_xlsx_p2)
            suffix = ".xlsx"
        except Exception:
            logger.warning("excel 流转换失败，返回原始文件: %s", resolved)

    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if suffix in (".xlsx", ".xlsm")
        else "text/csv"
    )

    def _iter_file():
        with open(actual_file, "rb") as f:  # type: ignore[arg-type]
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter_file(),
        media_type=content_type,
        headers={"Content-Disposition": _make_content_disposition(file_path.name)},
    )


@_router.get("/api/v1/files/spec")
async def get_spec_file(request: Request) -> JSONResponse:
    """返回 ReplicaSpec JSON 文件内容（供 VLM pipeline 迷你表格预览）。"""
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    if not path:
        return _error_json_response(400, "缺少 path 参数")  # type: ignore[return-value]

    from pathlib import Path as _Path

    ws_root = _resolve_workspace_root(request)
    file_path = _Path(path)
    if not file_path.is_absolute():
        file_path = _Path(ws_root) / file_path

    if not file_path.is_file() or file_path.suffix.lower() != ".json":
        return _error_json_response(404, f"Spec 文件不存在: {path}")  # type: ignore[return-value]

    import json as _json

    try:
        content = file_path.read_text(encoding="utf-8")
        data = _json.loads(content)
    except Exception as exc:
        return _error_json_response(500, f"读取 spec 失败: {exc}")  # type: ignore[return-value]

    return JSONResponse(content=data)


@_router.get("/api/v1/files/image")
async def get_image_file(request: Request) -> StreamingResponse:
    """返回 workspace 内图片文件的二进制流（供 VLM pipeline 原图预览）。"""
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    session_id = request.query_params.get("session_id")
    if not path:
        return _error_json_response(400, "缺少 path 参数")  # type: ignore[return-value]

    from pathlib import Path as _Path
    import mimetypes

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    
    # 使用与下载相同的路径解析逻辑
    resolved = _resolve_excel_path(path, session_id, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"图片文件不存在: {path}")  # type: ignore[return-value]
    
    file_path = _Path(resolved)

    _IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"}
    if not file_path.is_file() or file_path.suffix.lower() not in _IMAGE_EXTENSIONS:
        return _error_json_response(404, f"图片文件不存在: {path}")  # type: ignore[return-value]

    content_type = mimetypes.guess_type(file_path.name)[0] or "image/png"

    def _iter_image():
        with open(file_path, "rb") as f:
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter_image(),
        media_type=content_type,
        headers={"Cache-Control": "public, max-age=3600"},
    )


@_router.get("/api/v1/files/read")
async def read_text_file(request: Request) -> JSONResponse:
    """返回 workspace 内文本文件的内容（供代码/MD 预览）。"""
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    session_id = request.query_params.get("session_id")
    if not path:
        return _error_json_response(400, "缺少 path 参数")  # type: ignore[return-value]

    from pathlib import Path as _Path

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    
    # 使用与下载相同的路径解析逻辑
    resolved = _resolve_excel_path(path, session_id, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"文本文件不存在: {path}")  # type: ignore[return-value]
    
    file_path = _Path(resolved)

    # 支持的文本文件扩展名
    _TEXT_EXTENSIONS = {
        ".txt", ".md", ".markdown", ".json", ".js", ".jsx", ".ts", ".tsx",
        ".py", ".rb", ".go", ".rs", ".java", ".c", ".cpp", ".h", ".hpp",
        ".cs", ".php", ".swift", ".kt", ".scala", ".sh", ".bash", ".zsh",
        ".sql", ".html", ".css", ".scss", ".less", ".xml", ".yaml", ".yml",
        ".toml", ".ini", ".cfg", ".conf", ".log", ".env", ".gitignore",
        ".dockerignore", ".graphql", ".gql", ".vue", ".svelte", ".ex",
        ".exs", ".erl", ".hs", ".ml", ".fs", ".clj", ".lua", ".r", ".dart",
        ".groovy", ".txt", ".csv", ".tsv",
    }

    if not file_path.is_file() or file_path.suffix.lower() not in _TEXT_EXTENSIONS:
        return _error_json_response(404, f"文本文件不存在: {path}")  # type: ignore[return-value]

    # 限制文件大小（最大 1MB）
    if file_path.stat().st_size > 1024 * 1024:
        return _error_json_response(400, "文件过大，无法预览")  # type: ignore[return-value]

    try:
        content = file_path.read_text(encoding="utf-8")
        return JSONResponse(content={"content": content})
    except UnicodeDecodeError:
        return _error_json_response(400, "文件编码不支持，请使用 UTF-8 编码")  # type: ignore[return-value]
    except Exception as exc:
        return _error_json_response(500, f"读取文件失败: {exc}")  # type: ignore[return-value]


@_router.get("/api/v1/files/download")
async def download_file(request: Request) -> StreamingResponse:
    """通用文件下载：返回 workspace 内任意文件的二进制流。"""
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    session_id = request.query_params.get("session_id")
    if not path:
        return _error_json_response(400, "缺少 path 参数")  # type: ignore[return-value]

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    resolved = _resolve_excel_path(path, session_id, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"文件不存在或路径非法: {path}")  # type: ignore[return-value]

    from pathlib import Path as _Path
    import mimetypes

    file_path = _Path(resolved)
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"

    def _iter_file():
        with open(resolved, "rb") as f:  # type: ignore[arg-type]
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter_file(),
        media_type=content_type,
        headers={"Content-Disposition": _make_content_disposition(file_path.name)},
    )


@_router.get("/api/v1/files/dl/{token}")
async def download_file_by_token(token: str) -> StreamingResponse:
    """通过短效令牌下载文件（无需 auth，供 Bot 渠道分享下载链接）。"""
    assert _config is not None, "服务未初始化"

    from excelmanus.auth.security import decode_download_token

    claims = decode_download_token(token)
    if claims is None:
        return _error_json_response(403, "下载链接已过期或无效")  # type: ignore[return-value]

    file_path_str = claims.get("file_path", "")
    user_id = claims.get("sub", "")

    _auth_on = os.environ.get("EXCELMANUS_AUTH_ENABLED", "").strip().lower() in ("1", "true", "yes")
    from excelmanus.workspace import IsolatedWorkspace, SandboxConfig
    ws = IsolatedWorkspace.resolve(
        _config.workspace_root,
        user_id=user_id or None,
        auth_enabled=_auth_on,
        sandbox_config=SandboxConfig(docker_enabled=False),
        data_root=_config.data_root,
    )
    ws_root = str(ws.root_dir)

    resolved = _resolve_excel_path(file_path_str, None, workspace_root=ws_root, user_id=user_id)
    if resolved is None:
        return _error_json_response(404, "文件不存在或路径非法")  # type: ignore[return-value]

    from pathlib import Path as _Path
    import mimetypes

    fp = _Path(resolved)
    content_type = mimetypes.guess_type(fp.name)[0] or "application/octet-stream"

    def _iter():
        with open(resolved, "rb") as f:  # type: ignore[arg-type]
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter(),
        media_type=content_type,
        headers={"Content-Disposition": _make_content_disposition(fp.name)},
    )


@_router.post("/api/v1/files/download/link")
async def create_download_link(request: Request) -> JSONResponse:
    """生成文件的短效下载链接（供 Bot 渠道使用）。

    请求体: {"file_path": "...", "user_id": "..."}
    返回: {"url": "https://...", "token": "...", "expires_minutes": 30}
    """
    assert _config is not None, "服务未初始化"

    body = await request.json()
    file_path_str = body.get("file_path", "")
    user_id = body.get("user_id", "")

    if not file_path_str:
        return _error_json_response(400, "缺少 file_path 参数")

    # 验证文件存在
    _auth_on = os.environ.get("EXCELMANUS_AUTH_ENABLED", "").strip().lower() in ("1", "true", "yes")
    from excelmanus.workspace import IsolatedWorkspace, SandboxConfig
    ws = IsolatedWorkspace.resolve(
        _config.workspace_root,
        user_id=user_id or None,
        auth_enabled=_auth_on,
        sandbox_config=SandboxConfig(docker_enabled=False),
        data_root=_config.data_root,
    )
    ws_root = str(ws.root_dir)

    resolved = _resolve_excel_path(file_path_str, None, workspace_root=ws_root, user_id=user_id)
    if resolved is None:
        return _error_json_response(404, f"文件不存在或路径非法: {file_path_str}")

    from excelmanus.auth.security import create_download_token, DOWNLOAD_TOKEN_EXPIRE_MINUTES

    token = create_download_token(file_path_str, user_id=user_id)

    # 构建公开 URL
    public_url = _config.public_url
    if not public_url:
        # 回退：从请求 Host 推断
        scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
        host = request.headers.get("x-forwarded-host", request.headers.get("host", "localhost:8000"))
        public_url = f"{scheme}://{host}"

    download_url = f"{public_url}/api/v1/files/dl/{token}"

    return JSONResponse(content={
        "url": download_url,
        "token": token,
        "expires_minutes": DOWNLOAD_TOKEN_EXPIRE_MINUTES,
    })


@_router.get("/api/v1/files/excel/snapshot")
async def get_excel_snapshot(request: Request) -> JSONResponse:
    """返回 Excel 文件的轻量 JSON 快照（供聊天内嵌预览）。

    参数:
      - path: 文件路径
      - sheet: 指定工作表名（可选）
      - max_rows: 最大行数（默认 50）
      - session_id: 会话 ID（可选）
      - all_sheets: 设为 1 时一次返回所有工作表快照（减少 HTTP 往返）
    """
    assert _config is not None, "服务未初始化"

    path = request.query_params.get("path", "")
    sheet = request.query_params.get("sheet")
    max_rows = int(request.query_params.get("max_rows", "50"))
    session_id = request.query_params.get("session_id")
    all_sheets = request.query_params.get("all_sheets", "").strip() in ("1", "true")
    with_styles = request.query_params.get("with_styles", "1").strip() in ("1", "true")

    if not path:
        return _error_json_response(400, "缺少 path 参数")

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)
    resolved = _resolve_excel_path(path, session_id, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"文件不存在或路径非法: {path}")

    # ── CSV 快捷路径 ──────────────────────────────────────
    if os.path.splitext(resolved)[1].lower() == ".csv":
        try:
            import csv as _csv

            # 自动检测编码
            _enc = "utf-8"
            for _try_enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
                try:
                    with open(resolved, "r", encoding=_try_enc) as _f:
                        _f.read(4096)
                    _enc = _try_enc
                    break
                except (UnicodeDecodeError, LookupError):
                    continue

            with open(resolved, "r", encoding=_enc, newline="") as _f:
                reader = _csv.reader(_f)
                all_rows_raw: list[list[str]] = []
                for row in reader:
                    all_rows_raw.append(row)
                    if len(all_rows_raw) > max_rows + 1:
                        break

            total_rows = len(all_rows_raw)
            total_cols = max((len(r) for r in all_rows_raw), default=0)
            headers = all_rows_raw[0] if all_rows_raw else []
            col_letters = [chr(65 + i) if i < 26 else f"A{chr(65 + i - 26)}" for i in range(min(total_cols, 100))]
            data_rows = all_rows_raw[1: min(max_rows + 1, total_rows)]
            # 将纯数字字符串转换为数字
            converted_rows: list[list[Any]] = []
            for dr in data_rows:
                conv: list[Any] = []
                for v in dr:
                    if v == "":
                        conv.append(None)
                    else:
                        try:
                            conv.append(int(v))
                        except ValueError:
                            try:
                                conv.append(float(v))
                            except ValueError:
                                conv.append(v)
                converted_rows.append(conv)

            snap_csv: dict[str, Any] = {
                "file": os.path.basename(resolved),
                "sheet": "Sheet1",
                "sheets": ["Sheet1"],
                "shape": {"rows": total_rows, "columns": total_cols},
                "column_letters": col_letters,
                "headers": headers,
                "rows": converted_rows,
                "total_rows": total_rows,
                "truncated": total_rows > max_rows + 1,
            }
            if all_sheets:
                return JSONResponse(content={
                    "file": os.path.basename(resolved),
                    "sheets": ["Sheet1"],
                    "all_snapshots": [snap_csv],
                })
            return JSONResponse(content=snap_csv)
        except Exception as exc:
            logger.error("CSV snapshot 生成失败: %s", exc, exc_info=True)
            return _error_json_response(500, f"读取文件失败: {exc}")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        from excelmanus.tools._style_extract import extract_cell_style as _extract_cell_style

        # .xls/.xlsb → 透明转换为 xlsx 后再用 openpyxl 打开
        from excelmanus.xls_converter import needs_conversion as _nc, ensure_xlsx as _ensure
        _actual_path = resolved
        if _nc(resolved):
            try:
                _xlsx_p, _ = _ensure(resolved)
                _actual_path = str(_xlsx_p)
            except Exception:
                logger.warning("snapshot 转换失败，尝试直接打开: %s", resolved)

        wb = load_workbook(_actual_path, data_only=True, read_only=not with_styles)
        sheet_names = wb.sheetnames

        def _read_sheet(ws_obj: Any) -> dict:
            """读取单个工作表并返回快照 dict。"""
            s_total_rows = ws_obj.max_row or 0
            s_total_cols = ws_obj.max_column or 0
            s_headers: list[str] = []
            s_col_letters: list[str] = []
            for c in range(1, min(s_total_cols + 1, 101)):
                s_col_letters.append(get_column_letter(c))
                cell_val = ws_obj.cell(row=1, column=c).value
                s_headers.append(str(cell_val) if cell_val is not None else "")
            s_rows: list[list[Any]] = []
            s_row_limit = min(max_rows, s_total_rows, 200)
            for r in range(2, s_row_limit + 2):
                if r > s_total_rows:
                    break
                row_data: list[Any] = []
                for c in range(1, min(s_total_cols + 1, 101)):
                    val = ws_obj.cell(row=r, column=c).value
                    if val is None:
                        row_data.append(None)
                    elif isinstance(val, (int, float, bool)):
                        row_data.append(val)
                    else:
                        row_data.append(str(val))
                s_rows.append(row_data)

            result: dict[str, Any] = {
                "sheet": ws_obj.title,
                "shape": {"rows": s_total_rows, "columns": s_total_cols},
                "column_letters": s_col_letters,
                "headers": s_headers,
                "rows": s_rows,
                "total_rows": s_total_rows,
                "truncated": s_total_rows > s_row_limit,
            }

            # 提取样式（仅 with_styles=True 时）
            if with_styles:
                cell_styles: dict[str, dict] = {}
                merged: list[dict] = []
                for r in range(1, s_row_limit + 2):
                    if r > s_total_rows:
                        break
                    for c in range(1, min(s_total_cols + 1, 101)):
                        cell_obj = ws_obj.cell(row=r, column=c)
                        style = _extract_cell_style(cell_obj)
                        if style:
                            cell_styles[f"{r-1},{c-1}"] = style
                # 合并单元格
                try:
                    for merge_range in ws_obj.merged_cells.ranges:
                        merged.append({
                            "startRow": merge_range.min_row - 1,
                            "startColumn": merge_range.min_col - 1,
                            "endRow": merge_range.max_row - 1,
                            "endColumn": merge_range.max_col - 1,
                        })
                except Exception:
                    pass
                # 列宽
                col_widths: dict[str, float] = {}
                try:
                    for col_letter, dim in ws_obj.column_dimensions.items():
                        if dim.width and dim.width != 8.43:  # 默认宽度
                            col_idx = 0
                            for i, ch in enumerate(reversed(col_letter.upper())):
                                col_idx += (ord(ch) - 64) * (26 ** i)
                            col_widths[str(col_idx - 1)] = dim.width
                except Exception:
                    pass
                # 行高
                row_heights: dict[str, float] = {}
                try:
                    for row_idx, dim in ws_obj.row_dimensions.items():
                        if dim.height and dim.height != 15:  # 默认行高
                            row_heights[str(row_idx - 1)] = dim.height
                except Exception:
                    pass

                if cell_styles:
                    result["cell_styles"] = cell_styles
                if merged:
                    result["merged_cells"] = merged
                if col_widths:
                    result["column_widths"] = col_widths
                if row_heights:
                    result["row_heights"] = row_heights

            return result

        if all_sheets:
            # 一次返回所有工作表快照
            snapshots = []
            for sn in sheet_names:
                ws_obj = wb[sn]
                snapshots.append(_read_sheet(ws_obj))
            wb.close()
            return JSONResponse(content={
                "file": os.path.basename(resolved),
                "sheets": sheet_names,
                "all_snapshots": snapshots,
            })

        # 单 sheet 模式（向后兼容）
        ws = wb[sheet] if sheet and sheet in sheet_names else wb.active
        if ws is None:
            wb.close()
            return _error_json_response(404, "工作表不存在")

        snap = _read_sheet(ws)
        snap["file"] = os.path.basename(resolved)
        snap["sheets"] = sheet_names
        wb.close()
        return JSONResponse(content=snap)
    except Exception as exc:
        logger.error("Excel snapshot 生成失败: %s", exc, exc_info=True)
        return _error_json_response(500, f"读取文件失败: {exc}")


@_router.get("/api/v1/files/excel/compare")
async def get_excel_compare(request: Request) -> JSONResponse:
    """返回两个 Excel 文件的快照 + 跨文件列关系，供前端对比视图使用。

    参数:
      - path_a: 左侧文件路径
      - path_b: 右侧文件路径
      - session_id: 会话 ID（可选）
      - max_rows: 最大行数（默认 50）
    """
    assert _config is not None, "服务未初始化"

    path_a = request.query_params.get("path_a", "")
    path_b = request.query_params.get("path_b", "")
    session_id = request.query_params.get("session_id")
    max_rows = int(request.query_params.get("max_rows", "50"))

    if not path_a or not path_b:
        return _error_json_response(400, "缺少 path_a 或 path_b 参数")

    ws_root = _resolve_workspace_root(request)
    _iso_uid = _get_isolation_user_id(request)

    resolved_a = _resolve_excel_path(path_a, session_id, workspace_root=ws_root, user_id=_iso_uid)
    resolved_b = _resolve_excel_path(path_b, session_id, workspace_root=ws_root, user_id=_iso_uid)

    if resolved_a is None:
        return _error_json_response(404, f"文件不存在: {path_a}")
    if resolved_b is None:
        return _error_json_response(404, f"文件不存在: {path_b}")

    import asyncio

    def _load_snapshot(resolved: str) -> dict[str, Any]:
        """自包含的 snapshot 加载（支持 xlsx/xls/xlsb/csv）。"""
        basename = os.path.basename(resolved)
        ext = os.path.splitext(resolved)[1].lower()

        # ── CSV 快捷路径 ──
        if ext == ".csv":
            try:
                import csv as _csv
                _enc = "utf-8"
                for _try_enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
                    try:
                        with open(resolved, "r", encoding=_try_enc) as _f:
                            _f.read(4096)
                        _enc = _try_enc
                        break
                    except (UnicodeDecodeError, LookupError):
                        continue
                with open(resolved, "r", encoding=_enc, newline="") as _f:
                    reader = _csv.reader(_f)
                    all_rows_raw: list[list[str]] = []
                    for row in reader:
                        all_rows_raw.append(row)
                        if len(all_rows_raw) > max_rows + 1:
                            break
                total = len(all_rows_raw)
                total_cols = max((len(r) for r in all_rows_raw), default=0)
                headers = all_rows_raw[0] if all_rows_raw else []
                col_letters = [chr(65 + i) if i < 26 else f"A{chr(65 + i - 26)}" for i in range(min(total_cols, 100))]
                data_rows = all_rows_raw[1: min(max_rows + 1, total)]
                converted: list[list[Any]] = []
                for dr in data_rows:
                    conv: list[Any] = []
                    for v in dr:
                        if v == "":
                            conv.append(None)
                        else:
                            try:
                                conv.append(int(v))
                            except ValueError:
                                try:
                                    conv.append(float(v))
                                except ValueError:
                                    conv.append(v)
                    converted.append(conv)
                snap_csv = {
                    "file": basename, "sheet": "Sheet1", "sheets": ["Sheet1"],
                    "shape": {"rows": total, "columns": total_cols},
                    "column_letters": col_letters, "headers": headers,
                    "rows": converted, "total_rows": total,
                    "truncated": total > max_rows + 1,
                }
                return {"file": basename, "sheets": ["Sheet1"], "all_snapshots": [snap_csv]}
            except Exception as exc:
                logger.error("Compare CSV snapshot 失败: %s — %s", resolved, exc)
                return {"file": basename, "sheets": [], "all_snapshots": [], "error": str(exc)}

        # ── xls/xlsb 格式转换 ──
        _actual_path = resolved
        if ext in (".xls", ".xlsb"):
            try:
                from excelmanus.tools._helpers import ensure_openpyxl_compatible
                _actual_path = str(ensure_openpyxl_compatible(resolved))
            except Exception:
                logger.warning("Compare 格式转换失败，尝试直接打开: %s", resolved)

        # ── xlsx/xlsm 主路径 ──
        from openpyxl import load_workbook as _lwb

        def _read_ws(ws_obj: Any, _max_rows: int = max_rows) -> dict[str, Any]:
            from openpyxl.utils import get_column_letter as _gcl
            s_total_rows = ws_obj.max_row or 0
            s_total_cols = ws_obj.max_column or 0
            s_headers: list[str] = []
            s_col_letters: list[str] = []
            for c in range(1, min(s_total_cols + 1, 101)):
                s_col_letters.append(_gcl(c))
                cell_val = ws_obj.cell(row=1, column=c).value
                s_headers.append(str(cell_val) if cell_val is not None else "")
            s_rows: list[list[Any]] = []
            s_row_limit = min(_max_rows, s_total_rows, 200)
            for r in range(2, s_row_limit + 2):
                if r > s_total_rows:
                    break
                row_data: list[Any] = []
                for c in range(1, min(s_total_cols + 1, 101)):
                    val = ws_obj.cell(row=r, column=c).value
                    if val is None:
                        row_data.append(None)
                    elif isinstance(val, (int, float, bool)):
                        row_data.append(val)
                    else:
                        row_data.append(str(val))
                s_rows.append(row_data)
            return {
                "sheet": ws_obj.title,
                "shape": {"rows": s_total_rows, "columns": s_total_cols},
                "column_letters": s_col_letters,
                "headers": s_headers,
                "rows": s_rows,
                "total_rows": s_total_rows,
                "truncated": s_total_rows > s_row_limit,
            }

        try:
            wb = _lwb(_actual_path, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            snapshots = []
            for sn in sheet_names:
                ws_obj = wb[sn]
                snap = _read_ws(ws_obj)
                snap["file"] = basename
                snap["sheets"] = sheet_names
                snapshots.append(snap)
            wb.close()
            return {"file": basename, "sheets": sheet_names, "all_snapshots": snapshots}
        except Exception as exc:
            logger.error("Compare snapshot 失败: %s — %s", resolved, exc)
            return {"file": basename, "sheets": [], "all_snapshots": [], "error": str(exc)}

    snap_a, snap_b = await asyncio.gather(
        asyncio.to_thread(_load_snapshot, resolved_a),
        asyncio.to_thread(_load_snapshot, resolved_b),
    )

    # ── 跨文件关系检测 ──
    relationships: dict[str, Any] = {"shared_columns": []}
    try:
        from excelmanus.tools.data_tools import discover_file_relationships as _dfr
        import json as _json

        rel_json = await asyncio.to_thread(
            _dfr, file_paths=[resolved_a, resolved_b], max_files=2
        )
        rel_data = _json.loads(rel_json)
        pairs = rel_data.get("file_pairs", [])
        if pairs:
            relationships["shared_columns"] = pairs[0].get("shared_columns", [])
        hints = rel_data.get("merge_hints", [])
        if hints:
            relationships["merge_hint"] = hints[0]
    except Exception as exc:
        logger.debug("Compare 关系检测失败: %s", exc)

    return JSONResponse(content={
        "file_a": snap_a,
        "file_b": snap_b,
        "relationships": relationships,
    })


@_router.get("/api/v1/files/relationships")
async def get_file_relationships(request: Request) -> JSONResponse:
    """发现工作区内 Excel 文件之间的列关联关系。

    参数:
      - directory: 扫描目录（可选，默认 "."）
    """
    assert _config is not None, "服务未初始化"

    directory = request.query_params.get("directory", ".")

    import asyncio

    try:
        from excelmanus.tools.data_tools import discover_file_relationships as _dfr
        import json as _json

        rel_json = await asyncio.to_thread(_dfr, directory=directory, max_files=5)
        rel_data = _json.loads(rel_json)
        return JSONResponse(content=rel_data)
    except Exception as exc:
        logger.error("文件关系发现失败: %s", exc, exc_info=True)
        return _error_json_response(500, f"分析失败: {exc}")


class ExcelWriteRequest(BaseModel):
    """Excel 单元格写入请求。"""

    model_config = ConfigDict(extra="forbid")
    session_id: str | None = None
    path: str
    sheet: str | None = None
    changes: list[dict[str, Any]]


@_router.post("/api/v1/files/excel/write")
async def write_excel_cells(request: ExcelWriteRequest, raw_request: Request) -> JSONResponse:
    """侧边面板编辑回写：将单元格变更写入文件。

    当备份模式启用时，写操作会自动重定向到备份副本（通过 ensure_backup
    确保备份存在），避免直接修改原始文件。
    """
    assert _config is not None, "服务未初始化"

    ws_root = _resolve_workspace_root(raw_request)

    # 先解析原始路径（不经过 backup 重定向），用于 ensure_backup
    _iso_uid = _get_isolation_user_id(raw_request)
    resolved = _resolve_excel_path(request.path, None, workspace_root=ws_root, user_id=_iso_uid)
    if resolved is None:
        return _error_json_response(404, f"文件不存在或路径非法: {request.path}")

    # 备份模式下，写操作需要 ensure_backup（创建备份副本如果尚不存在），
    # 然后写入备份副本而非原始文件
    if request.session_id and _session_manager is not None:
        engine = _session_manager.get_engine(request.session_id, user_id=_iso_uid)
        if engine is not None and engine.backup_enabled:
            tx = engine.transaction
            if tx is not None:
                try:
                    resolved = tx.stage_for_write(resolved)
                except ValueError:
                    pass

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import coordinate_to_tuple

        # .xls/.xlsb → 透明转换为 xlsx
        from excelmanus.xls_converter import needs_conversion as _nc3, ensure_xlsx as _ensure3
        if _nc3(resolved):
            try:
                _xlsx_p3, _ = _ensure3(resolved)
                resolved = str(_xlsx_p3)
            except Exception:
                pass

        wb = load_workbook(resolved)
        sheet_names = wb.sheetnames
        ws = wb[request.sheet] if request.sheet and request.sheet in sheet_names else wb.active
        if ws is None:
            wb.close()
            return _error_json_response(404, "工作表不存在")

        cells_written = 0
        for change in request.changes:
            cell_ref = change.get("cell", "")
            value = change.get("value")
            if not cell_ref:
                continue
            row, col = coordinate_to_tuple(cell_ref.upper())
            ws.cell(row=row, column=col, value=value)
            cells_written += 1

        wb.save(resolved)
        wb.close()

        return JSONResponse(content={
            "status": "success",
            "cells_written": cells_written,
        })
    except Exception as exc:
        logger.error("Excel write 失败: %s", exc, exc_info=True)
        return _error_json_response(500, f"写入失败: {exc}")




def _safe_uploads_path(uploads_dir: "Path", relative: str) -> "Path | None":
    """在 uploads_dir 下解析 relative 路径并确保不越界。"""
    cleaned = relative.replace("\\", "/").strip("/")
    if ".." in cleaned.split("/"):
        return None
    target = (uploads_dir / cleaned).resolve()
    if not str(target).startswith(str(uploads_dir.resolve())):
        return None
    return target


_UPLOAD_MAX_PART_SIZE = 100 * 1024 * 1024  # 100 MB – 覆盖 Starlette 默认的 1 MB 限制


@_router.post("/api/v1/upload")
async def upload_file(raw_request: Request) -> JSONResponse:
    """上传文件到 workspace uploads 目录（支持可选 folder 参数指定子目录）。

    注意: 不使用 FastAPI 的 UploadFile 依赖注入，改为手动调用
    ``request.form(max_part_size=...)`` 以突破 Starlette 0.50+ 默认的
    1 MB multipart 大小限制。
    """
    assert _config is not None, "服务未初始化"

    form = await raw_request.form(max_part_size=_UPLOAD_MAX_PART_SIZE)
    file = form.get("file")
    logger.info("upload_file: form keys=%s, file type=%s, file=%r", list(form.keys()), type(file).__name__, file)
    if file is None or not hasattr(file, "read"):
        return _error_json_response(400, f"缺少 file 字段 (got {type(file).__name__})")

    filename = file.filename or "unnamed"

    from excelmanus.auth.dependencies import extract_user_id
    user_id = extract_user_id(raw_request)
    ws = _resolve_workspace(raw_request)

    content = await file.read()

    auth_enabled = getattr(raw_request.app.state, "auth_enabled", False)
    if auth_enabled and user_id:
        allowed, reason = ws.check_upload_allowed(len(content))
        if not allowed:
            return _error_json_response(413, reason)

    upload_dir = ws.get_upload_dir()

    # 支持可选的 folder= 表单字段或查询参数
    folder = raw_request.query_params.get("folder", "")
    if not folder:
        folder = str(form.get("folder", ""))

    if folder:
        target_dir = _safe_uploads_path(upload_dir, folder)
        if target_dir is None:
            return _error_json_response(400, "非法目标路径")
        target_dir.mkdir(parents=True, exist_ok=True)
    else:
        target_dir = upload_dir

    safe_name = f"{uuid.uuid4().hex[:8]}_{filename}"
    dest_path = target_dir / safe_name

    with open(dest_path, "wb") as f:
        f.write(content)

    if auth_enabled and user_id:
        ws.enforce_quota()

    # .xls/.xlsb → 自动转换为 .xlsx，转换成功后删除原始文件节省空间
    converted = False
    original_filename = filename
    _original_dest = dest_path  # 保留引用用于清理
    from excelmanus.xls_converter import needs_conversion, convert_to_xlsx, ConversionError
    if needs_conversion(dest_path):
        try:
            xlsx_path = convert_to_xlsx(dest_path, overwrite=True)
            dest_path = xlsx_path
            filename = xlsx_path.name
            converted = True
            # 清理原始 .xls/.xlsb 文件，避免双倍磁盘占用
            try:
                _original_dest.unlink(missing_ok=True)
            except OSError:
                pass
            logger.info("上传文件自动转换: %s → %s", original_filename, filename)
        except (ConversionError, Exception) as exc:
            logger.warning("上传文件转换失败，保留原始格式: %s (%s)", original_filename, exc)

    rel_path = f"./{dest_path.relative_to(ws.root_dir)}"

    # 注册到 FileRegistry
    registry = _get_file_registry(str(ws.root_dir), user_id=user_id)
    if registry is not None:
        try:
            entry = registry.register_upload(
                canonical_path=str(dest_path.relative_to(ws.root_dir)),
                original_name=original_filename,
                size_bytes=dest_path.stat().st_size,
            )
            # 转换后添加原始扩展名别名，方便用户用原名引用
            if converted:
                registry.add_alias(entry.id, "original_path", f"./{(target_dir / safe_name).relative_to(ws.root_dir)}")
        except Exception:
            logger.debug("FileRegistry register_upload 失败", exc_info=True)

    resp: dict[str, Any] = {
        "filename": original_filename,
        "path": rel_path,
        "size": dest_path.stat().st_size,
    }
    if converted:
        resp["converted_from"] = original_filename
        resp["converted_to"] = filename
    return JSONResponse(content=resp)


@_router.post("/api/v1/upload-from-url")
async def upload_file_from_url(raw_request: Request) -> JSONResponse:
    """从 URL 下载文件并保存到 workspace uploads 目录。

    请求体 JSON::

        {"url": "https://example.com/data.xlsx"}
    """
    assert _config is not None, "服务未初始化"

    try:
        body = await raw_request.json()
    except Exception:
        return _error_json_response(400, "请求体必须是 JSON")

    url: str = (body.get("url") or "").strip()
    if not url:
        return _error_json_response(400, "缺少 url 字段")

    # 仅允许 http/https
    if not url.lower().startswith(("http://", "https://")):
        return _error_json_response(400, "仅支持 http/https 链接")

    import httpx
    from urllib.parse import urlparse, unquote

    # 从 URL 路径推断文件名
    parsed = urlparse(url)
    url_path = unquote(parsed.path.rstrip("/"))
    raw_filename = url_path.split("/")[-1] if "/" in url_path else ""
    if not raw_filename or "." not in raw_filename:
        return _error_json_response(400, "无法从 URL 推断文件名（需带扩展名，如 .xlsx/.csv/.png）")

    # 下载文件（限制大小）
    max_download = _UPLOAD_MAX_PART_SIZE  # 复用上传限制
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=60.0) as client:
            resp = await client.get(url)
            resp.raise_for_status()
    except httpx.HTTPStatusError as exc:
        return _error_json_response(502, f"远程服务器返回 {exc.response.status_code}")
    except Exception as exc:
        return _error_json_response(502, f"下载失败: {exc}")

    content = resp.content
    if len(content) > max_download:
        return _error_json_response(413, f"文件过大 (>{max_download // (1024*1024)} MB)")
    if len(content) == 0:
        return _error_json_response(400, "下载到空文件")

    from excelmanus.auth.dependencies import extract_user_id
    user_id = extract_user_id(raw_request)
    ws = _resolve_workspace(raw_request)

    auth_enabled = getattr(raw_request.app.state, "auth_enabled", False)
    if auth_enabled and user_id:
        allowed, reason = ws.check_upload_allowed(len(content))
        if not allowed:
            return _error_json_response(413, reason)

    upload_dir = ws.get_upload_dir()
    safe_name = f"{uuid.uuid4().hex[:8]}_{raw_filename}"
    dest_path = upload_dir / safe_name

    with open(dest_path, "wb") as f:
        f.write(content)

    if auth_enabled and user_id:
        ws.enforce_quota()

    # .xls/.xlsb → 自动转换为 .xlsx，转换成功后删除原始文件节省空间
    converted = False
    original_filename = raw_filename
    _original_dest_url = dest_path
    from excelmanus.xls_converter import needs_conversion as _nc_url, convert_to_xlsx as _conv_url, ConversionError as _CE_url
    if _nc_url(dest_path):
        try:
            xlsx_path = _conv_url(dest_path, overwrite=True)
            dest_path = xlsx_path
            raw_filename = xlsx_path.name
            converted = True
            try:
                _original_dest_url.unlink(missing_ok=True)
            except OSError:
                pass
            logger.info("URL 上传文件自动转换: %s → %s", original_filename, raw_filename)
        except (_CE_url, Exception) as exc:
            logger.warning("URL 上传文件转换失败，保留原始格式: %s (%s)", original_filename, exc)

    rel_path = f"./{dest_path.relative_to(ws.root_dir)}"

    # 注册到 FileRegistry
    registry = _get_file_registry(str(ws.root_dir), user_id=user_id)
    if registry is not None:
        try:
            entry = registry.register_upload(
                canonical_path=str(dest_path.relative_to(ws.root_dir)),
                original_name=original_filename,
                size_bytes=dest_path.stat().st_size,
            )
            if converted:
                registry.add_alias(entry.id, "original_path", f"./{(upload_dir / safe_name).relative_to(ws.root_dir)}")
        except Exception:
            logger.debug("FileRegistry register_upload 失败 (from-url)", exc_info=True)

    resp: dict[str, Any] = {
        "filename": original_filename,
        "path": rel_path,
        "size": dest_path.stat().st_size,
    }
    if converted:
        resp["converted_from"] = original_filename
        resp["converted_to"] = raw_filename
    return JSONResponse(content=resp)


# ── 文件管理 API ─────────────────────────────────────


@_router.post("/api/v1/files/workspace/mkdir")
async def workspace_mkdir(request: Request) -> JSONResponse:
    """在 uploads/ 下创建子目录。"""
    assert _config is not None, "服务未初始化"
    body = await request.json()
    path = body.get("path", "").strip()
    if not path:
        return _error_json_response(400, "缺少 path 参数")

    ws = _resolve_workspace(request)
    uploads = ws.get_upload_dir()
    target = _safe_uploads_path(uploads, path)
    if target is None:
        return _error_json_response(400, "非法目标路径")
    if target.exists():
        return _error_json_response(409, "目录已存在")
    target.mkdir(parents=True, exist_ok=True)
    return JSONResponse(content={"status": "created", "path": path})


@_router.post("/api/v1/files/workspace/create")
async def workspace_create_file(request: Request) -> JSONResponse:
    """在 uploads/ 下创建空文件。"""
    assert _config is not None, "服务未初始化"
    body = await request.json()
    path = body.get("path", "").strip()
    if not path:
        return _error_json_response(400, "缺少 path 参数")

    ws = _resolve_workspace(request)
    uploads = ws.get_upload_dir()
    target = _safe_uploads_path(uploads, path)
    if target is None:
        return _error_json_response(400, "非法目标路径")
    if target.exists():
        return _error_json_response(409, "文件已存在")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.touch()
    return JSONResponse(content={"status": "created", "path": path})


@_router.delete("/api/v1/files/workspace/item")
async def workspace_delete_item(request: Request) -> JSONResponse:
    """删除 uploads/ 下的文件或文件夹。"""
    assert _config is not None, "服务未初始化"
    body = await request.json()
    path = body.get("path", "").strip()
    if not path:
        return _error_json_response(400, "缺少 path 参数")

    ws = _resolve_workspace(request)
    uploads = ws.get_upload_dir()
    target = _safe_uploads_path(uploads, path)
    if target is None:
        return _error_json_response(400, "非法目标路径")
    if not target.exists():
        return _error_json_response(404, "路径不存在")
    # 防止删除上传根目录本身
    if target.resolve() == uploads.resolve():
        return _error_json_response(400, "无法删除根目录")

    import shutil
    if target.is_dir():
        shutil.rmtree(target)
    else:
        target.unlink()
    # W4: 通知所有活跃 session 清理关联 staging 条目
    if _session_manager is not None:
        _session_manager.notify_file_deleted(str(target))
    return JSONResponse(content={"status": "deleted", "path": path})


@_router.post("/api/v1/files/workspace/rename")
async def workspace_rename_item(request: Request) -> JSONResponse:
    """重命名 uploads/ 下的文件或文件夹。"""
    assert _config is not None, "服务未初始化"
    body = await request.json()
    old_path = body.get("old_path", "").strip()
    new_path = body.get("new_path", "").strip()
    if not old_path or not new_path:
        return _error_json_response(400, "缺少 old_path 或 new_path 参数")

    ws = _resolve_workspace(request)
    uploads = ws.get_upload_dir()
    src = _safe_uploads_path(uploads, old_path)
    dst = _safe_uploads_path(uploads, new_path)
    if src is None or dst is None:
        return _error_json_response(400, "非法路径")
    if not src.exists():
        return _error_json_response(404, "源路径不存在")
    if dst.exists():
        return _error_json_response(409, "目标路径已存在")
    dst.parent.mkdir(parents=True, exist_ok=True)
    src.rename(dst)
    # W5: 通知所有活跃 session 更新 staging 映射
    if _session_manager is not None:
        _session_manager.notify_file_renamed(str(src), str(dst))
    return JSONResponse(content={"status": "renamed", "old_path": old_path, "new_path": new_path})


@_router.post("/api/v1/files/reveal")
async def reveal_file(request: Request) -> JSONResponse:
    """在本地文件管理器中打开文件所在目录。"""
    # 服务器/Docker 模式下此功能无意义（打开的是服务器端文件管理器，用户不可见）
    if _config is not None and _config.is_server:
        return _error_json_response(
            400,
            "此功能仅在本地部署模式下可用。服务器模式下无法打开本地文件管理器。",
        )

    import platform
    import subprocess

    body = await request.json()
    file_path = body.get("path", "").strip()
    if not file_path:
        return _error_json_response(400, "缺少 path 参数")

    target = os.path.abspath(file_path)
    # 安全校验：限制在工作区范围内，防止路径遍历
    if _config is not None:
        ws_root = os.path.abspath(_config.workspace_root)
        if not (target == ws_root or target.startswith(ws_root + os.sep)):
            return _error_json_response(403, "路径不在工作区范围内")
    if not os.path.exists(target):
        return _error_json_response(404, f"路径不存在: {target}")

    try:
        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", "-R", target])
        elif system == "Windows":
            subprocess.Popen(["explorer", "/select,", target])
        else:
            parent = os.path.dirname(target) if os.path.isfile(target) else target
            subprocess.Popen(["xdg-open", parent])
    except Exception as exc:
        logger.warning("打开文件管理器失败: %s", exc)
        return _error_json_response(500, f"打开失败: {exc}")

    return JSONResponse(content={"status": "ok", "path": target})


@_router.get("/api/v1/sessions")
async def list_sessions(request: Request) -> JSONResponse:
    """列出所有会话（含历史）。认证启用时仅返回当前用户的会话。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    include_archived = request.query_params.get("include_archived", "false").lower() == "true"
    sessions = await _session_manager.list_sessions(
        include_archived=include_archived, user_id=user_id
    )
    return JSONResponse(content={"sessions": sessions})


@_router.delete("/api/v1/sessions", responses={409: _error_responses[409]})
async def clear_all_sessions(request: Request) -> JSONResponse:
    """清空当前用户的会话历史。认证启用时仅删除当前用户的会话。若有会话正在处理中则返回 409。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    sess_count, msg_count = await _session_manager.clear_all_sessions(user_id=user_id)
    return JSONResponse(content={
        "status": "ok",
        "sessions_deleted": sess_count,
        "messages_deleted": msg_count,
    })


@_router.get("/api/v1/sessions/{session_id}/messages")
async def get_session_messages(session_id: str, request: Request) -> JSONResponse:
    """分页获取会话消息。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    try:
        limit = max(1, min(500, int(request.query_params.get("limit", "50"))))
        offset = max(0, int(request.query_params.get("offset", "0")))
    except (ValueError, TypeError):
        return JSONResponse(status_code=400, content={"detail": "limit/offset 必须为整数"})
    messages = await _session_manager.get_session_messages(
        session_id, limit=limit, offset=offset, user_id=user_id
    )
    return JSONResponse(content={"messages": messages, "session_id": session_id})


@_router.get("/api/v1/sessions/{session_id}/excel-events")
async def get_session_excel_events(session_id: str, request: Request) -> JSONResponse:
    """返回持久化的 Excel diff 和改动文件列表，供前端重启后恢复。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    ch = _session_manager.chat_history
    if ch is None:
        return JSONResponse(content={"diffs": [], "affected_files": [], "previews": []})
    if user_id is not None and not ch.session_owned_by(session_id, user_id):
        return JSONResponse(content={"diffs": [], "affected_files": [], "previews": []})
    diffs = ch.load_excel_diffs(session_id)
    affected_files = ch.load_affected_files(session_id)
    previews = ch.load_excel_previews(session_id)
    safe_mode = _is_external_safe_mode()
    safe_diffs = []
    for d in diffs:
        safe_diffs.append({
            "tool_call_id": d["tool_call_id"],
            "file_path": _public_excel_path(d["file_path"], safe_mode=safe_mode),
            "sheet": d["sheet"],
            "affected_range": d["affected_range"],
            "changes": d["changes"],
            "timestamp": d["timestamp"],
        })
    safe_previews = []
    for p in previews:
        safe_previews.append({
            "tool_call_id": p["tool_call_id"],
            "file_path": _public_excel_path(p["file_path"], safe_mode=safe_mode),
            "sheet": p["sheet"],
            "columns": p["columns"],
            "rows": p["rows"],
            "total_rows": p["total_rows"],
            "truncated": p["truncated"],
        })
    safe_files = [
        _public_excel_path(f, safe_mode=safe_mode)
        for f in affected_files if f
    ]
    return JSONResponse(content={
        "diffs": safe_diffs,
        "previews": safe_previews,
        "affected_files": safe_files,
    })


@_router.get("/api/v1/sessions/{session_id}/export")
async def export_session(session_id: str, request: Request) -> Response:
    """导出会话为 Markdown / 纯文本 / EMX 格式。

    Query params:
        format: md | txt | emx (默认 md)
        include_workspace: true | false (默认 true，仅 emx 有效)
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    fmt = (request.query_params.get("format") or "md").lower().strip()
    if fmt not in ("md", "txt", "emx"):
        return JSONResponse(status_code=400, content={"detail": f"不支持的格式: {fmt}，可选: md, txt, emx"})

    from excelmanus.session_export import export_markdown, export_text

    # ── EMX: 完整导出（v2.0），委托 SessionManager ──
    if fmt == "emx":
        include_ws = (request.query_params.get("include_workspace") or "true").lower() != "false"
        try:
            data = await _session_manager.export_full_session(
                session_id, user_id=user_id, include_workspace=include_ws,
            )
        except Exception as exc:
            if "不存在" in str(exc):
                return _error_json_response(404, str(exc))
            logger.warning("EMX 导出失败", exc_info=True)
            return _error_json_response(500, "导出失败")

        raw_title = data.get("session", {}).get("title", "session") or "session"
        from urllib.parse import quote
        ascii_title = "".join(c for c in raw_title if c.isascii() and (c.isalnum() or c in " _-")).strip()[:50] or "session"
        utf8_title = "".join(c for c in raw_title if c.isalnum() or c in " _-").strip()[:50] or "session"
        cd = f"attachment; filename=\"{ascii_title}.emx\"; filename*=UTF-8''{quote(utf8_title)}.emx"
        return Response(
            content=json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": cd},
        )

    # ── MD / TXT: 轻量导出（与 v1 行为一致）──
    messages = await _session_manager.get_session_messages(
        session_id, limit=100000, offset=0, user_id=user_id,
    )
    if not messages:
        return _error_json_response(404, f"会话 '{session_id}' 不存在或无消息")

    ch = _session_manager.chat_history
    session_meta: dict[str, Any] = {"id": session_id, "title": "未命名会话", "created_at": "", "updated_at": ""}
    if ch is not None:
        meta = ch.get_session_meta(session_id)
        if meta:
            session_meta.update(meta)

    excel_diffs: list[dict] = []
    excel_previews: list[dict] = []
    affected_files: list[str] = []
    if ch is not None and fmt == "md":
        excel_diffs = ch.load_excel_diffs(session_id)
        excel_previews = ch.load_excel_previews(session_id)
        affected_files = ch.load_affected_files(session_id)

    raw_title = session_meta.get("title", "session") or "session"
    ascii_title = "".join(c for c in raw_title if c.isascii() and (c.isalnum() or c in " _-")).strip()[:50] or "session"
    from urllib.parse import quote
    utf8_title = "".join(c for c in raw_title if c.isalnum() or c in " _-").strip()[:50] or "session"

    def _cd(ext: str) -> str:
        return (
            f"attachment; filename=\"{ascii_title}.{ext}\"; "
            f"filename*=UTF-8''{quote(utf8_title)}.{ext}"
        )

    if fmt == "md":
        content = export_markdown(session_meta, messages, excel_diffs, excel_previews, affected_files)
        return Response(
            content=content.encode("utf-8"),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": _cd("md")},
        )
    else:  # txt
        content = export_text(session_meta, messages)
        return Response(
            content=content.encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": _cd("txt")},
        )


@_router.post("/api/v1/sessions/import")
async def import_session(request: Request) -> JSONResponse:
    """从 EMX (.emx) 文件导入会话（v2.0 完整恢复）。

    接收 JSON body（EMX 格式），创建新会话并恢复所有状态：
    消息、SessionState checkpoint、持久记忆、工作区文件。
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.session_export import parse_emx, EMXImportError

    user_id = _get_isolation_user_id(request)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"detail": "无法解析 JSON body"})

    try:
        parsed = parse_emx(body)
    except EMXImportError as exc:
        return JSONResponse(status_code=400, content={"detail": str(exc)})

    try:
        result = await _session_manager.import_full_session(parsed, user_id=user_id)
    except RuntimeError as exc:
        return _error_json_response(503, str(exc))
    except Exception:
        logger.warning("导入会话失败", exc_info=True)
        return _error_json_response(500, "导入失败")

    return JSONResponse(content={"status": "ok", **result})


@_router.get("/api/v1/sessions/{session_id}/status")
async def get_session_status(session_id: str, request: Request) -> JSONResponse:
    """获取会话运行时状态（上下文压缩 + 文件注册表）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    def _normalize_registry_status(registry_payload: dict[str, Any]) -> dict[str, Any]:
        normalized = dict(registry_payload)
        state = str(normalized.get("state") or "idle").lower()
        if state == "ready":
            state = "built"
        if state not in {"idle", "building", "built", "error"}:
            state = "idle"
        normalized["state"] = state

        # 兼容前端旧字段约定：将 total_files 回填到 sheet_count。
        total_files = normalized.get("total_files")
        if normalized.get("sheet_count") is None and total_files is not None:
            try:
                normalized["sheet_count"] = int(total_files)
            except (TypeError, ValueError):
                pass
        return normalized

    # 性能优化：仅查询内存中已有的引擎，不触发引擎创建/恢复。
    # 如果会话被 TTL 清理出内存，返回 idle 默认值即可——
    # 引擎会在用户真正发消息时才创建，避免轮询导致不必要的重量级初始化。
    engine = _session_manager.get_engine(session_id, user_id=user_id)

    if engine is None:
        _idle = _normalize_registry_status({"state": "idle"})
        return JSONResponse(content={
            "session_id": session_id,
            "compaction": {"enabled": False},
            "registry": _idle,
        })

    # 上下文压缩状态
    compaction: dict[str, Any] = {"enabled": False}
    try:
        compaction = engine.get_compaction_status()
    except Exception:
        pass

    # 文件注册表扫描状态
    registry: dict[str, Any] = {"state": "idle"}
    try:
        registry = engine.registry_scan_status()
    except Exception:
        pass
    registry = _normalize_registry_status(registry)

    return JSONResponse(content={
        "session_id": session_id,
        "compaction": compaction,
        "registry": registry,
    })


@_router.post("/api/v1/sessions/{session_id}/compact")
async def compact_session_context(session_id: str, request: Request) -> JSONResponse:
    """在指定会话内执行 /compact，并返回执行结果。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    engine = await _session_manager.get_or_restore_engine(
        session_id, user_id=user_id
    )
    if engine is None:
        return _error_json_response(404, f"会话不存在: {session_id}")

    try:
        result = await engine._command_handler.handle("/compact")
    except Exception as exc:
        logger.warning("会话 %s 执行 /compact 失败: %s", session_id, exc)
        return _error_json_response(500, f"压缩执行失败: {exc}")

    if not result:
        result = "压缩命令已执行，但未返回可展示结果。"

    return JSONResponse(content={
        "session_id": session_id,
        "result": result,
    })


@_router.post("/api/v1/sessions/{session_id}/memory/extract")
async def extract_session_memory(session_id: str, request: Request) -> JSONResponse:
    """手动触发指定会话的记忆提取。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    engine = await _session_manager.get_or_restore_engine(
        session_id, user_id=user_id
    )
    if engine is None:
        return _error_json_response(404, f"会话不存在: {session_id}")

    try:
        entries = await engine.extract_and_save_memory(trigger="manual")
    except Exception as exc:
        logger.warning("会话 %s 手动记忆提取失败: %s", session_id, exc)
        return _error_json_response(500, f"记忆提取失败: {exc}")

    return JSONResponse(content={
        "session_id": session_id,
        "count": len(entries),
        "entries": [
            {"id": e.id, "content": e.content, "category": e.category.value}
            for e in entries
        ],
    })


@_router.post("/api/v1/sessions/{session_id}/registry/scan")
async def scan_session_registry(session_id: str, request: Request) -> JSONResponse:
    """触发指定会话的 FileRegistry 后台扫描（force=True）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    engine = await _session_manager.get_or_restore_engine(
        session_id, user_id=user_id
    )
    if engine is None:
        return _error_json_response(404, f"会话不存在: {session_id}")

    try:
        started = engine.start_registry_scan(force=True)
    except Exception as exc:
        logger.warning("会话 %s 触发 registry scan 失败: %s", session_id, exc)
        return _error_json_response(500, f"文件扫描失败: {exc}")

    return JSONResponse(content={
        "session_id": session_id,
        "started": started,
        "message": "文件扫描已启动" if started else "扫描正在进行中，请稍候",
    })


@_router.post("/api/v1/sessions/{session_id}/full-access")
async def toggle_full_access(session_id: str, request: Request) -> JSONResponse:
    """切换指定会话的 full_access 开关（供前端快捷按钮使用）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)

    body = await request.json()
    enabled = bool(body.get("enabled", True))

    engine = await _session_manager.get_or_restore_engine(
        session_id, user_id=user_id
    )
    if engine is not None:
        engine._full_access_enabled = enabled
        engine._persist_full_access(enabled)
        if not enabled:
            # 关闭时驱逐受限 skill，与 command_handler 保持一致
            blocked = set(engine._restricted_code_skillpacks)
            engine._active_skills = [
                s for s in engine._active_skills if s.name not in blocked
            ]
    else:
        # 会话尚未创建（local-first），仅持久化到 UserConfigStore
        if _database is not None:
            try:
                from excelmanus.stores.config_store import UserConfigStore
                uc = UserConfigStore(_database.conn, user_id=user_id)
                uc.set_full_access(enabled)
            except Exception:
                logger.debug("持久化 full_access 失败（无会话）", exc_info=True)

    return JSONResponse(content={
        "session_id": session_id,
        "full_access_enabled": enabled,
    })


@_router.get("/api/v1/sessions/{session_id}")
async def get_session(session_id: str, request: Request) -> JSONResponse:
    """获取会话详情含消息历史。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(request)
    try:
        detail = await _session_manager.get_session_detail(session_id, user_id=user_id)
    except SessionNotFoundError:
        # 会话可能尚未在后端创建（前端 local-first 乐观创建），返回空默认值。
        return JSONResponse(content={
            "id": session_id,
            "message_count": 0,
            "in_flight": False,
            "messages": [],
            "full_access_enabled": False,
            "chat_mode": "write",
            "current_model": None,
            "current_model_name": None,
            "vision_capable": False,
            "pending_approval": None,
            "pending_question": None,
            "last_route": None,
        })
    return JSONResponse(content=detail)


class ModelSwitchRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str


@_router.get("/api/v1/models")
async def list_models(request: Request) -> JSONResponse:
    """获取可用模型列表（含多模型配置档案）。

    注意：模型列表为只读信息，不需要管理员权限，所有已认证用户均可访问。
    """
    assert _config is not None, "服务未初始化"

    user_id = _get_isolation_user_id(request)
    active_name: str | None = None
    if _config_store is not None:
        try:
            from excelmanus.stores.config_store import UserConfigStore
            _user_cfg = UserConfigStore(
                _config_store._conn if hasattr(_config_store, '_conn') else _config_store._conn,
                user_id=user_id,
            )
            active_name = _user_cfg.get_active_model()
        except Exception:
            active_name = _config_store.get_active_model() if hasattr(_config_store, 'get_active_model') else None
    models: list[dict] = [{
        "name": "default",
        "model": _config.model,
        "display_name": _config.model,
        "description": "默认模型（主配置）",
        "active": active_name is None,
        "base_url": _config.base_url,
    }]
    db_profiles = _config_store.list_profiles() if _config_store else []
    for p in db_profiles:
        models.append({
            "name": p["name"],
            "model": p["model"],
            "display_name": p.get("name", ""),
            "description": p.get("description", ""),
            "active": p["name"] == active_name,
            "base_url": p.get("base_url", ""),
        })

    # 按用户 allowed_models 过滤（空列表 = 不限制）
    allowed = _get_user_allowed_models(request)
    if allowed:
        models = [m for m in models if m["name"] == "default" or m["name"] in allowed]

    return JSONResponse(content={"models": models})


@_router.put("/api/v1/models/active")
async def switch_model(request: ModelSwitchRequest, raw_request: Request) -> JSONResponse:
    """切换当前活跃模型并持久化到数据库，同时同步所有活跃会话的 engine。

    所有已认证用户均可切换模型（不再要求管理员权限）。
    如果用户配置了 allowed_models，则只能切换到允许的模型。
    """
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    assert _config is not None, "服务未初始化"

    name = request.name.strip()
    if not name:
        return _error_json_response(400, "请指定模型名称。")

    user_id = _get_isolation_user_id(raw_request)

    # 验证目标模型存在（统一走 DB profile 查找）
    if name.lower() != "default":
        profile = _config_store.get_profile(name) if _config_store else None
        if profile is None:
            available_names = _list_available_model_names_for_user(raw_request, user_id)
            return _error_json_response(
                404,
                f"未找到模型 {name!r}。可用模型：{', '.join(available_names)}",
            )

    # 校验用户模型权限（allowed_models 为空表示不限制）
    allowed = _get_user_allowed_models(raw_request)
    if allowed and name.lower() != "default" and name not in allowed:
        return _error_json_response(403, f"您没有使用模型 {name!r} 的权限。")

    # 持久化到用户级配置（auth 启用时隔离，匿名时写全局）
    if _config_store is not None:
        try:
            from excelmanus.stores.config_store import UserConfigStore
            _user_cfg = UserConfigStore(
                _config_store._conn if hasattr(_config_store, '_conn') else _config_store._conn,
                user_id=user_id,
            )
            _user_cfg.set_active_model(None if name.lower() == "default" else name)
        except Exception:
            _config_store.set_active_model(None if name.lower() == "default" else name)

    # 仅同步当前用户的活跃会话 engine
    result_msg = f"模型已切换为 {name}"
    sessions = await _session_manager.list_sessions(user_id=user_id)
    for session_info in sessions:
        try:
            engine = _session_manager.get_engine(session_info["id"], user_id=user_id)
            if engine is not None:
                result_msg = engine.switch_model(name)
        except Exception:
            pass

    # 模型切换后尝试从缓存加载新模型的能力探测结果
    db = _session_manager.database if _session_manager else None
    if db is not None:
        try:
            from excelmanus.model_probe import load_capabilities

            for session_info in sessions:
                engine = _session_manager.get_engine(session_info["id"], user_id=user_id)
                if engine is not None:
                    caps = load_capabilities(
                        db, engine.current_model, engine.active_base_url,
                    )
                    if caps is not None:
                        engine.set_model_capabilities(caps)
        except Exception:
            logger.debug("模型切换后加载能力缓存失败", exc_info=True)

    return JSONResponse(content={"message": result_msg})


# ── Thinking 配置 API ──────────────────────────────────


class ThinkingConfigRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    effort: str | None = None  # none|minimal|low|medium|high|xhigh
    budget: int | None = None  # 精确 token 预算（0 = 使用 effort 换算）


@_router.get("/api/v1/thinking")
async def get_thinking_config(raw_request: Request) -> JSONResponse:
    """获取当前 thinking 配置（等级 + 预算）。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    user_id = _get_isolation_user_id(raw_request)
    sessions = await _session_manager.list_sessions(user_id=user_id)
    # 取第一个活跃 session 的 thinking_config
    for s in sessions:
        engine = _session_manager.get_engine(s["id"], user_id=user_id)
        if engine is not None:
            tc = engine.thinking_config
            return JSONResponse(content={
                "effort": tc.effort,
                "budget": tc.budget_tokens,
                "effective_budget": tc.effective_budget(),
            })
    # 回退到全局配置
    assert _config is not None
    return JSONResponse(content={
        "effort": _config.thinking_effort,
        "budget": _config.thinking_budget,
        "effective_budget": 0,
    })


@_router.put("/api/v1/thinking")
async def set_thinking_config(request: ThinkingConfigRequest, raw_request: Request) -> JSONResponse:
    """设置 thinking 等级和/或预算，同步到所有活跃会话。"""
    if _session_manager is None:
        return _error_json_response(503, "服务未初始化")
    from excelmanus.engine import _EFFORT_RATIOS
    if request.effort is not None and request.effort not in _EFFORT_RATIOS:
        return _error_json_response(400, f"无效的 effort 值: {request.effort!r}。可选: {', '.join(sorted(_EFFORT_RATIOS))}")

    user_id = _get_isolation_user_id(raw_request)
    sessions = await _session_manager.list_sessions(user_id=user_id)
    updated = 0
    result_tc = None
    for s in sessions:
        engine = _session_manager.get_engine(s["id"], user_id=user_id)
        if engine is not None:
            engine.set_thinking_config(effort=request.effort, budget=request.budget)
            result_tc = engine.thinking_config
            updated += 1

    if result_tc is None:
        return _error_json_response(404, "无活跃会话。")

    return JSONResponse(content={
        "effort": result_tc.effort,
        "budget": result_tc.budget_tokens,
        "effective_budget": result_tc.effective_budget(),
        "sessions_updated": updated,
    })


# ── 模型配置管理 API（.env 持久化） ──────────────────────

_MODEL_ENV_KEYS = {
    "main": {"api_key": "EXCELMANUS_API_KEY", "base_url": "EXCELMANUS_BASE_URL", "model": "EXCELMANUS_MODEL", "protocol": "EXCELMANUS_PROTOCOL"},
    "aux": {"api_key": "EXCELMANUS_AUX_API_KEY", "base_url": "EXCELMANUS_AUX_BASE_URL", "model": "EXCELMANUS_AUX_MODEL", "enabled": "EXCELMANUS_AUX_ENABLED", "protocol": "EXCELMANUS_AUX_PROTOCOL"},
    "vlm": {"api_key": "EXCELMANUS_VLM_API_KEY", "base_url": "EXCELMANUS_VLM_BASE_URL", "model": "EXCELMANUS_VLM_MODEL", "enabled": "EXCELMANUS_VLM_ENABLED", "protocol": "EXCELMANUS_VLM_PROTOCOL"},
    "embedding": {"api_key": "EXCELMANUS_EMBEDDING_API_KEY", "base_url": "EXCELMANUS_EMBEDDING_BASE_URL", "model": "EXCELMANUS_EMBEDDING_MODEL", "enabled": "EXCELMANUS_EMBEDDING_ENABLED"},
}


def _find_env_file() -> str:
    """定位 .env 文件路径。"""
    candidates = [
        os.path.join(os.getcwd(), ".env"),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return candidates[0]


def _read_env_file(path: str) -> list[str]:
    """读取 .env 文件所有行。"""
    if not os.path.isfile(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return f.readlines()


def _write_env_file(path: str, lines: list[str]) -> None:
    """写回 .env 文件。"""
    with open(path, "w", encoding="utf-8") as f:
        f.writelines(lines)


def _update_env_var(lines: list[str], key: str, value: str) -> list[str]:
    """更新或追加环境变量行，保持注释和格式。"""
    new_lines = []
    found = False
    for line in lines:
        stripped = line.strip()
        # 匹配 KEY=... 或 # KEY=...
        if stripped.startswith(f"{key}=") or stripped.startswith(f"# {key}="):
            if value:
                new_lines.append(f"{key}={value}\n")
            else:
                new_lines.append(f"# {key}=\n")
            found = True
        else:
            new_lines.append(line)
    if not found and value:
        new_lines.append(f"{key}={value}\n")
    return new_lines


class ModelConfigUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    api_key: str | None = None
    base_url: str | None = None
    model: str | None = None
    enabled: bool | None = None
    protocol: str | None = None


class ModelProfileCreate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str
    model: str
    api_key: str = ""
    base_url: str = ""
    description: str = ""
    protocol: str = "auto"
    thinking_mode: str = "auto"
    model_family: str = ""
    custom_extra_body: str = ""
    custom_extra_headers: str = ""


@_router.get("/api/v1/config/models")
async def get_model_config(request: Request) -> JSONResponse:
    """获取全部模型配置（main/aux/vlm + profiles）。"""
    assert _config is not None, "服务未初始化"
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    result: dict = {
        "main": {
            "api_key": _mask_key(_config.api_key),
            "base_url": _config.base_url,
            "model": _config.model,
            "protocol": _config.protocol,
        },
        "aux": {
            "api_key": _mask_key(_config.aux_api_key or ""),
            "base_url": _config.aux_base_url or "",
            "model": _config.aux_model or "",
            "enabled": _config.aux_enabled,
            "protocol": _config.aux_protocol,
        },
        "vlm": {
            "api_key": _mask_key(_config.vlm_api_key or ""),
            "base_url": _config.vlm_base_url or "",
            "model": _config.vlm_model or "",
            "enabled": _config.vlm_enabled,
            "protocol": _config.vlm_protocol,
        },
        "embedding": {
            "api_key": _mask_key(_config.embedding_api_key or ""),
            "base_url": _config.embedding_base_url or "",
            "model": _config.embedding_model or "",
            "enabled": _config.embedding_enabled,
        },
        "profiles": [
            {
                "name": p["name"],
                "model": p["model"],
                "api_key": _mask_key(p.get("api_key", "")),
                "base_url": p.get("base_url", ""),
                "description": p.get("description", ""),
                "protocol": p.get("protocol", "auto"),
                "thinking_mode": p.get("thinking_mode", "auto"),
                "model_family": p.get("model_family", ""),
                "custom_extra_body": p.get("custom_extra_body", ""),
                "custom_extra_headers": p.get("custom_extra_headers", ""),
            }
            for p in (_config_store.list_profiles() if _config_store else [])
        ],
    }
    return JSONResponse(content=result)


def _mask_key(key: str) -> str:
    """脱敏 API Key：保留前4后4位。"""
    if not key or len(key) <= 12:
        return "****" if key else ""
    return f"{key[:4]}{'*' * (len(key) - 8)}{key[-4:]}"


def _sync_config_profiles_from_db() -> None:
    """从数据库读取 model_profiles 并同步到 _config.models。"""
    if _config is None or _config_store is None:
        return
    rows = _config_store.list_profiles()
    default_api_key = _config.api_key
    default_base_url = _config.base_url
    profiles: list[ModelProfile] = []
    for row in rows:
        profiles.append(ModelProfile(
            name=row["name"],
            model=row["model"],
            api_key=row.get("api_key") or default_api_key,
            base_url=row.get("base_url") or default_base_url,
            description=row.get("description", ""),
            protocol=row.get("protocol", "auto"),
            thinking_mode=row.get("thinking_mode", "auto"),
            model_family=row.get("model_family", ""),
            custom_extra_body=row.get("custom_extra_body", ""),
            custom_extra_headers=row.get("custom_extra_headers", ""),
        ))
    object.__setattr__(_config, "models", tuple(profiles))


@_router.put("/api/v1/config/models/{section}")
async def update_model_config(
    section: str,
    request: ModelConfigUpdate,
    raw_request: Request,
) -> JSONResponse:
    """更新指定模型配置区块并持久化到 .env。"""
    guard_error = await _require_admin_if_auth_enabled(raw_request)
    if guard_error is not None:
        return guard_error
    global _config_incomplete
    if section not in _MODEL_ENV_KEYS:
        return _error_json_response(400, f"未知配置区块: {section}")

    env_path = _find_env_file()
    lines = _read_env_file(env_path)
    key_map = _MODEL_ENV_KEYS[section]

    updates: dict[str, str] = {}
    if request.api_key is not None and "api_key" in key_map:
        updates[key_map["api_key"]] = request.api_key
    if request.base_url is not None and "base_url" in key_map:
        updates[key_map["base_url"]] = request.base_url
    if request.model is not None and "model" in key_map:
        updates[key_map["model"]] = request.model
    if request.enabled is not None and "enabled" in key_map:
        updates[key_map["enabled"]] = "true" if request.enabled else "false"
    if request.protocol is not None and "protocol" in key_map:
        updates[key_map["protocol"]] = request.protocol

    if not updates:
        return _error_json_response(400, "无有效更新字段")

    for env_key, env_val in updates.items():
        lines = _update_env_var(lines, env_key, env_val)

    _write_env_file(env_path, lines)

    # 同步更新环境变量使运行时生效
    for env_key, env_val in updates.items():
        if env_val:
            os.environ[env_key] = env_val
        elif env_key in os.environ:
            del os.environ[env_key]

    # 同步更新内存中的 _config 实例
    if _config is not None:
        _SECTION_CONFIG_FIELDS = {
            "main": {"api_key": "api_key", "base_url": "base_url", "model": "model", "protocol": "protocol"},
            "aux": {"api_key": "aux_api_key", "base_url": "aux_base_url", "model": "aux_model", "enabled": "aux_enabled", "protocol": "aux_protocol"},
            "vlm": {"api_key": "vlm_api_key", "base_url": "vlm_base_url", "model": "vlm_model", "enabled": "vlm_enabled", "protocol": "vlm_protocol"},
            "embedding": {"api_key": "embedding_api_key", "base_url": "embedding_base_url", "model": "embedding_model", "enabled": "embedding_enabled"},
        }
        field_map = _SECTION_CONFIG_FIELDS.get(section, {})
        for req_field, config_attr in field_map.items():
            val = getattr(request, req_field, None)
            if val is not None:
                object.__setattr__(_config, config_attr, val)

    # 主模型配置更新后，检查是否已补齐必填项 → 解除降级模式
    if section == "main" and _config is not None and _config_incomplete:
        if _config.api_key and _config.base_url and _config.model:
            _config_incomplete = False
            logger.info("模型配置已补齐，服务已从降级模式恢复为正常模式。")

    # AUX 变更需广播到所有活跃 engine，避免子代理/路由/顾问使用过时快照
    if section == "aux" and _session_manager is not None:
        await _session_manager.broadcast_aux_config(
            aux_enabled=_config.aux_enabled if _config else True,
            aux_model=_config.aux_model if _config else None,
            aux_api_key=_config.aux_api_key if _config else None,
            aux_base_url=_config.aux_base_url if _config else None,
        )

    return JSONResponse(content={"status": "ok", "section": section, "updated": list(updates.keys())})


@_router.post("/api/v1/config/models/profiles")
async def add_model_profile(request: ModelProfileCreate, raw_request: Request) -> JSONResponse:
    """新增多模型条目并持久化到数据库。"""
    guard_error = await _require_admin_if_auth_enabled(raw_request)
    if guard_error is not None:
        return guard_error
    if _config_store is None:
        return _error_json_response(503, "配置存储未初始化")

    if _config_store.get_profile(request.name):
        return _error_json_response(409, f"模型名称已存在: {request.name}")

    _config_store.add_profile(
        name=request.name,
        model=request.model,
        api_key=request.api_key or "",
        base_url=request.base_url or "",
        description=request.description or "",
        protocol=request.protocol or "auto",
        thinking_mode=request.thinking_mode or "auto",
        model_family=request.model_family or "",
        custom_extra_body=request.custom_extra_body or "",
        custom_extra_headers=request.custom_extra_headers or "",
    )
    _sync_config_profiles_from_db()
    if _session_manager is not None and _config is not None:
        await _session_manager.broadcast_model_profiles(_config.models)

    return JSONResponse(status_code=201, content={"status": "created", "name": request.name})


@_router.delete("/api/v1/config/models/profiles/{name:path}")
async def delete_model_profile(name: str, request: Request) -> JSONResponse:
    """删除多模型条目。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config_store is None:
        return _error_json_response(503, "配置存储未初始化")

    if not _config_store.delete_profile(name):
        return _error_json_response(404, f"未找到模型: {name}")

    _sync_config_profiles_from_db()
    if _session_manager is not None and _config is not None:
        await _session_manager.broadcast_model_profiles(_config.models)
    return JSONResponse(content={"status": "deleted", "name": name})


@_router.put("/api/v1/config/models/profiles/{name:path}")
async def update_model_profile(
    name: str,
    request: ModelProfileCreate,
    raw_request: Request,
) -> JSONResponse:
    """更新多模型条目。"""
    guard_error = await _require_admin_if_auth_enabled(raw_request)
    if guard_error is not None:
        return guard_error
    if _config_store is None:
        return _error_json_response(503, "配置存储未初始化")

    if not _config_store.get_profile(name):
        return _error_json_response(404, f"未找到模型: {name}")

    _config_store.update_profile(
        name,
        new_name=request.name if request.name != name else None,
        model=request.model,
        api_key=request.api_key or None,
        base_url=request.base_url or None,
        description=request.description or None,
        protocol=request.protocol or None,
        thinking_mode=request.thinking_mode,
        model_family=request.model_family,
        custom_extra_body=request.custom_extra_body,
        custom_extra_headers=request.custom_extra_headers,
    )
    _sync_config_profiles_from_db()
    if _session_manager is not None and _config is not None:
        await _session_manager.broadcast_model_profiles(_config.models)

    return JSONResponse(content={"status": "updated", "name": request.name})


# ── 模型配置导出/导入 API ──────────────────────────────


class ConfigExportRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    sections: list[str] = ["main", "aux", "vlm", "profiles"]
    mode: Literal["password", "simple"] = "password"
    password: str | None = None


class ConfigImportRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    token: str
    password: str | None = None


def _collect_raw_sections(section_names: list[str]) -> dict[str, Any]:
    """收集指定区块的原始（未脱敏）配置数据。"""
    assert _config is not None
    result: dict[str, Any] = {}
    if "main" in section_names:
        result["main"] = {
            "api_key": _config.api_key,
            "base_url": _config.base_url,
            "model": _config.model,
            "protocol": _config.protocol,
        }
    if "aux" in section_names:
        result["aux"] = {
            "api_key": _config.aux_api_key or "",
            "base_url": _config.aux_base_url or "",
            "model": _config.aux_model or "",
            "protocol": _config.aux_protocol,
        }
    if "vlm" in section_names:
        result["vlm"] = {
            "api_key": _config.vlm_api_key or "",
            "base_url": _config.vlm_base_url or "",
            "model": _config.vlm_model or "",
            "protocol": _config.vlm_protocol,
        }
    if "embedding" in section_names:
        result["embedding"] = {
            "api_key": _config.embedding_api_key or "",
            "base_url": _config.embedding_base_url or "",
            "model": _config.embedding_model or "",
            "enabled": _config.embedding_enabled,
        }
    if "profiles" in section_names and _config_store is not None:
        result["profiles"] = _config_store.list_profiles()
    return result


def _collect_user_section(request: Request, user_id: str) -> dict[str, Any]:
    """收集当前用户个人的 LLM 配置（users 表中的 llm_* 覆盖）。"""
    store = getattr(request.app.state, "user_store", None)
    if store is None:
        return {"api_key": "", "base_url": "", "model": ""}
    user = store.get_by_id(user_id)
    if user is None:
        return {"api_key": "", "base_url": "", "model": ""}
    return {
        "api_key": user.llm_api_key or "",
        "base_url": user.llm_base_url or "",
        "model": user.llm_model or "",
    }


@_router.get("/api/v1/config/models/user")
async def get_user_model_config(request: Request) -> JSONResponse:
    """获取当前用户的自定义 LLM 配置（api_key 脱敏返回）。"""
    user_id = getattr(request.state, "user_id", None)
    if not user_id:
        return JSONResponse(content={"api_key": "", "base_url": "", "model": ""})
    data = _collect_user_section(request, user_id)
    # 脱敏 API key
    if data.get("api_key"):
        data["api_key"] = _mask_key(data["api_key"])
    return JSONResponse(content=data)


async def _get_config_transfer_scope(request: Request) -> tuple[str | None, bool] | JSONResponse:
    """返回配置导出/导入的权限范围。成功返回 (user_id, is_admin_scope)，失败返回错误响应。

    - 认证关闭: (None, True) 视为管理员，可操作全局配置
    - 认证开启: 须登录（AuthMiddleware 已注入 user_id），管理员返回 (user_id, True)，普通用户返回 (user_id, False)
    """
    if not getattr(request.app.state, "auth_enabled", False):
        return (None, True)
    user_id = getattr(request.state, "user_id", None)
    if not user_id:
        return _error_json_response(401, "请先登录")
    user_role = getattr(request.state, "user_role", "user")
    return (user_id, user_role == "admin")


@_router.post("/api/v1/config/export")
async def export_model_config(
    request: ConfigExportRequest,
    raw_request: Request,
) -> JSONResponse:
    """将模型配置加密导出为令牌字符串。

    管理员可导出全局配置（main/aux/vlm/profiles），普通用户仅可导出自己的个人配置（user）。
    """
    scope_result = await _get_config_transfer_scope(raw_request)
    if isinstance(scope_result, JSONResponse):
        return scope_result
    user_id, is_admin_scope = scope_result

    if _config is None:
        return _error_json_response(503, "服务未初始化")

    if is_admin_scope:
        valid_sections = {"main", "aux", "vlm", "embedding", "profiles"}
        invalid = set(request.sections) - valid_sections
        if invalid:
            return _error_json_response(400, f"无效的配置区块: {', '.join(invalid)}")
        sections = _collect_raw_sections(request.sections)
    else:
        valid_sections = {"user"}
        invalid = set(request.sections) - valid_sections
        if invalid:
            return _error_json_response(400, "普通用户仅可导出个人配置（user）")
        if "user" not in request.sections:
            return _error_json_response(400, "请选择导出个人配置")
        sections = {"user": _collect_user_section(raw_request, user_id)}

    try:
        from excelmanus.config_transfer import export_config
        token = export_config(sections, password=request.password, mode=request.mode)
    except ValueError as exc:
        return _error_json_response(400, str(exc))

    return JSONResponse(content={"token": token, "sections": list(sections.keys()), "mode": request.mode})


@_router.post("/api/v1/config/import")
async def import_model_config(
    request: ConfigImportRequest,
    raw_request: Request,
) -> JSONResponse:
    """解密并导入模型配置令牌。

    管理员可导入全局配置（main/aux/vlm/profiles），普通用户仅可导入自己的个人配置（user）。
    """
    scope_result = await _get_config_transfer_scope(raw_request)
    if isinstance(scope_result, JSONResponse):
        return scope_result
    user_id, is_admin_scope = scope_result

    if _config is None:
        return _error_json_response(503, "服务未初始化")

    try:
        from excelmanus.config_transfer import import_config
        payload = import_config(request.token, password=request.password)
    except ValueError as exc:
        return _error_json_response(400, str(exc))

    sections = payload.get("sections", {})
    imported: dict[str, Any] = {}

    if is_admin_scope:
        env_path = _find_env_file()
        lines = _read_env_file(env_path)
        env_dirty = False

        for section_key in ("main", "aux", "vlm", "embedding"):
            section_data = sections.get(section_key)
            if not isinstance(section_data, dict):
                continue
            key_map = _MODEL_ENV_KEYS.get(section_key)
            if not key_map:
                continue

            updated_fields: list[str] = []
            for field_name in ("api_key", "base_url", "model", "protocol"):
                val = section_data.get(field_name)
                if val is None or not isinstance(val, str):
                    continue
                env_key = key_map.get(field_name)
                if not env_key:
                    continue
                lines = _update_env_var(lines, env_key, val)
                if val:
                    os.environ[env_key] = val
                elif env_key in os.environ:
                    del os.environ[env_key]
                updated_fields.append(field_name)
                env_dirty = True

            # enabled 字段为 bool，单独处理
            if "enabled" in section_data and isinstance(section_data["enabled"], bool):
                env_key = key_map.get("enabled")
                if env_key:
                    env_val = "true" if section_data["enabled"] else "false"
                    lines = _update_env_var(lines, env_key, env_val)
                    os.environ[env_key] = env_val
                    updated_fields.append("enabled")
                    env_dirty = True

            if _config is not None and updated_fields:
                field_map = {
                    "main": {"api_key": "api_key", "base_url": "base_url", "model": "model", "protocol": "protocol"},
                    "aux": {"api_key": "aux_api_key", "base_url": "aux_base_url", "model": "aux_model", "protocol": "aux_protocol"},
                    "vlm": {"api_key": "vlm_api_key", "base_url": "vlm_base_url", "model": "vlm_model", "protocol": "vlm_protocol"},
                    "embedding": {"api_key": "embedding_api_key", "base_url": "embedding_base_url", "model": "embedding_model", "enabled": "embedding_enabled"},
                }.get(section_key, {})
                for f in updated_fields:
                    config_attr = field_map.get(f)
                    if config_attr:
                        object.__setattr__(_config, config_attr, section_data.get(f) if f == "enabled" else (section_data.get(f) or None))

            if updated_fields:
                imported[section_key] = updated_fields

        if env_dirty:
            _write_env_file(env_path, lines)

        profiles_data = sections.get("profiles")
        if isinstance(profiles_data, list) and _config_store is not None:
            profile_names: list[str] = []
            for p in profiles_data:
                if not isinstance(p, dict):
                    continue
                name = p.get("name", "").strip()
                model = p.get("model", "").strip()
                if not name or not model:
                    continue
                existing = _config_store.get_profile(name)
                if existing:
                    _config_store.update_profile(
                        name,
                        model=model,
                        api_key=p.get("api_key", ""),
                        base_url=p.get("base_url", ""),
                        description=p.get("description", ""),
                        protocol=p.get("protocol", "auto"),
                        thinking_mode=p.get("thinking_mode", "auto"),
                        model_family=p.get("model_family", ""),
                        custom_extra_body=p.get("custom_extra_body", ""),
                        custom_extra_headers=p.get("custom_extra_headers", ""),
                    )
                else:
                    _config_store.add_profile(
                        name=name,
                        model=model,
                        api_key=p.get("api_key", ""),
                        base_url=p.get("base_url", ""),
                        description=p.get("description", ""),
                        protocol=p.get("protocol", "auto"),
                        thinking_mode=p.get("thinking_mode", "auto"),
                        model_family=p.get("model_family", ""),
                        custom_extra_body=p.get("custom_extra_body", ""),
                        custom_extra_headers=p.get("custom_extra_headers", ""),
                    )
                profile_names.append(name)
            if profile_names:
                imported["profiles"] = profile_names
                _sync_config_profiles_from_db()
    else:
        user_data = sections.get("user")
        if isinstance(user_data, dict):
            store = getattr(raw_request.app.state, "user_store", None)
            if store is not None:
                updates: dict[str, object] = {}
                for key, attr in (("api_key", "llm_api_key"), ("base_url", "llm_base_url"), ("model", "llm_model")):
                    val = user_data.get(key)
                    if val is not None and isinstance(val, str):
                        updates[attr] = val if val else None
                if updates:
                    store.update_user(user_id, **updates)
                    imported["user"] = [k for k in ("api_key", "base_url", "model") if k in user_data]

    return JSONResponse(content={
        "status": "ok",
        "imported": imported,
        "exported_at": payload.get("ts", ""),
    })


@_router.post("/api/v1/config/transfer/detect")
async def detect_config_token(request: Request) -> JSONResponse:
    """检测令牌的加密模式（用于前端判断是否需要密码输入框）。"""
    body = await request.json()
    token = body.get("token", "")
    if not token:
        return _error_json_response(400, "缺少 token 字段")

    from excelmanus.config_transfer import detect_token_mode
    mode = detect_token_mode(token)
    return JSONResponse(content={"mode": mode, "needs_password": mode == "password"})


# ── 模型能力探测 API ──────────────────────────────────


def _resolve_active_engine_info() -> tuple[str, str, str]:
    """返回最新的主模型配置（始终从 _config 读取，不用引擎缓存）。"""
    assert _config is not None
    return _config.model, _config.base_url, _config.api_key


def _resolve_model_info(
    req_name: str | None,
    req_model: str | None,
    req_base_url: str | None,
) -> tuple[str, str, str, str]:
    """根据请求参数解析模型信息。

    优先级：profile name > model ID 匹配 profile > 直接使用 model+base_url。
    返回 (model, base_url, api_key, protocol)。
    """
    assert _config is not None
    _default_protocol = _config.protocol or "auto"

    # 1) 无参数：返回主模型配置
    if not req_name and not req_model:
        m, b, a = _resolve_active_engine_info()
        return m, b, a, _default_protocol

    # 1.5) 内置 section 名称直接返回对应配置
    if req_name == "main":
        return _config.model, _config.base_url, _config.api_key, _default_protocol
    if req_name == "aux":
        return (_config.aux_model or _config.model,
                _config.aux_base_url or _config.base_url,
                _config.aux_api_key or _config.api_key,
                getattr(_config, 'aux_protocol', None) or _default_protocol)
    if req_name == "vlm":
        return (_config.vlm_model or _config.model,
                _config.vlm_base_url or _config.base_url,
                _config.vlm_api_key or _config.api_key,
                getattr(_config, 'vlm_protocol', None) or _default_protocol)

    # 2) 按 profile name 精确查找
    lookup_name = req_name or req_model

    # 2-1) DB profile 查找
    if _config_store is not None and lookup_name:
        profile = _config_store.get_profile(lookup_name)
        if profile is not None:
            p_base_url = profile.get("base_url") or _config.base_url
            p_api_key = profile.get("api_key") or _config.api_key
            p_protocol = profile.get("protocol") or _default_protocol
            return profile["model"], p_base_url, p_api_key, p_protocol

    # 3) 按 model ID 在所有 profiles 中查找（处理前端传 model ID 而非 name 的情况）
    if _config_store is not None and req_model:
        for p in _config_store.list_profiles():
            if p["model"] == req_model:
                p_base_url = p.get("base_url") or _config.base_url
                p_api_key = p.get("api_key") or _config.api_key
                p_protocol = p.get("protocol") or _default_protocol
                # 如果也指定了 base_url，需要匹配
                if req_base_url and p_base_url != req_base_url:
                    continue
                return p["model"], p_base_url, p_api_key, p_protocol

    # 4) 兜底：直接使用参数，API key 从最新 _config 取
    model = req_model or _config.model
    base_url = req_base_url or _config.base_url
    return model, base_url, _config.api_key, _default_protocol


@_router.get("/api/v1/config/models/capabilities")
async def get_model_capabilities(request: Request) -> JSONResponse:
    """获取模型能力探测结果。支持 ?model=xxx 查询指定模型，否则返回当前活跃模型。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import load_capabilities

    db = _session_manager.database if _session_manager else None
    if db is None:
        return JSONResponse(content={"capabilities": None, "reason": "数据库未启用"})

    req_name = request.query_params.get("name")
    req_model = request.query_params.get("model")
    req_base_url = request.query_params.get("base_url")
    model, base_url, _, _protocol = _resolve_model_info(req_name, req_model, req_base_url)

    caps = load_capabilities(db, model, base_url)
    return JSONResponse(content={
        "capabilities": caps.to_dict() if caps else None,
        "model": model,
        "base_url": base_url,
    })


@_router.get("/api/v1/config/models/capabilities/all")
async def get_all_model_capabilities(request: Request) -> JSONResponse:
    """获取所有已配置模型的能力探测结果（主模型 + profiles）。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import load_capabilities

    db = _session_manager.database if _session_manager else None
    if db is None:
        return JSONResponse(content={"items": []})

    result: list[dict] = []

    # 主模型
    main_caps = load_capabilities(db, _config.model, _config.base_url)
    result.append({
        "name": "main",
        "model": _config.model,
        "base_url": _config.base_url,
        "capabilities": main_caps.to_dict() if main_caps else None,
    })

    # 模型配置档案
    profiles = _config_store.list_profiles() if _config_store else []
    for p in profiles:
        p_model = p["model"]
        p_base_url = p.get("base_url") or _config.base_url
        caps = load_capabilities(db, p_model, p_base_url)
        result.append({
            "name": p["name"],
            "model": p_model,
            "base_url": p_base_url,
            "capabilities": caps.to_dict() if caps else None,
        })

    return JSONResponse(content={"items": result})


@_router.post("/api/v1/config/models/capabilities/probe")
async def probe_model_capabilities(request: Request) -> JSONResponse:
    """手动触发模型能力探测。支持 body 指定 model/base_url/api_key，否则探测当前活跃模型。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import delete_capabilities, run_full_probe
    from excelmanus.providers import create_client

    db = _session_manager.database if _session_manager else None

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass

    req_name = body.get("name")
    req_model = body.get("model")
    req_base_url = body.get("base_url")

    model, base_url, api_key, resolved_protocol = _resolve_model_info(req_name, req_model, req_base_url)

    # 优先尝试当前用户的运行时凭据（如 Codex OAuth access token），
    # 以便能通过 backend-api 执行真实探测并写入缓存。
    _resolver = getattr(getattr(request, "app", None).state, "credential_resolver", None)
    _user_id = _get_isolation_user_id(request)
    if _resolver is not None and _user_id:
        try:
            _resolved_cred = _resolver.resolve_sync(_user_id, model)
            if _resolved_cred:
                api_key = _resolved_cred.api_key or api_key
                if _resolved_cred.base_url:
                    base_url = _resolved_cred.base_url
                if _resolved_cred.protocol:
                    resolved_protocol = _resolved_cred.protocol
        except Exception:
            logger.debug("能力探测解析运行时凭证失败", exc_info=True)

    if body.get("api_key"):
        api_key = body["api_key"]

    # 解析 thinking_mode（来自请求体或 profile 配置）
    req_thinking_mode = body.get("thinking_mode", "auto")
    if req_thinking_mode == "auto" and _config_store and req_name:
        _prof = _config_store.get_profile(req_name)
        if _prof:
            req_thinking_mode = _prof.get("thinking_mode", "auto")

    if db is not None:
        delete_capabilities(db, model, base_url)

    # 非 ASCII 字符检测（httpx 用 ASCII 编码 HTTP header）
    for _fl, _fv in [("API Key", api_key), ("Base URL", base_url), ("Model", model)]:
        try:
            _fv.encode("ascii")
        except UnicodeEncodeError as _enc:
            _bc = _enc.object[_enc.start:_enc.end]
            return _error_json_response(
                400, f"{_fl} 包含非法字符 '{_bc}'（位置 {_enc.start}），请检查是否有多余的特殊字符"
            )

    req_protocol = body.get("protocol") or resolved_protocol
    client = create_client(api_key=api_key, base_url=base_url, protocol=req_protocol)

    try:
        caps = await run_full_probe(
            client=client,
            model=model,
            base_url=base_url,
            skip_if_cached=False,
            db=db,
            thinking_mode=req_thinking_mode,
        )
    except Exception as exc:
        return _error_json_response(500, f"探测失败: {exc}")

    # 同步到所有活跃会话的引擎
    if _session_manager is not None:
        await _session_manager.broadcast_model_capabilities(model, caps)

    return JSONResponse(content={"capabilities": caps.to_dict(), "model": model})


@_router.post("/api/v1/config/models/capabilities/probe-all")
async def probe_all_model_capabilities(request: Request) -> JSONResponse:
    """一键探测所有已配置模型的能力。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import delete_capabilities, run_full_probe
    from excelmanus.providers import create_client

    db = _session_manager.database if _session_manager else None

    # 收集所有需要探测的 (name, model, base_url, api_key, protocol) 元组
    targets: list[tuple[str, str, str, str, str]] = []
    targets.append(("main", _config.model, _config.base_url, _config.api_key, _config.protocol))

    profiles = _config_store.list_profiles() if _config_store else []
    for p in profiles:
        # Codex OAuth 档案使用用户订阅凭据，不走通用 API Key 探测
        if p.get("model", "").startswith("openai-codex/"):
            continue
        p_base_url = p.get("base_url") or _config.base_url
        p_api_key = p.get("api_key") or _config.api_key
        p_protocol = p.get("protocol") or _config.protocol
        targets.append((p["name"], p["model"], p_base_url, p_api_key, p_protocol))

    results: list[dict] = []
    # 构建 name → thinking_mode 映射
    _thinking_mode_map: dict[str, str] = {}
    for p in profiles:
        _thinking_mode_map[p["name"]] = p.get("thinking_mode", "auto")

    for name, model, base_url, api_key, protocol in targets:
        if db is not None:
            delete_capabilities(db, model, base_url)
        client = create_client(api_key=api_key, base_url=base_url, protocol=protocol)
        _tm = _thinking_mode_map.get(name, "auto")
        try:
            caps = await run_full_probe(
                client=client, model=model, base_url=base_url,
                skip_if_cached=False, db=db, thinking_mode=_tm,
            )
            results.append({
                "name": name, "model": model,
                "capabilities": caps.to_dict(),
            })
            if _session_manager is not None:
                await _session_manager.broadcast_model_capabilities(model, caps)
        except Exception as exc:
            results.append({
                "name": name, "model": model,
                "error": str(exc)[:200],
            })

    return JSONResponse(content={"results": results})


def _diagnose_connection_error(error: str, base_url: str, model: str) -> str:
    """根据错误信息和配置，返回用户可操作的修复建议。"""
    err_lower = error.lower()

    # ── 请求被拦截（常见于 base_url 缺少 /v1） ──
    if "blocked" in err_lower or "request was blocked" in err_lower:
        if base_url and not base_url.rstrip("/").endswith("/v1"):
            return f"请求被拦截，通常是 Base URL 缺少 /v1 路径。请尝试将 Base URL 改为：{base_url.rstrip('/')}/v1"
        return "请求被 API 服务商拦截，请检查 Base URL 是否正确、账号是否有访问权限。"

    # ── 404：路径错误 ──
    if "404" in err_lower or "not found" in err_lower:
        if base_url and not base_url.rstrip("/").endswith("/v1"):
            return f"API 端点不存在 (404)。请尝试在 Base URL 末尾加上 /v1：{base_url.rstrip('/')}/v1"
        return "API 端点不存在 (404)。请检查 Base URL 和模型名称是否正确。"

    # ── 认证失败 ──
    if any(k in err_lower for k in ("401", "unauthorized", "invalid api key", "authentication", "invalid.*key")):
        return "API Key 认证失败，请检查 Key 是否正确、是否已过期。"

    # ── 权限不足 ──
    if any(k in err_lower for k in ("403", "forbidden", "permission")):
        return "权限不足 (403)。该 API Key 可能无权访问此模型，请确认账号权限。"

    # ── 额度/计费 ──
    if any(k in err_lower for k in ("402", "quota", "billing", "balance", "insufficient", "payment")):
        return "账号额度不足或计费问题，请检查 API 账号余额。"

    # ── 限流 ──
    if any(k in err_lower for k in ("429", "rate limit", "too many")):
        return "请求被限流 (429)，请稍后重试或降低请求频率。"

    # ── 模型不存在 ──
    if any(k in err_lower for k in ("model not found", "model_not_found", "does not exist", "no such model")):
        return f"模型 '{model}' 不存在。请检查模型名称是否拼写正确。"

    # ── 账号池耗尽 ──
    if any(k in err_lower for k in ("no.*account.*available", "no available", "pool.*exhaust")):
        return "API 代理的账号池暂无可用账号，请稍后重试。"

    # ── 超时 ──
    if any(k in err_lower for k in ("timeout", "timed out")):
        return "请求超时。请检查网络连通性，或 Base URL 是否可访问。"

    # ── 连接失败 ──
    if any(k in err_lower for k in ("connection refused", "connection error", "connect error", "name resolution", "getaddrinfo", "dns")):
        return "无法连接到 API 服务器。请检查 Base URL 是否正确、网络是否畅通。"

    # ── SSL 错误 ──
    if any(k in err_lower for k in ("ssl", "certificate", "cert")):
        return "SSL 证书验证失败。请检查 Base URL 的 HTTPS 证书是否有效。"

    return ""


@_router.post("/api/v1/config/models/test-connection")
async def test_model_connection(request: Request) -> JSONResponse:
    """轻量级模型连通测试：仅发送 Hi 检查模型是否可达、鉴权是否正常。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import probe_health
    from excelmanus.providers import create_client

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass

    req_name = body.get("name")
    req_model = body.get("model")
    req_base_url = body.get("base_url")

    # Codex OAuth 档案使用用户订阅凭据，无法用通用 API Key 测试
    _test_model_id = req_model or ""
    if not _test_model_id and req_name and _config_store:
        _tp = _config_store.get_profile(req_name)
        if _tp:
            _test_model_id = _tp.get("model", "")
    if _test_model_id.startswith("openai-codex/"):
        return JSONResponse(content={
            "ok": True,
            "model": _test_model_id,
            "note": "Codex OAuth 模型使用用户订阅凭据，无需通用 API Key 测试",
        })

    model, base_url, api_key, resolved_protocol = _resolve_model_info(req_name, req_model, req_base_url)
    if body.get("api_key"):
        api_key = body["api_key"]

    # 格式校验
    if not model or not model.strip():
        return JSONResponse(content={"ok": False, "error": "模型标识符为空", "model": model})
    if not base_url or not base_url.strip():
        return JSONResponse(content={"ok": False, "error": "Base URL 为空", "model": model})
    if not api_key or not api_key.strip():
        return JSONResponse(content={"ok": False, "error": "API Key 为空", "model": model})

    # 占位符检测
    placeholder_patterns = [
        "sk-xxx", "your-api-key", "your_api_key", "api-key-here",
        "replace-with", "填入", "替换为", "请填写", "请输入",
    ]
    api_key_lower = api_key.strip().lower()
    for pat in placeholder_patterns:
        if pat in api_key_lower:
            return JSONResponse(content={
                "ok": False,
                "error": f"API Key 似乎是占位符（包含 '{pat}'），请填入真实的 API Key",
                "is_placeholder": True,
                "model": model,
            })

    # 非 ASCII 字符检测（httpx 用 ASCII 编码 HTTP header，非 ASCII 会导致 UnicodeEncodeError）
    for _field_label, _field_val in [("API Key", api_key), ("Base URL", base_url), ("Model", model)]:
        try:
            _field_val.encode("ascii")
        except UnicodeEncodeError as _enc_err:
            _bad_char = _enc_err.object[_enc_err.start:_enc_err.end]
            return JSONResponse(content={
                "ok": False,
                "error": f"{_field_label} 包含非法字符 '{_bad_char}'（位置 {_enc_err.start}），请检查是否有多余的特殊字符或空格",
                "model": model,
            })

    req_protocol = body.get("protocol") or resolved_protocol
    client = create_client(api_key=api_key, base_url=base_url, protocol=req_protocol)
    try:
        healthy, health_err = await probe_health(client, model, timeout=15.0)
    except Exception as exc:
        err_str = str(exc)[:200]
        hint = _diagnose_connection_error(err_str, base_url, model)
        return JSONResponse(content={"ok": False, "error": f"连通测试异常: {err_str}", "hint": hint, "model": model})

    result: dict = {
        "ok": healthy,
        "error": health_err if not healthy else "",
        "model": model,
        "base_url": base_url,
    }
    if not healthy and health_err:
        hint = _diagnose_connection_error(health_err, base_url, model)
        if hint:
            result["hint"] = hint
    return JSONResponse(content=result)


# Provider fallback model lists used when /models endpoint returns 404.
# Each entry: (base_url keyword, model list, user hint)
_PROVIDER_FALLBACK_MODELS: list[tuple[str, list[dict], str]] = [
    (
        "minimax",
        [
            {"id": "MiniMax-M2.5"},
            {"id": "MiniMax-M2.5-highspeed"},
            {"id": "MiniMax-M2.1"},
            {"id": "MiniMax-M2.1-highspeed"},
            {"id": "MiniMax-M2.1-lightning"},
            {"id": "MiniMax-M2"},
            {"id": "M2-her"},
        ],
        "MiniMax \u901a\u5e38\u4e0d\u652f\u6301 /models \u679a\u4e3e\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002"
        "\u82e5\u4ecd\u5f02\u5e38\uff0c\u8bf7\u786e\u8ba4 Base URL\uff08\u5efa\u8bae https://api.minimax.io/v1\uff09\u548c API Key\u3002",
    ),
    (
        "generativelanguage.googleapis.com",
        [
            {"id": "gemini-2.5-pro"},
            {"id": "gemini-2.5-flash"},
            {"id": "gemini-2.5-flash-lite"},
            {"id": "gemini-2.0-flash"},
            {"id": "gemini-2.0-flash-lite"},
            {"id": "gemini-1.5-pro"},
            {"id": "gemini-1.5-flash"},
        ],
        "Gemini OpenAI \u517c\u5bb9\u7aef\u70b9\u4e0d\u652f\u6301\u6807\u51c6 /models \u679a\u4e3e\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002",
    ),
    (
        "bigmodel.cn",
        [
            {"id": "glm-4-plus"},
            {"id": "glm-4"},
            {"id": "glm-4-long"},
            {"id": "glm-4-flash"},
            {"id": "glm-z1"},
            {"id": "glm-z1-flash"},
            {"id": "glm-z1-air"},
            {"id": "glm-4v-plus"},
            {"id": "glm-4v"},
        ],
        "\u667a\u8c31 GLM /models \u7aef\u70b9\u8def\u5f84\u4e0e\u6807\u51c6 OpenAI \u4e0d\u540c\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002",
    ),
    (
        "dashscope.aliyuncs.com",
        [
            {"id": "qwen-max"},
            {"id": "qwen-plus"},
            {"id": "qwen-turbo"},
            {"id": "qwen-long"},
            {"id": "qwen-flash"},
            {"id": "qwq-plus"},
            {"id": "qwen3-235b"},
            {"id": "qwen3-30b"},
            {"id": "qwen3-32b"},
            {"id": "qwen-coder-plus"},
            {"id": "qwen3-coder-plus"},
        ],
        "\u963f\u91cc\u4e91\u767e\u70bc DashScope /models \u679a\u4e3e\u901a\u5e38\u9700\u8981\u7279\u5b9a\u6743\u9650\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002",
    ),
    (
        "moonshot.cn",
        [
            {"id": "kimi-k2"},
            {"id": "kimi-k2-thinking"},
            {"id": "kimi-k2.5"},
            {"id": "moonshot-v1-128k"},
            {"id": "moonshot-v1-32k"},
            {"id": "moonshot-v1-8k"},
        ],
        "Kimi (Moonshot) /models \u7aef\u70b9\u4e0d\u53ef\u7528\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002",
    ),
    (
        "deepseek.com",
        [
            {"id": "deepseek-chat"},
            {"id": "deepseek-reasoner"},
        ],
        "DeepSeek /models \u7aef\u70b9\u4e0d\u53ef\u7528\uff0c\u5df2\u56de\u9000\u4e3a\u63a8\u8350\u6a21\u578b\u5217\u8868\u3002",
    ),
]


def _get_provider_fallback(base_url: str) -> tuple[list[dict], str] | None:
    """Return curated model list for a known provider when /models returns 404."""
    url_lower = base_url.lower()
    for pattern, models, hint in _PROVIDER_FALLBACK_MODELS:
        if pattern in url_lower:
            return models, hint
    return None


@_router.post("/api/v1/config/models/list-remote")
async def list_remote_models(request: Request) -> JSONResponse:
    """调用远程 API 端点列出可用模型（用于添加模型时自动检测）。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    body: dict = {}
    try:
        body = await request.json()
    except Exception:
        pass

    base_url = (body.get("base_url") or "").strip() or _config.base_url
    api_key = (body.get("api_key") or "").strip() or _config.api_key
    protocol = (body.get("protocol") or "auto").strip().lower()

    if not base_url:
        return JSONResponse(content={"models": [], "error": "Base URL \u4e3a\u7a7a"})
    if not api_key:
        return JSONResponse(content={"models": [], "error": "API Key \u4e3a\u7a7a"})

    import httpx

    # Always append /models directly to base_url.
    # Never insert an extra /v1 segment — the caller already provides the versioned prefix.
    url = base_url.rstrip("/")
    if not url.endswith("/models"):
        url = url + "/models"

    headers: dict[str, str] = {"Authorization": f"Bearer {api_key}"}
    # Anthropic native API uses x-api-key header
    if protocol == "anthropic" or "anthropic" in base_url.lower():
        headers = {"x-api-key": api_key, "anthropic-version": "2023-06-01"}
        url = base_url.rstrip("/").rstrip("/v1").rstrip("/") + "/v1/models"

    try:
        async with httpx.AsyncClient(timeout=15.0) as http:
            resp = await http.get(url, headers=headers)
            resp.raise_for_status()
            data = resp.json()
    except httpx.TimeoutException:
        return JSONResponse(content={"models": [], "error": "\u8bf7\u6c42\u8d85\u65f6\uff0c\u8bf7\u68c0\u67e5 Base URL \u662f\u5426\u53ef\u8bbf\u95ee"})
    except httpx.HTTPStatusError as exc:
        if exc.response.status_code == 404:
            fallback = _get_provider_fallback(base_url)
            if fallback is not None:
                models, hint = fallback
                return JSONResponse(content={"models": models, "hint": hint})
        hint = _diagnose_connection_error(str(exc)[:200], base_url, "")
        return JSONResponse(content={"models": [], "error": f"HTTP {exc.response.status_code}: {str(exc)[:150]}", "hint": hint})
    except Exception as exc:
        return JSONResponse(content={"models": [], "error": f"请求失败: {str(exc)[:200]}"})

    # 解析模型列表（兼容 OpenAI / Anthropic / 各类代理格式）
    models_raw: list = []
    if isinstance(data, dict):
        models_raw = data.get("data") or data.get("models") or []
    elif isinstance(data, list):
        models_raw = data

    models_out: list[dict] = []
    for m in models_raw:
        if isinstance(m, str):
            models_out.append({"id": m})
        elif isinstance(m, dict):
            mid = m.get("id") or m.get("name") or m.get("model") or ""
            if mid:
                models_out.append({"id": mid, "owned_by": m.get("owned_by", "")})

    # 按 id 排序
    models_out.sort(key=lambda x: x["id"])

    return JSONResponse(content={"models": models_out})


@_router.get("/api/v1/config/models/check-placeholder")
async def check_model_placeholder(request: Request) -> JSONResponse:
    """检测当前模型配置是否使用了默认占位符值。"""
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    placeholder_patterns = [
        "sk-xxx", "your-api-key", "your_api_key", "api-key-here",
        "replace-with", "填入", "替换为", "请填写", "请输入",
    ]

    def _is_placeholder(val: str) -> bool:
        if not val or not val.strip():
            return True
        lower = val.strip().lower()
        return any(p in lower for p in placeholder_patterns)

    results: list[dict] = []

    # 主模型
    if _is_placeholder(_config.api_key):
        results.append({"name": "main", "field": "api_key", "model": _config.model})
    if not _config.model or not _config.model.strip():
        results.append({"name": "main", "field": "model", "model": ""})

    # AUX
    if _config.aux_model:
        if _is_placeholder(_config.aux_api_key or ""):
            results.append({"name": "aux", "field": "api_key", "model": _config.aux_model})

    # VLM
    if _config.vlm_model:
        if _is_placeholder(_config.vlm_api_key or ""):
            results.append({"name": "vlm", "field": "api_key", "model": _config.vlm_model})

    # Profiles
    profiles = _config_store.list_profiles() if _config_store else []
    for p in profiles:
        p_api_key = p.get("api_key") or _config.api_key
        if _is_placeholder(p_api_key):
            results.append({"name": p["name"], "field": "api_key", "model": p["model"]})

    return JSONResponse(content={
        "has_placeholder": len(results) > 0,
        "items": results,
    })


@_router.put("/api/v1/config/models/capabilities")
async def update_model_capabilities(request: Request) -> JSONResponse:
    """手动覆盖模型能力标记（前端设置用）。"""
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    if _config is None:
        return _error_json_response(503, "服务未初始化")

    from excelmanus.model_probe import update_capabilities_override

    db = _session_manager.database if _session_manager else None
    if db is None:
        return _error_json_response(503, "数据库未启用")

    body = await request.json()
    overrides = body.get("overrides", {})
    if not overrides:
        return _error_json_response(400, "缺少 overrides 字段")

    req_name = body.get("name")
    req_model = body.get("model")
    req_base_url = body.get("base_url")
    model, base_url, _, _protocol = _resolve_model_info(req_name, req_model, req_base_url)

    caps = update_capabilities_override(db, model, base_url, overrides)

    if caps is not None and _session_manager is not None:
        await _session_manager.broadcast_model_capabilities(model, caps)

    return JSONResponse(content={
        "capabilities": caps.to_dict() if caps else None,
    })


# ── 运行时配置管理 API ──────────────────────────────────

_RUNTIME_ENV_KEYS: dict[str, str] = {
    # ── 会话与多用户 ──
    "auth_enabled": "EXCELMANUS_AUTH_ENABLED",
    "session_ttl_seconds": "EXCELMANUS_SESSION_TTL_SECONDS",
    "max_sessions": "EXCELMANUS_MAX_SESSIONS",
    "max_consecutive_failures": "EXCELMANUS_MAX_CONSECUTIVE_FAILURES",
    # ── 执行与安全 ──
    "subagent_enabled": "EXCELMANUS_SUBAGENT_ENABLED",
    "verifier_enabled": "EXCELMANUS_VERIFIER_ENABLED",
    "backup_enabled": "EXCELMANUS_BACKUP_ENABLED",
    "checkpoint_enabled": "EXCELMANUS_CHECKPOINT_ENABLED",
    "external_safe_mode": "EXCELMANUS_EXTERNAL_SAFE_MODE",
    "max_iterations": "EXCELMANUS_MAX_ITERATIONS",
    "guard_mode": "EXCELMANUS_GUARD_MODE",
    "friendly_error_messages": "EXCELMANUS_FRIENDLY_ERROR_MESSAGES",
    # ── AUX / VLM 开关 ──
    "aux_enabled": "EXCELMANUS_AUX_ENABLED",
    "vlm_enabled": "EXCELMANUS_VLM_ENABLED",
    # ── 上下文与记忆 ──
    "max_context_tokens": "EXCELMANUS_MAX_CONTEXT_TOKENS",
    "memory_enabled": "EXCELMANUS_MEMORY_ENABLED",
    "memory_auto_extract_interval": "EXCELMANUS_MEMORY_AUTO_EXTRACT_INTERVAL",
    "memory_auto_load_lines": "EXCELMANUS_MEMORY_AUTO_LOAD_LINES",
    "memory_expire_days": "EXCELMANUS_MEMORY_EXPIRE_DAYS",
    "chat_history_enabled": "EXCELMANUS_CHAT_HISTORY_ENABLED",
    # ── 记忆维护 ──
    "memory_maintenance_enabled": "EXCELMANUS_MEMORY_MAINTENANCE_ENABLED",
    "memory_maintenance_min_entries": "EXCELMANUS_MEMORY_MAINTENANCE_MIN_ENTRIES",
    "memory_maintenance_new_threshold": "EXCELMANUS_MEMORY_MAINTENANCE_NEW_THRESHOLD",
    "memory_maintenance_interval_hours": "EXCELMANUS_MEMORY_MAINTENANCE_INTERVAL_HOURS",
    "memory_maintenance_model": "EXCELMANUS_MEMORY_MAINTENANCE_MODEL",
    # ── 摘要与压缩 ──
    "summarization_enabled": "EXCELMANUS_SUMMARIZATION_ENABLED",
    "summarization_threshold_ratio": "EXCELMANUS_SUMMARIZATION_THRESHOLD_RATIO",
    "summarization_keep_recent_turns": "EXCELMANUS_SUMMARIZATION_KEEP_RECENT_TURNS",
    "compaction_enabled": "EXCELMANUS_COMPACTION_ENABLED",
    "compaction_threshold_ratio": "EXCELMANUS_COMPACTION_THRESHOLD_RATIO",
    "compaction_keep_recent_turns": "EXCELMANUS_COMPACTION_KEEP_RECENT_TURNS",
    "compaction_max_summary_tokens": "EXCELMANUS_COMPACTION_MAX_SUMMARY_TOKENS",
    "prompt_cache_key_enabled": "EXCELMANUS_PROMPT_CACHE_KEY_ENABLED",
    # ── 推理配置 ──
    "thinking_effort": "EXCELMANUS_THINKING_EFFORT",
    "thinking_budget": "EXCELMANUS_THINKING_BUDGET",
    # ── 子代理 ──
    "subagent_max_iterations": "EXCELMANUS_SUBAGENT_MAX_ITERATIONS",
    "subagent_timeout_seconds": "EXCELMANUS_SUBAGENT_TIMEOUT_SECONDS",
    "subagent_max_consecutive_failures": "EXCELMANUS_SUBAGENT_MAX_CONSECUTIVE_FAILURES",
    "parallel_subagent_max": "EXCELMANUS_PARALLEL_SUBAGENT_MAX",
    # ── LLM 重试 ──
    "llm_retry_max_attempts": "EXCELMANUS_LLM_RETRY_MAX_ATTEMPTS",
    "llm_retry_base_delay_seconds": "EXCELMANUS_LLM_RETRY_BASE_DELAY_SECONDS",
    "llm_retry_max_delay_seconds": "EXCELMANUS_LLM_RETRY_MAX_DELAY_SECONDS",
    # ── 感知与视觉 ──
    "window_perception_enabled": "EXCELMANUS_WINDOW_PERCEPTION_ENABLED",
    "window_perception_system_budget_tokens": "EXCELMANUS_WINDOW_PERCEPTION_SYSTEM_BUDGET_TOKENS",
    "window_perception_tool_append_tokens": "EXCELMANUS_WINDOW_PERCEPTION_TOOL_APPEND_TOKENS",
    "window_perception_max_windows": "EXCELMANUS_WINDOW_PERCEPTION_MAX_WINDOWS",
    "window_perception_default_rows": "EXCELMANUS_WINDOW_PERCEPTION_DEFAULT_ROWS",
    "window_perception_default_cols": "EXCELMANUS_WINDOW_PERCEPTION_DEFAULT_COLS",
    "window_perception_minimized_tokens": "EXCELMANUS_WINDOW_PERCEPTION_MINIMIZED_TOKENS",
    "window_perception_background_after_idle": "EXCELMANUS_WINDOW_PERCEPTION_BACKGROUND_AFTER_IDLE",
    "window_perception_suspend_after_idle": "EXCELMANUS_WINDOW_PERCEPTION_SUSPEND_AFTER_IDLE",
    "window_perception_terminate_after_idle": "EXCELMANUS_WINDOW_PERCEPTION_TERMINATE_AFTER_IDLE",
    "window_perception_advisor_mode": "EXCELMANUS_WINDOW_PERCEPTION_ADVISOR_MODE",
    "window_perception_advisor_timeout_ms": "EXCELMANUS_WINDOW_PERCEPTION_ADVISOR_TIMEOUT_MS",
    "window_perception_advisor_trigger_window_count": "EXCELMANUS_WINDOW_PERCEPTION_ADVISOR_TRIGGER_WINDOW_COUNT",
    "window_perception_advisor_trigger_turn": "EXCELMANUS_WINDOW_PERCEPTION_ADVISOR_TRIGGER_TURN",
    "window_perception_advisor_plan_ttl_turns": "EXCELMANUS_WINDOW_PERCEPTION_ADVISOR_PLAN_TTL_TURNS",
    "window_return_mode": "EXCELMANUS_WINDOW_RETURN_MODE",
    "window_full_max_rows": "EXCELMANUS_WINDOW_FULL_MAX_ROWS",
    "window_full_total_budget_tokens": "EXCELMANUS_WINDOW_FULL_TOTAL_BUDGET_TOKENS",
    "window_data_buffer_max_rows": "EXCELMANUS_WINDOW_DATA_BUFFER_MAX_ROWS",
    "window_intent_enabled": "EXCELMANUS_WINDOW_INTENT_ENABLED",
    "window_intent_sticky_turns": "EXCELMANUS_WINDOW_INTENT_STICKY_TURNS",
    "window_intent_repeat_warn_threshold": "EXCELMANUS_WINDOW_INTENT_REPEAT_WARN_THRESHOLD",
    "window_intent_repeat_trip_threshold": "EXCELMANUS_WINDOW_INTENT_REPEAT_TRIP_THRESHOLD",
    "window_rule_engine_version": "EXCELMANUS_WINDOW_RULE_ENGINE_VERSION",
    "vlm_enhance": "EXCELMANUS_VLM_ENHANCE",
    "main_model_vision": "EXCELMANUS_MAIN_MODEL_VISION",
    "vlm_timeout_seconds": "EXCELMANUS_VLM_TIMEOUT_SECONDS",
    "vlm_max_retries": "EXCELMANUS_VLM_MAX_RETRIES",
    "vlm_max_tokens": "EXCELMANUS_VLM_MAX_TOKENS",
    "vlm_image_max_long_edge": "EXCELMANUS_VLM_IMAGE_MAX_LONG_EDGE",
    "vlm_image_jpeg_quality": "EXCELMANUS_VLM_IMAGE_JPEG_QUALITY",
    "vlm_extraction_tier": "EXCELMANUS_VLM_EXTRACTION_TIER",
    "image_keep_rounds": "EXCELMANUS_IMAGE_KEEP_ROUNDS",
    "image_max_active": "EXCELMANUS_IMAGE_MAX_ACTIVE",
    "image_token_budget": "EXCELMANUS_IMAGE_TOKEN_BUDGET",
    # ── 系统消息与工具 ──
    "system_message_mode": "EXCELMANUS_SYSTEM_MESSAGE_MODE",
    "tool_result_hard_cap_chars": "EXCELMANUS_TOOL_RESULT_HARD_CAP_CHARS",
    "large_excel_threshold_bytes": "EXCELMANUS_LARGE_EXCEL_THRESHOLD_BYTES",
    "parallel_readonly_tools": "EXCELMANUS_PARALLEL_READONLY_TOOLS",
    "hooks_command_enabled": "EXCELMANUS_HOOKS_COMMAND_ENABLED",
    "hooks_command_timeout_seconds": "EXCELMANUS_HOOKS_COMMAND_TIMEOUT_SECONDS",
    "hooks_output_max_chars": "EXCELMANUS_HOOKS_OUTPUT_MAX_CHARS",
    "log_level": "EXCELMANUS_LOG_LEVEL",
    # ── 代码策略 ──
    "code_policy_enabled": "EXCELMANUS_CODE_POLICY_ENABLED",
    "code_policy_green_auto_approve": "EXCELMANUS_CODE_POLICY_GREEN_AUTO",
    "code_policy_yellow_auto_approve": "EXCELMANUS_CODE_POLICY_YELLOW_AUTO",
    "tool_schema_validation_mode": "EXCELMANUS_TOOL_SCHEMA_VALIDATION_MODE",
    "tool_schema_validation_canary_percent": "EXCELMANUS_TOOL_SCHEMA_VALIDATION_CANARY_PERCENT",
    "tool_schema_strict_path": "EXCELMANUS_TOOL_SCHEMA_STRICT_PATH",
    # ── 技能发现 ──
    "skills_context_char_budget": "EXCELMANUS_SKILLS_CONTEXT_CHAR_BUDGET",
    "skills_discovery_enabled": "EXCELMANUS_SKILLS_DISCOVERY_ENABLED",
    "skills_discovery_scan_workspace_ancestors": "EXCELMANUS_SKILLS_DISCOVERY_SCAN_WORKSPACE_ANCESTORS",
    "skills_discovery_include_agents": "EXCELMANUS_SKILLS_DISCOVERY_INCLUDE_AGENTS",
    "skills_discovery_scan_external_tool_dirs": "EXCELMANUS_SKILLS_DISCOVERY_SCAN_EXTERNAL_TOOL_DIRS",
    # ── Embedding / 语义检索 ──
    "embedding_enabled": "EXCELMANUS_EMBEDDING_ENABLED",
    "embedding_model": "EXCELMANUS_EMBEDDING_MODEL",
    "embedding_dimensions": "EXCELMANUS_EMBEDDING_DIMENSIONS",
    "embedding_timeout_seconds": "EXCELMANUS_EMBEDDING_TIMEOUT_SECONDS",
    "memory_semantic_top_k": "EXCELMANUS_MEMORY_SEMANTIC_TOP_K",
    "memory_semantic_threshold": "EXCELMANUS_MEMORY_SEMANTIC_THRESHOLD",
    "memory_semantic_fallback_recent": "EXCELMANUS_MEMORY_SEMANTIC_FALLBACK_RECENT",
    # ── Playbook ──
    "playbook_enabled": "EXCELMANUS_PLAYBOOK_ENABLED",
    "playbook_max_bullets": "EXCELMANUS_PLAYBOOK_MAX_BULLETS",
    "playbook_inject_top_k": "EXCELMANUS_PLAYBOOK_INJECT_TOP_K",
    "registry_semantic_top_k": "EXCELMANUS_REGISTRY_SEMANTIC_TOP_K",
    "registry_semantic_threshold": "EXCELMANUS_REGISTRY_SEMANTIC_THRESHOLD",
}


@_router.get("/api/v1/config/runtime")
async def get_runtime_config(request: Request) -> JSONResponse:
    """读取运行时行为配置。"""
    assert _config is not None, "服务未初始化"
    guard_error = await _require_admin_if_auth_enabled(request)
    if guard_error is not None:
        return guard_error
    return JSONResponse(content={
        # ── 会话与多用户 ──
        "auth_enabled": getattr(request.app.state, "auth_enabled", False),
        "session_ttl_seconds": _config.session_ttl_seconds,
        "max_sessions": _config.max_sessions,
        "max_consecutive_failures": _config.max_consecutive_failures,
        # ── 执行与安全 ──
        "subagent_enabled": _config.subagent_enabled,
        "verifier_enabled": _config.verifier_enabled,
        "backup_enabled": _config.backup_enabled,
        "checkpoint_enabled": _config.checkpoint_enabled,
        "external_safe_mode": _config.external_safe_mode,
        "max_iterations": _config.max_iterations,
        "guard_mode": _config.guard_mode,
        "friendly_error_messages": _config.friendly_error_messages,
        # ── AUX / VLM 开关 ──
        "aux_enabled": _config.aux_enabled,
        "vlm_enabled": _config.vlm_enabled,
        # ── 上下文与记忆 ──
        "max_context_tokens": _config.max_context_tokens,
        "memory_enabled": _config.memory_enabled,
        "memory_auto_extract_interval": _config.memory_auto_extract_interval,
        "memory_auto_load_lines": _config.memory_auto_load_lines,
        "memory_expire_days": _config.memory_expire_days,
        "chat_history_enabled": _config.chat_history_enabled,
        # ── 记忆维护 ──
        "memory_maintenance_enabled": _config.memory_maintenance_enabled,
        "memory_maintenance_min_entries": _config.memory_maintenance_min_entries,
        "memory_maintenance_new_threshold": _config.memory_maintenance_new_threshold,
        "memory_maintenance_interval_hours": _config.memory_maintenance_interval_hours,
        "memory_maintenance_model": _config.memory_maintenance_model or "",
        # ── 摘要与压缩 ──
        "summarization_enabled": _config.summarization_enabled,
        "summarization_threshold_ratio": _config.summarization_threshold_ratio,
        "summarization_keep_recent_turns": _config.summarization_keep_recent_turns,
        "compaction_enabled": _config.compaction_enabled,
        "compaction_threshold_ratio": _config.compaction_threshold_ratio,
        "compaction_keep_recent_turns": _config.compaction_keep_recent_turns,
        "compaction_max_summary_tokens": _config.compaction_max_summary_tokens,
        "prompt_cache_key_enabled": _config.prompt_cache_key_enabled,
        # ── 推理配置 ──
        "thinking_effort": _config.thinking_effort,
        "thinking_budget": _config.thinking_budget,
        # ── 子代理 ──
        "subagent_max_iterations": _config.subagent_max_iterations,
        "subagent_timeout_seconds": _config.subagent_timeout_seconds,
        "subagent_max_consecutive_failures": _config.subagent_max_consecutive_failures,
        "parallel_subagent_max": _config.parallel_subagent_max,
        # ── LLM 重试 ──
        "llm_retry_max_attempts": _config.llm_retry_max_attempts,
        "llm_retry_base_delay_seconds": _config.llm_retry_base_delay_seconds,
        "llm_retry_max_delay_seconds": _config.llm_retry_max_delay_seconds,
        # ── 感知与视觉 ──
        "window_perception_enabled": _config.window_perception_enabled,
        "window_perception_system_budget_tokens": _config.window_perception_system_budget_tokens,
        "window_perception_tool_append_tokens": _config.window_perception_tool_append_tokens,
        "window_perception_max_windows": _config.window_perception_max_windows,
        "window_perception_default_rows": _config.window_perception_default_rows,
        "window_perception_default_cols": _config.window_perception_default_cols,
        "window_perception_minimized_tokens": _config.window_perception_minimized_tokens,
        "window_perception_background_after_idle": _config.window_perception_background_after_idle,
        "window_perception_suspend_after_idle": _config.window_perception_suspend_after_idle,
        "window_perception_terminate_after_idle": _config.window_perception_terminate_after_idle,
        "window_perception_advisor_mode": _config.window_perception_advisor_mode,
        "window_perception_advisor_timeout_ms": _config.window_perception_advisor_timeout_ms,
        "window_perception_advisor_trigger_window_count": _config.window_perception_advisor_trigger_window_count,
        "window_perception_advisor_trigger_turn": _config.window_perception_advisor_trigger_turn,
        "window_perception_advisor_plan_ttl_turns": _config.window_perception_advisor_plan_ttl_turns,
        "window_return_mode": _config.window_return_mode,
        "window_full_max_rows": _config.window_full_max_rows,
        "window_full_total_budget_tokens": _config.window_full_total_budget_tokens,
        "window_data_buffer_max_rows": _config.window_data_buffer_max_rows,
        "window_intent_enabled": _config.window_intent_enabled,
        "window_intent_sticky_turns": _config.window_intent_sticky_turns,
        "window_intent_repeat_warn_threshold": _config.window_intent_repeat_warn_threshold,
        "window_intent_repeat_trip_threshold": _config.window_intent_repeat_trip_threshold,
        "window_rule_engine_version": _config.window_rule_engine_version,
        "vlm_enhance": _config.vlm_enhance,
        "main_model_vision": _config.main_model_vision,
        "vlm_timeout_seconds": _config.vlm_timeout_seconds,
        "vlm_max_retries": _config.vlm_max_retries,
        "vlm_max_tokens": _config.vlm_max_tokens,
        "vlm_image_max_long_edge": _config.vlm_image_max_long_edge,
        "vlm_image_jpeg_quality": _config.vlm_image_jpeg_quality,
        "vlm_extraction_tier": _config.vlm_extraction_tier,
        "image_keep_rounds": _config.image_keep_rounds,
        "image_max_active": _config.image_max_active,
        "image_token_budget": _config.image_token_budget,
        # ── 系统消息与工具 ──
        "system_message_mode": _config.system_message_mode,
        "tool_result_hard_cap_chars": _config.tool_result_hard_cap_chars,
        "large_excel_threshold_bytes": _config.large_excel_threshold_bytes,
        "parallel_readonly_tools": _config.parallel_readonly_tools,
        "hooks_command_enabled": _config.hooks_command_enabled,
        "hooks_command_timeout_seconds": _config.hooks_command_timeout_seconds,
        "hooks_output_max_chars": _config.hooks_output_max_chars,
        "log_level": _config.log_level,
        # ── 代码策略 ──
        "code_policy_enabled": _config.code_policy_enabled,
        "code_policy_green_auto_approve": _config.code_policy_green_auto_approve,
        "code_policy_yellow_auto_approve": _config.code_policy_yellow_auto_approve,
        "tool_schema_validation_mode": _config.tool_schema_validation_mode,
        "tool_schema_validation_canary_percent": _config.tool_schema_validation_canary_percent,
        "tool_schema_strict_path": _config.tool_schema_strict_path,
        # ── 技能发现 ──
        "skills_context_char_budget": _config.skills_context_char_budget,
        "skills_discovery_enabled": _config.skills_discovery_enabled,
        "skills_discovery_scan_workspace_ancestors": _config.skills_discovery_scan_workspace_ancestors,
        "skills_discovery_include_agents": _config.skills_discovery_include_agents,
        "skills_discovery_scan_external_tool_dirs": _config.skills_discovery_scan_external_tool_dirs,
        # ── Embedding / 语义检索 ──
        "embedding_enabled": _config.embedding_enabled,
        "embedding_model": _config.embedding_model,
        "embedding_dimensions": _config.embedding_dimensions,
        "embedding_timeout_seconds": _config.embedding_timeout_seconds,
        "memory_semantic_top_k": _config.memory_semantic_top_k,
        "memory_semantic_threshold": _config.memory_semantic_threshold,
        "memory_semantic_fallback_recent": _config.memory_semantic_fallback_recent,
        # ── Playbook ──
        "playbook_enabled": _config.playbook_enabled,
        "playbook_max_bullets": _config.playbook_max_bullets,
        "playbook_inject_top_k": _config.playbook_inject_top_k,
        "registry_semantic_top_k": _config.registry_semantic_top_k,
        "registry_semantic_threshold": _config.registry_semantic_threshold,
    })


class RuntimeConfigUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    # ── 会话与多用户 ──
    auth_enabled: bool | None = None
    session_ttl_seconds: int | None = Field(default=None, gt=0)
    max_sessions: int | None = Field(default=None, gt=0)
    max_consecutive_failures: int | None = Field(default=None, gt=0)
    # ── 执行与安全 ──
    subagent_enabled: bool | None = None
    verifier_enabled: bool | None = None
    backup_enabled: bool | None = None
    checkpoint_enabled: bool | None = None
    external_safe_mode: bool | None = None
    max_iterations: int | None = None
    guard_mode: Literal["off", "soft"] | None = None
    friendly_error_messages: bool | None = None
    # ── AUX / VLM 开关 ──
    aux_enabled: bool | None = None
    vlm_enabled: bool | None = None
    # ── 上下文与记忆 ──
    max_context_tokens: int | None = Field(default=None, gt=0)
    memory_enabled: bool | None = None
    memory_auto_extract_interval: int | None = Field(default=None, ge=0)
    memory_auto_load_lines: int | None = Field(default=None, gt=0)
    memory_expire_days: int | None = Field(default=None, ge=0)
    chat_history_enabled: bool | None = None
    # ── 记忆维护 ──
    memory_maintenance_enabled: bool | None = None
    memory_maintenance_min_entries: int | None = Field(default=None, ge=1)
    memory_maintenance_new_threshold: int | None = Field(default=None, ge=1)
    memory_maintenance_interval_hours: float | None = Field(default=None, ge=0.5)
    memory_maintenance_model: str | None = None
    # ── 摘要与压缩 ──
    summarization_enabled: bool | None = None
    summarization_threshold_ratio: float | None = Field(default=None, gt=0, lt=1)
    summarization_keep_recent_turns: int | None = Field(default=None, gt=0)
    compaction_enabled: bool | None = None
    compaction_threshold_ratio: float | None = Field(default=None, gt=0, lt=1)
    compaction_keep_recent_turns: int | None = Field(default=None, gt=0)
    compaction_max_summary_tokens: int | None = Field(default=None, gt=0)
    prompt_cache_key_enabled: bool | None = None
    # ── 推理配置 ──
    thinking_effort: Literal["none", "minimal", "low", "medium", "high", "xhigh"] | None = None
    thinking_budget: int | None = Field(default=None, ge=0)
    # ── 子代理 ──
    subagent_max_iterations: int | None = Field(default=None, gt=0)
    subagent_timeout_seconds: int | None = Field(default=None, gt=0)
    subagent_max_consecutive_failures: int | None = Field(default=None, gt=0)
    parallel_subagent_max: int | None = Field(default=None, gt=0)
    # ── LLM 重试 ──
    llm_retry_max_attempts: int | None = Field(default=None, ge=1)
    llm_retry_base_delay_seconds: float | None = Field(default=None, gt=0)
    llm_retry_max_delay_seconds: float | None = Field(default=None, gt=0)
    # ── 感知与视觉 ──
    window_perception_enabled: bool | None = None
    window_perception_system_budget_tokens: int | None = Field(default=None, gt=0)
    window_perception_tool_append_tokens: int | None = Field(default=None, gt=0)
    window_perception_max_windows: int | None = Field(default=None, gt=0)
    window_perception_default_rows: int | None = Field(default=None, gt=0)
    window_perception_default_cols: int | None = Field(default=None, gt=0)
    window_perception_minimized_tokens: int | None = Field(default=None, gt=0)
    window_perception_background_after_idle: int | None = Field(default=None, gt=0)
    window_perception_suspend_after_idle: int | None = Field(default=None, gt=0)
    window_perception_terminate_after_idle: int | None = Field(default=None, gt=0)
    window_perception_advisor_mode: Literal["rules", "hybrid"] | None = None
    window_perception_advisor_timeout_ms: int | None = Field(default=None, gt=0)
    window_perception_advisor_trigger_window_count: int | None = Field(default=None, gt=0)
    window_perception_advisor_trigger_turn: int | None = Field(default=None, gt=0)
    window_perception_advisor_plan_ttl_turns: int | None = Field(default=None, gt=0)
    window_return_mode: Literal["unified", "anchored", "enriched", "adaptive"] | None = None
    window_full_max_rows: int | None = Field(default=None, gt=0)
    window_full_total_budget_tokens: int | None = Field(default=None, gt=0)
    window_data_buffer_max_rows: int | None = Field(default=None, gt=0)
    window_intent_enabled: bool | None = None
    window_intent_sticky_turns: int | None = Field(default=None, gt=0)
    window_intent_repeat_warn_threshold: int | None = Field(default=None, gt=0)
    window_intent_repeat_trip_threshold: int | None = Field(default=None, gt=0)
    window_rule_engine_version: Literal["v1", "v2"] | None = None
    vlm_enhance: bool | None = None
    main_model_vision: Literal["auto", "true", "false"] | None = None
    vlm_timeout_seconds: int | None = Field(default=None, gt=0)
    vlm_max_retries: int | None = Field(default=None, ge=0)
    vlm_max_tokens: int | None = Field(default=None, gt=0)
    vlm_image_max_long_edge: int | None = Field(default=None, gt=0)
    vlm_image_jpeg_quality: int | None = Field(default=None, ge=1, le=100)
    vlm_extraction_tier: Literal["auto", "strong", "standard", "weak"] | None = None
    image_keep_rounds: int | None = Field(default=None, gt=0)
    image_max_active: int | None = Field(default=None, gt=0)
    image_token_budget: int | None = Field(default=None, gt=0)
    # ── 系统消息与工具 ──
    system_message_mode: Literal["auto", "merge", "replace"] | None = None
    tool_result_hard_cap_chars: int | None = Field(default=None, ge=0)
    large_excel_threshold_bytes: int | None = Field(default=None, gt=0)
    parallel_readonly_tools: bool | None = None
    hooks_command_enabled: bool | None = None
    hooks_command_timeout_seconds: int | None = Field(default=None, gt=0)
    hooks_output_max_chars: int | None = Field(default=None, gt=0)
    log_level: Literal["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"] | None = None
    # ── 代码策略 ──
    code_policy_enabled: bool | None = None
    code_policy_green_auto_approve: bool | None = None
    code_policy_yellow_auto_approve: bool | None = None
    tool_schema_validation_mode: Literal["off", "shadow", "enforce"] | None = None
    tool_schema_validation_canary_percent: int | None = Field(default=None, ge=0, le=100)
    tool_schema_strict_path: bool | None = None
    # ── 技能发现 ──
    skills_context_char_budget: int | None = Field(default=None, ge=0)
    skills_discovery_enabled: bool | None = None
    skills_discovery_scan_workspace_ancestors: bool | None = None
    skills_discovery_include_agents: bool | None = None
    skills_discovery_scan_external_tool_dirs: bool | None = None
    # ── Embedding / 语义检索 ──
    embedding_enabled: bool | None = None
    embedding_model: str | None = None
    embedding_dimensions: int | None = Field(default=None, gt=0)
    embedding_timeout_seconds: float | None = Field(default=None, gt=0)
    memory_semantic_top_k: int | None = Field(default=None, gt=0)
    memory_semantic_threshold: float | None = Field(default=None, ge=0, le=1)
    memory_semantic_fallback_recent: int | None = Field(default=None, gt=0)
    # ── Playbook ──
    playbook_enabled: bool | None = None
    playbook_max_bullets: int | None = Field(default=None, gt=0)
    playbook_inject_top_k: int | None = Field(default=None, gt=0)
    registry_semantic_top_k: int | None = Field(default=None, gt=0)
    registry_semantic_threshold: float | None = Field(default=None, ge=0, le=1)


@_router.put("/api/v1/config/runtime")
async def update_runtime_config(request: RuntimeConfigUpdate, raw_request: Request) -> JSONResponse:
    """更新运行时行为配置并持久化到 .env。"""
    assert _config is not None, "服务未初始化"
    guard_error = await _require_admin_if_auth_enabled(raw_request)
    if guard_error is not None:
        return guard_error

    env_path = _find_env_file()
    lines = _read_env_file(env_path)
    updates: dict[str, str] = {}

    payload = request.model_dump(exclude_none=True)
    if not payload:
        return _error_json_response(400, "无有效更新字段")

    for field, value in payload.items():
        env_key = _RUNTIME_ENV_KEYS.get(field)
        if env_key is None:
            continue
        if isinstance(value, bool):
            str_val = "true" if value else "false"
        else:
            str_val = str(value)
        updates[env_key] = str_val
        lines = _update_env_var(lines, env_key, str_val)

    _write_env_file(env_path, lines)

    for env_key, env_val in updates.items():
        os.environ[env_key] = env_val

    # 同步更新内存中的 config 实例
    for field, value in payload.items():
        if hasattr(_config, field):
            object.__setattr__(_config, field, value)

    # 需要重启才能生效的配置项集合
    _RESTART_REQUIRED_KEYS = {
        "auth_enabled",
        "embedding_enabled",
        "deploy_mode",
        "mcp_shared_manager",
    }
    # 人类可读的重启原因映射
    _RESTART_REASON_MAP: dict[str, str] = {
        "auth_enabled": "认证配置已更新",
        "embedding_enabled": "语义检索配置已更新",
        "deploy_mode": "部署模式已更改",
        "mcp_shared_manager": "MCP 管理器配置已更改",
    }
    restart_keys = payload.keys() & _RESTART_REQUIRED_KEYS
    need_restart = len(restart_keys) > 0

    # 构建重启原因描述
    restart_reason = ""
    if need_restart:
        reasons = [_RESTART_REASON_MAP.get(k, k) for k in restart_keys]
        restart_reason = "、".join(reasons)

    resp = JSONResponse(content={
        "status": "ok",
        "updated": list(payload.keys()),
        "restarting": need_restart,
        "restart_reason": restart_reason,
    })
    if need_restart:
        global _restart_reason
        _restart_reason = restart_reason
        from starlette.background import BackgroundTask
        from excelmanus.restart import schedule_restart
        resp.background = BackgroundTask(schedule_restart)
    return resp


# ── MCP Server 管理 API（已提取到 api_routes_mcp.py）──────


@_router.get("/api/v1/mentions")
async def list_mentions(request: Request, path: str = "") -> JSONResponse:
    """返回 @ 提及可选项。path 参数支持子目录扫描。"""
    tools: list[str] = []
    skills: list[dict] = []
    files: list[str] = []

    if _tool_registry is not None:
        tools = sorted(_tool_registry.get_tool_names())
    if _skillpack_loader is not None:
        for name, sp in _skillpack_loader.get_skillpacks().items():
            skills.append({
                "name": name,
                "description": sp.get("description", "") if isinstance(sp, dict) else "",
            })
    safe_path = path.replace("..", "").strip("/")
    if _config is not None:
        ws = _resolve_workspace_root(request)
        scan_dir = os.path.join(ws, safe_path) if safe_path else ws
        if os.path.isdir(scan_dir):
            try:
                for entry in os.scandir(scan_dir):
                    if entry.name.startswith(".") or entry.name in {"node_modules", "__pycache__", ".venv"}:
                        continue
                    rel = f"{safe_path}/{entry.name}" if safe_path else entry.name
                    if entry.is_dir():
                        files.append(rel + "/")
                    elif entry.is_file():
                        files.append(rel)
            except OSError:
                pass

    return JSONResponse(content={
        "tools": tools,
        "skills": [{"name": s["name"], "description": s.get("description", "")} for s in skills],
        "files": sorted(files),
        "path": safe_path if _config is not None else "",
    })


@_router.post("/api/v1/command")
async def execute_command(request: Request) -> JSONResponse:
    """执行展示型斜杠命令并返回结果（不走 chat 流）。"""
    body = await request.json()
    command = (body.get("command") or "").strip()
    if not command:
        return _error_json_response(400, "缺少 command 字段")

    # /help → 返回命令列表
    if command == "/help":
        from excelmanus.control_commands import CONTROL_COMMAND_SPECS
        lines = ["### ExcelManus 命令帮助\n"]
        lines.append("| 命令 | 说明 |")
        lines.append("|------|------|")
        base = [
            ("/help", "显示帮助"), ("/skills", "查看技能包"), ("/history", "对话历史摘要"),
            ("/clear", "清除对话历史"), ("/mcp", "MCP Server 状态"), ("/save", "保存对话记录"),
            ("/config", "环境变量配置"),
        ]
        for cmd, desc in base:
            lines.append(f"| `{cmd}` | {desc} |")
        for spec in CONTROL_COMMAND_SPECS:
            args = f" ({', '.join(spec.arguments)})" if spec.arguments else ""
            lines.append(f"| `{spec.command}{args}` | {spec.description} |")
        return JSONResponse(content={"result": "\n".join(lines), "format": "markdown"})

    # /skills → 返回技能包列表
    if command == "/skills":
        if _skillpack_loader is None:
            return JSONResponse(content={"result": "技能包未加载", "format": "text"})
        lines = ["### 已加载技能包\n"]
        for sp in _skillpack_loader.list_skillpacks():
            name = sp.name if hasattr(sp, "name") else str(sp.get("name", ""))
            desc = sp.description if hasattr(sp, "description") else str(sp.get("description", ""))
            source = sp.source if hasattr(sp, "source") else str(sp.get("source", ""))
            lines.append(f"- **{name}** ({source}) — {desc}")
        return JSONResponse(content={"result": "\n".join(lines), "format": "markdown"})

    # /history → 对话历史摘要
    if command == "/history":
        if _session_manager is None:
            return JSONResponse(content={"result": "服务未初始化", "format": "text"})
        try:
            sessions = await _session_manager.list_sessions()
            if not sessions:
                return JSONResponse(content={"result": "暂无活跃会话", "format": "text"})
            lines = ["### 活跃会话\n"]
            for s in sessions:
                status = "🔄" if s.get("in_flight") else "💬"
                lines.append(f"- {status} **{s['title']}** ({s['message_count']} 条消息) `{s['id'][:8]}...`")
            return JSONResponse(content={"result": "\n".join(lines), "format": "markdown"})
        except Exception:
            return JSONResponse(content={"result": "获取会话历史失败", "format": "text"})

    # /mcp → 返回 MCP 状态
    if command == "/mcp":
        if _session_manager is None:
            return JSONResponse(content={"result": "服务未初始化", "format": "text"})
        return JSONResponse(content={"result": "MCP 状态请通过设置面板查看", "format": "text"})

    # /config → 返回配置摘要
    if command in {"/config", "/config list", "/config get"}:
        if _config is None:
            return JSONResponse(content={"result": "配置未加载", "format": "text"})
        lines = ["### 当前配置\n"]
        lines.append(f"- **模型**: `{_config.model}`")
        lines.append(f"- **Base URL**: `{_config.base_url}`")
        lines.append(f"- **工作区**: `{_config.workspace_root}`")
        lines.append(f"- **最大迭代**: {_config.max_iterations}")
        lines.append(f"- **辅助模型**: `{_config.aux_model or '未配置'}`")
        lines.append(f"- **AUX Base URL**: `{_config.aux_base_url or '继承主配置'}`")
        lines.append(f"- **VLM 模型**: `{_config.vlm_model or '未配置'}`")
        lines.append(f"- **子代理**: {'开启' if _config.subagent_enabled else '关闭'}")
        lines.append(f"- **备份模式**: {'开启' if _config.backup_enabled else '关闭'}")
        lines.append(f"- **安全模式**: {'开启' if _config.external_safe_mode else '关闭'}")
        lines.append(f"- **多模型配置**: {len(_config.models)} 个")
        return JSONResponse(content={"result": "\n".join(lines), "format": "markdown"})

    # /model, /model list → 返回模型列表
    if command in {"/model", "/model list"}:
        try:
            assert _config is not None
            lines = ["### 可用模型\n"]
            lines.append(f"- **default** → `{_config.model}` — 默认模型（主配置） ✦")
            for profile in _config.models:
                desc = f" — {profile.description}" if profile.description else ""
                lines.append(f"- **{profile.name}** → `{profile.model}`{desc}")
            return JSONResponse(content={"result": "\n".join(lines), "format": "markdown"})
        except Exception:
            return JSONResponse(content={"result": "模型列表获取失败", "format": "text"})

    # /subagent list, /subagent status
    if command in {"/subagent list", "/subagent status"}:
        assert _config is not None
        status = "开启" if _config.subagent_enabled else "关闭"
        return JSONResponse(content={"result": f"子代理状态: **{status}**\n\n最大迭代: {_config.subagent_max_iterations}", "format": "markdown"})

    # /backup list, /backup status
    if command in {"/backup list", "/backup status"}:
        assert _config is not None
        status = "开启" if _config.backup_enabled else "关闭"
        return JSONResponse(content={"result": f"备份沙盒: **{status}**", "format": "markdown"})

    # /compact status
    if command == "/compact status":
        assert _config is not None
        status = "开启" if _config.compaction_enabled else "关闭"
        return JSONResponse(content={"result": f"上下文压缩: **{status}**\n\n阈值: {_config.compaction_threshold_ratio}", "format": "markdown"})

    # /fullaccess status
    if command == "/fullaccess status":
        # fullaccess 已改为跨会话持久化设置
        hint = "全权限模式: **关闭**\n\n使用 `/fullaccess on` 开启，开启后工具调用将跳过审批确认（跨会话生效）"
        if _database is not None:
            try:
                from excelmanus.stores.config_store import UserConfigStore
                _uc = UserConfigStore(_database.conn, user_id=_get_isolation_user_id(request))
                if _uc.get_full_access():
                    hint = "全权限模式: **开启**（跨会话生效）\n\n使用 `/fullaccess off` 关闭"
            except Exception:
                pass
        elif _session_manager is not None:
            try:
                sessions = await _session_manager.list_sessions()
                for s in sessions:
                    detail = await _session_manager.get_session_detail(s["id"])
                    if detail.get("full_access_enabled"):
                        hint = f"全权限模式: **开启**\n\n使用 `/fullaccess off` 关闭"
                        break
            except Exception:
                pass
        return JSONResponse(content={"result": hint, "format": "markdown"})

    # /plan status
    if command == "/plan status":
        hint = "计划模式: **关闭**（默认）\n\n使用 `/plan on` 开启，开启后 Agent 会先输出计划再执行"
        return JSONResponse(content={"result": hint, "format": "markdown"})

    # /registry status
    if command == "/registry status":
        return JSONResponse(content={"result": "文件注册表: 请通过 `/registry scan` 触发扫描\n\n注册表会在会话首轮自动扫描构建", "format": "markdown"})

    # /save
    if command == "/save":
        if _session_manager is None:
            return JSONResponse(content={"result": "服务未初始化", "format": "text"})
        try:
            sessions = await _session_manager.list_sessions()
            count = len(sessions)
            return JSONResponse(content={"result": f"当前有 **{count}** 个活跃会话\n\n网页端对话自动保存，无需手动操作", "format": "markdown"})
        except Exception:
            return JSONResponse(content={"result": "会话状态获取失败", "format": "text"})

    # ── 会话级命令：需要 engine 实例 ──
    # 当前端传入 session_id 时，委托给 command_handler 处理会话级控制命令
    # （如 /fullaccess on, /subagent off, /compact, /rules, /memory, /playbook 等）
    session_id = body.get("session_id") or ""
    if session_id and _session_manager is not None:
        user_id = _get_isolation_user_id(request)
        engine = _session_manager.get_engine(session_id, user_id=user_id)
        if engine is not None:
            try:
                result = await engine._command_handler.handle(command)
                if result is not None:
                    # 判断是否包含 markdown 格式
                    fmt = "markdown" if any(c in result for c in ("**", "##", "`", "- ")) else "text"
                    return JSONResponse(content={"result": result, "format": fmt})
            except Exception as exc:
                logger.warning("command_handler 执行 '%s' 异常: %s", command, exc, exc_info=True)
                return JSONResponse(content={"result": f"命令执行失败: {exc}", "format": "text"})

    return JSONResponse(content={"result": f"未知命令: {command}", "format": "text"})


# ── Rules / Memory API（已提取到 api_routes_rules.py）──────


@_router.get("/api/v1/health")
async def health(request: Request) -> dict:
    """健康检查：返回版本号和已加载的工具/技能包。"""
    if _draining:
        return {
            "status": "draining",
            "version": excelmanus.__version__,
            "restart_reason": _restart_reason,
        }

    if _is_external_safe_mode():
        return {
            "status": "ok",
            "version": excelmanus.__version__,
            "configured": not _config_incomplete,
            "model": "hidden",
            "tools": [],
            "skillpacks": [],
        }

    tools: list[str] = []
    skillpacks: list[str] = []
    if _tool_registry is not None:
        tools = sorted(_tool_registry.get_tool_names())
    if _skillpack_loader is not None:
        skillpacks = sorted(_skillpack_loader.get_skillpacks().keys())

    active_sessions = 0
    if _session_manager is not None:
        active_sessions = await _session_manager.get_active_count()

    auth_enabled = os.environ.get("EXCELMANUS_AUTH_ENABLED", "").strip().lower() in ("1", "true", "yes")

    # 登录方式配置（供前端登录/注册页动态渲染）
    login_methods = {
        "github_enabled": True,
        "google_enabled": True,
        "qq_enabled": False,
        "email_verify_required": False,
        "require_agreement": True,
    }
    if auth_enabled:
        try:
            from excelmanus.auth.router import get_login_config
            lc = get_login_config(request)
            login_methods["github_enabled"] = lc.get("login_github_enabled", True)
            login_methods["google_enabled"] = lc.get("login_google_enabled", True)
            login_methods["qq_enabled"] = lc.get("login_qq_enabled", False)
            login_methods["email_verify_required"] = lc.get("email_verify_required", False)
            login_methods["require_agreement"] = lc.get("require_agreement", True)
        except Exception:
            pass

    # 发布清单摘要（供前端版本轮询使用）
    from excelmanus.api_routes_version import get_manifest_data, _API_SCHEMA_VERSION
    _manifest = get_manifest_data()

    return {
        "status": "ok",
        "version": excelmanus.__version__,
        "configured": not _config_incomplete,
        "model": _config.model if _config is not None else "",
        "tools": tools,
        "skillpacks": skillpacks,
        "active_sessions": active_sessions,
        "auth_enabled": auth_enabled,
        "login_methods": login_methods,
        "session_isolation_enabled": getattr(request.app.state, "session_isolation_enabled", False),
        "docker_sandbox_enabled": getattr(request.app.state, "docker_sandbox_enabled", False),
        "build_id": _manifest.get("frontend_build_id"),
        "version_fingerprint": _manifest.get("version_fingerprint"),
        "api_schema_version": _API_SCHEMA_VERSION,
        "git_commit": _manifest.get("git_commit"),
        "min_frontend_build_id": _manifest.get("min_frontend_build_id"),
        "min_backend_version": _manifest.get("min_backend_version"),
        "deploy_mode": _config.deploy_mode if _config is not None else "standalone",
        "channels": _channel_launcher.active_channels if _channel_launcher is not None else [],
    }


@_router.get("/api/v1/server/public-ip")
async def server_public_ip() -> JSONResponse:
    """检测服务器的公网 IP 地址。"""
    import httpx

    _IP_SERVICES = [
        "https://api.ipify.org",
        "https://icanhazip.com",
        "https://checkip.amazonaws.com",
    ]
    async with httpx.AsyncClient(timeout=5) as client:
        for url in _IP_SERVICES:
            try:
                resp = await client.get(url)
                if resp.status_code == 200:
                    ip = resp.text.strip()
                    if ip:
                        return JSONResponse({"ip": ip})
            except Exception:
                continue
    return JSONResponse({"ip": None, "error": "无法检测公网 IP"}, status_code=503)


@_router.get("/api/v1/channels")
async def channels_status() -> dict:
    """查询渠道协同启动状态（含每个渠道的详细状态和配置信息）。"""
    from excelmanus.channels.config_store import CHANNEL_CREDENTIAL_FIELDS
    from excelmanus.channels.launcher import ChannelLauncher, _CHANNEL_BUILDERS

    statuses: dict[str, str] = {}
    if _channel_launcher is not None:
        statuses = _channel_launcher.all_channel_status()

    # 加载持久化配置
    saved_configs: dict = {}
    if _config_store is not None:
        try:
            from excelmanus.channels.config_store import ChannelConfigStore
            ccs = ChannelConfigStore(_config_store)
            for name, cfg in ccs.load_all().items():
                # 脱敏：secret 字段仅返回是否已填
                masked_creds: dict[str, str] = {}
                fields = CHANNEL_CREDENTIAL_FIELDS.get(name, [])
                secret_keys = {f["key"] for f in fields if f.get("secret")}
                for k, v in cfg.credentials.items():
                    if k in secret_keys and v:
                        masked_creds[k] = (v[:4] + "••••••••" + v[-4:]) if len(v) > 12 else "••••••••"
                    else:
                        masked_creds[k] = v
                saved_configs[name] = {
                    "enabled": cfg.enabled,
                    "credentials": masked_creds,
                    "has_required": cfg.has_required_credentials(),
                    "missing_fields": cfg.get_missing_fields(),
                    "updated_at": cfg.updated_at,
                }
        except Exception:
            logger.debug("加载渠道配置失败", exc_info=True)

    # 构建每个渠道的完整信息
    all_channels = sorted(set(list(_CHANNEL_BUILDERS.keys()) + list(saved_configs.keys())))
    channel_details = []
    for ch in all_channels:
        dep_ok, dep_hint = ChannelLauncher.check_dependency(ch)
        detail: dict = {
            "name": ch,
            "status": statuses.get(ch, "stopped"),
            "supported": ch in _CHANNEL_BUILDERS,
            "fields": CHANNEL_CREDENTIAL_FIELDS.get(ch, []),
            "dep_installed": dep_ok,
            "install_hint": dep_hint,
        }
        if ch in saved_configs:
            detail.update(saved_configs[ch])
        else:
            detail["enabled"] = False
            detail["credentials"] = {}
            detail["has_required"] = False
            detail["missing_fields"] = [
                f["key"] for f in CHANNEL_CREDENTIAL_FIELDS.get(ch, []) if f.get("required")
            ]
        channel_details.append(detail)

    # 读取 require_bind 状态（env var > config_kv > false）
    env_rb = os.environ.get("EXCELMANUS_CHANNEL_REQUIRE_BIND", "").strip().lower()
    if env_rb:
        require_bind = env_rb in ("1", "true", "yes")
        require_bind_source = "env"
    elif _config_store is not None:
        db_rb = _config_store.get("channel_require_bind", "")
        require_bind = db_rb.strip().lower() in ("1", "true", "yes") if db_rb else False
        require_bind_source = "config"
    else:
        require_bind = False
        require_bind_source = "default"

    # 读取速率限制配置（env > DB > defaults）
    from excelmanus.channels.rate_limit import RateLimitConfig
    rl_cfg = RateLimitConfig.from_store(_config_store)
    rl_env_overrides = RateLimitConfig.env_overrides()

    # ── 读取扩展渠道设置 ──

    def _read_setting(db_key: str, env_key: str, default: str = "") -> tuple[str, str]:
        """返回 (value, source)。source: 'env' | 'config' | 'default'。"""
        env_val = os.environ.get(env_key, "").strip()
        if env_val:
            return env_val, "env"
        if _config_store is not None:
            db_val = _config_store.get(db_key, "")
            if db_val:
                return db_val.strip(), "config"
        return default, "default"

    # 访问控制
    admin_users_val, admin_users_src = _read_setting(
        "channel_admin_users", "EXCELMANUS_CHANNEL_ADMINS",
    )
    group_policy_val, group_policy_src = _read_setting(
        "channel_group_policy", "EXCELMANUS_CHANNEL_GROUP_POLICY", "auto",
    )
    group_whitelist_val, _ = _read_setting(
        "channel_group_whitelist", "", "",
    )
    group_blacklist_val, _ = _read_setting(
        "channel_group_blacklist", "", "",
    )
    allowed_users_val, _ = _read_setting(
        "channel_allowed_users", "", "",
    )

    # 行为设置
    default_concurrency_val, default_concurrency_src = _read_setting(
        "channel_default_concurrency", "EXCELMANUS_CHANNEL_DEFAULT_CONCURRENCY", "queue",
    )
    default_chat_mode_val, default_chat_mode_src = _read_setting(
        "channel_default_chat_mode", "EXCELMANUS_CHANNEL_DEFAULT_CHAT_MODE", "write",
    )
    public_url_val, public_url_src = _read_setting(
        "channel_public_url", "EXCELMANUS_PUBLIC_URL", "",
    )

    # 输出调优
    tg_edit_interval_min_val, _ = _read_setting("channel_tg_edit_interval_min", "", "1.5")
    tg_edit_interval_max_val, _ = _read_setting("channel_tg_edit_interval_max", "", "3.0")
    qq_progressive_chars_val, _ = _read_setting("channel_qq_progressive_chars", "", "200")
    qq_progressive_interval_val, _ = _read_setting("channel_qq_progressive_interval", "", "3.0")
    feishu_update_interval_val, _ = _read_setting("channel_feishu_update_interval", "", "0.5")

    # 构建 env_overrides 映射（前端用于判断锁定状态）
    settings_env_overrides: dict[str, str] = {}
    if admin_users_src == "env":
        settings_env_overrides["admin_users"] = "EXCELMANUS_CHANNEL_ADMINS"
    if group_policy_src == "env":
        settings_env_overrides["group_policy"] = "EXCELMANUS_CHANNEL_GROUP_POLICY"
    if default_concurrency_src == "env":
        settings_env_overrides["default_concurrency"] = "EXCELMANUS_CHANNEL_DEFAULT_CONCURRENCY"
    if default_chat_mode_src == "env":
        settings_env_overrides["default_chat_mode"] = "EXCELMANUS_CHANNEL_DEFAULT_CHAT_MODE"
    if public_url_src == "env":
        settings_env_overrides["public_url"] = "EXCELMANUS_PUBLIC_URL"

    return {
        "enabled": _channel_launcher is not None,
        "channels": _channel_launcher.active_channels if _channel_launcher is not None else [],
        "details": channel_details,
        "require_bind": require_bind,
        "require_bind_source": require_bind_source,
        "rate_limit": rl_cfg.to_dict(),
        "rate_limit_env_overrides": rl_env_overrides,
        # 扩展设置
        "settings": {
            "admin_users": admin_users_val,
            "group_policy": group_policy_val,
            "group_whitelist": group_whitelist_val,
            "group_blacklist": group_blacklist_val,
            "allowed_users": allowed_users_val,
            "default_concurrency": default_concurrency_val,
            "default_chat_mode": default_chat_mode_val,
            "public_url": public_url_val,
            "tg_edit_interval_min": tg_edit_interval_min_val,
            "tg_edit_interval_max": tg_edit_interval_max_val,
            "qq_progressive_chars": qq_progressive_chars_val,
            "qq_progressive_interval": qq_progressive_interval_val,
            "feishu_update_interval": feishu_update_interval_val,
        },
        "settings_env_overrides": settings_env_overrides,
    }


@_router.put("/api/v1/channels/settings")
async def update_channel_settings(request: Request) -> JSONResponse:
    """更新渠道全局设置（管理员）。

    请求体支持以下字段（均可选，仅传入需修改的）：
    - require_bind: bool — 强制绑定前端账号
    - admin_users: str — 管理员用户 ID（逗号分隔）
    - group_policy: str — 群聊策略 deny/allow/whitelist/blacklist/auto
    - group_whitelist: str — 群白名单 JSON 数组
    - group_blacklist: str — 群黑名单 JSON 数组
    - allowed_users: str — 允许用户 JSON 数组
    - default_concurrency: str — 默认并发模式 queue/steer/guide
    - default_chat_mode: str — 默认聊天模式 write/read/plan
    - public_url: str — 公开访问 URL
    - tg_edit_interval_min: str — Telegram 编辑间隔最小值
    - tg_edit_interval_max: str — Telegram 编辑间隔最大值
    - qq_progressive_chars: str — QQ 渐进发送字符阈值
    - qq_progressive_interval: str — QQ 渐进发送间隔
    - feishu_update_interval: str — 飞书卡片更新间隔
    """
    if _config_store is None:
        return _error_json_response(503, "数据库未初始化。")

    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    body = await request.json()
    updated: list[str] = []
    locked: list[str] = []

    # ── require_bind ──
    if "require_bind" in body:
        env_rb = os.environ.get("EXCELMANUS_CHANNEL_REQUIRE_BIND", "").strip()
        if env_rb:
            locked.append("require_bind")
        else:
            val = "true" if body["require_bind"] else "false"
            _config_store.set("channel_require_bind", val)
            updated.append("require_bind")
            logger.info("渠道强制绑定设置已更新: %s", val)

    # ── 访问控制字段 ──

    # 检查 env 锁定的辅助函数
    _ENV_LOCK_MAP: dict[str, str] = {
        "admin_users": "EXCELMANUS_CHANNEL_ADMINS",
        "group_policy": "EXCELMANUS_CHANNEL_GROUP_POLICY",
        "default_concurrency": "EXCELMANUS_CHANNEL_DEFAULT_CONCURRENCY",
        "default_chat_mode": "EXCELMANUS_CHANNEL_DEFAULT_CHAT_MODE",
        "public_url": "EXCELMANUS_PUBLIC_URL",
    }
    _DB_KEY_MAP: dict[str, str] = {
        "admin_users": "channel_admin_users",
        "group_policy": "channel_group_policy",
        "group_whitelist": "channel_group_whitelist",
        "group_blacklist": "channel_group_blacklist",
        "allowed_users": "channel_allowed_users",
        "default_concurrency": "channel_default_concurrency",
        "default_chat_mode": "channel_default_chat_mode",
        "public_url": "channel_public_url",
        "tg_edit_interval_min": "channel_tg_edit_interval_min",
        "tg_edit_interval_max": "channel_tg_edit_interval_max",
        "qq_progressive_chars": "channel_qq_progressive_chars",
        "qq_progressive_interval": "channel_qq_progressive_interval",
        "feishu_update_interval": "channel_feishu_update_interval",
    }

    # 验证枚举值
    _VALID_GROUP_POLICIES = {"deny", "allow", "whitelist", "blacklist", "auto"}
    _VALID_CONCURRENCY = {"queue", "steer", "guide"}
    _VALID_CHAT_MODES = {"write", "read", "plan"}

    for field, db_key in _DB_KEY_MAP.items():
        if field not in body:
            continue

        # 检查 env 锁定
        env_var = _ENV_LOCK_MAP.get(field, "")
        if env_var and os.environ.get(env_var, "").strip():
            locked.append(field)
            continue

        value = body[field]

        # 枚举验证
        if field == "group_policy" and value not in _VALID_GROUP_POLICIES:
            return _error_json_response(400, f"无效的群聊策略: {value}，可选值: {', '.join(_VALID_GROUP_POLICIES)}")
        if field == "default_concurrency" and value not in _VALID_CONCURRENCY:
            return _error_json_response(400, f"无效的并发模式: {value}，可选值: {', '.join(_VALID_CONCURRENCY)}")
        if field == "default_chat_mode" and value not in _VALID_CHAT_MODES:
            return _error_json_response(400, f"无效的聊天模式: {value}，可选值: {', '.join(_VALID_CHAT_MODES)}")

        # 数值验证
        if field in ("tg_edit_interval_min", "tg_edit_interval_max", "qq_progressive_interval", "feishu_update_interval"):
            try:
                fval = float(value)
                if fval < 0.1 or fval > 60.0:
                    return _error_json_response(400, f"字段 {field} 超出合理范围 (0.1-60.0): {value}")
                value = str(fval)
            except (ValueError, TypeError):
                return _error_json_response(400, f"字段 {field} 必须为数字: {value}")
        if field == "qq_progressive_chars":
            try:
                ival = int(value)
                if ival < 50 or ival > 5000:
                    return _error_json_response(400, f"字段 {field} 超出合理范围 (50-5000): {value}")
                value = str(ival)
            except (ValueError, TypeError):
                return _error_json_response(400, f"字段 {field} 必须为整数: {value}")

        _config_store.set(db_key, str(value))
        updated.append(field)

    if updated:
        logger.info("渠道设置已更新: %s", updated)

    # 热更新运行中的 handler
    if _channel_launcher is not None:
        _propagate_channel_settings()

    result: dict = {"status": "ok", "updated_fields": updated}
    if locked:
        result["locked_fields"] = locked
        result["message"] = f"以下字段被环境变量锁定: {', '.join(locked)}"
    return JSONResponse(result)


@_router.put("/api/v1/channels/rate-limit")
async def update_rate_limit_settings(request: Request) -> JSONResponse:
    """更新渠道速率限制配置（管理员）。

    请求体: {"chat_per_minute": 5, "chat_per_hour": 30, ...}
    仅传入需要修改的字段，未传入的保持不变。
    环境变量锁定的字段不可通过此接口修改。
    """
    if _config_store is None:
        return _error_json_response(503, "数据库未初始化。")

    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    from excelmanus.channels.rate_limit import RateLimitConfig

    body = await request.json()

    # 加载当前持久化配置作为 base
    current = RateLimitConfig.from_store(_config_store)
    current_dict = current.to_dict()

    # 环境变量锁定的字段不允许修改
    env_overrides = RateLimitConfig.env_overrides()
    locked_fields = [k for k in body if k in env_overrides]

    # 合并：仅更新传入的非锁定字段
    valid_fields = set(current_dict.keys())
    updated_fields: list[str] = []
    for key, value in body.items():
        if key not in valid_fields:
            continue
        if key in env_overrides:
            continue  # 被环境变量锁定，跳过
        try:
            if key in ("reject_cooldown_seconds", "auto_ban_duration_seconds"):
                current_dict[key] = float(value)
            else:
                current_dict[key] = int(value)
            updated_fields.append(key)
        except (ValueError, TypeError):
            return _error_json_response(400, f"字段 {key} 的值无效: {value}")

    # 保存到 DB
    new_cfg = RateLimitConfig.from_dict(current_dict)
    RateLimitConfig.save_to_store(new_cfg, _config_store)
    logger.info("速率限制配置已更新: %s", updated_fields)

    # 动态更新运行中的渠道 Bot 的限流器
    if _channel_launcher is not None:
        _propagate_rate_limit_config(new_cfg)

    result: dict = {"status": "ok", "updated_fields": updated_fields}
    if locked_fields:
        result["locked_fields"] = locked_fields
        result["message"] = f"以下字段被环境变量锁定，未修改: {', '.join(locked_fields)}"
    return JSONResponse(content=result)


def _propagate_rate_limit_config(cfg: "RateLimitConfig") -> None:
    """将新的速率限制配置传播到所有运行中的渠道 Bot。"""
    if _channel_launcher is None:
        return
    for name, handler in _channel_launcher._handlers.items():
        try:
            if hasattr(handler, "_rate_limiter"):
                handler._rate_limiter.config = cfg
                logger.debug("渠道 %s 速率限制配置已热更新", name)
        except Exception:
            logger.debug("渠道 %s 速率限制配置热更新失败", name, exc_info=True)


def _propagate_channel_settings() -> None:
    """将渠道扩展设置传播到所有运行中的 MessageHandler。

    读取 config_store 中的最新值并更新 handler 内部状态。
    仅处理可安全热更新的字段（default_concurrency 等）。
    """
    if _channel_launcher is None or _config_store is None:
        return
    for name, handler in _channel_launcher._handlers.items():
        try:
            # 默认并发模式
            dc = _config_store.get("channel_default_concurrency", "")
            if dc and dc in ("queue", "steer", "guide"):
                handler._default_concurrency = dc

            logger.debug("渠道 %s 扩展设置已热更新", name)
        except Exception:
            logger.debug("渠道 %s 扩展设置热更新失败", name, exc_info=True)


@_router.put("/api/v1/channels/{channel_name}/config")
async def save_channel_config(channel_name: str, request: Request) -> JSONResponse:
    """保存单个渠道的配置（凭证 + 启用状态）。

    请求体: {"credentials": {"token": "xxx", ...}, "enabled": true}
    """
    if _config_store is None:
        return _error_json_response(503, "数据库未初始化，无法保存渠道配置。")

    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    from excelmanus.channels.config_store import (
        CHANNEL_CREDENTIAL_FIELDS,
        ChannelConfig,
        ChannelConfigStore,
    )

    if channel_name not in CHANNEL_CREDENTIAL_FIELDS:
        return _error_json_response(400, f"不支持的渠道: {channel_name}")

    body = await request.json()
    credentials = body.get("credentials", {})
    enabled = body.get("enabled", False)

    # 合并：如果前端传来 "••••••••" 占位符，保留原值
    ccs = ChannelConfigStore(_config_store)
    existing = ccs.get(channel_name)
    if existing:
        fields = CHANNEL_CREDENTIAL_FIELDS.get(channel_name, [])
        secret_keys = {f["key"] for f in fields if f.get("secret")}
        for k in secret_keys:
            if "••••••••" in (credentials.get(k) or "") and existing.credentials.get(k):
                credentials[k] = existing.credentials[k]

    cfg = ChannelConfig(
        name=channel_name,
        enabled=bool(enabled),
        credentials=credentials,
    )
    ccs.save(cfg)

    logger.info("渠道 %s 配置已保存 (enabled=%s)", channel_name, enabled)
    return JSONResponse(content={
        "status": "ok",
        "channel": channel_name,
        "enabled": enabled,
        "has_required": cfg.has_required_credentials(),
        "missing_fields": cfg.get_missing_fields(),
    })


@_router.delete("/api/v1/channels/{channel_name}/config")
async def delete_channel_config(channel_name: str, request: Request) -> JSONResponse:
    """删除单个渠道的持久化配置。"""
    if _config_store is None:
        return _error_json_response(503, "数据库未初始化。")

    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    from excelmanus.channels.config_store import ChannelConfigStore
    ccs = ChannelConfigStore(_config_store)
    deleted = ccs.delete(channel_name)
    if not deleted:
        return _error_json_response(404, f"渠道 {channel_name} 无保存的配置。")

    logger.info("渠道 %s 配置已删除", channel_name)
    return JSONResponse(content={"status": "ok", "channel": channel_name})


@_router.post("/api/v1/channels/{channel_name}/start")
async def start_channel(channel_name: str, request: Request) -> JSONResponse:
    """热启动单个渠道 Bot。

    使用已保存的持久化配置凭证启动，环境变量作为 fallback。
    """
    global _channel_launcher

    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    # 确保 launcher 存在
    if _channel_launcher is None:
        from excelmanus.channels.launcher import ChannelLauncher
        api_port = int(os.environ.get("EXCELMANUS_API_PORT", "8000"))
        _bind_mgr = getattr(request.app.state, "bind_manager", None)
        _svc_token = getattr(request.app.state, "service_token", None)
        _evt_bridge = getattr(request.app.state, "event_bridge", None)
        _channel_launcher = ChannelLauncher(
            [],
            api_port=api_port,
            bind_manager=_bind_mgr,
            service_token=_svc_token,
            event_bridge=_evt_bridge,
        )

    # 从持久化配置加载凭证
    credentials: dict[str, str] | None = None
    if _config_store is not None:
        try:
            from excelmanus.channels.config_store import ChannelConfigStore
            ccs = ChannelConfigStore(_config_store)
            cfg = ccs.get(channel_name)
            if cfg and cfg.credentials:
                credentials = cfg.credentials
                logger.debug(
                    "渠道 %s 加载到持久化凭证，字段: %s",
                    channel_name, list(credentials.keys()),
                )
            else:
                logger.warning(
                    "渠道 %s 无持久化凭证（cfg=%s），将依赖环境变量",
                    channel_name, cfg is not None,
                )
        except Exception:
            logger.debug("加载渠道 %s 持久化凭证失败", channel_name, exc_info=True)
    else:
        logger.warning("config_store 未初始化，渠道 %s 将仅使用环境变量凭证", channel_name)

    ok, msg = await _channel_launcher.start_channel(channel_name, credentials=credentials)
    if ok:
        return JSONResponse(content={"status": "ok", "message": msg})
    return _error_json_response(400, msg)


@_router.post("/api/v1/channels/{channel_name}/stop")
async def stop_channel(channel_name: str, request: Request) -> JSONResponse:
    """热停止单个渠道 Bot。"""
    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    if _channel_launcher is None:
        return _error_json_response(400, f"渠道 {channel_name} 未启动")

    ok, msg = await _channel_launcher.stop_channel(channel_name)
    if ok:
        return JSONResponse(content={"status": "ok", "message": msg})
    return _error_json_response(400, msg)


@_router.post("/api/v1/channels/{channel_name}/test")
async def test_channel_config(channel_name: str, request: Request) -> JSONResponse:
    """测试渠道凭证是否有效（不启动 Bot）。

    请求体（可选）: {"credentials": {"token": "xxx"}}
    如不传则使用已保存的配置。
    """
    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    from excelmanus.channels.config_store import CHANNEL_CREDENTIAL_FIELDS

    if channel_name not in CHANNEL_CREDENTIAL_FIELDS:
        return _error_json_response(400, f"不支持的渠道: {channel_name}")

    # 获取凭证（请求体 > 持久化配置 > 环境变量）
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass

    credentials = body.get("credentials", {})

    # 如果请求未传凭证，从持久化配置加载
    if not credentials and _config_store is not None:
        try:
            from excelmanus.channels.config_store import ChannelConfigStore
            ccs = ChannelConfigStore(_config_store)
            cfg = ccs.get(channel_name)
            if cfg:
                credentials = cfg.credentials
        except Exception:
            pass

    # 合并：如果前端传来 "••••••••" 占位符，从持久化配置还原原值
    if credentials and _config_store is not None:
        try:
            from excelmanus.channels.config_store import ChannelConfigStore
            ccs = ChannelConfigStore(_config_store)
            existing = ccs.get(channel_name)
            if existing:
                fields = CHANNEL_CREDENTIAL_FIELDS.get(channel_name, [])
                secret_keys = {f["key"] for f in fields if f.get("secret")}
                for k in secret_keys:
                    if "••••••••" in (credentials.get(k) or "") and existing.credentials.get(k):
                        credentials[k] = existing.credentials[k]
        except Exception:
            pass

    # 执行平台特定的凭证验证
    if channel_name == "telegram":
        token = credentials.get("token") or os.environ.get("EXCELMANUS_TG_TOKEN", "")
        if not token or "••••••••" in token:
            return _error_json_response(400, "未提供 Telegram Bot Token。")
        try:
            import httpx
            async with httpx.AsyncClient(timeout=10, trust_env=True) as client:
                resp = await client.get(f"https://api.telegram.org/bot{token}/getMe")
                data = resp.json()
                if data.get("ok"):
                    bot_info = data.get("result", {})
                    return JSONResponse(content={
                        "status": "ok",
                        "message": f"Token 有效！Bot: @{bot_info.get('username', '?')} ({bot_info.get('first_name', '')})",
                        "bot_info": {
                            "username": bot_info.get("username"),
                            "name": bot_info.get("first_name"),
                        },
                    })
                return _error_json_response(400, f"Token 无效: {data.get('description', '未知错误')}")
        except Exception as e:
            hint = "（提示：如需代理访问 Telegram，请设置 HTTPS_PROXY 环境变量）"
            return _error_json_response(502, f"连接 Telegram API 失败: {e} {hint}")

    elif channel_name == "qq":
        app_id = credentials.get("app_id") or os.environ.get("EXCELMANUS_QQ_APPID", "")
        secret = credentials.get("secret") or os.environ.get("EXCELMANUS_QQ_SECRET", "")
        if not app_id or not secret or "••••••••" in secret:
            return _error_json_response(400, "未提供 QQ Bot AppID 或 AppSecret。")
        # QQ Bot 的鉴权较复杂（WebSocket），这里只做基础格式验证
        try:
            int(app_id)
        except ValueError:
            return _error_json_response(400, "AppID 格式无效（应为数字）。")
        return JSONResponse(content={
            "status": "ok",
            "message": f"凭证格式验证通过 (AppID: {app_id})。完整验证需启动 Bot。",
        })

    elif channel_name == "feishu":
        app_id = credentials.get("app_id") or os.environ.get("EXCELMANUS_FEISHU_APP_ID", "")
        app_secret = credentials.get("app_secret") or os.environ.get("EXCELMANUS_FEISHU_APP_SECRET", "")
        if not app_id or not app_secret or "••••••••" in (app_secret or ""):
            return _error_json_response(400, "未提供飞书 App ID 或 App Secret。")
        try:
            import httpx
            # 使用飞书 API 获取 tenant_access_token 验证凭证
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.post(
                    "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
                    json={"app_id": app_id, "app_secret": app_secret},
                )
                data = resp.json()
                if data.get("code") == 0:
                    return JSONResponse(content={
                        "status": "ok",
                        "message": f"凭证有效！已获取 tenant_access_token (AppID: {app_id})",
                        "bot_info": {"app_id": app_id},
                    })
                return _error_json_response(
                    400,
                    f"凭证无效: {data.get('msg', '未知错误')} (code={data.get('code')})",
                )
        except Exception as e:
            return _error_json_response(502, f"连接飞书 API 失败: {e}")

    else:
        return JSONResponse(content={
            "status": "ok",
            "message": f"渠道 {channel_name} 的凭证验证尚未实现，请直接启动测试。",
        })


# ── 飞书 Webhook 回调 ─────────────────────────────────────

@_router.post("/api/v1/channels/feishu/webhook")
async def feishu_webhook(request: Request) -> JSONResponse:
    """飞书事件订阅回调端点。

    支持：
    1. URL 验证（challenge-response）
    2. im.message.receive_v1 消息事件
    3. 卡片按钮回调（card.action.trigger）
    """
    try:
        body = await request.json()
    except Exception:
        return _error_json_response(400, "无效的请求体")

    # 1) URL 验证（飞书事件订阅配置时的 challenge 验证）
    if body.get("type") == "url_verification":
        challenge = body.get("challenge", "")
        return JSONResponse(content={"challenge": challenge})

    # 2) 获取 launcher 中的飞书 adapter 和 handler
    launcher = getattr(request.app.state, "channel_launcher", None)
    if launcher is None:
        return _error_json_response(503, "渠道启动器未初始化")

    feishu_adapter = launcher._apps.get("feishu")
    feishu_handler = launcher._handlers.get("feishu")
    if feishu_adapter is None or feishu_handler is None:
        return _error_json_response(503, "飞书渠道未启动，请先在设置中启动飞书渠道")

    # 3) 处理事件（v2.0 事件格式）
    header = body.get("header", {})
    event_type = header.get("event_type", "")
    event = body.get("event", {})

    if event_type == "im.message.receive_v1":
        from excelmanus.channels.feishu.handlers import handle_feishu_event
        # 异步处理，不阻塞 webhook 响应
        _fire_and_forget(
            handle_feishu_event(feishu_adapter, feishu_handler, event),
            name="feishu_message",
        )
        return JSONResponse(content={"code": 0, "msg": "ok"})

    # 4) 卡片按钮回调
    if event_type == "card.action.trigger":
        from excelmanus.channels.feishu.handlers import handle_feishu_card_action
        _fire_and_forget(
            handle_feishu_card_action(feishu_adapter, feishu_handler, event),
            name="feishu_card_action",
        )
        return JSONResponse(content={"code": 0, "msg": "ok"})

    # 5) v1.0 事件格式兼容（部分老版本飞书使用）
    if body.get("event") and not header:
        v1_event = body.get("event", {})
        v1_type = v1_event.get("type", "")
        if v1_type == "message":
            from excelmanus.channels.feishu.handlers import handle_feishu_event
            _fire_and_forget(
                handle_feishu_event(feishu_adapter, feishu_handler, v1_event),
                name="feishu_v1_message",
            )
            return JSONResponse(content={"code": 0, "msg": "ok"})

    logger.debug("飞书未处理的事件类型: %s", event_type or body.get("type", "unknown"))
    return JSONResponse(content={"code": 0, "msg": "ok"})


@_router.get("/api/v1/settings/session-isolation")
async def get_session_isolation(request: Request) -> JSONResponse:
    """查询会话用户隔离开关状态。"""
    return JSONResponse(content={
        "session_isolation_enabled": getattr(
            request.app.state, "session_isolation_enabled", False
        ),
    })


@_router.put("/api/v1/settings/session-isolation")
async def set_session_isolation(request: Request) -> JSONResponse:
    """[DEPRECATED] 会话隔离现在随 auth 自动启用，此接口保留仅为兼容。"""
    auth_enabled = getattr(request.app.state, "auth_enabled", False)
    return JSONResponse(content={
        "status": "ok",
        "session_isolation_enabled": auth_enabled,
        "deprecated": True,
        "message": "Session isolation is now automatic when auth is enabled.",
    })


@_router.get("/api/v1/settings/docker-sandbox")
async def get_docker_sandbox(request: Request) -> JSONResponse:
    """查询 Docker 沙盒状态（含 daemon / 镜像可用性）。"""
    from excelmanus.security.docker_sandbox import is_docker_available, is_sandbox_image_ready

    return JSONResponse(content={
        "docker_sandbox_enabled": getattr(
            request.app.state, "docker_sandbox_enabled", False
        ),
        "docker_available": is_docker_available(),
        "sandbox_image_ready": is_sandbox_image_ready(),
    })


@_router.put("/api/v1/settings/docker-sandbox")
async def set_docker_sandbox(request: Request) -> JSONResponse:
    """管理员切换 Docker 沙盒开关（持久化到 config_kv）。

    请求体: {"enabled": true}  启用
    请求体: {"enabled": false} 关闭
    启用时自动检测 Docker 可用性和镜像状态。
    """
    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    body = await request.json()
    enabled = bool(body.get("enabled", False))

    if enabled:
        from excelmanus.security.docker_sandbox import (
            is_docker_available,
            is_sandbox_image_ready,
            build_sandbox_image,
        )

        if not is_docker_available():
            return _error_json_response(
                400, "Docker daemon 不可用，无法启用 Docker 沙盒。"
            )
        if not is_sandbox_image_ready():
            ok, msg = build_sandbox_image()
            if not ok:
                return _error_json_response(
                    400, f"沙盒镜像构建失败: {msg}"
                )

    if _config_store is not None:
        _config_store.set("docker_sandbox_enabled", "true" if enabled else "false")
    request.app.state.docker_sandbox_enabled = enabled

    from excelmanus.tools.code_tools import init_docker_sandbox
    init_docker_sandbox(enabled)

    logger.info("Docker 沙盒已%s（管理员操作）", "启用" if enabled else "关闭")
    return JSONResponse(content={
        "status": "ok",
        "docker_sandbox_enabled": enabled,
    })


@_router.post("/api/v1/settings/docker-sandbox/build")
async def build_docker_sandbox_image(request: Request) -> JSONResponse:
    """管理员触发构建/重建 Docker 沙盒镜像。"""
    if getattr(request.app.state, "auth_enabled", False):
        from excelmanus.auth.dependencies import get_current_user_from_request
        user = await get_current_user_from_request(request)
        if user.role != "admin":
            return _error_json_response(403, "需要管理员权限。")

    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    force = bool(body.get("force", False))

    from excelmanus.security.docker_sandbox import build_sandbox_image

    ok, msg = build_sandbox_image(force=force)
    if ok:
        return JSONResponse(content={"status": "ok", "message": msg})
    return _error_json_response(500, f"镜像构建失败: {msg}")


# ── 快捷方式管理 API ────────────────────────────────────
# 注意: 版本检查、备份管理、更新执行、安装注册表、数据迁移
# 已统一迁移到 api_routes_version.py（/api/v1/version/* 命名空间）


@_router.get("/api/v1/shortcut/info")
async def shortcut_info(request: Request) -> JSONResponse:
    """返回桌面快捷方式状态。"""
    from excelmanus.shortcuts import get_shortcut_info
    return JSONResponse(content=get_shortcut_info())


@_router.post("/api/v1/shortcut/create")
async def shortcut_create(request: Request) -> JSONResponse:
    """创建桌面快捷方式。

    服务器/Docker 模式下需传入 site_url 参数来创建浏览器书签，
    standalone 模式下创建启动脚本快捷方式。
    """
    from excelmanus.shortcuts import create_desktop_shortcut
    project_root = Path(__file__).resolve().parent.parent
    mode = _config.deploy_mode if _config is not None else "standalone"
    site_url = ""
    if mode in ("server", "docker"):
        try:
            body = await request.json()
            site_url = body.get("site_url", "")
        except Exception:
            pass
        if not site_url:
            return _error_json_response(
                400, "服务器模式下需要提供 site_url 参数（如 http://your-server:3000）"
            )
    result = create_desktop_shortcut(
        project_root, deploy_mode=mode, site_url=site_url,
    )
    if result:
        return JSONResponse(content={"status": "ok", "path": result})
    return _error_json_response(500, "创建快捷方式失败，请查看日志。")


@_router.post("/api/v1/shortcut/remove")
async def shortcut_remove(request: Request) -> JSONResponse:
    """删除桌面快捷方式。"""
    from excelmanus.shortcuts import remove_desktop_shortcut
    ok = remove_desktop_shortcut()
    if ok:
        return JSONResponse(content={"status": "ok"})
    return _error_json_response(404, "未找到桌面快捷方式。")


# 默认 ASGI app（供 `uvicorn excelmanus.api:app` 与测试直接导入）
app = create_app()


# ── 入口函数 ──────────────────────────────────────────────


def main() -> None:
    """API 服务入口函数（pyproject.toml 入口点）。

    支持 --channels 参数或 EXCELMANUS_CHANNELS 环境变量来协同启动渠道 Bot::

        # 启动 API + QQ Bot
        python -m excelmanus.api --channels qq

        # 多渠道
        python -m excelmanus.api --channels qq,telegram

        # 通过环境变量（适合 Docker / systemd）
        EXCELMANUS_CHANNELS=qq python -m excelmanus.api
    """
    import argparse

    parser = argparse.ArgumentParser(
        prog="excelmanus-api",
        description="ExcelManus API Server",
    )
    parser.add_argument(
        "--channels",
        type=str,
        default="",
        help="要协同启动的渠道 Bot，逗号分隔（如 qq,telegram）",
    )
    parser.add_argument("--host", type=str, default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8000)
    args, _ = parser.parse_known_args()

    # 将 CLI 参数提升为环境变量，供 lifespan 读取
    # （模块级 app = create_app() 已在 import 时创建，lifespan 延迟读取环境变量）
    os.environ["EXCELMANUS_API_PORT"] = str(args.port)
    if args.channels:
        os.environ["EXCELMANUS_CHANNELS"] = args.channels

    uvicorn.run(
        "excelmanus.api:app",
        host=args.host,
        port=args.port,
        log_level="info",
    )
