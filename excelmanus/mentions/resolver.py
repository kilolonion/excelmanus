"""上下文解析器：将 Mention 列表解析为实际内容。

对每种 Mention 类型（file/folder/skill/mcp/img）执行内容提取，
生成注入系统提示词的 context_block。所有 file/folder 引用通过
FileAccessGuard 进行安全校验。
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import TYPE_CHECKING

from excelmanus.excel_extensions import EXCEL_EXTENSIONS
from excelmanus.mentions.parser import Mention, ResolvedMention
from excelmanus.security.guard import FileAccessGuard, SecurityViolationError

if TYPE_CHECKING:
    from excelmanus.mcp.manager import MCPManager
    from excelmanus.skillpacks.loader import SkillpackLoader

# Excel 文件扩展名（向后兼容保留私有别名）
_EXCEL_EXTENSIONS = EXCEL_EXTENSIONS

# 目录树排除项
_EXCLUDED_NAMES = {".venv", "node_modules", "__pycache__"}


def _count_tokens(text: str) -> int:
    """使用 tiktoken 计算 token 数，失败时降级为字符估算。"""
    try:
        import tiktoken

        enc = tiktoken.encoding_for_model("gpt-4o")
        return len(enc.encode(text))
    except Exception:
        # 降级：1 token ≈ 4 字符
        return len(text) // 4


def _truncate_to_tokens(text: str, max_tokens: int) -> str:
    """将文本截断到 max_tokens 以内，按行截断。"""
    try:
        import tiktoken

        enc = tiktoken.encoding_for_model("gpt-4o")
        tokens = enc.encode(text)
        if len(tokens) <= max_tokens:
            return text
        truncated = enc.decode(tokens[:max_tokens])
        return truncated
    except Exception:
        # 降级：1 token ≈ 4 字符
        max_chars = max_tokens * 4
        if len(text) <= max_chars:
            return text
        return text[:max_chars]


class MentionResolver:
    """将 Mention 列表解析为实际内容。"""

    def __init__(
        self,
        workspace_root: str,
        guard: FileAccessGuard,
        skill_loader: SkillpackLoader | None = None,
        mcp_manager: MCPManager | None = None,
        max_file_tokens: int = 2000,
        max_folder_depth: int = 2,
    ) -> None:
        self._workspace_root = workspace_root
        self._guard = guard
        self._skill_loader = skill_loader
        self._mcp_manager = mcp_manager
        self._max_file_tokens = max_file_tokens
        self._max_folder_depth = max_folder_depth

    async def resolve(self, mentions: list[Mention]) -> list[ResolvedMention]:
        """解析所有 Mention，返回 ResolvedMention 列表。"""
        results: list[ResolvedMention] = []
        for mention in mentions:
            if mention.kind == "file":
                results.append(self._resolve_file(mention))
            elif mention.kind == "folder":
                results.append(self._resolve_folder(mention))
            elif mention.kind == "skill":
                results.append(self._resolve_skill(mention))
            elif mention.kind == "mcp":
                results.append(await self._resolve_mcp(mention))
            elif mention.kind == "img":
                # img 类型保持现有行为，不生成 context_block
                results.append(ResolvedMention(mention=mention))
            else:
                results.append(
                    ResolvedMention(
                        mention=mention,
                        error=f"未知的引用类型：{mention.kind}",
                    )
                )
        return results

    # ── file 解析 ─────────────────────────────────────────

    def _resolve_file(self, mention: Mention) -> ResolvedMention:
        """解析文件引用：安全校验 + 内容提取 + token 预算限制。"""
        # 安全校验
        try:
            resolved_path = self._guard.resolve_and_validate(mention.value)
        except SecurityViolationError as exc:
            return ResolvedMention(mention=mention, error=str(exc))

        if not resolved_path.exists():
            return ResolvedMention(
                mention=mention, error=f"文件不存在：{mention.value}"
            )

        if not resolved_path.is_file():
            return ResolvedMention(
                mention=mention, error=f"路径不是文件：{mention.value}"
            )

        # 根据文件类型选择解析方式
        suffix = resolved_path.suffix.lower()
        if suffix in _EXCEL_EXTENSIONS:
            return self._resolve_excel_file(mention, resolved_path)
        else:
            return self._resolve_text_file(mention, resolved_path)

    def _resolve_excel_file(
        self, mention: Mention, path: Path
    ) -> ResolvedMention:
        """解析 Excel 文件：sheet 列表 + 行列数 + 首个 sheet 表头。"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                lines: list[str] = []
                first_sheet_headers: list[str] | None = None

                for i, name in enumerate(wb.sheetnames):
                    ws = wb[name]
                    rows = ws.max_row or 0
                    cols = ws.max_column or 0
                    lines.append(f"  {name} ({rows}行×{cols}列)")

                    # 首个 sheet 的表头行
                    if i == 0 and rows > 0:
                        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                        if header_row:
                            first_sheet_headers = [
                                str(c) if c is not None else "" for c in header_row
                            ]

                summary_parts = [f"Sheets: {len(wb.sheetnames)}"]
                summary_parts.extend(lines)
                if first_sheet_headers:
                    headers_str = ", ".join(first_sheet_headers)
                    summary_parts.append(
                        f"  Headers({wb.sheetnames[0]}): {headers_str}"
                    )

                context = "\n".join(summary_parts)
            finally:
                wb.close()

            # token 预算限制
            context = _truncate_to_tokens(context, self._max_file_tokens)
            return ResolvedMention(mention=mention, context_block=context)

        except Exception as exc:
            return ResolvedMention(
                mention=mention,
                error=f"文件读取失败：{mention.value}：{exc}",
            )

    def _resolve_text_file(
        self, mention: Mention, path: Path
    ) -> ResolvedMention:
        """解析文本文件：读取前 N 行，受 token 预算限制。"""
        try:
            collected_lines: list[str] = []
            current_text = ""

            with open(path, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    candidate = current_text + line
                    token_count = _count_tokens(candidate)
                    if token_count > self._max_file_tokens:
                        # 当前行会超出预算，停止
                        break
                    collected_lines.append(line.rstrip("\n"))
                    current_text = candidate

            context = "\n".join(collected_lines)
            # 最终截断保障
            context = _truncate_to_tokens(context, self._max_file_tokens)
            return ResolvedMention(mention=mention, context_block=context)

        except Exception as exc:
            return ResolvedMention(
                mention=mention,
                error=f"文件读取失败：{mention.value}：{exc}",
            )

    # ── folder 解析 ───────────────────────────────────────

    def _resolve_folder(self, mention: Mention) -> ResolvedMention:
        """解析文件夹引用：安全校验 + 树形目录结构。"""
        # 安全校验
        try:
            resolved_path = self._guard.resolve_and_validate(mention.value)
        except SecurityViolationError as exc:
            return ResolvedMention(mention=mention, error=str(exc))

        if not resolved_path.exists():
            return ResolvedMention(
                mention=mention, error=f"目录不存在：{mention.value}"
            )

        if not resolved_path.is_dir():
            return ResolvedMention(
                mention=mention, error=f"路径不是目录：{mention.value}"
            )

        tree = self._build_tree(resolved_path, depth=0)
        return ResolvedMention(mention=mention, context_block=tree)

    def _build_tree(
        self,
        dir_path: Path,
        depth: int,
        prefix: str = "",
        is_last: bool = True,
    ) -> str:
        """递归构建目录树文本，深度 ≤ max_folder_depth，排除隐藏/排除项。"""
        lines: list[str] = []

        if depth == 0:
            # 根目录名
            lines.append(f"{dir_path.name}/")
        else:
            connector = "└── " if is_last else "├── "
            name = dir_path.name
            if dir_path.is_dir():
                name += "/"
            lines.append(f"{prefix}{connector}{name}")

        if depth >= self._max_folder_depth:
            return "\n".join(lines)

        if not dir_path.is_dir():
            return "\n".join(lines)

        # 列出子项，排除隐藏文件和排除目录
        try:
            entries = sorted(dir_path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        except PermissionError:
            return "\n".join(lines)

        filtered = [
            e for e in entries
            if not e.name.startswith(".") and e.name not in _EXCLUDED_NAMES
        ]

        if depth == 0:
            child_prefix = ""
        else:
            child_prefix = prefix + ("    " if is_last else "│   ")

        for i, entry in enumerate(filtered):
            is_entry_last = i == len(filtered) - 1
            if entry.is_dir():
                subtree = self._build_tree(
                    entry,
                    depth=depth + 1,
                    prefix=child_prefix,
                    is_last=is_entry_last,
                )
                lines.append(subtree)
            else:
                connector = "└── " if is_entry_last else "├── "
                lines.append(f"{child_prefix}{connector}{entry.name}")

        return "\n".join(lines)

    # ── skill 解析 ────────────────────────────────────────

    def _resolve_skill(self, mention: Mention) -> ResolvedMention:
        """解析 Skill 引用：加载 SKILL.md 内容。"""
        if self._skill_loader is None:
            return ResolvedMention(
                mention=mention, error=f"技能不存在：{mention.value}"
            )

        skill = self._skill_loader.get_skillpack(mention.value)
        if skill is None:
            return ResolvedMention(
                mention=mention, error=f"技能不存在：{mention.value}"
            )

        context = skill.render_context()
        return ResolvedMention(mention=mention, context_block=context)

    # ── mcp 解析 ──────────────────────────────────────────

    async def _resolve_mcp(self, mention: Mention) -> ResolvedMention:
        """解析 MCP 服务引用：查询工具列表。"""
        if self._mcp_manager is None:
            return ResolvedMention(
                mention=mention,
                error=f"MCP 服务未连接或不存在：{mention.value}",
            )

        connected = self._mcp_manager.connected_servers()
        if mention.value not in connected:
            return ResolvedMention(
                mention=mention,
                error=f"MCP 服务未连接或不存在：{mention.value}",
            )

        # 从 get_server_info 获取工具列表
        server_info_list = self._mcp_manager.get_server_info()
        tools_desc: list[str] = []
        for info in server_info_list:
            if info.get("name") == mention.value:
                tool_names = info.get("tools", [])
                tools_desc = tool_names
                break

        if tools_desc:
            tools_str = ", ".join(tools_desc)
            context = f"Server: {mention.value}\nTools: {tools_str}"
        else:
            context = f"Server: {mention.value}\nTools: (无工具)"

        return ResolvedMention(mention=mention, context_block=context)
