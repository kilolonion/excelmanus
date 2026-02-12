"""格式化 Skill：提供单元格格式化和列宽调整工具。"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excelmanus.logger import get_logger
from excelmanus.security import FileAccessGuard
from excelmanus.skills import ToolDef

logger = get_logger("skills.format")

# ── Skill 元数据 ──────────────────────────────────────────

SKILL_NAME = "format"
SKILL_DESCRIPTION = "格式化工具集：单元格样式设置与列宽调整"

# ── 模块级 FileAccessGuard（延迟初始化） ─────────────────

_guard: FileAccessGuard | None = None


def _get_guard() -> FileAccessGuard:
    """获取或创建 FileAccessGuard 单例。"""
    global _guard
    if _guard is None:
        _guard = FileAccessGuard(".")
    return _guard


def init_guard(workspace_root: str) -> None:
    """初始化文件访问守卫（供外部配置调用）。

    Args:
        workspace_root: 工作目录根路径。
    """
    global _guard
    _guard = FileAccessGuard(workspace_root)


# ── 工具函数 ──────────────────────────────────────────────


def format_cells(
    file_path: str,
    cell_range: str,
    sheet_name: str | None = None,
    font: dict[str, Any] | None = None,
    fill: dict[str, Any] | None = None,
    border: dict[str, Any] | None = None,
    alignment: dict[str, Any] | None = None,
    number_format: str | None = None,
) -> str:
    """对指定单元格范围应用格式化样式。

    Args:
        file_path: Excel 文件路径。
        cell_range: 单元格范围，如 "A1:C3" 或 "A1"。
        sheet_name: 工作表名称，默认活动工作表。
        font: 字体设置，支持 name/size/bold/italic/color。
        fill: 填充设置，支持 color（十六进制颜色码）。
        border: 边框设置，支持 style（thin/medium/thick）和 color。
        alignment: 对齐设置，支持 horizontal/vertical/wrap_text。
        number_format: 数字格式字符串，如 "#,##0.00"。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # 构建 openpyxl 样式对象
    style_font = _build_font(font) if font else None
    style_fill = _build_fill(fill) if fill else None
    style_border = _build_border(border) if border else None
    style_alignment = _build_alignment(alignment) if alignment else None

    # 应用样式到范围内每个单元格
    applied_count = 0
    cell_data = ws[cell_range]

    # ws[cell_range] 对单个单元格返回 Cell，对范围返回 tuple of tuples
    from openpyxl.cell.cell import Cell
    if isinstance(cell_data, Cell):
        rows = ((cell_data,),)
    elif isinstance(cell_data, tuple) and cell_data and not isinstance(cell_data[0], tuple):
        # 单行范围如 "A1:C1" 返回 tuple of Cell
        rows = (cell_data,)
    else:
        rows = cell_data

    for row in rows:
        cells = row if isinstance(row, tuple) else (row,)
        for cell in cells:
            if style_font:
                cell.font = style_font
            if style_fill:
                cell.fill = style_fill
            if style_border:
                cell.border = style_border
            if style_alignment:
                cell.alignment = style_alignment
            if number_format:
                cell.number_format = number_format
            applied_count += 1

    wb.save(safe_path)
    wb.close()

    logger.info("已格式化 %s 范围 %s（%d 个单元格）", safe_path.name, cell_range, applied_count)

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "range": cell_range,
            "cells_formatted": applied_count,
        },
        ensure_ascii=False,
    )


def adjust_column_width(
    file_path: str,
    columns: dict[str, float] | None = None,
    auto_fit: bool = False,
    sheet_name: str | None = None,
) -> str:
    """调整列宽：支持指定宽度或自动适配。

    Args:
        file_path: Excel 文件路径。
        columns: 列宽映射，如 {"A": 20, "B": 15}。与 auto_fit 互斥时优先使用。
        auto_fit: 是否自动适配所有列宽（基于内容最大长度）。
        sheet_name: 工作表名称，默认活动工作表。

    Returns:
        JSON 格式的操作结果。
    """
    guard = _get_guard()
    safe_path = guard.resolve_and_validate(file_path)

    wb = load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    adjusted: dict[str, float] = {}

    if columns:
        # 手动指定列宽
        for col_letter, width in columns.items():
            ws.column_dimensions[col_letter.upper()].width = width
            adjusted[col_letter.upper()] = width

    elif auto_fit:
        # 自动适配：遍历所有列，取每列最大内容长度
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            # 加 2 作为边距
            width = max_len + 2
            ws.column_dimensions[col_letter].width = width
            adjusted[col_letter] = width

    wb.save(safe_path)
    wb.close()

    logger.info("已调整 %s 列宽（%d 列）", safe_path.name, len(adjusted))

    return json.dumps(
        {
            "status": "success",
            "file": safe_path.name,
            "columns_adjusted": adjusted,
        },
        ensure_ascii=False,
    )


# ── 内部辅助函数 ──────────────────────────────────────────


def _build_font(config: dict[str, Any]) -> Font:
    """从配置字典构建 openpyxl Font 对象。"""
    return Font(
        name=config.get("name"),
        size=config.get("size"),
        bold=config.get("bold"),
        italic=config.get("italic"),
        color=config.get("color"),
    )


def _build_fill(config: dict[str, Any]) -> PatternFill:
    """从配置字典构建 openpyxl PatternFill 对象。"""
    return PatternFill(
        start_color=config.get("color", "FFFFFF"),
        end_color=config.get("color", "FFFFFF"),
        fill_type="solid",
    )


def _build_border(config: dict[str, Any]) -> Border:
    """从配置字典构建 openpyxl Border 对象。"""
    style = config.get("style", "thin")
    color = config.get("color", "000000")
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _build_alignment(config: dict[str, Any]) -> Alignment:
    """从配置字典构建 openpyxl Alignment 对象。"""
    return Alignment(
        horizontal=config.get("horizontal"),
        vertical=config.get("vertical"),
        wrap_text=config.get("wrap_text"),
    )


# ── get_tools() 导出 ──────────────────────────────────────


def get_tools() -> list[ToolDef]:
    """返回格式化 Skill 的所有工具定义。"""
    return [
        ToolDef(
            name="format_cells",
            description="对 Excel 单元格范围应用格式化样式（字体、填充、边框、对齐、数字格式）",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "单元格范围，如 'A1:C3' 或 'A1'",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                    "font": {
                        "type": "object",
                        "description": "字体设置：name/size/bold/italic/color",
                        "properties": {
                            "name": {"type": "string"},
                            "size": {"type": "number"},
                            "bold": {"type": "boolean"},
                            "italic": {"type": "boolean"},
                            "color": {"type": "string", "description": "十六进制颜色码，如 'FF0000'"},
                        },
                    },
                    "fill": {
                        "type": "object",
                        "description": "填充设置：color（十六进制颜色码）",
                        "properties": {
                            "color": {"type": "string", "description": "十六进制颜色码，如 'FFFF00'"},
                        },
                    },
                    "border": {
                        "type": "object",
                        "description": "边框设置：style（thin/medium/thick）和 color",
                        "properties": {
                            "style": {"type": "string", "enum": ["thin", "medium", "thick"]},
                            "color": {"type": "string", "description": "十六进制颜色码"},
                        },
                    },
                    "alignment": {
                        "type": "object",
                        "description": "对齐设置：horizontal/vertical/wrap_text",
                        "properties": {
                            "horizontal": {"type": "string", "enum": ["left", "center", "right"]},
                            "vertical": {"type": "string", "enum": ["top", "center", "bottom"]},
                            "wrap_text": {"type": "boolean"},
                        },
                    },
                    "number_format": {
                        "type": "string",
                        "description": "数字格式字符串，如 '#,##0.00'",
                    },
                },
                "required": ["file_path", "cell_range"],
                "additionalProperties": False,
            },
            func=format_cells,
        ),
        ToolDef(
            name="adjust_column_width",
            description="调整 Excel 列宽：支持手动指定宽度或自动适配内容",
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "columns": {
                        "type": "object",
                        "description": "列宽映射，如 {'A': 20, 'B': 15}",
                        "additionalProperties": {"type": "number"},
                    },
                    "auto_fit": {
                        "type": "boolean",
                        "description": "是否自动适配所有列宽",
                        "default": False,
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称，默认活动工作表",
                    },
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            func=adjust_column_width,
        ),
    ]
