"""Bench 断言校验引擎：对 BenchResult 执行声明式断言规则。

suite JSON 中通过 ``assertions`` 字段声明规则，runner 执行完毕后
调用本模块自动校验，结果嵌入输出 JSON 的 ``validation`` 字段。

支持 **suite 级默认** + **case 级覆盖** 的两层合并策略。
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ── 断言结果 ──────────────────────────────────────────────


@dataclass
class AssertionResult:
    """单条断言的校验结果。"""

    rule: str
    passed: bool
    expected: Any = None
    actual: Any = None
    message: str = ""
    severity: str = "error"  # "error" | "warning"

    def to_dict(self) -> dict[str, Any]:
        d: dict[str, Any] = {
            "rule": self.rule,
            "passed": self.passed,
        }
        if self.expected is not None:
            d["expected"] = self.expected
        if self.actual is not None:
            d["actual"] = self.actual
        if self.message:
            d["message"] = self.message
        if self.severity != "error":
            d["severity"] = self.severity
        return d


@dataclass
class ValidationSummary:
    """一个 case 的断言校验汇总。"""

    total: int = 0
    passed: int = 0
    failed: int = 0
    results: list[AssertionResult] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "total": self.total,
            "passed": self.passed,
            "failed": self.failed,
            "results": [r.to_dict() for r in self.results],
        }


@dataclass
class SuiteValidationSummary:
    """一个 suite 的断言校验汇总。"""

    total_assertions: int = 0
    passed: int = 0
    failed: int = 0
    pass_rate: float = 0.0
    failed_cases: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_assertions": self.total_assertions,
            "passed": self.passed,
            "failed": self.failed,
            "pass_rate": self.pass_rate,
            "failed_cases": self.failed_cases,
        }


# ── 断言合并 ──────────────────────────────────────────────


def merge_assertions(
    suite_assertions: dict[str, Any] | None,
    case_assertions: dict[str, Any] | None,
) -> dict[str, Any]:
    """合并 suite 级和 case 级断言，case 级覆盖 suite 级同名字段。"""
    merged: dict[str, Any] = {}
    if suite_assertions:
        merged.update(suite_assertions)
    if case_assertions:
        merged.update(case_assertions)
    return merged


# ── 校验规则实现 ──────────────────────────────────────────


def _check_status(result_dict: dict[str, Any], expected: str) -> AssertionResult:
    actual = result_dict.get("execution", {}).get("status", "unknown")
    return AssertionResult(
        rule="status",
        passed=actual == expected,
        expected=expected,
        actual=actual,
        message="" if actual == expected else f"状态 {actual!r} != 期望 {expected!r}",
    )


def _check_max_int(
    result_dict: dict[str, Any],
    rule_name: str,
    path: tuple[str, ...],
    limit: int,
) -> AssertionResult:
    """通用的 max_xxx 校验：按 path 从 result_dict 取值，判断 <= limit。"""
    obj = result_dict
    for key in path:
        obj = obj.get(key, {}) if isinstance(obj, dict) else {}
    actual = obj if isinstance(obj, (int, float)) else 0
    passed = actual <= limit
    return AssertionResult(
        rule=rule_name,
        passed=passed,
        expected=f"<= {limit}",
        actual=actual,
        message="" if passed else f"{rule_name}: {actual} 超过上限 {limit}",
    )


def _check_expected_skill(
    result_dict: dict[str, Any],
    expected_skill: str,
) -> AssertionResult:
    skills_used = result_dict.get("execution", {}).get("skills_used", [])
    route_mode = result_dict.get("execution", {}).get("route_mode", "")
    # 在 skills_used 或 route_mode 中匹配
    matched = expected_skill in skills_used or route_mode == expected_skill
    actual_str = ", ".join(skills_used) if skills_used else route_mode
    return AssertionResult(
        rule="expected_skill",
        passed=matched,
        expected=expected_skill,
        actual=actual_str,
        message="" if matched else f"路由未命中期望技能 {expected_skill!r}，实际: {actual_str}",
    )


def _check_required_tools(
    result_dict: dict[str, Any],
    required: list[str],
) -> AssertionResult:
    tool_calls = result_dict.get("artifacts", {}).get("tool_calls", [])
    called_names = {tc.get("tool_name", "") for tc in tool_calls}
    missing = [t for t in required if t not in called_names]
    return AssertionResult(
        rule="required_tools",
        passed=len(missing) == 0,
        expected=required,
        actual=sorted(called_names),
        message="" if not missing else f"缺少必要工具调用: {missing}",
    )


def _check_forbidden_tools(
    result_dict: dict[str, Any],
    forbidden: list[str],
) -> AssertionResult:
    tool_calls = result_dict.get("artifacts", {}).get("tool_calls", [])
    called_names = {tc.get("tool_name", "") for tc in tool_calls}
    violations = [t for t in forbidden if t in called_names]
    return AssertionResult(
        rule="forbidden_tools",
        passed=len(violations) == 0,
        expected=f"不应调用 {forbidden}",
        actual=sorted(called_names),
        message="" if not violations else f"调用了禁止的工具: {violations}",
    )


def _check_no_empty_promise(result_dict: dict[str, Any]) -> AssertionResult:
    """首轮 LLM 响应不应有"空承诺"：content 非空 + 无 tool_calls。"""
    llm_calls = result_dict.get("artifacts", {}).get("llm_calls", [])
    if not llm_calls:
        return AssertionResult(
            rule="no_empty_promise",
            passed=True,
            message="无 LLM 调用记录，跳过检查",
            severity="warning",
        )
    first_resp = llm_calls[0].get("response", {})
    content = first_resp.get("content") or ""
    tool_calls = first_resp.get("tool_calls") or []
    # 空承诺 = 有文字回复但没有工具调用
    is_empty_promise = bool(content.strip()) and not tool_calls
    return AssertionResult(
        rule="no_empty_promise",
        passed=not is_empty_promise,
        expected="首轮应直接行动（tool_calls）或纯文本回复（无需工具时）",
        actual=f"content={len(content)}chars, tool_calls={len(tool_calls)}",
        message="" if not is_empty_promise else "首轮存在空承诺：有文字回复但未发起工具调用",
    )

def _check_no_silent_first_turn(result_dict: dict[str, Any]) -> AssertionResult:
    """首轮 LLM 响应不应完全静默：content 和 tool_calls 均为空。"""
    llm_calls = result_dict.get("artifacts", {}).get("llm_calls", [])
    if not llm_calls:
        return AssertionResult(
            rule="no_silent_first_turn",
            passed=True,
            message="无 LLM 调用记录，跳过检查",
            severity="warning",
        )
    first_resp = llm_calls[0].get("response", {})
    content = first_resp.get("content") or ""
    tool_calls = first_resp.get("tool_calls") or []
    is_silent = not content.strip() and not tool_calls
    return AssertionResult(
        rule="no_silent_first_turn",
        passed=not is_silent,
        expected="首轮应有响应内容（content 或 tool_calls）",
        actual=f"content={len(content)}chars, tool_calls={len(tool_calls)}",
        message="" if not is_silent else "首轮完全静默：既无文字回复也无工具调用",
    )



def _check_reply_contains(
    result_dict: dict[str, Any],
    keywords: list[str],
) -> AssertionResult:
    reply = result_dict.get("result", {}).get("reply", "")
    missing = [kw for kw in keywords if kw not in reply]
    return AssertionResult(
        rule="reply_contains",
        passed=len(missing) == 0,
        expected=keywords,
        actual=reply[:200] if reply else "(空回复)",
        message="" if not missing else f"回复缺少关键词: {missing}",
    )


def _check_reply_not_contains(
    result_dict: dict[str, Any],
    keywords: list[str],
) -> AssertionResult:
    reply = result_dict.get("result", {}).get("reply", "")
    violations = [kw for kw in keywords if kw in reply]
    return AssertionResult(
        rule="reply_not_contains",
        passed=len(violations) == 0,
        expected=f"不应包含 {keywords}",
        actual=reply[:200] if reply else "(空回复)",
        message="" if not violations else f"回复包含不期望的关键词: {violations}",
    )


# ── Golden 文件单元格比对 ─────────────────────────────────


def _normalize_to_date(val: Any):
    """尝试将值归一化为 date 对象。"""
    from datetime import date, datetime

    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _check_golden_cells(
    workfile_dir: Path,
    golden_file: str | Path,
    answer_position: str,
    answer_sheet: str | None = None,
) -> AssertionResult:
    """对比输出文件与 golden 文件在指定范围内的单元格值。

    Args:
        workfile_dir: 工作文件目录，包含 agent 修改后的 xlsx 文件。
        golden_file: golden 文件路径（相对于项目根或绝对路径）。
        answer_position: 待比对的单元格范围，格式如 "'Sheet1'!H1:I370"。
        answer_sheet: 可选的回答工作表名称（当 answer_position 中无 sheet 前缀时使用）。

    Returns:
        AssertionResult 包含匹配统计和不匹配样本。
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return AssertionResult(
            rule="golden_cells",
            passed=False,
            message="openpyxl 未安装，无法执行 golden 文件比对",
        )

    # ── 解析 answer_position ──
    sheet_name: str | None = answer_sheet
    cell_range = answer_position
    if "!" in answer_position:
        sheet_part, cell_range = answer_position.rsplit("!", 1)
        # 去掉引号，如 'Sheet1' -> Sheet1
        sheet_name = sheet_part.strip("'\"")

    # sheet_name 延迟解析：若无前缀，打开文件后回退到首个/唯一 sheet
    _sheet_name_deferred = sheet_name is None

    # ── 定位输出文件 ──
    output_file: Path | None = None
    if workfile_dir and workfile_dir.is_dir():
        xlsx_files = list(workfile_dir.glob("*.xlsx"))
        if len(xlsx_files) == 1:
            output_file = xlsx_files[0]
        elif len(xlsx_files) > 1:
            # 尝试匹配 golden 文件名
            golden_name = Path(golden_file).name
            # 查找与源文件同名的（init 文件被复制到 workdir）
            for f in xlsx_files:
                if "init" in f.name or f.stem in golden_name:
                    output_file = f
                    break
            if output_file is None:
                output_file = xlsx_files[0]

    if output_file is None or not output_file.exists():
        return AssertionResult(
            rule="golden_cells",
            passed=False,
            message=f"未找到输出 xlsx 文件: {workfile_dir}",
        )

    # ── 定位 golden 文件 ──
    golden_path = Path(golden_file)
    if not golden_path.is_absolute():
        # 尝试相对于项目根目录解析
        for candidate in [
            golden_path,
            Path.cwd() / golden_path,
        ]:
            if candidate.exists():
                golden_path = candidate
                break

    if not golden_path.exists():
        return AssertionResult(
            rule="golden_cells",
            passed=False,
            message=f"golden 文件不存在: {golden_file}",
        )

    # ── 批量加载单元格数据（优化：减少 I/O 次数）────
    try:
        # 先只加载 data_only=True 版本（用于值比对）
        wb_out = load_workbook(output_file, data_only=True, read_only=True)
        wb_gold = load_workbook(golden_path, data_only=True, read_only=True)
    except Exception as exc:
        return AssertionResult(
            rule="golden_cells",
            passed=False,
            message=f"加载 xlsx 文件失败: {exc}",
        )

    try:
        # ── 延迟解析 sheet_name：无前缀时回退到首个/唯一 sheet ──
        if _sheet_name_deferred:
            if len(wb_out.sheetnames) == 1:
                sheet_name = wb_out.sheetnames[0]
            elif wb_out.active is not None:
                sheet_name = wb_out.active.title
            else:
                sheet_name = wb_out.sheetnames[0]
            logger.debug(
                "answer_position 无 sheet 前缀，回退到 %r (from %r)",
                sheet_name, answer_position,
            )

        if sheet_name not in wb_out.sheetnames:
            return AssertionResult(
                rule="golden_cells",
                passed=False,
                message=f"输出文件缺少工作表 {sheet_name!r}",
            )
        if sheet_name not in wb_gold.sheetnames:
            return AssertionResult(
                rule="golden_cells",
                passed=False,
                message=f"golden 文件缺少工作表 {sheet_name!r}",
            )

        ws_out = wb_out[sheet_name]
        ws_gold = wb_gold[sheet_name]

        # ── 解析范围边界 ──
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        # ── 批量读取（优化：使用迭代器代替逐个单元格访问）────
        total = 0
        matched = 0
        formula_written = 0  # 公式已写入但未求值（容错计数）
        mismatches: list[dict[str, Any]] = []
        max_mismatches = 20  # 限制不匹配样本数量

        # 批量读取输出和 golden 值到内存
        out_values = []
        gold_values = []

        for row in range(min_row, max_row + 1):
            out_row = []
            gold_row = []
            for col in range(min_col, max_col + 1):
                out_row.append(ws_out.cell(row=row, column=col).value)
                gold_row.append(ws_gold.cell(row=row, column=col).value)
            out_values.append(out_row)
            gold_values.append(gold_row)

        # ── 逐单元格比对 ──
        for row_idx, (out_row, gold_row) in enumerate(zip(out_values, gold_values)):
            row = min_row + row_idx
            for col_idx, (val_out, val_gold) in enumerate(zip(out_row, gold_row)):
                total += 1
                col = min_col + col_idx

                # 归一化: None 和 '' 视为等价
                norm_out = None if (val_out is None or val_out == "") else val_out
                norm_gold = None if (val_gold is None or val_gold == "") else val_gold

                # 数值宽松比对: float vs int
                if isinstance(norm_out, float) and isinstance(norm_gold, int):
                    values_equal = abs(norm_out - norm_gold) < 1e-9
                elif isinstance(norm_out, int) and isinstance(norm_gold, float):
                    values_equal = abs(norm_out - norm_gold) < 1e-9
                elif isinstance(norm_out, float) and isinstance(norm_gold, float):
                    values_equal = abs(norm_out - norm_gold) < 1e-9
                else:
                    # 日期归一化比较
                    d_out = _normalize_to_date(norm_out)
                    d_gold = _normalize_to_date(norm_gold)
                    if d_out is not None and d_gold is not None:
                        values_equal = d_out == d_gold
                    else:
                        values_equal = norm_out == norm_gold

                if values_equal:
                    matched += 1
                else:
                    # ── 公式容错：输出为 None 但有公式写入 ──
                    # 延迟加载公式（仅在需要时）
                    if norm_out is None and norm_gold is not None:
                        # 只有在第一次发现不匹配且需要检查公式时才加载公式文件
                        if formula_written == 0 and mismatches == []:
                            try:
                                wb_out_formula = load_workbook(output_file, data_only=False, read_only=True)
                                ws_out_formula = wb_out_formula[sheet_name] if sheet_name in wb_out_formula.sheetnames else None
                                if ws_out_formula is not None:
                                    formula_val = ws_out_formula.cell(row=row, column=col).value
                                    if isinstance(formula_val, str) and formula_val.startswith("="):
                                        formula_written += 1
                                        matched += 1  # 计入 matched
                                        continue  # 不计入 mismatches
                                if ws_out_formula:
                                    wb_out_formula.close()
                            except Exception:
                                pass  # 忽略公式加载失败

                    if len(mismatches) < max_mismatches:
                        from openpyxl.utils import get_column_letter
                        cell_ref = f"{get_column_letter(col)}{row}"
                        mismatches.append({
                            "cell": cell_ref,
                            "output": _serialize_cell_value(val_out),
                            "golden": _serialize_cell_value(val_gold),
                        })

    finally:
        wb_out.close()
        wb_gold.close()

    effective_matched = matched + formula_written
    accuracy = round(effective_matched / total * 100, 1) if total > 0 else 0.0
    # 公式容错：matched + formula_written 覆盖全部单元格即视为通过
    passed = effective_matched == total

    message_parts: list[str] = []
    if formula_written > 0:
        message_parts.append(f"公式已写入但未求值: {formula_written} 个单元格")
    if not passed:
        strict_mismatches = total - effective_matched
        message_parts.append(
            f"单元格比对: {matched}/{total} 精确匹配"
            + (f", {formula_written} 公式容错" if formula_written else "")
            + f", {strict_mismatches} 不匹配"
        )
        if mismatches:
            samples = "; ".join(
                f"{m['cell']}: 输出={m['output']!r} vs 期望={m['golden']!r}"
                for m in mismatches[:5]
            )
            message_parts.append(f"示例: {samples}")

    return AssertionResult(
        rule="golden_cells",
        passed=passed,
        expected=f"{total} 单元格全部匹配 ({answer_position})",
        actual={
            "total_cells": total,
            "matched": matched,
            "formula_written": formula_written,
            "accuracy_pct": accuracy,
            "mismatches_sample": mismatches,
        },
        message=" | ".join(message_parts),
    )


def _serialize_cell_value(val: Any) -> Any:
    """将单元格值序列化为 JSON 友好格式。"""
    if val is None:
        return None
    if isinstance(val, (int, float, bool, str)):
        return val
    return str(val)

def _check_min_match_rate(
    result_dict: dict[str, Any],
    threshold: float,
) -> AssertionResult:
    """检查 verify_excel_replica 工具返回的 match_rate 是否达标。

    从 tool_calls 中找到最后一次 verify_excel_replica 调用，
    解析其 result JSON 字符串中的 match_rate 字段，与阈值比较。

    Args:
        result_dict: BenchResult.to_dict() 的输出。
        threshold: match_rate 最低阈值（如 0.95）。

    Returns:
        AssertionResult 包含实际 match_rate 和比较结果。
    """
    tool_calls = result_dict.get("artifacts", {}).get("tool_calls", [])

    # 找到所有 verify_excel_replica 调用
    verify_calls = [
        tc for tc in tool_calls
        if tc.get("tool_name") == "verify_excel_replica"
    ]

    if not verify_calls:
        return AssertionResult(
            rule="min_match_rate",
            passed=False,
            expected=f"match_rate >= {threshold}",
            actual=None,
            message="未找到 verify_excel_replica 工具调用",
        )

    # 取最后一次调用（最终验证结果）
    last_call = verify_calls[-1]
    raw_result = last_call.get("result", "")

    try:
        parsed = json.loads(raw_result) if isinstance(raw_result, str) else raw_result
        actual_rate = parsed["match_rate"]
    except (json.JSONDecodeError, KeyError, TypeError) as exc:
        return AssertionResult(
            rule="min_match_rate",
            passed=False,
            expected=f"match_rate >= {threshold}",
            actual=raw_result,
            message=f"解析 verify_excel_replica 结果失败: {exc}",
        )

    passed = actual_rate >= threshold
    message = "" if passed else (
        f"match_rate {actual_rate} 低于阈值 {threshold}"
    )

    return AssertionResult(
        rule="min_match_rate",
        passed=passed,
        expected=f"match_rate >= {threshold}",
        actual=actual_rate,
        message=message,
    )




def _split_composite_position(raw: str) -> list[str]:
    """将可能包含逗号分隔的多 sheet 复合 position 拆分为独立子 position。

    示例:
        "'Sheet1'!A1:A50,'Sheet2'!A1:E20" → ["'Sheet1'!A1:A50", "'Sheet2'!A1:E20"]
        "Sheet1!B2:B6" → ["Sheet1!B2:B6"]
        "B2" → ["B2"]
    """
    # 简单 position（无逗号或无 !）直接返回
    if "," not in raw:
        return [raw]
    # 有逗号但无 ! → 可能是单个 range（不太可能，保险起见直接返回）
    if "!" not in raw:
        return [raw]
    # 按逗号拆分，重新组装带 sheet 前缀的子 position
    parts = [p.strip() for p in raw.split(",")]
    result: list[str] = []
    for part in parts:
        if part:
            result.append(part)
    return result if result else [raw]


# ── 主校验函数 ────────────────────────────────────────────


def validate_case(
    result_dict: dict[str, Any],
    assertions: dict[str, Any],
    *,
    expected: dict[str, Any] | None = None,
    workfile_dir: Path | None = None,
) -> ValidationSummary:
    """对单个用例的输出 dict 执行所有断言规则。

    Args:
        result_dict: ``BenchResult.to_dict()`` 的输出。
        assertions: 合并后的断言规则字典。
        expected: case 级 ``expected`` 字段（含 golden_file / answer_position 等）。
        workfile_dir: 工作文件目录，用于 golden 文件比对。

    Returns:
        ValidationSummary 包含所有断言结果。
    """
    has_golden = bool(
        expected and expected.get("golden_file") and expected.get("answer_position")
    )
    if not assertions and not has_golden:
        return ValidationSummary()

    results: list[AssertionResult] = []

    # status
    if "status" in assertions:
        results.append(_check_status(result_dict, assertions["status"]))

    # max_iterations
    if "max_iterations" in assertions:
        results.append(_check_max_int(
            result_dict, "max_iterations",
            ("execution", "iterations"), assertions["max_iterations"],
        ))

    # max_llm_calls
    if "max_llm_calls" in assertions:
        results.append(_check_max_int(
            result_dict, "max_llm_calls",
            ("stats", "llm_call_count"), assertions["max_llm_calls"],
        ))

    # max_tool_calls
    if "max_tool_calls" in assertions:
        results.append(_check_max_int(
            result_dict, "max_tool_calls",
            ("stats", "tool_call_count"), assertions["max_tool_calls"],
        ))

    # max_tool_failures
    if "max_tool_failures" in assertions:
        results.append(_check_max_int(
            result_dict, "max_tool_failures",
            ("stats", "tool_failures"), assertions["max_tool_failures"],
        ))

    # max_tokens
    if "max_tokens" in assertions:
        results.append(_check_max_int(
            result_dict, "max_tokens",
            ("stats", "total_tokens"), assertions["max_tokens"],
        ))

    # max_duration_seconds
    if "max_duration_seconds" in assertions:
        results.append(_check_max_int(
            result_dict, "max_duration_seconds",
            ("execution", "duration_seconds"), assertions["max_duration_seconds"],
        ))

    # expected_skill
    if "expected_skill" in assertions:
        results.append(_check_expected_skill(result_dict, assertions["expected_skill"]))

    # required_tools
    if "required_tools" in assertions:
        results.append(_check_required_tools(result_dict, assertions["required_tools"]))

    # forbidden_tools
    if "forbidden_tools" in assertions:
        results.append(_check_forbidden_tools(result_dict, assertions["forbidden_tools"]))

    # no_empty_promise
    if assertions.get("no_empty_promise"):
        results.append(_check_no_empty_promise(result_dict))

    # no_silent_first_turn
    if assertions.get("no_silent_first_turn"):
        results.append(_check_no_silent_first_turn(result_dict))

    # reply_contains
    if "reply_contains" in assertions:
        results.append(_check_reply_contains(result_dict, assertions["reply_contains"]))

    # reply_not_contains
    if "reply_not_contains" in assertions:
        results.append(_check_reply_not_contains(result_dict, assertions["reply_not_contains"]))

    # min_match_rate
    if "min_match_rate" in assertions:
        results.append(_check_min_match_rate(result_dict, assertions["min_match_rate"]))

    # golden_cells — 自动从 expected 推导，无需在 assertions 中显式声明
    if expected and expected.get("golden_file") and expected.get("answer_position"):
        if workfile_dir is not None:
            raw_position = expected["answer_position"]
            # 支持多 sheet 复合 position（逗号分隔的多个 'SheetN'!Range）
            sub_positions = _split_composite_position(raw_position)
            for sub_pos in sub_positions:
                results.append(_check_golden_cells(
                    workfile_dir=workfile_dir,
                    golden_file=expected["golden_file"],
                    answer_position=sub_pos,
                    answer_sheet=expected.get("answer_sheet"),
                ))
        else:
            logger.debug(
                "跳过 golden_cells 断言: workfile_dir 未提供",
            )

    passed = sum(1 for r in results if r.passed)
    failed = len(results) - passed

    return ValidationSummary(
        total=len(results),
        passed=passed,
        failed=failed,
        results=results,
    )


def aggregate_suite_validation(
    case_validations: list[tuple[str, ValidationSummary]],
) -> SuiteValidationSummary:
    """聚合多个 case 的校验结果为 suite 级汇总。

    Args:
        case_validations: [(case_id, ValidationSummary), ...]
    """
    total = sum(v.total for _, v in case_validations)
    passed = sum(v.passed for _, v in case_validations)
    failed = total - passed
    failed_cases = [cid for cid, v in case_validations if v.failed > 0]
    pass_rate = round(passed / total * 100, 1) if total > 0 else 100.0

    return SuiteValidationSummary(
        total_assertions=total,
        passed=passed,
        failed=failed,
        pass_rate=pass_rate,
        failed_cases=failed_cases,
    )
